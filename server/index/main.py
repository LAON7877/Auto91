# ========== 自動套件檢查和安裝 ==========
# 在所有其他import之前執行套件檢查，確保所有依賴都已安裝
import sys
import os
import json

# 添加當前目錄到Python路徑，以便導入package_checker
current_dir = os.path.dirname(os.path.abspath(__file__))
if current_dir not in sys.path:
    sys.path.insert(0, current_dir)

# ========== 動態版本載入系統 ==========
def load_version_info():
    """從version.json載入版本資訊"""
    try:
        version_file = os.path.join(current_dir, 'version.json')
        with open(version_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError) as e:
        print(f"⚠️  無法載入版本資訊: {e}")
        # 返回預設版本資訊
        return {
            "version": "1.0.0",
            "build": "unknown",
            "description": "Auto91雙系統自動交易平台"
        }

# 載入版本資訊
VERSION_INFO = load_version_info()
CURRENT_VERSION = VERSION_INFO.get('version', '1.0.0')
CURRENT_BUILD = VERSION_INFO.get('build', 'unknown')
SYSTEM_DESCRIPTION = VERSION_INFO.get('description', 'Auto91雙系統自動交易平台')

# ========== 套件依賴檢查 ==========
# 在所有其他導入之前執行依賴檢查，確保所有套件都已安裝
def check_and_install_dependencies():
    """檢查並安裝依賴套件"""
    import subprocess
    import sys
    
    
    try:
        # 首先嘗試導入依賴管理器
        from dependencymanager import auto_install_dependencies_on_startup
        
        # 執行自動依賴安裝檢查
        dependency_success = auto_install_dependencies_on_startup()
        
        if dependency_success:
            print(f"✅ 依賴檢查完成 - {CURRENT_VERSION}")
        else:
            print(f"⚠️  依賴檢查部分失敗 - {CURRENT_VERSION}")
        
        return dependency_success
        
    except ImportError as e:
        if "dependencymanager" in str(e):
            
            # 從 requirements.txt 讀取依賴列表
            requirements_file = os.path.join(os.path.dirname(__file__), 'requirements.txt')
            basic_dependencies = []
            
            if os.path.exists(requirements_file):
                try:
                    with open(requirements_file, 'r', encoding='utf-8') as f:
                        for line in f:
                            line = line.strip()
                            # 跳過註釋和空行
                            if line and not line.startswith('#'):
                                # 提取套件名稱（去除版本限制）
                                package_name = line.split('>=')[0].split('==')[0].split('[')[0]
                                basic_dependencies.append(package_name)
                    print(f"📋 從 requirements.txt 讀取到 {len(basic_dependencies)} 個依賴")
                except Exception as e:
                    print(f"⚠️ 讀取 requirements.txt 失敗: {e}")
            
            # 如果無法讀取 requirements.txt，使用預設清單
            if not basic_dependencies:
                basic_dependencies = [
                    'flask', 'requests', 'shioaji', 'openpyxl', 
                    'python-dotenv', 'schedule', 'websocket-client', 'pytz', 'psutil'
                ]
                print("📋 使用預設依賴清單")
            
            # 檢查並安裝基本依賴
            missing_packages = []
            
            # 套件名稱映射（安裝名稱 -> 導入名稱）
            package_import_map = {
                'python-dotenv': 'dotenv',
                'websocket-client': 'websocket',
                'pytz': 'pytz',
                'openpyxl': 'openpyxl',
                'requests': 'requests',
                'flask': 'flask',
                'shioaji': 'shioaji',
                'schedule': 'schedule',
                'psutil': 'psutil'
            }
            
            for package in basic_dependencies:
                # 獲取正確的導入名稱
                import_name = package_import_map.get(package, package.replace('-', '_'))
                
                try:
                    __import__(import_name)
                    print(f"✅ {package} 已安裝")
                except ImportError:
                    missing_packages.append(package)
                    print(f"❌ {package} 未安裝")
            
            # 安裝缺失的套件
            if missing_packages:
                print(f"🔄 正在安裝 {len(missing_packages)} 個缺失的套件...")
                for package in missing_packages:
                    try:
                        print(f"安裝 {package}...")
                        subprocess.check_call([sys.executable, '-m', 'pip', 'install', package])
                        print(f"✅ {package} 安裝成功")
                    except Exception as install_error:
                        print(f"❌ {package} 安裝失敗: {install_error}")
                        return False
                
                print("✅ 所有基本依賴已安裝完成")
            else:
                print("✅ 所有基本依賴已存在")
            
            # 檢查可選依賴
            optional_dependencies = ['flask-cors']
            missing_optional = []
            
            for package in optional_dependencies:
                try:
                    __import__(package.replace('-', '_'))
                except ImportError:
                    missing_optional.append(package)
            
            if missing_optional:
                pass
            
            return True
        else:
            raise e
            
    except Exception as e:
        return False
    
    finally:
        pass

# 執行依賴檢查
try:
    dependency_check_success = check_and_install_dependencies()
except Exception as e:
    dependency_check_success = False

# ========== 正常的程式導入開始 ==========
from flask import Flask, send_from_directory, request, jsonify, abort
from flask_cors import CORS
import threading
import webview
import re
import requests
import subprocess
import json
import time
import logging
import atexit
import signal
import platform
import zipfile
import shutil
from datetime import datetime, timezone, timedelta
import csv
import schedule
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter

# ========== 進程管理器導入和初始化 ==========
# 進程管理功能已簡化，使用標準的threading和subprocess
process_manager = None

def create_managed_thread(target, name="未知線程", daemon=True, args=(), kwargs={}):
    """創建線程（兼容函數）"""
    return threading.Thread(target=target, name=name, daemon=daemon, args=args, kwargs=kwargs)

def create_managed_subprocess(*args, **kwargs):
    """創建subprocess（兼容函數）"""
    kwargs.pop('process_name', None)  # 移除自定義參數
    return subprocess.Popen(*args, **kwargs)

# ========== 日誌配置 ==========
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger('main_system')

def register_subprocess(process, name="未知進程", description=""):
    """註冊子進程到ALL_CHILD_PROCESSES列表以便清理"""
    global ALL_CHILD_PROCESSES
    if process and process not in ALL_CHILD_PROCESSES:
        ALL_CHILD_PROCESSES.append(process)
        logger.info(f"📋 已註冊子進程: {name} (PID: {process.pid}) - {description}")

def unregister_subprocess(process):
    """從ALL_CHILD_PROCESSES列表移除子進程"""
    global ALL_CHILD_PROCESSES
    if process in ALL_CHILD_PROCESSES:
        ALL_CHILD_PROCESSES.remove(process)
        logger.info(f"📋 已移除子進程註冊: PID {process.pid}")

def register_thread(thread, name="未知線程"):
    """註冊非daemon線程到ALL_ACTIVE_THREADS列表以便清理"""
    global ALL_ACTIVE_THREADS
    if thread and not thread.daemon and thread not in ALL_ACTIVE_THREADS:
        ALL_ACTIVE_THREADS.append(thread)
        logger.info(f"🧵 已註冊活動線程: {name}")

def signal_shutdown():
    """設置全域停止標誌，通知所有線程準備結束"""
    global SHUTDOWN_FLAG
    SHUTDOWN_FLAG.set()
    logger.info("🚩 已設置全域停止標誌")

# ========================== 美化輸出系統 ==========================

def format_console_output(category, status, message, detail=None):
    """
    統一的後端輸出格式化函數
    category: 分類 (SYSTEM, API, TRADE, TG, TUNNEL 等)
    status: 狀態 (SUCCESS, ERROR, WARNING, INFO, START, STOP)
    message: 主要訊息
    detail: 詳細信息 (可選)
    """
    # 狀態圖標和顏色
    status_icons = {
        'SUCCESS': '',
        'ERROR': '', 
        'WARNING': '',
        'INFO': '',
        'START': '',
        'STOP': '',
        'LOADING': '',
        'TRADE': '',
        'MONEY': ''
    }
    
    # 格式化分類 (8個字符寬度，左對齊)
    formatted_category = f"[{category:<7}]"
    
    # 格式化狀態 (8個字符寬度)
    icon = status_icons.get(status, '')
    formatted_status = f"{status:<8}"
    
    # 主要輸出
    output = f"{formatted_category} {formatted_status} │ {message}"
    
    # 如果有詳細信息，添加第二行
    if detail:
        padding = " " * (len(formatted_category) + len(formatted_status) + 3)
        output += f"\n{padding}│ {detail}"
    
    return output

def print_console(category, status, message, detail=None):
    """便捷的格式化輸出函數"""
    print(format_console_output(category, status, message, detail))

# 導入 Tunnel
try:
    from tunnel import CloudflareTunnel, TunnelManager, QuantTradingTunnelManager
    CLOUDFLARE_TUNNEL_AVAILABLE = True
except ImportError:
    CLOUDFLARE_TUNNEL_AVAILABLE = False
    print_console("SYSTEM", "WARNING", "Tunnel 模組未找到")

# 導入 BTC 模組
try:
    import btcmain
    BTC_MODULE_AVAILABLE = True
except ImportError:
    BTC_MODULE_AVAILABLE = False
    print_console("SYSTEM", "WARNING", "BTC 模組未找到")

# 導入 psutil 用於進程管理
try:
    import psutil
    PSUTIL_AVAILABLE = True
except ImportError:
    PSUTIL_AVAILABLE = False
    print_console("SYSTEM", "WARNING", "psutil 模組未安裝，進程清理功能將受限")

# 永豐API相關
try:
    import shioaji as sj
    from dotenv import load_dotenv
    SHIOAJI_AVAILABLE = True
    DOTENV_AVAILABLE = True
except ImportError as e:
    if 'shioaji' in str(e):
        SHIOAJI_AVAILABLE = False
        print_console("SYSTEM", "WARNING", "shioaji 模組未安裝，永豐API功能將無法使用")
    if 'dotenv' in str(e):
        DOTENV_AVAILABLE = False
        print_console("SYSTEM", "WARNING", "python-dotenv 模組未安裝")

# 創建兼容的常數處理類別
class SafeConstants:
    """安全的常數處理類別，避免shioaji版本相容性問題"""
    
    @staticmethod
    def get_action(action_str):
        """獲取動作常數"""
        if not SHIOAJI_AVAILABLE:
            return action_str
        try:
            if action_str.upper() == 'BUY':
                return sj.constant.Action.Buy
            else:
                return sj.constant.Action.Sell
        except AttributeError:
            return action_str
    
    @staticmethod
    def get_price_type(price_type_str):
        """獲取價格類型常數"""
        if not SHIOAJI_AVAILABLE:
            return price_type_str
        try:
            if price_type_str.upper() == 'MKT':
                return sj.constant.FuturesPriceType.MKT
            else:
                return sj.constant.FuturesPriceType.LMT
        except AttributeError:
            return price_type_str
    
    @staticmethod
    def get_order_type(order_type_str):
        """獲取訂單類型常數"""
        if not SHIOAJI_AVAILABLE:
            return order_type_str
        try:
            if order_type_str.upper() == 'IOC':
                return sj.constant.FuturesOrderType.IOC
            else:
                return sj.constant.FuturesOrderType.ROD
        except AttributeError:
            return order_type_str
    
    @staticmethod
    def get_oc_type(oc_type='Auto'):
        """獲取開平倉類型常數"""
        if not SHIOAJI_AVAILABLE:
            return oc_type
        try:
            if oc_type.upper() == 'NEW':
                return sj.constant.FuturesOCType.New
            elif oc_type.upper() == 'COVER':
                return sj.constant.FuturesOCType.Cover
            else:
                return sj.constant.FuturesOCType.Auto
        except AttributeError:
            return oc_type
    
    @staticmethod
    def check_order_status(status, target_status):
        """檢查訂單狀態"""
        if not SHIOAJI_AVAILABLE:
            return str(status).upper() == target_status.upper()
        try:
            if target_status.upper() == 'FILLED':
                return status == sj.constant.OrderStatus.Filled
            elif target_status.upper() == 'FAILED':
                return status == sj.constant.OrderStatus.Failed
            else:
                return False
        except AttributeError:
            return str(status).upper() == target_status.upper()

safe_constants = SafeConstants()

def compare_versions(version1, version2):
    """
    比較兩個版本號
    返回值: 1 表示 version1 > version2, 0 表示相等, -1 表示 version1 < version2
    """
    def version_to_tuple(v):
        return tuple(map(int, (v.split("."))))
    
    v1_tuple = version_to_tuple(version1)
    v2_tuple = version_to_tuple(version2)
    
    if v1_tuple > v2_tuple:
        return 1
    elif v1_tuple < v2_tuple:
        return -1
    else:
        return 0

app = Flask(__name__, static_folder='../web', static_url_path='')
CORS(app, origins=['*'])  # 允許所有域名的跨域請求

# 全局變數初始化
tunnel_type = 'tx'  # 預設隧道類型
tunnel_service = None
tunnel_manager = None

# 請求記錄中間件
def identify_tunnel_type(host):
    """根據Host頭部識別隧道類型"""
    global tunnel_manager
    
    if not tunnel_manager:
        return 'tx'  # 默認為TX
    
    # 獲取兩個隧道的URL
    tx_tunnel = tunnel_manager.get_tunnel('tx')
    btc_tunnel = tunnel_manager.get_tunnel('btc')
    
    # 檢查host是否匹配隧道URL
    if tx_tunnel and tx_tunnel.tunnel_url:
        tx_hostname = tx_tunnel.tunnel_url.replace('https://', '').replace('http://', '')
        if tx_hostname in host:
            return 'tx'
    
    if btc_tunnel and btc_tunnel.tunnel_url:
        btc_hostname = btc_tunnel.tunnel_url.replace('https://', '').replace('http://', '')
        if btc_hostname in host:
            return 'btc'
    
    # 如果無法識別，根據請求路徑判斷
    if hasattr(request, 'path') and '/api/btc' in request.path:
        return 'btc'
    
    return 'tx'  # 默認為TX

@app.before_request
def log_request():
    """記錄所有進入的請求"""
    try:
        # 記錄請求開始時間
        request.start_time = time.time()
        
        # 識別請求來源隧道
        host = request.headers.get('Host', '')
        request.tunnel_type = identify_tunnel_type(host)
        
    except Exception as e:
        print_console("SYSTEM", "ERROR", "記錄請求開始時間失敗", str(e))

@app.after_request
def log_response(response):
    """記錄所有請求的響應"""
    global tunnel_manager
    
    try:
        # 計算響應時間
        response_time = None
        if hasattr(request, 'start_time'):
            response_time = round((time.time() - request.start_time) * 1000, 2)  # 轉換為毫秒
        
        # 獲取隧道類型
        tunnel_type = getattr(request, 'tunnel_type', 'tx')
        
        # 記錄到對應的隧道日誌
        if tunnel_manager:
            tunnel = tunnel_manager.get_tunnel(tunnel_type)
            if tunnel and hasattr(tunnel, 'add_request_log'):
                tunnel.add_request_log(
                    method=request.method,
                    path=request.path,
                    status_code=response.status_code,
                    response_time=response_time
                )
        
        # 保持向後兼容，也記錄到主tunnel_service
        if tunnel_service and hasattr(tunnel_service, 'add_request_log'):
            tunnel_service.add_request_log(
                method=request.method,
                path=request.path,
                status_code=response.status_code,
                response_time=response_time
            )
        
        # 對於重要的API請求和動態端點，也記錄到前端顯示的custom_request_logs中
        important_paths = ['/api/manual/order', '/api/btc/manual_order', '/webhook']
        is_dynamic_endpoint = (request.method == 'POST' and 
                              request.path.startswith('/') and 
                              request.path.count('/') == 1 and 
                              len(request.path.split('/')[1]) > 5)
        
        if request.path in important_paths or is_dynamic_endpoint:
            # 為動態端點添加請求記錄，確保顯示在請求日誌中
            if is_dynamic_endpoint:
                add_custom_request_log(
                    method=request.method,
                    uri=request.path,
                    status=response.status_code,
                    extra_info={
                        'response_time': response_time,
                        'tunnel_type': tunnel_type,
                        'is_webhook': True
                    }
                )
        
    except Exception as e:
        print_console("SYSTEM", "ERROR", "記錄請求響應失敗", str(e))
    
    return response

# Cloudflare Tunnel 相關變數（直接替換 ngrok）
tunnel_manager = None  # 隧道管理器實例
tunnel_service = None  # 保持向後兼容
tunnel_process = None
tunnel_status = "stopped"  # stopped, starting, running, error
tunnel_version = None
tunnel_update_available = False
tunnel_auto_restart_timer = None  # 自動重啟定時器


# 永豐API相關變數
sinopac_api = None
sinopac_connected = False
sinopac_account = None
sinopac_login_status = False
sinopac_login_time = None  # 新增：記錄登入時間

# 期貨合約相關變數
futures_contracts = {
    'TXF': None,  # 大台指
    'MXF': None,  # 小台指
    'TMF': None   # 微台指
}
margin_requirements = {
    '大台': 0,
    '小台': 0,
    '微台': 0
}

# 新增：12小時自動登出相關變數
AUTO_LOGOUT_HOURS = 12  # 12小時自動登出
auto_logout_timer = None  # 自動登出定時器

# 新增：記錄上一次的保證金金額
last_margin_requirements = {
    '大台': 0,
    '小台': 0,
    '微台': 0
}

# 新增：訂單映射管理（參考TXserver.py架構）
order_octype_map = {}  # 記錄訂單詳細資訊
duplicate_signal_window = {}  # 重複訊號時間窗口記錄
global_lock = threading.Lock()  # 線程鎖

# 子進程和線程管理
ALL_CHILD_PROCESSES = []  # 記錄所有啟動的子進程，用於程式關閉時清理
ALL_ACTIVE_THREADS = []   # 記錄所有非daemon線程，用於程式關閉時清理
SHUTDOWN_FLAG = threading.Event()  # 全域停止標誌
flask_server_thread = None  # Flask服務器線程
flask_server_shutdown = False  # Flask服務器關閉標誌

# TXserver 風格的全域變數和狀態管理
contract_txf = None
contract_mxf = None 
contract_tmf = None
active_trades = {"txf": None, "mxf": None, "tmf": None}
recent_signals = set()
contract_key_map = {"大台": "txf", "小台": "mxf", "微台": "tmf"}
# 轉倉相關變數
has_processed_delivery_exit = False
rollover_mode = False  # 是否處於轉倉模式
next_month_contracts = {}  # 次月合約快取
rollover_start_date = None  # 轉倉開始日期
rollover_processed_signals = set()  # 已處理的轉倉訊號ID

# 斷線重連相關變數
connection_monitor_timer = None  # 連線監控定時器
connection_check_interval = 60  # 每1分鐘檢查一次連線狀態
max_reconnect_attempts = 999  # 無限重連嘗試
reconnect_attempts = 0  # 當前重連嘗試次數
last_connection_check = None  # 上次連線檢查時間
is_reconnecting = False  # 是否正在重連中

# Webhook活動追蹤變數（用於自動/手動判斷）
last_webhook_time = 0

# 自定義請求日誌系統（用於前端顯示）
custom_request_logs = []
MAX_CUSTOM_LOGS = 50  # 最多保留50筆記錄

# 新增：錯誤訊息翻譯對照表
OP_MSG_TRANSLATIONS = {
    "Order not found": "訂單未找到",
    "Price not satisfied": "價格未滿足",
    "Insufficient margin": "保證金不足",
    "Invalid quantity": "無效數量",
    "Invalid price": "無效價格",
    "Market closed": "市場已關閉",
    "非該商品可下單時間": "非交易時間",
    "可委託金額不足": "保證金不足",
    "委託價格不符合規範": "委託價格不符合規範",
    "委託數量不符合規範": "委託數量不符合規範",
    "帳戶未開啟": "帳戶未開啟",
    "未開放交易": "未開放交易",
    "已超過風控限制": "已超過風控限制",
    "該商品已停止交易": "該商品已停止交易",
    "無足夠庫存": "無足夠庫存",
    "委託失敗": "委託失敗",
    "系統忙碌中": "系統忙碌中",
    "超過單筆委託數量限制": "超過單筆委託數量限制",
    "超過單日委託數量限制": "超過單日委託數量限制",
    "不可做空": "不可做空",
    "暫停交易": "暫停交易",
    "漲跌停限制": "漲跌停限制",
    "未知錯誤": "未知錯誤",
    "Order Cancelled": "手動取消訂單",
    "cancelled": "手動取消訂單",
    "Cancelled": "手動取消訂單",
    "CANCELLED": "手動取消訂單",
    "手動取消": "手動取消訂單",
    "取消訂單": "手動取消訂單"
}

ENV_TEMPLATE = '''# Telegram Bot
BOT_TOKEN=7202376519:AAF-i3MbuMEpz0W7nFE9KmieqVw7L5s0xK4

# Telegram ID
CHAT_ID=

# 永豐 API Key
API_KEY=

# 永豐 Secret Key
SECRET_KEY=

# 身分證字號
PERSON_ID=

# 台股日曆
HOLIDAY_DIR=Desktop/AutoTX/server/holiday

# 憑證檔案
CA_PATH=Desktop/AutoTX/server/holiday

# 憑證密碼
CA_PASSWD=

# 憑證起始日
CERT_START=

# 憑證到期日
CERT_END=

# 登入狀態
LOGIN=0
'''

CONFIG_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config')
TX_ENV_PATH = os.path.join(CONFIG_DIR, 'tx.env')
ENV_PATH = TX_ENV_PATH  # 為了向後兼容，保持TX為默認
os.makedirs(CONFIG_DIR, exist_ok=True)
if not os.path.exists(ENV_PATH):
    with open(ENV_PATH, 'w', encoding='utf-8') as f:
        f.write(ENV_TEMPLATE)

def update_login_status(status):
    """更新LOGIN狀態的通用函數"""
    if os.path.exists(ENV_PATH):
        with open(ENV_PATH, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        found = False
        for i, line in enumerate(lines):
            if line.startswith('LOGIN='):
                lines[i] = f'LOGIN={status}\n'
                found = True
                break
        if not found:
            lines.append(f'LOGIN={status}\n')
        with open(ENV_PATH, 'w', encoding='utf-8') as f:
            f.writelines(lines)





def init_tunnel_service(mode="temporary"):
    """初始化隧道服務（使用隧道管理器）"""
    global tunnel_manager, tunnel_service
    
    try:
        if CLOUDFLARE_TUNNEL_AVAILABLE:
            # 使用量化交易隧道管理器（帶智能URL變化通知）
            tunnel_manager = QuantTradingTunnelManager()
            # 為保持向後兼容，創建TX隧道作為默認隧道服務
            tunnel_service = tunnel_manager.create_tunnel('tx', mode)
            print_console("TUNNEL", "SUCCESS", f"已初始化智能隧道管理器 (模式: {mode})")
        else:
            print_console("TUNNEL", "WARNING", "Cloudflare Tunnel 不可用，請檢查模組")
    except Exception as e:
        print_console("TUNNEL", "ERROR", "初始化隧道管理器失敗", str(e))
        tunnel_manager = None
        tunnel_service = None

def start_tunnel():
    """啟動隧道服務（直接使用 Cloudflare Tunnel）"""
    return start_cloudflare_tunnel()


def start_cloudflare_tunnel():
    """啟動 Cloudflare Tunnel"""
    global tunnel_service, tunnel_status
    
    try:
        if not tunnel_service:
            print_console("TUNNEL", "WARNING", "Cloudflare Tunnel 服務未初始化")
            tunnel_status = "error"
            return False
        
        # 設定狀態為啟動中
        tunnel_status = "starting"
        
        # 根據模式決定啟動方式
        if tunnel_service.mode == 'temporary':
            # 臨時模式：直接啟動，無需token
            print_console("TUNNEL", "INFO", "使用臨時域名模式，無需token")
            success = tunnel_service.start_tunnel()
        else:
            # 其他模式：使用隧道實例的 load_token 方法
            token = tunnel_service.load_token()
            if not token:
                print_console("TUNNEL", "WARNING", "未找到 Cloudflare Token，請先設定")
                tunnel_status = "error"
                return False
                
            print_console("TUNNEL", "INFO", f"使用自訂域名模式，Token: {token[:20]}...")
            # 使用快速設定
            success = tunnel_service.quick_setup(token)
        
        if success:
            tunnel_status = "running"
            print_console("TUNNEL", "SUCCESS", "Cloudflare Tunnel 啟動成功!")
        else:
            tunnel_status = "error"
            print_console("TUNNEL", "ERROR", "Cloudflare Tunnel 啟動失敗!")
        
        return success
        
    except Exception as e:
        print_console("TUNNEL", "ERROR", "啟動 Cloudflare Tunnel 失敗", str(e))
        tunnel_status = "error"
        return False


def stop_tunnel():
    """停止隧道服務（直接使用 Cloudflare Tunnel）"""
    return stop_cloudflare_tunnel()


def stop_cloudflare_tunnel():
    """停止 Cloudflare Tunnel"""
    global tunnel_service, tunnel_status
    
    try:
        if tunnel_service:
            success = tunnel_service.stop_tunnel()
            if success:
                tunnel_status = "stopped"
            return success
        
        # 即使沒有服務實例，也更新狀態
        tunnel_status = "stopped"
        return True
    except Exception as e:
        print_console("TUNNEL", "ERROR", "停止 Cloudflare Tunnel 失敗", str(e))
        return False

def get_tunnel_status():
    """獲取隧道狀態（直接使用 Cloudflare Tunnel）"""
    return get_cloudflare_tunnel_status()


def get_cloudflare_tunnel_status():
    """獲取 Cloudflare Tunnel 狀態"""
    global tunnel_service, tunnel_status
    
    try:
        if tunnel_service:
            status_info = tunnel_service.get_status()
            tunnel_status = status_info.get('status', 'unknown')
            return {
                'status': tunnel_status,
                'public_url': status_info.get('url', ''),
                'tunnel_name': status_info.get('tunnel_name', ''),
                'service_type': 'cloudflare',
                'port': status_info.get('port', CURRENT_PORT),
                'message': status_info.get('message', ''),
                'timestamp': status_info.get('timestamp', '')
            }
        else:
            return {
                'status': 'error',
                'message': 'Cloudflare Tunnel 服務未初始化',
                'service_type': 'cloudflare'
            }
    except Exception as e:
        return {
            'status': 'error',
            'message': f'獲取 Cloudflare Tunnel 狀態失敗: {e}',
            'service_type': 'cloudflare'
        }

def get_cloudflare_tunnel_status_as_ngrok():
    """獲取 Cloudflare Tunnel 狀態並轉換為 ngrok 格式"""
    global tunnel_service, tunnel_status
    
    try:
        if tunnel_service:
            status_info = tunnel_service.get_status()
            tunnel_status = status_info.get('status', 'unknown')
            
            # 轉換為 ngrok 格式
            tunnels = []
            if status_info.get('url'):
                tunnels.append({
                    'public_url': status_info.get('url'),
                    'config': {
                        'addr': f"localhost:{CURRENT_PORT}",
                        'inspect': False
                    },
                    'proto': 'https',
                    'name': status_info.get('tunnel_name', 'cloudflare-tunnel')
                })
            
            return {
                'status': tunnel_status,
                'tunnels': tunnels,
                'message': f"Cloudflare Tunnel - {status_info.get('message', '')}" if status_info.get('message') else "Cloudflare Tunnel 運行中",
                'version': 'Cloudflare Tunnel',
                'service_type': 'cloudflare-tunnel'
            }
        else:
            return {
                'status': 'error',
                'tunnels': [],
                'message': 'Cloudflare Tunnel 服務未初始化',
                'version': 'Cloudflare Tunnel',
                'service_type': 'cloudflare-tunnel'
            }
    except Exception as e:
        return {
            'status': 'error',
            'tunnels': [],
            'message': f'獲取隧道狀態失敗: {e}',
            'version': 'Cloudflare Tunnel',
            'service_type': 'cloudflare-tunnel'
        }




def _verify_user_id(user_id, system_type):
    """驗證用戶ID是否與API獲取的真實帳戶ID匹配
    
    Args:
        user_id: 請求URL中的用戶ID
        system_type: 系統類型 ('tx' 或 'btc')
    
    Returns:
        bool: 驗證是否成功
    """
    try:
        # 獲取真實的用戶帳戶ID
        real_user_ids = _get_user_ids()
        real_user_id = real_user_ids.get(system_type)
        
        if not real_user_id:
            logger.warning(f"{system_type.upper()}用戶ID不存在或API未連接")
            return False
        
        # 比較請求的用戶ID是否與真實帳戶ID匹配
        is_valid = str(user_id) == str(real_user_id)
        
        if is_valid:
            logger.info(f"{system_type.upper()}用戶ID驗證成功: {user_id}")
        else:
            logger.warning(f"{system_type.upper()}用戶ID驗證失敗: 請求ID={user_id}, 真實ID={real_user_id}")
        
        return is_valid
        
    except Exception as e:
        logger.error(f"用戶ID驗證失敗: {e}")
        return False

def add_custom_request_log(method, uri, status, extra_info=None):
    """添加自定義請求記錄"""
    global custom_request_logs
    
    # 過濾不需要在前端顯示的請求記錄
    if uri in ['/api/btc_system_log', '/api/system_log'] and method in ['POST', 'GET']:
        # 系統日誌API調用本身不記錄，但其內容已經通過特殊處理添加
        logger.info(f"[SYSTEM_LOG_API] {method} {uri} - {status}")
        return
    
    # 註：移除webhook過濾邏輯，讓TX和BTC的webhook請求能顯示在前端請求日誌中
    # 用戶需要看到webhook請求記錄來監控系統運行狀況
    
    # 建立時間戳（加上日期但前端會只顯示時分秒）
    now = datetime.now()
    time_str = now.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3] + ' CST'
    display_time_str = now.strftime('%H:%M:%S.%f')[:-3] + ' CST'
    
    # 建立請求記錄
    log_entry = {
        'timestamp': time_str,
        'display_timestamp': display_time_str,
        'method': method,
        'uri': uri,
        'status': status,
        'status_text': get_status_text(status),
        'type': 'custom',
        'extra_info': extra_info or {}
    }
    
    # 添加到日誌列表
    with global_lock:
        custom_request_logs.append(log_entry)
        # 保持日誌數量在限制內，優先保留系統日誌
        if len(custom_request_logs) > MAX_CUSTOM_LOGS:
            # 分離系統日誌和其他日誌
            system_logs = [log for log in custom_request_logs if log.get('method') == 'TX_LOG' or log.get('method') == 'BTC_LOG']
            other_logs = [log for log in custom_request_logs if not (log.get('method') == 'TX_LOG' or log.get('method') == 'BTC_LOG')]
            
            # 優先保留系統日誌
            max_other_logs = MAX_CUSTOM_LOGS - len(system_logs)
            if max_other_logs > 0:
                other_logs = other_logs[-max_other_logs:]
            
            # 重新合併，系統日誌在前
            custom_request_logs = system_logs + other_logs


def format_timestamp(timestamp_str):
    """格式化時間戳為顯示用格式"""
    if not timestamp_str:
        return ''
    
    try:
        # 嘗試解析 ISO 格式的時間戳
        if 'T' in timestamp_str:
            dt = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
        else:
            # 如果不是 ISO 格式，嘗試其他常見格式
            dt = datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S')
        
        # 轉換為本地時間並格式化
        local_dt = dt.replace(tzinfo=timezone.utc).astimezone()
        return local_dt.strftime('%H:%M:%S')
    except:
        # 如果解析失敗，返回原始字符串的最後8個字符（通常是時間部分）
        return timestamp_str[-8:] if len(timestamp_str) >= 8 else timestamp_str

def get_status_text(status_code):
    """根據狀態碼獲取狀態文字"""
    status_texts = {
        200: 'OK',
        201: 'Created',
        204: 'No Content',
        301: 'Moved Permanently',
        302: 'Found',
        304: 'Not Modified',
        400: 'Bad Request',
        401: 'Unauthorized',
        403: 'Forbidden',
        404: 'Not Found',
        405: 'Method Not Allowed',
        500: 'Internal Server Error',
        502: 'Bad Gateway',
        503: 'Service Unavailable',
        504: 'Gateway Timeout'
    }
    return status_texts.get(status_code, 'Unknown')


def get_tunnel_requests():
    """獲取請求記錄（結合自定義記錄和 Cloudflare Tunnel 記錄）"""
    global custom_request_logs, tunnel_service
    
    # 複製自定義記錄列表（避免並發修改）
    with global_lock:
        custom_logs = custom_request_logs.copy()
    
    # 獲取 Cloudflare Tunnel 記錄
    tunnel_logs = []
    if tunnel_service:
        try:
            # 獲取 Cloudflare Tunnel 請求記錄
            tunnel_request_logs = tunnel_service.get_request_logs()
            if tunnel_request_logs:
                for log in tunnel_request_logs:
                    # 轉換 Cloudflare Tunnel 記錄格式為統一格式
                    tunnel_logs.append({
                        'timestamp': log.get('timestamp', ''),
                        'display_timestamp': format_timestamp(log.get('timestamp', '')),
                        'method': log.get('method', 'GET'),
                        'uri': log.get('path', log.get('uri', '')),  # 使用path字段
                        'status': log.get('status_code', log.get('status', 200)),  # 使用status_code字段
                        'status_text': get_status_text(log.get('status_code', log.get('status', 200))),
                        'source': 'cloudflare_tunnel',
                        'type': log.get('type', 'cloudflare'),
                        'extra_info': {
                            'response_time': log.get('response_time', ''),
                            'latency': log.get('latency', ''),
                            'user_agent': log.get('user_agent', ''),
                            'ip': log.get('ip', ''),
                            'response_size': log.get('response_size', '')
                        }
                    })
        except Exception as e:
            print_console("TUNNEL", "ERROR", "獲取 Cloudflare Tunnel 記錄失敗", str(e))
    
    # 為自定義記錄添加來源標識
    for log in custom_logs:
        log['source'] = 'custom'
    
    # 合併所有記錄
    all_logs = custom_logs + tunnel_logs
    
    # 按時間戳排序（最新的在前）
    try:
        def parse_timestamp_for_sort(timestamp_str):
            """解析時間戳用於排序"""
            if not timestamp_str:
                return datetime.min
            try:
                # 移除 CST 後綴
                clean_timestamp = timestamp_str.replace(' CST', '').strip()
                
                # 嘗試解析完整時間戳（包含日期）
                if ' ' in clean_timestamp and len(clean_timestamp) > 10:
                    # 格式：2025-07-17 15:49:14.570
                    try:
                        return datetime.strptime(clean_timestamp, '%Y-%m-%d %H:%M:%S.%f')
                    except ValueError:
                        # 如果沒有微秒，嘗試不含微秒的格式
                        return datetime.strptime(clean_timestamp, '%Y-%m-%d %H:%M:%S')
                else:
                    # 只有時分秒部分，需要推斷日期
                    try:
                        now = datetime.now()
                        
                        # 先嘗試解析時間部分
                        try:
                            time_part = datetime.strptime(clean_timestamp, '%H:%M:%S.%f').time()
                        except ValueError:
                            time_part = datetime.strptime(clean_timestamp, '%H:%M:%S').time()
                        
                        # 更精確的跨日邏輯
                        current_time = now.time()
                        current_date = now.date()
                        
                        # 計算今天的日期時間和昨天的日期時間
                        today_dt = datetime.combine(current_date, time_part)
                        yesterday_dt = datetime.combine(current_date - timedelta(days=1), time_part)
                        
                        # 判斷哪個更接近當前時間
                        time_diff_today = abs((today_dt - now).total_seconds())
                        time_diff_yesterday = abs((yesterday_dt - now).total_seconds())
                        
                        # 特殊處理：如果現在是凌晨（00:00-06:00），且日誌時間是晚上（18:00-23:59）
                        # 那麼日誌時間應該是昨天的
                        if (now.hour < 6 and time_part.hour >= 18):
                            return yesterday_dt
                        # 如果現在是晚上（18:00-23:59），且日誌時間是凌晨（00:00-06:00）
                        # 那麼日誌時間應該是明天的
                        elif (now.hour >= 18 and time_part.hour < 6):
                            tomorrow_dt = datetime.combine(current_date + timedelta(days=1), time_part)
                            return tomorrow_dt
                        # 其他情況，選擇時間差較小的
                        elif time_diff_yesterday < time_diff_today:
                            return yesterday_dt
                        else:
                            return today_dt
                            
                    except ValueError:
                        # 解析失敗，返回最小時間戳
                        return datetime.min
            except Exception as e:
                # 解析失敗，使用最小時間戳
                return datetime.min
        
        all_logs.sort(key=lambda x: parse_timestamp_for_sort(x.get('timestamp', '')), reverse=False)
    except:
        # 如果排序失敗，保持原順序
        pass
    
    # 優先保留系統日誌，然後限制總記錄數
    max_logs = 100
    if len(all_logs) > max_logs:
        # 分別提取系統日誌和其他日誌
        system_logs = [log for log in all_logs if log.get('method') == 'TX_LOG' or log.get('method') == 'BTC_LOG' or log.get('uri') == '/api/system_log' or log.get('uri') == 'system_log']
        other_logs = [log for log in all_logs if not (log.get('method') == 'TX_LOG' or log.get('method') == 'BTC_LOG' or log.get('uri') == '/api/system_log' or log.get('uri') == 'system_log')]
        
        # 限制其他日誌數量，優先保留系統日誌
        max_other_logs = max_logs - len(system_logs)
        if max_other_logs > 0:
            other_logs = other_logs[:max_other_logs]
        
        # 重新合併，保持時間排序
        all_logs = system_logs + other_logs
        # 重新按時間排序，確保合併後的日誌仍按時間順序
        all_logs.sort(key=lambda x: parse_timestamp_for_sort(x.get('timestamp', '')), reverse=False)
    
    return all_logs

@app.route('/')
def index():
    response = send_from_directory(app.static_folder, 'index.html')
    # 針對 WebView 設置不緩存頭
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@app.route('/api/save_env', methods=['POST'])
def save_env():
    data = request.get_json()
    # 讀取原始 env 模板，保留註解與順序
    with open(ENV_PATH, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    new_lines = []
    has_empty_required_fields = False
    
    # 必填欄位列表
    required_fields = ['CHAT_ID', 'API_KEY', 'SECRET_KEY', 'PERSON_ID', 'CA_PASSWD', 'CERT_START', 'CERT_END']
    
    for line in lines:
        m = re.match(r'^(\w+)=.*$', line)
        if m:
            key = m.group(1)
            # BOT_TOKEN不允許被覆蓋，保持原值
            if key == 'BOT_TOKEN':
                new_lines.append(line)
            # 處理其他欄位
            elif key in data:
                val = data.get(key, '').strip()
                new_lines.append(f'{key}={val}\n')
                # 檢查必填欄位是否為空
                if key in required_fields and not val:
                    has_empty_required_fields = True
            else:
                new_lines.append(f'{key}=\n')
        else:
            new_lines.append(line)
    
    # 如果有必填欄位為空，自動登出
    if has_empty_required_fields:
        update_login_status(0)
    
    with open(ENV_PATH, 'w', encoding='utf-8') as f:
        f.writelines(new_lines)
    
    return jsonify({'status': 'ok', 'has_empty_fields': has_empty_required_fields})

@app.route('/api/load_env', methods=['GET'])
def load_env():
    env_data = {}
    if os.path.exists(ENV_PATH):
        with open(ENV_PATH, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    # 載入所有欄位，包括空值
                    env_data[key] = value.strip()
    return jsonify(env_data)

@app.route('/api/upload/holiday', methods=['POST'])
def upload_holiday():
    try:
        if 'file' not in request.files:
            return jsonify({'error': '沒有檔案'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '沒有選擇檔案'}), 400
        
        if not file.filename.endswith('.csv'):
            return jsonify({'error': '只支援 CSV 檔案'}), 400
        
        # 確保目錄存在
        holiday_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'holiday')
        os.makedirs(holiday_dir, exist_ok=True)
        
        # 儲存檔案
        file_path = os.path.join(holiday_dir, file.filename)
        file.save(file_path)
        
        return jsonify({'status': 'success', 'message': '檔案上傳成功'})
    except Exception as e:
        return jsonify({'error': f'上傳失敗: {str(e)}'}), 500

@app.route('/api/upload/certificate', methods=['POST'])
def upload_certificate():
    try:
        if 'file' not in request.files:
            return jsonify({'error': '沒有檔案'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '沒有選擇檔案'}), 400
        
        if not file.filename.endswith('.pfx'):
            return jsonify({'error': '只支援 PFX 檔案'}), 400
        
        # 確保目錄存在
        cert_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'certificate')
        os.makedirs(cert_dir, exist_ok=True)
        
        # 儲存檔案
        file_path = os.path.join(cert_dir, file.filename)
        file.save(file_path)
        
        return jsonify({'status': 'success', 'message': '檔案上傳成功'})
    except Exception as e:
        return jsonify({'error': f'上傳失敗: {str(e)}'}), 500

@app.route('/api/uploaded_files', methods=['GET'])
def get_uploaded_files():
    try:
        holiday_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'holiday')
        cert_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'certificate')
        
        holiday_file = None
        cert_file = None
        
        # 檢查台股日曆檔案
        if os.path.exists(holiday_dir):
            csv_files = [f for f in os.listdir(holiday_dir) if f.endswith('.csv')]
            if csv_files:
                holiday_file = csv_files[0]  # 取第一個CSV檔案
        
        # 檢查憑證檔案
        if os.path.exists(cert_dir):
            cert_files = [f for f in os.listdir(cert_dir) if not f.endswith('.txt')]
            if cert_files:
                cert_file = cert_files[0]  # 取第一個檔案
        
        # 計算實際使用的憑證路徑（與登入時邏輯一致）
        ca_path = os.getenv('CA_PATH', '')
        actual_cert_path = None
        if ca_path:
            if os.path.isabs(ca_path):
                actual_cert_path = ca_path
            else:
                program_root = os.path.dirname(os.path.dirname(__file__))
                actual_cert_path = os.path.join(program_root, ca_path)
        
        return jsonify({
            'holiday_file': holiday_file,
            'certificate_file': cert_file,
            'cert_path_info': {
                'env_value': ca_path,
                'actual_path': actual_cert_path,
                'exists': os.path.exists(actual_cert_path) if actual_cert_path else False
            }
        })
    except Exception as e:
        return jsonify({'error': f'獲取檔案資訊失敗: {str(e)}'}), 500

@app.route('/api/bot_username', methods=['POST'])
def get_bot_username():
    try:
        # 從.env文件讀取token，如果沒有則使用硬編碼值
        token = None
        if os.path.exists(ENV_PATH):
            with open(ENV_PATH, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if line.startswith('BOT_TOKEN='):
                        token = line.split('=', 1)[1]
                        break
        
        # 如果.env中沒有token，使用硬編碼值
        if not token:
            token = '7202376519:AAF-i3MbuMEpz0W7nFE9KmieqVw7L5s0xK4'
        
        if not token:
            return jsonify({'username': None})
        
        # 呼叫 Telegram Bot API 獲取 Bot 資訊
        url = f'https://api.telegram.org/bot{token}/getMe'
        response = requests.get(url, timeout=10)
        
        if response.status_code == 200:
            bot_data = response.json()
            if bot_data.get('ok'):
                username = bot_data['result'].get('username', '')
                if username:
                    return jsonify({'username': f'@{username}'})
        
        return jsonify({'username': None})
    except Exception as e:
        return jsonify({'error': f'查詢失敗: {str(e)}'}), 500

@app.route('/api/send-telegram', methods=['POST'])
def api_send_telegram():
    """發送Telegram通知的API端點"""
    try:
        data = request.get_json()
        message = data.get('message', '')
        
        if not message:
            return jsonify({'success': False, 'error': '訊息內容不能為空'}), 400
        
        # 發送Telegram訊息
        success = send_telegram_message(message)
        
        if success:
            return jsonify({'success': True, 'message': '訊息發送成功'})
        else:
            return jsonify({'success': False, 'error': '訊息發送失敗'})
    
    except Exception as e:
        return jsonify({'success': False, 'error': f'發送失敗: {str(e)}'}), 500

@app.route('/api/login', methods=['POST'])
def api_login():
    # 將 LOGIN=1 寫入 .env
    update_login_status(1)
    
    # 在背景線程中啟動隧道服務，不阻塞主請求
    def start_tunnel_background():
        start_tunnel()
    
    create_managed_thread(target=start_tunnel_background, name="隧道啟動線程").start()
    
    # 同時登入永豐API
    def login_sinopac_background():
        login_sinopac()
    
    create_managed_thread(target=login_sinopac_background, name="永豐API登入線程").start()
    
    return jsonify({'status': 'ok'})

@app.route('/api/logout', methods=['POST'])
def api_logout():
    # 重置LOGIN狀態
    update_login_status(0)
    
    # 停止隧道服務
    stop_tunnel()
    
    # 登出永豐API
    logout_sinopac()
    
    return jsonify({'status': 'ok'})

@app.route('/api/reconnect', methods=['POST'])
def api_reconnect():
    """手動觸發API重連"""
    global is_reconnecting, reconnect_attempts
    
    try:
        logger.info("收到手動重連請求...")
        
        
        # 標記重連狀態
        is_reconnecting = True
        reconnect_attempts += 1
        
        # 執行重連
        if reconnect_api():
            is_reconnecting = False
            reconnect_attempts = 0
            return jsonify({
                'success': True, 
                'message': '手動重連成功！',
                'status': 'connected'
            })
        else:
            return jsonify({
                'success': False, 
                'message': '手動重連失敗，系統將繼續自動重試',
                'status': 'reconnecting'
            })
            
    except Exception as e:
        logger.error(f"手動重連API時發生錯誤: {e}")
        return jsonify({
            'success': False, 
            'message': f'手動重連時發生錯誤: {str(e)}',
            'status': 'error'
        })

@app.route('/api/health', methods=['GET'])
def api_health():
    """API健康檢查"""
    health_status = check_api_health()
    return jsonify({
        'healthy': health_status,
        'connected': sinopac_connected,
        'login_status': sinopac_login_status,
        'timestamp': datetime.now().isoformat()
    })

@app.route('/api/btc/reconnect', methods=['POST'])
def api_btc_reconnect():
    """手動觸發BTC API重連"""
    if BTC_MODULE_AVAILABLE:
        try:
            logger.info("收到BTC手動重連請求...")
            
            # 調用BTC重連功能
            success = btcmain.btc_reconnect_api()
            
            if success:
                return jsonify({
                    'success': True,
                    'message': 'BTC API重連成功'
                })
            else:
                return jsonify({
                    'success': False,
                    'message': 'BTC API重連失敗，請檢查網路連接和API配置'
                })
                
        except Exception as e:
            logger.error(f"BTC手動重連失敗: {e}")
            return jsonify({
                'success': False,
                'message': f'BTC重連異常: {str(e)}'
            })
    else:
        return jsonify({'success': False, 'message': 'BTC模組不可用'})

@app.route('/api/tunnel/start', methods=['POST'])
def api_start_tunnel():
    """啟動隧道服務 API"""
    success = start_tunnel()
    return jsonify({
        'success': success,
        'status': get_tunnel_status()
    })

@app.route('/api/ngrok/start', methods=['POST'])
def api_start_ngrok():
    """啟動ngrok API (保持兼容性)"""
    success = start_cloudflare_tunnel()
    return jsonify({
        'success': success,
        'status': get_cloudflare_tunnel_status()
    })

@app.route('/api/tunnel/stop', methods=['POST'])
def api_stop_tunnel():
    """停止隧道服務 API"""
    success = stop_tunnel()
    return jsonify({
        'success': success,
        'status': get_tunnel_status()
    })

@app.route('/api/ngrok/stop', methods=['POST'])
def api_stop_ngrok():
    """停止ngrok API (保持兼容性)"""
    success = stop_cloudflare_tunnel()
    return jsonify({
        'success': success,
        'status': get_cloudflare_tunnel_status()
    })

@app.route('/api/tunnel/status', methods=['GET'])
def api_tunnel_status():
    """獲取隧道狀態 API"""
    return jsonify(get_tunnel_status())

@app.route('/api/tunnel/<tunnel_type>/start', methods=['POST'])
def api_start_tunnel_by_type(tunnel_type):
    """啟動指定類型的隧道服務 API"""
    global tunnel_manager
    
    if not tunnel_manager:
        return jsonify({
            'success': False,
            'error': '隧道管理器未初始化'
        })
    
    try:
        success = tunnel_manager.start_tunnel(tunnel_type)
        return jsonify({
            'success': success,
            'status': tunnel_manager.get_tunnel_status(tunnel_type)
        })
    except ValueError as e:
        return jsonify({
            'success': False,
            'error': str(e)
        })

@app.route('/api/tunnel/<tunnel_type>/stop', methods=['POST'])
def api_stop_tunnel_by_type(tunnel_type):
    """停止指定類型的隧道服務 API"""
    global tunnel_manager
    
    if not tunnel_manager:
        return jsonify({
            'success': False,
            'error': '隧道管理器未初始化'
        })
    
    try:
        success = tunnel_manager.stop_tunnel(tunnel_type)
        return jsonify({
            'success': success,
            'status': tunnel_manager.get_tunnel_status(tunnel_type)
        })
    except ValueError as e:
        return jsonify({
            'success': False,
            'error': str(e)
        })

@app.route('/api/tunnel/<tunnel_type>/requests', methods=['GET'])
def api_tunnel_requests_by_type(tunnel_type):
    """獲取指定類型隧道的請求日誌 API"""
    global tunnel_manager
    
    if not tunnel_manager:
        return jsonify([])
    
    try:
        tunnel = tunnel_manager.get_tunnel(tunnel_type)
        if tunnel:
            requests = tunnel.get_request_logs()
            return jsonify(requests)
        else:
            return jsonify([])
    except Exception as e:
        print_console("TUNNEL", "ERROR", f"獲取{tunnel_type}隧道請求日誌失敗", str(e))
        return jsonify([])

@app.route('/api/tunnel/<tunnel_type>/status', methods=['GET'])
def api_tunnel_status_by_type(tunnel_type):
    """獲取指定類型隧道狀態 API"""
    global tunnel_manager
    
    if not tunnel_manager:
        return jsonify({
            'status': 'stopped',
            'error': '隧道管理器未初始化'
        })
    
    try:
        return jsonify(tunnel_manager.get_tunnel_status(tunnel_type))
    except ValueError as e:
        return jsonify({
            'status': 'error',
            'error': str(e)
        })

@app.route('/api/tunnels/status', methods=['GET'])
def api_all_tunnels_status():
    """獲取所有隧道狀態 API"""
    global tunnel_manager
    
    if not tunnel_manager:
        return jsonify({
            'tx': {'status': 'stopped', 'error': '隧道管理器未初始化'},
            'btc': {'status': 'stopped', 'error': '隧道管理器未初始化'}
        })
    
    return jsonify(tunnel_manager.get_all_status())

@app.route('/api/tunnel/<tunnel_type>/health', methods=['GET'])
def api_tunnel_health_report(tunnel_type):
    """獲取隧道健康報告 API（量化交易專用）"""
    global tunnel_manager
    
    if not tunnel_manager:
        return jsonify({
            'error': '隧道管理器未初始化',
            'tunnel_type': tunnel_type
        })
    
    # 檢查是否為量化交易專用管理器
    if hasattr(tunnel_manager, 'get_tunnel_health_report'):
        health_report = tunnel_manager.get_tunnel_health_report(tunnel_type)
        if health_report:
            return jsonify(health_report)
    
    return jsonify({
        'error': f'無法獲取{tunnel_type}隧道健康報告',
        'tunnel_type': tunnel_type
    })

@app.route('/api/tunnels/health/summary', methods=['GET'])
def api_tunnels_health_summary():
    """獲取所有隧道健康狀況總覽（量化交易監控面板）"""
    global tunnel_manager
    
    if not tunnel_manager or not hasattr(tunnel_manager, 'get_tunnel_health_report'):
        return jsonify({
            'error': '量化交易隧道管理器未初始化',
            'overall_status': 'unknown'
        })
    
    try:
        summary = {
            'overall_status': 'healthy',
            'tunnels': {},
            'total_uptime_hours': 0,
            'avg_availability': 0,
            'last_updated': time.time()
        }
        
        tunnel_types = ['tx', 'btc']
        availability_scores = []
        total_uptime = 0
        
        for tunnel_type in tunnel_types:
            health_report = tunnel_manager.get_tunnel_health_report(tunnel_type)
            if health_report:
                summary['tunnels'][tunnel_type] = health_report
                availability_scores.append(health_report['availability_percentage'])
                total_uptime += health_report['uptime_hours']
                
                # 判斷整體狀態
                if health_report['availability_percentage'] < 95 or health_report['status'] != 'running':
                    summary['overall_status'] = 'degraded'
                elif health_report['availability_percentage'] < 90:
                    summary['overall_status'] = 'unhealthy'
        
        # 計算平均值
        if availability_scores:
            summary['avg_availability'] = round(sum(availability_scores) / len(availability_scores), 2)
        summary['total_uptime_hours'] = round(total_uptime, 2)
        
        return jsonify(summary)
        
    except Exception as e:
        return jsonify({
            'error': f'獲取健康狀況總覽失敗: {str(e)}',
            'overall_status': 'error'
        })

@app.route('/api/health', methods=['GET'])
def api_system_health():
    """系統健康檢查端點（供隧道監控使用）"""
    try:
        # 檢查Flask應用狀態
        flask_status = 'healthy'
        
        # 檢查重要服務狀態
        services_status = {
            'flask': 'running',
            'tunnel_manager': 'running' if tunnel_manager else 'stopped',
            'btc_module': 'available' if BTC_MODULE_AVAILABLE else 'unavailable'
        }
        
        return jsonify({
            'status': flask_status,
            'services': services_status,
            'timestamp': time.time(),
            'version': CURRENT_VERSION
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': str(e),
            'timestamp': time.time()
        }), 500

@app.route('/api/tunnel/<tunnel_type>/persistent_url', methods=['GET'])
def api_get_persistent_url(tunnel_type):
    """獲取固定URL（量化交易專用）"""
    global tunnel_manager
    
    if not tunnel_manager:
        return jsonify({
            'error': '隧道管理器未初始化',
            'tunnel_type': tunnel_type
        })
    
    # 檢查是否為持久化隧道管理器
    if hasattr(tunnel_manager, 'get_persistent_url'):
        persistent_url = tunnel_manager.get_persistent_url(tunnel_type)
        if persistent_url:
            return jsonify({
                'tunnel_type': tunnel_type,
                'persistent_url': persistent_url,
                'url_fixed': True,
                'message': '此URL在重連後不會改變，適合TradingView等外部服務使用'
            })
    
    # 降級到普通隧道URL
    tunnel = tunnel_manager.get_tunnel(tunnel_type)
    if tunnel:
        status = tunnel.get_status()
        return jsonify({
            'tunnel_type': tunnel_type,
            'persistent_url': status.get('url'),
            'url_fixed': False,
            'warning': '此為臨時URL，重連後會改變'
        })
    
    return jsonify({
        'error': f'{tunnel_type}隧道未啟動',
        'tunnel_type': tunnel_type
    })

def _get_user_ids():
    """從API獲取真實的用戶帳戶ID"""
    try:
        user_ids = {}
        
        # 獲取TX用戶真實期貨帳戶ID
        try:
            global sinopac_connected, sinopac_account
            if sinopac_connected and sinopac_account and sinopac_account not in ["無期貨帳戶", "帳戶設定失敗"]:
                user_ids['tx'] = str(sinopac_account)
                logger.info(f"TX用戶ID (永豐期貨帳戶): {sinopac_account}")
        except Exception as e:
            logger.warning(f"獲取TX帳戶ID失敗: {e}")
        
        # 獲取BTC用戶ID（從API和配置文件雙重獲取）
        try:
            if BTC_MODULE_AVAILABLE:
                # 方法1：從配置文件獲取BINANCE_USER_ID（用戶填入的UID）
                btc_config_path = os.path.join(CONFIG_DIR, 'btc.env')
                if os.path.exists(btc_config_path):
                    import dotenv
                    btc_config = dotenv.dotenv_values(btc_config_path)
                    if btc_config.get('BINANCE_USER_ID'):
                        user_ids['btc'] = str(btc_config.get('BINANCE_USER_ID'))
                        logger.info(f"BTC用戶ID (配置文件): {btc_config.get('BINANCE_USER_ID')}")
                        
                        # 可選：驗證這個UID是否與當前登入的帳戶匹配
                        # 這裡可以添加API驗證邏輯
        except Exception as e:
            logger.warning(f"獲取BTC帳戶ID失敗: {e}")
        
        return user_ids
    except Exception as e:
        logger.error(f"獲取用戶ID失敗: {e}")
        return {}

@app.route('/api/tunnels/webhook_urls', methods=['GET'])
def api_get_webhook_urls():
    """獲取所有隧道的Webhook URL（TradingView等外部服務專用）- 支持多用戶架構"""
    global tunnel_manager
    
    if not tunnel_manager:
        return jsonify({
            'error': '隧道管理器未初始化',
            'webhook_urls': {}
        })
    
    webhook_urls = {}
    user_ids = _get_user_ids()  # 獲取用戶ID
    
    try:
        for tunnel_type in ['tx', 'btc']:
            user_id = user_ids.get(tunnel_type)
            
            # 優先獲取固定URL
            if hasattr(tunnel_manager, 'get_persistent_url'):
                persistent_url = tunnel_manager.get_persistent_url(tunnel_type)
                if persistent_url:
                    # 只提供真實用戶ID的URL格式，不提供備用格式
                    if user_id:
                        webhook_url = f"{persistent_url}/{user_id}"
                        api_webhook_url = f"{persistent_url}/{user_id}"  # 統一為簡潔格式
                    else:
                        # 沒有用戶ID時不提供URL
                        continue
                    
                    webhook_urls[tunnel_type] = {
                        'base_url': persistent_url,
                        'webhook_url': webhook_url,
                        'api_webhook_url': api_webhook_url,
                        'user_id': user_id,
                        'url_fixed': True,
                        'status': 'persistent'
                    }
                    continue
            
            # 降級到普通隧道
            tunnel = tunnel_manager.get_tunnel(tunnel_type)
            if tunnel:
                status = tunnel.get_status()
                base_url = status.get('url')
                if base_url:
                    # 只提供真實用戶ID的URL格式，不提供備用格式
                    if user_id:
                        webhook_url = f"{base_url}/{user_id}"
                        api_webhook_url = f"{base_url}/{user_id}"  # 統一為簡潔格式
                    else:
                        # 沒有用戶ID時跳過此tunnel
                        continue
                    
                    webhook_urls[tunnel_type] = {
                        'base_url': base_url,
                        'webhook_url': webhook_url,
                        'api_webhook_url': api_webhook_url,
                        'user_id': user_id,
                        'url_fixed': False,
                        'status': tunnel.status,
                        'warning': '臨時URL，重連後會改變'
                    }
        
        return jsonify({
            'webhook_urls': webhook_urls,
            'last_updated': time.time(),
            'note': '固定URL適合配置在TradingView等外部服務中'
        })
        
    except Exception as e:
        return jsonify({
            'error': f'獲取Webhook URL失敗: {str(e)}',
            'webhook_urls': {}
        })

@app.route('/api/tunnels/url_history', methods=['GET'])
def api_get_tunnel_url_history():
    """獲取隧道URL變化歷史"""
    try:
        history_file = os.path.join(os.path.dirname(__file__), 'tunnel_url_history.json')
        
        if os.path.exists(history_file):
            with open(history_file, 'r', encoding='utf-8') as f:
                history = json.load(f)
            
            return jsonify({
                'success': True,
                'url_changes': history,
                'total_changes': len(history),
                'last_updated': history[-1]['timestamp'] if history else None
            })
        else:
            return jsonify({
                'success': True,
                'url_changes': [],
                'total_changes': 0,
                'message': '尚無URL變化記錄'
            })
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'獲取URL變化歷史失敗: {str(e)}',
            'url_changes': []
        })

@app.route('/api/tunnels/current_urls', methods=['GET'])
def api_get_current_tunnel_urls():
    """獲取當前所有隧道URL（TX和BTC分別顯示）- 支持多用戶架構"""
    global tunnel_manager
    
    if not tunnel_manager:
        return jsonify({
            'error': '隧道管理器未初始化',
            'urls': {}
        })
    
    try:
        current_urls = {}
        user_ids = _get_user_ids()  # 獲取用戶ID
        
        for tunnel_type in ['tx', 'btc']:
            user_id = user_ids.get(tunnel_type)
            tunnel = tunnel_manager.get_tunnel(tunnel_type)
            if tunnel:
                status = tunnel.get_status()
                if status.get('url'):
                    # 只提供真實用戶ID的URL格式，不提供備用格式
                    if user_id:
                        webhook_url = f"{status['url']}/{user_id}"
                        api_webhook_url = f"{status['url']}/{user_id}"  # 統一為簡潔格式
                    else:
                        # 沒有用戶ID時跳過
                        continue
                    
                    current_urls[tunnel_type] = {
                        'url': status['url'],
                        'status': status['status'],
                        'webhook_url': webhook_url,
                        'api_webhook_url': api_webhook_url,
                        'user_id': user_id,
                        'port': status.get('port', 5000),
                        'last_updated': time.time()
                    }
                else:
                    current_urls[tunnel_type] = {
                        'url': None,
                        'status': status.get('status', 'stopped'),
                        'message': f'{tunnel_type.upper()}隧道未啟動或URL未生成'
                    }
            else:
                current_urls[tunnel_type] = {
                    'url': None,
                    'status': 'not_created',
                    'message': f'{tunnel_type.upper()}隧道未創建'
                }
        
        return jsonify({
            'success': True,
            'current_urls': current_urls,
            'note': '這些URL在隧道重連後可能會改變，系統會自動發送Telegram通知',
            'timestamp': time.time()
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'獲取當前隧道URL失敗: {str(e)}',
            'current_urls': {}
        })

# ========================== BTC 相關 API ==========================


@app.route('/api/btc/login', methods=['POST'])
def api_btc_login():
    """BTC帳戶登入/連接"""
    if BTC_MODULE_AVAILABLE:
        # 在背景線程中啟動隧道服務
        def start_tunnel_background():
            start_tunnel()
        
        create_managed_thread(target=start_tunnel_background, name="BTC隧道啟動線程").start()
        
        # BTC帳戶登入
        return btcmain.btc_login()
    else:
        return jsonify({'success': False, 'message': 'BTC模組不可用'})

@app.route('/api/btc/logout', methods=['POST'])
def api_btc_logout():
    """BTC帳戶登出"""
    if BTC_MODULE_AVAILABLE:
        # 停止隧道服務
        stop_tunnel()
        
        # BTC帳戶登出
        return btcmain.btc_logout()
    else:
        return jsonify({'success': False, 'message': 'BTC模組不可用'})

@app.route('/api/btc/save_env', methods=['POST'])
def api_btc_save_env():
    """保存BTC環境變量"""
    if BTC_MODULE_AVAILABLE:
        return btcmain.save_btc_env()
    else:
        return jsonify({'success': False, 'message': 'BTC模組不可用'})

@app.route('/api/btc/get_config', methods=['GET'])
def api_btc_get_config():
    """獲取BTC配置"""
    if BTC_MODULE_AVAILABLE:
        return btcmain.get_btc_config()
    else:
        return jsonify({'success': False, 'message': 'BTC模組不可用'})

@app.route('/api/btc/account_info', methods=['GET'])
def api_btc_account_info():
    """獲取BTC帳戶資訊"""
    if BTC_MODULE_AVAILABLE:
        return btcmain.get_btc_account_info()
    else:
        return jsonify({'success': False, 'message': 'BTC模組不可用'})

@app.route('/api/btc/positions', methods=['GET'])
def api_btc_positions():
    """獲取BTC持倉資訊"""
    if BTC_MODULE_AVAILABLE:
        return btcmain.get_btc_positions()
    else:
        return jsonify({'success': False, 'message': 'BTC模組不可用'})

@app.route('/api/btc/startup_notification', methods=['POST'])
def api_btc_startup_notification():
    """發送BTC啟動通知"""
    if not BTC_MODULE_AVAILABLE:
        return jsonify({'success': False, 'message': 'BTC模組不可用'})
    
    try:
        # 調用btcmain模組的統一通知函數
        btcmain.send_btc_daily_startup_notification()
        
        return jsonify({
            'success': True,
            'message': '啟動通知發送成功',
            'telegram_sent': True
        })
        
    except Exception as e:
        logger.error(f"BTC啟動通知失敗: {e}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/btc/trading_statistics', methods=['POST'])
def api_btc_trading_statistics():
    """發送BTC交易統計通知 - 調用btcmain模組統一處理"""
    if not BTC_MODULE_AVAILABLE:
        return jsonify({'success': False, 'message': 'BTC模組不可用'})
    
    try:
        # 調用btcmain模組的統一交易統計函數
        btcmain.send_btc_trading_statistics()
        
        return jsonify({
            'success': True,
            'message': '交易統計通知發送成功',
            'telegram_sent': True
        })
        
    except Exception as e:
        logger.error(f"發送BTC交易統計通知失敗: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        })

@app.route('/api/btc/generate_daily_report', methods=['POST'])
def api_btc_generate_daily_report():
    """生成BTC日報Excel文件"""
    if not BTC_MODULE_AVAILABLE:
        return jsonify({'success': False, 'message': 'BTC模組不可用'})
    
    try:
        # 獲取請求中的日期（可選）
        data = request.get_json() or {}
        date_str = data.get('date')
        
        # 生成日報
        result = btcmain.generate_btc_daily_report(date_str)
        
        if result['success']:
            print_console("REPORT", "SUCCESS", f"BTC日報生成成功: {result['filename']}")
            return jsonify({
                'success': True,
                'message': 'BTC日報生成成功',
                'filename': result['filename'],
                'filepath': result['filepath'],
                'date': result['date']
            })
        else:
            return jsonify(result)
        
    except Exception as e:
        logger.error(f"生成BTC日報失敗: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        })

@app.route('/api/btc/generate_monthly_report', methods=['POST'])
def api_btc_generate_monthly_report():
    """生成BTC月報Excel文件"""
    if not BTC_MODULE_AVAILABLE:
        return jsonify({'success': False, 'message': 'BTC模組不可用'})
    
    try:
        # 獲取請求中的年份和月份
        data = request.get_json() or {}
        year = data.get('year')
        month = data.get('month')
        
        # 如果沒有指定年月，使用當前年月
        if not year or not month:
            from datetime import datetime
            now = datetime.now()
            year = year or now.year
            month = month or now.month
        
        # 生成月報
        result = btcmain.generate_btc_monthly_report(int(year), int(month))
        
        if result['success']:
            print_console("REPORT", "SUCCESS", f"BTC月報生成成功: {result['filename']}")
            return jsonify({
                'success': True,
                'message': 'BTC月報生成成功',
                'filename': result['filename'],
                'filepath': result['filepath'],
                'year': result['year'],
                'month': result['month']
            })
        else:
            return jsonify(result)
        
    except Exception as e:
        logger.error(f"生成BTC月報失敗: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        })


@app.route('/api/btc/manual_order', methods=['POST'])
def api_btc_manual_order():
    """BTC手動下單接口"""
    logger.info(f"🚀 收到BTC手動下單請求")
    
    if not BTC_MODULE_AVAILABLE:
        logger.error(f"❌ BTC模組不可用")
        return jsonify({'success': False, 'message': 'BTC模組不可用'})
    
    try:
        order_data = request.get_json()
        logger.info(f"📋 BTC下單數據: {order_data}")
        
        if not order_data:
            logger.error(f"❌ 缺少訂單數據")
            return jsonify({'success': False, 'error': '缺少訂單數據'})
        
        quantity = order_data.get('quantity')
        action = order_data.get('action')  # new, cover
        side = order_data.get('side')      # buy, sell
        order_type = order_data.get('order_type', 'MARKET')
        
        if not all([quantity, action, side]):
            logger.error(f"❌ 缺少必要參數: quantity={quantity}, action={action}, side={side}")
            return jsonify({'success': False, 'error': '缺少必要的交易參數'})
        
        logger.info(f"📞 調用btcmain.place_btc_futures_order: quantity={quantity}, action={action}, side={side}, order_type={order_type}")
        
        # 執行手動下單
        result = btcmain.place_btc_futures_order(
            symbol='BTCUSDT',
            side=side,
            quantity=quantity,
            order_type=order_type,
            is_manual=True
        )
        
        logger.info(f"📋 btc_place_order結果: {result}")
        
        # 記錄手動下單日誌
        success = result.get('success', False) if isinstance(result, dict) else True
        status_code = 200 if success else 500
        
        add_custom_request_log('POST', '/api/btc/manual_order', status_code, {
            'reason': 'BTC手動下單' + ('成功' if success else '失敗'),
            'quantity': quantity,
            'action': action,
            'side': side,
            'order_type': order_type,
            'system': 'BTC',
            'is_manual': True
        })
        
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"BTC手動下單失敗: {e}")
        
        # 記錄異常日誌
        add_custom_request_log('POST', '/api/btc/manual_order', 500, {
            'reason': f'BTC手動下單異常: {str(e)}',
            'system': 'BTC',
            'is_manual': True,
            'error': str(e)
        })
        
        return jsonify({
            'success': False,
            'error': str(e)
        })


@app.route('/api/btc_bot_username', methods=['GET'])
def api_get_btc_bot_username():
    try:
        if BTC_MODULE_AVAILABLE:
            return btcmain.get_btc_bot_username()
        else:
            # 如果BTC模組不可用，直接在這裡處理
            import os
            btc_env_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config', 'btc.env')
            bot_token = None
            
            if os.path.exists(btc_env_path):
                with open(btc_env_path, 'r', encoding='utf-8') as f:
                    for line in f:
                        line = line.strip()
                        if line.startswith('BOT_TOKEN_BTC='):
                            bot_token = line.split('=', 1)[1]
                            break
            
            if not bot_token:
                bot_token = "7912873826:AAFPPDwuwspKVdyDGwh1oVqxH6u1gQ_N-jU"
            
            if bot_token:
                import requests
                try:
                    response = requests.get(f"https://api.telegram.org/bot{bot_token}/getMe", timeout=10)
                    if response.status_code == 200:
                        bot_info = response.json()
                        if bot_info.get('ok'):
                            username = bot_info['result'].get('username', 'Auto91_BtcBot')
                            return jsonify({'username': f'@{username}'})
                except:
                    pass
            
            return jsonify({'username': '@Auto91_BtcBot'})
    except Exception as e:
        return jsonify({'username': '@Auto91_BtcBot'})

@app.route('/api/save_btc_env', methods=['POST'])
def api_save_btc_env():
    """保存BTC環境變量 (前端調用路由)"""
    try:
        if BTC_MODULE_AVAILABLE:
            return btcmain.save_btc_env()
        else:
            # 如果BTC模組不可用，直接在這裡處理保存
            import os
            from flask import request
            
            data = request.get_json()
            btc_env_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config', 'btc.env')
            
            # 檢查是否有空值欄位
            required_fields = ['CHAT_ID_BTC', 'BINANCE_API_KEY', 'BINANCE_SECRET_KEY', 'BINANCE_USER_ID', 'TRADING_PAIR', 'LEVERAGE', 'POSITION_SIZE', 'MARGIN_TYPE', 'CONTRACT_TYPE']
            has_empty_fields = False
            
            for field in required_fields:
                if not data.get(field, '').strip():
                    has_empty_fields = True
                    break
            
            # 讀取當前登入狀態
            current_login_status = '0'
            if os.path.exists(btc_env_path):
                try:
                    with open(btc_env_path, 'r', encoding='utf-8') as f:
                        for line in f:
                            line = line.strip()
                            if line.startswith('LOGIN_BTC='):
                                current_login_status = line.split('=', 1)[1]
                                break
                except Exception:
                    current_login_status = '0'
            
            # 如果有空欄位，強制登出狀態設為0，否則保持當前狀態
            final_login_status = '0' if has_empty_fields else current_login_status
            
            # 創建BTC環境文件內容
            btc_env_content = f"""# Telegram Bot
BOT_TOKEN_BTC=7912873826:AAFPPDwuwspKVdyDGwh1oVqxH6u1gQ_N-jU

# Telegram ID
CHAT_ID_BTC={data.get('CHAT_ID_BTC', '')}

# 幣安 API Key
BINANCE_API_KEY={data.get('BINANCE_API_KEY', '')}

# 幣安 Secret Key
BINANCE_SECRET_KEY={data.get('BINANCE_SECRET_KEY', '')}

# 幣安用戶ID
BINANCE_USER_ID={data.get('BINANCE_USER_ID', '')}

# 交易對
TRADING_PAIR={data.get('TRADING_PAIR', 'BTCUSDT')}

# 合約類型
CONTRACT_TYPE={data.get('CONTRACT_TYPE', 'PERPETUAL')}

# 槓桿倍數
LEVERAGE={data.get('LEVERAGE', '10')}

# 風險比例百分比
POSITION_SIZE={data.get('POSITION_SIZE', '20')}

# 保證金模式
MARGIN_TYPE={data.get('MARGIN_TYPE', 'CROSS')}

# 登入狀態
LOGIN_BTC={final_login_status}
"""
            
            # 確保配置目錄存在
            os.makedirs(os.path.dirname(btc_env_path), exist_ok=True)
            
            # 儲存到btc.env文件
            with open(btc_env_path, 'w', encoding='utf-8') as f:
                f.write(btc_env_content)
            
            return jsonify({
                'success': True, 
                'message': 'BTC配置儲存成功',
                'has_empty_fields': has_empty_fields
            })
    except Exception as e:
        return jsonify({'success': False, 'message': f'儲存失敗: {str(e)}'})

@app.route('/api/load_btc_env', methods=['GET'])
def api_load_btc_env():
    if BTC_MODULE_AVAILABLE:
        return btcmain.load_btc_env()
    else:
        return jsonify({'success': False, 'message': 'BTC模組不可用'})


# 統一webhook處理器 - 動態帳戶ID路由
@app.route('/<user_id>', methods=['POST'])
def unified_webhook(system=None, user_id=None):
    """統一webhook處理器，支持TX和BTC系統，支持多用戶架構（簡潔路由）"""
    
    # 🚨 調試：記錄所有參數
    logger.info(f"🔍 Webhook調試 - user_id: {user_id}, system: {system}")
    
    # 當直接使用 /<user_id> 路由時，需要判斷是TX還是BTC用戶
    if user_id and not system:
        # 自動判斷用戶類型
        user_ids = _get_user_ids()
        logger.info(f"🔍 用戶ID映射: {user_ids}")
        if user_id == user_ids.get('tx'):
            system = 'tx'
            logger.info(f"🔍 自動識別為TX系統: {user_id}")
        elif user_id == user_ids.get('btc'):
            system = 'btc'
            logger.info(f"🔍 自動識別為BTC系統: {user_id}")
        else:
            logger.warning(f"🚨 未知用戶ID: {user_id}")
            # 🔧 修復：使用內容識別作為fallback機制
            try:
                raw = request.data.decode('utf-8')
                if raw.strip():
                    data = json.loads(raw)
                    # 如果有TX特有字段，則假設為TX訊號
                    if 'direction' in data and any(d in data.get('direction', '') for d in ['開多', '開空', '平多', '平空']):
                        system = 'tx'
                        logger.info(f"🔍 通過內容識別為TX系統: {user_id}")
                    # 如果有BTC特有字段，則假設為BTC訊號  
                    elif 'symbol' in data or 'action' in data:
                        system = 'btc'
                        logger.info(f"🔍 通過內容識別為BTC系統: {user_id}")
                    else:
                        # 仍然無法識別，返回錯誤
                        return jsonify({'success': False, 'message': f'未知用戶ID且無法識別訊號類型: {user_id}'}), 404
                else:
                    return jsonify({'success': False, 'message': f'未知用戶ID且請求內容為空: {user_id}'}), 404
            except Exception as e:
                logger.error(f"內容識別失敗: {e}")
                return jsonify({'success': False, 'message': f'未知用戶ID: {user_id}'}), 404
    
    # 用戶驗證（如果提供了user_id）
    if user_id:
        # 根據系統類型進行驗證
        system_type = 'btc' if system == 'btc' else 'tx'
        
        # 檢查是否通過內容識別確定的系統類型（允許較寬鬆的驗證）
        user_ids = _get_user_ids()
        is_content_identified = user_id not in [user_ids.get('tx'), user_ids.get('btc')]
        
        if is_content_identified:
            # 通過內容識別的用戶，記錄警告但允許通過
            logger.warning(f"⚠️ {system_type.upper()} 系統未登入，通過內容識別處理用戶 {user_id} 的請求")
        else:
            # 正常的用戶ID映射，進行嚴格驗證
            if not _verify_user_id(user_id, system_type):
                logger.warning(f"{system_type.upper()} Webhook 用戶驗證失敗: user_id={user_id}")
                return jsonify({'success': False, 'message': '用戶驗證失敗'}), 403
        
        logger.info(f"{system_type.upper()} Webhook 收到來自用戶 {user_id} 的請求")
    
    # 🚨 調試：記錄最終系統類型
    logger.info(f"🔍 最終系統類型: {system}")
    
    # 如果明確指定BTC系統
    if system == 'btc':
        logger.info(f"🔍 執行BTC專用路徑")
        # 解析請求數據
        try:
            raw = request.data.decode('utf-8')
            data = json.loads(raw) if raw.strip() else {}
            action = data.get('action', '未知')
        except:
            add_custom_request_log('POST', '/webhook/btc', 400, {
                'reason': 'BTC訊號解析失敗',
                'system': 'BTC'
            })
            return jsonify({'success': False, 'message': '請求數據解析失敗'}), 400
        
        if BTC_MODULE_AVAILABLE:
            try:
                # 設置 Flask request 的數據，讓 btcmain.py 能正確處理
                from flask import g
                g.webhook_data = data
                # 🔧 不在這裡設置標記，讓 btc_webhook() 自己管理
                result = btcmain.btc_webhook()
                # 檢查返回結果判斷是否成功
                if hasattr(result, 'get_json'):
                    response_data = result.get_json()
                    success = response_data.get('success', True) if response_data else True
                else:
                    success = True  # 預設成功
                
                status_code = 200 if success else 500
                add_custom_request_log('POST', '/webhook/btc', status_code, {
                    'reason': 'BTC訊號處理成功' if success else 'BTC訊號處理失敗',
                    'action': action,
                    'system': 'BTC'
                })
                return result
            except Exception as e:
                add_custom_request_log('POST', '/webhook/btc', 500, {
                    'reason': f'BTC訊號處理異常: {str(e)[:50]}',
                    'action': action,
                    'system': 'BTC'
                })
                return jsonify({'success': False, 'message': f'處理失敗: {str(e)}'})
        else:
            add_custom_request_log('POST', '/webhook/btc', 503, {
                'reason': 'BTC模組不可用',
                'action': action,
                'system': 'BTC'
            })
            return jsonify({'success': False, 'message': 'BTC模組不可用'})
    
    # 如果明確指定TX系統
    elif system == 'tx':
        return tradingview_webhook_tx()
    
    # 不支援自動識別，已移除該功能
    elif system == 'auto':
        logger.error("❌ 不支援自動識別，請使用標準路由")
        return jsonify({'success': False, 'message': '不支援自動識別，請使用明確的TX或BTC路由'}), 400
    
    # 註釋的自動識別邏輯
    elif False:  # system == 'auto':
        try:
            # 嘗試解析請求數據
            raw = request.data.decode('utf-8')
            if not raw.strip():
                return jsonify({'success': False, 'message': '無效的請求數據'}), 400
                
            data = json.loads(raw)
            
            # 自動識別訊號類型
            if is_btc_signal(data):
                # 🚨 緊急修復：檢查是否已被處理，避免重複處理導致雙重下單
                from flask import g
                if hasattr(g, 'webhook_processed') and g.webhook_processed:
                    print_console("WEBHOOK", "INFO", "BTC訊號已被處理，跳過自動識別")
                    return jsonify({'success': True, 'message': '訊號已處理，跳過重複處理'})
                
                # 🚨 緊急修復：檢查是否來自直接BTC路由，避免重複處理
                if request.endpoint and 'btc' in request.endpoint:
                    print_console("WEBHOOK", "INFO", "檢測到直接BTC路由調用，跳過自動識別以避免重複處理")
                    return jsonify({'success': True, 'message': '直接BTC路由已處理，跳過自動識別'})
                
                print_console("WEBHOOK", "INFO", "自動識別為BTC訊號")
                action = data.get('action', '未知')
                
                if BTC_MODULE_AVAILABLE:
                    try:
                        # 設置 Flask request 的數據，讓 btcmain.py 能正確處理
                        g.webhook_data = data
                        # 🎯 企業級修復：實現正確的BTC交易邏輯
                        # 設置 Flask request 的數據，讓 btcmain.py 能正確處理
                        g.webhook_data = data
                        
                        # 🎯 核心邏輯：基於持倉狀況和訊號類型的智能處理
                        result = btcmain.btc_webhook()
                        
                        # 檢查返回結果判斷是否成功
                        if hasattr(result, 'get_json'):
                            response_data = result.get_json()
                            success = response_data.get('success', True) if response_data else True
                        else:
                            success = True  # 預設成功
                        
                        status_code = 200 if success else 500
                        add_custom_request_log('POST', f'/webhook/{user_id}', status_code, {
                            'reason': 'BTC訊號自動識別處理成功' if success else 'BTC訊號自動識別處理失敗',
                            'action': action,
                            'system': 'BTC'
                        })
                        return result
                    except Exception as e:
                        add_custom_request_log('POST', '/webhook/btc', 500, {
                            'reason': f'BTC訊號自動識別處理異常: {str(e)[:50]}',
                            'action': action,
                            'system': 'BTC'
                        })
                        return jsonify({'success': False, 'message': f'BTC訊號自動識別處理失敗: {str(e)}'})
                else:
                    add_custom_request_log('POST', '/webhook/btc', 503, {
                        'reason': 'BTC模組不可用（自動識別）',
                        'action': action,
                        'system': 'BTC'
                    })
                    return jsonify({'success': False, 'message': 'BTC模組不可用'})
            else:
                print_console("WEBHOOK", "INFO", "自動識別為TX訊號")
                # TX訊號的日誌記錄已經在 tradingview_webhook_tx() 函數中處理
                return tradingview_webhook_tx()
                
        except json.JSONDecodeError:
            return jsonify({'success': False, 'message': 'JSON格式錯誤'}), 400
        except Exception as e:
            return jsonify({'success': False, 'message': f'處理錯誤: {str(e)}'}), 500
    
    else:
        return jsonify({'success': False, 'message': '不支援的系統類型'}), 400

def is_btc_signal(data):
    """判斷是否為BTC訊號"""
    # BTC訊號特徵：包含symbol字段且action為BTC特有的動作
    btc_actions = ['LONG', 'SHORT', 'CLOSE', 'EXIT', 'CLOSE_LONG', 'CLOSE_SHORT', 'NEW', 'COVER']
    
    # 檢查是否有symbol字段（BTC訊號特有）
    has_symbol = 'symbol' in data
    
    # 檢查action是否為BTC特有的動作
    action = data.get('action', '').upper()
    is_btc_action = action in btc_actions
    
    # 檢查是否沒有TX特有的字段
    has_contract = 'contract' in data
    has_trade_id = 'tradeId' in data
    
    # BTC訊號：有symbol字段或者是BTC特有動作，且沒有TX特有字段
    return (has_symbol or is_btc_action) and not (has_contract and has_trade_id)

def tradingview_webhook_tx():
    """處理TX系統的webhook邏輯"""
    global has_processed_delivery_exit, active_trades, recent_signals, rollover_processed_signals
    
    client_ip = request.remote_addr
    # 暫時不進行IP限制，但記錄來源IP
    print_console("WEBHOOK", "INFO", f"客戶端IP: {client_ip}")
    
    try:
        raw = request.data.decode('utf-8')
        current_time = datetime.now().strftime('%Y/%m/%d %H:%M:%S')
        print_console("WEBHOOK", "INFO", f"[{current_time}] 收到TradingView Webhook請求 (TX)")
        print_console("WEBHOOK", "INFO", "原始數據", raw)
        
        # 檢查無效訊號
        if '{{strategy.order.alert_message}}' in raw or not raw.strip():
            print_console("WEBHOOK", "WARNING", "無效訊號")
            add_custom_request_log('POST', '/webhook/tx', 400, {
                'reason': 'TX無效訊號',
                'system': 'TX'
            })
            return '無效訊號', 400
            
        data = json.loads(raw)
        signal_id = data.get('tradeId')
        action = data.get('action', '')
        contract_code = data.get('contract', '')
        
        # 重複訊號檢查（優化功能）
        if is_duplicate_signal(signal_id, action, contract_code):
            print_console("WEBHOOK", "WARNING", f"忽略重複訊號: {signal_id}")
            add_custom_request_log('POST', '/webhook/tx', 200, {
                'reason': 'TX重複訊號已忽略',
                'signal_id': signal_id,
                'system': 'TX'
            })
            return '重複訊號已忽略', 200
        
        # 轉倉邏輯檢查
        if process_rollover_signal(data):
            print_console("TRADE", "INFO", f"轉倉模式: 處理訊號 {signal_id}")
            # 在轉倉模式下，強制使用次月合約
            data['rollover_mode'] = True
        
        # 重複訊號檢查
        with global_lock:
            print_console("WEBHOOK", "INFO", f"檢查重複訊號: recent_signals={recent_signals}")
            if signal_id in recent_signals:
                print_console("WEBHOOK", "WARNING", f"重複訊號 {signal_id}，忽略")
                add_custom_request_log('POST', '/webhook/tx', 400, {
                    'reason': 'TX重複訊號',
                    'signal_id': signal_id,
                    'system': 'TX'
                })
                return '重複訊號', 400
            recent_signals.add(signal_id)
            # 10秒後清除記錄
            timer_thread = create_managed_thread(target=lambda: (time.sleep(10), recent_signals.discard(signal_id)), name=f"信號清理定時器-{signal_id}")
            timer_thread.start()
        
        data['receive_time'] = datetime.now()
        
        # 記錄webhook活動時間，用於自動/手動判斷
        global last_webhook_time
        last_webhook_time = time.time()
        
        process_signal(data)
        
        signal_type = data.get('type', 'entry')
        direction = data.get('direction', '未知')
        
        if signal_type == 'entry':
            if direction == '開多':
                log_message = f"來自webhook開倉訊號：開多"
            elif direction == '開空':
                log_message = f"來自webhook開倉訊號：開空"
            else:
                log_message = f"來自webhook開倉訊號：{direction}"
        elif signal_type == 'exit':
            if direction == '平多':
                log_message = f"來自webhook平倉訊號：平多"
            elif direction == '平空':
                log_message = f"來自webhook平倉訊號：平空"
            else:
                log_message = f"來自webhook平倉訊號：{direction}"
        else:
            log_message = f"來自webhook訊號：{signal_type} - {direction}"
        
        # 後端日誌記錄
        logger.info(f"[WEBHOOK] {log_message}")
        
        # 添加TX webhook請求日誌記錄到前端請求日誌
        add_custom_request_log('POST', '/webhook/tx', 200, {
            'reason': 'TX訊號處理成功',
            'action': f"{signal_type} - {direction}",
            'signal_id': signal_id,
            'system': 'TX'
        })
        
        return 'OK', 200
        
    except Exception as e:
        error_msg = str(e)
        print_console("WEBHOOK", "ERROR", "TX Webhook 處理錯誤", error_msg)
        import traceback
        traceback.print_exc()
        
        # 嘗試使用統一格式，如果無法解析data就用簡單訊息
        try:
            send_unified_failure_message(data, f"TX Webhook解析錯誤：{error_msg[:100]}")
        except:
            send_telegram_message(f"❌ TX Webhook 錯誤：{error_msg[:100]}")
        
        # 添加TX webhook錯誤請求日誌記錄到前端請求日誌
        add_custom_request_log('POST', '/webhook/tx', 500, {
            'reason': f'TX訊號處理異常: {error_msg[:50]}',
            'system': 'TX'
        })
        
        return f'TX錯誤：{error_msg}', 500


@app.route('/api/btc/strategy/status', methods=['GET'])
def api_get_btc_strategy_status():
    if BTC_MODULE_AVAILABLE:
        return btcmain.get_btc_strategy_status()
    else:
        return jsonify({'success': False, 'message': 'BTC模組不可用'})

@app.route('/api/btc/account/balance', methods=['GET'])
def api_get_btc_account_balance():
    """獲取BTC帳戶餘額"""
    if BTC_MODULE_AVAILABLE:
        return btcmain.get_btc_account_balance()
    else:
        return jsonify({'success': False, 'message': 'BTC模組不可用'})

@app.route('/api/btc/position', methods=['GET'])
def api_get_btc_position():
    """獲取BTC持倉信息"""
    if BTC_MODULE_AVAILABLE:
        return btcmain.get_btc_position()
    else:
        return jsonify({'success': False, 'message': 'BTC模組不可用'})

@app.route('/api/btc/version', methods=['GET'])
def api_get_btc_version():
    """獲取幣安版本信息"""
    if BTC_MODULE_AVAILABLE:
        return btcmain.get_btc_version()
    else:
        return jsonify({'success': False, 'message': 'BTC模組不可用'})

@app.route('/api/btc/trading/status', methods=['GET'])
def api_get_btc_trading_status():
    """獲取BTC交易狀態"""
    if BTC_MODULE_AVAILABLE:
        return btcmain.get_btc_trading_status()
    else:
        return jsonify({'success': False, 'message': 'BTC模組不可用'})

@app.route('/api/btc/risk/status', methods=['GET'])
def api_get_btc_risk_status():
    """獲取BTC風險狀態"""
    if BTC_MODULE_AVAILABLE:
        return btcmain.get_btc_risk_status()
    else:
        return jsonify({'success': False, 'message': 'BTC模組不可用'})

@app.route('/api/btc/realtime', methods=['GET'])
def api_get_btc_realtime_data():
    """獲取BTC實時數據"""
    if BTC_MODULE_AVAILABLE:
        return btcmain.get_btc_realtime_data()
    else:
        return jsonify({'success': False, 'message': 'BTC模組不可用'})

@app.route('/api/btc/websocket/start', methods=['POST'])
def api_start_btc_websocket():
    """啟動BTC WebSocket"""
    if BTC_MODULE_AVAILABLE:
        try:
            result = btcmain.start_btc_websocket()
            return jsonify({
                'success': result,
                'message': 'WebSocket啟動成功' if result else 'WebSocket啟動失敗'
            })
        except Exception as e:
            return jsonify({'success': False, 'message': f'啟動失敗: {str(e)}'})
    else:
        return jsonify({'success': False, 'message': 'BTC模組不可用'})

@app.route('/api/btc/websocket/stop', methods=['POST'])
def api_stop_btc_websocket():
    """停止BTC WebSocket"""
    if BTC_MODULE_AVAILABLE:
        try:
            btcmain.stop_btc_websocket()
            return jsonify({'success': True, 'message': 'WebSocket已停止'})
        except Exception as e:
            return jsonify({'success': False, 'message': f'停止失敗: {str(e)}'})
    else:
        return jsonify({'success': False, 'message': 'BTC模組不可用'})

@app.route('/api/btc/statistics', methods=['POST'])
def api_send_btc_statistics():
    """發送BTC交易統計"""
    if BTC_MODULE_AVAILABLE:
        try:
            btcmain.send_btc_trading_statistics()
            return jsonify({'success': True, 'message': 'BTC交易統計已發送'})
        except Exception as e:
            return jsonify({'success': False, 'message': f'發送BTC統計失敗: {str(e)}'})
    else:
        return jsonify({'success': False, 'message': 'BTC模組不可用'})

@app.route('/api/btc/generate_report', methods=['POST'])
def api_generate_btc_report():
    """生成BTC交易報表"""
    if BTC_MODULE_AVAILABLE:
        try:
            result = btcmain.generate_btc_daily_report()
            if result:
                return jsonify({'success': True, 'message': 'BTC交易報表生成成功', 'filepath': result})
            else:
                return jsonify({'success': False, 'message': 'BTC交易報表生成失敗'})
        except Exception as e:
            return jsonify({'success': False, 'message': f'生成BTC日報失敗: {str(e)}'})
    else:
        return jsonify({'success': False, 'message': 'BTC模組不可用'})

@app.route('/api/btc/generate_monthly_report', methods=['POST'])
def api_generate_btc_monthly_report():
    """生成BTC交易報表"""
    if BTC_MODULE_AVAILABLE:
        try:
            year = datetime.now().year
            month = datetime.now().month
            result = btcmain.generate_btc_monthly_report(year, month)
            if result:
                return jsonify({'success': True, 'message': 'BTC交易報表生成成功', 'filepath': result})
            else:
                return jsonify({'success': False, 'message': 'BTC交易報表生成失敗'})
        except Exception as e:
            return jsonify({'success': False, 'message': f'生成BTC月報失敗: {str(e)}'})
    else:
        return jsonify({'success': False, 'message': 'BTC模組不可用'})


# ========== TX交易報表API ==========
@app.route('/api/trading_statistics', methods=['GET'])
def api_trading_statistics():
    """獲取TX交易統計數據"""
    try:
        # 獲取交易記錄（參考send_daily_trading_statistics的邏輯）
        trades = []
        today = datetime.now()
        
        for i in range(7):  # 讀取過去7天的記錄
            check_date = today - timedelta(days=i)
            date_str = check_date.strftime('%Y%m%d')
            trades_file = os.path.join(TX_DATA_DIR, f'TXtransdata_{date_str}.json')
            
            if os.path.exists(trades_file):
                try:
                    with open(trades_file, 'r', encoding='utf-8') as f:
                        daily_trades = json.load(f)
                        trades.extend(daily_trades)
                except Exception as e:
                    logger.error(f"讀取 {date_str} 交易記錄失敗: {e}")
        
        # 分析交易統計
        cover_trades, total_cover_quantity, contract_pnl = analyze_simple_trading_stats(trades)
        
        # 統計委託、取消、成交次數
        total_orders = 0
        total_cancels = 0
        total_trades = 0
        
        for trade in trades:
            trade_type = trade.get('type', '')
            if trade_type in ['submit', 'deal']:
                total_orders += 1
            if trade_type == 'deal':
                total_trades += 1
            elif trade_type in ['cancel', 'fail']:
                total_cancels += 1
        
        # 獲取帳戶狀態
        try:
            account_response = requests.get(f'http://127.0.0.1:{CURRENT_PORT}/api/account/status', timeout=5)
            account_data = account_response.json().get('data', {}) if account_response.status_code == 200 else {}
        except:
            account_data = {}
        
        # 獲取持倉狀態
        try:
            position_response = requests.get(f'http://127.0.0.1:{CURRENT_PORT}/api/position/status', timeout=5)
            position_data = position_response.json() if position_response.status_code == 200 else []
        except:
            position_data = []
        
        return jsonify({
            'success': True,
            'total_orders': total_orders,
            'total_cancels': total_cancels,
            'total_trades': total_trades,
            'total_cover_quantity': total_cover_quantity,
            'contract_pnl': contract_pnl,
            'account_data': account_data,
            'trades': cover_trades[:10],  # 只返回前10筆平倉交易
            'position_data': position_data[:10] if isinstance(position_data, list) else [],
            'total_trade_count': len(cover_trades),
            'total_position_count': len(position_data) if isinstance(position_data, list) else 0
        })
    except Exception as e:
        logger.error(f"獲取TX交易統計失敗: {e}")
        return jsonify({
            'success': False,
            'error': str(e),
            'message': '獲取交易統計數據失敗'
        })

@app.route('/api/generate_trading_report', methods=['POST'])
def api_generate_trading_report():
    """生成TX日報Excel文件"""
    try:
        data = request.get_json() or {}
        date_str = data.get('date')  # 可選的日期參數
        
        # 獲取交易記錄（參考send_daily_trading_statistics的邏輯）
        trades = []
        today = datetime.now()
        
        for i in range(7):  # 讀取過去7天的記錄
            check_date = today - timedelta(days=i)
            date_str_check = check_date.strftime('%Y%m%d')
            trades_file = os.path.join(TX_DATA_DIR, f'TXtransdata_{date_str_check}.json')
            
            if os.path.exists(trades_file):
                try:
                    with open(trades_file, 'r', encoding='utf-8') as f:
                        daily_trades = json.load(f)
                        trades.extend(daily_trades)
                except Exception as e:
                    logger.error(f"讀取 {date_str_check} 交易記錄失敗: {e}")
        
        # 分析交易統計
        cover_trades, total_cover_quantity, contract_pnl = analyze_simple_trading_stats(trades)
        
        # 計算統計數據
        total_orders = total_cancels = total_trades = 0
        for trade in trades:
            trade_type = trade.get('type', '')
            if trade_type in ['submit', 'deal']:
                total_orders += 1
            if trade_type == 'deal':
                total_trades += 1
            elif trade_type in ['cancel', 'fail']:
                total_cancels += 1
        
        # 獲取帳戶狀態
        try:
            account_response = requests.get(f'http://127.0.0.1:{CURRENT_PORT}/api/account/status', timeout=5)
            account_data = account_response.json().get('data', {}) if account_response.status_code == 200 else {}
        except:
            account_data = {}
        
        # 獲取持倉數據
        try:
            position_response = requests.get(f'http://127.0.0.1:{CURRENT_PORT}/api/position/status', timeout=5)
            if position_response.status_code == 200:
                position_result = position_response.json()
                position_data = {
                    'has_positions': position_result.get('has_positions', False),
                    'data': position_result.get('data', {})
                }
            else:
                position_data = {'has_positions': False, 'data': {}}
        except:
            position_data = {'has_positions': False, 'data': {}}
        
        # 生成報表
        result = generate_trading_report(
            trades=trades,
            account_data=account_data,
            position_data=position_data,
            cover_trades=cover_trades,
            total_orders=total_orders,
            total_cancels=total_cancels,
            total_trades=total_trades,
            total_cover_quantity=total_cover_quantity,
            contract_pnl=contract_pnl
        )
        
        if result and result.get('success'):
            return jsonify({
                'success': True,
                'message': 'TX日報生成成功',
                'file_path': result.get('file_path'),
                'filename': result.get('filename'),
                'statistics': {
                    'total_orders': total_orders,
                    'total_trades': total_trades,
                    'total_positions': len(position_data)
                }
            })
        else:
            return jsonify({
                'success': False,
                'error': result.get('error') if result else '未知錯誤',
                'message': 'TX日報生成失敗'
            })
            
    except Exception as e:
        logger.error(f"生成TX日報失敗: {e}")
        return jsonify({
            'success': False,
            'error': str(e),
            'message': 'API調用失敗'
        })

@app.route('/api/generate_monthly_trading_report', methods=['POST'])
def api_generate_monthly_trading_report():
    """生成TX月報Excel文件"""
    try:
        data = request.get_json() or {}
        year = data.get('year')
        month = data.get('month')
        
        # 如果沒有指定年月，使用當前年月
        if not year or not month:
            now = datetime.now()
            year = year or now.year
            month = month or now.month
        
        # 生成月報
        result = generate_monthly_trading_report()
        
        if result and result.get('success'):
            return jsonify({
                'success': True,
                'message': f'TX月報生成及發送成功 ({year}年{month}月)',
                'file_path': result.get('file_path'),
                'filename': result.get('filename'),
                'period': f'{year}-{month:02d}'
            })
        else:
            return jsonify({
                'success': False,
                'error': result.get('error') if result else '未知錯誤',
                'message': 'TX月報生成失敗'
            })
    except Exception as e:
        print(f"TX月報生成API錯誤: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e),
            'message': 'TX月報生成過程發生錯誤'
        })

@app.route('/api/diagnose_tx_opening_records', methods=['POST'])
def api_diagnose_tx_opening_records():
    """診斷TX開倉記錄搜尋功能API"""
    try:
        result = diagnose_tx_opening_records()
        if result:
            return jsonify({
                'success': True,
                'message': 'TX開倉記錄診斷完成，請查看日誌'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'TX開倉記錄診斷失敗'
            })
    except Exception as e:
        logger.error(f"API診斷TX開倉記錄失敗: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        })
            
    except Exception as e:
        logger.error(f"生成TX月報失敗: {e}")
        return jsonify({
            'success': False,
            'error': str(e),
            'message': 'API調用失敗'
        })

@app.route('/api/ngrok/status', methods=['GET'])
def api_ngrok_status():
    """獲取ngrok狀態 API (保持兼容性)"""
    return jsonify(get_cloudflare_tunnel_status())


@app.route('/api/ngrok/requests', methods=['GET'])
def api_ngrok_requests():
    """獲取請求記錄 API（結合自定義記錄和 Cloudflare Tunnel 記錄）"""
    try:
        requests_data = get_tunnel_requests()
        
        # 添加額外的統計信息
        latency_info = {}
        if tunnel_service:
            try:
                latency_info = tunnel_service.get_latency()
            except Exception as e:
                logger.error(f"獲取延遲信息失敗: {e}")
        
        
        return jsonify({
            'requests': requests_data,
            'total_count': len(requests_data),
            'latency_info': latency_info,
            'status': 'success'
        })
    except Exception as e:
        return jsonify({
            'error': f'獲取請求記錄失敗: {str(e)}',
            'status': 'error'
        }), 500


@app.route('/api/ngrok/setup', methods=['POST'])
def api_ngrok_setup():
    """設置ngrok authtoken並啟動"""
    try:
        data = request.get_json() or {}
        authtoken = data.get('authtoken', '').strip()
        
        mode = data.get('mode', 'temporary')
        
        # 根據模式處理
        if authtoken in ['workers-mode', 'temporary-mode']:
            pass
        elif not authtoken:
            return jsonify({
                'status': 'error',
                'message': '請提供有效的 Cloudflare Tunnel token'
            }), 400
        elif len(authtoken) < 20:
            return jsonify({
                'status': 'error',
                'message': 'Cloudflare Tunnel token 格式不正確'
            }), 400
        
        # 儲存 Cloudflare token（免費模式除外）
        if authtoken not in ['workers-mode', 'temporary-mode']:
            try:
                config_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config')
                if not os.path.exists(config_dir):
                    os.makedirs(config_dir)
                
                token_file = os.path.join(config_dir, 'token.txt')
                with open(token_file, 'w') as f:
                    f.write(authtoken)
            except Exception as e:
                return jsonify({
                    'status': 'error',
                    'message': f'儲存 token 失敗: {str(e)}'
                }), 500
        
        # 在背景執行設置
        def setup_tunnel_background():
            try:
                # 初始化並啟動 Cloudflare Tunnel
                init_tunnel_service(mode)
                success = start_cloudflare_tunnel()
                
                if success:
                    return True
                else:
                    logger.error("Cloudflare Tunnel 設置失敗！")
                    return False
                    
            except Exception as e:
                logger.error(f"設置 Cloudflare Tunnel 失敗: {e}")
                return False
        
        create_managed_thread(target=setup_tunnel_background, name="隧道設置背景線程").start()
        
        return jsonify({
            'status': 'success',
            'message': '正在設置 Cloudflare Tunnel，請稍候...'
        })
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'設置失敗: {str(e)}'
        }), 500


@app.route('/api/ngrok/token/save', methods=['POST'])
def api_ngrok_token_save():
    """保存ngrok authtoken到服務器端"""
    try:
        data = request.get_json() or {}
        authtoken = data.get('authtoken', '').strip()
        
        if not authtoken:
            return jsonify({
                'status': 'error',
                'message': '請提供authtoken'
            }), 400
        
        # 保存到配置目錄
        config_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config')
        os.makedirs(config_dir, exist_ok=True)
        
        token_file = os.path.join(config_dir, 'token.txt')
        with open(token_file, 'w', encoding='utf-8') as f:
            f.write(authtoken)
        
        return jsonify({
            'status': 'success',
            'message': 'Token已保存'
        })
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'保存失敗: {str(e)}'
        }), 500

@app.route('/api/ngrok/token/load', methods=['GET'])
def api_ngrok_token_load():
    """從服務器端載入 Cloudflare Tunnel token（替換原 ngrok token）"""
    try:
        config_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config')
        token_file = os.path.join(config_dir, 'token.txt')
        
        if os.path.exists(token_file):
            with open(token_file, 'r', encoding='utf-8') as f:
                authtoken = f.read().strip()
            
            if authtoken:
                return jsonify({
                    'status': 'success',
                    'authtoken': authtoken
                })
        
        return jsonify({
            'status': 'not_found',
            'message': '未找到保存的 Cloudflare Tunnel token'
        })
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'載入失敗: {str(e)}'
        }), 500

@app.route('/api/tunnel/setup', methods=['POST'])
def api_tunnel_setup():
    """設置隧道服務 token 並啟動"""
    global tunnel_type, tunnel_service, tunnel_manager
    
    try:
        data = request.get_json()
        token = data.get('token', '').strip()
        service_type = data.get('service_type', 'cloudflare')
        target_tunnel_type = data.get('tunnel_type', 'tx')  # tx 或 btc
        origin_cert = data.get('origin_certificate', '').strip()
        private_key = data.get('private_key', '').strip()
        mode = data.get('mode', 'temporary')  # temporary 或 custom
        
        # 只在自訂域名模式下檢查token
        if mode == 'custom' and not token:
            return jsonify({
                'success': False,
                'message': '請提供有效的 token'
            })
        
        # 檢查自訂域名模式是否提供憑證
        if mode == 'custom' and (not origin_cert or not private_key):
            return jsonify({
                'success': False,
                'message': '自訂域名模式需要提供 Origin Certificate 和 Private Key'
            })
        
        # 創建配置目錄
        config_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config')
        if not os.path.exists(config_dir):
            os.makedirs(config_dir)
        
        # 保存 token (統一格式 - 純共用模式)
        # 只在自訂域名模式或有token時才寫入token檔案
        if token:
            if service_type == "cloudflare":
                token_file = os.path.join(config_dir, 'token.txt')  # 統一token檔案
            else:
                token_file = os.path.join(config_dir, f'{target_tunnel_type}_ngrok_token.txt')
            
            with open(token_file, 'w') as f:
                f.write(token)
        
        # 設置隧道類型
        tunnel_type = service_type
        
        # 在背景線程中執行設置
        def setup_tunnel_background():
            try:
                if service_type == "cloudflare":
                    # 初始化 Cloudflare Tunnel (使用正確的模式)
                    init_tunnel_service(mode)
                    
                    # 使用隧道管理器來獲取或創建指定類型的隧道實例
                    if tunnel_manager:
                        tunnel_instance = tunnel_manager.get_tunnel(target_tunnel_type)
                        if not tunnel_instance:
                            tunnel_instance = tunnel_manager.create_tunnel(target_tunnel_type, mode)
                        
                        # 保存隧道配置（包含憑證）
                        if tunnel_instance:
                            tunnel_instance.save_tunnel_config(token, origin_cert, private_key)
                    
                    success = start_cloudflare_tunnel()
                else:
                    # 使用 Cloudflare Tunnel 設置
                    success = start_cloudflare_tunnel()
                
                if success:
                    pass
                else:
                    pass
            except Exception as e:
                logger.error(f"設置 {service_type} 失敗: {e}")
        
        create_managed_thread(target=setup_tunnel_background, name="隧道設置背景線程").start()
        
        return jsonify({
            'success': True,
            'message': f'正在設置 {target_tunnel_type.upper()} {service_type}，請稍候...'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'設置失敗: {str(e)}'
        })

@app.route('/api/tunnel/token/load', methods=['GET'])
def api_tunnel_token_load():
    """載入隧道服務 token"""
    try:
        config_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config')
        
        # 檢查 Cloudflare token
        cf_token_file = os.path.join(config_dir, 'token.txt')
        ngrok_token_file = os.path.join(config_dir, 'ngrok_token.txt')
        
        result = {
            'cloudflare_token': '',
            'ngrok_token': '',
            'current_service': tunnel_type
        }
        
        if os.path.exists(cf_token_file):
            with open(cf_token_file, 'r') as f:
                result['cloudflare_token'] = f.read().strip()
        
        if os.path.exists(ngrok_token_file):
            with open(ngrok_token_file, 'r') as f:
                result['ngrok_token'] = f.read().strip()
        
        return jsonify({
            'success': True,
            'data': result
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'載入 token 失敗: {str(e)}'
        })

@app.route('/api/tunnel/settings', methods=['GET', 'POST'])
def api_tunnel_settings():
    """量化交易系統隧道設定管理API"""
    global tunnel_type
    
    if request.method == 'GET':
        # 讀取設定
        try:
            config_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config')
            current_tunnel_type = request.args.get('tunnel_type', tunnel_type or 'tx')
            
            # 讀取設定檔案
            settings_file = os.path.join(config_dir, 'tunnel_settings.json')
            
            if os.path.exists(settings_file):
                with open(settings_file, 'r', encoding='utf-8') as f:
                    settings = json.load(f)
                
                # 檢查是否有已保存的共用數據
                has_saved_data = False
                if settings.get('domain_mode') == 'custom':
                    token_file = os.path.join(config_dir, 'token.txt')
                    cert_file = os.path.join(config_dir, 'cert.pem')
                    key_file = os.path.join(config_dir, 'key.pem')
                    
                    has_saved_data = (os.path.exists(token_file) and 
                                    os.path.exists(cert_file) and 
                                    os.path.exists(key_file))
                    
                    # 載入保存的數據
                    if has_saved_data:
                        try:
                            if os.path.exists(token_file):
                                with open(token_file, 'r', encoding='utf-8') as f:
                                    content = f.read().strip()
                                    if content:
                                        settings['token'] = content
                        except Exception as e:
                            logger.warning(f"讀取token失敗: {e}")
                        
                        try:
                            if os.path.exists(cert_file):
                                with open(cert_file, 'r', encoding='utf-8') as f:
                                    content = f.read().strip()
                                    if content:
                                        settings['origin_cert'] = content
                        except Exception as e:
                            logger.warning(f"讀取cert失敗: {e}")
                        
                        try:
                            if os.path.exists(key_file):
                                with open(key_file, 'r', encoding='utf-8') as f:
                                    content = f.read().strip()
                                    if content:
                                        settings['private_key'] = content
                        except Exception as e:
                            logger.warning(f"讀取key失敗: {e}")
                
                settings['has_saved_data'] = has_saved_data
                
                return jsonify({
                    'success': True,
                    'settings': settings
                })
            else:
                # 預設設定
                return jsonify({
                    'success': True,
                    'settings': {
                        'domain_mode': 'temporary',
                        'tunnel_type': current_tunnel_type,
                        'has_saved_data': False
                    }
                })
                
        except Exception as e:
            logger.error(f"讀取隧道設定失敗: {e}")
            return jsonify({
                'success': False,
                'message': f'讀取設定失敗: {str(e)}'
            })
    
    elif request.method == 'POST':
        # 保存設定
        try:
            data = request.get_json()
            domain_mode = data.get('domain_mode', 'temporary')
            current_tunnel_type = data.get('tunnel_type', tunnel_type or 'tx')
            
            config_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config')
            os.makedirs(config_dir, exist_ok=True)
            
            # 準備設定數據
            settings = {
                'domain_mode': domain_mode,
                'tunnel_type': current_tunnel_type,
                'last_updated': datetime.now().isoformat()
            }
            
            # 只在自訂域名模式時才保存憑證檔案
            if domain_mode == 'custom':
                token = data.get('token', '').strip()
                origin_cert = data.get('origin_cert', '').strip()
                private_key = data.get('private_key', '').strip()
                
                # 檢查現有檔案
                token_file = os.path.join(config_dir, 'token.txt')
                cert_file = os.path.join(config_dir, 'cert.pem')
                key_file = os.path.join(config_dir, 'key.pem')
                
                # 保存到統一的共用檔案（只在有內容時保存，避免覆蓋現有檔案）
                logger.info(f"檢查保存資料 - Token長度: {len(token) if token else 0}, Cert長度: {len(origin_cert) if origin_cert else 0}, Key長度: {len(private_key) if private_key else 0}")
                
                if token and len(token) > 50:  # 確保token有實際內容且不是placeholder
                    with open(token_file, 'w', encoding='utf-8') as f:
                        f.write(token)
                    logger.info(f"Token已保存到統一檔案，長度: {len(token)}")
                elif not token:
                    logger.info("Token為空，不覆蓋現有檔案")
                
                if origin_cert and len(origin_cert) > 50:  # 確保憑證有實際內容
                    with open(cert_file, 'w', encoding='utf-8') as f:
                        f.write(origin_cert)
                    logger.info(f"Origin Certificate已保存到統一檔案，長度: {len(origin_cert)}")
                elif not origin_cert:
                    logger.info("Certificate為空，不覆蓋現有檔案")
                
                if private_key and len(private_key) > 50:  # 確保私鑰有實際內容
                    with open(key_file, 'w', encoding='utf-8') as f:
                        f.write(private_key)
                    os.chmod(key_file, 0o600)  # 設置私鑰權限
                    logger.info("Private Key已保存到統一檔案")
                elif not private_key:
                    logger.info("Private Key為空，不覆蓋現有檔案")
                
                # 標記為共用域名模式
                settings['shared_domain'] = True
                
                # 檢查檔案實際存在狀態，而不是基於傳入資料
                has_token_file = os.path.exists(token_file) and os.path.getsize(token_file) > 0
                has_cert_file = os.path.exists(cert_file) and os.path.getsize(cert_file) > 0
                has_key_file = os.path.exists(key_file) and os.path.getsize(key_file) > 0
                
                settings.update({
                    'has_cert': has_cert_file,
                    'has_key': has_key_file,
                    'has_token': has_token_file
                })
                
                logger.info(f"檔案檢查結果 - Token: {has_token_file}, Cert: {has_cert_file}, Key: {has_key_file}")
            else:
                # 臨時域名模式 - 明確標記為沒有這些資源
                settings.update({
                    'has_cert': False,
                    'has_key': False,
                    'has_token': False,
                    'shared_domain': False
                })
            
            # 保存設定檔案
            settings_file = os.path.join(config_dir, 'tunnel_settings.json')
            with open(settings_file, 'w', encoding='utf-8') as f:
                json.dump(settings, f, indent=2, ensure_ascii=False)
            
            logger.info(f"隧道設定已保存: {current_tunnel_type} - {domain_mode}")
            
            return jsonify({
                'success': True,
                'message': '設定已保存'
            })
            
        except Exception as e:
            logger.error(f"保存隧道設定失敗: {e}")
            return jsonify({
                'success': False,
                'message': f'保存設定失敗: {str(e)}'
            })

@app.route('/api/ngrok/token/clear', methods=['POST'])
def api_ngrok_token_clear():
    """清除服務器端保存的ngrok authtoken"""
    try:
        config_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config')
        token_file = os.path.join(config_dir, 'ngrok_token.txt')
        
        if os.path.exists(token_file):
            os.remove(token_file)
        
        return jsonify({
            'status': 'success',
            'message': 'Token已清除'
        })
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'清除失敗: {str(e)}'
        }), 500



@app.route('/api/sinopac/status', methods=['GET'])
def api_sinopac_status():
    """獲取永豐API連線狀態和期貨帳號"""
    return jsonify(get_sinopac_status())

@app.route('/api/app/version', methods=['GET'])
def api_app_version():
    """獲取程式版本信息"""
    try:
        # 使用已載入的版本資訊
        return jsonify({
            'version': CURRENT_VERSION,
            'build': CURRENT_BUILD,
            'release_date': VERSION_INFO.get('release_date', 'unknown'),
            'description': SYSTEM_DESCRIPTION,
            'updated_at': VERSION_INFO.get('updated_at', 'unknown')
        })
    except Exception as e:
        return jsonify({
            'version': 'error',
            'build': 'error',
            'release_date': 'error',
            'description': 'Auto91 交易系統',
            'error': str(e)
        })

@app.route('/api/sinopac/version', methods=['GET'])
def api_sinopac_version():
    """獲取永豐shioaji版本信息"""
    try:
        if SHIOAJI_AVAILABLE:
            import shioaji as sj
            version = getattr(sj, '__version__', 'unknown')
            return jsonify({
                'version': version,
                'available': True
            })
        else:
            return jsonify({
                'version': 'N/A',
                'available': False
            })
    except Exception as e:
        return jsonify({
            'version': 'Error',
            'available': False,
            'error': str(e)
        })

@app.route('/api/sinopac/check_update', methods=['POST'])
def api_sinopac_check_update():
    """檢查shioaji更新"""
    try:
        import requests
        import json
        
        # 獲取當前版本
        current_version = None
        if SHIOAJI_AVAILABLE:
            import shioaji as sj
            current_version = getattr(sj, '__version__', 'unknown')
        
        # 查詢PyPI獲取最新版本
        response = requests.get('https://pypi.org/pypi/shioaji/json', timeout=10)
        if response.status_code == 200:
            pypi_data = response.json()
            latest_version = pypi_data.get('info', {}).get('version', 'unknown')
            
            if current_version and latest_version != 'unknown':
                # 簡單版本比較
                if compare_versions(latest_version, current_version) > 0:
                    return jsonify({
                        'status': 'success',
                        'data': {
                            'update_available': True,
                            'current_version': current_version,
                            'latest_version': latest_version,
                            'update_command': 'pip install --upgrade shioaji'
                        }
                    })
                else:
                    return jsonify({
                        'status': 'success',
                        'data': {
                            'update_available': False,
                            'current_version': current_version,
                            'latest_version': latest_version
                        }
                    })
        
        return jsonify({
            'status': 'error',
            'message': '無法檢查更新'
        })
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'檢查更新失敗: {str(e)}'
        }), 500

@app.route('/api/sinopac/update', methods=['POST'])
def api_sinopac_update():
    """更新shioaji (僅提供更新指令，不自動執行)"""
    try:
        # 注意：不直接執行pip命令，而是提供指令給用戶
        return jsonify({
            'status': 'success',
            'message': '請在終端機中執行以下指令來更新shioaji:',
            'command': 'pip install --upgrade shioaji',
            'note': '更新後請重新啟動應用程序'
        })
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'更新失敗: {str(e)}'
        }), 500

@app.route('/api/sinopac/auto_update', methods=['POST'])
def api_sinopac_auto_update():
    """自動更新shioaji"""
    try:
        import subprocess
        import sys
        import os
        
        # 執行pip更新
        print_console("SYSTEM", "LOADING", "開始自動更新shioaji...")
        result = subprocess.run(
            [sys.executable, '-m', 'pip', 'install', '--upgrade', 'shioaji'],
            capture_output=True,
            text=True,
            timeout=300  # 5分鐘超時
        )
        
        if result.returncode == 0:
            # 更新成功
            update_output = result.stdout
            
            return jsonify({
                'status': 'success',
                'message': 'shioaji更新成功！',
                'output': update_output,
                'note': '請重啟應用程序以應用新版本'
            })
        else:
            # 更新失敗
            error_output = result.stderr
            return jsonify({
                'status': 'error',
                'message': 'shioaji更新失敗',
                'error': error_output
            }), 500
            
    except subprocess.TimeoutExpired:
        return jsonify({
            'status': 'error',
            'message': '更新超時，請手動執行: pip install --upgrade shioaji'
        }), 500
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'自動更新失敗: {str(e)}'
        }), 500

@app.route('/api/futures/contracts', methods=['GET'])
def api_futures_contracts():
    """獲取期貨合約資訊"""
    global futures_contracts, margin_requirements, sinopac_connected, sinopac_api
    
    if not sinopac_connected or not sinopac_api:
        return jsonify({
            'status': 'disconnected',
            'selected_contracts': {
                'TXF': '-',
                'MXF': '-', 
                'TMF': '-'
            },
            'available_contracts': {
                'TXF': [],
                'MXF': [],
                'TMF': []
            }
        })
    
    try:
        # 先更新保證金資訊（確保獲取最新數據）
        update_margin_requirements_from_api()
        
        # 獲取所有可用合約
        available_contracts = {}
        selected_contracts = {}
        
        for code in ['TXF', 'MXF', 'TMF']:
            try:
                contracts = sinopac_api.Contracts.Futures.get(code)
                if contracts:
                    # 按交割日期排序（確保正確的日期格式排序）
                    sorted_contracts = sorted(contracts, key=get_sort_date)
                    
                    # 可用合約列表
                    available_list = []
                    for c in sorted_contracts:
                        delivery_date = format_delivery_date(c.delivery_date)
                        
                        available_list.append({
                            'code': c.code,
                            'delivery_date': delivery_date,
                            'delivery_month': getattr(c, 'delivery_month', ''),
                            'name': getattr(c, 'name', '')
                        })
                    
                    available_contracts[code] = available_list
                    
                    # 選用合約：轉倉模式下使用次月合約(R2)，否則使用當月合約(R1)
                    if sorted_contracts:
                        # 在轉倉模式下，優先使用次月合約（R2）
                        if rollover_mode and next_month_contracts.get(code):
                            selected_contract = next_month_contracts[code]
                        else:
                            # 非轉倉模式：尋找R1合約作為當月合約
                            r1_contract = None
                            for contract in sorted_contracts:
                                if contract.code.endswith('R1'):
                                    r1_contract = contract
                                    break
                            # 如果找到R1合約則使用，否則使用第一個合約
                            selected_contract = r1_contract if r1_contract else sorted_contracts[0]
                        
                        contract_name = '大台' if code == 'TXF' else '小台' if code == 'MXF' else '微台'
                        margin = margin_requirements.get(contract_name, 0)
                        
                        delivery_date = format_delivery_date(selected_contract.delivery_date)
                        
                        selected_contracts[code] = f"{selected_contract.code}　交割日：{delivery_date}　保證金 ${margin:,}"
                    else:
                        selected_contracts[code] = '-'
                else:
                    available_contracts[code] = []
                    selected_contracts[code] = '-'
                    
            except Exception as e:
                print_console("API", "ERROR", f"獲取{code}合約失敗", str(e))
                available_contracts[code] = []
                selected_contracts[code] = '-'
        
        return jsonify({
            'status': 'connected',
            'selected_contracts': selected_contracts,
            'available_contracts': available_contracts
        })
        
    except Exception as e:
        print_console("API", "ERROR", "獲取期貨合約資訊失敗", str(e))
        return jsonify({
            'status': 'error',
            'selected_contracts': {
                'TXF': '-',
                'MXF': '-',
                'TMF': '-'
            },
            'available_contracts': {
                'TXF': [],
                'MXF': [],
                'TMF': []
            }
        })

@app.route('/api/account/status', methods=['GET'])
def api_account_status():
    """獲取帳戶狀態資訊"""
    global sinopac_api, sinopac_connected
    
    if not sinopac_connected or not sinopac_api:
        return jsonify({
            'status': 'disconnected',
            'error': '永豐API未連線'
        }), 400
    
    try:
        # 獲取保證金資訊
        margin_data = sinopac_api.margin()
        
        # 獲取持倉資訊計算未實現損益
        total_pnl = 0.0
        try:
            positions = sinopac_api.list_positions(sinopac_api.futopt_account)
            for pos in positions:
                total_pnl += pos.pnl
        except:
            total_pnl = 0.0
        
        return jsonify({
            'status': 'success',
            'data': {
                '權益總值': getattr(margin_data, 'equity', 0) or 0,
                '權益總額': getattr(margin_data, 'equity_amount', 0) or 0,
                '今日餘額': getattr(margin_data, 'today_balance', 0) or 0,
                '昨日餘額': getattr(margin_data, 'yesterday_balance', 0) or 0,
                '可用保證金': getattr(margin_data, 'available_margin', 0) or 0,
                '原始保證金': getattr(margin_data, 'initial_margin', 0) or 0,
                '維持保證金': getattr(margin_data, 'maintenance_margin', 0) or 0,
                '風險指標': getattr(margin_data, 'risk_indicator', 0) or 0,
                '手續費': getattr(margin_data, 'fee', 0) or 0,
                '期交稅': getattr(margin_data, 'tax', 0) or 0,
                '本日平倉損益': getattr(margin_data, 'future_settle_profitloss', 0) or 0,
                '未實現損益': total_pnl
            },
            'last_updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': f'獲取帳戶狀態失敗: {str(e)}'
        }), 500

@app.route('/api/trading/status', methods=['GET'])
def api_trading_status():
    """獲取交易日和交割日狀態"""
    try:
        import csv
        
        today = datetime.now()
        
        # 判斷交易日 - 直接調用TXserver中的源頭方法
        def is_trading_day_advanced(check_date=None):
            """交易日判斷邏輯，獨立實現"""
            if check_date is None:
                check_date = today.date()
            
            # 週日固定為非交易日（週六有夜盤交易到凌晨05:00，所以週六是交易日）
            if check_date.weekday() == 6:  # 週日
                return False
            
            # 檢查假期表 - 尋找當年度的holidaySchedule_XXX.csv檔案（民國年）
            try:
                holiday_dir = os.path.join(os.path.dirname(__file__), 'holiday')
                if os.path.exists(holiday_dir):
                    # 轉換西元年為民國年（民國年 = 西元年 - 1911）
                    current_year = check_date.year
                    roc_year = current_year - 1911
                    
                    # 尋找當年度的假期檔案（民國年格式）
                    holiday_files = [f for f in os.listdir(holiday_dir) 
                                   if f.startswith('holidaySchedule_') and f.endswith('.csv')]
                    
                    # 嘗試找到包含當年民國年的檔案
                    target_file = None
                    for filename in holiday_files:
                        # 檔案名稱可能包含民國年資訊（如 holidaySchedule_114.csv）
                        if str(roc_year) in filename:
                            target_file = filename
                            break
                    
                    # 如果沒找到年份檔案，使用最新的檔案
                    if not target_file and holiday_files:
                        target_file = max(holiday_files, key=lambda f: os.path.getctime(os.path.join(holiday_dir, f)))
                    
                    if target_file:
                        csv_path = os.path.join(holiday_dir, target_file)
                        holidays = {}
                        
                        with open(csv_path, 'r', encoding='big5') as f:
                            reader = csv.DictReader(f)
                            for row in reader:
                                try:
                                    # 解析日期
                                    date_str = row.get('日期', '')
                                    if '/' in date_str:
                                        holiday_date = datetime.strptime(date_str, '%Y/%m/%d').date()
                                    elif '-' in date_str:
                                        holiday_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                                    else:
                                        continue
                                    
                                    # 判斷是否為交易日：'o'表示交易日，其他或空白表示非交易日
                                    remark = row.get('備註', '').strip().lower()
                                    is_trading = (remark == 'o')
                                    holidays[holiday_date] = is_trading
                                except (ValueError, KeyError):
                                    continue
                        
                        # 檢查今天是否在假期表中
                        if check_date in holidays:
                            return holidays[check_date]
            except Exception as e:
                print_console("SYSTEM", "ERROR", "讀取假期檔案失敗", str(e))
                pass
            
            # 未在假期表中的工作日視為交易日
            return True
        
        # 判斷交割日 - 檢查現在使用的合約交割日期
        def is_delivery_day_advanced(check_date=None):
            """交割日判斷邏輯，檢查現在使用的合約是否今天交割"""
            if check_date is None:
                check_date = today.date()
            
            try:
                global sinopac_api, sinopac_connected
                if not sinopac_connected or not sinopac_api:
                    return False
                
                # 檢查目前選用的合約（最近交割日的合約）的交割日
                for code in ['TXF', 'MXF', 'TMF']:
                    try:
                        contracts = sinopac_api.Contracts.Futures.get(code)
                        if contracts:
                            # 按交割日期排序，取得最近的合約（即選用合約）
                            sorted_contracts = sorted(contracts, key=get_sort_date)
                            if sorted_contracts:
                                # 檢查最近的合約（選用合約）是否今天交割
                                selected_contract = sorted_contracts[0]
                                delivery_date_str = selected_contract.delivery_date
                                
                                try:
                                    # 解析交割日期
                                    if isinstance(delivery_date_str, str):
                                        if len(delivery_date_str) == 8:  # YYYYMMDD
                                            delivery_date = datetime.strptime(delivery_date_str, '%Y%m%d').date()
                                        elif '/' in delivery_date_str:  # YYYY/MM/DD
                                            delivery_date = datetime.strptime(delivery_date_str, '%Y/%m/%d').date()
                                        elif '-' in delivery_date_str:  # YYYY-MM-DD
                                            delivery_date = datetime.strptime(delivery_date_str, '%Y-%m-%d').date()
                                        else:
                                            continue
                                    else:
                                        continue
                                    
                                    # 如果今天是任一選用合約的交割日，就返回True
                                    if check_date == delivery_date:
                                        return True
                                except ValueError:
                                    continue
                    except Exception as e:
                        print_console("TRADE", "ERROR", f"檢查{code}合約交割日失敗", str(e))
                        continue
                
                return False
            except Exception as e:
                print_console("TRADE", "ERROR", "檢查交割日失敗", str(e))
                return False
        
        # 執行判斷
        is_trading = is_trading_day_advanced()
        is_delivery = is_delivery_day_advanced()
        
        # 判斷開市/關市狀態
        def is_market_open():
            """判斷是否為開市時間"""
            # 如果不是交易日，直接返回休市
            if not is_trading:
                return False
            
            current_hour = today.hour
            current_minute = today.minute
            current_time = current_hour * 100 + current_minute  # HHMM格式
            current_weekday = today.weekday()  # 0=週一, 6=週日
            
            # 週六特殊處理：週六凌晨05:00後為休市
            if current_weekday == 5:  # 週六
                if current_time >= 500:  # 05:00後
                    return False  # 休市
            
            # 早盤：8:45-13:45
            morning_start = 845
            morning_end = 1345
            
            # 午盤：14:50-次日05:00
            afternoon_start = 1450
            afternoon_end = 500  # 次日05:00
            
            # 判斷是否在交易時段
            if current_time >= morning_start and current_time <= morning_end:
                return True  # 早盤時段
            elif current_time >= afternoon_start or current_time <= afternoon_end:
                return True  # 午盤時段（跨日）
            
            return False
        
        is_open = is_market_open()
        
        # 週幾的中文對應
        weekday_chinese = ['週一', '週二', '週三', '週四', '週五', '週六', '週日']
        weekday_display = weekday_chinese[today.weekday()]
        
        return jsonify({
            'status': 'success',
            'current_datetime': today.strftime('%Y/%m/%d %H:%M:%S'),
            'weekday': weekday_display,
            'trading_day_status': '交易日' if is_trading else '非交易日',
            'delivery_day_status': '交割日' if is_delivery else '非交割日',
            'market_status': '開市' if is_open else '休市',
            'is_trading_day': is_trading,
            'is_delivery_day': is_delivery,
            'is_market_open': is_open
        })
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': f'獲取交易狀態失敗: {str(e)}'
        }), 500

@app.route('/api/position/status', methods=['GET'])
def api_position_status():
    """獲取持倉狀態資訊"""
    global sinopac_api, sinopac_connected
    
    if not sinopac_connected or not sinopac_api:
        return jsonify({
            'status': 'disconnected',
            'error': '永豐API未連線'
        }), 400

    try:
        # 獲取持倉資訊
        positions = sinopac_api.list_positions(sinopac_api.futopt_account)
        logger.info(f"🔍 原始持倉數據 - 總數: {len(positions) if positions else 0}")
        
        # 詳細記錄每個持倉的原始數據
        if positions:
            for i, pos in enumerate(positions):
                logger.info(f"📊 持倉 {i+1}: code={pos.code}, direction={pos.direction}, quantity={pos.quantity}, price={pos.price}")
                logger.info(f"    額外屬性: {[attr for attr in dir(pos) if not attr.startswith('_')]}")
        else:
            logger.warning("⚠️ 未獲取到任何持倉數據")
        
        # 初始化三種合約的持倉資料
        position_data = {
            'TXF': {'動作': '-', '數量': '-', '均價': '-', '市價': '-', '未實現損益': '-'},
            'MXF': {'動作': '-', '數量': '-', '均價': '-', '市價': '-', '未實現損益': '-'},
            'TMF': {'動作': '-', '數量': '-', '均價': '-', '市價': '-', '未實現損益': '-'}
        }
        
        # 讀取最近7天交易記錄以獲取開倉詳細信息（修復隔夜持倉問題）
        opening_trades = {}  # 儲存開倉交易的詳細信息
        today = datetime.now()
        logger.info("🔍 開始搜尋最近7天的開倉記錄以修復隔夜持倉顯示問題...")
        
        # 遍歷最近7天的交易記錄
        for i in range(7):  # 檢查最近7天
            check_date = today - timedelta(days=i)
            date_str = check_date.strftime('%Y%m%d')
            json_file = os.path.join(TX_DATA_DIR, f'TXtransdata_{date_str}.json')
            
            if os.path.exists(json_file):
                try:
                    with open(json_file, 'r', encoding='utf-8') as f:
                        trades_data = json.load(f)
                        logger.info(f"📄 讀取交易記錄: {date_str} ({len(trades_data)}筆)")
                except Exception as e:
                    logger.error(f"讀取 {date_str} 交易記錄失敗: {e}")
                    continue
                    
                # 查找開倉交易
                for trade in trades_data:
                    if trade.get('type') == 'deal':
                        raw_data = trade.get('raw_data', {})
                        order = raw_data.get('order', {})
                        contract = raw_data.get('contract', {})
                        
                        if order.get('oc_type') == 'New':  # 開倉交易
                            contract_code = contract.get('code', '')
                            contract_type = None
                            
                            if 'TXF' in contract_code:
                                contract_type = 'TXF'
                            elif 'MXF' in contract_code:
                                contract_type = 'MXF'
                            elif 'TMF' in contract_code:
                                contract_type = 'TMF'
                            
                            if contract_type:
                                # 優先從contract對象獲取完整的交割月份信息
                                delivery_month = contract.get('delivery_month', '')
                                
                                # 如果沒有，嘗試從合約代碼提取
                                if not delivery_month and len(contract_code) >= 8:
                                    delivery_month = contract_code[-6:]  # 取最後6位數字
                                
                                # 避免覆蓋已有記錄（優先使用較新的記錄用於隔夜持倉）
                                trade_timestamp = trade.get('timestamp', '')
                                
                                # 如果還沒有該合約類型的記錄，或找到更新的記錄，則更新
                                if (contract_type not in opening_trades or 
                                    trade_timestamp > opening_trades[contract_type].get('開倉時間', '')):
                                    
                                    opening_trades[contract_type] = {
                                        '開倉時間': trade_timestamp,
                                        '成交單號': order.get('id', ''),  # 真實成交單號
                                        '委託單號': order.get('ordno', ''),
                                        '訂單類型': order.get('order_type', ''),  # IOC, ROD等
                                        '委託價格類型': order.get('price_type', ''),
                                        '商品名稱': trade.get('contract_name', ''),
                                        '到期月份': delivery_month,
                                        '交割日': delivery_month,  # 用於計算實際交割日
                                        '合約代號': contract_code,
                                        '開倉價格': trade.get('real_opening_price', order.get('price', 0)),
                                        '數據來源': f'{date_str} ({i}天前)' if i > 0 else f'{date_str} (今日)'
                                    }
                                    logger.info(f"🔍 找到 {contract_type} 開倉記錄: {trade_timestamp[:19]} (來源: {date_str})")
        
        # 記錄搜尋結果
        logger.info(f"✅ 開倉記錄搜尋完成，共找到 {len(opening_trades)} 個合約類型的記錄:")
        for contract_type, info in opening_trades.items():
            logger.info(f"   - {contract_type}: {info['開倉時間'][:19]} ({info['數據來源']})")
            logger.info(f"     成交單號: {info.get('成交單號', 'N/A')}, 訂單類型: {info.get('訂單類型', 'N/A')}")
        
        # 初始化總損益
        total_pnl = 0.0
        has_positions = False
        
        if positions and len(positions) > 0:
            logger.info(f"📊 發現 {len(positions)} 筆持倉記錄")
            # 遍歷所有持倉，按合約類型分類
            for position in positions:
                contract_code = position.code
                logger.info(f"🔍 分析持倉合約代碼: {contract_code}")
                contract_type = None
                
                # 擴展合約識別邏輯，支援R1、G5等格式
                if 'TXF' in contract_code or contract_code.startswith('TX'):
                    contract_type = 'TXF'
                    logger.info(f"✅ 識別為大台指 TXF: {contract_code}")
                elif 'MXF' in contract_code or contract_code.startswith('MX'):
                    contract_type = 'MXF'
                    logger.info(f"✅ 識別為小台指 MXF: {contract_code}")
                elif 'TMF' in contract_code or contract_code.startswith('TM'):
                    contract_type = 'TMF'
                    logger.info(f"✅ 識別為微台指 TMF: {contract_code}")
                else:
                    logger.warning(f"❌ 未識別的合約代碼: {contract_code}")
                    continue  # 跳過非期貨合約
                
                try:
                    # 判斷多空方向
                    direction = '多單' if position.direction == 'Buy' else '空單'
                    
                    # 獲取該持倉的未實現損益
                    unrealized_pnl = float(position.pnl) if hasattr(position, 'pnl') and position.pnl is not None else 0.0
                    
                    # 獲取市價
                    last_price = float(position.last_price) if hasattr(position, 'last_price') else 0.0
                    
                    # 獲取數量和均價
                    quantity = abs(int(position.quantity))
                    avg_price = float(position.price)
                    
                    # 從JSON配對系統獲取開倉詳細信息 - 更強健的方案
                    def get_position_opening_info(contract_type, quantity, avg_price):
                        """從JSON配對系統獲取開倉詳細信息"""
                        try:
                            from datetime import timedelta
                            from trading_config import TradingConfig
                            
                            TX_RECORDS_DIR = TradingConfig.TX_RECORDS_DIR
                            today = datetime.now()
                            
                            # 查找最近30天的開倉記錄
                            for i in range(30):
                                check_date = today - timedelta(days=i)
                                date_str = check_date.strftime('%Y%m%d')
                                open_file = os.path.join(TX_RECORDS_DIR, f'open_positions_{date_str}.json')
                                
                                if os.path.exists(open_file):
                                    try:
                                        with open(open_file, 'r', encoding='utf-8') as f:
                                            opens = json.load(f)
                                        
                                        # 查找匹配的開倉記錄
                                        for open_record in opens:
                                            record_contract = open_record.get('contract_code', '')
                                            record_status = open_record.get('status', '')
                                            record_remaining = float(open_record.get('remaining_quantity', 0))
                                            
                                            # 判斷合約類型匹配
                                            record_type = None
                                            if 'TXF' in record_contract:
                                                record_type = 'TXF'
                                            elif 'MXF' in record_contract:
                                                record_type = 'MXF'
                                            elif 'TMF' in record_contract:
                                                record_type = 'TMF'
                                            
                                            # 匹配條件：同合約類型、未完全平倉、數量匹配
                                            if (record_type == contract_type and 
                                                record_status in ['open', 'partial_covered'] and
                                                record_remaining > 0):
                                                
                                                # 提取交割月份
                                                delivery_month = ''
                                                if len(record_contract) >= 8:
                                                    delivery_month = record_contract[-6:]
                                                
                                                # 格式化時間
                                                timestamp = open_record.get('timestamp', '')
                                                formatted_time = ''
                                                if timestamp:
                                                    try:
                                                        if 'T' in timestamp:
                                                            dt = datetime.fromisoformat(timestamp.replace('Z', '+00:00'))
                                                        else:
                                                            dt = datetime.fromisoformat(timestamp)
                                                        formatted_time = dt.strftime('%Y/%m/%d %H:%M:%S')
                                                    except:
                                                        formatted_time = timestamp
                                                
                                                logger.info(f"✅ 從JSON配對系統找到 {contract_type} 開倉記錄: {formatted_time}")
                                                
                                                return {
                                                    '開倉時間': formatted_time,
                                                    '成交單號': open_record.get('order_id', ''),
                                                    '委託單號': open_record.get('order_id', ''),
                                                    '訂單類型': 'IOC',  # 默認
                                                    '委託價格類型': 'MKT',  # 默認
                                                    '商品名稱': f"{contract_type}期貨",
                                                    '到期月份': delivery_month,
                                                    '交割日': delivery_month,
                                                    '合約代號': record_contract,
                                                    '開倉價格': open_record.get('price', avg_price)
                                                }
                                    except Exception as e:
                                        logger.warning(f"讀取開倉記錄失敗 {open_file}: {e}")
                                        continue
                            
                            # 找不到匹配記錄，返回空值
                            logger.warning(f"未在JSON配對系統中找到 {contract_type} 的開倉記錄")
                            return {}
                            
                        except Exception as e:
                            logger.error(f"從JSON配對系統獲取開倉信息失敗: {e}")
                            return {}
                    
                    # 獲取開倉詳細信息：優先使用JSON配對系統
                    opening_info = get_position_opening_info(contract_type, quantity, avg_price)
                    
                    # 如果 JSON 系統沒有找到，回退到交易記錄搜尋方法
                    if not opening_info:
                        opening_info = opening_trades.get(contract_type, {})
                        data_source = opening_info.get('數據來源', '無記錄')
                        logger.info(f"回退使用交易記錄獲取 {contract_type} 開倉信息 (來源: {data_source})")
                    
                    # 優先使用JSON配對系統的交割月份數據
                    delivery_month = opening_info.get('到期月份', '')
                    
                    logger.info(f"🔍 開始解析合約 {contract_code} 的交割月份，優先使用JSON: {delivery_month}")
                    
                    # 如果JSON系統沒有提供，嘗試從永豐API合約代碼提取
                    if not delivery_month:
                        # 方法1: 標準格式 (TXF202508, MXF202508, TMF202508)
                        if len(contract_code) >= 8:
                            potential_month = contract_code[-6:]
                            if potential_month.isdigit() and len(potential_month) == 6:
                                delivery_month = potential_month
                                logger.info(f"✅ 從永豐API合約代碼提取: {contract_code} -> {delivery_month}")
                        
                        # 方法2: 正則表達式尋找6位數字
                        if not delivery_month:
                            import re
                            digit_match = re.search(r'(\d{6})', contract_code)
                            if digit_match:
                                delivery_month = digit_match.group(1)
                                logger.info(f"✅ 正則表達式提取: {contract_code} -> {delivery_month}")
                    
                    # 方法3: 處理R1、G5等特殊格式
                    if not delivery_month:
                        current_date = datetime.now()
                        current_year = current_date.year
                        current_month = current_date.month
                        
                        # R1通常表示當月合約，R2表示次月合約，G5可能表示特定月份
                        if 'R1' in contract_code:
                            delivery_month = f"{current_year}{current_month:02d}"
                            logger.info(f"✅ 方法3成功 - R1格式當月: {contract_code} -> {delivery_month}")
                        elif 'R2' in contract_code:
                            next_month = current_month + 1 if current_month < 12 else 1
                            next_year = current_year if current_month < 12 else current_year + 1
                            delivery_month = f"{next_year}{next_month:02d}"
                            logger.info(f"✅ 方法3成功 - R2格式次月: {contract_code} -> {delivery_month}")
                        elif 'G5' in contract_code:
                            # G5可能表示5月份，但需要確定年份
                            delivery_month = f"{current_year}05"
                            logger.info(f"✅ 方法3成功 - G5格式5月: {contract_code} -> {delivery_month}")
                        else:
                            # 其他未知格式，使用當前月份
                            delivery_month = f"{current_year}{current_month:02d}"
                            logger.info(f"⚠️ 方法3備用 - 使用當前月份: {contract_code} -> {delivery_month}")
                    
                    logger.info(f"🎯 最終交割月份: {contract_code} -> {delivery_month}")
                    
                    # (已在上方優先使用JSON數據)
                    
                    # 如果還是沒有交割月份，嘗試從當前選用合約獲取
                    if not delivery_month:
                        try:
                            global selected_contracts
                            if 'selected_contracts' in globals() and selected_contracts:
                                contract_info = selected_contracts.get(contract_type, '')
                                if contract_info and '交割日：' in contract_info:
                                    # 解析選用合約信息 (如 "TXF　交割日：2025/08/20　保證金 $240,000")
                                    parts = contract_info.split('　')
                                    for part in parts:
                                        if '交割日：' in part:
                                            date_str = part.replace('交割日：', '')
                                            # 轉換為YYYYMM格式
                                            if '/' in date_str:
                                                year, month, day = date_str.split('/')
                                                delivery_month = f"{year}{month.zfill(2)}"
                                                logger.info(f"從選用合約信息獲取 {contract_type} 交割月份: {delivery_month}")
                                            break
                        except Exception as e:
                            logger.warning(f"解析選用合約交割日期失敗: {e}")
                    
                    # 更新對應合約的資料
                    position_data[contract_type] = {
                        '動作': direction,
                        '數量': f"{quantity} 口",
                        '均價': f"{avg_price:,.0f}",
                        '市價': f"{last_price:,.0f}",
                        '未實現損益': f"{unrealized_pnl:,.0f}",
                        # 新增開倉詳細信息
                        '開倉時間': opening_info.get('開倉時間', ''),
                        '成交單號': opening_info.get('成交單號', ''),  # 真實成交單號
                        '委託單號': opening_info.get('委託單號', ''),
                        '訂單類型': opening_info.get('訂單類型', ''),  # IOC, ROD等
                        '委託條件': opening_info.get('訂單類型', ''),  # 兼容舊欄位名稱
                        '委託價格類型': opening_info.get('委託價格類型', ''),
                        '商品名稱': opening_info.get('商品名稱', ''),
                        '到期月份': delivery_month,  # 確保有交割月份
                        '交割日': delivery_month,  # 交割日信息
                        '合約代號': opening_info.get('合約代號', contract_code),
                        '開倉價格': opening_info.get('開倉價格', position.price)  # 真實開倉價格
                    }
                    
                    # 累計總損益
                    total_pnl += unrealized_pnl
                    has_positions = True
                    
                except Exception as e:
                    logger.error(f"處理持倉 {contract_code} 時發生錯誤: {e}")
                    continue
        
        # 格式化總損益
        total_pnl_display = f"{total_pnl:,.0f} TWD" if has_positions else "-"
        
        return jsonify({
            'status': 'success',
            'data': position_data,
            'total_pnl': total_pnl_display,
            'total_pnl_value': total_pnl,
            'has_positions': has_positions,
            'last_updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': f'獲取持倉狀態失敗: {str(e)}'
        }), 500



@app.route('/api/logs', methods=['GET'])
def api_get_logs():
    """獲取系統日誌"""
    try:
        # 使用 get_tunnel_requests() 來獲取合併的請求記錄
        all_logs = get_tunnel_requests()
        return jsonify({
            'status': 'success',
            'logs': all_logs
        })
    except Exception as e:
        # 如果出錯，降級使用只有 custom_request_logs
        global custom_request_logs
        return jsonify({
            'status': 'error', 
            'message': str(e),
            'logs': custom_request_logs
        })

@app.route('/api/rollover/status', methods=['GET'])
def api_rollover_status():
    """獲取轉倉狀態"""
    global rollover_mode, rollover_start_date, next_month_contracts
    
    try:
        # 檢查轉倉模式
        check_rollover_mode()
        
        # 獲取當前合約資訊
        current_contracts = {
            'TXF': contract_txf.code if contract_txf else None,
            'MXF': contract_mxf.code if contract_mxf else None,
            'TMF': contract_tmf.code if contract_tmf else None
        }
        
        # 獲取次月合約資訊
        next_contracts = {
            'TXF': next_month_contracts.get('TXF', {}).code if next_month_contracts.get('TXF') else None,
            'MXF': next_month_contracts.get('MXF', {}).code if next_month_contracts.get('MXF') else None,
            'TMF': next_month_contracts.get('TMF', {}).code if next_month_contracts.get('TMF') else None
        }
        
        return jsonify({
            'status': 'success',
            'rollover_mode': rollover_mode,
            'rollover_start_date': rollover_start_date.isoformat() if rollover_start_date else None,
            'current_contracts': current_contracts,
            'next_month_contracts': next_contracts,
            'last_updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': f'獲取轉倉狀態失敗: {str(e)}'
        }), 500

@app.route('/api/connection/duration', methods=['GET'])
def api_connection_duration():
    """獲取連線時長信息"""
    try:
        duration_hours = get_connection_duration()
        login_time = sinopac_login_time.isoformat() if sinopac_login_time else None
        
        return jsonify({
            'status': 'success',
            'duration_hours': round(duration_hours, 2),
            'login_time': login_time,
            'auto_logout_hours': AUTO_LOGOUT_HOURS,
            'remaining_hours': max(0, AUTO_LOGOUT_HOURS - duration_hours)
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/connection/monitor', methods=['GET'])
def api_connection_monitor_status():
    """查詢連線監控狀態"""
    global reconnect_attempts, last_connection_check, max_reconnect_attempts, is_reconnecting
    
    try:
        # 獲取當前動態檢查間隔
        current_interval = get_dynamic_check_interval()
        
        # 判斷當前檢查模式
        if is_reconnecting:
            check_mode = "重連中"
            interval_display = "30秒"
        elif current_interval == 60:
            check_mode = "交易時間"
            interval_display = "1分鐘"
        else:
            check_mode = "非交易時間"
            interval_display = "10分鐘"
        
        return jsonify({
            'status': 'success',
            'monitor_active': connection_monitor_timer is not None,
            'check_interval_seconds': current_interval,
            'check_interval_display': interval_display,
            'check_mode': check_mode,
            'max_reconnect_attempts': max_reconnect_attempts,
            'current_reconnect_attempts': reconnect_attempts,
            'is_reconnecting': is_reconnecting,
            'last_check_time': last_connection_check.strftime('%Y-%m-%d %H:%M:%S') if last_connection_check else None,
            'connection_status': sinopac_connected and sinopac_login_status
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/close_application', methods=['POST'])
def api_close_application():
    """關閉整個應用程式"""
    try:
        # 執行清理工作
        cleanup_on_exit()
        
        # 延遲一秒後關閉程式，確保清理工作完成
        def delayed_exit():
            time.sleep(1)
            os._exit(0)  # 強制關閉整個程式
        
        create_managed_thread(target=delayed_exit, name="延遲退出線程").start()
        
        return jsonify({
            'status': 'success',
            'message': '應用程式正在關閉...'
        })
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'關閉程式失敗: {str(e)}'
        }), 500

@app.route('/shutdown', methods=['POST'])
def shutdown_flask_server():
    """優雅關閉Flask服務器"""
    try:
        print_console("FLASK", "INFO", "收到關閉信號，正在關閉Flask服務器...")
        
        # 設置關閉標誌
        global flask_server_shutdown
        flask_server_shutdown = True
        
        # 使用werkzeug的shutdown功能
        shutdown_func = request.environ.get('werkzeug.server.shutdown')
        if shutdown_func is None:
            # 如果無法獲取shutdown函數，使用強制退出
            def force_shutdown():
                time.sleep(0.5)
                os._exit(0)
            create_managed_thread(target=force_shutdown, name="強制關閉線程").start()
            return jsonify({'message': 'Flask服務器正在強制關閉...'})
        
        shutdown_func()
        return jsonify({'message': 'Flask服務器正在關閉...'})
        
    except Exception as e:
        print_console("FLASK", "ERROR", f"關閉Flask服務器失敗: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/<path:path>')
def static_files(path):
    if path.startswith('api/'):
        abort(404)
    return send_from_directory(app.static_folder, path)

@app.route('/favicon/<path:filename>')
def serve_favicon(filename):
    favicon_dir = os.path.join(os.path.dirname(__file__), 'favicon')
    return send_from_directory(favicon_dir, filename)

# 端口設置
def get_port():
    """從根目錄的 port.txt 檔案讀取端口設置，若無則自動建立"""
    try:
        # 獲取根目錄路徑（server 資料夾的上一層）
        root_dir = os.path.dirname(os.path.dirname(__file__))
        port_file = os.path.join(root_dir, 'port.txt')
        
        if not os.path.exists(port_file):
            # 自動建立預設 port.txt
            with open(port_file, 'w', encoding='utf-8') as f:
                f.write('port:5000\nlog_console:0\n')
            return 5000, 0
        
        with open(port_file, 'r', encoding='utf-8') as f:
            lines = f.readlines()
            port = 5000
            log_console = 0
            
            for line in lines:
                line = line.strip()
                if line.startswith('port:'):
                    try:
                        port_str = line.split(':')[1].strip()
                        port = int(port_str)
                        if not (1024 <= port <= 65535):  # 檢查端口範圍
                            port = 5000
                    except ValueError:
                        port = 5000
                elif line.startswith('log_console:'):
                    try:
                        log_str = line.split(':')[1].strip()
                        log_console = int(log_str)
                        if log_console not in [0, 1]:  # 檢查值範圍
                            log_console = 0
                    except ValueError:
                        log_console = 0
            
            return port, log_console
    except Exception as e:
        print_console("SYSTEM", "WARNING", "讀取設置失敗，使用預設設置", str(e))
        return 5000, 0

# 獲取當前端口和日誌設置
CURRENT_PORT, LOG_CONSOLE = get_port()

def start_flask():
    """啟動Flask伺服器並註冊到進程管理"""
    global flask_server_shutdown
    
    # 配置werkzeug日誌格式，移除重複時間戳
    werkzeug_logger = logging.getLogger('werkzeug')
    werkzeug_logger.setLevel(logging.INFO)
    
    # 清除任何現有的處理器，避免重複日誌
    werkzeug_logger.handlers.clear()
    
    # 創建自定義格式化器，只顯示IP和請求信息（移除werkzeug內建時間戳）
    class CleanHTTPFormatter(logging.Formatter):
        def format(self, record):
            # 使用正則表達式移除werkzeug內建的時間戳部分 [24/Jul/2025 18:07:28]
            import re
            message = record.getMessage()
            # 移除 [日期/月份/年份 時間] 格式的時間戳
            clean_message = re.sub(r'\s*\[\d{1,2}/\w{3}/\d{4}\s+\d{1,2}:\d{2}:\d{2}\]\s*', ' ', message)
            # 移除多餘的空格和破折號
            clean_message = re.sub(r'\s+-\s+-\s+', ' ', clean_message)
            clean_message = clean_message.strip()
            
            # 返回我們自定義的格式：時間戳 - werkzeug - INFO - 清理後的請求信息
            return f"{self.formatTime(record)} - {record.name} - {record.levelname} - {clean_message}"
    
    handler = logging.StreamHandler()
    handler.setFormatter(CleanHTTPFormatter())
    werkzeug_logger.addHandler(handler)
    
    # 防止向父級傳播（避免重複輸出）
    werkzeug_logger.propagate = False
    
    try:
        print_console("FLASK", "START", f"Flask服務器啟動於端口 {CURRENT_PORT}")
        
        # 檢查是否收到關閉信號
        if not flask_server_shutdown and not SHUTDOWN_FLAG.is_set():
            app.run(port=CURRENT_PORT, threaded=True, use_reloader=False)
    except Exception as e:
        print_console("FLASK", "ERROR", f"Flask服務器發生錯誤: {e}")
    finally:
        print_console("FLASK", "STOP", "Flask服務器已停止")

def start_webview():
    # 根據 log_console 設定決定是否隱藏命令行視窗
    if LOG_CONSOLE == 0:
        # 隱藏命令行視窗（背景執行）
        import ctypes
        try:
            # 獲取當前進程的句柄
            hwnd = ctypes.windll.kernel32.GetConsoleWindow()
            if hwnd:
                # 隱藏命令行視窗
                ctypes.windll.user32.ShowWindow(hwnd, 0)  # SW_HIDE = 0
                print_console("SYSTEM", "INFO", "已隱藏命令行視窗，程式在背景執行")
        except Exception as e:
            print_console("SYSTEM", "ERROR", "隱藏命令行視窗失敗", str(e))
    
    # 🔧 增強的GUI初始化 - 支援無GUI環境運行
    try:
        # 創建webview視窗
        window = webview.create_window(
            'Auto91－交易系統', 
            f'http://127.0.0.1:{CURRENT_PORT}',
            width=1280,
            height=960,
            min_size=(1280, 960),
            maximized=True
        )
        
        # 綁定視窗關閉事件
        def on_window_closing():
            print_console("SYSTEM", "INFO", "視窗關閉中，正在清理資源...")
            
            # 執行清理工作（同步執行，確保完成）
            try:
                cleanup_on_exit()
                print_console("SYSTEM", "SUCCESS", "視窗關閉清理完成")
            except Exception as e:
                print_console("SYSTEM", "ERROR", f"視窗關閉清理失敗: {e}")
            
            return True  # 允許關閉
        
        # 使用closing事件來確保在關閉前執行清理
        window.events.closing += on_window_closing
        
        # 啟動webview
        print_console("SYSTEM", "INFO", f"正在啟動GUI界面: http://127.0.0.1:{CURRENT_PORT}")
        webview.start(debug=False)
        
    except Exception as e:
        # GUI初始化失敗時，以無GUI模式運行
        print_console("SYSTEM", "WARNING", f"GUI初始化失敗，切換到無GUI模式運行: {str(e)}")
        print_console("SYSTEM", "INFO", f"系統繼續以純後台模式運行，可通過 http://127.0.0.1:{CURRENT_PORT} 訪問")
        
        # 無GUI模式下，保持程序運行
        try:
            # 等待中斷信號
            import signal
            import time
            
            def signal_handler(signum, frame):
                print_console("SYSTEM", "INFO", "收到停止信號，正在關閉系統...")
                cleanup_on_exit()
                exit(0)
            
            # 註冊信號處理器
            signal.signal(signal.SIGINT, signal_handler)  # Ctrl+C
            signal.signal(signal.SIGTERM, signal_handler)  # 終止信號
            
            print_console("SYSTEM", "INFO", "無GUI模式啟動完成，使用 Ctrl+C 停止系統")
            print_console("SYSTEM", "INFO", f"Web界面地址: http://127.0.0.1:{CURRENT_PORT}")
            
            # 保持運行直到收到停止信號
            while True:
                time.sleep(1)
                
        except KeyboardInterrupt:
            print_console("SYSTEM", "INFO", "收到鍵盤中斷，正在關閉系統...")
            cleanup_on_exit()
            exit(0)
    
    # webview關閉後，程式應該退出
    print_console("SYSTEM", "INFO", "webview已關閉，程式即將退出...")
    
    # 確保清理工作執行完成後再退出
    try:
        cleanup_on_exit()
        time.sleep(0.5)  # 給清理工作一點時間
    except Exception as e:
        print_console("SYSTEM", "ERROR", f"最終清理失敗: {e}")
    
    # 使用sys.exit而不是os._exit，確保atexit處理器被調用
    sys.exit(0)

# 永豐API相關函數
def init_sinopac_api():
    """初始化永豐API對象（不設置callback）"""
    global sinopac_api
    try:
        if not SHIOAJI_AVAILABLE:
            print_console("API", "WARNING", "shioaji模組未安裝，無法初始化永豐API")
            return False
            
        sinopac_api = sj.Shioaji()
        print_console("API", "SUCCESS", "永豐API對象創建成功")
        return True
    except Exception as e:
        print_console("API", "ERROR", "初始化永豐API失敗", str(e))
        sinopac_api = None
        return False

def update_futures_contracts():
    """更新期貨合約資訊"""
    global futures_contracts, sinopac_api, sinopac_connected
    
    if not sinopac_api or not sinopac_connected:
        return False
    
    try:
        # 獲取各期貨合約的最新資訊
        for code in ['TXF', 'MXF', 'TMF']:
            contracts = sinopac_api.Contracts.Futures.get(code)
            if contracts:
                # 根據轉倉模式選擇合約
                sorted_contracts = sorted(contracts, key=lambda x: x.delivery_date)
                if rollover_mode and next_month_contracts.get(code):
                    futures_contracts[code] = next_month_contracts[code]
                else:
                    # 非轉倉模式：尋找R1合約作為當月合約
                    r1_contract = None
                    for contract in sorted_contracts:
                        if contract.code.endswith('R1'):
                            r1_contract = contract
                            break
                    futures_contracts[code] = r1_contract if r1_contract else sorted_contracts[0]
        
        return True
    except Exception as e:
        print_console("API", "ERROR", "更新期貨合約失敗", str(e))
        return False

def update_margin_requirements_from_api():
    """從台期所API更新保證金資訊"""
    global margin_requirements
    
    try:
        import requests
        url = "https://openapi.taifex.com.tw/v1/IndexFuturesAndOptionsMargining"
        response = requests.get(url)
        
        if response.status_code == 200:
            data = response.json()
            new_margins = {}
            
            for item in data:
                contract = item.get('Contract', '')
                margin = int(item.get('InitialMargin', 0))
                
                if contract == '臺股期貨':
                    new_margins['大台'] = margin
                elif contract == '小型臺指':
                    new_margins['小台'] = margin  
                elif contract == '微型臺指期貨':
                    new_margins['微台'] = margin
            
            if new_margins:
                margin_requirements.update(new_margins)
                return True
        
        return False
    except Exception as e:
        print_console("API", "ERROR", "更新保證金失敗", str(e))
        return False

def order_callback(state, deal, order=None):
    """訂單回調函數處理 - 參照BTC系統的統一數據源策略"""
    global order_octype_map, contract_txf, contract_mxf, contract_tmf
        
    try:
        logger.info(f"🔔 收到TX回調事件: {state}")
        
        # 🔧 重新設計：統一數據提取邏輯，支援永豐金API的不同回調格式
        if str(state) == 'OrderState.FuturesDeal':
            # 成交回調 - 檢查多種可能的數據路徑
            order_id = (deal.get('trade_id') or 
                       deal.get('order_id') or 
                       deal.get('id') or '未知').strip()
            contract_code = (deal.get('code') or 
                           deal.get('contract', {}).get('code', '') or '')
        else:
            # 訂單提交回調 - 檢查多種可能的數據路徑  
            order_id = (deal.get('order', {}).get('id') or 
                       deal.get('id') or 
                       deal.get('order_id') or '未知').strip()
            contract_code = (deal.get('contract', {}).get('code') or 
                           deal.get('code') or '')
        
        contract_name = get_contract_name_from_code(contract_code)
        logger.info(f"📋 處理訂單: {order_id}, 合約: {contract_name}({contract_code})")
        
        # ⭐ 關鍵修復：統一從order_octype_map獲取訂單原始資訊
        octype_info = order_octype_map.get(order_id)
        logger.info(f"🔍 查詢訂單映射結果: order_id={order_id}, found={octype_info is not None}")
        if octype_info:
            logger.info(f"🔍 映射內容: {octype_info}")
        
        if octype_info is None:
            logger.warning(f"⚠️ 訂單映射中找不到 {order_id}，這不應該發生！")
            logger.warning("🚨 這表明訂單提交時沒有正確保存到order_octype_map中")
            
            # 🎯 強制延遲處理 - 不再進行同步應急推斷
            # 所有沒有映射的情況都等待提交回調創建映射
            logger.warning(f"🔄 強制延遲處理，等待提交回調創建映射: {order_id}")
            
            # 🎯 企業級修復：提交通知延遲到映射創建後發送
            # 不創建Unknown映射，而是等待正確映射創建後再發送通知
            octype_info = {
                'octype': 'PENDING_SUBMIT',  # 標記為等待提交處理
                'direction': 'PENDING_SUBMIT',
                'contract_name': contract_name,
                'order_type': 'PENDING_SUBMIT',
                'price_type': 'PENDING_SUBMIT',
                'is_manual': True,
                'requires_delay': True  # 標記需要延遲處理
            }
            
            # 下面的同步應急推斷邏輯被移除，改為延遲處理
            """
            ⚠️ 應急推斷機制已移除，改為延遲機制處理
            try:
                # 🔍 詳細記錄API數據結構以便調試
                logger.info(f"🔍 完整deal數據: {deal}")
                logger.info(f"🔍 order參數內容: {order}")
                
                # 🎯 重新設計：多路徑數據提取，支援永豐金API的複雜結構
                # 優先順序：order參數 > deal.order > deal直接屬性 > 預設值
                
                # 🎯 Action (Buy/Sell) 強化提取和驗證
                action = None
                action_sources = []  # 記錄所有可能的數據源
                
                # 檢查所有可能的數據源
                if order and hasattr(order, 'action'):
                    order_action = str(order.action).replace('Action.', '').upper()
                    action_sources.append(f"order參數: {order_action}")
                elif order and isinstance(order, dict) and order.get('action'):
                    order_action = str(order.get('action')).upper()
                    action_sources.append(f"order字典: {order_action}")
                    
                if deal.get('order', {}).get('action'):
                    deal_order_action = str(deal.get('order', {}).get('action')).upper()
                    action_sources.append(f"deal.order: {deal_order_action}")
                    
                if deal.get('action'):
                    deal_action = str(deal.get('action')).upper()
                    action_sources.append(f"deal直接: {deal_action}")
                
                logger.info(f"🔍 所有可能的action數據源: {action_sources}")
                
                # 選擇最可靠的數據源 (優先順序: deal.order > order參數 > deal直接)
                if deal.get('order', {}).get('action'):
                    action = str(deal.get('order', {}).get('action')).upper()
                    logger.info(f"🔍 使用deal.order的action: {action}")
                elif order and hasattr(order, 'action'):
                    action = str(order.action).replace('Action.', '').upper()
                    logger.info(f"🔍 使用order參數的action: {action}")
                elif order and isinstance(order, dict) and order.get('action'):
                    action = str(order.get('action')).upper()
                    logger.info(f"🔍 使用order字典的action: {action}")
                elif deal.get('action'):
                    action = str(deal.get('action')).upper()
                    logger.info(f"🔍 使用deal直接的action: {action}")
                else:
                    action = 'SELL'  # 預設值
                    logger.warning(f"🔍 使用預設action: {action}")
                
                # ⚠️ 數據一致性驗證
                if len(action_sources) > 1:
                    unique_actions = set([src.split(': ')[1] for src in action_sources])
                    if len(unique_actions) > 1:
                        logger.error(f"🚨 Action數據不一致! {action_sources}")
                        logger.error(f"🚨 建議人工檢查API返回數據")
                
                # OrderType 提取
                extracted_order_type = None
                if order and hasattr(order, 'order_type'):
                    extracted_order_type = str(order.order_type).replace('FuturesOrderType.', '').upper()
                    logger.info(f"🔍 從order參數提取order_type: {extracted_order_type}")
                elif deal.get('order', {}).get('order_type'):
                    extracted_order_type = str(deal.get('order', {}).get('order_type')).upper()
                    logger.info(f"🔍 從deal.order提取order_type: {extracted_order_type}")
                else:
                    extracted_order_type = 'IOC'
                    logger.warning(f"🔍 使用預設order_type: {extracted_order_type}")
                
                # PriceType 提取  
                extracted_price_type = None
                if order and hasattr(order, 'price_type'):
                    extracted_price_type = str(order.price_type).replace('FuturesPriceType.', '').upper()
                    logger.info(f"🔍 從order參數提取price_type: {extracted_price_type}")
                elif deal.get('order', {}).get('price_type'):
                    extracted_price_type = str(deal.get('order', {}).get('price_type')).upper()
                    logger.info(f"🔍 從deal.order提取price_type: {extracted_price_type}")
                else:
                    extracted_price_type = 'MKT'
                    logger.warning(f"🔍 使用預設price_type: {extracted_price_type}")
                
                # 只做臨時推斷，不保存到映射中，避免覆蓋正確資料
                positions = sinopac_api.list_positions(sinopac_api.futopt_account)
                current_position = 0
                for pos in positions:
                    if pos.code.startswith(contract_code[:3]):
                        current_position += float(pos.quantity)
                
                # 🎯 重新設計開平倉判斷邏輯 - 確保action映射正確
                if action.upper() == 'BUY':
                    oc_type = 'Cover' if current_position < 0 else 'New'
                elif action.upper() == 'SELL':
                    oc_type = 'Cover' if current_position > 0 else 'New'
                else:
                    # 未知動作，根據持倉推斷
                    oc_type = 'Cover' if current_position != 0 else 'New'
                    logger.warning(f"🔧 未知動作 {action}，根據持倉推斷")
                    
                logger.warning(f"🔧 應急推斷: 持倉={current_position}, 動作={action}, 判斷={oc_type}")
                
                # 🎯 訂單類型轉換 - 使用提取的真實數據
                real_order_type = 'IOC'
                real_price_type = 'MKT'
                
                # 使用前面提取的數據
                logger.info(f"🔍 提取的原始數據: order_type={extracted_order_type}, price_type={extracted_price_type}")
                
                # 永豐API訂單類型轉換
                if extracted_price_type and extracted_price_type.upper() in ['LMT', 'LIMIT']:
                    real_price_type = 'LMT'
                    # 限價單時檢查時效類型
                    if extracted_order_type and extracted_order_type.upper() in ['ROD', 'REST_OF_DAY']:
                        real_order_type = 'ROD'
                    else:
                        real_order_type = 'IOC'
                elif extracted_price_type and extracted_price_type.upper() in ['MKT', 'MARKET']:
                    real_price_type = 'MKT'
                    real_order_type = 'IOC'
                
                logger.info(f"🎯 最終轉換結果: order_type={real_order_type}, price_type={real_price_type}")
                
                # 只使用臨時資訊，不保存到全局映射
                octype_info = {
                    'octype': oc_type,
                    'direction': action.upper(),  # 確保統一格式
                    'contract_name': contract_name, 
                    'order_type': real_order_type,
                    'price_type': real_price_type,
                    'is_manual': True,  # 應急推斷預設為手動
                    'emergency_inferred': True  # 標記為應急推斷
                }
            """
            # 舊的應急推斷邏輯結束
        else:
            logger.info(f"✅ 從映射中找到訂單資訊: {octype_info}")
            # 檢查是否為預先保存的資料
            if octype_info.get('pre_saved'):
                logger.info("🔒 使用預先保存的訂單資訊，確保時序一致性")
        
        # ✅ 確保使用一致的訂單資訊，不再重新推斷
        octype = octype_info['octype']
        direction = octype_info['direction']
        order_type = octype_info['order_type']
        price_type = octype_info['price_type']
        is_manual = octype_info.get('is_manual', False)
        
        # 記錄資料來源供除錯參考
        data_source = "應急推斷" if octype_info.get('emergency_inferred') else "預先保存"
        logger.info(f"📋 使用{data_source}資訊: octype={octype}, direction={direction}, manual={is_manual}")
        
        
        # 獲取訂單數量和操作信息
        qty = deal.get('order', {}).get('quantity', 0)
        op_code = deal.get('operation', {}).get('op_code', '00')
        op_msg = deal.get('operation', {}).get('op_msg', '')
        op_type = deal.get('operation', {}).get('op_type', '')
        
        current_time = datetime.now().strftime('%Y/%m/%d %H:%M')
        
        # 處理成交回調 - 🎯 修復：如果沒有映射或需要延遲，延遲處理等待提交回調創建映射
        if str(state) == 'OrderState.FuturesDeal' and deal.get('code') and deal.get('quantity'):
            if octype_info.get('requires_delay', False):
                # 沒有映射，延遲3秒處理，給提交回調時間創建映射
                logger.warning(f"🔄 成交回調沒有映射，延遲3秒等待提交回調創建映射: {order_id}")
                def delayed_deal_callback():
                    time.sleep(3)
                    # 重新查詢映射
                    delayed_octype_info = order_octype_map.get(order_id)
                    if delayed_octype_info:
                        logger.info(f"✅ 延遲後找到映射，處理成交回調: {order_id}")
                        handle_futures_deal_callback(deal, delayed_octype_info)
                    else:
                        logger.warning(f"⚠️ 延遲後仍無映射，使用應急推斷處理: {order_id}")
                        # 重新執行應急推斷邏輯
                        emergency_octype_info = None
                        # 重新查詢訂單映射
                        emergency_octype_info = order_octype_map.get(order_id)
                        if emergency_octype_info is None:
                            # 執行應急推斷（從前面複製）
                            try:
                                # 詳細記錄API數據結構以便調試
                                logger.info(f"🔍 延遲處理完整deal數據: {deal}")
                                deal_order = deal.get('order', {})
                                logger.info(f"🔍 延遲處理deal.order內容: {deal_order}")
                                
                                # 🎯 延遲處理：同樣使用永豐金API提供的完整資訊
                                deal_order = deal.get('order', {})
                                action = str(deal_order.get('action', deal.get('action', 'Sell'))).upper()
                                emergency_order_type = str(deal_order.get('order_type', 'IOC')).upper()
                                emergency_price_type = str(deal_order.get('price_type', 'MKT')).upper()
                                oc_type = str(deal_order.get('oc_type', 'New')).upper()  # 🔥 使用永豐金提供的開平倉
                                
                                logger.info(f"🔍 延遲處理API原始數據:")
                                logger.info(f"  - action: {action}")
                                logger.info(f"  - oc_type: {oc_type} (永豐金提供)")
                                logger.info(f"  - order_type: {emergency_order_type}")
                                logger.info(f"  - price_type: {emergency_price_type}")
                                
                                logger.info(f"🎯 延遲處理使用永豐金API提供的開平倉資訊: {oc_type} (無需推斷)")
                                
                                # ✅ 直接使用永豐金提供的完整資訊，無需額外轉換
                                emergency_octype_info = {
                                    'octype': oc_type,  # 直接使用永豐金的判斷
                                    'direction': action,
                                    'contract_name': contract_name, 
                                    'order_type': emergency_order_type,  # 直接使用永豐金提供的
                                    'price_type': emergency_price_type,  # 直接使用永豐金提供的
                                    'is_manual': True,
                                    'emergency_inferred': True
                                }
                                
                            except Exception as e:
                                logger.error(f"延遲應急推斷失敗: {e}")
                                emergency_octype_info = {
                                    'octype': 'New',
                                    'direction': 'Sell',
                                    'contract_name': contract_name,
                                    'order_type': 'IOC',
                                    'price_type': 'MKT',
                                    'is_manual': True,
                                    'emergency_inferred': True
                                }
                        
                        handle_futures_deal_callback(deal, emergency_octype_info)
                
                create_managed_thread(target=delayed_deal_callback, name=f"延遲成交回調-{order_id}").start()
            else:
                # 有映射，正常處理
                handle_futures_deal_callback(deal, octype_info)
            
        # 處理訂單提交回調
        elif str(state) in ['OrderState.Submitted', 'OrderState.FuturesOrder']:
            if op_type == 'Cancel' or op_code != '00':
                # 訂單失敗或取消
                if op_type == 'Cancel':
                    if order_type == 'IOC' and not op_msg:
                        fail_reason = "價格未滿足"
                    else:
                        fail_reason = "手動取消掛單"  # 手動取消訂單
                else:
                    fail_reason = OP_MSG_TRANSLATIONS.get(op_msg, op_msg or '未知錯誤')
                
                price_value = order.get('price', 0.0) if order else deal.get('order', {}).get('price', 0.0)
                
                # 獲取完整合約代碼和交割日期用於失敗通知
                full_contract_code = None
                delivery_date_for_fail = None
                try:
                    # 根據合約代碼前綴找到對應的全域合約對象
                    if contract_code.startswith('TXF') and contract_txf:
                        full_contract_code = contract_txf.code
                        if hasattr(contract_txf, 'delivery_date'):
                            if hasattr(contract_txf.delivery_date, 'strftime'):
                                delivery_date_for_fail = contract_txf.delivery_date.strftime('%Y/%m/%d')
                            else:
                                delivery_date_for_fail = str(contract_txf.delivery_date)
                    elif contract_code.startswith('MXF') and contract_mxf:
                        full_contract_code = contract_mxf.code
                        if hasattr(contract_mxf, 'delivery_date'):
                            if hasattr(contract_mxf.delivery_date, 'strftime'):
                                delivery_date_for_fail = contract_mxf.delivery_date.strftime('%Y/%m/%d')
                            else:
                                delivery_date_for_fail = str(contract_mxf.delivery_date)
                    elif contract_code.startswith('TMF') and contract_tmf:
                        full_contract_code = contract_tmf.code
                        if hasattr(contract_tmf, 'delivery_date'):
                            if hasattr(contract_tmf.delivery_date, 'strftime'):
                                delivery_date_for_fail = contract_tmf.delivery_date.strftime('%Y/%m/%d')
                            else:
                                delivery_date_for_fail = str(contract_tmf.delivery_date)
                except:
                    pass
                
                # 保存交易記錄（參考TXserver.py）
                save_trade({
                    'type': 'order',
                    'trade_category': 'normal',
                    'raw_data': deal,
                    'deal_order_id': order_id,
                    'contract_name': contract_name,
                    'timestamp': datetime.now().isoformat(),
                    'is_manual': is_manual
                })
                
                # 保存交易記錄（參考TXserver.py）
                save_trade({
                    'type': 'cancel',
                    'trade_category': 'normal',
                    'raw_data': deal,
                    'deal_order_id': order_id,
                    'contract_name': contract_name,
                    'timestamp': datetime.now().isoformat(),
                    'is_manual': is_manual
                })
                
                # 記錄掛單失敗日誌
                log_message = get_simple_order_log_message(
                    contract_name=contract_name,
                    direction=direction,
                    qty=qty,
                    price=price_value,
                    order_id=order_id,
                    octype=octype,
                    is_manual=is_manual,
                    is_success=False,
                    order_type=order_type,
                    price_type=price_type
                )
                # 系統日誌記錄（JSON數據已註解，暫時移除POST請求）
                # try:
                #     requests.post(
                #         f'http://127.0.0.1:{CURRENT_PORT}/api/system_log',
                #         json={'message': log_message, 'type': 'error'},
                #         timeout=5
                #     )
                # except:
                #     pass
                
                # 發送失敗通知
                msg = get_formatted_order_message(
                    is_success=False,
                    order_id=order_id,
                    contract_name=contract_name,
                    qty=qty,
                    price=price_value,
                    octype=octype,
                    direction=direction,
                    order_type=order_type,
                    price_type=price_type,
                    is_manual=is_manual,
                    reason=fail_reason,
                    contract_code=full_contract_code or contract_code,
                    delivery_date=delivery_date_for_fail
                )
                send_telegram_message(msg)
                
                # 清理映射
                with global_lock:
                    order_octype_map.pop(order_id, None)
            else:
                # 訂單提交成功
                price_value = order.get('price', 0.0) if order else deal.get('order', {}).get('price', 0.0)
                
                # 獲取完整合約代碼和交割日期
                full_contract_code = None
                delivery_date = None
                try:
                    # 首先嘗試從deal或order獲取delivery_month
                    delivery_month = None
                    if order and order.get('delivery_month'):
                        delivery_month = order.get('delivery_month')
                    elif deal and deal.get('contract', {}).get('delivery_month'):
                        delivery_month = deal.get('contract', {}).get('delivery_month')
                    
                    if delivery_month and len(delivery_month) == 6:
                        year = int(delivery_month[:4])
                        month = int(delivery_month[4:6])
                        # 計算該月的第三個星期三（台指期貨交割日）
                        third_wednesday = get_third_wednesday(year, month)
                        delivery_date = f"{year}/{month:02d}/{third_wednesday:02d}"
                    else:
                        # 如果沒有delivery_month，從全域合約對象獲取
                        if contract_code.startswith('TXF') and contract_txf:
                            full_contract_code = contract_txf.code
                            if hasattr(contract_txf, 'delivery_date'):
                                if hasattr(contract_txf.delivery_date, 'strftime'):
                                    delivery_date = contract_txf.delivery_date.strftime('%Y/%m/%d')
                                else:
                                    delivery_date = str(contract_txf.delivery_date)
                        elif contract_code.startswith('MXF') and contract_mxf:
                            full_contract_code = contract_mxf.code
                            if hasattr(contract_mxf, 'delivery_date'):
                                if hasattr(contract_mxf.delivery_date, 'strftime'):
                                    delivery_date = contract_mxf.delivery_date.strftime('%Y/%m/%d')
                                else:
                                    delivery_date = str(contract_mxf.delivery_date)
                        elif contract_code.startswith('TMF') and contract_tmf:
                            full_contract_code = contract_tmf.code
                            if hasattr(contract_tmf, 'delivery_date'):
                                if hasattr(contract_tmf.delivery_date, 'strftime'):
                                    delivery_date = contract_tmf.delivery_date.strftime('%Y/%m/%d')
                                else:
                                    delivery_date = str(contract_tmf.delivery_date)
                except:
                    pass
                
                # 保存交易記錄（參考TXserver.py）
                save_trade({
                    'type': 'order',
                    'trade_category': 'normal',
                    'raw_data': deal,  # 保存完整的deal對象，包含order信息
                    'deal_order_id': order_id,
                    'contract_name': contract_name,
                    'timestamp': datetime.now().isoformat(),
                    'is_manual': is_manual
                })
                
                # 記錄掛單成功日誌
                log_message = get_simple_order_log_message(
                    contract_name=contract_name,
                    direction=direction,
                    qty=qty,
                    price=price_value,
                    order_id=order_id,
                    octype=octype,
                    is_manual=is_manual,
                    is_success=False,
                    order_type=order_type,
                    price_type=price_type
                )
                # 系統日誌記錄（JSON數據已註解，暫時移除POST請求）
                # try:
                #     requests.post(
                #         f'http://127.0.0.1:{CURRENT_PORT}/api/system_log',
                #         json={'message': log_message, 'type': 'success'},
                #         timeout=5
                #     )
                # except:
                #     pass
                
                # 🎯 關鍵修復：手機APP下單時需要先創建映射，再發送通知
                if order_id not in order_octype_map:
                    # 手機APP下單，需要從提交回調的deal數據中提取正確資訊
                    logger.info(f"🔍 提交回調完整deal數據: {deal}")
                    
                    # 🎯 企業級修復：直接使用永豐金API提供的完整資訊
                    deal_order = deal.get('order', {})
                    real_action = str(deal_order.get('action', 'SELL')).upper()
                    real_order_type = str(deal_order.get('order_type', 'IOC')).upper()
                    real_price_type = str(deal_order.get('price_type', 'MKT')).upper()
                    real_octype = str(deal_order.get('oc_type', 'New')).upper()  # 🔥 關鍵：使用永豐金提供的開平倉資訊
                    
                    logger.info(f"🔍 提交回調API原始數據:")
                    logger.info(f"  - action: {real_action}")
                    logger.info(f"  - oc_type: {real_octype} (永豐金提供)")
                    logger.info(f"  - order_type: {real_order_type}")
                    logger.info(f"  - price_type: {real_price_type}")
                    
                    # ✅ 不再需要持倉查詢和推斷！直接使用永豐金的判斷
                    logger.info(f"🎯 使用永豐金API提供的開平倉資訊: {real_octype} (無需自行推斷)")
                    
                    with global_lock:
                        order_octype_map[order_id] = {
                            'octype': real_octype,
                            'direction': real_action,
                            'contract_name': contract_name,
                            'order_type': real_order_type,
                            'price_type': real_price_type,
                            'is_manual': is_manual,
                            'created_at_submit': True  # 標記為提交時創建
                        }
                    logger.info(f"✅ 手機APP下單：在提交成功時創建訂單映射: {order_id}")
                    logger.info(f"✅ 映射內容: {order_octype_map[order_id]}")
                
                # 保存訂單映射到文件
                if order_id in order_octype_map:
                    save_order_mapping()
                    logger.info(f"✅ 提交成功後保存訂單映射: {order_id}")
                    
                    # ✅ 現在用正確的映射資訊生成提交通知
                    mapping_info = order_octype_map[order_id]
                    msg = get_formatted_order_message(
                        is_success=True,
                        order_id=order_id,
                        contract_name=contract_name,
                        qty=qty,
                        price=price_value,
                        octype=mapping_info.get('octype', octype),
                        direction=mapping_info.get('direction', direction),
                        order_type=mapping_info.get('order_type', order_type),
                        price_type=mapping_info.get('price_type', price_type),
                        is_manual=mapping_info.get('is_manual', is_manual),
                        contract_code=full_contract_code or contract_code,
                        delivery_date=delivery_date
                    )
                    logger.info(f"🎯 提交通知使用正確映射資訊: octype={mapping_info.get('octype')}, direction={mapping_info.get('direction')}")
                    
                    # 延遲1秒發送提交成功通知，避免與其他通知重疊
                    def delayed_submit_notification():
                        time.sleep(1)
                        send_telegram_message(msg)
                    
                    create_managed_thread(target=delayed_submit_notification, name="延遲提交通知線程").start()
                else:
                    logger.warning(f"⚠️ 提交成功後仍找不到訂單映射，跳過提交通知: {order_id}")
        
    except Exception as e:
        logger.error(f"回調函數處理失敗: {e}")

def handle_futures_deal_callback(deal, octype_info):
    """處理期貨成交回調"""
    global order_octype_map, contract_txf, contract_mxf, contract_tmf
    
    try:
        order_id = deal.get('trade_id', deal.get('order_id', '未知'))
        contract_code = deal.get('code', '')
        contract_name = get_contract_name_from_code(contract_code)
        
        deal_price = deal.get('price', 0)
        deal_quantity = deal.get('quantity', 0)
        
        # 增強數據收集用於XLSX生成
        deal_number = deal.get('trade_id', deal.get('deal_id', deal.get('id', order_id)))  # 成交單號
        ts = deal.get('ts', deal.get('timestamp', int(time.time() * 1000)))  # 成交時間戳
        
        # 收集所有可能的訂單類型資訊
        original_order_type = deal.get('order_type', '')
        original_price_type = deal.get('price_type', '')
        
        logger.info(f"[數據收集] 成交數據完整性檢查:")
        logger.info(f"  - 成交單號: {deal_number}")
        logger.info(f"  - 訂單ID: {order_id}")
        logger.info(f"  - 合約代碼: {contract_code}")
        logger.info(f"  - 成交價格: {deal_price}")
        logger.info(f"  - 成交數量: {deal_quantity}")
        logger.info(f"  - 原始訂單類型: {original_order_type}")
        logger.info(f"  - 原始價格類型: {original_price_type}")
        logger.info(f"  - 時間戳: {ts}")
        
        current_time = datetime.now().strftime('%Y/%m/%d %H:%M')
        
        # 直接使用octype_info中已經處理好的資訊
        octype = octype_info.get('octype', 'New')
        direction = octype_info.get('direction', 'Sell')
        order_type = octype_info.get('order_type', 'IOC')
        price_type = octype_info.get('price_type', 'MKT')
        is_manual = octype_info.get('is_manual', True)
        
        logger.info(f"🎯 使用訂單資訊: octype={octype}, direction={direction}, order_type={order_type}, price_type={price_type}")
        
        
        # 獲取完整合約代碼和交割日期用於成交通知
        # 使用轉倉邏輯選擇合約
        full_contract_code = None
        delivery_date_for_deal = None
        try:
            # 根據合約代碼前綴和轉倉模式找到對應的合約對象
            target_contract = None
            if contract_code.startswith('TXF'):
                target_contract = get_contract_for_rollover('TXF')
            elif contract_code.startswith('MXF'):
                target_contract = get_contract_for_rollover('MXF')
            elif contract_code.startswith('TMF'):
                target_contract = get_contract_for_rollover('TMF')
            
            if target_contract:
                full_contract_code = target_contract.code
                if hasattr(target_contract, 'delivery_date'):
                    delivery_date_for_deal = format_delivery_date(target_contract.delivery_date)
            
            # 如果無法獲取交割日期，嘗試從 delivery_month 計算
            if not delivery_date_for_deal:
                # 從其他來源獲取 delivery_month 或其他交割日期信息
                if hasattr(deal, 'contract') and hasattr(deal.contract, 'delivery_month'):
                    delivery_month = deal.contract.delivery_month
                    if delivery_month and len(delivery_month) == 6:
                        year = int(delivery_month[:4])
                        month = int(delivery_month[4:6])
                        # 計算該月的第三個星期三（台指期貨交割日）
                        third_wednesday = get_third_wednesday(year, month)
                        delivery_date_for_deal = f"{year}/{month:02d}/{third_wednesday:02d}"
        except:
            pass
        
        # 記錄成交成功日誌（延遲5秒，與TG通知同步）
        log_message = get_simple_order_log_message(
            contract_name=contract_name,
            direction=direction,
            qty=deal_quantity,
            price=deal_price,
            order_id=order_id,
            octype=octype,
            is_manual=is_manual,
            is_success=True,
            order_type=order_type,
            price_type=price_type
        )
        
        def delayed_log():
            time.sleep(5)
            # 系統日誌記錄（JSON數據已註解，暫時移除POST請求）
            # try:
            #     requests.post(
            #         f'http://127.0.0.1:{CURRENT_PORT}/api/system_log',
            #         json={'message': log_message, 'type': 'success'},
            #         timeout=5
            #     )
            # except:
            #     pass
        
        create_managed_thread(target=delayed_log, name="延遲日誌線程").start()
        
        # 發送成交通知 - 使用映射中的統一資訊
        msg = get_formatted_trade_message(
            order_id=order_id,
            contract_name=octype_info.get('contract_name', contract_name),
            qty=deal_quantity,
            price=deal_price,
            octype=octype_info.get('octype', 'New'),
            direction=octype_info.get('direction', 'Sell'),
            order_type=octype_info.get('order_type', 'IOC'),
            price_type=octype_info.get('price_type', 'MKT'),
            is_manual=octype_info.get('is_manual', True),
            contract_code=full_contract_code or contract_code,
            delivery_date=delivery_date_for_deal
        )
        
        # 使用JSON配對記錄系統
        from trade_pairing_TX import record_opening_trade, record_covering_trade
        
        try:
            if octype.upper() == 'NEW':
                # 記錄開倉交易
                trade_id = record_opening_trade(
                    contract_code=contract_code,
                    action=direction.title(),  # Buy/Sell
                    quantity=deal_quantity,
                    price=deal_price,  # 使用實際成交價格
                    order_id=order_id
                )
                logger.info(f"✅ 開倉記錄已建立: {trade_id}")
                
            elif octype.upper() == 'COVER':
                # 記錄平倉交易並自動配對
                cover_record = record_covering_trade(
                    contract_code=contract_code,
                    action=direction.title(),  # Buy/Sell  
                    quantity=deal_quantity,
                    price=deal_price,  # 使用實際成交價格
                    order_id=order_id
                )
                if cover_record:
                    logger.info(f"✅ 平倉記錄已建立並完成配對，總損益: {cover_record.get('total_pnl', 0)}")
                
        except Exception as e:
            logger.error(f"JSON配對記錄失敗: {e}")
            import traceback
            traceback.print_exc()
        
        # 保存成交記錄（增強版本，包含成交單號、交割日期等完整資訊）
        save_trade({
            'type': 'deal',
            'trade_category': 'normal',
            'raw_data': {
                'operation': {
                    'op_type': 'Deal',
                    'op_code': '00',
                    'op_msg': ''
                },
                'order': {
                    'id': order_id,
                    'action': direction,
                    'price': deal_price,
                    'quantity': deal_quantity,
                    'oc_type': octype,
                    'order_type': order_type,
                    'price_type': price_type
                },
                'contract': {
                    'code': contract_code
                },
                'deal': {
                    'deal_number': deal_number,  # 成交單號
                    'original_order_type': original_order_type,  # 原始訂單類型
                    'original_price_type': original_price_type,  # 原始價格類型
                    'timestamp': ts,  # 原始時間戳
                    'delivery_date': delivery_date_for_deal  # 交割日期
                }
            },
            'deal_order_id': order_id,
            'deal_number': deal_number,  # 頂層成交單號欄位供XLSX使用
            'contract_name': contract_name,
            'contract_code': contract_code,  # 完整合約代碼
            'delivery_date': delivery_date_for_deal,  # 頂層交割日期欄位供XLSX使用
            'order_type_display': f"市價單（{order_type}）" if price_type == 'MKT' else f"限價單（{order_type}）" if price_type == 'LMT' else f"{price_type}（{order_type}）",  # 格式化的訂單類型顯示
            'timestamp': datetime.now().isoformat(),
            'deal_timestamp': ts,  # 實際成交時間戳
            'is_manual': is_manual
        })
        
        # TX系統部位追蹤
        try:
            trade_data_for_tracking = {
                'deal_order_id': order_id,
                'timestamp': datetime.now().isoformat(),
                'raw_data': {
                    'order': {
                        'action': direction,
                        'oc_type': octype,
                        'quantity': deal_quantity,
                        'price': deal_price
                    },
                    'contract': {
                        'code': contract_code
                    }
                }
            }
            
            if octype.upper() == 'NEW':
                # 開倉：保存未平倉部位
                save_tx_open_position(trade_data_for_tracking)
            elif octype.upper() == 'COVER':
                # 平倉：更新未平倉部位
                update_tx_position_on_close(trade_data_for_tracking)
                
        except Exception as e:
            logger.error(f"TX部位追蹤失敗: {str(e)}")
        
        # 延遲5秒發送成交通知，確保在提交通知之後，並在發送後清理映射
        def delayed_send_and_cleanup():
            time.sleep(5)
            send_telegram_message(msg)
            # 成交通知發送後才清理映射 - 確保線程安全
            with global_lock:
                if order_id in order_octype_map:
                    order_octype_map.pop(order_id, None)
                    logger.info(f"✅ 成交通知發送後清理訂單映射: {order_id}")
                else:
                    logger.warning(f"⚠️ 成交通知發送後映射已不存在: {order_id}")
        
        create_managed_thread(target=delayed_send_and_cleanup, name="延遲發送清理線程").start()
            
    except Exception as e:
        logger.error(f"處理期貨成交回調失敗: {e}")

def get_contract_name_from_code(contract_code):
    """根據合約代碼獲取合約名稱"""
    if not contract_code:
        return "未知"
    
    if contract_code.startswith('TXF'):
        return "大台"
    elif contract_code.startswith('MXF'):
        return "小台"
    elif contract_code.startswith('TMF'):
        return "微台"
    else:
        return "未知"

def get_contract_display_with_delivery(contract_code, delivery_month):
    """統一的合約顯示函數，包含合約代碼和交割日期
    
    Args:
        contract_code: 合約代碼 (如 'TXF' 或 'TXF202508')
        delivery_month: 到期月份 (如 '202508' 或 '')
    
    Returns:
        格式化的合約顯示字串 (如 'TXF (2025/08/20)' 或 'TXF (日期未知)')
    """
    try:
        # 提取基本合約代碼 (如 TXF202508 → TXF)
        base_contract_code = contract_code[:3] if len(contract_code) >= 3 else contract_code
        
        if delivery_month and len(delivery_month) == 6:
            # 轉換 202508 格式為 2025/08/20 (第三個星期三)
            year = delivery_month[:4]
            month = delivery_month[4:6]
            
            # 計算該月的第三個星期三
            first_day = datetime(int(year), int(month), 1)
            # 找到第一個星期三 (weekday 2 = 星期三)
            days_until_wednesday = (2 - first_day.weekday()) % 7
            first_wednesday = first_day + timedelta(days=days_until_wednesday)
            # 第三個星期三 = 第一個星期三 + 14天
            third_wednesday = first_wednesday + timedelta(days=14)
            formatted_date = third_wednesday.strftime('%Y/%m/%d')
            
            return f"{base_contract_code} ({formatted_date})"
        elif delivery_month:
            # 其他格式的日期，嘗試用現有函數格式化
            formatted_date = format_delivery_date(delivery_month)
            return f"{base_contract_code} ({formatted_date})"
        else:
            # 最後備用方案：使用當前月份
            from datetime import datetime
            current_date = datetime.now()
            current_year = current_date.year
            current_month = current_date.month
            
            # 計算當月第三個星期三
            first_day = datetime(current_year, current_month, 1)
            days_until_wednesday = (2 - first_day.weekday()) % 7
            first_wednesday = first_day + timedelta(days=days_until_wednesday)
            third_wednesday = first_wednesday + timedelta(days=14)
            formatted_date = third_wednesday.strftime('%Y/%m/%d')
            
            logger.warning(f"無法獲取交割日期，使用當前月份: {base_contract_code} ({formatted_date})")
            return f"{base_contract_code} ({formatted_date})"
    except Exception as e:
        logger.warning(f"格式化合約顯示失敗: {contract_code}, {delivery_month}, 錯誤: {e}")
        # 提取基本合約代碼作為備用
        base_contract_code = contract_code[:3] if len(contract_code) >= 3 else contract_code
        # 使用當前月份作為備用日期
        try:
            current_date = datetime.now()
            current_year = current_date.year
            current_month = current_date.month
            
            # 計算當月第三個星期三
            first_day = datetime(current_year, current_month, 1)
            days_until_wednesday = (2 - first_day.weekday()) % 7
            first_wednesday = first_day + timedelta(days=days_until_wednesday)
            third_wednesday = first_wednesday + timedelta(days=14)
            fallback_date = third_wednesday.strftime('%Y/%m/%d')
            
            logger.info(f"使用當月第三個星期三作為備用: {base_contract_code} ({fallback_date})")
            return f"{base_contract_code} ({fallback_date})"
        except:
            logger.error(f"備用日期計算也失敗，返回基本合約: {base_contract_code}")
            return f"{base_contract_code} (日期未知)"

# 新增：參考TXserver.py的動作顯示邏輯
def get_action_display_by_rule(octype, direction):
    """根據開平倉類型和方向判斷動作顯示，企業級交易系統邏輯"""
    # 統一轉換為大寫進行比較
    octype_upper = str(octype).upper()
    direction_upper = str(direction).upper()
    
    # 🎯 企業級修復：根據開平倉類型和方向的組合判斷部位方向
    # 核心邏輯：部位方向 = 開倉時的交易方向，平倉時相反
    if octype_upper == 'NEW':  # 開倉
        # 開倉時：Buy=多單部位, Sell=空單部位
        if direction_upper == 'BUY':
            return '多單'
        else:  # SELL
            return '空單'
    elif octype_upper == 'COVER':  # 平倉
        # 平倉時：需要看原本的部位方向
        # Buy平倉 = 平空單部位 -> 顯示空單
        # Sell平倉 = 平多單部位 -> 顯示多單
        if direction_upper == 'BUY':
            return '空單'  # 買入平倉 = 平空單
        else:  # SELL
            return '多單'  # 賣出平倉 = 平多單
    else:
        # 未知開平倉類型，回退到原始邏輯
        if direction_upper == 'BUY':
            return '多單'
        else:  # SELL
            return '空單'

def get_formatted_order_message(is_success, order_id, contract_name, qty, price, octype, direction, order_type, price_type, is_manual, reason=None, contract_code=None, delivery_date=None):
    """格式化訂單提交訊息（參考TXserver.py）"""
    current_time = datetime.now().strftime('%Y/%m/%d')
    
    # 獲取完整合約資訊
    try:
        if contract_code and delivery_date:
            # 簡化合約代碼顯示：TMFG5 -> TMF, MXFG5 -> MXF, TXFG5 -> TXF
            if contract_code.startswith('TMF'):
                simple_code = 'TMF'
            elif contract_code.startswith('MXF'):
                simple_code = 'MXF'
            elif contract_code.startswith('TXF'):
                simple_code = 'TXF'
            else:
                simple_code = contract_code
            contract_display = f"{simple_code} ({delivery_date})"
        elif contract_code:
            # 簡化合約代碼顯示
            if contract_code.startswith('TMF'):
                simple_code = 'TMF'
            elif contract_code.startswith('MXF'):
                simple_code = 'MXF'
            elif contract_code.startswith('TXF'):
                simple_code = 'TXF'
            else:
                simple_code = contract_code
            contract_display = f"{simple_code} (日期未知)"
        else:
            contract_display = "未知"
    except:
        contract_display = "未知"
    
    # 訂單類型顯示
    try:
        if str(price_type).upper() == 'MKT':
            price_text = "市價單"
        else:
            price_text = "限價單"
        
        if str(order_type).upper() == 'IOC':
            type_text = "IOC"
        else:
            type_text = "ROD"
        
        order_type_display = f"{price_text}（{type_text}）"
    except:
        order_type_display = f"未知 ({order_type})"
    
    # 提交類型 - 移除手動/自動標示，只顯示開倉/平倉
    if str(octype).upper() == 'NEW':
        submit_type = "開倉"
    elif str(octype).upper() == 'COVER':
        submit_type = "平倉"
    else:
        submit_type = f"未知({octype})"  # 不預設為開倉，顯示實際值
    
    # 提交動作 - 使用TXserver.py的get_action_display_by_rule邏輯
    submit_action = get_action_display_by_rule(str(octype).upper(), str(direction).upper())
    
    # 提交價格
    if str(price_type).upper() == "MKT":
        price_display = "市價"
    else:
        price_display = f"{price:.0f}"
    
    # 失敗原因翻譯 - 參考TXserver.py的OP_MSG_TRANSLATIONS
    if reason:
        translated_reason = OP_MSG_TRANSLATIONS.get(reason, reason)
    else:
        translated_reason = reason
    
    if is_success:
        msg = (f"⭕ 提交成功（{current_time}）\n"
               f"選用合約：{contract_display}\n"
               f"訂單類型：{order_type_display}\n"
               f"提交單號：{order_id}\n"
               f"提交類型：{submit_type}\n"
               f"提交動作：{submit_action}\n"
               f"提交部位：{contract_name}\n"
               f"提交數量：{qty} 口\n"
               f"提交價格：{price_display}")
    else:
        msg = (f"❌ 提交失敗（{current_time}）\n"
               f"選用合約：{contract_display}\n"
               f"訂單類型：{order_type_display}\n"
               f"提交單號：{order_id if order_id != '未知' else '未知'}\n"
               f"提交類型：{submit_type}\n"
               f"提交動作：{submit_action}\n"
               f"提交部位：{contract_name}\n"
               f"提交數量：{qty} 口\n"
               f"提交價格：{price_display}\n"
               f"原因：{translated_reason}")
    
    return msg

def get_formatted_trade_message(order_id, contract_name, qty, price, octype, direction, order_type, price_type, is_manual, contract_code=None, delivery_date=None):
    """格式化成交訊息"""
    current_time = datetime.now().strftime('%Y/%m/%d')
    
    # 獲取完整合約資訊 - 修正：簡化合約代碼顯示
    try:
        if contract_code and delivery_date:
            # 簡化合約代碼顯示：TMFG5 -> TMF, MXFG5 -> MXF, TXFG5 -> TXF
            if contract_code.startswith('TMF'):
                simple_code = 'TMF'
            elif contract_code.startswith('MXF'):
                simple_code = 'MXF'
            elif contract_code.startswith('TXF'):
                simple_code = 'TXF'
            else:
                simple_code = contract_code
            contract_display = f"{simple_code} ({delivery_date})"
        elif contract_code:
            # 簡化合約代碼顯示
            if contract_code.startswith('TMF'):
                simple_code = 'TMF'
            elif contract_code.startswith('MXF'):
                simple_code = 'MXF'
            elif contract_code.startswith('TXF'):
                simple_code = 'TXF'
            else:
                simple_code = contract_code
            contract_display = f"{simple_code} (日期未知)"
        else:
            contract_display = "未知"
    except:
        contract_display = "未知"
    
    # 訂單類型顯示
    try:
        if str(price_type).upper() == 'MKT':
            price_text = "市價單"
        else:
            price_text = "限價單"
        
        if str(order_type).upper() == 'IOC':
            type_text = "IOC"
        else:
            type_text = "ROD"
        
        order_type_display = f"{price_text}（{type_text}）"
    except:
        order_type_display = f"未知 ({order_type})"
    
    # 成交類型
    # 成交類型 - 移除手動/自動標示，只顯示開倉/平倉
    if str(octype).upper() == 'NEW':
        trade_type = "開倉"
    elif str(octype).upper() == 'COVER':
        trade_type = "平倉"
    else:
        trade_type = f"未知({octype})"  # 不預設為開倉，顯示實際值
    
    # 成交動作 - 使用TXserver.py的get_action_display_by_rule邏輯
    trade_action = get_action_display_by_rule(str(octype).upper(), str(direction).upper())
    
    msg = (f"✅ 成交通知（{current_time}）\n"
           f"選用合約：{contract_display}\n"
           f"訂單類型：{order_type_display}\n"
           f"成交單號：{order_id}\n"
           f"成交類型：{trade_type}\n"
           f"成交動作：{trade_action}\n"
           f"成交部位：{contract_name}\n"
           f"成交數量：{qty} 口\n"
           f"成交價格：{price:.0f}")
    
    return msg

def get_simple_order_log_message(contract_name, direction, qty, price, order_id, octype, is_manual, is_success=False, order_type=None, price_type=None):
    """生成簡化的訂單日誌訊息"""
    try:
        # 簡化合約名稱
        if '微台' in contract_name or 'TMF' in contract_name:
            simple_contract = '微台'
        elif '小台' in contract_name or 'MXF' in contract_name:
            simple_contract = '小台'
        elif '大台' in contract_name or 'TXF' in contract_name:
            simple_contract = '大台'
        else:
            simple_contract = contract_name
        
        # 判斷開平倉
        if str(octype).upper() == 'NEW':
            action_type = '開倉'
        elif str(octype).upper() == 'COVER':
            action_type = '平倉'
        else:
            action_type = '未知'
        
        # 移除手動/自動判斷（不再使用）
        
        # 格式化價格
        if price == 0:
            price_display = '市價'
        else:
            price_display = f'{price:,.0f}'
        
        # 格式化方向 - 修正邏輯
        if str(octype).upper() == 'NEW':
            # 開倉：BUY=多單, SELL=空單
            if str(direction).upper() == 'BUY':
                direction_display = '多單'
            else:
                direction_display = '空單'
        elif str(octype).upper() == 'COVER':
            # 平倉：BUY=平空單, SELL=平多單
            if str(direction).upper() == 'BUY':
                direction_display = '空單'  # 平空單
            else:
                direction_display = '多單'  # 平多單
        else:
            # 備援邏輯
            if str(direction).upper() == 'BUY':
                direction_display = '多單'
            else:
                direction_display = '空單'
        
        # 格式化訂單類型和價格類型
        order_type_display = order_type or 'ROD'
        price_type_display = price_type or 'LMT'
        
        # 組合訂單類型顯示
        if price_type_display.upper() == 'MKT':
            order_info = f"市價 ({order_type_display})"
        else:
            order_info = f"限價 ({order_type_display})"
        
        if is_success:
            # 成交成功格式
            return f"{action_type}成功：{simple_contract}｜{direction_display}｜{qty} 口｜{price_display}｜{order_info}"
        else:
            # 掛單格式（移除手動/自動標示）
            return f"{action_type}：{simple_contract}｜{direction_display}｜{qty} 口｜{price_display}｜{order_info}"
            
    except Exception as e:
        logger.error(f"生成簡化日誌訊息失敗: {e}")
        return f"日誌生成失敗: {order_id}"

def login_sinopac():
    global sinopac_api, sinopac_connected, sinopac_account, sinopac_login_status, sinopac_login_time, auto_logout_timer
    
    try:
        # 使用與main.py相同的方式載入.env設定
        if DOTENV_AVAILABLE and os.path.exists(ENV_PATH):
            load_dotenv(ENV_PATH)
        
        api_key = os.getenv('API_KEY', '')
        secret_key = os.getenv('SECRET_KEY', '')
        person_id = os.getenv('PERSON_ID', '')
        ca_passwd = os.getenv('CA_PASSWD', '')
        ca_path = os.getenv('CA_PATH', '')
        
        # 初始化API
        if sinopac_api:
            try:
                sinopac_api.logout()
            except:
                pass
        
        sinopac_api = sj.Shioaji()
        
        # API登入
        sinopac_api.login(api_key=api_key, secret_key=secret_key)
        
        # 設置回調函數
        try:
            sinopac_api.set_order_callback(order_callback)
            logger.info("回調函數設置成功")
        except Exception as e:
            logger.error(f"回調函數設置失敗: {e}")
            # 回調函數設置失敗，但繼續使用基本功能
        
        # 激活CA憑證 - 智能路徑處理
        cert_file = None
        
        # 優先檢查 server/certificate/ 目錄
        cert_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'certificate')
        if os.path.exists(cert_dir):
            for file in os.listdir(cert_dir):
                if file.endswith('.pfx'):
                    cert_file = os.path.join(cert_dir, file)
                    logger.info(f"找到憑證檔案: {cert_file}")
                    break
        
        # 如果 server/certificate/ 目錄沒有找到，再檢查 ca_path
        if not cert_file and ca_path:
            # 如果是絕對路徑，直接使用
            if os.path.isabs(ca_path):
                final_ca_path = ca_path
            else:
                # 如果是相對路徑，轉換為絕對路徑
                final_ca_path = os.path.abspath(os.path.join(os.path.dirname(__file__), ca_path))
            
            logger.info(f"憑證路徑: {final_ca_path}")
            
            # 檢查路徑是否存在
            if os.path.isfile(final_ca_path):
                cert_file = final_ca_path
            elif os.path.isdir(final_ca_path):
                for file in os.listdir(final_ca_path):
                    if file.endswith('.pfx'):
                        cert_file = os.path.join(final_ca_path, file)
                        logger.info(f"找到憑證檔案: {cert_file}")
                        break
        
        if cert_file and os.path.exists(cert_file):
            try:
                sinopac_api.activate_ca(ca_path=cert_file, ca_passwd=ca_passwd, person_id=person_id)
                logger.info("憑證激活成功")
            except Exception as e:
                error_msg = str(e).lower()
                logger.error(f"憑證激活失敗: {e}")
                
                # 詳細的錯誤分析和建議（僅控制台輸出，不發送TG通知）
                if "password" in error_msg or "passwd" in error_msg or "密碼" in error_msg:
                    logger.error("❌ 憑證密碼錯誤！請檢查前端輸入的憑證密碼是否正確")
                elif "person_id" in error_msg or "身分證字號" in error_msg or "id" in error_msg:
                    logger.error("❌ 身分證字號錯誤！請檢查格式是否為：1個英文字母+9個數字")
                elif "file" in error_msg or "path" in error_msg or "not found" in error_msg:
                    logger.error("❌ 憑證檔案問題！請檢查 .pfx 檔案是否正確上傳")
                elif "expired" in error_msg or "過期" in error_msg:
                    logger.error("❌ 憑證已過期！請聯絡永豐證券更新憑證")
                else:
                    logger.error(f"❌ 憑證激活失敗：{e}")
        else:
            error_msg = f"找不到憑證檔案，請確認 {final_ca_path if 'final_ca_path' in locals() else ca_path} 目錄下有 .pfx 檔案"
            logger.error(f"❌ {error_msg}")
            # 不發送TG通知，只記錄到前端
            try:
                requests.post(
                    f'http://127.0.0.1:{CURRENT_PORT}/api/system_log',
                    # json={'message': error_msg, 'type': 'error'},
                    timeout=5
                )
            except:
                pass
        
        # 設定期貨帳戶
        try:
            accounts = [acc for acc in sinopac_api.list_accounts() if acc.account_type == 'F']
            if accounts:
                sinopac_api.futopt_account = accounts[0]
                sinopac_account = accounts[0].account_id
            else:
                sinopac_account = "無期貨帳戶"
        except Exception:
            sinopac_account = "帳戶設定失敗"
        
        sinopac_connected = True
        sinopac_login_status = True
        sinopac_login_time = datetime.now()  # 記錄登入時間
        
        
        # 啟動12小時自動登出定時器
        start_auto_logout_timer()
        
        # 登入成功後更新期貨合約和保證金資訊
        update_futures_contracts()
        update_margin_requirements_from_api()
        
        # 初始化TXserver風格的合約對象
        init_contracts()
        
        print_console("API", "SUCCESS", "永豐API 登入成功!!!")
        return True
        
    except Exception as e:
        sinopac_connected = False
        sinopac_login_status = False
        sinopac_account = None
        sinopac_login_time = None
        print_console("API", "ERROR", "永豐API 登入失敗!!!", str(e))
        return False

def check_api_health():
    """檢查API健康狀態"""
    global sinopac_api, sinopac_connected
    
    if not sinopac_api or not sinopac_connected:
        return False
    
    try:
        # 嘗試獲取帳戶餘額來測試API連接性
        balance = sinopac_api.account_balance()
        if balance is not None:
            print_console("API", "SUCCESS", f"API健康檢查正常 - 餘額: {balance}")
            return True
        else:
            print_console("API", "WARNING", "API健康檢查失敗 - 無法獲取帳戶餘額")
            return False
    except Exception as e:
        print_console("API", "ERROR", "API健康檢查失敗", str(e))
        return False

def save_order_mapping():
    """保存訂單映射到文件"""
    global order_octype_map
    
    try:
        order_mapping_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'order_mapping.json')
        with open(order_mapping_file, 'w', encoding='utf-8') as f:
            json.dump(order_octype_map, f, ensure_ascii=False, indent=2)
        logger.info(f"訂單映射已保存到 {order_mapping_file}")
    except Exception as e:
        logger.error(f"保存訂單映射失敗: {str(e)}")

def load_order_mapping():
    """從文件加載訂單映射"""
    global order_octype_map
    
    try:
        order_mapping_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'order_mapping.json')
        if os.path.exists(order_mapping_file):
            with open(order_mapping_file, 'r', encoding='utf-8') as f:
                order_octype_map = json.load(f)
            logger.info(f"訂單映射已從 {os.path.basename(order_mapping_file)} 加載，共 {len(order_octype_map)} 個訂單")
        else:
            logger.info("訂單映射文件不存在，使用空映射")
    except Exception as e:
        logger.error(f"加載訂單映射失敗: {str(e)}")
        order_octype_map = {}

def is_duplicate_signal(signal_id, action, contract_code, time_window=30):
    """檢查是否為重複訊號"""
    global duplicate_signal_window
    
    current_time = datetime.now()
    signal_key = f"{signal_id}_{action}_{contract_code}"
    
    # 清理過期的記錄
    expired_keys = []
    for key, timestamp in duplicate_signal_window.items():
        if (current_time - timestamp).total_seconds() > time_window:
            expired_keys.append(key)
    
    for key in expired_keys:
        del duplicate_signal_window[key]
    
    # 檢查是否為重複訊號
    if signal_key in duplicate_signal_window:
        time_diff = (current_time - duplicate_signal_window[signal_key]).total_seconds()
        logger.info(f"檢測到重複訊號: {signal_key}, 時間差: {time_diff:.2f}秒")
        return True
    
    # 記錄新訊號
    duplicate_signal_window[signal_key] = current_time
    return False

def logout_sinopac():
    """登出永豐API"""
    global sinopac_api, sinopac_connected, sinopac_account, sinopac_login_status, sinopac_login_time, auto_logout_timer, order_octype_map
    
    try:
        # 停止自動登出定時器
        stop_auto_logout_timer()
        
        # 登出API
        if sinopac_api and sinopac_connected:
            try:
                sinopac_api.logout()
            except Exception as api_e:
                print_console("API", "WARNING", "API登出時發生錯誤", str(api_e))
        
        # 清理所有狀態
        sinopac_api = None  # 重要：清除API物件
        sinopac_connected = False
        sinopac_login_status = False
        sinopac_account = None
        sinopac_login_time = None
        
        # 清理訂單映射
        with global_lock:
            order_octype_map.clear()
        
        print_console("API", "SUCCESS", "永豐API登出成功!!!")
        return True
        
    except Exception as e:
        print_console("API", "ERROR", "永豐API登出失敗", str(e))
        # 即使出錯也要清理狀態
        sinopac_api = None
        sinopac_connected = False
        sinopac_login_status = False
        sinopac_account = None
        sinopac_login_time = None
        
        # 即使登出失敗也清理映射
        with global_lock:
            order_octype_map.clear()
        
        return False

def start_auto_logout_timer():
    """啟動12小時自動登出定時器"""
    global auto_logout_timer
    
    # 如果已有定時器在運行，先停止它
    stop_auto_logout_timer()
    
    # 計算12小時後的時間
    logout_time = datetime.now() + timedelta(hours=AUTO_LOGOUT_HOURS)
    
    def auto_logout_task():
        """12小時後自動登出並重新登入"""
        global sinopac_connected, sinopac_login_status, is_reconnecting
        
        if sinopac_connected and sinopac_login_status:
            print_console("API", "WARNING", f"目前連線已滿{AUTO_LOGOUT_HOURS}個小時，將自動登出並重新登入!")
            
            # 登出
            logout_sinopac()
            
            # 等待1秒後重新登入
            time.sleep(1)
            
            # 持續重新登入直到成功
            retry_count = 0
            while True:
                retry_count += 1
                print_console("API", "INFO", f"嘗試第{retry_count}次自動重新登入...")
                
                if login_sinopac():
                    print_console("API", "SUCCESS", f"12小時自動重新登入成功！(第{retry_count}次嘗試)")
                    break  # 登入成功，跳出循環
                else:
                    print_console("API", "WARNING", f"第{retry_count}次自動重新登入失敗，30秒後重試...")
                    
                    # 等待30秒後重試
                    time.sleep(30)
    
    # 計算延遲時間（秒）
    delay_seconds = AUTO_LOGOUT_HOURS * 3600
    
    # 啟動定時器
    # 創建並註冊自動登出定時器線程
    auto_logout_timer = create_managed_thread(target=lambda: (time.sleep(delay_seconds), auto_logout_task()), name="自動登出定時器")
    auto_logout_timer.daemon = True
    auto_logout_timer.start()
    
    print_console("SYSTEM", "INFO", f"已啟動{AUTO_LOGOUT_HOURS}小時自動登出定時器，將於 {logout_time.strftime('%Y-%m-%d %H:%M:%S')} 自動登出")

def stop_auto_logout_timer():
    """停止自動登出定時器"""
    global auto_logout_timer
    
    try:
        if auto_logout_timer and auto_logout_timer.is_alive():
            # Thread對象沒有cancel方法，需要用其他方式停止
            # 這裡我們只是將其設為None，讓其自然結束
            auto_logout_timer = None
            print_console("SYSTEM", "INFO", "已標記自動登出定時器停止")
    except Exception as e:
        print_console("SYSTEM", "ERROR", "停止自動登出定時器失敗", str(e))
        auto_logout_timer = None

def get_connection_duration():
    """獲取當前連線時長（小時）"""
    global sinopac_login_time, sinopac_connected
    
    if sinopac_login_time and sinopac_connected:
        duration = datetime.now() - sinopac_login_time
        return duration.total_seconds() / 3600  # 轉換為小時
    elif not sinopac_connected:
        return -1  # 未連線
    else:
        return 0  # 連線但沒有登入時間記錄

def get_sinopac_status():
    """獲取永豐API狀態"""
    global sinopac_connected, sinopac_account, sinopac_login_status
    
    # 處理期貨帳戶顯示
    if sinopac_account and sinopac_account not in ["無期貨帳戶", "帳戶設定失敗"]:
        futures_display = sinopac_account  # 真實帳戶號碼
    elif sinopac_account == "無期貨帳戶":
        futures_display = "無期貨帳戶"
    else:
        futures_display = "未獲取帳戶"  # 包含未獲取和設定失敗的情況
    
    return {
        "connected": sinopac_connected,
        "status": sinopac_login_status,
        "futures_account": futures_display,
        "api_ready": sinopac_connected and sinopac_account is not None and sinopac_account != "無期貨帳戶" and sinopac_account != "帳戶設定失敗"
    }

def reset_login_flag():
    update_login_status(0)
    # 重置LOGIN時停止隧道
    stop_cloudflare_tunnel()
    # 重置時也登出永豐API（如果已經初始化的話）
    if sinopac_api is not None:
        logout_sinopac()

# 程式啟動時重置登入狀態
reset_login_flag()

def ensure_tx_env_exists():
    """確保TX環境配置文件存在，如果不存在則創建預設配置"""
    try:
        if not os.path.exists(TX_ENV_PATH):
            print_console("SYSTEM", "INFO", "TX環境配置文件不存在，正在創建預設配置...")
            
            # 創建預設TX環境配置
            default_tx_config = """# Telegram Bot
BOT_TOKEN=7202376519:AAF-i3MbuMEpz0W7nFE9KmieqVw7L5s0xK4

# Telegram ID
CHAT_ID=

# 永豐 API Key
API_KEY=

# 永豐 Secret Key
SECRET_KEY=

# 身分證字號
PERSON_ID=

# 台股日曆
HOLIDAY_DIR=Desktop/AutoTX/holiday

# 憑證檔案
CA_PATH=Desktop/AutoTX/certificate

# 憑證密碼
CA_PASSWD=

# 憑證起始日
CERT_START=

# 憑證到期日
CERT_END=

# 登入狀態
LOGIN=0
"""
            # 確保config目錄存在
            os.makedirs(os.path.dirname(TX_ENV_PATH), exist_ok=True)
            
            # 寫入預設配置
            with open(TX_ENV_PATH, 'w', encoding='utf-8') as f:
                f.write(default_tx_config)
            
            print_console("SYSTEM", "SUCCESS", f"TX環境配置文件已創建: {TX_ENV_PATH}")
        else:
            print_console("SYSTEM", "INFO", "TX環境配置文件已存在")
    except Exception as e:
        print_console("SYSTEM", "ERROR", "創建TX環境配置失敗", str(e))

def ensure_btc_env_exists():
    """確保BTC環境配置文件存在，如果不存在則創建預設配置"""
    try:
        btc_env_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config', 'btc.env')
        
        if not os.path.exists(btc_env_path):
            print_console("SYSTEM", "INFO", "BTC環境配置文件不存在，正在創建預設配置...")
            
            # 創建預設BTC環境配置
            default_btc_config = """# Telegram Bot
BOT_TOKEN_BTC=7912873826:AAFPPDwuwspKVdyDGwh1oVqxH6u1gQ_N-jU

# Telegram ID
CHAT_ID_BTC=

# 幣安 API Key
BINANCE_API_KEY=

# 幣安 Secret Key
BINANCE_SECRET_KEY=

# 幣安用戶ID
BINANCE_USER_ID=

# 交易對
TRADING_PAIR=BTCUSDT

# 合約類型
CONTRACT_TYPE=PERPETUAL

# 槓桿倍數
LEVERAGE=10

# 風險比例百分比
POSITION_SIZE=20

# 保證金模式
MARGIN_TYPE=CROSS

# 登入狀態
LOGIN_BTC=0
"""
            # 確保config目錄存在
            os.makedirs(os.path.dirname(btc_env_path), exist_ok=True)
            
            # 寫入預設配置
            with open(btc_env_path, 'w', encoding='utf-8') as f:
                f.write(default_btc_config)
            
            print_console("SYSTEM", "SUCCESS", f"BTC環境配置文件已創建: {btc_env_path}")
        else:
            print_console("SYSTEM", "INFO", "BTC環境配置文件已存在")
    except Exception as e:
        print_console("SYSTEM", "ERROR", "創建BTC環境配置失敗", str(e))

# 環境配置文件檢查將在主程式啟動時執行

def cleanup_on_exit():
    """程式退出時的清理工作"""
    global sinopac_api, sinopac_connected, sinopac_account, sinopac_login_status, auto_logout_timer, order_octype_map, connection_monitor_timer, ALL_CHILD_PROCESSES, ALL_ACTIVE_THREADS, SHUTDOWN_FLAG, tunnel_manager, flask_server_thread, flask_server_shutdown
    
    try:
        # ========== 基本清理（優先執行） ==========
        print_console("SYSTEM", "INFO", "開始程式清理工作...")
        
        # 設置全域停止標誌
        signal_shutdown()
        
        # ========== 關閉Flask服務器 ==========
        try:
            print_console("FLASK", "INFO", "正在關閉Flask服務器...")
            flask_server_shutdown = True
            
            # 嘗試通過API端點關閉Flask服務器
            try:
                import requests
                requests.post(f'http://127.0.0.1:{CURRENT_PORT}/shutdown', timeout=2)
                print_console("FLASK", "SUCCESS", "Flask服務器已通過API關閉")
            except:
                # 如果API關閉失敗，嘗試其他方法
                print_console("FLASK", "WARNING", "API關閉失敗，嘗試強制關閉")
                
        except Exception as e:
            print_console("FLASK", "ERROR", f"關閉Flask服務器時發生錯誤: {e}")
        
        # 停止BTC模組
        try:
            if BTC_MODULE_AVAILABLE:
                btcmain.stop_btc_module()
                print_console("SYSTEM", "SUCCESS", "BTC模組已停止")
        except Exception as e:
            print_console("SYSTEM", "WARNING", "停止BTC模組時發生錯誤", str(e))
        
        # 停止自動登出定時器
        stop_auto_logout_timer()
        
        # 停止連線監控器
        stop_connection_monitor()
        
        
        # 關閉隧道服務
        try:
            # 關閉主隧道服務
            if tunnel_service:
                print_console("TUNNEL", "STOP", "正在關閉主隧道服務...")
                tunnel_service.stop_tunnel()
                print_console("TUNNEL", "SUCCESS", "主隧道服務已關閉")
            
            # 關閉隧道管理器中的所有隧道
            if tunnel_manager:
                print_console("TUNNEL", "STOP", "正在關閉所有隧道服務...")
                tunnel_manager.stop_all_tunnels()
                print_console("TUNNEL", "SUCCESS", "所有隧道服務已關閉")
                
        except Exception as e:
            print_console("TUNNEL", "ERROR", "關閉隧道服務時發生錯誤", str(e))
        
        # 額外強制終止所有 cloudflared.exe 進程
        try:
            import platform
            if platform.system() == "Windows":
                import subprocess
                print_console("SYSTEM", "INFO", "檢查並終止殘留的 cloudflared.exe 進程...")
                subprocess.run(["taskkill", "/F", "/IM", "cloudflared.exe"], 
                             capture_output=True, text=True, check=False)
                print_console("SYSTEM", "SUCCESS", "已清理 cloudflared.exe 進程")
            else:
                # Linux/WSL 環境
                import subprocess
                result = subprocess.run(["pkill", "-f", "cloudflared"], 
                                      capture_output=True, text=True, check=False)
                if result.returncode == 0:
                    print_console("SYSTEM", "SUCCESS", "已清理 cloudflared 進程")
        except Exception as e:
            print_console("SYSTEM", "WARNING", "清理 cloudflared 進程時發生錯誤", str(e))
        
        # ========== 清理所有活動線程 ========== #
        if ALL_ACTIVE_THREADS:
            print_console("SYSTEM", "INFO", f"正在等待活動線程結束... (共 {len(ALL_ACTIVE_THREADS)} 個)")
            for i, thread in enumerate(ALL_ACTIVE_THREADS.copy()):
                try:
                    if thread and thread.is_alive():
                        print_console("SYSTEM", "INFO", f"等待線程結束: {thread.name} ({i+1}/{len(ALL_ACTIVE_THREADS)})")
                        thread.join(timeout=3)  # 等待3秒
                        if thread.is_alive():
                            print_console("SYSTEM", "WARNING", f"線程 {thread.name} 仍在運行，將被強制結束")
                        else:
                            print_console("SYSTEM", "SUCCESS", f"線程 {thread.name} 已正常結束")
                    else:
                        print_console("SYSTEM", "INFO", f"線程 {thread.name} 已結束")
                except Exception as e:
                    print_console("SYSTEM", "WARNING", f"處理線程 {getattr(thread, 'name', '未知')} 時發生錯誤", str(e))
            
            ALL_ACTIVE_THREADS.clear()
            print_console("SYSTEM", "SUCCESS", "所有活動線程清理完畢")
        else:
            print_console("SYSTEM", "INFO", "沒有需要清理的活動線程")
        
        # ========== 統一結束所有子進程 ========== #
        if ALL_CHILD_PROCESSES:
            print_console("SYSTEM", "INFO", f"正在結束所有由主程式啟動的子進程... (共 {len(ALL_CHILD_PROCESSES)} 個)")
            for i, p in enumerate(ALL_CHILD_PROCESSES.copy()):  # 使用copy避免迭代時修改列表
                try:
                    if p and hasattr(p, 'pid'):
                        if PSUTIL_AVAILABLE:
                            try:
                                proc = psutil.Process(p.pid)
                                # 先終止所有子進程
                                for child in proc.children(recursive=True):
                                    child.terminate()
                                    try:
                                        child.wait(timeout=2)
                                    except psutil.TimeoutExpired:
                                        child.kill()
                                # 再終止主進程
                                proc.terminate()
                                try:
                                    proc.wait(timeout=3)
                                except psutil.TimeoutExpired:
                                    proc.kill()
                                print_console("SYSTEM", "SUCCESS", f"已結束子進程 {i+1}/{len(ALL_CHILD_PROCESSES)} PID={p.pid}")
                            except psutil.NoSuchProcess:
                                print_console("SYSTEM", "INFO", f"子進程 PID={p.pid} 已不存在")
                            except Exception as e:
                                print_console("SYSTEM", "WARNING", f"使用psutil結束進程失敗 PID={p.pid}", str(e))
                                # 降級到基本方法
                                p.terminate()
                                try:
                                    p.wait(timeout=3)
                                except Exception:
                                    p.kill()
                        else:
                            # 沒有psutil時的基本清理
                            p.terminate()
                            try:
                                p.wait(timeout=3)
                                print_console("SYSTEM", "SUCCESS", f"已結束子進程 {i+1}/{len(ALL_CHILD_PROCESSES)} PID={p.pid}")
                            except Exception:
                                p.kill()
                                print_console("SYSTEM", "WARNING", f"強制結束子進程 PID={p.pid}")
                except Exception as e:
                    print_console("SYSTEM", "WARNING", f"無法結束子進程 {i+1}/{len(ALL_CHILD_PROCESSES)} PID={getattr(p, 'pid', '?')}", str(e))
            
            # 清空子進程列表
            ALL_CHILD_PROCESSES.clear()
            print_console("SYSTEM", "SUCCESS", "所有子進程清理完畢，進程列表已清空")
        else:
            print_console("SYSTEM", "INFO", "沒有需要清理的子進程")
        
        # 永豐API登出（靜默）
        if sinopac_api and sinopac_connected:
            sinopac_api.logout()
            sinopac_connected = False
            sinopac_account = None
            sinopac_login_status = False
        
        # 清理訂單映射
        with global_lock:
            order_octype_map.clear()
            
    except Exception as e:
        pass  # 靜默處理錯誤
    
    # 重置TX LOGIN狀態（靜默重試）
    for attempt in range(3):
        try:
            update_login_status(0)
            # 驗證是否成功重置
            if os.path.exists(ENV_PATH):
                with open(ENV_PATH, 'r', encoding='utf-8') as f:
                    content = f.read()
                    if 'LOGIN=0' in content:
                        break
            time.sleep(0.1)  # 縮短重試間隔
        except Exception as e:
            pass  # 靜默處理錯誤
    
    # 重置BTC LOGIN_BTC狀態（靜默重試）
    if BTC_MODULE_AVAILABLE:
        try:
            btc_env_path = os.path.join(CONFIG_DIR, 'btc.env')
            if os.path.exists(btc_env_path):
                for attempt in range(3):
                    try:
                        # 讀取BTC配置文件
                        with open(btc_env_path, 'r', encoding='utf-8') as f:
                            lines = f.readlines()
                        
                        # 更新LOGIN_BTC狀態
                        for i, line in enumerate(lines):
                            if line.startswith('LOGIN_BTC='):
                                lines[i] = 'LOGIN_BTC=0\n'
                                break
                        
                        # 寫回文件
                        with open(btc_env_path, 'w', encoding='utf-8') as f:
                            f.writelines(lines)
                        
                        # 驗證重置是否成功
                        with open(btc_env_path, 'r', encoding='utf-8') as f:
                            content = f.read()
                            if 'LOGIN_BTC=0' in content:
                                break
                        time.sleep(0.1)
                    except Exception as e:
                        pass  # 靜默處理錯誤
        except Exception as e:
            pass  # 靜默處理錯誤
    
    print_console("SYSTEM", "SUCCESS", "清理工作完成")

def terminate_all_processes():
    """最後手段：強制終止所有相關進程"""
    try:
        print_console("SYSTEM", "WARNING", "執行強制進程清理...")
        
        if PSUTIL_AVAILABLE:
            current_pid = os.getpid()
            current_process = psutil.Process(current_pid)
            
            # 獲取當前進程的所有子進程
            children = current_process.children(recursive=True)
            
            # 終止所有子進程
            for child in children:
                try:
                    child.terminate()
                    print_console("SYSTEM", "INFO", f"終止子進程: PID={child.pid}")
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    pass
            
            # 等待子進程終止
            time.sleep(1)
            
            # 強制殺死仍然運行的子進程
            for child in children:
                try:
                    if child.is_running():
                        child.kill()
                        print_console("SYSTEM", "WARNING", f"強制殺死進程: PID={child.pid}")
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    pass
        
        # 使用系統命令強制清理特定進程
        import platform
        if platform.system() == "Windows":
            try:
                # 清理所有Python進程（除了當前進程）
                import subprocess
                result = subprocess.run(
                    ["tasklist", "/FI", "IMAGENAME eq python.exe", "/FO", "CSV"],
                    capture_output=True, text=True, check=False
                )
                
                if result.returncode == 0:
                    lines = result.stdout.strip().split('\n')
                    current_pid = os.getpid()
                    
                    for line in lines[1:]:  # 跳過標題行
                        if line:
                            parts = line.split(',')
                            if len(parts) >= 2:
                                pid_str = parts[1].strip('"')
                                try:
                                    pid = int(pid_str)
                                    if pid != current_pid:
                                        subprocess.run(["taskkill", "/F", "/PID", str(pid)], 
                                                     capture_output=True, check=False)
                                        print_console("SYSTEM", "INFO", f"清理Python進程: PID={pid}")
                                except ValueError:
                                    pass
                
                # 再次清理cloudflared進程
                subprocess.run(["taskkill", "/F", "/IM", "cloudflared.exe"], 
                             capture_output=True, check=False)
                print_console("SYSTEM", "SUCCESS", "強制進程清理完成")
                
            except Exception as e:
                print_console("SYSTEM", "ERROR", f"Windows進程清理失敗: {e}")
                
        else:
            # Linux/WSL環境
            try:
                import subprocess
                # 清理cloudflared進程
                subprocess.run(["pkill", "-f", "cloudflared"], 
                             capture_output=True, check=False)
                print_console("SYSTEM", "SUCCESS", "Linux進程清理完成")
            except Exception as e:
                print_console("SYSTEM", "ERROR", f"Linux進程清理失敗: {e}")
                
    except Exception as e:
        print_console("SYSTEM", "ERROR", f"強制進程清理失敗: {e}")

def signal_handler(signum, frame):
    """信號處理函數"""
    print_console("SYSTEM", "WARNING", f"收到信號 {signum}，開始清理程式...")
    try:
        cleanup_on_exit()
        time.sleep(0.5)
        print_console("SYSTEM", "SUCCESS", "信號處理清理完成")
    except Exception as e:
        print_console("SYSTEM", "ERROR", f"信號處理過程中發生錯誤: {e}")
    finally:
        print_console("SYSTEM", "INFO", "信號處理完成，正常退出")
        sys.exit(0)

# 註冊程序關閉時的清理函數
atexit.register(cleanup_on_exit)  # 確保程式正常退出時也會執行清理

# 註冊信號處理器 - 移至main.py主線程處理
# signal.signal(signal.SIGINT, signal_handler)
# signal.signal(signal.SIGTERM, signal_handler)




# 移除重複的signal_handler函數，保留第一個

def check_daily_startup_notification():
    """檢查是否需要發送每日啟動通知"""
    try:
        global notification_sent_date
        
        # 檢查是否為交易日且開市
        now = datetime.now()
        today = now.date()
        current_time = now.hour * 100 + now.minute  # HHMM格式
        
        # 只在08:45-08:46之間檢查，確保只在開市時發送
        if 845 <= current_time <= 846:
            # 檢查今天是否為交易日且開市
            response = requests.get(f'http://127.0.0.1:{CURRENT_PORT}/api/trading/status', timeout=5)
            if response.status_code == 200:
                data = response.json()
                if data.get('is_trading_day', False) and data.get('is_market_open', False):
                    # 確保今天還沒發送過通知
                    if notification_sent_date != today:
                        send_daily_startup_notification()
                        notification_sent_date = today
                    else:
                        pass
                else:
                    pass
            else:
                pass
            
    except Exception as e:
        pass

def is_btc_logged_in():
    """檢查BTC系統是否已登入"""
    try:
        btc_env_path = os.path.join(os.path.dirname(__file__), '..', 'config', 'btc.env')
        if os.path.exists(btc_env_path):
            with open(btc_env_path, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if line.startswith('LOGIN_BTC='):
                        login_status = line.split('=', 1)[1]
                        return login_status == '1'
        return False
    except Exception as e:
        logger.error(f"檢查BTC登入狀態失敗: {e}")
        return False

def send_btc_notification_if_logged_in(notification_type):
    """僅在BTC已登入時發送通知"""
    try:
        if not BTC_MODULE_AVAILABLE:
            logger.info("BTC模組不可用，跳過通知")
            return
        
        if not is_btc_logged_in():
            logger.info(f"BTC系統未登入，跳過{notification_type}通知")
            return
        
        logger.info(f"BTC系統已登入，發送{notification_type}通知")
        
        if notification_type == 'startup':
            btcmain.send_btc_daily_startup_notification()
        elif notification_type == 'statistics':
            btcmain.check_btc_daily_trading_statistics()
        else:
            logger.warning(f"未知的通知類型: {notification_type}")
            
    except Exception as e:
        logger.error(f"BTC條件通知發送失敗: {e}")

def schedule_next_check():
    """排程下一次檢查"""
    # 清除所有現有的排程
    schedule.clear()
    
    # 設定明天早上 8:45 的檢查
    tomorrow = datetime.now() + timedelta(days=1)
    schedule.every().day.at("08:45").do(check_daily_startup_notification)
    
    # 設定BTC每日啟動通知 00:00 (24/7無交易日限制，僅在BTC登入時發送)
    if BTC_MODULE_AVAILABLE:
        schedule.every().day.at("00:00").do(lambda: send_btc_notification_if_logged_in('startup'))
        
        # 設定BTC每日交易統計 23:58 (24/7無交易日限制，僅在BTC登入時發送，統計後會自動延遲生成日報和月報)
        schedule.every().day.at("23:58").do(lambda: send_btc_notification_if_logged_in('statistics'))
        
        print_console("SYSTEM", "SUCCESS", "已設定BTC定時任務")
        print_console("SYSTEM", "INFO", "  - 00:00: BTC每日啟動通知 (僅在BTC登入時發送)")
        print_console("SYSTEM", "INFO", "  - 23:58: BTC每日交易統計 (僅在BTC登入時發送，統計後延遲30秒生成日報，月末再延遲30秒生成月報)")
    
    # 設定今天下午 14:50 的夜盤檢查
    schedule.every().day.at("14:50").do(check_night_session_notification)
    
    # 設定今天晚上 23:58 的交易統計檢查
    schedule.every().day.at("23:58").do(check_daily_trading_statistics)
    
    print_console("SYSTEM", "INFO", f"已排程下一次啟動通知檢查：{tomorrow.strftime('%Y-%m-%d')} 08:45")
    print_console("SYSTEM", "INFO", f"已排程下一次夜盤通知檢查：{datetime.now().strftime('%Y-%m-%d')} 14:50")
    print_console("SYSTEM", "INFO", f"已排程下一次交易統計檢查：{datetime.now().strftime('%Y-%m-%d')} 23:58")

def start_notification_checker():
    """啟動通知檢查器"""
    global notification_sent_date
    notification_sent_date = None
    
    def schedule_loop():
        # 初始排程
        schedule_next_check()
        
        while True:
            schedule.run_pending()
            time.sleep(30)  # 每30秒檢查一次排程
    
    notification_thread = create_managed_thread(target=schedule_loop, name="排程通知線程")
    notification_thread.start()

def format_number_for_notification(value):
    """格式化數字用於通知，移除整數的 .0"""
    if value is None or value == '':
        return '0'
    
    try:
        num = float(value)
        # 如果是整數，返回整數格式
        if num == int(num):
            return str(int(num))
        else:
            # 如果是小數，保留小數位
            return str(num)
    except (ValueError, TypeError):
        return str(value)

def send_daily_startup_notification():
    """發送每日啟動通知"""
    try:
        # 讀取 .env 檔案獲取 bot token 和 chat id
        if not os.path.exists(ENV_PATH):
            return
        
        load_dotenv(ENV_PATH)
        bot_token = os.getenv('BOT_TOKEN')
        chat_id = os.getenv('CHAT_ID')
        
        if not bot_token or not chat_id:
            return
        
        # 獲取憑證到期日
        cert_end = os.getenv('CERT_END', '')
        
        # 獲取API狀態
        api_status_response = requests.get(f'http://127.0.0.1:{CURRENT_PORT}/api/sinopac/status', timeout=5)
        api_status_data = api_status_response.json()
        api_status = "已連線" if api_status_data.get('connected', False) else "未連線"
        futures_account = api_status_data.get('futures_account', '-')
        
        # 獲取合約資訊
        contracts_response = requests.get(f'http://127.0.0.1:{CURRENT_PORT}/api/futures/contracts', timeout=5)
        contracts_data = contracts_response.json()
        selected_contracts = contracts_data.get('selected_contracts', {})
        
        # 獲取帳戶狀態
        account_response = requests.get(f'http://127.0.0.1:{CURRENT_PORT}/api/account/status', timeout=5)
        account_data = account_response.json().get('data', {})
        
        # 獲取持倉狀態
        position_response = requests.get(f'http://127.0.0.1:{CURRENT_PORT}/api/position/status', timeout=5)
        position_data = position_response.json()
        
        # 構建訊息
        message = "✅ 自動交易台指期正在啟動中.....\n"
        message += "═════ 系統資訊 ═════\n"
        message += f"憑證效期：{cert_end[:10]}\n"  # 只取年月日
        message += f"綁定帳戶：{futures_account}\n"
        message += f"API 狀態：{api_status}\n"
        
        message += "═════ 選用合約 ═════\n"
        # 按照指定順序顯示合約：大台指、小台指、微台指
        contract_order = [
            ('TXF', '大台指'),
            ('MXF', '小台指'), 
            ('TMF', '微台指')
        ]
        
        for code, contract_name in contract_order:
            # 在轉倉模式下，優先使用次月合約資訊
            if rollover_mode and next_month_contracts.get(code):
                next_contract = next_month_contracts[code]
                code_part = next_contract.code
                delivery_part = format_delivery_date(next_contract.delivery_date)
                # 獲取次月合約的保證金
                margin = margin_requirements.get(contract_name.replace('指', ''), 0)
                message += f"{contract_name} {code_part} ({delivery_part}) ${margin:,}\n"
            else:
                # 使用常規選用合約
                contract_info = selected_contracts.get(code, '-')
                if contract_info != '-':
                    # 解析合約資訊
                    parts = contract_info.split('　')
                    code_part = parts[0]
                    delivery_part = parts[1].replace('交割日：', '')
                    margin_part = parts[2].replace('保證金 $', '').replace(',', '')
                    
                    message += f"{contract_name} {code_part} ({delivery_part}) ${int(margin_part):,}\n"
        
        message += "═════ 帳戶狀態 ═════\n"
        message += f"權益總值：{format_number_for_notification(account_data.get('權益總值', 0))}\n"
        message += f"權益總額：{format_number_for_notification(account_data.get('權益總額', 0))}\n"
        message += f"今日餘額：{format_number_for_notification(account_data.get('今日餘額', 0))}\n"
        message += f"昨日餘額：{format_number_for_notification(account_data.get('昨日餘額', 0))}\n"
        message += f"可用保證金：{format_number_for_notification(account_data.get('可用保證金', 0))}\n"
        message += f"原始保證金：{format_number_for_notification(account_data.get('原始保證金', 0))}\n"
        message += f"維持保證金：{format_number_for_notification(account_data.get('維持保證金', 0))}\n"
        message += f"風險指標：{format_number_for_notification(account_data.get('風險指標', 0))}%\n"
        message += f"手續費：{format_number_for_notification(account_data.get('手續費', 0))}\n"
        message += f"期交稅：{format_number_for_notification(account_data.get('期交稅', 0))}\n"
        message += f"本日平倉損益＄{format_number_for_notification(account_data.get('本日平倉損益', 0))} TWD\n"
        
        message += "═════ 持倉狀態 ═════\n"
        if not position_data.get('has_positions', False):
            message += "❌ 無持倉部位"
        else:
            positions = position_data.get('data', {})
            total_pnl = position_data.get('total_pnl_value', 0)
            
            # 按照指定順序顯示持倉：大台、小台、微台
            position_order = [
                ('TXF', '大台'),
                ('MXF', '小台'),
                ('TMF', '微台')
            ]
            
            for code, contract_name in position_order:
                pos = positions.get(code, {})
                if pos.get('動作', '-') != '-':  # 有持倉的才顯示
                    # 獲取該持倉的未實現損益
                    unrealized_pnl = pos.get('未實現損益', '0')
                    # 移除千分位符號並轉換為數字
                    pnl_value = int(unrealized_pnl.replace(',', '')) if unrealized_pnl != '-' else 0
                    message += f"{contract_name}｜{pos['動作']}｜{pos['數量']}｜{pos['均價']}｜＄{pnl_value:,} TWD\n"
            
            message += f"未實現總損益＄{int(total_pnl):,} TWD"
        
        # 發送 Telegram 訊息
        url = f'https://api.telegram.org/bot{bot_token}/sendMessage'
        payload = {
            'chat_id': chat_id,
            'text': message,
            'parse_mode': 'HTML'
        }
        
        response = requests.post(url, json=payload, timeout=10)
        response.raise_for_status()
        
        
    except Exception as e:
        pass

def check_margin_changes():
    """檢查保證金是否有變更"""
    try:
        global margin_requirements, last_margin_requirements
        
        # 更新保證金資訊
        if update_margin_requirements_from_api():
            # 檢查是否有變更
            has_changes = False
            changes = []
            
            for contract in ['大台', '小台', '微台']:
                current = margin_requirements.get(contract, 0)
                previous = last_margin_requirements.get(contract, 0)
                
                # 如果有之前的記錄，且金額不同，就是有變更
                if previous > 0 and current != previous:
                    has_changes = True
                    changes.append((contract, previous, current))
                
                # 更新記錄
                last_margin_requirements[contract] = current
            
            # 如果有變更，發送通知
            if has_changes:
                send_margin_change_notification(changes)
                
    except Exception as e:
        logger.error(f"檢查保證金變更失敗: {e}")

def generate_trading_report(trades, account_data, position_data, cover_trades, total_orders, total_cancels, total_trades, total_cover_quantity, contract_pnl, custom_filename=None):
    """生成交易報表Excel文件"""
    try:
        # 創建TX交易報表目錄
        report_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'TX交易報表')
        os.makedirs(report_dir, exist_ok=True)
        
        # 創建工作簿和工作表
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # 設置所有欄寬為19
        for col in range(1, 12):
            ws.column_dimensions[get_column_letter(col)].width = 25
            
        # 設置樣式（與BTC報表一致）
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # 背景色（與BTC一致）
        blue_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # 字體（與BTC一致）
        white_font = Font(color="FFFFFF", bold=True)
        black_font = Font(color="000000", bold=True)
        
        # 對齊
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # 邊框（與BTC一致）
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 交易總覽區塊（四大區塊用藍色，A-K欄位）
        ws.merge_cells('A1:K1')
        ws['A1'] = '交易總覽'
        # 應用藍色背景到A1:K1（與BTC一致）
        for col in range(1, 12):  # A到K
            cell = ws[f'{get_column_letter(col)}1']
            cell.fill = blue_fill
            cell.font = white_font
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # 交易總覽標題（橫向）
        titles = ['委託次數', '取消次數', '成交次數', '平倉口數', '大台損益', '小台損益', '微台損益']
        for i, title in enumerate(titles):
            col = get_column_letter(i + 1)
            ws[f'{col}2'] = title
            ws[f'{col}2'].alignment = center_alignment
            ws[f'{col}2'].fill = gray_fill
            ws[f'{col}2'].font = black_font
            ws[f'{col}2'].border = thin_border
        
        # 交易總覽內容（加上文字置中）
        values = [
            f"{total_orders} 筆",
            f"{total_cancels} 筆", 
            f"{total_trades} 筆",
            f"{total_cover_quantity} 口",
            f"＄{format_number_for_notification(contract_pnl['TXF'])} TWD",
            f"＄{format_number_for_notification(contract_pnl['MXF'])} TWD",
            f"＄{format_number_for_notification(contract_pnl['TMF'])} TWD"
        ]
        for i, value in enumerate(values):
            col = get_column_letter(i + 1)
            ws[f'{col}3'] = value
            ws[f'{col}3'].alignment = center_alignment
            ws[f'{col}3'].border = thin_border
        
        # 帳戶狀態區塊（四大區塊用藍色，A-K欄位）
        current_row = 5
        ws.merge_cells(f'A{current_row}:K{current_row}')
        ws[f'A{current_row}'] = '帳戶狀態'
        # 應用藍色背景到A5:K5（與BTC一致）
        for col in range(1, 12):  # A到K
            cell = ws[f'{get_column_letter(col)}{current_row}']
            cell.fill = blue_fill
            cell.font = white_font
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # 帳戶狀態標題（橫向）
        account_titles = ['權益總值', '權益總額', '今日餘額', '昨日餘額', '可用保證金', '原始保證金', 
                         '維持保證金', '風險指標', '手續費', '期交稅', '本日平倉損益']
        for i, title in enumerate(account_titles):
            col = get_column_letter(i + 1)
            ws[f'{col}{current_row + 1}'] = title
            ws[f'{col}{current_row + 1}'].alignment = center_alignment
            ws[f'{col}{current_row + 1}'].fill = gray_fill
            ws[f'{col}{current_row + 1}'].font = black_font
            ws[f'{col}{current_row + 1}'].border = thin_border
            ws[f'{col}{current_row + 1}'].font = black_font
            ws[f'{col}{current_row + 1}'].border = thin_border
        
        # 帳戶狀態內容（加上文字置中）
        for i, title in enumerate(account_titles):
            col = get_column_letter(i + 1)
            value = account_data.get(title, 0)
            if title == '風險指標':
                ws[f'{col}{current_row + 2}'] = f"{format_number_for_notification(value)}%"
            elif title == '本日平倉損益':
                ws[f'{col}{current_row + 2}'] = f"＄{format_number_for_notification(value)} TWD"
            else:
                ws[f'{col}{current_row + 2}'] = format_number_for_notification(value)
            ws[f'{col}{current_row + 2}'].alignment = center_alignment
            ws[f'{col}{current_row + 2}'].border = thin_border
        
        # 交易明細區塊（四大區塊用藍色，A-K欄位）
        current_row += 4
        ws.merge_cells(f'A{current_row}:K{current_row}')
        ws[f'A{current_row}'] = '交易明細'
        # 應用藍色背景到A:K（與BTC一致）
        for col in range(1, 12):  # A到K
            cell = ws[f'{get_column_letter(col)}{current_row}']
            cell.fill = blue_fill
            cell.font = white_font
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # 交易明細標題
        detail_titles = ['平倉時間', '交易單號', '選用合約', '訂單類型', '成交類型', '成交部位', 
                        '成交動作', '成交數量', '開倉價格', '平倉價格', '已實現損益']
        for i, title in enumerate(detail_titles):
            col = get_column_letter(i + 1)
            ws[f'{col}{current_row + 1}'] = title
            ws[f'{col}{current_row + 1}'].alignment = center_alignment
            ws[f'{col}{current_row + 1}'].fill = gray_fill
            ws[f'{col}{current_row + 1}'].font = black_font
            ws[f'{col}{current_row + 1}'].border = thin_border
        
        # 交易明細內容（加上文字置中）
        if cover_trades:
            for i, trade in enumerate(cover_trades):
                row = current_row + i + 2
                
                # 平倉時間（格式化時間，移除毫秒）
                timestamp = trade.get('timestamp', '')
                if timestamp:
                    try:
                        # 移除毫秒部分，只保留到秒
                        if '.' in timestamp:
                            timestamp = timestamp.split('.')[0]
                        # 確保格式為 YYYY-MM-DD HH:MM:SS
                        if 'T' in timestamp:
                            timestamp = timestamp.replace('T', ' ')
                        ws[f'A{row}'] = timestamp
                    except:
                        ws[f'A{row}'] = timestamp
                else:
                    ws[f'A{row}'] = ''
                ws[f'A{row}'].alignment = center_alignment
                
                # 平倉單號 - 優先使用成交單號，回退到訂單ID
                deal_number = trade.get('deal_number', '')
                order_id = trade.get('order_id', '')
                # 優先顯示成交單號，如果沒有則顯示訂單ID
                display_id = deal_number if deal_number and deal_number != order_id else order_id
                ws[f'B{row}'] = display_id
                ws[f'B{row}'].alignment = center_alignment
                
                # 選用合約顯示：優先使用完整合約代碼，回退到基礎代碼
                contract_code = trade.get('contract_code', '')
                if not contract_code:
                    # 回退到基礎合約代碼
                    contract_code = 'TXF' if trade['contract_name'] == '大台' else 'MXF' if trade['contract_name'] == '小台' else 'TMF'
                
                # 優先使用增強數據中的交割日期
                delivery_date = trade.get('delivery_date', '')
                
                # 如果沒有交割日，嘗試從全域合約對象獲取
                if not delivery_date:
                    try:
                        if contract_code == 'TXF' and 'contract_txf' in globals() and contract_txf:
                            if hasattr(contract_txf, 'delivery_date'):
                                delivery_date = contract_txf.delivery_date.strftime('%Y/%m/%d') if hasattr(contract_txf.delivery_date, 'strftime') else str(contract_txf.delivery_date)
                        elif contract_code == 'MXF' and 'contract_mxf' in globals() and contract_mxf:
                            if hasattr(contract_mxf, 'delivery_date'):
                                delivery_date = contract_mxf.delivery_date.strftime('%Y/%m/%d') if hasattr(contract_mxf.delivery_date, 'strftime') else str(contract_mxf.delivery_date)
                        elif contract_code == 'TMF' and 'contract_tmf' in globals() and contract_tmf:
                            if hasattr(contract_tmf, 'delivery_date'):
                                delivery_date = contract_tmf.delivery_date.strftime('%Y/%m/%d') if hasattr(contract_tmf.delivery_date, 'strftime') else str(contract_tmf.delivery_date)
                    except:
                        pass
                
                # 如果還是沒有交割日，嘗試從API獲取目前合約的交割日
                if not delivery_date and sinopac_connected and sinopac_api:
                    try:
                        contracts = sinopac_api.Contracts.Futures
                        available_contracts = []
                        for contract in contracts:
                            if contract_code in contract.code:
                                available_contracts.append(contract)
                        
                        if available_contracts:
                            # 按交割日期排序，取得最近的合約（即選用合約）
                            available_contracts.sort(key=lambda x: x.delivery_date)
                            selected_contract = available_contracts[0]
                            delivery_date = format_delivery_date(selected_contract.delivery_date)
                    except:
                        pass
                
                # 格式化合約顯示
                if delivery_date:
                    formatted_date = format_delivery_date(delivery_date)
                    # 如果是完整合約代碼（如TMFR1），顯示完整代碼+交割日
                    # 如果是基礎代碼（如TMF），顯示基礎代碼+交割日
                    display_contract = contract_code
                    ws[f'C{row}'] = f"{display_contract}（{formatted_date}）"
                else:
                    # 沒有交割日期時，顯示合約代碼
                    ws[f'C{row}'] = contract_code
                ws[f'C{row}'].alignment = center_alignment
                
                # 訂單類型顯示 - 直接使用已格式化的數據
                formatted_order_type = trade.get('order_type_display', '')
                if formatted_order_type:
                    # 直接使用已格式化的訂單類型（避免重複括號）
                    ws[f'D{row}'] = formatted_order_type
                else:
                    # 回退邏輯：自行格式化
                    price_type = trade.get('price_type', '')
                    order_type = trade.get('order_type', '')
                    
                    if price_type == 'MKT':
                        ws[f'D{row}'] = f"市價單（{order_type}）" if order_type else '市價單'
                    elif price_type == 'LMT':
                        ws[f'D{row}'] = f"限價單（{order_type}）" if order_type else '限價單'
                    elif order_type:
                        ws[f'D{row}'] = order_type
                    else:
                        ws[f'D{row}'] = '-'
                ws[f'D{row}'].alignment = center_alignment
                
                # 成交類型 - 移除手動/自動標示，只顯示平倉
                ws[f'E{row}'] = '平倉'
                ws[f'E{row}'].alignment = center_alignment
                
                # 成交部位
                ws[f'F{row}'] = trade['contract_name']
                ws[f'F{row}'].alignment = center_alignment
                
                # 成交動作顯示簡化動作（只顯示多單/空單）
                ws[f'G{row}'] = trade['action']  # 直接使用從trade記錄中來的完整動作
                ws[f'G{row}'].alignment = center_alignment
                
                # 成交數量（修復重複單位問題）
                quantity_str = str(trade['quantity'])
                # 移除所有現有的"口"字，然後重新添加
                quantity_clean = quantity_str.replace(' 口', '').strip()
                ws[f'H{row}'] = f"{quantity_clean} 口"
                ws[f'H{row}'].alignment = center_alignment
                
                # 開倉價格 - 使用真實數據，"未知"表示確實無法獲取開倉價格
                open_price = trade.get('open_price', '')
                ws[f'I{row}'] = open_price
                ws[f'I{row}'].alignment = center_alignment
                
                # 平倉價格 - 使用真實數據
                cover_price = trade.get('cover_price', '')
                ws[f'J{row}'] = cover_price
                ws[f'J{row}'].alignment = center_alignment
                
                # 已實現損益（總是顯示）
                pnl_value = trade.get('pnl', 0)
                if trade.get('open_price') == "未知" or pnl_value == 0:
                    ws[f'K{row}'] = f"＄{pnl_value:,} TWD" if pnl_value != 0 else "＄0 TWD"
                else:
                    ws[f'K{row}'] = f"＄{pnl_value:,} TWD"
                ws[f'K{row}'].alignment = center_alignment
        
        # 持倉狀態區塊（四大區塊用藍色，A-K欄位）
        current_row = current_row + (len(cover_trades) if cover_trades else 0) + 3
        ws.merge_cells(f'A{current_row}:K{current_row}')
        ws[f'A{current_row}'] = '持倉狀態'
        # 應用藍色背景和白色字體到A:K
        for col in range(1, 12):  # A到K
            cell = ws[f'{get_column_letter(col)}{current_row}']
            cell.fill = blue_fill
            cell.font = white_font
            cell.alignment = center_alignment
        
        # 持倉狀態標題
        position_titles = ['開倉時間', '交易單號', '選用合約', '訂單類型', '成交類型', '成交部位', 
                         '成交動作', '成交數量', '開倉價格', '平倉價格', '未實現損益']
        for i, title in enumerate(position_titles):
            col = get_column_letter(i + 1)
            ws[f'{col}{current_row + 1}'] = title
            ws[f'{col}{current_row + 1}'].alignment = center_alignment
            ws[f'{col}{current_row + 1}'].fill = gray_fill
            ws[f'{col}{current_row + 1}'].font = black_font
            ws[f'{col}{current_row + 1}'].border = thin_border
        
        # 持倉狀態內容（加上文字置中）
        if isinstance(position_data, dict) and position_data.get('has_positions', False):
            positions = position_data.get('data', {})
            row_offset = 2
            for code, pos in positions.items():
                if pos.get('動作', '-') != '-':
                    # 平倉時間（格式化開倉時間）
                    opening_time = pos.get('開倉時間', '')
                    if opening_time:
                        try:
                            # 轉換ISO時間格式為可讀格式
                            dt = datetime.fromisoformat(opening_time.replace('Z', '+00:00'))
                            formatted_time = dt.strftime('%Y/%m/%d %H:%M:%S')
                            ws[f'A{current_row + row_offset}'] = formatted_time
                        except:
                            ws[f'A{current_row + row_offset}'] = opening_time
                    else:
                        ws[f'A{current_row + row_offset}'] = ''
                    ws[f'A{current_row + row_offset}'].alignment = center_alignment
                    
                    # 成交單號（真實成交單號）
                    ws[f'B{current_row + row_offset}'] = pos.get('成交單號', pos.get('委託單號', ''))
                    ws[f'B{current_row + row_offset}'].alignment = center_alignment
                    
                    # 選用合約顯示：使用歷史記錄數據源（與交易明細保持一致）
                    contract_code = code  # 直接使用code作為合約代號
                    
                    # 使用與交易明細相同的獲取邏輯
                    delivery_date = get_delivery_date_for_contract(contract_code)
                    
                    # 統一的合約顯示邏輯
                    contract_display = get_contract_display_with_delivery(contract_code, delivery_date)
                    ws[f'C{current_row + row_offset}'] = contract_display
                    ws[f'C{current_row + row_offset}'].alignment = center_alignment
                    
                    # 訂單類型（顯示開倉的訂單類型）
                    price_type = pos.get('委託價格類型', '')
                    order_type = pos.get('訂單類型', pos.get('委託條件', ''))
                    
                    if price_type == 'MKT':
                        order_type_str = f'市價單（{order_type}）' if order_type else '市價單'
                    elif price_type == 'LMT':
                        order_type_str = f'限價單（{order_type}）' if order_type else '限價單'
                    else:
                        order_type_str = order_type or price_type or '未知'
                    
                    ws[f'D{current_row + row_offset}'] = order_type_str
                    ws[f'D{current_row + row_offset}'].alignment = center_alignment
                    
                    # 成交類型 - 移除手動/自動標示，只顯示開倉
                    ws[f'E{current_row + row_offset}'] = '開倉'
                    ws[f'E{current_row + row_offset}'].alignment = center_alignment
                    
                    # 成交部位（大台/小台/微台）
                    contract_position = pos.get('商品名稱', '')
                    if not contract_position:
                        if code == 'TXF':
                            contract_position = '大台'
                        elif code == 'MXF':
                            contract_position = '小台'
                        elif code == 'TMF':
                            contract_position = '微台'
                    ws[f'F{current_row + row_offset}'] = contract_position
                    ws[f'F{current_row + row_offset}'].alignment = center_alignment
                    
                    # 成交動作顯示簡化動作（移除買入/賣出標示）
                    action = pos.get('動作', '')
                    if '多單' in action:
                        action_text = '多單'
                    elif '空單' in action:
                        action_text = '空單'
                    else:
                        action_text = action  # 保持原值作為備援
                    ws[f'G{current_row + row_offset}'] = action_text
                    ws[f'G{current_row + row_offset}'].alignment = center_alignment
                    
                    # 成交數量（修復重複單位問題）
                    quantity_str = str(pos.get('數量', ''))
                    quantity_clean = quantity_str.replace(' 口', '').strip()
                    ws[f'H{current_row + row_offset}'] = f"{quantity_clean} 口"
                    ws[f'H{current_row + row_offset}'].alignment = center_alignment
                    
                    # 開倉價格（優先使用真實開倉價格）
                    opening_price = pos.get('開倉價格', pos.get('均價', ''))
                    ws[f'I{current_row + row_offset}'] = opening_price
                    ws[f'I{current_row + row_offset}'].alignment = center_alignment
                    
                    # 平倉價格（持倉狀態顯示0）
                    ws[f'J{current_row + row_offset}'] = '0'
                    ws[f'J{current_row + row_offset}'].alignment = center_alignment
                    
                    # 未實現損益
                    unrealized_pnl = pos.get('未實現損益', '0')
                    pnl_value = int(unrealized_pnl.replace(',', '')) if unrealized_pnl != '-' else 0
                    ws[f'K{current_row + row_offset}'] = f"＄{pnl_value:,} TWD"
                    ws[f'K{current_row + row_offset}'].alignment = center_alignment
                    
                    row_offset += 1
        
        # 保存文件
        if custom_filename:
            filename = custom_filename
        else:
            today = datetime.now().strftime('%Y-%m-%d')
            filename = f"TX_{today}.xlsx"
        filepath = os.path.join(report_dir, filename)
        wb.save(filepath)
        
        # 日誌記錄已移除
        
        # 發送 Telegram 通知並附上檔案（合併發送）
        date_str = datetime.now().strftime('%Y年%m月%d日')
        caption = f"{date_str} | 交易結算報表已生成！！！"
        send_telegram_file(filepath, caption)
            
        return filepath
        
    except Exception as e:
        logger.error(f"生成交易報表失敗: {e}")
        import traceback
        traceback.print_exc()
        
        # 錯誤日誌記錄已移除
        
        return {
            'success': False,
            'error': str(e),
            'message': '交易報表生成失敗'
        }

def send_margin_change_notification(changes):
    """發送保證金變更通知"""
    try:
        # 讀取 .env 檔案獲取 bot token 和 chat id
        if not os.path.exists(ENV_PATH):
            return
        
        load_dotenv(ENV_PATH)
        bot_token = os.getenv('BOT_TOKEN')
        chat_id = os.getenv('CHAT_ID')
        
        if not bot_token or not chat_id:
            return
        
        # 構建訊息
        message = "⚠️ 保證金已更新！！！\n"
        
        # 顯示所有合約的最新保證金
        for contract in ['大台', '小台', '微台']:
            margin = margin_requirements.get(contract, 0)
            contract_name = contract + "指" if contract != "微台" else contract + "指"
            message += f"{contract_name}　${margin:,}\n"
        
        # 發送 Telegram 訊息
        url = f'https://api.telegram.org/bot{bot_token}/sendMessage'
        payload = {
            'chat_id': chat_id,
            'text': message,
            'parse_mode': 'HTML'
        }
        
        response = requests.post(url, json=payload, timeout=10)
        response.raise_for_status()
        
        # 保證金日誌記錄已移除
        
    except Exception as e:
        logger.error(f"發送保證金變更通知失敗: {e}")

def check_night_session_notification():
    """檢查是否需要發送夜盤通知"""
    try:
        now = datetime.now()
        current_time = now.hour * 100 + now.minute  # HHMM格式
        
        # 只在14:49-14:51之間檢查
        if 1449 <= current_time <= 1451:
            # 檢查今天是否為交易日且開市
            response = requests.get(f'http://127.0.0.1:{CURRENT_PORT}/api/trading/status', timeout=5)
            if response.status_code == 200:
                data = response.json()
                if data.get('is_trading_day', False) and data.get('is_market_open', False):
                    # 檢查保證金變更
                    check_margin_changes()

    except Exception as e:
        logger.error(f"檢查夜盤通知失敗: {e}")

def check_daily_trading_statistics():
    """檢查是否需要發送每日交易統計"""
    try:
        today = datetime.now()
        current_weekday = today.weekday()  # 0=週一, 5=週六, 6=週日
        
        # 週六特殊處理：檢查週五是否為交易日，決定是否發送夜盤統計
        if current_weekday == 5:  # 週六
            friday = today - timedelta(days=1)  # 週六減一天=週五
            
            # 判斷週五是否為交易日
            friday_is_trading_day = False
            try:
                response = requests.get(f'http://127.0.0.1:{CURRENT_PORT}/api/trading/status', 
                                      params={'date': friday.strftime('%Y-%m-%d')}, timeout=5)
                if response.status_code == 200:
                    friday_is_trading_day = response.json().get('is_trading_day', False)
            except:
                # 如果API失敗，直接判斷週五（weekday=4）
                friday_is_trading_day = (friday.weekday() != 6)  # 週日=6為非交易日
            
            # 只有週五是交易日時，週六才會有夜盤，才發送統計
            if friday_is_trading_day:
                logger.info(f"週五({friday.strftime('%Y-%m-%d')})是交易日，週六有夜盤，發送交易統計")
                send_daily_trading_statistics()
                logger.info(f"已發送週六交易統計：{today.strftime('%Y-%m-%d')}")
            else:
                logger.info(f"週五({friday.strftime('%Y-%m-%d')})非交易日，週六無夜盤，跳過統計")
        else:
            # 非週六：檢查今天是否為交易日
            response = requests.get(f'http://127.0.0.1:{CURRENT_PORT}/api/trading/status', timeout=5)
            if response.status_code == 200:
                data = response.json()
                if data.get('is_trading_day', False):
                    send_daily_trading_statistics()
                else:
                    logger.info(f"非交易日，跳過交易統計：{today.strftime('%Y-%m-%d')}")
            else:
                logger.info(f"無法獲取交易狀態，跳過交易統計：{today.strftime('%Y-%m-%d')}")
            
    except Exception as e:
        logger.error(f"檢查每日交易統計失敗: {e}")


def is_last_trading_day_of_month():
    """檢查今天是否為當月最後一個交易日（簡化版本）"""
    try:
        today = datetime.now()
        tomorrow = today + timedelta(days=1)
        
        # 如果明天是下個月，且今天是交易日，則今天是本月最後一個交易日
        if today.month != tomorrow.month:
            response = requests.get(f'http://127.0.0.1:{CURRENT_PORT}/api/trading/status', timeout=5)
            if response.status_code == 200:
                return response.json().get('is_trading_day', False)
        
        return False
        
    except Exception as e:
        logger.error(f"檢查月末交易日失敗: {e}")
        return False

def diagnose_tx_opening_records():
    """診斷TX開倉記錄搜尋功能（用於測試隔夜持倉修復）"""
    try:
        logger.info("🔍 開始診斷TX開倉記錄搜尋功能...")
        
        opening_trades = {}  
        today = datetime.now()
        
        # 遍歷最近7天的交易記錄
        for i in range(7):
            check_date = today - timedelta(days=i)
            date_str = check_date.strftime('%Y%m%d')
            json_file = os.path.join(TX_DATA_DIR, f'TXtransdata_{date_str}.json')
            
            if os.path.exists(json_file):
                try:
                    with open(json_file, 'r', encoding='utf-8') as f:
                        trades_data = json.load(f)
                        logger.info(f"📄 診斷讀取: {date_str} ({len(trades_data)}筆)")
                        
                    # 查找開倉交易
                    opening_count = 0
                    for trade in trades_data:
                        if trade.get('type') == 'deal':
                            raw_data = trade.get('raw_data', {})
                            order = raw_data.get('order', {})
                            contract = raw_data.get('contract', {})
                            
                            if order.get('oc_type') == 'New':  # 開倉交易
                                contract_code = contract.get('code', '')
                                contract_type = None
                                
                                if 'TXF' in contract_code:
                                    contract_type = 'TXF'
                                elif 'MXF' in contract_code:
                                    contract_type = 'MXF'
                                elif 'TMF' in contract_code:
                                    contract_type = 'TMF'
                                
                                if contract_type:
                                    opening_count += 1
                                    trade_timestamp = trade.get('timestamp', '')
                                    order_id = order.get('id', '')
                                    order_type = order.get('order_type', '')
                                    
                                    logger.info(f"🎯 找到開倉: {contract_type} - {trade_timestamp[:19]} - ID:{order_id} - 類型:{order_type}")
                    
                    if opening_count > 0:
                        logger.info(f"   該日共找到 {opening_count} 筆開倉記錄")
                                    
                except Exception as e:
                    logger.error(f"診斷讀取 {date_str} 失敗: {e}")
        
        logger.info("✅ TX開倉記錄診斷完成")
        return True
        
    except Exception as e:
        logger.error(f"TX開倉記錄診斷失敗: {e}")
        return False

def generate_monthly_trading_report():
    """生成當月交易報表"""
    try:
        # 獲取當月日期範圍
        today = datetime.now()
        year = today.year
        month = today.month
        
        # 收集當月所有交易記錄
        all_trades = []
        
        # 獲取當月第一天和最後一天
        first_day = datetime(year, month, 1)
        if month == 12:
            last_day = datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            last_day = datetime(year, month + 1, 1) - timedelta(days=1)
        
        # 遍歷當月每一天收集交易記錄
        current_date = first_day
        while current_date <= last_day:
            date_str = current_date.strftime('%Y%m%d')
            trades_file = os.path.join(TX_DATA_DIR, f'TXtransdata_{date_str}.json')
            
            if os.path.exists(trades_file):
                try:
                    with open(trades_file, 'r', encoding='utf-8') as f:
                        daily_trades = json.load(f)
                        if daily_trades:
                            all_trades.extend(daily_trades)
                            logger.info(f"讀取 {date_str} 交易記錄：{len(daily_trades)} 筆")
                except Exception as e:
                    logger.error(f"讀取 {date_str} 交易記錄失敗: {e}")
            
            current_date += timedelta(days=1)
        
        if not all_trades:
            logger.info("當月無交易記錄，不生成月報")
            return {
                'success': False,
                'error': '當月無交易記錄',
                'message': '當月無交易記錄，不生成月報'
            }
        
        # 統計月度累計數據（第一區塊、第三區塊）
        total_orders = 0
        total_cancels = 0
        total_trades = 0
        
        for trade in all_trades:
            trade_type = trade.get('type', '')
            if trade_type == 'order':
                total_orders += 1
            elif trade_type == 'deal':
                total_trades += 1
            elif trade_type == 'cancel' or trade_type == 'fail':
                total_cancels += 1
        
        # 分析平倉交易明細（整月所有平倉）
        logger.info(f"正在分析 {len(all_trades)} 筆月度交易記錄...")
        cover_trades, total_cover_quantity, contract_pnl = analyze_simple_trading_stats(all_trades)
        logger.info(f"月度統計完成：平倉 {total_cover_quantity} 口，{len(cover_trades)} 筆交易")
        
        # 獲取當日帳戶狀態和持倉狀態（第二區塊、第四區塊）
        try:
            account_response = requests.get(f'http://127.0.0.1:{CURRENT_PORT}/api/account/status', timeout=5)
            position_response = requests.get(f'http://127.0.0.1:{CURRENT_PORT}/api/position/status', timeout=5)
            
            account_data = account_response.json().get('data', {}) if account_response.status_code == 200 else {}
            position_data = position_response.json().get('data', {}) if position_response.status_code == 200 else {}
            
        except Exception as e:
            logger.error(f"獲取帳戶和持倉狀態失敗: {e}")
            account_data = {}
            position_data = {}
        
        # 使用與日報相同的函數生成報表，但檔名標註為月報
        date_str = today.strftime('%Y-%m')
        filename = f'TX_{date_str}_月報.xlsx'
        
        # 直接調用日報生成函數，使用相同的數據結構和格式
        report_path = generate_trading_report(
            all_trades, 
            account_data, 
            position_data, 
            cover_trades, 
            total_orders, 
            total_cancels, 
            total_trades, 
            total_cover_quantity, 
            contract_pnl,
            custom_filename=filename
        )
        
        if report_path:
            logger.info(f"TX月報生成成功: {report_path}")
            
            # 📤 添加Telegram月報發送功能（優化訊息內容）
            month_str = today.strftime('%Y年%m月')
            caption = f"{month_str} | 月度交易結算報表已生成！！！"
            send_telegram_file(report_path, caption)
            
            return {
                'success': True,
                'file_path': report_path,
                'filename': filename
            }
        else:
            logger.error("TX月報生成失敗")
            return {
                'success': False,
                'error': 'TX月報生成失敗'
            }
        
    except Exception as e:
        logger.error(f"生成TX月報失敗: {e}")
        return {
            'success': False,
            'error': str(e)
        }


def send_daily_trading_statistics():
    """發送每日交易統計"""
    try:
        # 讀取 .env 檔案獲取 bot token 和 chat id
        if not os.path.exists(ENV_PATH):
            return
        
        load_dotenv(ENV_PATH)
        bot_token = os.getenv('BOT_TOKEN')
        chat_id = os.getenv('CHAT_ID')
        
        if not bot_token or not chat_id:
            return
        
        # 獲取最近7天的交易記錄（包含今天），以便正確配對跨日交易的開倉價格
        trades = []
        today = datetime.now()
        
        for i in range(7):  # 讀取過去7天的記錄
            check_date = today - timedelta(days=i)
            date_str = check_date.strftime('%Y%m%d')
            trades_file = os.path.join(TX_DATA_DIR, f'TXtransdata_{date_str}.json')
            
            if os.path.exists(trades_file):
                try:
                    with open(trades_file, 'r', encoding='utf-8') as f:
                        daily_trades = json.load(f)
                        trades.extend(daily_trades)
                        logger.info(f"讀取 {date_str} 交易記錄：{len(daily_trades)} 筆")
                except Exception as e:
                    logger.error(f"讀取 {date_str} 交易記錄失敗: {e}")
        
        # 統計變數
        total_orders = 0  # 委託單量
        total_trades = 0  # 成交單量
        total_cancels = 0  # 取消單量
        total_cover_quantity = 0  # 平倉口數
        cover_trades = []  # 平倉交易明細
        
        # 各合約類型的損益統計
        contract_pnl = {
            'TXF': 0,  # 大台損益
            'MXF': 0,  # 小台損益
            'TMF': 0   # 微台損益
        }
        
        # 用於追蹤已統計的訂單ID，避免重複計算
        processed_orders = set()
        
        if trades:  # 如果有讀取到交易記錄
            # 🔧 企業級統計系統 - 增強版當天交易統計
            today_str = today.strftime('%Y%m%d')
            today_trades_file = os.path.join(TX_DATA_DIR, f'TXtransdata_{today_str}.json')
            
            logger.info(f"📊 開始統計當天交易數據: {today_trades_file}")
            
            if os.path.exists(today_trades_file):
                try:
                    with open(today_trades_file, 'r', encoding='utf-8') as f:
                        today_trades = json.load(f)
                    
                    # 🔧 增強數據驗證和統計
                    if not isinstance(today_trades, list):
                        logger.error(f"❌ 交易記錄格式錯誤：期待list，得到{type(today_trades)}")
                        today_trades = []
                    
                    logger.info(f"📄 讀取到 {len(today_trades)} 筆當天交易記錄")
                    
                    # 詳細分類統計
                    type_counts = {'order': 0, 'deal': 0, 'cancel': 0, 'fail': 0, 'unknown': 0}
                    
                    for i, trade in enumerate(today_trades):
                        if not isinstance(trade, dict):
                            logger.warning(f"⚠️ 記錄 #{i} 格式異常：{type(trade)}")
                            continue
                            
                        trade_type = trade.get('type', 'unknown')
                        
                        # 統計委託次數（提交成功的訂單）
                        if trade_type == 'order':
                            total_orders += 1
                            type_counts['order'] += 1
                        
                        # 統計成交次數（成交通知）
                        elif trade_type == 'deal':
                            total_trades += 1
                            type_counts['deal'] += 1
                        
                        # 統計取消次數（主動取消）
                        elif trade_type == 'cancel':
                            total_cancels += 1
                            type_counts['cancel'] += 1
                        
                        # 統計失敗次數（提交失敗）
                        elif trade_type == 'fail':
                            total_cancels += 1
                            type_counts['fail'] += 1
                        
                        else:
                            type_counts['unknown'] += 1
                            logger.warning(f"⚠️ 未知交易類型: {trade_type} in record #{i}")
                    
                    # 📊 詳細統計日誌
                    logger.info(f"📈 詳細統計結果:")
                    logger.info(f"   ├─ 委託成功(order): {type_counts['order']} 筆")
                    logger.info(f"   ├─ 成交通知(deal): {type_counts['deal']} 筆") 
                    logger.info(f"   ├─ 主動取消(cancel): {type_counts['cancel']} 筆")
                    logger.info(f"   ├─ 提交失敗(fail): {type_counts['fail']} 筆")
                    if type_counts['unknown'] > 0:
                        logger.info(f"   └─ 未知類型(unknown): {type_counts['unknown']} 筆")
                            
                except (json.JSONDecodeError, FileNotFoundError, PermissionError) as e:
                    logger.error(f"❌ 讀取當天交易記錄失敗: {today_trades_file}, 錯誤: {str(e)}")
                    send_telegram_message(f"❌ TX統計數據讀取失敗: {str(e)[:50]}")
                except Exception as e:
                    logger.error(f"❌ 統計處理異常: {str(e)}")
                    
            else:
                logger.warning(f"⚠️ 當天交易記錄文件不存在: {today_trades_file}")
                # 無交易記錄時直接略過，不發送TG通知
            
            # 📊 最終統計摘要
            logger.info(f"📋 當天統計摘要：委託{total_orders}筆，成交{total_trades}筆，取消{total_cancels}筆")
                
            # 使用基本統計分析平倉口數（不重新計算損益，使用永豐API提供的數據）
            # 只統計今天的平倉交易明細，但使用7天數據做開倉價格配對
            today_str = today.strftime('%Y%m%d')
            logger.info(f"正在統計 {len(trades)} 筆交易記錄（篩選當天 {today_str} 的平倉明細）...")
            cover_trades, total_cover_quantity, contract_pnl = analyze_simple_trading_stats(trades, filter_date=today_str)
            logger.info(f"統計完成：平倉 {total_cover_quantity} 口，{len(cover_trades)} 筆交易")
        else:
            logger.info("沒有找到交易記錄")
            cover_trades = []
            total_cover_quantity = 0
        
        # 獲取帳戶狀態
        account_response = requests.get(f'http://127.0.0.1:{CURRENT_PORT}/api/account/status', timeout=5)
        account_data = account_response.json().get('data', {}) if account_response.status_code == 200 else {}
        
        # 獲取持倉狀態
        position_response = requests.get(f'http://127.0.0.1:{CURRENT_PORT}/api/position/status', timeout=5)
        position_data = position_response.json() if position_response.status_code == 200 else {}
        
        # 構建訊息
        today_str = datetime.now().strftime('%Y/%m/%d')
        message = f"📊 交易統計（{today_str}）\n"
        message += "═════ 交易總覽 ═════\n"
        message += f"委託次數：{total_orders} 筆\n"
        message += f"取消次數：{total_cancels} 筆\n"
        message += f"成交次數：{total_trades} 筆\n"
        message += f"平倉口數：{total_cover_quantity} 口\n"
        message += f"大台損益＄{format_number_for_notification(contract_pnl['TXF'])} TWD\n"
        message += f"小台損益＄{format_number_for_notification(contract_pnl['MXF'])} TWD\n"
        message += f"微台損益＄{format_number_for_notification(contract_pnl['TMF'])} TWD\n"
        message += "═════ 帳戶狀態 ═════\n"
        message += f"權益總值：{format_number_for_notification(account_data.get('權益總值', 0))}\n"
        message += f"權益總額：{format_number_for_notification(account_data.get('權益總額', 0))}\n"
        message += f"今日餘額：{format_number_for_notification(account_data.get('今日餘額', 0))}\n"
        message += f"昨日餘額：{format_number_for_notification(account_data.get('昨日餘額', 0))}\n"
        message += f"可用保證金：{format_number_for_notification(account_data.get('可用保證金', 0))}\n"
        message += f"原始保證金：{format_number_for_notification(account_data.get('原始保證金', 0))}\n"
        message += f"維持保證金：{format_number_for_notification(account_data.get('維持保證金', 0))}\n"
        message += f"風險指標：{format_number_for_notification(account_data.get('風險指標', 0))}%\n"
        message += f"手續費：{format_number_for_notification(account_data.get('手續費', 0))}\n"
        message += f"期交稅：{format_number_for_notification(account_data.get('期交稅', 0))}\n"
        message += f"本日平倉損益＄{format_number_for_notification(account_data.get('本日平倉損益', 0))} TWD\n"
        
        message += "═════ 交易明細 ═════\n"
        if not cover_trades:
            message += "❌ 無平倉交易\n"
        else:
            # 按照指定順序排序：大台、小台、微台
            def get_contract_order(contract_name):
                order_map = {'大台': 0, '小台': 1, '微台': 2}
                return order_map.get(contract_name, 3)
            
            # 排序交易明細
            cover_trades.sort(key=lambda x: get_contract_order(x['contract_name']))
            
            for trade in cover_trades:
                # 原格式：微台｜多單｜1口｜22,902｜22,902
                action_text = '多單' if '多' in trade['action'] or 'Buy' in trade['action'] else '空單'
                # 格式化價格顯示千分位
                open_price = f"{trade['open_price']:,}" if isinstance(trade['open_price'], (int, float)) else trade['open_price']
                cover_price = f"{trade['cover_price']:,}" if isinstance(trade['cover_price'], (int, float)) else trade['cover_price']
                
                message += f"{trade['contract_name']}｜{action_text}｜{trade['quantity']}｜{open_price}｜{cover_price}\n"
                # 始終顯示損益，包括0和未知的情況
                if trade.get('open_price') == "未知":
                    message += "損益未知（無對應開倉記錄）\n"
                else:
                    message += f"損益＄{trade['pnl']:,} TWD\n"
        
        message += "═════ 持倉狀態 ═════\n"
        if not position_data.get('has_positions', False):
            message += "❌ 無持倉部位"
        else:
            positions = position_data.get('data', {})
            total_pnl = position_data.get('total_pnl_value', 0)
            
            # 按照指定順序顯示持倉：大台、小台、微台
            position_order = [
                ('TXF', '大台'),
                ('MXF', '小台'),
                ('TMF', '微台')
            ]
            
            for code, contract_name in position_order:
                pos = positions.get(code, {})
                if pos.get('動作', '-') != '-':  # 有持倉的才顯示
                    # 獲取該持倉的未實現損益
                    unrealized_pnl = pos.get('未實現損益', '0')
                    # 移除千分位符號並轉換為數字
                    pnl_value = int(unrealized_pnl.replace(',', '')) if unrealized_pnl != '-' else 0
                    message += f"{contract_name}｜{pos['動作']}｜{pos['數量']}｜{pos['均價']}｜＄{pnl_value:,} TWD\n"
            
            message += f"未實現總損益＄{int(total_pnl):,} TWD"
        
        # 發送 Telegram 訊息
        url = f'https://api.telegram.org/bot{bot_token}/sendMessage'
        payload = {
            'chat_id': chat_id,
            'text': message,
            'parse_mode': 'HTML'
        }
        
        response = requests.post(url, json=payload, timeout=10)
        response.raise_for_status()
        
        logger.info(f"已發送交易統計：{today_str}")
        
        # try:
            # requests.post(
                # f'http://127.0.0.1:{CURRENT_PORT}/api/system_log',
                # json={'message': 'Telegram［交易統計］訊息發送成功！！！', 'type': 'success'},
                # timeout=5
            # )
        # except:
            # pass
        
        # 延遲生成報表 - 統一時間控制
        def delayed_generate_reports():
            # 等待30秒後生成日報 (23:58:30)
            time.sleep(30)
            daily_report_result = generate_trading_report(
                trades=trades,
                account_data=account_data,
                position_data=position_data,
                cover_trades=cover_trades,
                total_orders=total_orders,
                total_cancels=total_cancels,
                total_trades=total_trades,
                total_cover_quantity=total_cover_quantity,
                contract_pnl=contract_pnl
            )
            
            # 🔧 修復：如果是月末且日報生成成功，再等待30秒後生成月報 (23:59:00)
            # 簡化月末檢測邏輯，避免API調用失敗導致月報無法生成
            today = datetime.now()
            tomorrow = today + timedelta(days=1)
            is_month_end = today.month != tomorrow.month
            
            if daily_report_result and is_month_end:
                time.sleep(30)
                monthly_result = generate_monthly_trading_report()
                if monthly_result and monthly_result.get('success'):
                    logger.info(f"📅 TX月報自動生成並發送成功: {monthly_result.get('filename')}")
                else:
                    logger.error("❌ TX月報自動生成失敗")
        
        # 在新線程中執行延遲生成報表
        create_managed_thread(target=delayed_generate_reports, name="延遲報表生成線程").start()
        
    except Exception as e:
        logger.error(f"發送每日交易統計失敗: {e}")


@app.route('/api/manual/order', methods=['POST'])
def manual_order():
    """手動下單API"""
    logger.info(f"🔥 收到TX手動下單API請求")
    try:
        # 驗證API是否已連線
        if not sinopac_connected or not sinopac_api:
            return jsonify({
                'status': 'error',
                'message': '永豐API未連線'
            }), 400
        
        # 獲取下單數據
        data = request.get_json()
        if not data:
            return jsonify({
                'status': 'error',
                'message': '無效的請求數據'
            }), 400
        
        # 解析下單參數
        contract_code = data.get('contract_code', 'TXF')  # 合約代碼
        quantity = int(data.get('quantity', 1))  # 數量
        direction = data.get('direction', '')  # 開多、開空、平多、平空（兼容性）
        price = float(data.get('price', 0))  # 價格
        price_type = data.get('price_type', None)  # MKT或LMT
        order_type = data.get('order_type', None)  # IOC或ROD
        position_type = data.get('position_type', None)  # None、long、short
        action_param = data.get('action', None)  # Buy或Sell（永豐官方參數）
        octype_param = data.get('octype', None)  # New或Cover（永豐官方參數）
        
        
        # 驗證必要欄位
        if not action_param or not octype_param:
            return jsonify({
                'status': 'error',
                'message': '永豐手動下單需要提供 action (Buy/Sell) 和 octype (New/Cover) 參數'
            }), 400
        
        # 檢查是否為交易時間
        trading_status = requests.get(f'http://127.0.0.1:{CURRENT_PORT}/api/trading/status', timeout=5).json()
        if not trading_status.get('is_trading_day', False) or not trading_status.get('is_market_open', False):
            return jsonify({
                'status': 'error',
                'message': '非交易時間'
            }), 400
        
        # 執行手動下單
        try:
            order_result = place_futures_order(
                contract_code=contract_code,
                quantity=quantity,
                direction=direction,
                price=price,
                is_manual=True,  # 手動下單
                position_type=position_type,
                price_type=price_type,  # 傳遞實際的price_type
                order_type=order_type,  # 傳遞實際的order_type
                action_param=action_param,  # 傳遞永豐官方action參數
                octype_param=octype_param   # 傳遞永豐官方octype參數
            )
            
            # 記錄手動下單成功日誌
            add_custom_request_log('POST', '/api/manual/order', 200, {
                'reason': 'TX手動下單成功',
                'contract_code': contract_code,
                'quantity': quantity,
                'action': action_param,
                'octype': octype_param,
                'price': price,
                'price_type': price_type,
                'system': 'TX',
                'is_manual': True
            })
            
            return jsonify({
                'status': 'success',
                'message': '手動下單成功',
                'order': order_result
            })
            
        except Exception as e:
            error_msg = str(e)
            
            # 記錄手動下單失敗日誌
            add_custom_request_log('POST', '/api/manual/order', 500, {
                'reason': f'TX手動下單失敗: {error_msg}',
                'contract_code': contract_code,
                'quantity': quantity,
                'action': action_param,
                'octype': octype_param,
                'system': 'TX',
                'is_manual': True,
                'error': error_msg
            })
            
            # 保存提交失敗記錄
            save_trade({
                'type': 'fail',
                'trade_category': 'manual',
                'raw_data': {
                    'operation': {
                        'op_type': 'OrderFail',
                        'op_code': '99',
                        'op_msg': error_msg
                    },
                    'order': {
                        'action': action_param,
                        'quantity': quantity,
                        'price': price,
                        'oc_type': octype_param,
                        'order_type': order_type or 'IOC',
                        'price_type': price_type or 'MKT'
                    },
                    'contract': {
                        'code': contract_code
                    }
                },
                'contract_name': '大台' if contract_code.startswith('TXF') else '小台' if contract_code.startswith('MXF') else '微台',
                'contract_code': contract_code,
                'error_reason': error_msg,
                'is_manual': True
            })
            
            return jsonify({
                'status': 'error',
                'message': f'手動下單失敗: {error_msg}'
            }), 500
            
    except Exception as e:
        # 記錄請求處理失敗日誌
        add_custom_request_log('POST', '/api/manual/order', 500, {
            'reason': f'TX手動下單請求處理失敗: {str(e)}',
            'system': 'TX',
            'is_manual': True,
            'error': str(e)
        })
        
        return jsonify({
            'status': 'error',
            'message': f'處理請求失敗: {str(e)}'
        }), 500

def is_ip_allowed(ip):
    """檢查IP是否在白名單中"""
    allowed_ips = {"127.0.0.1", "::1"}  # 本地IP白名單，可根據需要擴充
    return ip in allowed_ips


def send_unified_failure_message(data, reason, order_id="未知"):
    """發送統一的提交失敗訊息"""
    global contract_txf, contract_mxf, contract_tmf
    
    try:
        # 保存提交失敗記錄
        save_trade({
            'type': 'fail',
            'trade_category': 'auto',
            'raw_data': {
                'operation': {
                    'op_type': 'SignalFail',
                    'op_code': '99',
                    'op_msg': reason
                },
                'order': {
                    'id': order_id
                },
                'signal_data': data
            },
            'error_reason': reason,
            'is_manual': False,
            'signal_data': data
        })
        # 解析訊號數據
        qty_txf = int(float(data.get('txf', 0)))
        qty_mxf = int(float(data.get('mxf', 0)))
        qty_tmf = int(float(data.get('tmf', 0)))
        price = float(data.get('price', 0))
        direction = data.get('direction', '未知')
        msg_type = data.get('type', 'entry')
        
        # 確定交易動作
        if direction == "開多":
            expected_action = safe_constants.get_action('BUY')
        elif direction == "開空":
            expected_action = safe_constants.get_action('SELL')
        else:
            expected_action = direction
        
        # 確定開平倉類型 - webhook邏輯
        if direction in ['開多', '開空']:
            octype = 'New'  # 開倉
        elif direction in ['平多', '平空']:
            octype = 'Cover'  # 平倉
        else:
            # 根據msg_type判斷
            octype = 'New' if msg_type == 'entry' else 'Cover'
        
        # 獲取合約資訊
        contracts = [
            (contract_txf, qty_txf, "大台", "TXF"),
            (contract_mxf, qty_mxf, "小台", "MXF"), 
            (contract_tmf, qty_tmf, "微台", "TMF")
        ]
        
        # 對每個有數量的合約發送失敗訊息
        for contract, qty, name, code in contracts:
            if qty > 0:
                # 如果合約存在，獲取交割日期
                if contract:
                    delivery_date = contract.delivery_date
                    if hasattr(delivery_date, 'strftime'):
                        delivery_date_str = delivery_date.strftime('%Y/%m/%d')
                    else:
                        delivery_date_str = str(delivery_date)
                    contract_code = contract.code
                else:
                    # 如果合約不存在，使用預設值
                    delivery_date_str = "未知"
                    contract_code = f"{code}XX" 
                
                # 先記錄前端交易日誌
                log_message = get_simple_order_log_message(
                    contract_name=name,
                    direction=str(expected_action),
                    qty=qty,
                    price=price,
                    order_id=order_id,
                    octype=octype,
                    is_manual=False,
                    is_success=False,
                    order_type="IOC",
                    price_type="MKT"
                )
                
                add_custom_request_log('POST', '/webhook', 500, {
                    'reason': log_message,
                    'system': 'TX'
                })
                
                # 然後發送 Telegram 訊息
                fail_message = get_formatted_order_message(
                    is_success=False,
                    order_id=order_id,
                    contract_name=name,
                    qty=qty,
                    price=price,
                    octype=octype,
                    direction=str(expected_action),
                    order_type="IOC",
                    price_type="MKT",
                    is_manual=False,
                    reason=reason,
                    contract_code=contract_code,
                    delivery_date=delivery_date_str
                )
                send_telegram_message(fail_message)
                
    except Exception as e:
        logger.error(f"發送統一失敗訊息錯誤: {e}")
        # 如果統一格式失敗，回退到簡單訊息
        send_telegram_message(f"❌ 提交失敗：{reason}")

def process_signal(data):
    """處理TradingView訊號（參考TXserver.py邏輯）"""
    global has_processed_delivery_exit, active_trades, contract_txf, contract_mxf, contract_tmf, rollover_mode
    
    logger.info(f"[process_signal] 開始處理訊號數據: {data}")
    
    try:
        # 🔧 檢查API連接狀態（記錄但不阻止）
        if not sinopac_connected or not sinopac_api:
            logger.warning(f"⚠️ 永豐API可能未完全連線")
            logger.warning(f"   sinopac_connected: {sinopac_connected}")
            logger.warning(f"   sinopac_api: {sinopac_api is not None}")
            logger.info("🔄 嘗試繼續執行下單...")
        else:
            logger.info("✅ 永豐API連線正常")
            
        # 解析訊號數據
        signal_id = data.get('tradeId')
        msg_type = data.get('type')  # entry 或 exit
        direction = data.get('direction', '未知')
        is_rollover_mode = data.get('rollover_mode', False)
        
        # 獲取價格和時間
        time_ms = int(data.get('time', 0)) / 1000 if data.get('time') else 0
        if time_ms > 0:
            time_str = (datetime.utcfromtimestamp(time_ms) + timedelta(hours=8)).strftime('%Y/%m/%d %H:%M')
        else:
            time_str = datetime.now().strftime('%Y/%m/%d %H:%M')
            
        qty_txf = int(float(data.get('txf', 0)))
        qty_mxf = int(float(data.get('mxf', 0)))
        qty_tmf = int(float(data.get('tmf', 0)))
        price = float(data.get('price', 0))
        
        # 強制使用市價單IOC（參考TXserver.py）
        order_type = "IOC"
        price_type = "MKT"
        
        logger.info(f"解析結果: type={msg_type}, direction={direction}, price={price}")
        logger.info(f"合約數量: TXF={qty_txf}, MXF={qty_mxf}, TMF={qty_tmf}")
        logger.info(f"轉倉模式: {is_rollover_mode}")
        
        # 🔧 價格檢查（使用市價單時價格可以為0）
        if price <= 0:
            logger.info(f"💰 訊號價格為 {price}，使用市價單模式")
            price = 0  # 市價單使用0作為價格
            
        # 🔧 移除交易時間限制 - 允許24小時交易
        logger.info("⏰ 已移除交易時間限制，允許24小時接收訊號")
            
        # 取得持倉資訊
        positions = sinopac_api.list_positions(sinopac_api.futopt_account)
        
        # 初始化合約對象（如果尚未設置）
        # 更新TXF合約（無論是否已存在，都重新選擇以確保轉倉邏輯正確）
        txf_contracts = sinopac_api.Contracts.Futures.get("TXF")
        if txf_contracts:
            sorted_contracts = sorted(txf_contracts, key=lambda x: x.delivery_date)
            if rollover_mode and next_month_contracts.get('TXF'):
                contract_txf = next_month_contracts['TXF']
            else:
                # 非轉倉模式：尋找R1合約作為當月合約
                r1_contract = None
                for contract in sorted_contracts:
                    if contract.code.endswith('R1'):
                        r1_contract = contract
                        break
                contract_txf = r1_contract if r1_contract else sorted_contracts[0]
                
        # 更新MXF合約（無論是否已存在，都重新選擇以確保轉倉邏輯正確）
        mxf_contracts = sinopac_api.Contracts.Futures.get("MXF")
        if mxf_contracts:
            sorted_contracts = sorted(mxf_contracts, key=lambda x: x.delivery_date)
            if rollover_mode and next_month_contracts.get('MXF'):
                contract_mxf = next_month_contracts['MXF']
            else:
                # 非轉倉模式：尋找R1合約作為當月合約
                r1_contract = None
                for contract in sorted_contracts:
                    if contract.code.endswith('R1'):
                        r1_contract = contract
                        break
                contract_mxf = r1_contract if r1_contract else sorted_contracts[0]
                
        # 更新TMF合約（無論是否已存在，都重新選擇以確保轉倉邏輯正確）
        tmf_contracts = sinopac_api.Contracts.Futures.get("TMF")
        if tmf_contracts:
            sorted_contracts = sorted(tmf_contracts, key=lambda x: x.delivery_date)
            if rollover_mode and next_month_contracts.get('TMF'):
                contract_tmf = next_month_contracts['TMF']
            else:
                # 非轉倉模式：尋找R1合約作為當月合約
                r1_contract = None
                for contract in sorted_contracts:
                    if contract.code.endswith('R1'):
                        r1_contract = contract
                        break
                contract_tmf = r1_contract if r1_contract else sorted_contracts[0]
        
        logger.info(f"當前合約: TXF={contract_txf.code if contract_txf else None}, "
                    f"MXF={contract_mxf.code if contract_mxf else None}, "
                    f"TMF={contract_tmf.code if contract_tmf else None}")
        
        # 處理進場訊號
        if msg_type == "entry":
            process_entry_signal(data, qty_txf, qty_mxf, qty_tmf, direction, price, order_type, price_type, positions)
            
        # 處理出場訊號  
        elif msg_type == "exit":
            process_exit_signal(data, qty_txf, qty_mxf, qty_tmf, direction, price, order_type, price_type, positions)
            
        else:
            # 僅支援標準格式：entry/exit
            logger.error(f"❌ 不支援的訊號類型: {msg_type}，僅支援 'entry' 或 'exit'")
            send_unified_failure_message(data, f"不支援的訊號類型: {msg_type}，僅支援標準格式")
            return
            
    except Exception as e:
        error_msg = str(e)
        logger.error(f"[process_signal] 處理訊號失敗：{error_msg}")
        import traceback
        traceback.print_exc()
        send_unified_failure_message(data, error_msg[:100])

def process_entry_signal(data, qty_txf, qty_mxf, qty_tmf, direction, price, order_type, price_type, positions):
    """處理進場訊號"""
    global active_trades, contract_txf, contract_mxf, contract_tmf, rollover_mode
    
    logger.info(f"[process_entry_signal] 處理進場訊號: direction={direction}")
    
    # 標準格式檢查：僅支援「開多」、「開空」
    if direction not in ["開多", "開空"]:
        logger.error(f"❌ 不支援的進場方向: {direction}，僅支援 '開多' 或 '開空'")
        send_unified_failure_message(data, f"不支援的進場方向: {direction}，僅支援標準格式")
        return
        
    # 確定交易動作
    if direction == "開多":
        expected_action = safe_constants.get_action('BUY')
    else:  # 開空
        expected_action = safe_constants.get_action('SELL')
        
    # 🔧 檢查是否有相反持倉 - 改為自動對翻
    opposite_positions = [p for p in positions if p.direction != expected_action and p.quantity != 0]
    
    if opposite_positions:
        logger.info(f"🔄 檢測到反向持倉，執行自動對翻")
        logger.info(f"   目標動作: {expected_action}")
        
        # 先平倉反向持倉
        for position in opposite_positions:
            logger.info(f"   平倉反向持倉: {position.contract.code} {position.direction} {position.quantity}口")
            
            # 確定平倉方向
            if position.direction == safe_constants.get_action('BUY'):
                close_direction = "平多"
            else:
                close_direction = "平空"
            
            # 執行平倉
            try:
                place_futures_order_tx_style(
                    contract=position.contract,
                    quantity=abs(position.quantity),
                    direction=close_direction,
                    price=price,
                    order_type=order_type,
                    price_type=price_type,
                    is_manual=False,
                    position=position
                )
                logger.info(f"✅ 反向持倉平倉成功: {position.contract.code}")
            except Exception as e:
                logger.error(f"❌ 反向持倉平倉失敗: {e}")
        
        # 短暫延遲確保平倉完成
        time.sleep(0.5)
        logger.info("🔄 平倉完成，繼續執行開倉")
        
        # 繼續執行開倉邏輯
    
    # 轉倉邏輯：根據轉倉模式選擇合約
    is_rollover_mode = data.get('rollover_mode', False)
    
    # 執行開倉下單
    contracts = [
        (get_contract_for_rollover('TXF'), qty_txf, "大台", "TXF"),
        (get_contract_for_rollover('MXF'), qty_mxf, "小台", "MXF"), 
        (get_contract_for_rollover('TMF'), qty_tmf, "微台", "TMF")
    ]
    
    for contract, qty, name, code in contracts:
        if qty > 0 and contract:
            try:
                # 判斷使用的合約類型
                if contract.code.endswith('R2'):
                    contract_type = "R2合約"
                elif contract.code.endswith('R1'):
                    contract_type = "R1合約"
                else:
                    contract_type = f"合約{contract.code}"
                    
                logger.info(f"[process_entry_signal] 開始處理{name}進場: {qty} 口 {direction}，開倉合約: {contract.code} ({contract_type})")
                
                result = place_futures_order_tx_style(
                    contract=contract,
                    quantity=qty,
                    direction=direction,
                    price=price,
                    order_type=order_type,
                    price_type=price_type,
                    is_manual=False
                )
                
                if result.get('success'):
                    active_trades[contract_key_map[name]] = data.get('tradeId')
                    logger.info(f"{name}進場成功，開倉合約: {contract.code} ({contract_type})")
                else:
                    logger.error(f"{name}進場失敗: {result.get('message', '未知錯誤')}")
                    
            except Exception as e:
                error_msg = str(e)
                logger.error(f"{name}進場異常: {error_msg}")
                
                # 創建單個合約的data用於發送失敗訊息
                single_data = data.copy()
                single_data['txf'] = qty if code == 'TXF' else 0
                single_data['mxf'] = qty if code == 'MXF' else 0
                single_data['tmf'] = qty if code == 'TMF' else 0
                send_unified_failure_message(single_data, error_msg)

def process_exit_signal(data, qty_txf, qty_mxf, qty_tmf, direction, price, order_type, price_type, positions):
    """處理出場訊號"""
    global active_trades, rollover_mode
    
    logger.info(f"[process_exit_signal] 處理出場訊號: direction={direction}")
    
    # 平倉邏輯：使用實際持倉的合約進行平倉
    is_rollover_mode = data.get('rollover_mode', False)
    
    # 檢查是否有對應方向的持倉 - 修正邏輯
    # 平多：只平多單(Buy方向)，平空：只平空單(Sell方向)
    if direction == "平多":
        target_direction = safe_constants.get_action('BUY')
    elif direction == "平空":
        target_direction = safe_constants.get_action('SELL')
    else:
        logger.error(f"❌ 不支援的平倉方向: {direction}，僅支援 '平多' 或 '平空'")
        send_unified_failure_message(data, f"不支援的平倉方向: {direction}，僅支援標準格式")
        return
    
    # 修改持倉查詢邏輯：只要有持倉就查詢，不受webhook數量限制
    position_txf = next((p for p in positions if p.code.startswith("TXF") and p.quantity != 0 and p.direction == target_direction), None)
    position_mxf = next((p for p in positions if p.code.startswith("MXF") and p.quantity != 0 and p.direction == target_direction), None)
    position_tmf = next((p for p in positions if p.code.startswith("TMF") and p.quantity != 0 and p.direction == target_direction), None)
    
    # 但在實際平倉時，仍然按照webhook指定的數量執行
    if qty_txf == 0: position_txf = None
    if qty_mxf == 0: position_mxf = None  
    if qty_tmf == 0: position_tmf = None
    
    has_position = bool(position_txf or position_mxf or position_tmf)
    
    # 調試信息
    logger.info(f"[平倉檢查] 訊號方向: {direction}, 目標持倉方向: {target_direction}")
    logger.info(f"[平倉檢查] 找到持倉: TXF={position_txf is not None}, MXF={position_mxf is not None}, TMF={position_tmf is not None}")
    if position_txf:
        logger.info(f"[平倉檢查] TXF持倉: {position_txf.code}, 方向: {position_txf.direction}, 數量: {position_txf.quantity}")
    if position_mxf:
        logger.info(f"[平倉檢查] MXF持倉: {position_mxf.code}, 方向: {position_mxf.direction}, 數量: {position_mxf.quantity}")
    if position_tmf:
        logger.info(f"[平倉檢查] TMF持倉: {position_tmf.code}, 方向: {position_tmf.direction}, 數量: {position_tmf.quantity}")
            
    if not has_position:
        logger.warning(f"警告: 無對應方向的持倉，取消平倉。訊號: {direction}")
        logger.info(f"當前所有持倉: {[(p.code, p.direction, p.quantity) for p in positions if p.quantity != 0]}")
        
        # 發送提交失敗訊息 - 使用轉倉邏輯選擇合約（用於通知顯示）
        contracts_to_check = [
            (get_contract_for_rollover('TXF'), qty_txf, "大台", "TXF"),
            (get_contract_for_rollover('MXF'), qty_mxf, "小台", "MXF"), 
            (get_contract_for_rollover('TMF'), qty_tmf, "微台", "TMF")
        ]
        
        # 對每個有數量的合約發送失敗訊息
        for contract, qty, name, code in contracts_to_check:
            if qty > 0 and contract:
                # 獲取交割日期
                delivery_date = contract.delivery_date
                if hasattr(delivery_date, 'strftime'):
                    delivery_date_str = delivery_date.strftime('%Y/%m/%d')
                else:
                    delivery_date_str = str(delivery_date)
                
                # 先記錄前端交易日誌
                log_message = get_simple_order_log_message(
                    contract_name=name,
                    direction=direction,
                    qty=qty,
                    price=price,
                    order_id="未知",
                    octype='Cover',
                    is_manual=False,
                    is_success=False,
                    order_type="IOC",
                    price_type="MKT"
                )
                add_custom_request_log('POST', '/webhook', 500, {
                    'reason': log_message,
                    'system': 'TX'
                })
                
                # 然後發送 Telegram 訊息
                fail_message = get_formatted_order_message(
                    is_success=False,
                    order_id="未知",
                    contract_name=name,
                    qty=qty,
                    price=price,
                    octype='Cover',  # 出場訊號都是平倉
                    direction=direction,
                    order_type="IOC",
                    price_type="MKT",
                    is_manual=False,  # webhook來源為自動
                    reason="無對應持倉",
                    contract_code=contract.code,
                    delivery_date=delivery_date_str
                )
                send_telegram_message(fail_message)
        return
        
    # 執行平倉 - 使用實際持倉的合約
    def get_contract_from_position(position, contract_type):
        """根據持倉獲取實際的合約對象"""
        if not position:
            return None
            
        # 從持倉中獲取合約代碼，然後找到對應的合約對象
        position_code = position.code
        logger.info(f"持倉合約代碼: {position_code}")
        
        try:
            # 獲取該類型的所有合約
            contracts = sinopac_api.Contracts.Futures.get(contract_type)
            if contracts:
                # 找到與持倉代碼匹配的合約對象
                for contract in contracts:
                    if contract.code == position_code:
                        logger.info(f"找到匹配的合約對象: {contract.code}")
                        return contract
                        
            logger.warning(f"警告: 未找到持倉 {position_code} 對應的合約對象")
            return None
        except Exception as e:
            logger.error(f"獲取持倉合約對象失敗: {e}")
            return None
    
    contracts_positions = [
        (get_contract_from_position(position_txf, 'TXF'), qty_txf, "大台", "TXF", position_txf),
        (get_contract_from_position(position_mxf, 'MXF'), qty_mxf, "小台", "MXF", position_mxf),
        (get_contract_from_position(position_tmf, 'TMF'), qty_tmf, "微台", "TMF", position_tmf)
    ]
    
    for contract, qty, name, code, position in contracts_positions:
        if qty > 0 and contract and position:
            try:
                # 判斷使用的合約類型
                if contract.code.endswith('R2'):
                    contract_type = "R2合約"
                elif contract.code.endswith('R1'):
                    contract_type = "R1合約"
                else:
                    contract_type = f"合約{contract.code}"
                    
                logger.info(f"[process_exit_signal] 開始處理{name}出場: {qty} 口，平倉合約: {contract.code} ({contract_type})")
                
                result = place_futures_order_tx_style(
                    contract=contract,
                    quantity=qty,
                    direction=direction,
                    price=price,
                    order_type=order_type,
                    price_type=price_type,
                    is_manual=False,
                    position=position
                )
                
                if result.get('success'):
                    active_trades[contract_key_map[name]] = None
                    logger.info(f"{name}出場成功，平倉合約: {contract.code} ({contract_type})")
                else:
                    logger.error(f"{name}出場失敗: {result.get('message', '未知錯誤')}")
            
            except Exception as e:
                error_msg = str(e)
                logger.error(f"{name}出場異常: {error_msg}")
                
                # 創建單個合約的data用於發送失敗訊息
                single_data = data.copy()
                single_data['txf'] = qty if code == 'TXF' else 0
                single_data['mxf'] = qty if code == 'MXF' else 0
                single_data['tmf'] = qty if code == 'TMF' else 0
                send_unified_failure_message(single_data, error_msg)

def place_futures_order_tx_style(contract, quantity, direction, price, order_type="IOC", price_type="MKT", is_manual=False, position=None):
    """TXserver風格的下單函數"""
    try:
        # 判斷開平倉類型 - 根據direction判斷
        if direction in ["開多", "開空"]:
            is_entry = True
            octype = 'New'
        elif direction in ["平多", "平空"]:
            is_entry = False
            octype = 'Cover'
        else:
            # 如果direction不是中文，則根據position判斷
            is_entry = position is None
            octype = 'New' if is_entry else 'Cover'
        
        # 確定交易動作
        if is_entry:
            # 開倉
            if direction == "開多":
                action = safe_constants.get_action('BUY')
            elif direction == "開空":
                action = safe_constants.get_action('SELL')
            else:
                raise Exception(f"無效的開倉方向: {direction}")
        else:
            # 平倉：根據持倉方向決定平倉動作
            if position.direction == safe_constants.get_action('BUY'):
                # 持多單，平多
                action = safe_constants.get_action('SELL')
            else:
                # 持空單，平空
                action = safe_constants.get_action('BUY')
        
        # 🔧 在API調用前就預先準備所有需要的變數和映射
        contract_name = "大台" if contract.code.startswith('TXF') else "小台" if contract.code.startswith('MXF') else "微台"
        
        # 使用臨時訂單ID機制，確保映射的唯一性和時序正確性
        temp_order_id = f"temp_{int(time.time() * 1000)}_{contract.code}"
        
        # 修正：將永豐API常數轉換為字符串（統一格式）
        direction_str = 'Buy' if action == sj.constant.Action.Buy else 'Sell'
        # 統一使用永豐API常數比較
        if hasattr(sj.constant, 'FuturesOCType'):
            octype_str = 'New' if octype == sj.constant.FuturesOCType.New else 'Cover'
        else:
            octype_str = 'New' if str(octype).upper() in ['NEW', 'OPEN'] else 'Cover'
        
        # 預先準備訂單資訊，確保資料的一致性
        order_info_template = {
            'octype': octype_str,
            'direction': direction_str,
            'contract_name': contract_name,
            'order_type': order_type,
            'price_type': price_type,
            'is_manual': is_manual,
            'timestamp': datetime.now().strftime('%H:%M:%S'),
            'pre_saved': True  # 標記為預先保存
        }
        
        # 預先保存到臨時映射，避免API調用和callback之間的競態條件
        with global_lock:
            order_octype_map[temp_order_id] = order_info_template.copy()
            logger.info(f"🔒 預先保存臨時訂單映射: {temp_order_id} -> {octype_str}/{direction_str}")
        
        # 建立訂單
        order = sinopac_api.Order(
            price=price if price_type == "LMT" else 0,  # 市價單價格設為0
            quantity=quantity,
            action=action,
            price_type=safe_constants.get_price_type(price_type),
            order_type=safe_constants.get_order_type(order_type),
            octype=safe_constants.get_oc_type(octype),
            account=sinopac_api.futopt_account
        )
        
        # 送出訂單
        trade = sinopac_api.place_order(contract, order)
        
        # 檢查訂單結果
        if not trade or not hasattr(trade, 'order') or not trade.order.id:
            # 清理臨時映射
            with global_lock:
                if temp_order_id in order_octype_map:
                    del order_octype_map[temp_order_id]
            raise Exception("訂單提交失敗")
            
        order_id = trade.order.id
        
        # ✅ 立即將臨時映射轉換為正式映射，確保callback能找到正確資訊
        with global_lock:
            # 將預先保存的資訊轉移到真實訂單ID
            order_octype_map[order_id] = order_info_template.copy()
            order_octype_map[order_id]['order_id'] = order_id
            order_octype_map[order_id]['pre_saved'] = False
            
            # 清理臨時映射
            if temp_order_id in order_octype_map:
                del order_octype_map[temp_order_id]
            
            # 立即保存，確保callback時能讀取到
            save_order_mapping()
            logger.info(f"✅ 正式保存訂單映射: {order_id} -> {octype_str}/{direction_str}")
        
        # 檢查操作結果
        if hasattr(trade, 'operation') and trade.operation.get('op_msg'):
            error_msg = trade.operation.get('op_msg')
            
            # 記錄掛單失敗日誌（僅後端顯示）
            log_message = get_simple_order_log_message(
                contract_name=contract_name,
                direction=direction_str,
                qty=quantity,
                price=price,
                order_id=order_id,
                octype=octype_str,
                is_manual=is_manual,
                is_success=False,
                order_type=order_type,
                price_type=price_type
            )
            # 僅在後端控制台顯示，不發送到前端
            
            fail_message = get_formatted_order_message(
                is_success=False,
                order_id=order_id,
                contract_name=contract_name,
                qty=quantity,
                price=price,
                octype=octype_str,  # 使用轉換後的字符串
                direction=direction_str,  # 使用轉換後的字符串
                order_type=order_type,
                price_type=price_type,
                is_manual=is_manual,
                reason=OP_MSG_TRANSLATIONS.get(error_msg, error_msg),
                contract_code=contract.code,
                delivery_date=contract.delivery_date.strftime('%Y/%m/%d')
            )
            send_telegram_message(fail_message)
            
            return {
                'success': False,
                'message': OP_MSG_TRANSLATIONS.get(error_msg, error_msg),
                'order_id': order_id
            }
        
        # 記錄掛單成功日誌（僅後端顯示）
        log_message = get_simple_order_log_message(
            contract_name=contract_name,
            direction=direction_str,
            qty=quantity,
            price=price,
            order_id=order_id,
            octype=octype_str,
            is_manual=is_manual,
            is_success=False,
            order_type=order_type,
            price_type=price_type
        )
        # 僅在後端控制台顯示，不發送到前端
        
        # 訂單提交成功 - 不需要立即發送通知，等callback處理
        logger.info(f"訂單提交成功: {order_id} - {contract_name} {quantity} 口 {direction}")
        
        return {
            'success': True,
            'message': '訂單提交成功',
            'order_id': order_id,
            'contract_name': contract_name
        }
            
    except Exception as e:
        error_msg = str(e)
        logger.error(f"下單失敗: {error_msg}")
        
        # 記錄掛單失敗日誌
        contract_name = "大台" if contract.code.startswith('TXF') else "小台" if contract.code.startswith('MXF') else "微台"
        
        # 判斷octype
        if direction in ["開多", "開空"]:
            error_octype = 'New'
        elif direction in ["平多", "平空"]:
            error_octype = 'Cover'
        else:
            error_octype = 'New' if position is None else 'Cover'
        
        log_message = get_simple_order_log_message(
            contract_name=contract_name,
            direction=direction,
            qty=quantity,
            price=price,
            order_id="未知",
            octype=error_octype,
            is_manual=is_manual,
            is_success=False,
            order_type=order_type,
            price_type=price_type
        )
        # 錯誤日誌記錄已移除
            
        # 發送錯誤通知
        fail_message = get_formatted_order_message(
            is_success=False,
            order_id="未知",
            contract_name=contract_name,
            qty=quantity,
            price=price,
            octype=error_octype,
            direction=direction,
            order_type=order_type,
            price_type=price_type,
            is_manual=is_manual,
            reason=error_msg,
            contract_code=contract.code if contract else None
        )
        send_telegram_message(fail_message)
        
        return {
            'success': False,
            'message': error_msg,
            'order_id': None
        }

def init_contracts():
    """初始化TXserver風格的合約對象"""
    global contract_txf, contract_mxf, contract_tmf
    
    try:
        logger.info("[init_contracts] 開始初始化合約對象...")
        
        # 初始化大台指合約（使用轉倉邏輯）
        txf_contracts = sinopac_api.Contracts.Futures.get("TXF")
        if txf_contracts:
            sorted_contracts = sorted(txf_contracts, key=lambda x: x.delivery_date)
            if rollover_mode and next_month_contracts.get('TXF'):
                contract_txf = next_month_contracts['TXF']
                logger.info(f"大台指合約（轉倉模式）: {contract_txf.code} (交割日: {contract_txf.delivery_date})")
            else:
                # 非轉倉模式：尋找R1合約作為當月合約
                r1_contract = None
                for contract in sorted_contracts:
                    if contract.code.endswith('R1'):
                        r1_contract = contract
                        break
                contract_txf = r1_contract if r1_contract else sorted_contracts[0]
                logger.info(f"大台指合約: {contract_txf.code} (交割日: {contract_txf.delivery_date})")
        else:
            logger.warning("警告: 無法獲取大台指合約")
            
        # 初始化小台指合約（使用轉倉邏輯）
        mxf_contracts = sinopac_api.Contracts.Futures.get("MXF")
        if mxf_contracts:
            sorted_contracts = sorted(mxf_contracts, key=lambda x: x.delivery_date)
            if rollover_mode and next_month_contracts.get('MXF'):
                contract_mxf = next_month_contracts['MXF']
                logger.info(f"小台指合約（轉倉模式）: {contract_mxf.code} (交割日: {contract_mxf.delivery_date})")
            else:
                # 非轉倉模式：尋找R1合約作為當月合約
                r1_contract = None
                for contract in sorted_contracts:
                    if contract.code.endswith('R1'):
                        r1_contract = contract
                        break
                contract_mxf = r1_contract if r1_contract else sorted_contracts[0]
                logger.info(f"小台指合約: {contract_mxf.code} (交割日: {contract_mxf.delivery_date})")
        else:
            logger.warning("警告: 無法獲取小台指合約")
            
        # 初始化微台指合約（使用轉倉邏輯）
        tmf_contracts = sinopac_api.Contracts.Futures.get("TMF")
        if tmf_contracts:
            sorted_contracts = sorted(tmf_contracts, key=lambda x: x.delivery_date)
            if rollover_mode and next_month_contracts.get('TMF'):
                contract_tmf = next_month_contracts['TMF']
                logger.info(f"微台指合約（轉倉模式）: {contract_tmf.code} (交割日: {contract_tmf.delivery_date})")
            else:
                # 非轉倉模式：尋找R1合約作為當月合約
                r1_contract = None
                for contract in sorted_contracts:
                    if contract.code.endswith('R1'):
                        r1_contract = contract
                        break
                contract_tmf = r1_contract if r1_contract else sorted_contracts[0]
                logger.info(f"微台指合約: {contract_tmf.code} (交割日: {contract_tmf.delivery_date})")
        else:
            logger.warning("警告: 無法獲取微台指合約")
            
        logger.info("[init_contracts] 合約對象初始化完成")
        
        # 合約對象初始化完成後，立即執行轉倉狀態檢查
        print_console("SYSTEM", "INFO", "合約對象初始化完成，執行轉倉狀態檢查...")
        try:
            check_rollover_mode()
            print_console("SYSTEM", "SUCCESS", "轉倉狀態檢查完成")
        except Exception as check_error:
            print_console("SYSTEM", "ERROR", "轉倉狀態檢查失敗", str(check_error))
        
    except Exception as e:
        logger.error(f"[init_contracts] 初始化合約對象失敗: {e}")
        contract_txf = None
        contract_mxf = None
        contract_tmf = None



def place_futures_order(contract_code, quantity, direction, price=0, is_manual=False, position_type=None, price_type=None, order_type=None, action_param=None, octype_param=None):
    """執行期貨下單
    contract_code: 合約代碼 (TXF, MXF, TMF)
    quantity: 數量
    direction: 方向 ('開多', '開空', '平多', '平空')
    price: 價格 (0為市價)
    is_manual: 是否為手動下單
    position_type: 持倉類型 (當direction為平倉時使用)
    price_type: 價格類型 (LMT/MKT)
    order_type: 單別 (ROD/IOC/FOK)
    """
    logger.info(f"🔥 place_futures_order 開始執行")
    logger.info(f"  - contract_code: {contract_code}")
    logger.info(f"  - quantity: {quantity}")
    logger.info(f"  - direction: {direction}")
    logger.info(f"  - price: {price}")
    logger.info(f"  - is_manual: {is_manual}")
    logger.info(f"  - action_param: {action_param}")
    logger.info(f"  - octype_param: {octype_param}")
    
    try:
        # 獲取合約資訊
        contracts = sinopac_api.Contracts.Futures.get(contract_code)
        if not contracts:
            raise Exception(f'無法獲取{contract_code}合約資訊')
        
        # 根據轉倉模式選擇合約
        if rollover_mode and next_month_contracts.get(contract_code):
            target_contract = next_month_contracts[contract_code]
        else:
            # 非轉倉模式：尋找R1合約作為當月合約
            sorted_contracts = sorted(contracts, key=lambda x: x.delivery_date)
            r1_contract = None
            for contract in sorted_contracts:
                if contract.code.endswith('R1'):
                    r1_contract = contract
                    break
            target_contract = r1_contract if r1_contract else sorted_contracts[0]
        
        
        # 永豐手動下單：使用永豐官方參數格式
        if is_manual:
            # 永豐手動下單應該使用永豐官方的參數格式
            # 前端應該傳遞 action (Buy/Sell) 和 octype (New/Cover) 參數
            # 而不是中文的 direction 參數
            
            # 使用傳入的永豐官方參數
            logger.info(f"永豐手動下單參數檢查:")
            logger.info(f"  action_param: '{action_param}'")
            logger.info(f"  octype_param: '{octype_param}'")
            
            if action_param and octype_param:
                # 使用永豐官方參數
                final_action = safe_constants.get_action(action_param)
                final_octype = safe_constants.get_oc_type(octype_param)
                logger.info(f"永豐手動下單使用官方參數: action={action_param} -> {final_action}, octype={octype_param} -> {final_octype}")
            else:
                # 如果沒有官方參數，顯示未知
                logger.error(f"錯誤: 永豐手動下單缺少官方參數")
                raise Exception('永豐手動下單缺少官方參數 action 和 octype')
        # WEBHOOK下單：使用 direction 參數
        else:
            if direction:
                if direction == "開多":
                    final_action = safe_constants.get_action('BUY')
                    final_octype = safe_constants.get_oc_type('New')
                    logger.info(f"WEBHOOK開多 -> BUY/New")
                elif direction == "開空":
                    final_action = safe_constants.get_action('SELL')
                    final_octype = safe_constants.get_oc_type('New')
                    logger.info(f"WEBHOOK開空 -> SELL/New")
                elif direction == "平多":
                    final_action = safe_constants.get_action('SELL')
                    final_octype = safe_constants.get_oc_type('Cover')
                    logger.info(f"WEBHOOK平多 -> SELL/Cover")
                elif direction == "平空":
                    final_action = safe_constants.get_action('BUY')
                    final_octype = safe_constants.get_oc_type('Cover')
                    logger.info(f"WEBHOOK平空 -> BUY/Cover")
                else:
                    logger.info(f"無效的WEBHOOK direction: '{direction}'")
                    raise Exception(f'無效的WEBHOOK交易方向: {direction}')
            else:
                logger.info(f"WEBHOOK缺少direction參數")
                raise Exception('WEBHOOK缺少direction參數')
        
        logger.info(f"最終: action={final_action}, octype={final_octype}")
        
        # 判斷訂單類型
        if is_manual:
            # 手動下單，使用傳入的參數或根據price判斷
            if price_type is None:
                final_price_type = safe_constants.get_price_type('MKT' if price == 0 else 'LMT')
            else:
                final_price_type = safe_constants.get_price_type(price_type)
            
            if order_type is None:
                final_order_type = safe_constants.get_order_type('ROD')  # 手動預設使用ROD
            else:
                final_order_type = safe_constants.get_order_type(order_type)
        else:
            # webhook下單，強制使用市價單IOC
            final_price_type = safe_constants.get_price_type('MKT')
            final_order_type = safe_constants.get_order_type('IOC')
            price = 0  # 確保使用市價
        
        # 🚨 關鍵修復：在API調用前準備所有變數和映射
        contract_name = '大台' if contract_code == 'TXF' else '小台' if contract_code == 'MXF' else '微台'
        
        # 修正：將永豐API常數轉換為字符串（統一格式）
        direction_str = 'Buy' if final_action == sj.constant.Action.Buy else 'Sell'
        octype_str = 'New' if final_octype == sj.constant.FuturesOCType.New else 'Cover'
        
        # 預先準備訂單資訊，確保資料的一致性
        order_info_template = {
            'octype': octype_str,
            'direction': direction_str,
            'contract_name': contract_name,
            'order_type': str(final_order_type),
            'price_type': str(final_price_type),
            'is_manual': is_manual,
            'timestamp': datetime.now().strftime('%H:%M:%S'),
            'pre_saved': False  # 標記為正式保存
        }
        
        # 建立訂單
        order = sinopac_api.Order(
            price=price,
            quantity=quantity,
            action=final_action,
            price_type=final_price_type,
            order_type=final_order_type,
            octype=final_octype,
            account=sinopac_api.futopt_account
        )
        
        # 送出訂單
        logger.info(f"🚀 開始提交訂單到永豐API...")
        trade = sinopac_api.place_order(target_contract, order)
        logger.info(f"🎯 訂單API調用完成，trade對象: {trade}")
        
        # 🚨 關鍵修復：立即獲取真實訂單ID並建立映射，搶在callback前執行
        if trade and hasattr(trade, 'order') and trade.order and trade.order.id:
            real_order_id = trade.order.id
            logger.info(f"⚡ 立即獲取真實訂單ID: {real_order_id}")
            
            # 立即建立真實訂單ID的映射
            with global_lock:
                order_octype_map[real_order_id] = order_info_template.copy()
                order_octype_map[real_order_id]['order_id'] = real_order_id
                logger.info(f"⚡ 緊急建立真實訂單映射: {real_order_id} -> {octype_str}/{direction_str}")
                
                # 立即保存到文件，確保callback能讀取到
                save_order_mapping()
                logger.info(f"⚡ 緊急保存映射文件完成")
        
        # 檢查是否有操作訊息
        if hasattr(trade, 'operation') and trade.operation.get('op_msg'):
            error_msg = trade.operation.get('op_msg')
            logger.error(f"訂單操作訊息: {error_msg}")
            
            # 準備訂單資訊用於失敗通知
            contract_name = '大台' if contract_code == 'TXF' else '小台' if contract_code == 'MXF' else '微台'
            order_id = trade.order.id if trade and trade.order else "未知"
            
            # 發送失敗通知 - 延遲1秒發送
            # 修正：將永豐API常數轉換為字符串
            direction_str = 'Buy' if final_action == sj.constant.Action.Buy else 'Sell'
            octype_str = 'New' if final_octype == sj.constant.FuturesOCType.New else 'Cover'
            
            fail_message = get_formatted_order_message(
                is_success=False,
                order_id=order_id,
                contract_name=contract_name,
                qty=quantity,
                price=price,
                octype=octype_str,
                direction=direction_str,
                order_type=str(final_order_type),
                price_type=str(final_price_type),
                is_manual=is_manual,
                reason=OP_MSG_TRANSLATIONS.get(error_msg, error_msg),
                contract_code=target_contract.code,
                delivery_date=target_contract.delivery_date.strftime('%Y/%m/%d')
            )
            
            # 記錄掛單失敗日誌
            log_message = get_simple_order_log_message(
                contract_name=contract_name,
                direction=direction_str,
                qty=quantity,
                price=price,
                order_id=order_id,
                octype=octype_str,
                is_manual=is_manual,
                is_success=False,
                order_type=str(final_order_type),
                price_type=str(final_price_type)
            )
            try:
                requests.post(
                    f'http://127.0.0.1:{CURRENT_PORT}/api/system_log',
                    # json={'message': log_message, 'type': 'error'},
                    timeout=5
                )
            except:
                pass
            
            # 延遲5秒發送失敗通知
            def delayed_send_fail():
                time.sleep(5)
                send_telegram_message(fail_message)
            
            create_managed_thread(target=delayed_send_fail, name="延遲發送失敗通知線程").start()
            
            raise Exception(error_msg)
        
        # 檢查訂單是否成功提交
        if not trade or not hasattr(trade, 'order') or not trade.order.id:
            raise Exception("訂單提交失敗")
        
        order_id = trade.order.id
        
        # 驗證映射是否已正確建立（前面應該已經建立了）
        if order_id not in order_octype_map:
            logger.warning(f"⚠️ 緊急映射建立失敗，使用備用方案: {order_id}")
            # 備用建立映射
            with global_lock:
                order_octype_map[order_id] = order_info_template.copy()
                order_octype_map[order_id]['order_id'] = order_id
                save_order_mapping()
                logger.info(f"🔧 備用保存訂單映射: {order_id} -> {octype_str}/{direction_str}")
        else:
            logger.info(f"✅ 確認訂單映射已存在: {order_id}")
        
        logger.info(f"訂單提交成功，單號: {order_id}")
        logger.info(f"訂單映射已建立: {order_info_template}")
        logger.info(f"當前 order_octype_map 內容: {order_octype_map}")
        
        # 記錄掛單成功日誌（僅後端顯示）
        log_message = get_simple_order_log_message(
            contract_name=contract_name,
            direction=direction_str,
            qty=quantity,
            price=price,
            order_id=order_id,
            octype=octype_str,
            is_manual=is_manual,
            is_success=True,  # 修復：提交成功應該是True
            order_type=str(final_order_type),
            price_type=str(final_price_type)
        )
        # 僅在後端控制台顯示，不發送到前端
        
        # 返回訂單資訊
        return {
            'contract_code': contract_code,
            'contract_name': target_contract.code,
            'delivery_date': target_contract.delivery_date.strftime('%Y/%m/%d'),
            'quantity': quantity,
            'action': final_action,
            'position_type': position_type,
            'order_id': order_id,
            'status': 'submitted',
            'contract_type': contract_name,
            'price': price,
            'price_type': final_price_type,
            'order_type': final_order_type
        }
        
    except Exception as e:
        raise Exception(f'{contract_code}下單失敗: {str(e)}')



# 移除舊的通知函數，改用新的回調機制

def send_telegram_file(file_path, caption=""):
    """發送Telegram檔案"""
    try:
        if not os.path.exists(ENV_PATH):
            logger.info(f"找不到 .env 檔案，路徑: {ENV_PATH}")
            return False
            
        if not os.path.exists(file_path):
            logger.info(f"找不到檔案，路徑: {file_path}")
            return False
        
        # 讀取環境變數
        with open(ENV_PATH, 'r', encoding='utf-8') as f:
            env_content = f.read()
        
        # 解析環境變數
        bot_token = None
        chat_id_raw = None
        
        for line in env_content.split('\n'):
            if line.strip().startswith('BOT_TOKEN='):
                bot_token = line.split('=', 1)[1].strip().strip('"')
            elif line.strip().startswith('CHAT_ID='):
                chat_id_raw = line.split('=', 1)[1].strip().strip('"')
        
        if not bot_token or not chat_id_raw:
            logger.warning("Telegram設定不完整")
            return False
        
        # 支援多個CHAT_ID，用逗號分隔
        chat_ids = [id.strip() for id in chat_id_raw.split(',') if id.strip()]
        
        logger.info(f"準備發送檔案到 {len(chat_ids)} 個接收者")
        
        # 發送檔案
        url = f"https://api.telegram.org/bot{bot_token}/sendDocument"
        
        # 記錄發送結果
        success_count = 0
        total_count = len(chat_ids)
        
        # 對每個CHAT_ID發送檔案
        for chat_id in chat_ids:
            with open(file_path, 'rb') as f:
                files = {'document': f}
                data = {
                    'chat_id': chat_id,
                    'caption': caption
                }
                
                logger.info(f"發送檔案到 Chat ID: {chat_id}")
                response = requests.post(url, files=files, data=data, timeout=30)
                
                if response.status_code == 200:
                    logger.info(f"Telegram檔案發送成功 (Chat ID: {chat_id})")
                    success_count += 1
                else:
                    logger.error(f"Telegram檔案發送失敗 (Chat ID: {chat_id}): {response.status_code}")
            
        # 判斷整體發送結果
        if success_count == total_count:
            logger.info(f"Telegram檔案發送完成！成功發送到 {success_count}/{total_count} 個接收者")
            
            try:
                # 生成報表通知
                requests.post(
                    f'http://127.0.0.1:{CURRENT_PORT}/api/system_log',
                    # json={'message': 'Telegram［生成報表］訊息發送成功！！！', 'type': 'success'},
                    timeout=5
                )
                
                # 檔案發送通知
                requests.post(
                    f'http://127.0.0.1:{CURRENT_PORT}/api/system_log',
                    # json={'message': 'Telegram［檔案發送］訊息發送成功！！！', 'type': 'success'},
                    timeout=5
                )
            except:
                pass
            
            return True
        else:
            logger.warning(f"Telegram檔案部分發送失敗！成功發送到 {success_count}/{total_count} 個接收者")
            
            # 記錄失敗日誌
            try:
                status_type = 'warning' if success_count > 0 else 'error'
                requests.post(
                    f'http://127.0.0.1:{CURRENT_PORT}/api/system_log',
                    # json={'message': f'Telegram檔案部分發送失敗！成功：{success_count}/{total_count}', 'type': status_type},
                    timeout=5
                )
            except:
                pass
            
            return success_count > 0  # 至少有一個成功就返回True
            
    except Exception as e:
        logger.error(f"發送Telegram檔案失敗: {e}")
        return False

def send_telegram_message(message, log_type="info"):
    """發送Telegram訊息"""
    try:
        
        if not os.path.exists(ENV_PATH):
            logger.info(f"找不到 .env 檔案，路徑: {ENV_PATH}")
            return False
        
        load_dotenv(ENV_PATH)
        bot_token = os.getenv('BOT_TOKEN')
        chat_id_raw = os.getenv('CHAT_ID')
        
        if not bot_token:
            logger.info("找不到 BOT_TOKEN")
            return False
        if not chat_id_raw:
            logger.info("找不到 CHAT_ID")
            return False
        
        # 支援多個CHAT_ID，用逗號分隔
        chat_ids = [id.strip() for id in chat_id_raw.split(',') if id.strip()]
        
        logger.info(f"BOT_TOKEN: {bot_token[:10]}...")
        logger.info(f"CHAT_IDs: {chat_ids}")
        
        url = f'https://api.telegram.org/bot{bot_token}/sendMessage'
        
        # 記錄發送結果
        success_count = 0
        total_count = len(chat_ids)
        
        # 對每個CHAT_ID發送訊息
        for chat_id in chat_ids:
            payload = {
                'chat_id': chat_id,
                'text': message,
                'parse_mode': 'HTML'
            }
            
            logger.info(f"發送請求到 Telegram API (Chat ID: {chat_id})...")
            response = requests.post(url, json=payload, timeout=10)
            
            logger.info(f"Telegram API 回應 (Chat ID: {chat_id}): {response.status_code}")
            
            if response.status_code == 200:
                logger.info(f"Telegram 訊息發送成功 (Chat ID: {chat_id})！")
                success_count += 1
            else:
                logger.error(f"Telegram 訊息發送失敗 (Chat ID: {chat_id}): {response.text}")
        
        # 判斷整體發送結果
        if success_count == total_count:
            logger.info(f"Telegram 訊息發送完成！成功發送到 {success_count}/{total_count} 個接收者")
            
            # 根據訊息內容判斷發送狀態類型
            if "提交成功" in message:
                log_message = "Telegram［提交成功］訊息發送成功！！！"
            elif "提交失敗" in message:
                log_message = "Telegram［提交失敗］訊息發送成功！！！"
            elif "成交通知" in message:
                log_message = "Telegram［成交通知］訊息發送成功！！！"
            elif "API連線異常" in message:
                log_message = "Telegram［API連線異常］訊息發送成功！！！"
            elif "API連線成功" in message or "API重新連線成功" in message:
                log_message = "Telegram［API連線成功］訊息發送成功！！！"
            elif "交易統計" in message:
                log_message = "Telegram［交易統計］訊息發送成功！！！"
            elif "交易報表" in message or "交易報表" in message:
                log_message = "Telegram［生成報表］訊息發送成功！！！"
            elif "保證金管理警報" in message or "保證金不足" in message:
                log_message = "Telegram［保證金管理］訊息發送成功！！！"
            elif "保證金" in message or "轉倉" in message:
                log_message = "Telegram［系統通知］訊息發送成功！！！"
            else:
                log_message = "Telegram 訊息發送成功！！！"
            
            # 日誌記錄已移除
            
            return True
        else:
            logger.error(f"Telegram 訊息部分發送失敗！成功發送到 {success_count}/{total_count} 個接收者")
            # 發送失敗也要記錄日誌
            try:
                error_log_message = f"Telegram 訊息部分發送失敗！成功：{success_count}/{total_count}"
                requests.post(
                    f'http://127.0.0.1:{CURRENT_PORT}/api/system_log',
                    # json={'message': error_log_message, 'type': 'warning'},
                    timeout=5
                )
                logger.error(f"TX系統錯誤日誌已發送: {error_log_message}")
            except Exception as e:
                logger.error(f"發送TX系統錯誤日誌失敗: {e}")
            return success_count > 0  # 至少有一個成功就返回True
            
    except Exception as e:
        logger.error(f"發送Telegram訊息失敗: {e}")
        logger.error(f"錯誤類型: {str(e.__class__.__name__)}")
        if hasattr(e, 'response'):
            logger.info(f"回應內容: {e.response.text}")
        import traceback
        traceback.print_exc()
        return False

# 模擬通知函數已移除

# 移除舊的send_order_notification函數，新的回調機制會自動處理

# 使用標準化的交易記錄目錄
from trading_config import TradingConfig
TX_DATA_DIR = TradingConfig.TX_DATA_DIR
TX_RECORDS_DIR = TradingConfig.TX_RECORDS_DIR
BTC_DATA_DIR = TradingConfig.BTC_DATA_DIR
BTC_RECORDS_DIR = TradingConfig.BTC_RECORDS_DIR

# 為了向後兼容
TX_LOG_DIR = TX_DATA_DIR
BTC_LOG_DIR = BTC_DATA_DIR
LOG_DIR = TX_LOG_DIR

def save_trade(data):
    """保存交易記錄到JSON文件（企業級改進版本）"""
    try:
        today = datetime.now().strftime("%Y%m%d")
        filename = os.path.join(TX_DATA_DIR, f"TXtransdata_{today}.json")
        os.makedirs(TX_DATA_DIR, exist_ok=True)
        
        # 🔧 修復文件句柄洩漏問題 - 使用with語句確保文件正確關閉
        trades = []
        if os.path.exists(filename):
            try:
                with open(filename, 'r', encoding='utf-8') as f:
                    trades = json.load(f)
            except (json.JSONDecodeError, FileNotFoundError, PermissionError) as e:
                logger.error(f"讀取交易記錄檔案失敗: {filename}, 錯誤: {str(e)}")
                send_telegram_message(f"❌ 交易記錄檔案讀取失敗，已重置為空: {str(e)[:50]}")
                trades = []
        
        # 🔧 增強數據完整性檢查
        if not isinstance(trades, list):
            logger.warning(f"交易記錄格式異常，重置為空列表")
            trades = []
            
        # 添加時間戳和數據驗證
        data['timestamp'] = datetime.now().isoformat()
        data['save_attempt'] = len(trades) + 1  # 記錄保存順序
        
        trades.append(data)
        
        # 🔧 安全的文件寫入機制 - 先寫臨時文件再重命名（原子操作）
        temp_filename = f"{filename}.tmp"
        try:
            with open(temp_filename, 'w', encoding='utf-8') as f:
                json.dump(trades, f, indent=2, ensure_ascii=False)
            
            # 原子操作 - 重命名替換原文件
            import shutil
            shutil.move(temp_filename, filename)
            
            # 📊 記錄成功保存的統計信息
            logger.info(f"✅ 交易記錄保存成功: {data.get('type', '未知')}類型, 檔案: {os.path.basename(filename)}, 記錄數: {len(trades)}")
            
        except Exception as write_error:
            # 清理臨時文件
            if os.path.exists(temp_filename):
                try:
                    os.remove(temp_filename)
                except:
                    pass
            raise write_error
        
        # 清理舊的交易記錄檔案（保留30個交易日）
        cleanup_old_trade_files()
        
    except Exception as e:
        error_msg = f"儲存交易記錄失敗：{str(e)}"
        logger.error(error_msg)
        send_telegram_message(f"❌ {error_msg[:100]}")
        
        # 🔧 緊急情況下的數據備份機制
        try:
            backup_filename = os.path.join(TX_DATA_DIR, f"emergency_backup_{today}_{int(datetime.now().timestamp())}.json")
            with open(backup_filename, 'w', encoding='utf-8') as f:
                json.dump([data], f, indent=2, ensure_ascii=False)
            logger.warning(f"🆘 緊急備份已創建: {backup_filename}")
        except:
            logger.critical("💥 緊急備份也失敗了！數據可能丟失！")

def diagnose_trading_records():
    """🔍 專業級交易記錄診斷系統"""
    try:
        logger.info("🔍 開始執行交易記錄系統診斷...")
        
        today = datetime.now().strftime("%Y%m%d")
        today_file = os.path.join(TX_DATA_DIR, f"TXtransdata_{today}.json")
        
        # 檢查目錄狀態
        if not os.path.exists(TX_DATA_DIR):
            logger.error(f"❌ TX數據目錄不存在: {TX_DATA_DIR}")
            return {'status': 'error', 'message': '數據目錄不存在'}
        
        # 檢查目錄權限
        try:
            test_file = os.path.join(TX_DATA_DIR, "test_write.tmp")
            with open(test_file, 'w') as f:
                f.write("test")
            os.remove(test_file)
            logger.info(f"✅ 目錄寫入權限正常: {TX_DATA_DIR}")
        except Exception as e:
            logger.error(f"❌ 目錄權限問題: {str(e)}")
            return {'status': 'error', 'message': f'目錄權限問題: {str(e)}'}
        
        # 檢查今日文件
        if os.path.exists(today_file):
            try:
                with open(today_file, 'r', encoding='utf-8') as f:
                    records = json.load(f)
                
                logger.info(f"📊 今日記錄文件分析:")
                logger.info(f"   ├─ 文件路徑: {today_file}")
                logger.info(f"   ├─ 文件大小: {os.path.getsize(today_file)} bytes")
                logger.info(f"   ├─ 記錄總數: {len(records)}")
                
                # 類型統計
                type_stats = {}
                for record in records:
                    record_type = record.get('type', 'unknown')
                    type_stats[record_type] = type_stats.get(record_type, 0) + 1
                
                for record_type, count in type_stats.items():
                    logger.info(f"   ├─ {record_type}: {count} 筆")
                
                return {
                    'status': 'success',
                    'file_exists': True,
                    'record_count': len(records),
                    'type_stats': type_stats
                }
                
            except json.JSONDecodeError as e:
                logger.error(f"❌ JSON格式錯誤: {str(e)}")
                return {'status': 'error', 'message': f'JSON格式錯誤: {str(e)}'}
                
        else:
            logger.warning(f"⚠️ 今日交易記錄文件不存在: {today_file}")
            return {
                'status': 'warning',
                'file_exists': False,
                'message': '今日尚無交易記錄'
            }
            
    except Exception as e:
        logger.error(f"❌ 診斷系統異常: {str(e)}")
        return {'status': 'error', 'message': f'診斷異常: {str(e)}'}

def cleanup_old_trade_files():
    """清理舊的交易記錄檔案，保留30個交易日"""
    try:
        if not os.path.exists(TX_DATA_DIR):
            return
        
        # 獲取所有交易記錄檔案
        trade_files = []
        for filename in os.listdir(TX_DATA_DIR):
            if filename.startswith('TXtransdata_') and filename.endswith('.json'):
                try:
                    # 從檔案名提取日期
                    date_str = filename.replace('TXtransdata_', '').replace('.json', '')
                    file_date = datetime.strptime(date_str, '%Y%m%d')
                    trade_files.append((filename, file_date))
                except ValueError:
                    # 如果檔案名格式不正確，跳過
                    continue
        
        # 按日期排序
        trade_files.sort(key=lambda x: x[1], reverse=True)
        
        # 保留最新的30個檔案，刪除其餘的
        if len(trade_files) > 30:
            files_to_delete = trade_files[30:]
            for filename, file_date in files_to_delete:
                file_path = os.path.join(TX_DATA_DIR, filename)
                try:
                    os.remove(file_path)
                    logger.info(f"已刪除舊交易記錄檔案：{filename}")
                except Exception as e:
                    logger.error(f"刪除檔案失敗 {filename}：{e}")
            
            logger.info(f"清理完成：保留 {len(trade_files) - len(files_to_delete)} 個檔案，刪除 {len(files_to_delete)} 個舊檔案")
    
    except Exception as e:
        logger.error(f"清理舊交易記錄檔案失敗：{e}")

# 公共工具函數
def get_sort_date(contract):
    """按交割日期排序合約的公共函數"""
    date_str = contract.delivery_date
    if isinstance(date_str, str):
        if len(date_str) == 8:  # YYYYMMDD
            return date_str
        elif '-' in date_str:  # YYYY-MM-DD
            return date_str.replace('-', '')
    return str(date_str)

def get_third_wednesday(year, month):
    """計算該月的第三個星期三（台指期貨交割日）"""
    import calendar
    
    # 找到該月第一天是星期幾 (0=星期一, 6=星期日)
    first_day = datetime(year, month, 1)
    first_weekday = first_day.weekday()  # 0=星期一, 6=星期日
    
    # 找到第一個星期三 (星期三是weekday=2)
    if first_weekday <= 2:  # 星期一、星期二、星期三
        first_wednesday = 3 - first_weekday
    else:  # 星期四到星期日
        first_wednesday = 10 - first_weekday
    
    # 第三個星期三 = 第一個星期三 + 14天
    third_wednesday = first_wednesday + 14
    
    return third_wednesday

def format_delivery_date(delivery_date):
    """格式化交割日期的公共函數，統一格式為 YYYY/MM/DD"""
    if isinstance(delivery_date, str):
        if len(delivery_date) == 8:  # YYYYMMDD
            return f"{delivery_date[:4]}/{delivery_date[4:6]}/{delivery_date[6:8]}"
        elif '-' in delivery_date:  # YYYY-MM-DD
            return delivery_date.replace('-', '/')
        elif '/' in delivery_date:  # 已經是正確格式
            return delivery_date
    elif hasattr(delivery_date, 'strftime'):  # datetime 對象
        return delivery_date.strftime('%Y/%m/%d')
    return str(delivery_date)

# 轉倉相關函數
def get_next_month_contracts():
    """獲取次月合約"""
    global next_month_contracts, sinopac_api
    
    try:
        if not sinopac_connected or not sinopac_api:
            return {}
        
        next_month_contracts = {}
        
        for code in ['TXF', 'MXF', 'TMF']:
            try:
                contracts = sinopac_api.Contracts.Futures.get(code)
                if contracts:
                    # 按交割日期排序
                    sorted_contracts = sorted(contracts, key=get_sort_date)
                    
                    # 優先使用R2合約作為次月合約，如果沒有R2則使用第二個合約
                    next_month_contract = None
                    
                    # 方法1: 尋找R2合約
                    for contract in sorted_contracts:
                        if contract.code.endswith('R2'):
                            next_month_contract = contract
                            break
                    
                    # 方法2: 如果沒有R2合約，使用第二個合約
                    if not next_month_contract and len(sorted_contracts) >= 2:
                        next_month_contract = sorted_contracts[1]
                    
                    if next_month_contract:
                        next_month_contracts[code] = next_month_contract
                        
            except Exception as e:
                logger.error(f"獲取{code}次月合約失敗: {e}")
                
        return next_month_contracts
        
    except Exception as e:
        logger.error(f"獲取次月合約失敗: {e}")
        return {}

def check_rollover_mode():
    """檢查是否應該進入轉倉模式"""
    global rollover_mode, rollover_start_date, contract_txf, contract_mxf, contract_tmf
    
    try:
        today = datetime.now().date()
        # 檢查當前合約的交割日
        current_contracts = [contract_txf, contract_mxf, contract_tmf]
        delivery_dates = []
        
        for contract in current_contracts:
            if contract and hasattr(contract, 'delivery_date'):
                delivery_date_str = contract.delivery_date
                contract_code = getattr(contract, 'code', 'Unknown')
                
                if isinstance(delivery_date_str, str):
                    if len(delivery_date_str) == 8:  # YYYYMMDD
                        delivery_date = datetime.strptime(delivery_date_str, '%Y%m%d').date()
                    elif '/' in delivery_date_str:  # YYYY/MM/DD
                        delivery_date = datetime.strptime(delivery_date_str, '%Y/%m/%d').date()
                    elif '-' in delivery_date_str:  # YYYY-MM-DD
                        delivery_date = datetime.strptime(delivery_date_str, '%Y-%m-%d').date()
                    else:
                        continue
                    delivery_dates.append(delivery_date)
        
        if delivery_dates:
            # 找到最近的交割日
            nearest_delivery = min(delivery_dates)
            # 轉倉開始日期 = 交割日前一天
            rollover_start = nearest_delivery - timedelta(days=1)
            
            formatted_nearest = format_delivery_date(str(nearest_delivery))
            formatted_rollover = format_delivery_date(str(rollover_start))
            
            # 檢查轉倉模式狀態
            # 轉倉期間：交割日前一天到交割日當天夜盤開始（15:00）前
            # 交割日當天15:00後才結束轉倉模式（夜盤開始使用新合約）
            current_time = datetime.now()
            if today == nearest_delivery and current_time.hour >= 15:
                # 交割日當天15:00後，結束轉倉模式
                rollover_end = nearest_delivery
            else:
                # 其他情況，交割日隔天才結束轉倉模式
                rollover_end = nearest_delivery + timedelta(days=1)
            
            # 更精確的轉倉期間判斷
            in_rollover_period = False
            if today < nearest_delivery:
                # 交割日前一天及之前
                in_rollover_period = (today >= rollover_start)
            elif today == nearest_delivery:
                # 交割日當天，15:00前仍在轉倉期間
                in_rollover_period = (current_time.hour < 15)
            else:
                # 交割日之後
                in_rollover_period = False
            
            if in_rollover_period:
                # 轉倉期間：交割日前一天到交割日當天之前
                if not rollover_mode:
                    rollover_mode = True
                    rollover_start_date = rollover_start
                    print_console("TRADE", "INFO", f"進入轉倉模式，交割日: {formatted_nearest}")
                    
                    # 獲取次月合約
                    get_next_month_contracts()
                else:
                    
                    # 系統重啟後的轉倉檢查，不發送通知（避免重複發送）
                    # 因為轉倉通知應該只在定時檢查（00:05）時發送一次
                    should_send_notification = False
                    
                    # 發送轉倉通知（僅在首次進入轉倉模式時）
                    if should_send_notification:
                        # 獲取次月合約的交割日期
                        next_month_delivery = None
                        if next_month_contracts:
                            # 獲取任一次月合約的交割日期（所有合約交割日相同）
                            for contract in next_month_contracts.values():
                                if contract and hasattr(contract, 'delivery_date'):
                                    next_month_delivery = format_delivery_date(contract.delivery_date)
                                    break
                        
                        rollover_message = f"🔄 自動轉倉已啟動！！！\n本月合約交割日: {formatted_nearest}\n下次開倉將使用次月合約"
                        if next_month_delivery:
                            rollover_message += f"\n次月合約交割日：{next_month_delivery}"
                        
                        send_telegram_message(rollover_message)
                        
                        # 轉倉日誌記錄已移除
                    else:
                        # 系統重啟後的轉倉檢查，不發送通知
                        print_console("SYSTEM", "INFO", f"系統重啟後檢測到轉倉模式，交割日: {formatted_nearest}，不重複發送通知")
                    
                    # 通知前端更新選用合約顯示為次月合約
                    print_console("SYSTEM", "INFO", "轉倉模式已啟動，前端合約顯示將切換至次月合約")
                    
                return True
            else:
                # 不在轉倉期間
                if rollover_mode:
                    # 檢查是否應該結束轉倉模式
                    should_end_rollover = False
                    
                    if today > nearest_delivery:
                        # 交割日之後，結束轉倉模式
                        should_end_rollover = True
                    elif today == nearest_delivery and current_time.hour >= 15:
                        # 交割日當天15:00後，結束轉倉模式
                        should_end_rollover = True
                    elif today < rollover_start:
                        # 還未到轉倉時間，結束轉倉模式
                        should_end_rollover = True
                    
                    if should_end_rollover:
                        rollover_mode = False
                        rollover_start_date = None
                        next_month_contracts.clear()
                        rollover_processed_signals.clear()
                        
                        if today >= nearest_delivery:
                            print_console("TRADE", "SUCCESS", f"交割日夜盤開始，結束轉倉模式，切換至新月份合約")
                            # 強制重新初始化合約對象，以更新為新的當月合約
                            contract_txf = None
                            contract_mxf = None 
                            contract_tmf = None
                            print_console("SYSTEM", "INFO", "已重置合約對象，將重新初始化為新當月合約")
                        else:
                            print_console("SYSTEM", "INFO", "退出轉倉模式（尚未到轉倉時間）")
                    
                return False
        else:
            return False
        
    except Exception as e:
        print_console("SYSTEM", "ERROR", "檢查轉倉模式失敗", str(e))
        return False

# ========================== TX風險管理功能 ==========================
# 風險警報發送頻率控制
tx_last_risk_alert_time = {}

def calculate_tx_risk_metrics():
    """計算TX風險指標"""
    try:
        # 獲取帳戶資訊
        if not sinopac_connected or not sinopac_api:
            return None
            
        # 獲取保證金資訊
        margin_data = sinopac_api.margin()
        
        # 獲取持倉資訊計算未實現損益
        total_pnl = 0.0
        try:
            positions = sinopac_api.list_positions(sinopac_api.futopt_account)
            for pos in positions:
                total_pnl += pos.pnl
        except:
            total_pnl = 0.0
            
        account_data = {
            '權益總值': getattr(margin_data, 'equity', 0) or 0,
            '權益總額': getattr(margin_data, 'equity_amount', 0) or 0,
            '今日餘額': getattr(margin_data, 'today_balance', 0) or 0,
            '昨日餘額': getattr(margin_data, 'yesterday_balance', 0) or 0,
            '可用保證金': getattr(margin_data, 'available_margin', 0) or 0,
            '原始保證金': getattr(margin_data, 'initial_margin', 0) or 0,
            '維持保證金': getattr(margin_data, 'maintenance_margin', 0) or 0,
            '風險指標': getattr(margin_data, 'risk_indicator', 0) or 0,
            '手續費': getattr(margin_data, 'fee', 0) or 0,
            '期交稅': getattr(margin_data, 'tax', 0) or 0,
            '本日平倉損益': getattr(margin_data, 'future_settle_profitloss', 0) or 0,
            '未實現損益': total_pnl
        }
        
        # 提取關鍵風險指標
        risk_metrics = {
            'available_margin': account_data.get('可用保證金', 0),
            'maintenance_margin': account_data.get('維持保證金', 0),
            'initial_margin': account_data.get('原始保證金', 0),
            'equity_amount': account_data.get('權益總額', 0),
            'today_balance': account_data.get('今日餘額', 0),
            'risk_indicator': account_data.get('風險指標', 0),
            'unrealized_pnl': account_data.get('未實現損益', 0)
        }
        
        # 判斷風險等級 - 修復：使用權益總額判斷
        equity_total = risk_metrics['equity']
        maintenance_margin = risk_metrics['maintenance_margin']
        
        if equity_total < maintenance_margin:
            risk_metrics['risk_level'] = 'HIGH'
        elif equity_total < maintenance_margin * 1.5:
            risk_metrics['risk_level'] = 'MEDIUM'
        else:
            risk_metrics['risk_level'] = 'SAFE'
        
        return risk_metrics
    except Exception as e:
        logger.error(f"計算TX風險指標失敗: {e}")
        return None

def check_tx_risk_alerts():
    """檢查TX風險並發送警報"""
    global tx_last_risk_alert_time
    
    try:
        risk_metrics = calculate_tx_risk_metrics()
        if not risk_metrics:
            return
        
        alerts = []
        current_time = datetime.now()
        
        # 保證金不足警報 - 修復：使用權益總額而不是可用保證金
        equity_total = risk_metrics['equity']  # 權益總額
        maintenance_margin = risk_metrics['maintenance_margin']
        
        if equity_total < maintenance_margin:  # 權益總額低於維持保證金發警報
            alert_key = 'MARGIN_INSUFFICIENT'
            
            # 檢查是否需要發送警報（每小時最多一次）
            last_alert = tx_last_risk_alert_time.get(alert_key)
            if not last_alert or (current_time - last_alert).total_seconds() > 3600:
                alerts.append({
                    'type': 'MARGIN_INSUFFICIENT',
                    'level': 'WARNING',
                    'message': f'權益總額低於維持保證金\n目前權益總額：{equity_total:,.0f} TWD\n維持保證金需求：{maintenance_margin:,.0f} TWD'
                })
                tx_last_risk_alert_time[alert_key] = current_time
        
        # 發送警報
        for alert in alerts:
            send_tx_risk_alert(alert)
            
    except Exception as e:
        logger.error(f"檢查TX風險警報失敗: {e}")

def send_tx_risk_alert(alert):
    """發送TX風險警報"""
    try:
        current_time = datetime.now()
        current_time_str = current_time.strftime('%Y/%m/%d')
        
        if alert['type'] == 'MARGIN_INSUFFICIENT':
            message = f"⚠️ TX保證金管理警報 ({current_time_str})\n{alert['message']}\n建議減倉或增加保證金\n請及時處理以避免損失！"
        else:
            message = f"⚠️ TX風險管理警報 ({current_time_str})\n{alert['message']}\n請及時處理以避免損失！"
        
        # 發送Telegram通知
        result = send_telegram_message(message)
        
        # 記錄到風險日誌
        if result:
            logger.info(f"TX風險警報已發送: {message[:50]}...")
        else:
            logger.error(f"TX風險警報發送失敗: {message[:50]}...")
            
    except Exception as e:
        logger.error(f"發送TX風險警報失敗: {e}")

def start_tx_risk_monitor():
    """啟動TX風險監控"""
    def tx_risk_check_loop():
        while True:
            try:
                # 每小時檢查一次
                check_tx_risk_alerts()
                time.sleep(3600)  # 休眠1小時
            except Exception as e:
                logger.error(f"TX風險監控循環錯誤: {e}")
                time.sleep(300)  # 出錯時休眠5分鐘後重試
    
    risk_thread = threading.Thread(target=tx_risk_check_loop, daemon=True)
    risk_thread.start()
    logger.info("TX風險監控已啟動")

def get_contract_for_rollover(contract_type):
    """根據轉倉模式獲取合約"""
    global rollover_mode, next_month_contracts, contract_txf, contract_mxf, contract_tmf
    
    if not rollover_mode:
        # 非轉倉模式，使用當前合約
        if contract_type == 'TXF':
            return contract_txf
        elif contract_type == 'MXF':
            return contract_mxf
        elif contract_type == 'TMF':
            return contract_tmf
        else:
            return None
    
    # 轉倉模式，使用次月合約
    next_month_contract = next_month_contracts.get(contract_type)
    if next_month_contract:
        logger.info(f"轉倉模式: 使用次月{contract_type}合約 {next_month_contract.code}")
        return next_month_contract
    else:
        # 如果沒有次月合約，回退到當前合約
        logger.warning(f"警告: 沒有次月{contract_type}合約，使用當前合約")
        if contract_type == 'TXF':
            return contract_txf
        elif contract_type == 'MXF':
            return contract_mxf
        elif contract_type == 'TMF':
            return contract_tmf
        else:
            return None

def process_rollover_signal(data):
    """處理轉倉訊號"""
    global rollover_processed_signals
    
    signal_id = data.get('tradeId')
    if signal_id in rollover_processed_signals:
        logger.info(f"轉倉訊號 {signal_id} 已處理，跳過")
        return True
    
    # 檢查是否為轉倉模式
    if not check_rollover_mode():
        return False
    
    # 檢查是否為第一個webhook訊號
    if len(rollover_processed_signals) == 0:
        logger.info("收到轉倉模式下的第一個webhook訊號")
        rollover_processed_signals.add(signal_id)
        return True
    
    return False

def start_rollover_checker():
    """啟動轉倉檢查器"""
    def rollover_check_loop():
        while True:
            try:
                # 每天凌晨00:05檢查一次轉倉模式（避免與其他檢查衝突）
                now = datetime.now()
                if now.hour == 0 and now.minute >= 5:
                    check_rollover_mode()
                    # 檢查後等待到明天凌晨
                    tomorrow = now.replace(hour=0, minute=5, second=0, microsecond=0) + timedelta(days=1)
                    sleep_seconds = (tomorrow - now).total_seconds()
                    time.sleep(sleep_seconds)
                else:
                    # 計算到明天凌晨00:05的時間
                    tomorrow = now.replace(hour=0, minute=5, second=0, microsecond=0) + timedelta(days=1)
                    sleep_seconds = (tomorrow - now).total_seconds()
                    time.sleep(sleep_seconds)
                    
            except Exception as e:
                logger.error(f"轉倉檢查器錯誤: {e}")
                time.sleep(3600)  # 發生錯誤時等待1小時
    
    rollover_thread = create_managed_thread(target=rollover_check_loop, name="轉倉檢查線程")
    rollover_thread.start()
    print_console("SYSTEM", "SUCCESS", "轉倉檢查器已啟動（每天凌晨00:05檢查）")

# 斷線重連相關函數
def check_api_connection():
    """檢查API連線狀態"""
    global sinopac_connected, sinopac_api, reconnect_attempts, last_connection_check
    
    try:
        last_connection_check = datetime.now()
        
        # 如果API未初始化，跳過檢查
        if not sinopac_api:
            return True
        
        # 嘗試獲取帳戶資訊來測試連線
        try:
            # 簡單的API調用來測試連線
            test_result = sinopac_api.list_positions(sinopac_api.futopt_account)
            # 如果成功執行，重置重連計數
            reconnect_attempts = 0
            return True
        except Exception as e:
            logger.error(f"API連線檢查失敗: {e}")
            return False
            
    except Exception as e:
        logger.error(f"檢查API連線時發生錯誤: {e}")
        return False

def get_dynamic_check_interval():
    """根據交易時間動態調整檢查間隔"""
    try:
        # 檢查是否為交易時間
        response = requests.get(f'http://127.0.0.1:{CURRENT_PORT}/api/trading/status', timeout=5)
        if response.status_code == 200:
            data = response.json()
            is_trading_day = data.get('is_trading_day', False)
            is_market_open = data.get('is_market_open', False)
            
            # 交易時間：每1分鐘檢查
            if is_trading_day and is_market_open:
                return 60  # 1分鐘
            # 非交易時間：每10分鐘檢查
            else:
                return 3600  # 60分鐘
        else:
            # 如果無法獲取交易狀態，預設使用較長的間隔
            return 600
    except Exception as e:
        logger.error(f"獲取動態檢查間隔失敗: {e}")
        # 發生錯誤時使用較長的間隔
        return 600

def start_connection_monitor():
    """啟動連線監控器（智能檢查頻率）"""
    global connection_monitor_timer
    
    def connection_monitor_loop():
        global reconnect_attempts, is_reconnecting
        
        while True:
            try:
                # 動態獲取檢查間隔
                check_interval = get_dynamic_check_interval()
                
                # 根據當前狀態決定檢查頻率
                if is_reconnecting:
                    # 重連中：每30秒檢查一次
                    sleep_time = 30
                    logger.info(f"重連中，{sleep_time}秒後檢查連線狀態...")
                else:
                    # 正常狀態：使用動態間隔
                    sleep_time = check_interval
                    if check_interval == 60:
                        logger.info(f"交易時間，{sleep_time}秒後檢查連線狀態...")
                    else:
                        logger.info(f"非交易時間，{sleep_time}秒後檢查連線狀態...")
                
                time.sleep(sleep_time)
                
                # 只有在已登入的情況下才檢查連線
                if sinopac_connected and sinopac_login_status:
                    if not check_api_connection():
                        # 如果還沒開始重連，發送斷線通知
                        if not is_reconnecting:
                            logger.info("檢測到API斷線，開始重連...")
                            send_telegram_message("⚠️ API連線異常！！！\n正在嘗試重新連線．．．")
                            is_reconnecting = True
                            reconnect_attempts = 0
                        
                        # 嘗試重連
                        if reconnect_api():
                            logger.info("API重連成功！")
                            send_telegram_message("✅ API連線成功！！！")
                            reconnect_attempts = 0
                            is_reconnecting = False
                        else:
                            reconnect_attempts += 1
                            logger.error(f"重連失敗，將在下次監控週期繼續嘗試... (第{reconnect_attempts}次)")
                            
                            # 發送警告通知（但不停止重連嘗試）
                            if reconnect_attempts % 5 == 0:  # 每5次失敗發送一次通知
                                send_telegram_message(f"⚠️ API重連失敗！已嘗試{reconnect_attempts}次，系統將持續重試...")
                            
                            # 註意：不將 is_reconnecting 設為 False，繼續保持重連狀態
                
            except Exception as e:
                logger.error(f"連線監控器錯誤: {e}")
                time.sleep(60)  # 發生錯誤時等待1分鐘
    
    connection_thread = create_managed_thread(target=connection_monitor_loop, name="連線監控線程")
    connection_thread.start()
    print_console("SYSTEM", "SUCCESS", "智能連線監控器已啟動（交易時間每1分鐘，非交易時間每10分鐘）")

def reconnect_api():
    """重連API - 增強版重連機制"""
    global sinopac_connected, sinopac_login_status
    
    max_retries = 3  # 單次重連嘗試最多3次
    
    for attempt in range(1, max_retries + 1):
        try:
            logger.info(f"開始第{attempt}次重連API...")
            
            # 先登出（包含token清理）
            if sinopac_connected:
                try:
                    sinopac_api.logout()
                    logger.info("已執行API登出")
                except Exception as e:
                    logger.error(f"API登出時發生錯誤: {e}")
                sinopac_connected = False
                sinopac_login_status = False
            
            # 等待間隔時間（重試次數越多等待越久）
            wait_time = attempt * 2
            logger.info(f"等待{wait_time}秒後重新初始化...")
            time.sleep(wait_time)
            
            # 重新初始化API並登入
            try:
                # 重新初始化API
                if sinopac_api:
                    try:
                        sinopac_api.logout()
                    except:
                        pass
                
                logger.info("重新初始化API...")
                init_sinopac_api()
                
                # 等待1秒讓API完全初始化
                time.sleep(1)
                
                # 重新登入
                logger.info("嘗試重新登入...")
                if login_sinopac():
                    logger.info(f"API重連成功！(第{attempt}次嘗試)")
                    
                    # API重連日誌記錄已移除
                    
                    return True
                else:
                    logger.error(f"第{attempt}次重連失敗：登入失敗")
                    
            except Exception as e:
                logger.error(f"第{attempt}次重連失敗：初始化錯誤 - {e}")
                
        except Exception as e:
            logger.error(f"第{attempt}次重連時發生嚴重錯誤: {e}")
        
        # 如果不是最後一次嘗試，記錄失敗並繼續
        if attempt < max_retries:
            logger.error(f"第{attempt}次重連失敗，準備下一次嘗試...")
    
    # 所有嘗試都失敗
    logger.error(f"API重連失敗！已嘗試{max_retries}次")
    
    try:
        requests.post(
            f'http://127.0.0.1:{CURRENT_PORT}/api/system_log',
            json={'message': f'API重連失敗！已嘗試{max_retries}次', 'type': 'error'},
            timeout=5
        )
    except:
        pass
    
    return False

def stop_connection_monitor():
    """停止連線監控器"""
    global connection_monitor_timer
    
    if connection_monitor_timer and connection_monitor_timer.is_alive():
        connection_monitor_timer.cancel()
        connection_monitor_timer = None
        logger.info("已停止連線監控器")

def get_contract_point_value(contract_code):
    """獲取合約點值"""
    if 'TXF' in contract_code:
        return 200  # 大台每點200元
    elif 'MXF' in contract_code:
        return 50   # 小台每點50元
    elif 'TMF' in contract_code:
        return 10   # 微台每點10元
    else:
        return 200  # 預設值

def analyze_position_changes(trades):
    """分析持倉變化來改善開平倉判斷"""
    position_tracker = {}  # 追蹤每個合約的持倉狀態
    enhanced_trades = []
    
    for trade in trades:
        raw_data = trade.get('raw_data', {})
        order = raw_data.get('order', {})
        
        # 只處理有成交的訂單
        if trade.get('type') != 'deal':
            continue
            
        # 嘗試多種方式獲取合約代碼
        contract_code = ''
        contract_data = order.get('contract', {})
        
        if isinstance(contract_data, dict):
            contract_code = contract_data.get('code', '')
        elif isinstance(contract_data, str):
            contract_code = contract_data
        
        # 如果仍然沒有找到，嘗試從原始交易記錄中獲取
        if not contract_code:
            contract_code = trade.get('raw_data', {}).get('contract', {}).get('code', '')
        
        if not contract_code:
            continue
            
        action = order.get('action', '')
        quantity = order.get('quantity', 0)
        price = order.get('price', 0)
        declared_oc_type = order.get('oc_type', '')
        
        # 初始化合約追蹤器
        if contract_code not in position_tracker:
            position_tracker[contract_code] = {'net_position': 0, 'trades': []}
        
        # 計算實際的開平倉類型
        current_position = position_tracker[contract_code]['net_position']
        actual_oc_type = declared_oc_type
        
        # 如果沒有明確的 oc_type，根據持倉狀態推斷
        if not declared_oc_type or declared_oc_type not in ['New', 'Cover']:
            if action == 'Buy':
                # 買入：如果目前是空單持倉，則為平倉；否則為開倉
                actual_oc_type = 'Cover' if current_position < 0 else 'New'
            elif action == 'Sell':
                # 賣出：如果目前是多單持倉，則為平倉；否則為開倉
                actual_oc_type = 'Cover' if current_position > 0 else 'New'
        
        # 更新持倉
        if action == 'Buy':
            position_tracker[contract_code]['net_position'] += quantity
        elif action == 'Sell':
            position_tracker[contract_code]['net_position'] -= quantity
        
        # 記錄交易
        trade_info = {
            'original_trade': trade,
            'contract_code': contract_code,
            'action': action,
            'quantity': quantity,
            'price': price,
            'actual_oc_type': actual_oc_type,
            'declared_oc_type': declared_oc_type,
            'timestamp': trade.get('timestamp', ''),
            'order_id': order.get('id', '')
        }
        
        position_tracker[contract_code]['trades'].append(trade_info)
        enhanced_trades.append(trade_info)
    
    return enhanced_trades, position_tracker

def calculate_trade_pnl(open_trades, close_trade):
    """計算平倉損益（FIFO 先進先出原則）"""
    if not open_trades or close_trade['actual_oc_type'] != 'Cover':
        return 0, []
    
    contract_code = close_trade['contract_code']
    point_value = get_contract_point_value(contract_code)
    close_action = close_trade['action']
    close_price = close_trade['price']
    close_quantity = close_trade['quantity']
    
    total_pnl = 0
    used_opens = []
    remaining_quantity = close_quantity
    
    # 根據平倉方向確定需要配對的開倉類型
    required_open_action = 'Buy' if close_action == 'Sell' else 'Sell'
    
    # 按時間順序處理開倉（FIFO）
    for open_trade in open_trades:
        if (remaining_quantity <= 0 or 
            open_trade['actual_oc_type'] != 'New' or 
            open_trade['action'] != required_open_action):
            continue
        
        # 計算可配對的數量
        available_quantity = open_trade['quantity'] - open_trade.get('used_quantity', 0)
        if available_quantity <= 0:
            continue
        
        paired_quantity = min(remaining_quantity, available_quantity)
        
        # 計算損益
        open_price = open_trade['price']
        if close_action == 'Sell':  # 平多倉
            pnl = (close_price - open_price) * paired_quantity * point_value
        else:  # 平空倉
            pnl = (open_price - close_price) * paired_quantity * point_value
        
        total_pnl += pnl
        
        # 記錄使用的開倉
        used_opens.append({
            'open_trade': open_trade,
            'paired_quantity': paired_quantity,
            'pnl': pnl
        })
        
        # 更新剩餘數量
        remaining_quantity -= paired_quantity
        open_trade['used_quantity'] = open_trade.get('used_quantity', 0) + paired_quantity
    
    return total_pnl, used_opens

def load_historical_open_trades(close_trades):
    """從歷史記錄中載入開倉交易"""
    import os
    import glob
    
    historical_opens = []
    
    # 取得所有歷史交易記錄檔案
    trade_files = glob.glob(os.path.join(TX_DATA_DIR, 'TXtransdata_*.json'))
    
    # 按日期排序（從最新到最舊）
    trade_files.sort(reverse=True)
    
    logger.info(f"搜尋歷史開倉記錄，共找到 {len(trade_files)} 個交易記錄檔案")
    
    for close_trade in close_trades:
        logger.info(f"搜尋 {close_trade['contract_code']} 的歷史開倉記錄...")
        
        # 為每個平倉交易尋找對應的開倉記錄
        for trade_file in trade_files:
            try:
                with open(trade_file, 'r', encoding='utf-8') as f:
                    historical_trades = json.load(f)
                
                # 在歷史記錄中尋找開倉交易
                for trade in historical_trades:
                    if trade.get('type') != 'deal':
                        continue
                        
                    raw_data = trade.get('raw_data', {})
                    order = raw_data.get('order', {})
                    
                    # 檢查是否為開倉交易
                    if order.get('oc_type', '').upper() != 'NEW':
                        continue
                    
                    # 獲取合約資訊
                    contract_code = ''
                    contract_data = order.get('contract', {})
                    if isinstance(contract_data, dict):
                        contract_code = contract_data.get('code', '')
                    elif isinstance(contract_data, str):
                        contract_code = contract_data
                    
                    if not contract_code:
                        contract_code = trade.get('raw_data', {}).get('contract', {}).get('code', '')
                    
                    # 檢查是否為相同合約且方向相反
                    if (contract_code == close_trade['contract_code'] and 
                        order.get('action') != close_trade['action'] and
                        order.get('quantity') == close_trade['quantity']):  # 方向相反且數量相同
                        
                        historical_opens.append({
                            'contract_code': contract_code,
                            'contract_name': get_contract_name_from_code(contract_code),
                            'action': order.get('action', ''),
                            'quantity': order.get('quantity', 0),
                            'price': order.get('price', 0),
                            'timestamp': trade.get('timestamp', ''),
                            # 新增：真實開倉價格（優先使用，如果沒有則使用price）
                            'real_opening_price': trade.get('real_opening_price', order.get('price', 0))
                        })
                        
                        logger.info(f"  找到歷史開倉記錄: {contract_code} {order.get('action')} {order.get('quantity')} 口 @ {order.get('price')}")
                        break  # 找到一個就停止搜尋該合約
                
                # 如果找到了開倉記錄，停止搜尋其他檔案
                if any(h['contract_code'] == close_trade['contract_code'] for h in historical_opens):
                    break
                
            except Exception as e:
                logger.error(f"讀取歷史記錄失敗 {trade_file}: {e}")
                continue
    
    return historical_opens


def get_delivery_date_for_contract(contract_code):
    """獲取合約交割日期"""
    try:
        contract_prefix = contract_code[:3]  # TXF, MXF, TMF
        
        # 嘗試從全域合約對象獲取
        if contract_prefix == 'TXF' and 'contract_txf' in globals() and contract_txf:
            if hasattr(contract_txf, 'delivery_date'):
                return contract_txf.delivery_date.strftime('%Y/%m/%d') if hasattr(contract_txf.delivery_date, 'strftime') else str(contract_txf.delivery_date)
        elif contract_prefix == 'MXF' and 'contract_mxf' in globals() and contract_mxf:
            if hasattr(contract_mxf, 'delivery_date'):
                return contract_mxf.delivery_date.strftime('%Y/%m/%d') if hasattr(contract_mxf.delivery_date, 'strftime') else str(contract_mxf.delivery_date)
        elif contract_prefix == 'TMF' and 'contract_tmf' in globals() and contract_tmf:
            if hasattr(contract_tmf, 'delivery_date'):
                return contract_tmf.delivery_date.strftime('%Y/%m/%d') if hasattr(contract_tmf.delivery_date, 'strftime') else str(contract_tmf.delivery_date)
        
        # 從永豐API查詢
        if sinopac_connected and sinopac_api:
            try:
                contracts = sinopac_api.Contracts.Futures
                for contract in contracts:
                    if contract_prefix in contract.code:
                        return contract.delivery_date.strftime('%Y/%m/%d') if hasattr(contract.delivery_date, 'strftime') else str(contract.delivery_date)
            except:
                pass
                
        return ''
    except:
        return ''


def analyze_simple_trading_stats(trades=None, filter_date=None):
    """使用JSON配對系統分析交易統計
    
    Args:
        trades: 廢棄參數，保持兼容性
        filter_date: 篩選日期（格式：YYYYMMDD），只統計該日期的平倉明細，None則不篩選
    """
    try:
        from trade_pairing_TX import get_tx_cover_trades_for_report, get_trading_statistics
        
        # 獲取JSON配對系統的統計數據
        date_range = 1 if filter_date else 7  # 如果有日期篩選則只查當天，否則查7天
        stats = get_trading_statistics(date_range)
        
        # 獲取平倉交易明細
        detailed_covers = get_tx_cover_trades_for_report(date_range)
        
        # 如果有日期篩選，過濾指定日期的數據
        if filter_date:
            filtered_covers = []
            for cover in detailed_covers:
                timestamp = cover.get('cover_timestamp', '')
                if timestamp:
                    try:
                        trade_date = timestamp[:10].replace('-', '')  # 2025-07-26T... -> 20250726
                        if trade_date == filter_date:
                            filtered_covers.append(cover)
                    except:
                        pass
            detailed_covers = filtered_covers
        
        # 轉換為舊格式兼容
        cover_trades = []
        total_cover_quantity = 0
        
        for cover in detailed_covers:
            total_cover_quantity += cover['matched_quantity']
            
            # 格式化為報表需要的格式
            cover_trades.append({
                'contract_name': get_contract_name_from_code(cover['contract_code']),
                'action': get_action_display_by_rule('COVER', cover['cover_action']),
                'quantity': f"{cover['matched_quantity']} 口",
                'open_price': f"{int(cover['open_price']):,}",
                'cover_price': f"{int(cover['cover_price']):,}",
                'pnl': int(cover['pnl']),
                'timestamp': cover['cover_timestamp'],
                'order_id': cover['cover_order_id'],
                'price_type': 'MKT',  # JSON系統記錄的都是成交價
                'order_type': 'IOC',
                'delivery_date': get_delivery_date_for_contract(cover['contract_code']),
                'is_manual': True  # 手動交易
            })
        
        logger.info(f"JSON配對系統統計: 平倉{len(detailed_covers)}筆, 總數量{total_cover_quantity}口")
        logger.info(f"各合約損益: TXF={stats['contract_pnl']['TXF']}, MXF={stats['contract_pnl']['MXF']}, TMF={stats['contract_pnl']['TMF']}")
        
        return cover_trades, total_cover_quantity, stats['contract_pnl']
        
    except Exception as e:
        logger.error(f"JSON配對系統分析失敗: {e}")
        # 回退到空結果
        return [], 0, {'TXF': 0, 'MXF': 0, 'TMF': 0}

def analyze_daily_trades_with_pnl(trades):
    """分析每日交易並計算損益"""
    logger.info(f"  開始分析 {len(trades)} 筆交易記錄...")
    enhanced_trades, position_tracker = analyze_position_changes(trades)
    logger.info(f"  識別出 {len(enhanced_trades)} 筆有效成交記錄")
    
    # 為每個合約建立開倉列表
    open_positions = {}
    cover_trades_with_pnl = []
    total_cover_quantity = 0
    
    # 統計開平倉數量
    open_count = sum(1 for t in enhanced_trades if t['actual_oc_type'] == 'New')
    close_count = sum(1 for t in enhanced_trades if t['actual_oc_type'] == 'Cover')
    logger.info(f"  開倉交易：{open_count} 筆，平倉交易：{close_count} 筆")
    
    for trade_info in enhanced_trades:
        contract_code = trade_info['contract_code']
        
        if contract_code not in open_positions:
            open_positions[contract_code] = []
        
        if trade_info['actual_oc_type'] == 'New':
            # 開倉交易：加入開倉列表
            open_positions[contract_code].append(trade_info)
        
        elif trade_info['actual_oc_type'] == 'Cover':
            # 平倉交易：計算損益
            pnl, used_opens = calculate_trade_pnl(
                open_positions[contract_code], 
                trade_info
            )
            
            total_cover_quantity += trade_info['quantity']
            
            # 構建平倉交易明細
            contract_name = get_contract_name_from_code(contract_code)
            # 使用正確的動作顯示邏輯
            action_display = get_action_display_by_rule('COVER', trade_info['action'])
            
            cover_trades_with_pnl.append({
                'contract_name': contract_name,
                'action': action_display,
                'quantity': f"{trade_info['quantity']} 口",
                'order_price': f"{int(trade_info['price']):,}",
                'cover_price': f"{int(trade_info['price']):,}",
                'pnl': int(pnl),
                'used_opens': used_opens,
                                 'timestamp': trade_info['timestamp']
                          })
     
     # 最終統計
    total_pnl = sum(trade.get('pnl', 0) for trade in cover_trades_with_pnl)
    logger.info(f"  損益計算完成：{len(cover_trades_with_pnl)} 筆平倉，總損益 {total_pnl:,} TWD")
     
    return cover_trades_with_pnl, total_cover_quantity

def start_tx_service():
    """啟動TX交易服務"""
    global notification_sent_date
    
    # 在其他初始化代碼之前添加
    notification_sent_date = None
    
    # 顯示啟動設定
    print_console("SYSTEM", "START", "=== TX 期貨交易系統啟動 ===")
    print_console("SYSTEM", "INFO", f"端口設定: {CURRENT_PORT}")
    print_console("SYSTEM", "INFO", "================================")
    
    # 啟動時清理舊的交易記錄檔案
    cleanup_old_trade_files()
    
    # 啟動時清理舊的BTC交易記錄檔案
    if BTC_MODULE_AVAILABLE:
        btcmain.cleanup_old_btc_trade_files()
    
    
    # 加載訂單映射
    load_order_mapping()
    
    # 初始化保證金記錄
    if update_margin_requirements_from_api():
        last_margin_requirements = margin_requirements.copy()
    
    # 程式啟動時強制重置LOGIN為0，確保乾淨狀態
    try:
        update_login_status(0)
        # 驗證重置是否成功（靜默檢查）
        if os.path.exists(ENV_PATH):
            with open(ENV_PATH, 'r', encoding='utf-8') as f:
                content = f.read()
                if 'LOGIN=0' not in content:
                    print_console("SYSTEM", "ERROR", "LOGIN狀態重置失敗")
    except Exception as e:
        print_console("SYSTEM", "ERROR", "重置LOGIN狀態時發生錯誤", str(e))
    
    # 初始化永豐API
    init_sinopac_api()
    
    # 啟動通知檢查器
    start_notification_checker()
    
    # 啟動轉倉檢查器
    start_rollover_checker()
    
    # 延遲執行轉倉檢查，等待合約對象初始化
    # 在登入後會自動執行轉倉檢查
    
    # 啟動連線監控器
    start_connection_monitor()
    
    # 啟動TX風險監控
    start_tx_risk_monitor()
    
    # 信號處理器只能在主線程中設置，在 start_tx_service 中不設置
    
    print_console("SYSTEM", "SUCCESS", "TX交易服務初始化完成")

if __name__ == '__main__':
    # ========== 自動更新檢查 ==========
    # 在程式啟動前先檢查程式更新 (支援Telegram通知、配置保護、優雅重啟)
    try:
        from updater import check_and_update
        
        print("=" * 60)
        print("Auto91 自動更新檢查")
        print("=" * 60)
        
        # 執行自動更新檢查（自動確認模式）
        update_success = check_and_update(auto_confirm=True, silent_mode=True)
        
        if update_success:
            print("版本檢查完成")
        else:
            print("更新檢查過程中發生問題，但將繼續啟動程式")
        
        print("=" * 60)
        
    except ImportError as e:
        print(f"警告: 無法導入更新器，跳過更新檢查")
    except Exception as e:
        print(f"警告: 更新檢查過程中發生錯誤，跳過更新檢查")
    
    try:
        # 設置信號處理器 (只能在主線程中設置)
        signal.signal(signal.SIGINT, signal_handler)
        signal.signal(signal.SIGTERM, signal_handler)
        
        # 確保環境配置文件存在
        ensure_tx_env_exists()
        ensure_btc_env_exists()
        
        # 初始化隧道管理器
        init_tunnel_service()
        
        # 啟動完整服務
        start_tx_service()
        
        # 啟動Flask伺服器和webview
        flask_server_thread = create_managed_thread(target=start_flask, name="Flask服務器線程")
        flask_server_thread.start()
        
        # 將Flask線程註冊到活動線程列表
        register_thread(flask_server_thread, "Flask服務器線程")
        
        time.sleep(2)  # 等待伺服器啟動
        start_webview()
    except KeyboardInterrupt:
        cleanup_on_exit()

# TX系統未平倉部位追蹤機制
def save_tx_open_position(trade_data):
    """保存TX開倉部位到追蹤檔案"""
    try:
        today = datetime.now().strftime('%Y%m%d')
        open_file = os.path.join(TX_RECORDS_DIR, f'tx_open_positions_{today}.json')
        os.makedirs(TX_RECORDS_DIR, exist_ok=True)
        
        # 讀取現有開倉記錄
        if os.path.exists(open_file):
            with open(open_file, 'r', encoding='utf-8') as f:
                open_positions = json.load(f)
        else:
            open_positions = []
        
        # 創建開倉記錄
        position_record = {
            'trade_id': f"TX_{trade_data.get('raw_data', {}).get('order', {}).get('oc_type', 'Open')}_{trade_data.get('raw_data', {}).get('order', {}).get('action', 'Buy')}_{today}_{trade_data.get('deal_order_id', '')}",
            'timestamp': trade_data.get('timestamp', datetime.now().isoformat()),
            'symbol': trade_data.get('raw_data', {}).get('contract', {}).get('code', ''),
            'action': trade_data.get('raw_data', {}).get('order', {}).get('action', ''),
            'oc_type': trade_data.get('raw_data', {}).get('order', {}).get('oc_type', ''),
            'quantity': trade_data.get('raw_data', {}).get('order', {}).get('quantity', 0),
            'price': trade_data.get('raw_data', {}).get('order', {}).get('price', 0),
            'order_id': trade_data.get('deal_order_id', ''),
            'pair_key': f"TX_{trade_data.get('raw_data', {}).get('order', {}).get('action', 'Buy')}",
            'remaining_quantity': trade_data.get('raw_data', {}).get('order', {}).get('quantity', 0),
            'status': 'open',
            'matched_covers': []
        }
        
        open_positions.append(position_record)
        
        with open(open_file, 'w', encoding='utf-8') as f:
            json.dump(open_positions, f, indent=2, ensure_ascii=False)
            
        logger.info(f"TX開倉部位已保存: {position_record['trade_id']}")
        
    except Exception as e:
        logger.error(f"保存TX開倉部位失敗: {str(e)}")

def get_tx_open_positions_today():
    """獲取TX當前未平倉部位"""
    try:
        open_positions = []
        
        if api:
            # 獲取真實持倉信息
            positions = api.get_position()
            
            for pos in positions:
                if hasattr(pos, 'quantity') and pos.quantity != 0:
                    # 有持倉
                    position_info = {
                        'contract_code': pos.code,
                        'quantity': pos.quantity,
                        'price': getattr(pos, 'price', 0),
                        'pnl': getattr(pos, 'pnl', 0),
                        'direction': 'Buy' if pos.quantity > 0 else 'Sell'
                    }
                    open_positions.append(position_info)
        
        return open_positions
        
    except Exception as e:
        logger.error(f"獲取TX持倉狀態失敗: {str(e)}")
        return []

def update_tx_position_on_close(close_trade_data):
    """平倉時更新TX部位追蹤"""
    try:
        today = datetime.now().strftime('%Y%m%d')
        open_file = os.path.join(TX_RECORDS_DIR, f'tx_open_positions_{today}.json')
        
        if not os.path.exists(open_file):
            return
            
        with open(open_file, 'r', encoding='utf-8') as f:
            open_positions = json.load(f)
        
        # 查找對應的開倉部位進行配對
        close_quantity = close_trade_data.get('raw_data', {}).get('order', {}).get('quantity', 0)
        close_action = close_trade_data.get('raw_data', {}).get('order', {}).get('action', '')
        
        # TX系統的對應關係：平多單時action是Sell，平空單時action是Buy
        target_action = 'Buy' if close_action == 'Sell' else 'Sell'
        
        updated = False
        for position in open_positions:
            if (position['status'] == 'open' and 
                position['action'] == target_action and 
                position['remaining_quantity'] > 0):
                
                # 執行配對
                matched_quantity = min(close_quantity, position['remaining_quantity'])
                position['remaining_quantity'] -= matched_quantity
                
                # 添加配對記錄
                position['matched_covers'].append({
                    'close_trade_id': close_trade_data.get('deal_order_id', ''),
                    'close_timestamp': close_trade_data.get('timestamp', ''),
                    'matched_quantity': matched_quantity,
                    'close_price': close_trade_data.get('raw_data', {}).get('order', {}).get('price', 0)
                })
                
                # 如果完全平倉，標記為已關閉
                if position['remaining_quantity'] <= 0:
                    position['status'] = 'closed'
                
                close_quantity -= matched_quantity
                updated = True
                
                if close_quantity <= 0:
                    break
        
        if updated:
            with open(open_file, 'w', encoding='utf-8') as f:
                json.dump(open_positions, f, indent=2, ensure_ascii=False)
            logger.info(f"TX部位配對更新完成")
        
    except Exception as e:
        logger.error(f"更新TX部位配對失敗: {str(e)}")
        return False
    
    return True

# 🔍 企業級系統診斷API端點
@app.route('/api/diagnose_trading_system', methods=['GET'])
def api_diagnose_trading_system():
    """🔍 交易系統診斷API - 專業級系統健康檢查"""
    try:
        logger.info("🔍 收到系統診斷請求")
        
        # 執行診斷
        diagnosis_result = diagnose_trading_records()
        
        # 添加系統運行狀態
        diagnosis_result['system_info'] = {
            'tx_data_dir': TX_DATA_DIR,
            'directory_exists': os.path.exists(TX_DATA_DIR),
            'current_time': datetime.now().isoformat(),
            'today_str': datetime.now().strftime('%Y%m%d'),
            'login_status': get_env_var('LOGIN', '0'),
            'system_version': CURRENT_VERSION,
            'build_info': CURRENT_BUILD,
            'system_description': SYSTEM_DESCRIPTION
        }
        
        # 檢查最近的交易活動
        recent_files = []
        if os.path.exists(TX_DATA_DIR):
            for i in range(7):  # 檢查最近7天
                check_date = datetime.now() - timedelta(days=i)
                date_str = check_date.strftime('%Y%m%d')
                file_path = os.path.join(TX_DATA_DIR, f'TXtransdata_{date_str}.json')
                if os.path.exists(file_path):
                    try:
                        with open(file_path, 'r', encoding='utf-8') as f:
                            records = json.load(f)
                        recent_files.append({
                            'date': date_str,
                            'record_count': len(records),
                            'file_size': os.path.getsize(file_path)
                        })
                    except:
                        pass
        
        diagnosis_result['recent_activity'] = recent_files
        
        return jsonify({
            'success': True,
            'diagnosis': diagnosis_result,
            'message': '系統診斷完成'
        })
        
    except Exception as e:
        logger.error(f"❌ 系統診斷API異常: {e}")
        return jsonify({
            'success': False,
            'message': f'診斷異常: {str(e)}'
        })