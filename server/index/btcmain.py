#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BTC交易系統主程式
處理BTC加密貨幣交易相關功能
"""

from flask import request, jsonify
import os
import json
import time
import hmac
import hashlib
import requests
import threading
import glob
import logging
from datetime import datetime, timedelta, timezone
from urllib.parse import urlencode
from collections import OrderedDict
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
from trade_pairing_BTC import record_btc_opening_trade, record_btc_covering_trade, save_btc_transdata
from trading_config import TradingConfig

# 配置目錄
CONFIG_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config')
BTC_ENV_PATH = os.path.join(CONFIG_DIR, 'btc.env')

# 幣安API配置
BINANCE_BASE_URL = "https://api.binance.com"
BINANCE_FAPI_URL = "https://fapi.binance.com"  # 期貨API

# 配置日誌
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger('btc_module')

# 全局變量
binance_client = None
account_info = None
btc_shutdown_flag = threading.Event()  # BTC模組停止標誌
btc_active_threads = []  # BTC模組活動線程列表
btc_active_trades = {}  # 活躍交易記錄
processed_orders = set()  # 記錄已處理的訂單，避免重複處理
btc_processed_signals = set()  # 記錄已處理的BTC webhook訊號，避免重複處理

# 線程安全鎖
processed_orders_lock = threading.Lock()  # 保護 processed_orders 的線程鎖
btc_signals_lock = threading.Lock()  # 保護 btc_processed_signals 的線程鎖
btc_order_lock = threading.Lock()  # 保護下單操作的線程鎖
btc_active_order_requests = set()  # 記錄正在處理的下單請求，防止重複下單

# LOGIN狀態管理函數
def update_btc_login_status(status):
    """更新BTC登入狀態
    Args:
        status (int): 登入狀態，1=已登入，0=未登入
    """
    try:
        if not os.path.exists(BTC_ENV_PATH):
            logger.warning(f"BTC配置文件不存在: {BTC_ENV_PATH}")
            return False
        
        # 讀取現有配置
        with open(BTC_ENV_PATH, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        # 更新LOGIN_BTC狀態
        updated = False
        for i, line in enumerate(lines):
            if line.strip().startswith('LOGIN_BTC='):
                lines[i] = f"LOGIN_BTC={status}\n"
                updated = True
                break
        
        # 如果沒找到LOGIN_BTC行，則添加
        if not updated:
            lines.append(f"LOGIN_BTC={status}\n")
        
        # 寫回文件
        with open(BTC_ENV_PATH, 'w', encoding='utf-8') as f:
            f.writelines(lines)
        
        logger.info(f"✅ BTC登入狀態已更新: LOGIN_BTC={status}")
        return True
        
    except Exception as e:
        logger.error(f"❌ 更新BTC登入狀態失敗: {e}")
        return False

def update_tx_login_status(status):
    """更新TX登入狀態
    Args:
        status (int): 登入狀態，1=已登入，0=未登入
    """
    try:
        TX_ENV_PATH = os.path.join(CONFIG_DIR, 'tx.env')
        if not os.path.exists(TX_ENV_PATH):
            logger.warning(f"TX配置文件不存在: {TX_ENV_PATH}")
            return False
        
        # 讀取現有配置
        with open(TX_ENV_PATH, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        # 更新LOGIN狀態
        updated = False
        for i, line in enumerate(lines):
            if line.strip().startswith('LOGIN='):
                lines[i] = f"LOGIN={status}\n"
                updated = True
                break
        
        # 如果沒找到LOGIN行，則添加
        if not updated:
            lines.append(f"LOGIN={status}\n")
        
        # 寫回文件
        with open(TX_ENV_PATH, 'w', encoding='utf-8') as f:
            f.writelines(lines)
        
        logger.info(f"✅ TX登入狀態已更新: LOGIN={status}")
        return True
        
    except Exception as e:
        logger.error(f"❌ 更新TX登入狀態失敗: {e}")
        return False

def initialize_btc_directories():
    """初始化BTC系統所需的配置（不預先創建交易數據目錄）"""
    try:
        # BTC交易數據目錄（traderecordsBTC, tradedataBTC）只在有數據時才創建
        # 這樣與BTC交易報表目錄的邏輯保持一致
        logger.info("🚀 BTC系統初始化完成（交易數據目錄將在有數據時自動創建）")
        
    except Exception as e:
        logger.error(f"❌ BTC系統初始化失敗: {e}")

def auto_start_btc_system():
    """自動啟動BTC系統（如果配置為已登入狀態）"""
    global binance_client
    
    try:
        # 載入配置檢查登入狀態
        env_data = load_btc_env_data()
        if env_data.get('LOGIN_BTC') != '1':
            logger.info("BTC系統未設為登入狀態，跳過自動啟動")
            return
        
        # 如果客戶端已初始化，跳過
        if binance_client:
            logger.info("BTC客戶端已初始化，跳過自動啟動")
            return
        
        logger.info("🚀 檢測到BTC系統已登入，開始自動啟動...")
        
        # 初始化客戶端
        api_key = env_data.get('BINANCE_API_KEY')
        secret_key = env_data.get('BINANCE_SECRET_KEY')
        
        if not api_key or not secret_key:
            logger.error("❌ BTC API金鑰缺失，無法自動啟動")
            return
        
        binance_client = BinanceClient(api_key, secret_key)
        logger.info("✅ BTC客戶端自動初始化完成")
        
        # 啟動WebSocket
        start_btc_websocket()
        logger.info("✅ BTC WebSocket自動啟動完成")
        
        # 啟動訂單監控
        success = start_btc_order_monitoring()
        if success:
            logger.info("✅ BTC訂單監控自動啟動完成")
        else:
            logger.error("❌ BTC訂單監控自動啟動失敗")
        
        logger.info("🎉 BTC系統自動啟動完成！現在可以接收委託和成交通知")
        
        # 啟動定時調度器
        start_btc_daily_scheduler()
        
    except Exception as e:
        logger.error(f"❌ BTC系統自動啟動失敗: {e}")

# 在模組載入時自動初始化
initialize_btc_directories()

# 延遲自動啟動BTC系統（確保所有依賴都已載入）
import threading
def delayed_auto_start():
    import time
    time.sleep(2)  # 等待2秒確保模組完全載入
    auto_start_btc_system()

threading.Thread(target=delayed_auto_start, daemon=True).start()

# 訂單監控相關
pending_orders = {}  # 存儲待成交訂單 {order_id: order_info}
btc_order_history = {}  # 存儲訂單歷史狀態，用於檢測修改 {order_id: {price, quantity, status}}
order_monitor_ws = None
order_monitor_thread = None
user_data_stream_key = None

# BTC斷線重連相關變數
btc_connection_check_interval = 60  # 每1分鐘檢查一次連線狀態
btc_max_reconnect_attempts = 999  # 無限重連嘗試
btc_reconnect_attempts = 0  # 當前重連嘗試次數
btc_last_connection_check = None  # 上次連線檢查時間
btc_is_reconnecting = False  # 是否正在重連中
btc_auto_logout_timer = None  # 自動登出計時器

# BTC定時調度相關變量
btc_daily_scheduler = None
btc_scheduler_thread = None

# BTC通知排隊系統
btc_notification_queue = []  # 通知排隊列表
btc_notification_queue_lock = threading.Lock()  # 排隊鎖
btc_notification_worker_running = False  # 排隊工作器狀態

def stop_btc_module():
    """停止BTC模組的所有線程和連接"""
    global btc_shutdown_flag, btc_active_threads, btc_auto_logout_timer, btc_notification_manager
    logger.info("🔴 正在停止BTC模組...")
    
    # 設置停止標誌
    btc_shutdown_flag.set()
    
    # 停止定時調度器
    stop_btc_daily_scheduler()
    
    # 停止通知排隊系統
    try:
        if 'btc_notification_manager' in globals():
            btc_notification_manager.shutdown()
            logger.info("BTC通知排隊系統已停止")
    except Exception as e:
        logger.error(f"停止BTC通知排隊系統失敗: {e}")
    
    # 停止自動登出Timer
    if btc_auto_logout_timer and btc_auto_logout_timer.is_alive():
        btc_auto_logout_timer.cancel()
        logger.info("已取消BTC自動登出Timer")
    
    # 停止Listen Key心跳維持
    stop_btc_listen_key_heartbeat()
    
    # 停止WebSocket連接
    stop_btc_websocket()
    
    # 將所有線程設為daemon，讓主程序可以退出
    active_count = 0
    for thread in btc_active_threads:
        if thread.is_alive():
            thread.daemon = True
            active_count += 1
            logger.info(f"已將BTC線程設為daemon: {thread.name}")
    
    if active_count > 0:
        logger.warning(f"有 {active_count} 個BTC線程仍在運行，已設為daemon模式")
    
    btc_active_threads.clear()
    logger.info("✅ BTC模組已停止")

def register_btc_thread(thread, name="未知BTC線程"):
    """註冊BTC線程"""
    global btc_active_threads
    btc_active_threads.append(thread)
    logger.info(f"🧵 已註冊BTC線程: {name}")

# ============ 統一工具函數 ============

class APIRequestHandler:
    """統一的API請求處理類"""
    
    @staticmethod
    def make_request(method, url, data=None, headers=None, timeout=10, **kwargs):
        """統一的API請求處理函數"""
        try:
            response = getattr(requests, method.lower())(url, json=data, headers=headers, timeout=timeout, **kwargs)
            return response
        except requests.RequestException as e:
            logger.error(f"API請求失敗 [{method}] {url}: {e}")
            return None
    
    @staticmethod
    def handle_telegram_request(bot_token, chat_ids, message, message_type="info"):
        """統一的Telegram消息發送"""
        if not bot_token or not chat_ids:
            logger.warning("Telegram配置不完整，跳過發送")
            return False
            
        success_count = 0
        for chat_id in chat_ids:
            try:
                url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
                payload = {
                    'chat_id': chat_id,
                    'text': message,
                    'parse_mode': 'HTML'
                }
                response = APIRequestHandler.make_request('post', url, data=payload)
                if response and response.status_code == 200:
                    success_count += 1
                    logger.info(f"Telegram消息發送成功到 {chat_id}")
                else:
                    logger.error(f"Telegram消息發送失敗到 {chat_id}")
            except Exception as e:
                logger.error(f"Telegram發送異常到 {chat_id}: {e}")
        
        return success_count > 0

def load_config_file(file_path, required_fields=None):
    """統一的配置文件讀取函數"""
    config = {}
    if not os.path.exists(file_path):
        logger.error(f"配置文件不存在: {file_path}")
        return config
        
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and '=' in line and not line.startswith('#'):
                    key, value = line.split('=', 1)
                    config[key.strip()] = value.strip()
        
        if required_fields:
            missing_fields = [field for field in required_fields if field not in config]
            if missing_fields:
                logger.error(f"配置文件缺少必需字段: {missing_fields}")
        
        logger.info(f"成功載入配置文件: {file_path}")
        return config
        
    except Exception as e:
        logger.error(f"讀取配置文件失敗 {file_path}: {e}")
        return config

def handle_api_error(error, context="API操作", raise_error=False):
    """統一的錯誤處理函數"""
    error_msg = f"{context}失敗: {error}"
    logger.error(error_msg)
    
    if raise_error:
        raise Exception(error_msg)
    
    return error_msg

class BinanceClient:
    """幣安API客戶端"""
    
    def __init__(self, api_key, secret_key, testnet=False):
        self.api_key = api_key
        self.secret_key = secret_key
        self.base_url = BINANCE_FAPI_URL if not testnet else "https://testnet.binancefuture.com"
        self.session = requests.Session()
        self.session.headers.update({
            'X-MBX-APIKEY': api_key,
            'Content-Type': 'application/json'
        })
    
    def _generate_signature(self, query_string):
        """生成API簽名"""
        return hmac.new(
            self.secret_key.encode('utf-8'),
            query_string.encode('utf-8'),
            hashlib.sha256
        ).hexdigest()
    
    def _get_server_time(self):
        """獲取幣安服務器時間"""
        try:
            url = f"{self.base_url}/fapi/v1/time"
            response = self.session.get(url)
            if response.status_code == 200:
                return response.json().get('serverTime')
        except:
            pass
        return None
    
    def _make_request(self, method, endpoint, params=None, signed=True):
        """發送API請求"""
        url = f"{self.base_url}{endpoint}"
        
        if params is None:
            params = {}
        
        if signed:
            # 使用幣安服務器時間來避免時間同步問題
            server_time = self._get_server_time()
            if server_time:
                params['timestamp'] = server_time
            else:
                # 如果無法獲取服務器時間，使用本地時間並減去1秒作為安全邊際
                params['timestamp'] = int(time.time() * 1000) - 1000
            
            query_string = urlencode(params)
            params['signature'] = self._generate_signature(query_string)
        
        try:
            if method == 'GET':
                response = self.session.get(url, params=params)
            elif method == 'POST':
                response = self.session.post(url, params=params)
            else:
                response = self.session.request(method, url, params=params)
            
            response.raise_for_status()
            return response.json()
        
        except requests.exceptions.RequestException as e:
            logger.error(f"幣安API請求失敗: {e}")
            if hasattr(e, 'response') and e.response is not None:
                logger.error(f"HTTP狀態碼: {e.response.status_code}")
                logger.error(f"錯誤詳情: {e.response.text}")
                # 嘗試解析JSON錯誤
                try:
                    error_json = e.response.json()
                    logger.error(f"API錯誤: {error_json}")
                except:
                    pass
            return None
    
    def test_connection(self):
        """測試API連接"""
        try:
            # 首先測試基本連接（無需API Key）
            url = f"{self.base_url}/fapi/v1/ping"
            logger.info(f"測試連接到: {url}")
            
            response = self.session.get(url, timeout=10)
            logger.info(f"Ping響應狀態碼: {response.status_code}")
            
            if response.status_code == 200:
                logger.info("基本連接測試成功")
                
                # 再測試服務器時間（無需API Key）
                time_url = f"{self.base_url}/fapi/v1/time"
                time_response = self.session.get(time_url, timeout=10)
                
                if time_response.status_code == 200:
                    server_time_data = time_response.json()
                    logger.info(f"服務器時間: {server_time_data}")
                    return True
                else:
                    logger.error(f"獲取服務器時間失敗: {time_response.status_code}")
                    return False
            elif response.status_code == 418:
                logger.error("收到HTTP 418狀態碼 - 可能是IP被限制或請求過於頻繁")
                logger.error("建議解決方案:")
                logger.error("1. 等待5-10分鐘後重試")
                logger.error("2. 檢查是否使用VPN或代理")
                logger.error("3. 確認IP地址是否在幣安白名單中")
                logger.error("4. 避免過於頻繁的API請求")
                return False
            else:
                logger.error(f"基本連接失敗，狀態碼: {response.status_code}")
                return False
                
        except requests.exceptions.Timeout:
            logger.error("連接超時，請檢查網絡連接")
            return False
        except requests.exceptions.ConnectionError as e:
            logger.error(f"網絡連接錯誤: {e}")
            return False
        except Exception as e:
            logger.error(f"幣安連接測試異常: {e}")
            return False
    
    def get_account_info(self):
        """獲取帳戶信息"""
        return self._make_request('GET', '/fapi/v2/account')
    
    def get_balance(self):
        """獲取餘額信息"""
        return self._make_request('GET', '/fapi/v2/balance')
    
    def get_position_info(self):
        """獲取持倉信息"""
        return self._make_request('GET', '/fapi/v2/positionRisk')
    
    def get_account_trades(self, symbol='BTCUSDT', startTime=None, endTime=None, limit=500):
        """獲取帳戶交易記錄"""
        params = {
            'symbol': symbol,
            'limit': limit
        }
        
        if startTime:
            params['startTime'] = startTime
        if endTime:
            params['endTime'] = endTime
            
        return self._make_request('GET', '/fapi/v1/userTrades', params)
    
    def get_income_history(self, startTime=None, endTime=None, incomeType=None, limit=1000, **kwargs):
        """獲取收入歷史 - 支援新舊參數名稱"""
        params = {
            'limit': limit
        }
        
        # 支援舊參數名稱的兼容性
        if 'start_time' in kwargs:
            startTime = kwargs['start_time']
        if 'end_time' in kwargs:
            endTime = kwargs['end_time']
        if 'income_type' in kwargs:
            incomeType = kwargs['income_type']
            
        if startTime:
            params['startTime'] = startTime
        if endTime:
            params['endTime'] = endTime
        if incomeType:
            params['incomeType'] = incomeType
            
        result = self._make_request('GET', '/fapi/v1/income', params)
        return result
    
    
    def get_server_time(self):
        """獲取服務器時間"""
        return self._make_request('GET', '/fapi/v1/time', signed=False)
    
    def get_exchange_info(self):
        """獲取交易所信息"""
        return self._make_request('GET', '/fapi/v1/exchangeInfo', signed=False)
    
    def place_order(self, symbol, side, order_type, quantity, price=None, time_in_force='GTC', 
                   reduce_only=False, close_position=False, position_side='BOTH'):
        """下單 - 期貨訂單"""
        params = {
            'symbol': symbol,
            'side': side,  # BUY 或 SELL
            'type': order_type,  # MARKET, LIMIT, STOP, TAKE_PROFIT等
            'quantity': str(quantity),
            'positionSide': position_side,  # BOTH, LONG, SHORT
            'reduceOnly': reduce_only,
            'closePosition': close_position
        }
        
        # 只有限價單等需要 timeInForce，市價單不需要
        if order_type in ['LIMIT', 'STOP', 'TAKE_PROFIT', 'STOP_MARKET', 'TAKE_PROFIT_MARKET']:
            params['timeInForce'] = time_in_force
        
        if order_type == 'LIMIT' and price:
            params['price'] = str(price)
        
        return self._make_request('POST', '/fapi/v1/order', params)
    
    def cancel_order(self, symbol, order_id=None, client_order_id=None):
        """取消訂單"""
        params = {'symbol': symbol}
        
        if order_id:
            params['orderId'] = order_id
        elif client_order_id:
            params['origClientOrderId'] = client_order_id
        else:
            raise ValueError("必須提供 order_id 或 client_order_id")
        
        return self._make_request('DELETE', '/fapi/v1/order', params)
    
    def get_order_status(self, symbol, order_id=None, client_order_id=None):
        """查詢訂單狀態"""
        params = {'symbol': symbol}
        
        if order_id:
            params['orderId'] = order_id
        elif client_order_id:
            params['origClientOrderId'] = client_order_id
        else:
            raise ValueError("必須提供 order_id 或 client_order_id")
        
        return self._make_request('GET', '/fapi/v1/order', params)
    
    def get_all_orders(self, symbol, limit=500):
        """獲取所有訂單"""
        params = {'symbol': symbol, 'limit': limit}
        return self._make_request('GET', '/fapi/v1/allOrders', params)
    
    def get_trades(self, symbol, limit=500):
        """獲取交易歷史"""
        params = {'symbol': symbol, 'limit': limit}
        return self._make_request('GET', '/fapi/v1/userTrades', params)
    
    def get_income_history(self, symbol=None, income_type=None, limit=100):
        """獲取收益歷史"""
        params = {'limit': limit}
        
        if symbol:
            params['symbol'] = symbol
        if income_type:
            params['incomeType'] = income_type  # TRANSFER, WELCOME_BONUS, REALIZED_PNL, etc.
        
        return self._make_request('GET', '/fapi/v1/income', params)
    
    def change_leverage(self, symbol, leverage):
        """調整槓桿倍數"""
        params = {
            'symbol': symbol,
            'leverage': leverage
        }
        return self._make_request('POST', '/fapi/v1/leverage', params)
    
    def change_margin_type(self, symbol, margin_type):
        """調整保證金模式"""
        params = {
            'symbol': symbol,
            'marginType': margin_type  # ISOLATED 或 CROSSED
        }
        return self._make_request('POST', '/fapi/v1/marginType', params)
    
    def get_open_orders(self, symbol=None):
        """獲取當前掛單"""
        params = {}
        if symbol:
            params['symbol'] = symbol
        return self._make_request('GET', '/fapi/v1/openOrders', params)
    
    def cancel_all_orders(self, symbol):
        """取消所有掛單"""
        params = {'symbol': symbol}
        return self._make_request('DELETE', '/fapi/v1/allOpenOrders', params)
    
    def get_position_risk(self, symbol=None):
        """獲取持倉風險"""
        params = {}
        if symbol:
            params['symbol'] = symbol
        return self._make_request('GET', '/fapi/v2/positionRisk', params)
    
    def get_ticker_price(self, symbol=None):
        """獲取最新價格"""
        params = {}
        if symbol:
            params['symbol'] = symbol
        return self._make_request('GET', '/fapi/v1/ticker/price', params, signed=False)
    
    def get_24hr_ticker(self, symbol=None):
        """獲取24小時價格變動"""
        params = {}
        if symbol:
            params['symbol'] = symbol
        return self._make_request('GET', '/fapi/v1/ticker/24hr', params, signed=False)

def save_btc_env():
    """保存BTC環境變量"""
    try:
        data = request.get_json()
        
        # 檢查是否有空值欄位
        required_fields = ['CHAT_ID_BTC', 'BINANCE_API_KEY', 'BINANCE_SECRET_KEY', 'BINANCE_USER_ID', 'TRADING_PAIR', 'LEVERAGE', 'POSITION_SIZE', 'MARGIN_TYPE', 'CONTRACT_TYPE']
        has_empty_fields = False
        
        for field in required_fields:
            if not data.get(field, '').strip():
                has_empty_fields = True
                break
        
        # 讀取當前登入狀態（如果檔案存在的話）
        current_login_status = '0'
        if os.path.exists(BTC_ENV_PATH):
            try:
                with open(BTC_ENV_PATH, 'r', encoding='utf-8') as f:
                    for line in f:
                        line = line.strip()
                        if line.startswith('LOGIN_BTC='):
                            current_login_status = line.split('=', 1)[1]
                            break
            except Exception:
                current_login_status = '0'
        
        # 如果有空欄位，強制登出狀態設為0，否則保持當前狀態
        final_login_status = '0' if has_empty_fields else current_login_status
        
        # 創建BTC環境文件內容（允許空值）
        btc_env_content = f"""# Telegram Bot
BOT_TOKEN_BTC=7912873826:AAFPPDwuwspKVdyDGwh1oVqxH6u1gQ_N-jU

# Telegram ID
CHAT_ID_BTC={data.get('CHAT_ID_BTC', '')}

# 幣安 API Key
BINANCE_API_KEY={data.get('BINANCE_API_KEY', '')}

# 幣安 Secret Key
BINANCE_SECRET_KEY={data.get('BINANCE_SECRET_KEY', '')}

# 幣安用戶ID
BINANCE_USER_ID={data.get('BINANCE_USER_ID', '')}

# 交易對
TRADING_PAIR={data.get('TRADING_PAIR', 'BTCUSDT')}

# 合約類型
CONTRACT_TYPE={data.get('CONTRACT_TYPE', 'PERPETUAL')}

# 槓桿倍數
LEVERAGE={data.get('LEVERAGE', '10')}

# 風險比例百分比
POSITION_SIZE={data.get('POSITION_SIZE', '20')}

# 保證金模式
MARGIN_TYPE={data.get('MARGIN_TYPE', 'CROSS')}

# 登入狀態
LOGIN_BTC={final_login_status}
"""
        
        # 確保配置目錄存在
        os.makedirs(os.path.dirname(BTC_ENV_PATH), exist_ok=True)
        
        # 儲存到btc.env文件
        with open(BTC_ENV_PATH, 'w', encoding='utf-8') as f:
            f.write(btc_env_content)
        
        logger.info(f"BTC配置已儲存至: {BTC_ENV_PATH}")
        
        return jsonify({
            'success': True, 
            'message': 'BTC配置儲存成功',
            'has_empty_fields': has_empty_fields
        })
        
    except Exception as e:
        logger.error(f"儲存BTC配置失敗: {e}")
        return jsonify({'success': False, 'message': f'儲存失敗: {str(e)}'})

def btc_login():
    """BTC帳戶登入/連接"""
    global binance_client, account_info
    
    try:
        # 載入BTC配置
        if not os.path.exists(BTC_ENV_PATH):
            return jsonify({'success': False, 'message': 'BTC配置不存在，請先儲存配置'})
        
        btc_env = {}
        with open(BTC_ENV_PATH, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    btc_env[key] = value
        
        # 檢查所有必填欄位
        required_fields = ['CHAT_ID_BTC', 'BINANCE_API_KEY', 'BINANCE_SECRET_KEY', 'BINANCE_USER_ID', 'TRADING_PAIR', 'LEVERAGE', 'POSITION_SIZE', 'MARGIN_TYPE', 'CONTRACT_TYPE']
        missing_fields = []
        
        for field in required_fields:
            if not btc_env.get(field, '').strip():
                missing_fields.append(field)
        
        if missing_fields:
            return jsonify({'success': False, 'message': f'以下必填欄位為空: {", ".join(missing_fields)}'})
        
        # 獲取API配置
        api_key = btc_env.get('BINANCE_API_KEY', '').strip()
        secret_key = btc_env.get('BINANCE_SECRET_KEY', '').strip()
        trading_pair = btc_env.get('TRADING_PAIR', 'BTCUSDT')
        
        # 創建幣安客戶端
        binance_client = BinanceClient(api_key, secret_key)
        
        # 檢查連接
        logger.info("開始測試幣安服務器連接...")
        connection_test = binance_client.test_connection()
        
        if not connection_test:
            # 檢查是否是418錯誤（IP限制）
            try:
                test_response = binance_client.session.get(f"{binance_client.base_url}/fapi/v1/ping", timeout=5)
                if test_response.status_code == 418:
                    error_msg = 'IP被幣安限制或請求過於頻繁'
                    possible_solutions = [
                        '1. 等待5-10分鐘後重試',
                        '2. 檢查是否使用VPN或代理',
                        '3. 確認IP地址在幣安白名單中',
                        '4. 避免過於頻繁的API請求',
                        '5. 聯繫幣安客服解除IP限制'
                    ]
                    return jsonify({'success': False, 'message': error_msg, 'details': possible_solutions, 'error_code': 418})
            except:
                pass
            
            error_msg = '無法連接到幣安服務器'
            possible_causes = [
                '1. 網絡連接問題',
                '2. 防火牆阻擋',
                '3. DNS解析問題',
                '4. 幣安服務器維護中',
                '5. IP地址被限制'
            ]
            full_error = f"{error_msg}\n可能原因:\n" + "\n".join(possible_causes)
            logger.error(full_error)
            return jsonify({'success': False, 'message': error_msg, 'details': possible_causes})
        
        # 獲取帳戶信息
        fresh_account_info = binance_client.get_account_info()
        
        if not fresh_account_info:
            return jsonify({'success': False, 'message': 'API認證失敗，請檢查API Key和Secret Key'})
        
        # 更新全局帳戶信息
        account_info = fresh_account_info
        logger.info(f"成功獲取帳戶信息，總錢包餘額: {account_info.get('totalWalletBalance', 'N/A')}")
        
        # 獲取餘額信息
        balance_info = binance_client.get_balance()
        
        # 🔧 設置槓桿倍數和保證金模式（確保與配置一致）
        try:
            leverage = btc_env.get('LEVERAGE', '10')
            margin_type = btc_env.get('MARGIN_TYPE', 'CROSS')
            
            logger.info(f"開始設置BTCUSDT槓桿為{leverage}X，保證金模式為{margin_type}")
            
            # 設置保證金模式
            margin_result = binance_client.change_margin_type('BTCUSDT', margin_type)
            logger.info(f"保證金模式設置結果: {margin_result}")
            
            # 設置槓桿倍數
            leverage_result = binance_client.change_leverage('BTCUSDT', int(leverage))
            logger.info(f"槓桿設置結果: {leverage_result}")
            
        except Exception as e:
            # 如果設置失敗，通常是因為已經有持倉或已經是相同設置
            logger.warning(f"設置槓桿/保證金模式失敗（可能已經是相同設置）: {e}")
        
        # 將登入狀態寫入btc.env
        updated_content = []
        login_found = False
        
        with open(BTC_ENV_PATH, 'r', encoding='utf-8') as f:
            for line in f:
                if line.startswith('LOGIN_BTC='):
                    updated_content.append('LOGIN_BTC=1\n')
                    login_found = True
                else:
                    updated_content.append(line)
        
        if not login_found:
            updated_content.append('LOGIN_BTC=1\n')
        
        with open(BTC_ENV_PATH, 'w', encoding='utf-8') as f:
            f.writelines(updated_content)
        
        # 準備返回信息
        total_wallet_balance = account_info.get('totalWalletBalance', '0')
        available_balance = account_info.get('availableBalance', '0')
        
        logger.info(f"BTC帳戶連接成功 - 交易對: {trading_pair}, 餘額: {total_wallet_balance} USDT")
        
        # 啟動WebSocket實時數據連接
        try:
            start_btc_websocket()
            logger.info("BTC WebSocket連接已啟動")
        except Exception as e:
            logger.error(f"啟動BTC WebSocket失敗: {e}")
        
        # 啟動風險監控
        try:
            risk_thread = threading.Thread(target=start_btc_risk_monitoring, daemon=True, name="BTC風險監控")
            register_btc_thread(risk_thread, "BTC風險監控")
            risk_thread.start()
            logger.info("BTC風險監控已啟動")
        except Exception as e:
            logger.error(f"啟動BTC風險監控失敗: {e}")
        
        # 啟動連接監控
        try:
            start_btc_connection_monitor()
        except Exception as e:
            logger.error(f"啟動BTC連接監控失敗: {e}")
        
        # 啟動訂單監控
        try:
            start_btc_order_monitoring()
            logger.info("BTC訂單監控已啟動")
        except Exception as e:
            logger.error(f"啟動BTC訂單監控失敗: {e}")
        
        # 啟動定時調度器
        start_btc_daily_scheduler()
        
        # BTC不需要12小時自動重連，只在斷線時自動重連
        # 12小時自動重連功能已移除
        
        return jsonify({
            'success': True, 
            'message': f'BTC帳戶連接成功 ({trading_pair})',
            'trading_pair': trading_pair,
            'total_balance': total_wallet_balance,
            'available_balance': available_balance,
            'account_alias': account_info.get('alias', '主帳戶')
        })
        
    except Exception as e:
        logger.error(f"BTC登入失敗: {e}")
        binance_client = None
        account_info = None
        return jsonify({'success': False, 'message': f'連接失敗: {str(e)}'})

def btc_logout():
    """BTC帳戶登出"""
    global binance_client, account_info
    
    try:
        if not os.path.exists(BTC_ENV_PATH):
            return jsonify({'success': False, 'message': 'BTC配置不存在'})
        
        # BTC沒有12小時自動重連功能，無需停止計時器
        
        # 停止Listen Key心跳維持
        stop_btc_listen_key_heartbeat()
        
        # 停止定時調度器
        stop_btc_daily_scheduler()
        
        # 清除BTC API連接
        binance_client = None
        account_info = None
        
        # 將登入狀態改為0
        updated_content = []
        with open(BTC_ENV_PATH, 'r', encoding='utf-8') as f:
            for line in f:
                if line.startswith('LOGIN_BTC='):
                    updated_content.append('LOGIN_BTC=0\n')
                else:
                    updated_content.append(line)
        
        with open(BTC_ENV_PATH, 'w', encoding='utf-8') as f:
            f.writelines(updated_content)
        
        logger.info("BTC帳戶登出成功")
        return jsonify({'success': True, 'message': 'BTC帳戶登出成功'})
        
    except Exception as e:
        logger.error(f"BTC登出失敗: {e}")
        return jsonify({'success': False, 'message': f'登出失敗: {str(e)}'})

def get_btc_bot_username():
    """獲取BTC Bot用戶名"""
    try:
        # 從BTC環境文件中讀取Bot Token
        bot_token = None
        if os.path.exists(BTC_ENV_PATH):
            with open(BTC_ENV_PATH, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if line.startswith('BOT_TOKEN_BTC='):
                        bot_token = line.split('=', 1)[1]
                        break
        
        # 如果.env中沒有token，使用硬編碼值
        if not bot_token:
            bot_token = "7912873826:AAFPPDwuwspKVdyDGwh1oVqxH6u1gQ_N-jU"
        
        if not bot_token:
            return jsonify({'username': None})
        
        # 使用Telegram Bot API獲取Bot信息
        bot_api_url = f"https://api.telegram.org/bot{bot_token}/getMe"
        
        try:
            response = requests.get(bot_api_url, timeout=10)
            if response.status_code == 200:
                bot_info = response.json()
                if bot_info.get('ok'):
                    username = bot_info['result'].get('username', 'Auto91_BtcBot')
                    return jsonify({'username': f'@{username}'})
        except:
            pass
        
        # 如果API調用失敗，返回默認值
        return jsonify({'username': '@Auto91_BtcBot'})
        
    except Exception as e:
        logger.error(f"獲取BTC Bot用戶名失敗: {e}")
        return jsonify({'username': '@Auto91_BtcBot'})

def load_btc_env():
    """載入BTC環境變量"""
    try:
        if not os.path.exists(BTC_ENV_PATH):
            # 如果文件不存在，創建預設配置
            create_default_btc_env()
        
        btc_env = {}
        with open(BTC_ENV_PATH, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    btc_env[key] = value
        return jsonify(btc_env)
            
    except Exception as e:
        logger.error(f"載入BTC配置失敗: {e}")
        return jsonify({})

def create_default_btc_env():
    """創建預設BTC環境配置文件"""
    try:
        # 確保配置目錄存在
        os.makedirs(os.path.dirname(BTC_ENV_PATH), exist_ok=True)
        
        default_config = """# Telegram Bot
BOT_TOKEN_BTC=7912873826:AAFPPDwuwspKVdyDGwh1oVqxH6u1gQ_N-jU

# Telegram ID
CHAT_ID_BTC=

# 幣安 API Key
BINANCE_API_KEY=

# 幣安 Secret Key
BINANCE_SECRET_KEY=

# 幣安用戶ID
BINANCE_USER_ID=

# 交易對
TRADING_PAIR=BTCUSDT

# 合約類型
CONTRACT_TYPE=PERPETUAL

# 槓桿倍數
LEVERAGE=10

# 風險比例百分比
POSITION_SIZE=20

# 保證金模式
MARGIN_TYPE=CROSS

# 登入狀態
LOGIN_BTC=0
"""
        
        with open(BTC_ENV_PATH, 'w', encoding='utf-8') as f:
            f.write(default_config)
        
        logger.info(f"已創建預設BTC配置文件: {BTC_ENV_PATH}")
        return True
        
    except Exception as e:
        logger.error(f"創庺預設BTC配置失敗: {e}")
        return False

def reset_btc_env_to_default():
    """重設BTC配置為預設值（保持現有的API資訊）"""
    try:
        # 讀取現有配置
        existing_config = {}
        if os.path.exists(BTC_ENV_PATH):
            with open(BTC_ENV_PATH, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith('#') and '=' in line:
                        key, value = line.split('=', 1)
                        existing_config[key] = value
        
        # 使用現有資訊或預設值
        config_content = f"""# Telegram Bot
BOT_TOKEN_BTC=7912873826:AAFPPDwuwspKVdyDGwh1oVqxH6u1gQ_N-jU

# Telegram ID
CHAT_ID_BTC={existing_config.get('CHAT_ID_BTC', '')}

# 幣安 API Key
BINANCE_API_KEY={existing_config.get('BINANCE_API_KEY', '')}

# 幣安 Secret Key
BINANCE_SECRET_KEY={existing_config.get('BINANCE_SECRET_KEY', '')}

# 幣安用戶ID
BINANCE_USER_ID={existing_config.get('BINANCE_USER_ID', '')}

# 交易對
TRADING_PAIR=BTCUSDT

# 合約類型
CONTRACT_TYPE=PERPETUAL

# 槓桿倍數
LEVERAGE=10

# 風險比例百分比
POSITION_SIZE=20

# 保證金模式
MARGIN_TYPE=CROSS

# 登入狀態
LOGIN_BTC=0
"""
        
        # 確保目錄存在
        os.makedirs(os.path.dirname(BTC_ENV_PATH), exist_ok=True)
        
        # 寫入配置
        with open(BTC_ENV_PATH, 'w', encoding='utf-8') as f:
            f.write(config_content)
        
        logger.info(f"BTC配置已重設為預設值: {BTC_ENV_PATH}")
        return True
        
    except Exception as e:
        logger.error(f"重設BTC配置失敗: {e}")
        return False

def load_btc_env_data():
    """載入BTC環境變量數據（內部使用）"""
    try:
        if not os.path.exists(BTC_ENV_PATH):
            create_default_btc_env()
        
        env_data = {}
        with open(BTC_ENV_PATH, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    env_data[key] = value
        return env_data
    except Exception as e:
        logger.error(f"載入BTC環境數據失敗: {e}")
        return {}

def calculate_btc_quantity(signal_data, account_balance):
    """計算BTC交易數量"""
    try:
        # 獲取配置的槓桿倍數
        env_data = load_btc_env_data()
        leverage = float(env_data.get('LEVERAGE', 5))
        
        # 獲取倉位大小百分比(從環境變數讀取，預設5%風險)
        position_size_pct = float(env_data.get('POSITION_SIZE', 10.0)) / 100
        
        # 計算可用餘額
        available_balance = float(account_balance.get('availableBalance', 0))
        
        # 🔧 修正計算邏輯：40%資金作為保證金，再開槓桿倍數
        # 保證金 = 可用餘額 * 倉位比例
        margin_used = available_balance * position_size_pct
        # 下單金額 = 保證金 * 槓桿倍數
        order_value = margin_used * leverage
        
        logger.info(f"💰 交易計算: 可用餘額${available_balance:.2f} × {position_size_pct*100}%倉位 = ${margin_used:.2f}保證金")
        logger.info(f"🎯 槓桿計算: ${margin_used:.2f}保證金 × {leverage}倍槓桿 = ${order_value:.2f}下單金額")
        
        # 獲取當前價格
        current_price = float(signal_data.get('price', 0))
        if current_price <= 0:
            # 如果信號沒有價格，獲取市場價格
            if binance_client:
                ticker = binance_client.get_ticker_price('BTCUSDT')
                current_price = float(ticker.get('price', 0)) if ticker else 0
        
        if current_price <= 0:
            raise ValueError("無法獲取有效價格")
        
        # 計算數量 = 下單金額 / 價格
        quantity = order_value / current_price
        
        # 幣安BTCUSDT最小下單單位(0.001)
        min_qty = 0.001
        quantity = max(min_qty, round(quantity, 3))
        
        return quantity
        
    except Exception as e:
        logger.error(f"計算BTC數量失敗: {e}")
        return 0.001  # 返回最小單位

def place_btc_futures_order(symbol, side, quantity, price=None, order_type="MARKET", reduce_only=False, is_manual=False, suppress_submit_notification=False):
    """BTC期貨下單 - 支援完整通知功能"""
    global binance_client, btc_active_trades, btc_active_order_requests
    
    # 生成下單請求ID用於重複檢查
    order_request_id = f"{symbol}_{side}_{quantity}_{order_type}_{reduce_only}_{int(datetime.now().timestamp() * 1000)}"
    
    try:
        if not binance_client:
            raise ValueError("BTC客戶端未初始化")
        
        # 重複下單保護 - 檢查是否有相同的下單請求正在處理
        with btc_order_lock:
            if order_request_id in btc_active_order_requests:
                logger.warning(f"重複下單請求被阻止: {order_request_id}")
                return {
                    'success': False,
                    'error': '重複下單請求已被阻止'
                }
            
            # 檢查是否有相似的下單請求（相同symbol+side+quantity+type）
            current_time = int(datetime.now().timestamp() * 1000)
            similar_requests = [req for req in btc_active_order_requests 
                              if req.startswith(f"{symbol}_{side}_{quantity}_{order_type}_{reduce_only}_")]
            
            if similar_requests:
                # 檢查是否在5秒內有相似請求
                for req in similar_requests:
                    req_time = int(req.split('_')[-1])
                    if current_time - req_time < 5000:  # 5秒內
                        logger.warning(f"5秒內相似下單請求被阻止: {order_request_id}")
                        return {
                            'success': False,
                            'error': '5秒內相似下單請求已被阻止'
                        }
            
            # 添加到活躍請求列表
            btc_active_order_requests.add(order_request_id)
            
            # 清理超過30秒的舊請求記錄
            btc_active_order_requests = {req for req in btc_active_order_requests 
                                       if current_time - int(req.split('_')[-1]) < 30000}
        
        # 解析動作和方向
        action = 'cover' if reduce_only else 'new'
        parsed_action = '平倉' if reduce_only else '開倉'
        
        # 方向邏輯修復：平倉時方向要相反顯示
        if reduce_only:
            # 平倉時：BUY = 平空單，SELL = 平多單
            direction = '空單' if side == 'BUY' else '多單'
        else:
            # 開倉時：BUY = 多單，SELL = 空單
            direction = '多單' if side == 'BUY' else '空單'
        # 移除手動/自動標示
        order_source = ''
        
        # 格式化數量
        formatted_quantity = f"{float(quantity):.8f}"
        
        # 記錄委託訂單日誌（僅後端日誌，不發送到前端避免重複）
        order_type_text = convert_order_type_to_display(order_type)
        commit_log = f"{order_source}{parsed_action}：{direction}｜{formatted_quantity} BTC｜市價｜{order_type_text}"
        
        # 只記錄後端日誌，前端日誌由Websocket成交回調統一處理
        logger.info(f"({order_source}委託) {commit_log}")
        
        # 執行下單
        result = binance_client.place_order(
            symbol=symbol,
            side=side,
            order_type=order_type,
            quantity=formatted_quantity,
            price=price if order_type == 'LIMIT' else None,
            reduce_only=reduce_only
        )
        
        if result:
            order_id = result.get('orderId')
            client_order_id = result.get('clientOrderId')
            
            logger.info(f"BTC訂單提交成功: OrderID={order_id}, ClientOrderID={client_order_id}")
            
            # 立即標記為已處理，避免WebSocket重複處理
            with processed_orders_lock:
                processed_orders.add(f"{order_id}_NEW")
                if order_type == 'MARKET':
                    processed_orders.add(f"{order_id}_FILLED")
            
            # 立即將訂單加入pending_orders，確保WebSocket能識別為系統訂單
            if str(order_id) not in pending_orders:
                pending_orders[str(order_id)] = {
                    'order_id': str(order_id),
                    'timestamp': datetime.now().isoformat(),
                    'symbol': symbol,
                    'action': parsed_action,
                    'direction': direction,
                    'quantity': formatted_quantity,
                    'order_type': order_type_text,
                    'source': order_source
                }
                logger.info(f"系統下單訂單{order_id}已立即加入pending_orders")
            
            # 獲取提交價格用於通知
            submitted_price = 0
            if order_type == 'LIMIT' and price:
                submitted_price = float(price)
                logger.info(f"系統限價單提交價格: {submitted_price}")
            else:
                # 市價單獲取當前市價作為參考
                try:
                    ticker = binance_client.get_ticker_price(symbol=symbol)
                    if ticker and 'price' in ticker:
                        submitted_price = float(ticker['price'])
                        logger.info(f"系統市價單參考價格獲取成功: {submitted_price}")
                    else:
                        logger.warning(f"系統獲取市價失敗，ticker回應: {ticker}")
                        submitted_price = 0
                except Exception as e:
                    logger.error(f"系統獲取市價異常: {e}")
                    submitted_price = 0
            
            logger.info(f"系統最終設置的提交價格: {submitted_price}")
            
            # 構建trade_record用於通知
            trade_record = {
                'symbol': symbol,
                'side': side,
                'quantity': formatted_quantity,
                'price': submitted_price,  # 添加提交價格
                'order_id': order_id,
                'order_type': order_type_text,
                'source': 'manual' if is_manual else 'webhook',  # 修正source字段用於判斷手動/自動
                'action_type': parsed_action,
                'reduceOnly': reduce_only,
                'is_manual': is_manual
            }
            
            # 立即標記為已處理的系統訂單，避免WebSocket重複處理
            with processed_orders_lock:
                processed_orders.add(f"{order_id}_NEW")
                if order_type == 'MARKET':
                    processed_orders.add(f"{order_id}_FILLED")  # 市價單可能立即成交
            
            # 立即將訂單加入pending_orders，確保WebSocket能識別為系統訂單
            if str(order_id) not in pending_orders:
                pending_orders[str(order_id)] = {
                    'order_id': str(order_id),
                    'timestamp': datetime.now().isoformat(),
                    'symbol': symbol,
                    'action': parsed_action,
                    'direction': direction,
                    'quantity': formatted_quantity,
                    'order_type': order_type_text,
                    'source': order_source
                }
                logger.info(f"系統下單訂單{order_id}已加入pending_orders")
            else:
                logger.info(f"系統下單訂單{order_id}已存在於pending_orders中")
            
            # 使用排隊系統發送提交成功通知（除非被抑制）
            if not suppress_submit_notification:
                logger.info(f"準備將BTC提交成功通知加入排隊: {trade_record}")
                btc_notification_manager.add_notification('submit', trade_record, True, 1)
                logger.info(f"BTC提交成功通知已加入排隊系統（延遲1秒）")
            else:
                logger.info("對翻操作：跳過額外的提交成功通知（由排隊機制處理）")
            
            # 啟動訂單成交監控
            handle_btc_order_fill(order_id, parsed_action, direction, formatted_quantity, order_type_text, order_source)
            
            # 記錄到活躍交易
            active_trade_record = {
                'order_id': order_id,
                'client_order_id': client_order_id,
                'symbol': symbol,
                'side': side,
                'quantity': formatted_quantity,
                'order_type': order_type,
                'is_manual': is_manual,
                'action_type': parsed_action,
                'reduceOnly': reduce_only,  # 🔧 修復：統一字段名為reduceOnly
                'status': result.get('status', 'NEW'),
                'timestamp': datetime.now().isoformat()
            }
            
            btc_active_trades[str(order_id)] = active_trade_record
            logger.info(f"📝 訂單已添加到btc_active_trades: {str(order_id)}")
            
            # 下單成功，清理請求記錄
            with btc_order_lock:
                btc_active_order_requests.discard(order_request_id)
            
            return {
                'success': True,
                'message': f'{order_source}{parsed_action}成功',
                'order_id': order_id,
                'symbol': symbol,
                'side': side,
                'quantity': formatted_quantity,
                'result': result
            }
        else:
            # 下單返回空結果，清理請求記錄
            with btc_order_lock:
                btc_active_order_requests.discard(order_request_id)
            raise Exception("下單返回空結果")
            
    except Exception as e:
        logger.error(f"BTC下單失敗: {e}")
        
        # 解析動作和方向（用於失敗通知）
        action = 'cover' if reduce_only else 'new'
        direction = '多單' if side == 'BUY' else '空單'
        parsed_action = '平倉' if reduce_only else '開倉'
        # 移除手動/自動標示
        order_source = ''
        formatted_quantity = f"{float(quantity):.8f}"
        order_type_text = convert_order_type_to_display(order_type)
        
        # 構建失敗的trade_record用於通知
        error_record = {
            'symbol': symbol,
            'side': side,
            'quantity': formatted_quantity,
            'order_id': '--',
            'order_type': order_type_text,
            'source': 'manual' if is_manual else 'webhook',  # 修正source字段用於判斷手動/自動
            'action_type': parsed_action,
            'reduceOnly': reduce_only,
            'is_manual': is_manual,
            'error': str(e)
        }
        
        # 使用排隊系統發送提交失敗通知（5秒延遲）
        logger.info(f"準備將BTC提交失敗通知加入排隊: {error_record}")
        btc_notification_manager.add_notification('submit', error_record, False, 5)
        logger.info(f"BTC提交失敗通知已加入排隊系統（延遲5秒）")
        
        # 下單失敗，清理請求記錄
        with btc_order_lock:
            btc_active_order_requests.discard(order_request_id)
        
        return {
            'success': False,
            'message': f'{order_source}下單失敗: {str(e)}',
            'symbol': symbol,
            'side': side,
            'quantity': formatted_quantity,
            'fail_message_sent': fail_success
        }

def check_btc_order_fill(order_id, symbol):
    """檢查BTC訂單成交狀態"""
    global binance_client, btc_active_trades
    
    try:
        if not binance_client:
            return
        
        # 查詢訂單狀態
        order_status = binance_client.get_order_status(symbol, order_id=order_id)
        
        if order_status:
            status = order_status.get('status')
            
            if status == 'FILLED':
                # 訂單已成交
                trade_record = btc_active_trades.get(order_id, {})
                trade_record.update({
                    'fill_price': order_status.get('avgPrice'),
                    'fill_quantity': order_status.get('executedQty'),
                    'fill_time': order_status.get('updateTime'),
                    'status': 'FILLED'
                })
                
                logger.info(f"BTC訂單成交: OrderID={order_id}, 成交價={trade_record.get('fill_price')}")
                
                # 記錄成交日誌到系統日誌（參考TX格式）
                fill_price = float(trade_record.get('fill_price', '0'))
                fill_quantity = trade_record.get('fill_quantity', '0')
                symbol = trade_record.get('symbol', 'BTCUSDT')
                side = trade_record.get('side', 'BUY')
                
                # 生成詳細的成交日誌（使用保存的訂單信息）
                detailed_fill_log = get_btc_order_log_message(
                    symbol=symbol,
                    side=side,
                    quantity=fill_quantity,
                    price=fill_price,
                    order_id=order_id,
                    order_type=trade_record.get('order_type', 'MARKET'),
                    is_manual=trade_record.get('is_manual', False),
                    action_type=trade_record.get('action_type', '開倉'),
                    is_success=True  # 成交時使用成功格式
                )
                log_message = detailed_fill_log
                # 僅在後端控制台顯示，不發送到前端
                
                # 使用排隊系統發送成交通知（5秒延遲）
                logger.info(f"準備將BTC成交通知加入排隊: 訂單{order_id}")
                btc_notification_manager.add_notification('fill', trade_record, True, 5)
                logger.info(f"BTC成交通知已加入排隊系統（延遲5秒）")
                
                # 延遲記錄成交成功到系統日誌（與TX系統格式統一）
                def delayed_success_log():
                    try:
                        time.sleep(5)  # 延遲5秒，在提交通知(1秒)之後，與成交通知(5秒)同步
                        success_log = get_btc_order_log_message(
                            symbol=symbol,
                            side=trade_record.get('side', ''),
                            quantity=float(executed_qty),
                            price=fill_price,
                            order_id=order_id,
                            order_type=trade_record.get('order_type', 'MARKET'),
                            is_manual=trade_record.get('is_manual', False),
                            action_type=trade_record.get('action_type', '開倉'),
                            is_success=True
                        )
                        log_btc_frontend_message(success_log, "success")
                        logger.info(f"延遲5秒後記錄BTC成交日誌: {success_log}")
                    except Exception as e:
                        logger.error(f"延遲成交日誌記錄失敗: {e}")
                
                # 在背景執行緒中執行延遲日誌記錄
                thread = threading.Thread(target=delayed_success_log)
                thread.daemon = True
                thread.start()
                
                # 保存交易記錄
                save_btc_trade_record(trade_record)
                
            elif status in ['CANCELED', 'REJECTED', 'EXPIRED']:
                # 訂單失效 - 這也屬於"掛單沒成功"，應觸發提交失敗通知
                trade_record = btc_active_trades.get(str(order_id), {})
                trade_record.update({'status': status})
                logger.info(f"🔴 BTC訂單失效觸發: OrderID={order_id}, 狀態={status}")
                logger.info(f"🔍 從btc_active_trades查找 {str(order_id)}: 找到={str(order_id) in btc_active_trades}")
                logger.info(f"🔍 當前btc_active_trades keys: {list(btc_active_trades.keys())}")
                logger.info(f"🔍 訂單記錄內容: {trade_record}")
                
                # 記錄失效日誌到系統日誌
                symbol = trade_record.get('symbol', 'BTCUSDT')
                side = trade_record.get('side', '未知')
                quantity = trade_record.get('quantity', '0')
                
                log_message = f"BTC訂單失效：{symbol}｜{side}｜{quantity}｜狀態：{status}"
                log_btc_frontend_message(log_message, "warning")
                
                # 🔥 新增：發送提交失敗通知（因為掛單沒成功）
                try:
                    # 構建失敗通知記錄
                    status_mapping = {
                        'CANCELED': '訂單被取消',
                        'REJECTED': '訂單被拒絕', 
                        'EXPIRED': '訂單已過期'
                    }
                    failure_reason = status_mapping.get(status, f'訂單失效({status})')
                    
                    # 使用現有的訂單信息構建失敗記錄，包含價格信息
                    error_record = {
                        'symbol': symbol,
                        'side': side,
                        'quantity': quantity,
                        'order_id': order_id,
                        'price': trade_record.get('price', 0),  # 添加價格信息
                        'order_type': trade_record.get('order_type', '市價單'),
                        'source': trade_record.get('source', '手動'),
                        'action_type': trade_record.get('action_type', '開倉'),
                        'reduceOnly': trade_record.get('reduceOnly', False),
                        'is_manual': trade_record.get('is_manual', False),
                        'error': failure_reason
                    }
                    
                    # 使用排隊系統發送提交失敗通知（5秒延遲）
                    logger.info(f"🚀 準備將BTC失效通知加入排隊: {error_record}")
                    btc_notification_manager.add_notification('submit', error_record, False, 5)
                    logger.info(f"📱 訂單失效提交失敗通知已加入排隊系統: {order_id} - {failure_reason}")
                    
                except Exception as e:
                    logger.error(f"發送訂單失效提交失敗通知失敗: {e}")
                
        # 從活躍交易中移除已完成的訂單
        if str(order_id) in btc_active_trades:
            final_status = btc_active_trades[str(order_id)].get('status')
            if final_status in ['FILLED', 'CANCELED', 'REJECTED', 'EXPIRED']:
                logger.info(f"從活躍交易中移除已完成訂單: {order_id}")
                del btc_active_trades[str(order_id)]
                
    except Exception as e:
        logger.error(f"檢查BTC訂單狀態失敗: {e}")

def auto_close_opposite_position(symbol, target_action):
    """自動平倉反向持倉
    
    Args:
        symbol: 交易對，如 'BTCUSDT'
        target_action: 目標動作 'LONG' 或 'SHORT'
        
    Returns:
        dict: 平倉結果
    """
    try:
        if not binance_client:
            return {'error': 'BTC客戶端未初始化'}
            
        positions = binance_client.get_position_info()
        if not positions:
            return {'error': '無法獲取持倉信息'}
            
        # 找到對應的持倉
        target_position = None
        for pos in positions:
            if pos.get('symbol') == symbol and float(pos.get('positionAmt', 0)) != 0:
                target_position = pos
                break
        
        if not target_position:
            logger.info(f"沒有找到{symbol}的活躍持倉，跳過平倉")
            return {'success': True, 'message': '無持倉需要平倉'}
        
        # 計算平倉方向和數量
        position_amt = float(target_position.get('positionAmt', 0))
        current_side = 'LONG' if position_amt > 0 else 'SHORT'
        
        # 確認是反向持倉
        if (target_action == 'LONG' and current_side == 'SHORT') or \
           (target_action == 'SHORT' and current_side == 'LONG'):
            
            # 計算平倉參數
            if position_amt > 0:
                # 多頭持倉，需要賣出平倉
                close_side = 'SELL'
                quantity = abs(position_amt)
            else:
                # 空頭持倉，需要買入平倉
                close_side = 'BUY'
                quantity = abs(position_amt)
            
            logger.info(f"執行自動平倉：{current_side}持倉 {quantity} BTC")
            
            # 執行平倉單（抑制提交通知，由排隊機制處理）
            close_result = place_btc_futures_order(
                symbol=symbol,
                side=close_side,
                quantity=quantity,
                order_type='MARKET',
                reduce_only=True,   # 平倉
                is_manual=False,    # 自動交易
                suppress_submit_notification=True  # 抑制提交通知
            )
            
            if close_result and not close_result.get('error'):
                logger.info(f"自動平倉成功：訂單ID {close_result.get('orderId')}")
                return {'success': True, 'order_result': close_result}
            else:
                error_msg = close_result.get('error') if close_result else '未知錯誤'
                logger.error(f"自動平倉失敗：{error_msg}")
                return {'error': f'平倉下單失敗：{error_msg}'}
        else:
            return {'success': True, 'message': '持倉方向與目標一致，無需平倉'}
            
    except Exception as e:
        logger.error(f"自動平倉執行失敗: {e}")
        return {'error': f'平倉執行異常：{str(e)}'}

def check_opposite_position(action):
    """檢查是否存在反向持倉
    
    Args:
        action: 'LONG' 或 'SHORT'
        
    Returns:
        bool: True if 存在反向持倉
    """
    try:
        if not binance_client:
            return False
            
        positions = binance_client.get_position_info()
        if not positions:
            return False
            
        for pos in positions:
            if pos.get('symbol') == 'BTCUSDT':
                position_amt = float(pos.get('positionAmt', 0))
                
                # 檢查是否有持倉
                if abs(position_amt) > 0.00000001:
                    current_side = 'LONG' if position_amt > 0 else 'SHORT'
                    
                    # 檢查是否為反向
                    if (action == 'LONG' and current_side == 'SHORT') or \
                       (action == 'SHORT' and current_side == 'LONG'):
                        logger.info(f"發現反向持倉: 當前{current_side}，嘗試開{action}")
                        return True
                        
        return False
        
    except Exception as e:
        logger.error(f"檢查反向持倉失敗: {e}")
        return False

def process_btc_entry_signal(signal_data, parsed_action=None):
    """處理BTC進場信號"""
    try:
        action = parsed_action or signal_data.get('action', '').upper()
        symbol = signal_data.get('symbol', 'BTCUSDT')
        
        if action not in ['LONG', 'SHORT']:
            raise ValueError(f"無效的進場動作: {action}")
        
        # 標準化方向（BTC期貨邏輯）
        side = 'BUY' if action == 'LONG' else 'SELL'
        
        # 獲取帳戶餘額計算數量
        if not binance_client:
            raise ValueError("BTC客戶端未初始化")
        
        balance_info = binance_client.get_balance()
        if not balance_info:
            raise ValueError("無法獲取帳戶餘額")
        
        # 找到USDT餘額
        usdt_balance = None
        for balance in balance_info:
            if balance.get('asset') == 'USDT':
                usdt_balance = balance
                break
        
        if not usdt_balance:
            raise ValueError("找不到USDT餘額")
        
        # 計算交易數量
        quantity = calculate_btc_quantity(signal_data, usdt_balance)
        
        # 檢查反向持倉
        has_opposite_position = check_opposite_position(action)
        
        if has_opposite_position:
            logger.info(f"檢測到反向持倉，開始執行自動對翻：當前持倉 -> {action}")
            
            # 先執行平倉
            close_result = auto_close_opposite_position(symbol, action)
            if not close_result or close_result.get('error'):
                # 平倉失敗，發送失敗通知
                trade_record = {
                    'symbol': symbol,
                    'side': side,
                    'quantity': quantity,
                    'order_type': 'MARKET',
                    'source': 'webhook',
                    'action_type': '開倉',
                    'is_manual': False,
                    'order_id': '平倉失敗',  # 添加order_id避免顯示"未知"
                    'error': f"平倉失敗：{close_result.get('error', '未知錯誤')}"
                }
                send_btc_order_submit_notification_delayed(trade_record, success=False, delay_seconds=5)
                return {'error': f'自動對翻失敗：平倉階段出錯'}
            
            # 平倉成功，準備排隊發送對翻通知
            close_order_result = close_result.get('order_result', {})
            
            # 使用排隊系統管理對翻通知順序
            # 1. 平倉提交成功通知（立即發送）
            close_trade_record = {
                'symbol': symbol,
                'side': close_order_result.get('side', ''),
                'quantity': close_order_result.get('origQty', 0),
                'order_type': 'MARKET',
                'source': 'webhook',
                'action_type': '平倉',
                'is_manual': False,
                'order_id': close_order_result.get('orderId')
            }
            btc_notification_manager.add_notification('submit', close_trade_record, True, 0)
            logger.info("平倉提交成功通知已加入排隊系統（立即發送）")
            
            # 2. 開倉提交成功通知（延遲6秒，在平倉成交通知之後）
            open_trade_record = {
                'symbol': symbol,
                'side': side,
                'quantity': quantity,
                'order_type': 'MARKET',
                'source': 'webhook',
                'action_type': '開倉',
                'is_manual': False,
                'order_id': ''  # 開倉訂單號將由實際下單結果提供
            }
            btc_notification_manager.add_notification('submit', open_trade_record, True, 6)
            logger.info("開倉提交成功通知已加入排隊系統（延遲6秒）")
            
            # 注意：平倉和開倉的成交通知會由WebSocket自動處理並加入排隊
            
            # 平倉成功，等待一小段時間確保訂單處理完成
            time.sleep(0.5)
            
            logger.info(f"平倉完成，開始開{action}倉")
        
        # 執行下單（webhook為自動交易，預設為開倉）
        # 如果是對翻操作，跳過額外的提交通知（已由排隊機制處理）
        suppress_submit_notification = has_opposite_position
        
        order_result = place_btc_futures_order(
            symbol=symbol,
            side=side,
            quantity=quantity,
            order_type='MARKET',
            reduce_only=False,  # 開倉
            is_manual=False,    # 自動交易
            suppress_submit_notification=suppress_submit_notification
        )
        
        return order_result
        
    except Exception as e:
        logger.error(f"處理BTC進場信號失敗: {e}")
        return None

def process_btc_exit_signal(signal_data, parsed_action=None):
    """處理BTC出場信號"""
    try:
        action = parsed_action or signal_data.get('action', '').upper()
        symbol = signal_data.get('symbol', 'BTCUSDT')
        
        if action not in ['CLOSE', 'CLOSE_LONG', 'CLOSE_SHORT']:
            raise ValueError(f"無效的出場動作: {action}")
        
        # 獲取當前持倉
        if not binance_client:
            raise ValueError("BTC客戶端未初始化")
        
        positions = binance_client.get_position_info()
        if not positions:
            logger.info("沒有找到持倉信息")
            return None
        
        # 找到對應的持倉
        target_position = None
        for pos in positions:
            if pos.get('symbol') == symbol and float(pos.get('positionAmt', 0)) != 0:
                target_position = pos
                break
        
        if not target_position:
            logger.info(f"沒有找到{symbol}的活躍持倉")
            return None
        
        # 計算平倉方向和數量
        position_amt = float(target_position.get('positionAmt', 0))
        
        if position_amt > 0:
            # 多頭持倉，需要賣出平倉
            side = 'SELL'
            quantity = abs(position_amt)
        else:
            # 空頭持倉，需要買入平倉
            side = 'BUY' 
            quantity = abs(position_amt)
        
        # 執行平倉單
        order_result = place_btc_futures_order(
            symbol=symbol,
            side=side,
            quantity=quantity,
            order_type='MARKET',
            reduce_only=True,   # 平倉
            is_manual=False     # 自動交易
        )
        
        return order_result
        
    except Exception as e:
        logger.error(f"處理BTC出場信號失敗: {e}")
        return None


def save_btc_trade_record(trade_record):
    """保存BTC交易記錄 - 使用日期格式文件名"""
    try:
        today = datetime.now().strftime("%Y%m%d")
        trades_file = os.path.join(os.path.dirname(__file__), 'BTCtransdata', f'BTCtrades_{today}.json')
        
        # 確保目錄存在
        os.makedirs(os.path.dirname(trades_file), exist_ok=True)
        
        # 讀取現有交易記錄
        existing_trades = []
        if os.path.exists(trades_file):
            try:
                with open(trades_file, 'r', encoding='utf-8') as f:
                    existing_trades = json.load(f)
            except:
                existing_trades = []
        
        # 添加時間戳
        trade_record['timestamp'] = datetime.now().isoformat()
        
        # 添加新交易記錄
        existing_trades.append(trade_record)
        
        # 保存回文件
        with open(trades_file, 'w', encoding='utf-8') as f:
            json.dump(existing_trades, f, ensure_ascii=False, indent=2)
        
        # 清理舊的BTC交易記錄檔案（保留30個交易日）
        cleanup_old_btc_trade_files()
            
    except Exception as e:
        logger.error(f"保存BTC交易記錄失敗: {e}")

def cleanup_old_btc_trade_files():
    """清理舊的BTC交易記錄檔案，保留30個交易日"""
    try:
        btc_log_dir = os.path.join(os.path.dirname(__file__), 'BTCtransdata')
        if not os.path.exists(btc_log_dir):
            return
        
        # 獲取所有BTC交易記錄檔案
        trade_files = []
        for filename in os.listdir(btc_log_dir):
            if filename.startswith('BTCtrades_') and filename.endswith('.json'):
                try:
                    # 從檔案名提取日期
                    date_str = filename.replace('BTCtrades_', '').replace('.json', '')
                    file_date = datetime.strptime(date_str, '%Y%m%d')
                    trade_files.append((filename, file_date))
                except ValueError:
                    # 如果檔案名格式不正確，跳過
                    continue
        
        # 按日期排序（最新的在前）
        trade_files.sort(key=lambda x: x[1], reverse=True)
        
        # 保留最新的30個檔案，刪除其餘的
        if len(trade_files) > 30:
            files_to_delete = trade_files[30:]
            for filename, _ in files_to_delete:
                try:
                    file_path = os.path.join(btc_log_dir, filename)
                    os.remove(file_path)
                    logger.info(f"已刪除舊BTC交易記錄檔案：{filename}")
                except Exception as e:
                    logger.error(f"刪除BTC檔案失敗 {filename}：{e}")
            
        logger.info(f"BTC交易記錄檔案清理完成，保留 {min(len(trade_files), 30)} 個檔案")
    
    except Exception as e:
        logger.error(f"清理舊BTC交易記錄檔案失敗：{e}")


def save_btc_order_record(order_data, is_manual=True):
    """保存BTC訂單記錄到tradedataBTC目錄（與TX系統格式一致）
    
    Args:
        order_data: 幣安WebSocket訂單數據
        is_manual: 是否為手動訂單
    """
    try:
        # 獲取當日日期
        today = datetime.now().strftime("%Y%m%d")
        
        # 創建目錄（只在有數據時創建）
        btc_data_dir = TradingConfig.BTC_DATA_DIR
        os.makedirs(btc_data_dir, exist_ok=True)
        
        # 檔案路徑
        transdata_file = os.path.join(btc_data_dir, f'BTCtransdata_{today}.json')
        
        # 讀取現有數據
        existing_data = []
        if os.path.exists(transdata_file):
            try:
                with open(transdata_file, 'r', encoding='utf-8') as f:
                    existing_data = json.load(f)
            except Exception as e:
                logger.error(f"讀取現有BTCtransdata失敗: {e}")
                existing_data = []
        
        # 構建訂單記錄（仿照TX格式）
        order_record = {
            "type": "order",
            "trade_category": "normal",
            "raw_data": {
                "operation": {
                    "op_type": "New",
                    "op_code": "00", 
                    "op_msg": ""
                },
                "order": {
                    "id": order_data.get('i', ''),  # 訂單ID
                    "account": {
                        "account_type": "F",  # 期貨
                        "person_id": "",
                        "broker_id": "BINANCE",
                        "account_id": "BTC_FUTURES",
                        "signed": True
                    },
                    "action": "Buy" if order_data.get('S') == "BUY" else "Sell",
                    "price": float(order_data.get('p', 0)),
                    "quantity": float(order_data.get('q', 0)),
                    "order_type": convert_order_type_to_display(order_data.get('o', 'MARKET')),  # 使用轉換函數
                    "price_type": "LMT" if order_data.get('o') == "LIMIT" else "MKT",
                    "market_type": "Day",
                    "oc_type": "Open" if not order_data.get('R', False) else "Cover",
                    "subaccount": "",
                    "combo": False
                },
                "status": {
                    "id": order_data.get('i', ''),
                    "exchange_ts": float(order_data.get('T', 0)) / 1000,  # 轉換為秒
                    "modified_price": 0.0,
                    "cancel_quantity": 0,
                    "order_quantity": float(order_data.get('q', 0)),
                    "web_id": "M" if is_manual else "W"  # M=Manual, W=Webhook
                },
                "contract": {
                    "security_type": "FUT",
                    "code": "BTC",
                    "exchange": "BINANCE",
                    "delivery_month": "",
                    "delivery_date": "",
                    "strike_price": 0.0,
                    "option_right": "Future"
                }
            },
            "deal_order_id": order_data.get('i', ''),
            "contract_name": "比特幣",
            "timestamp": datetime.now().isoformat(),
            "is_manual": is_manual
        }
        
        # 添加新記錄
        existing_data.append(order_record)
        
        # 保存檔案
        with open(transdata_file, 'w', encoding='utf-8') as f:
            json.dump(existing_data, f, ensure_ascii=False, indent=2)
        
        logger.info(f"✅ BTC訂單記錄已保存: {order_data.get('i', '')} 到 {transdata_file}")
        
    except Exception as e:
        logger.error(f"保存BTC訂單記錄失敗: {e}")


def send_btc_telegram_message(message, chat_id=None, bot_token=None):
    """發送BTC Telegram訊息（使用統一工具函數）"""
    try:
        
        # 載入BTC配置
        env_data = load_config_file(BTC_ENV_PATH, ['CHAT_ID_BTC', 'BOT_TOKEN_BTC'])
        
        if not chat_id:
            chat_id_raw = env_data.get('CHAT_ID_BTC')
        else:
            chat_id_raw = chat_id
        if not bot_token:
            bot_token = env_data.get('BOT_TOKEN_BTC')
        
        if not chat_id_raw or not bot_token:
            logger.info("BTC Telegram配置不完整，跳過通知")
            log_btc_frontend_message("Telegram配置不完整", "error")
            return False
        
        # 支援多個CHAT_ID，用逗號分隔
        chat_ids = [id.strip() for id in str(chat_id_raw).split(',') if id.strip()]
        
        logger.info(f"BTC BOT_TOKEN: {bot_token[:10]}...")
        logger.info(f"BTC CHAT_IDs: {chat_ids}")
        
        url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
        
        # 記錄發送結果
        success_count = 0
        total_count = len(chat_ids)
        
        # 對每個CHAT_ID發送訊息
        for chat_id in chat_ids:
            payload = {
                'chat_id': chat_id,
                'text': message,
                'parse_mode': 'HTML'
            }
            
            logger.info(f"發送請求到 BTC Telegram API (Chat ID: {chat_id})...")
            response = requests.post(url, json=payload, timeout=10)
            
            logger.info(f"BTC Telegram API 回應 (Chat ID: {chat_id}): {response.status_code}")
            
            if response.status_code == 200:
                logger.info(f"BTC Telegram 訊息發送成功 (Chat ID: {chat_id})！")
                success_count += 1
            else:
                logger.error(f"BTC Telegram 訊息發送失敗 (Chat ID: {chat_id}): {response.text}")
        
        # 判斷整體發送結果
        if success_count == total_count:
            logger.info(f"BTC Telegram 訊息發送完成！成功發送到 {success_count}/{total_count} 個接收者")
            
            # 根據訊息內容判斷發送狀態類型（參考TX系統）
            if "訂單提交成功" in message or "提交成功" in message or "⭕ 提交成功" in message:
                log_message = "Telegram［提交成功］訊息發送成功！！！"
            elif "訂單提交失敗" in message or "提交失敗" in message or "❌ 提交失敗" in message:
                log_message = "Telegram［提交失敗］訊息發送成功！！！"
            elif "成交通知" in message or "訂單成交" in message:
                log_message = "Telegram［成交通知］訊息發送成功！！！"
            elif "API連線異常" in message or "連線失敗" in message:
                log_message = "Telegram［API連線異常］訊息發送成功！！！"
            elif "API連線成功" in message or "連線成功" in message:
                log_message = "Telegram［API連線成功］訊息發送成功！！！"
            elif "交易統計" in message or "統計報告" in message:
                log_message = "Telegram［交易統計］訊息發送成功！！！"
            elif "日報" in message or "月報" in message or "報表" in message:
                log_message = "Telegram［生成報表］訊息發送成功！！！"
            elif "系統啟動" in message or "啟動通知" in message or "正在啟動中" in message:
                log_message = "Telegram［啟動通知］訊息發送成功！！！"
            elif "風險管理警報" in message or "虧損金額" in message:
                log_message = "Telegram［風險管理］訊息發送成功！！！"
            else:
                log_message = "Telegram 訊息發送成功！！！"
            
            # 記錄到BTC前端系統日誌
            log_type = 'warning' if 'API連線異常' in log_message else 'success'
            log_btc_frontend_message(log_message, log_type)
            logger.info(f"BTC系統日誌已發送: {log_message}")
            
            return True
        else:
            logger.error(f"BTC Telegram 訊息部分發送失敗！成功發送到 {success_count}/{total_count} 個接收者")
            # 發送失敗也要記錄日誌
            status_type = 'warning' if success_count > 0 else 'error'
            error_log_message = f"Telegram 訊息部分發送失敗！成功：{success_count}/{total_count}"
            log_btc_frontend_message(error_log_message, status_type)
            logger.info(f"BTC系統錯誤日誌已發送: {error_log_message}")
            return success_count > 0  # 至少有一個成功就返回True
            
    except Exception as e:
        logger.error(f"發送BTC Telegram訊息失敗: {e}")
        logger.error(f"錯誤類型: {str(e.__class__.__name__)}")
        if hasattr(e, 'response'):
            logger.error(f"回應內容: {e.response.text}")
        import traceback
        traceback.print_exc()
        
        # 記錄異常到系統日誌
        error_log_message = f"Telegram 訊息發送異常：{str(e)[:100]}"
        log_btc_frontend_message(error_log_message, "error")
        return False

def send_btc_telegram_file(file_path, caption=""):
    """發送BTC Telegram檔案（與TX系統格式一致）"""
    try:
        if not os.path.exists(file_path):
            logger.info(f"找不到檔案，路徑: {file_path}")
            return False
        
        # 載入BTC配置
        env_data = load_btc_env_data()
        
        bot_token = env_data.get('BOT_TOKEN_BTC')
        chat_id_raw = env_data.get('CHAT_ID_BTC')
        
        if not bot_token or not chat_id_raw:
            logger.info("BTC Telegram配置不完整，跳過檔案發送")
            log_btc_frontend_message("Telegram配置不完整", "error")
            return False
        
        # 支援多個CHAT_ID，用逗號分隔
        chat_ids = [id.strip() for id in str(chat_id_raw).split(',') if id.strip()]
        
        logger.info(f"準備發送BTC檔案到 {len(chat_ids)} 個接收者")
        
        # 發送檔案
        url = f"https://api.telegram.org/bot{bot_token}/sendDocument"
        
        # 記錄發送結果
        success_count = 0
        total_count = len(chat_ids)
        
        # 對每個CHAT_ID發送檔案
        for chat_id in chat_ids:
            with open(file_path, 'rb') as f:
                files = {'document': f}
                data = {
                    'chat_id': chat_id,
                    'caption': caption
                }
                
                logger.info(f"發送BTC檔案到 Chat ID: {chat_id}")
                response = requests.post(url, files=files, data=data, timeout=30)
                
                if response.status_code == 200:
                    logger.info(f"BTC Telegram檔案發送成功 (Chat ID: {chat_id})")
                    success_count += 1
                else:
                    logger.error(f"BTC Telegram檔案發送失敗 (Chat ID: {chat_id}): {response.status_code}")
        
        # 判斷整體發送結果
        if success_count == total_count:
            logger.info(f"BTC Telegram檔案發送完成！成功發送到 {success_count}/{total_count} 個接收者")
            
            # 記錄前端系統日誌（合併發送：生成報表 + 檔案發送）
            log_btc_frontend_message("Telegram［生成報表］訊息發送成功！！！", "success")
            log_btc_frontend_message("Telegram［檔案發送］訊息發送成功！！！", "success")
            
            return True
        else:
            logger.error(f"BTC Telegram檔案部分發送失敗！成功發送到 {success_count}/{total_count} 個接收者")
            status_type = 'warning' if success_count > 0 else 'error'
            log_btc_frontend_message(f"Telegram檔案部分發送失敗！成功：{success_count}/{total_count}", status_type)
            return success_count > 0  # 至少有一個成功就返回True
            
    except Exception as e:
        logger.error(f"發送BTC Telegram檔案失敗: {e}")
        log_btc_frontend_message(f"Telegram檔案發送失敗！異常錯誤：{str(e)[:50]}", "error")
        return False

def get_btc_order_log_message(symbol, side, quantity, price, order_id, order_type, is_manual, action_type, is_success=False):
    """生成BTC訂單日誌訊息（完全參考TX邏輯）"""
    try:
        # 使用完整交易對名稱（如BTCUSDT），與TX系統的合約代碼對應
        simple_symbol = symbol
        
        # 判斷手動/自動（與TX一致）
        manual_type = '手動' if is_manual else '自動'
        
        # 格式化價格（與TX邏輯一致）
        if price == 0 or order_type == 'MARKET':
            price_display = '市價'
        else:
            price_display = f'{price:,.0f}'
        
        # 格式化方向 - 完全遵循TX邏輯
        if action_type == '開倉':
            # 開倉：BUY=多單, SELL=空單（與TX NEW邏輯一致）
            if str(side).upper() == 'BUY':
                direction_display = '多單'
            else:
                direction_display = '空單'
        elif action_type == '平倉':
            # 平倉：SELL=平多單, BUY=平空單（與get_btc_action_and_direction邏輯保持一致）
            if str(side).upper() == 'SELL':
                direction_display = '多單'  # 平多單
            else:  # BUY
                direction_display = '空單'  # 平空單
        else:
            # 備援邏輯（與TX一致）
            if str(side).upper() == 'BUY':
                direction_display = '多單'
            else:
                direction_display = '空單'
        
        # 格式化數量（BTC格式）
        quantity_display = f'{quantity} BTC'
        
        # 格式化訂單類型 - 統一使用轉換函數
        order_info = convert_order_type_to_display(order_type)
        
        # 返回格式（與TX格式完全一致）
        if is_success:
            # 成交成功格式 - 顯示成交價格
            return f"{action_type}成功：{simple_symbol}｜{direction_display}｜{quantity_display}｜{price_display}｜{order_info}"
        else:
            # 掛單格式 - 市價單顯示"市價"作為價格，限價單顯示實際價格
            if order_type.upper() == 'MARKET':
                return f"{manual_type}{action_type}：{simple_symbol}｜{direction_display}｜{quantity_display}｜市價｜{order_info}"
            else:
                return f"{manual_type}{action_type}：{simple_symbol}｜{direction_display}｜{quantity_display}｜{price_display}｜{order_info}"
            
    except Exception as e:
        logger.error(f"生成BTC日誌訊息失敗: {e}")
        return f"日誌生成失敗: {order_id}"

def log_btc_system_message(message, log_type="info"):
    """記錄BTC系統日誌（僅後端控制台，不發送到前端）"""
    try:
        # 只在後端控制台顯示，不發送到前端
        logger.info(f"[BTC系統日誌] {log_type.upper()}: {message}")
    except:
        pass

def log_btc_frontend_message(message, log_type="info"):
    """記錄BTC前端系統日誌"""
    try:
        # 動態獲取當前端口
        current_port = 5000  # 預設端口
        try:
            # 嘗試從主模組獲取當前端口
            import sys
            if 'main' in sys.modules:
                main_module = sys.modules['main']
                if hasattr(main_module, 'CURRENT_PORT'):
                    current_port = main_module.CURRENT_PORT
            else:
                # 嘗試讀取port.txt文件
                import os
                port_file = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'port.txt')
                if os.path.exists(port_file):
                    with open(port_file, 'r', encoding='utf-8') as f:
                        for line in f:
                            if line.startswith('port:'):
                                current_port = int(line.split(':')[1].strip())
                                break
        except:
            pass
        
        # 準備日誌數據
        log_data = {
            'message': message, 
            'type': log_type, 
            'system': 'BTC'
        }
        
        # BTC前端日誌已移除，直接返回成功
        logger.info(f"BTC日誌: {message}")
        return True
        
    except Exception as e:
        logger.error(f"BTC日誌處理異常: {e}")
        return False
    
    # 無論如何都記錄到後端日誌作為備份
    try:
        logger.info(f"[BTC {log_type.upper()}] {message}")
    except:
            pass

def get_btc_position_info():
    """獲取BTC持倉信息，包括USDT持倉數量和強平價格"""
    try:
        if not binance_client:
            return None, None, None
            
        # 獲取持倉資訊
        positions = binance_client.get_position_info()
        for pos in positions:
            if pos.get('symbol') == 'BTCUSDT':
                position_amt = float(pos.get('positionAmt', 0))
                if position_amt != 0:
                    entry_price = float(pos.get('entryPrice', 0))
                    mark_price = float(pos.get('markPrice', 0))
                    liquidation_price = float(pos.get('liquidationPrice', 0))
                    
                    # 計算USDT持倉數量 = BTC數量 × 標記價格
                    usdt_position = abs(position_amt) * mark_price
                    
                    return usdt_position, entry_price, liquidation_price
        
        return None, None, None
        
    except Exception as e:
        logger.error(f"獲取BTC持倉信息失敗: {e}")
        return None, None, None

def get_btc_leverage_info():
    """獲取BTC槓桿和保證金模式信息 - 優先從API獲取實際設置"""
    try:
        # 如果BTC客戶端已初始化，優先從API獲取實際設置
        if binance_client:
            try:
                position_info = binance_client.get_position_risk(symbol='BTCUSDT')
                if position_info and len(position_info) > 0:
                    pos_data = position_info[0]  # BTCUSDT的持倉數據
                    actual_leverage = pos_data.get('leverage', '10')
                    actual_margin_type = pos_data.get('marginType', 'cross')
                    margin_mode = "全倉" if actual_margin_type.lower() == 'cross' else "逐倉"
                    logger.info(f"從API獲取實際槓桿設置: {actual_leverage}X {margin_mode}")
                    return f"{actual_leverage}x", margin_mode
            except Exception as e:
                logger.warning(f"從API獲取槓桿信息失敗: {e}")
        
        # 回退到配置文件
        env_data = load_btc_env_data()
        leverage = env_data.get('LEVERAGE', '10')
        margin_type = env_data.get('MARGIN_TYPE', 'CROSS')
        margin_mode = "全倉" if margin_type == 'CROSS' else "逐倉"
        logger.info(f"從配置文件獲取槓桿設置: {leverage}X {margin_mode}")
        return f"{leverage}x", margin_mode
    except:
        return "20x", "全倉"

def send_btc_order_submit_notification_delayed(trade_record, success=True, delay_seconds=1):
    """發送延遲的BTC訂單提交通知"""
    def delayed_send():
        time.sleep(delay_seconds)
        logger.info(f"延遲{delay_seconds}秒後發送BTC提交通知: {trade_record}")
        result = send_btc_order_submit_notification(trade_record, success)
        logger.info(f"延遲BTC提交通知發送結果: {result}")
        return result
    
    # 在背景執行緒中執行延遲發送
    thread = threading.Thread(target=delayed_send)
    thread.daemon = True
    thread.start()
    return True

# ============ BTC通知排隊系統 ============

class BTCNotificationQueue:
    """BTC通知排隊管理器 - 企業級排隊系統"""
    
    def __init__(self):
        self.queue = []
        self.lock = threading.Lock()
        self.worker_running = False
        self.worker_thread = None
        self.shutdown_flag = False
    
    def add_notification(self, notification_type, trade_record, success=True, delay_seconds=0):
        """添加通知到排隊列表
        
        Args:
            notification_type: 'submit' 或 'fill'
            trade_record: 交易記錄
            success: 成功狀態
            delay_seconds: 延遲秒數
        """
        with self.lock:
            notification_item = {
                'type': notification_type,
                'trade_record': trade_record.copy(),
                'success': success,
                'delay_seconds': delay_seconds,
                'scheduled_time': time.time() + delay_seconds,
                'order_id': trade_record.get('order_id', 'Unknown')
            }
            
            self.queue.append(notification_item)
            logger.info(f"📋 BTC通知已加入排隊: {notification_type} 訂單{notification_item['order_id']} (延遲{delay_seconds}秒)")
            
            # 啟動工作器（如果未運行）
            self._ensure_worker_running()
    
    def _ensure_worker_running(self):
        """確保工作器線程正在運行"""
        if not self.worker_running and not self.shutdown_flag:
            self.worker_running = True
            self.worker_thread = threading.Thread(target=self._worker_loop)
            self.worker_thread.daemon = True
            self.worker_thread.start()
            logger.info("🚀 BTC通知排隊工作器已啟動")
    
    def _worker_loop(self):
        """工作器主循環 - 按時間順序處理排隊通知"""
        logger.info("🔄 BTC通知排隊工作器開始運行")
        
        while not self.shutdown_flag and not btc_shutdown_flag.is_set():
            try:
                current_time = time.time()
                
                with self.lock:
                    # 找到所有應該執行的通知
                    ready_notifications = [
                        (i, item) for i, item in enumerate(self.queue) 
                        if item['scheduled_time'] <= current_time
                    ]
                    
                    # 按排隊順序處理（先進先出）
                    if ready_notifications:
                        # 移除已處理的通知（從後往前刪除，避免索引錯誤）
                        for i, notification in sorted(ready_notifications, reverse=True):
                            del self.queue[i]
                
                # 處理準備好的通知
                for _, notification in ready_notifications:
                    try:
                        self._process_notification(notification)
                        # 每個通知間隔0.5秒，避免Telegram API限流
                        time.sleep(0.5)
                    except Exception as e:
                        logger.error(f"處理BTC排隊通知失敗: {e}")
                
                # 檢查是否還有排隊通知
                with self.lock:
                    if not self.queue:
                        # 沒有更多通知，停止工作器
                        self.worker_running = False
                        logger.info("📋 BTC通知排隊列表已清空，工作器停止")
                        break
                
                # 短暫休眠避免CPU過度使用
                time.sleep(0.1)
                
            except Exception as e:
                logger.error(f"BTC通知排隊工作器異常: {e}")
                time.sleep(1)
        
        self.worker_running = False
        logger.info("🔴 BTC通知排隊工作器已停止")
    
    def _process_notification(self, notification):
        """處理單個通知"""
        notification_type = notification['type']
        trade_record = notification['trade_record']
        success = notification['success']
        order_id = notification['order_id']
        
        logger.info(f"📤 處理BTC排隊通知: {notification_type} 訂單{order_id}")
        
        if notification_type == 'submit':
            send_btc_order_submit_notification(trade_record, success)
        elif notification_type == 'fill':
            send_btc_trade_notification(trade_record)
        else:
            logger.error(f"未知的通知類型: {notification_type}")
    
    def shutdown(self):
        """關閉排隊系統"""
        self.shutdown_flag = True
        with self.lock:
            self.queue.clear()
        logger.info("🔴 BTC通知排隊系統已關閉")

# 全局排隊管理器實例
btc_notification_manager = BTCNotificationQueue()

def send_btc_order_submit_notification(trade_record, success=True):
    """發送BTC訂單提交通知 - 新格式"""
    try:
        current_date = datetime.now().strftime('%Y/%m/%d')
        
        # 提取基本信息
        symbol = trade_record.get('symbol', 'BTCUSDT')
        side = trade_record.get('side', '')
        quantity = trade_record.get('quantity', 0)
        order_id = trade_record.get('order_id', '' if success else '未知')
        price = trade_record.get('price', 0)
        order_type = trade_record.get('order_type', 'MARKET')
        is_manual = trade_record.get('is_manual', False)
        reduce_only = trade_record.get('reduceOnly', False)
        
        # 從Binance持倉信息獲取真實數據
        position_data = get_btc_position_data_for_notification(symbol)
        trading_pair = position_data.get('symbol', 'BTCUSDT')
        contract_type = position_data.get('contract_type', '永續')
        margin_mode = position_data.get('margin_mode', '全倉')
        leverage = position_data.get('leverage', '10')
        
        # 判斷交易動作和方向（按照你的邏輯）
        # new buy = 開倉多單, new sell = 開倉空單, cover sell = 平倉多單, cover buy = 平倉空單
        action_type, direction_display = get_btc_action_and_direction(trade_record)
        
        # 判斷是否為手動交易（來自webhook是自動，其他都是手動）
        source = trade_record.get('source', 'manual')
        # 移除手動/自動標示
        submit_type = ""
        
        # 交易類別 - 統一使用轉換函數確保中文顯示
        order_type_display = convert_order_type_to_display(order_type)
        
        # 判斷是否為平倉操作
        is_close_action = reduce_only or 'cover' in action_type.lower()
        
        # 🔥 修復：對於平倉操作，使用實際持倉信息而不是訂單信息
        if is_close_action:
            # 獲取實際持倉數量和開倉價格
            actual_position_amt, actual_entry_price = get_btc_actual_position_info(symbol)
            
            if actual_position_amt > 0 and actual_entry_price > 0:
                # 使用實際持倉信息
                btc_quantity = actual_position_amt
                entry_price = actual_entry_price
                position_usdt = btc_quantity * entry_price
                
                logger.info(f"🔧 使用實際持倉信息 - BTC數量: {btc_quantity:.8f}, 開倉價格: {entry_price:.2f}")
            else:
                # 回退到訂單信息
                btc_quantity = float(quantity)
                entry_price = position_data.get('entry_price', 0)
                position_usdt = btc_quantity * entry_price if entry_price > 0 else 0
                logger.warning(f"⚠️ 無法獲取實際持倉信息，使用訂單信息")
        else:
            # 開倉操作：使用訂單信息
            btc_quantity = float(quantity)
            entry_price = position_data.get('entry_price', 0)
            position_usdt = btc_quantity * entry_price if entry_price > 0 else 0
        
        liquidation_price = position_data.get('liquidation_price', 0)
        
        if is_close_action:
            # 平倉操作：顯示實際持倉信息
            # 持倉數量(USDT) = 實際BTC數量 * 實際開倉價格
            position_usdt_display = f"{position_usdt:,.2f}"
            
            # 平倉時嘗試獲取當前強平價格，無法獲取則顯示N/A
            try:
                position_info = binance_client.get_position_risk(symbol='BTCUSDT')
                real_liquidation_price = 0
                
                if position_info:
                    for pos in position_info:
                        if pos['symbol'] == 'BTCUSDT' and float(pos['positionAmt']) != 0:
                            real_liquidation_price = float(pos.get('liquidationPrice', 0))
                            break
                
                if real_liquidation_price > 0:
                    liquidation_price_display = f"{real_liquidation_price:,.2f}"
                else:
                    liquidation_price_display = "N/A"
                    
            except Exception as e:
                logger.error(f"平倉時獲取強平價格失敗: {e}")
                liquidation_price_display = "N/A"
        else:
            # 開倉操作：計算持倉數量(USDT)
            submitted_price = trade_record.get('price', 0)
            if submitted_price > 0:
                position_usdt_display = f"{btc_quantity * submitted_price:,.2f}"
            else:
                position_usdt_display = f"{position_usdt:,.2f}" if position_usdt > 0 else "N/A"
            
            # 獲取強平價格
            try:
                position_info = binance_client.get_position_risk(symbol='BTCUSDT')
                real_liquidation_price = 0
                
                if position_info:
                    for pos in position_info:
                        if pos['symbol'] == 'BTCUSDT' and float(pos['positionAmt']) != 0:
                            real_liquidation_price = float(pos.get('liquidationPrice', 0))
                            break
                
                liquidation_price_display = f"{real_liquidation_price:,.2f}" if real_liquidation_price > 0 else "N/A"
                        
            except Exception as e:
                logger.error(f"獲取強平價格失敗: {e}")
                liquidation_price_display = "N/A"
        
        price_label = "平倉價格(USDT)" if is_close_action else "開倉價格(USDT)"
        
        # 🔥 新增：獲取止盈止損信息
        tp_sl_display = ""
        take_profit_price = None
        stop_loss_price = None
        
        # 首先檢查訂單記錄中是否包含止盈止損價格（適用於取消的訂單）
        if reduce_only and 'stopPrice' in trade_record:
            stop_price = float(trade_record.get('stopPrice', 0))
            if stop_price > 0:
                # 從訂單本身判斷是止盈還是止損
                side = trade_record.get('side', '')
                try:
                    current_price = float(binance_client.get_mark_price(symbol).get('markPrice', 0)) if binance_client else 0
                    
                    if side == 'SELL':  # 多頭持倉的平倉訂單
                        if stop_price > current_price:
                            take_profit_price = stop_price  # 止盈
                        else:
                            stop_loss_price = stop_price   # 止損
                    elif side == 'BUY':  # 空頭持倉的平倉訂單
                        if stop_price < current_price:
                            take_profit_price = stop_price  # 止盈
                        else:
                            stop_loss_price = stop_price   # 止損
                    
                    logger.info(f"🎯 從訂單記錄獲取止盈止損: stopPrice={stop_price}, side={side}, 判定為={'止盈' if take_profit_price else '止損'}")
                except Exception as e:
                    logger.error(f"判斷止盈止損類型失敗: {e}")
                    # 如果無法判斷，預設為止盈
                    take_profit_price = stop_price
        
        # 如果從訂單記錄中沒有獲取到，再查詢現有的開放訂單
        if take_profit_price is None and stop_loss_price is None:
            try:
                take_profit_price, stop_loss_price = get_btc_tp_sl_orders(symbol)
                logger.info(f"📊 從開放訂單獲取止盈止損信息")
            except Exception as e:
                logger.error(f"獲取開放訂單止盈止損信息失敗: {e}")
        
        # 🔥 統一邏輯：確保成功和失敗都有相同的止盈止損顯示
        # 如果是平倉操作（reduceOnly=True），無論是否獲取到止盈止損，都顯示查詢結果
        if reduce_only:
            tp_text = f"{take_profit_price:,.2f}" if take_profit_price is not None else "--"
            sl_text = f"{stop_loss_price:,.2f}" if stop_loss_price is not None else "--"
            tp_sl_display = f"\n止盈/止損：{tp_text}/{sl_text}"
            logger.info(f"📊 平倉操作止盈止損顯示: 止盈={tp_text}, 止損={sl_text}")
        else:
            # 開倉操作：只有獲取到止盈止損時才顯示
            if take_profit_price is not None or stop_loss_price is not None:
                tp_text = f"{take_profit_price:,.2f}" if take_profit_price else "--"
                sl_text = f"{stop_loss_price:,.2f}" if stop_loss_price else "--"
                tp_sl_display = f"\n止盈/止損：{tp_text}/{sl_text}"
                logger.info(f"📊 開倉操作止盈止損顯示: 止盈={tp_text}, 止損={sl_text}")
        
        # 🔥 修正：設置平倉價格顯示邏輯（成功和失敗都使用同樣邏輯）
        if is_close_action:
            # 對於平倉操作，優先使用止盈止損價格
            if reduce_only and (take_profit_price is not None or stop_loss_price is not None):
                # 如果有止盈價格，優先顯示止盈價格作為平倉價格
                if take_profit_price is not None:
                    entry_price_display = f"{take_profit_price:,.2f}"
                    logger.info(f"🎯 使用止盈價格作為平倉價格: {take_profit_price:,.2f}")
                else:
                    entry_price_display = f"{stop_loss_price:,.2f}"
                    logger.info(f"🛑 使用止損價格作為平倉價格: {stop_loss_price:,.2f}")
            else:
                # 沒有止盈止損，使用提交的限價
                submitted_price = trade_record.get('price', 0)
                if order_type_display == "市價單":
                    entry_price_display = "市價"
                elif submitted_price > 0:
                    entry_price_display = f"{submitted_price:,.2f}"
                else:
                    entry_price_display = "限價"
        else:
            # 開倉操作的價格顯示邏輯
            submitted_price = trade_record.get('price', 0)
            if order_type_display == "市價單":
                entry_price_display = "市價"
            elif submitted_price > 0:
                entry_price_display = f"{submitted_price:,.2f}"
            else:
                entry_price_display = "限價"
        
        if success:
            msg = (f"⭕ 提交成功（{current_date}）\n"
                   f"{trading_pair}｜{contract_type}｜{margin_mode}{leverage}X\n"
                   f"交易訂單：{order_id}\n"
                   f"交易類別：{order_type_display}\n"
                   f"交易動作：{action_type}\n"
                   f"交易方向：{direction_display}\n"
                   f"交易數量：{btc_quantity:.8f} BTC\n"
                   f"持倉數量(USDT)：{position_usdt_display}\n"
                   f"{price_label}：{entry_price_display}\n"
                   f"強平價格(USDT)：{liquidation_price_display}{tp_sl_display}")
        else:
            error = trade_record.get('error', '未知錯誤')
            msg = (f"❌ 提交失敗（{current_date}）\n"
                   f"{trading_pair}｜{contract_type}｜{margin_mode}{leverage}X\n"
                   f"交易訂單：{order_id}\n"
                   f"交易類別：{order_type_display}\n"
                   f"交易動作：{action_type}\n"
                   f"交易方向：{direction_display}\n"
                   f"交易數量：{btc_quantity:.8f} BTC\n"
                   f"持倉數量(USDT)：{position_usdt_display}\n"
                   f"{price_label}：{entry_price_display}\n"
                   f"強平價格(USDT)：{liquidation_price_display}{tp_sl_display}\n"
                   f"原因：{error}")
        
        # 先添加前端日誌記錄（包含交易信息）
        # 為前端日誌添加USDT單位（除非是"市價"或"限價"文字）
        frontend_price_display = entry_price_display
        if entry_price_display not in ["市價", "限價"] and not entry_price_display.endswith("USDT"):
            frontend_price_display = f"{entry_price_display} USDT"
        
        # 為前端日誌格式化止盈止損信息
        frontend_tp_sl_display = ""
        if tp_sl_display:
            # 移除換行符，用｜分隔
            frontend_tp_sl_display = f"｜{tp_sl_display.strip()}"
        
        if success:
            # 提交成功時，前端日誌記錄交易詳情
            frontend_log = f"{submit_type}{action_type}：{direction_display}｜{btc_quantity:.8f} BTC｜{frontend_price_display}｜{order_type_display}{frontend_tp_sl_display}"
            frontend_result = log_btc_frontend_message(frontend_log, "success")
            logger.info(f"BTC前端日誌發送結果: {frontend_result}")
        else:
            # 提交失敗時，前端日誌記錄交易詳情和失敗原因（紅色顯示）
            error_reason = trade_record.get('error', '未知錯誤')
            frontend_log = f"{submit_type}{action_type}：{direction_display}｜{btc_quantity:.8f} BTC｜{frontend_price_display}｜{order_type_display}｜失敗原因：{error_reason}{frontend_tp_sl_display}"
            frontend_result = log_btc_frontend_message(frontend_log, "error")
            logger.info(f"BTC前端日誌發送結果: {frontend_result}")
        
        # 然後發送 Telegram 訊息
        logger.info(f"準備發送BTC Telegram訊息: {msg[:100]}...")
        result = send_btc_telegram_message(msg)
        logger.info(f"BTC Telegram發送結果: {result}")
        
        # 添加到系統日誌
        logger.info(f"BTC訂單通知: {msg}")
        
        return result
        
    except Exception as e:
        logger.error(f"發送BTC訂單提交通知失敗: {e}")

def get_btc_position_data_for_notification(symbol):
    """獲取BTC持倉數據用於通知格式"""
    try:
        if not binance_client:
            return {
                'symbol': 'BTCUSDT',
                'contract_type': '永續',
                'margin_mode': '全倉',
                'leverage': '10',
                'entry_price': 0,
                'liquidation_price': 0
            }
        
        # 獲取持倉信息
        positions = binance_client.get_position_info()
        if positions:
            for pos in positions:
                if pos.get('symbol') == symbol:
                    # 獲取保證金模式
                    margin_type = pos.get('marginType', 'cross')
                    margin_mode = '全倉' if margin_type.lower() == 'cross' else '逐倉'
                    
                    return {
                        'symbol': pos.get('symbol', 'BTCUSDT'),
                        'contract_type': '永續',  # Binance期貨都是永續合約
                        'margin_mode': margin_mode,
                        'leverage': str(int(float(pos.get('leverage', 10)))),
                        'entry_price': float(pos.get('entryPrice', 0)),
                        'liquidation_price': float(pos.get('liquidationPrice', 0))
                    }
        
        # 如果沒有持倉，使用預設值
        return {
            'symbol': 'BTCUSDT',
            'contract_type': '永續',
            'margin_mode': '全倉',
            'leverage': '10',
            'entry_price': 0,
            'liquidation_price': 0
        }
        
    except Exception as e:
        logger.error(f"獲取BTC持倉數據失敗: {e}")
        return {
            'symbol': 'BTCUSDT',
            'contract_type': '永續',
            'margin_mode': '全倉',
            'leverage': '10',
            'entry_price': 0,
            'liquidation_price': 0
        }

def get_btc_tp_sl_orders(symbol='BTCUSDT'):
    """獲取BTC止盈止損訂單信息"""
    try:
        if not binance_client:
            return None, None
        
        # 獲取所有未成交訂單
        open_orders = binance_client.get_open_orders(symbol)
        if not open_orders:
            return None, None
        
        take_profit_price = None
        stop_loss_price = None
        
        for order in open_orders:
            # 檢查是否為止盈止損訂單（reduceOnly=True）
            if order.get('reduceOnly', False):
                order_type = order.get('type', '')
                stop_price = order.get('stopPrice', 0)
                
                # 判斷是止盈還是止損
                # 通常止盈止損訂單會有stopPrice
                if stop_price and float(stop_price) > 0:
                    side = order.get('side', '')
                    
                    # 根據訂單方向判斷是止盈還是止損
                    # 多頭持倉：賣出止盈（價格高於當前）、賣出止損（價格低於當前）
                    # 空頭持倉：買入止盈（價格低於當前）、買入止損（價格高於當前）
                    
                    # 獲取當前標記價格作為參考
                    try:
                        current_price = float(binance_client.get_mark_price(symbol).get('markPrice', 0))
                        stop_price_float = float(stop_price)
                        
                        if side == 'SELL':  # 多頭持倉的平倉訂單
                            if stop_price_float > current_price:
                                take_profit_price = stop_price_float  # 止盈
                            else:
                                stop_loss_price = stop_price_float   # 止損
                        elif side == 'BUY':  # 空頭持倉的平倉訂單
                            if stop_price_float < current_price:
                                take_profit_price = stop_price_float  # 止盈
                            else:
                                stop_loss_price = stop_price_float   # 止損
                    except:
                        # 如果無法判斷，根據訂單類型猜測
                        if 'TAKE_PROFIT' in order_type:
                            take_profit_price = float(stop_price)
                        elif 'STOP' in order_type:
                            stop_loss_price = float(stop_price)
        
        return take_profit_price, stop_loss_price
        
    except Exception as e:
        logger.error(f"獲取止盈止損訂單失敗: {e}")
        return None, None

def convert_order_type_to_display(order_type):
    """統一轉換訂單類型格式為中文顯示格式（強化版）"""
    if not order_type:
        return '市價單'
    
    # 轉換為字符串並轉為大寫進行比較
    order_type_str = str(order_type).upper().strip()
    
    # 統一轉換邏輯 - 包含所有可能的格式
    market_types = ['MARKET', 'MKT', '市價單', '市價', 'IOC', 'FOK']
    limit_types = ['LIMIT', 'LMT', '限價單', '限價', 'ROD', 'GTD', 'GTC']
    
    if order_type_str in market_types:
        return '市價單'
    elif order_type_str in limit_types:
        return '限價單'
    else:
        # 如果是其他格式，嘗試智能判斷
        if any(market_key in order_type_str for market_key in ['MARKET', 'MKT', '市價']):
            return '市價單'
        else:
            return '限價單'

def get_btc_actual_position_info(symbol='BTCUSDT'):
    """獲取BTC實際持倉信息（數量和開倉價格）"""
    try:
        if not binance_client:
            return 0, 0
        
        # 獲取持倉信息
        positions = binance_client.get_position_info()
        if positions:
            for pos in positions:
                if pos.get('symbol') == symbol:
                    position_amt = float(pos.get('positionAmt', 0))
                    entry_price = float(pos.get('entryPrice', 0))
                    
                    # 返回絕對值的持倉數量和開倉價格
                    return abs(position_amt), entry_price
        
        return 0, 0
        
    except Exception as e:
        logger.error(f"獲取BTC實際持倉信息失敗: {e}")
        return 0, 0

def send_btc_order_modify_notification(order_data, old_data, new_data):
    """發送BTC訂單修改通知"""
    try:
        current_date = datetime.now().strftime('%Y/%m/%d')
        
        # 提取基本信息
        symbol = order_data.get('s', 'BTCUSDT')
        side = order_data.get('S', '')
        order_id = order_data.get('i', '')
        order_type = order_data.get('o', 'LIMIT')
        
        # 判斷修改類型
        price_changed = old_data.get('price') != new_data.get('price')
        quantity_changed = old_data.get('quantity') != new_data.get('quantity')
        
        # 格式化變更信息
        changes = []
        if price_changed:
            changes.append(f"價格：{old_data.get('price')} → {new_data.get('price')}")
        if quantity_changed:
            changes.append(f"數量：{old_data.get('quantity')} → {new_data.get('quantity')}")
        
        change_info = "、".join(changes) if changes else "未知變更"
        
        # 判斷動作方向
        direction = "買入" if side == 'BUY' else "賣出"
        order_type_display = convert_order_type_to_display(order_type)
        
        # 構建通知消息
        msg = f"""⭕ 修改成功（{current_date}）
交易對：{symbol}
動作：{direction}
類型：{order_type_display}
修改內容：{change_info}
訂單號：{order_id}"""
        
        # 發送Telegram通知
        result = send_btc_telegram_message(msg)
        logger.info(f"BTC修改訂單Telegram通知發送結果: {result}")
        
        # 記錄前端日誌
        log_btc_frontend_message(f"⭕ 修改成功 - {direction} {order_type_display} - {change_info}", "success")
        logger.info(f"BTC訂單修改通知: {msg}")
        
        return result
        
    except Exception as e:
        logger.error(f"發送BTC訂單修改通知失敗: {e}")
        return False

def get_btc_action_and_direction(trade_record):
    """判斷BTC交易動作和方向"""
    side = trade_record.get('side', '').upper()
    reduce_only = trade_record.get('reduceOnly', False)
    action_type = trade_record.get('action_type', '')
    
    # 根據量化交易邏輯：
    # new buy = 開倉多單, new sell = 開倉空單, cover sell = 平倉多單, cover buy = 平倉空單
    
    logger.info(f"🔍 BTC方向判斷: side={side}, reduce_only={reduce_only}, action_type={action_type}")
    
    if reduce_only or 'cover' in action_type.lower():
        # 平倉
        action_type = "平倉"
        if side == 'SELL':
            direction_display = "多單"  # cover sell = 平倉多單
        else:  # BUY
            direction_display = "空單"  # cover buy = 平倉空單
    else:
        # 開倉
        action_type = "開倉"
        if side == 'BUY':
            direction_display = "多單"  # new buy = 開倉多單
        else:  # SELL
            direction_display = "空單"  # new sell = 開倉空單
    
    logger.info(f"🔍 BTC方向判斷結果: action_type={action_type}, direction_display={direction_display}")
    return action_type, direction_display

def send_btc_trade_notification(trade_record):
    """發送BTC成交通知 - 新格式"""
    try:
        current_date = datetime.now().strftime('%Y/%m/%d')
        
        # 提取基本信息
        symbol = trade_record.get('symbol', 'BTCUSDT')
        side = trade_record.get('side', '')
        quantity = trade_record.get('fill_quantity', trade_record.get('quantity', 0))
        fill_price = float(trade_record.get('fill_price', 0))
        order_id = trade_record.get('order_id', '')
        order_type = trade_record.get('order_type', 'MARKET')
        is_manual = trade_record.get('is_manual', False)
        reduce_only = trade_record.get('reduceOnly', False)
        
        # 從Binance持倉信息獲取真實數據
        position_data = get_btc_position_data_for_notification(symbol)
        trading_pair = position_data.get('symbol', 'BTCUSDT')
        contract_type = position_data.get('contract_type', '永續')
        margin_mode = position_data.get('margin_mode', '全倉')
        leverage = position_data.get('leverage', '10')
        
        # 判斷交易動作和方向（按照你的邏輯）
        action_type, direction_display = get_btc_action_and_direction(trade_record)
        
        # 判斷是否為手動交易（來自webhook是自動，其他都是手動）
        source = trade_record.get('source', 'manual')
        # 移除手動/自動標示
        submit_type = ""
        
        # 交易類別 - 統一使用轉換函數確保中文顯示
        order_type_display = convert_order_type_to_display(order_type)
        
        # 計算持倉數量(USDT) = BTC數量 * 成交價格
        btc_quantity = float(quantity)
        position_usdt = btc_quantity * fill_price if fill_price > 0 else 0
        entry_price = position_data.get('entry_price', fill_price)
        liquidation_price = position_data.get('liquidation_price', 0)
        
        # 格式化顯示
        position_usdt_display = f"{position_usdt:,.2f}" if position_usdt > 0 else "0.00"
        entry_price_display = f"{entry_price:,.2f}" if entry_price > 0 else f"{fill_price:,.2f}"
        liquidation_price_display = f"{liquidation_price:,.2f}" if liquidation_price > 0 else "0.00"
        
        msg = (f"✅ 成交通知（{current_date}）\n"
               f"{trading_pair}｜{contract_type}｜{margin_mode}{leverage}X\n"
               f"交易訂單：{order_id}\n"
               f"交易類別：{order_type_display}\n"
               f"交易動作：{action_type}\n"
               f"交易方向：{direction_display}\n"
               f"交易數量：{btc_quantity:.8f} BTC\n"
               f"持倉數量(USDT)：{position_usdt_display}\n"
               f"開倉價格(USDT)：{entry_price_display}\n"
               f"強平價格(USDT)：{liquidation_price_display}")
        
        send_btc_telegram_message(msg)
        
        # 添加到系統日誌
        logger.info(f"BTC成交通知: {msg}")
        
    except Exception as e:
        logger.error(f"發送BTC成交通知失敗: {e}")

def send_btc_trading_statistics():
    """發送BTC每日交易統計 - 2025新格式"""
    try:
        current_date = datetime.now().strftime('%Y/%m/%d')
        
        # 1. 獲取今日前端日誌統計（從前端日誌計算真實數量）
        total_orders = get_btc_message_count_today('提交成功')  # 委託次數：提交成功訊息則數
        total_cancels = get_btc_message_count_today('提交失敗')    # 取消次數：提交失敗訊息則數
        total_deals = get_btc_message_count_today('成交通知')     # 成交次數：成交通知訊息則數
        
        # 2. 從Binance API計算交易統計（與XLSX報表使用相同邏輯和數據源）
        binance_stats = {
            'order_count': 0,       # 委託次數
            'cancel_count': 0,      # 取消次數  
            'fill_count': 0,        # 成交次數
            'buy_volume': 0.0,      # 買入總量(USDT)
            'sell_volume': 0.0,     # 賣出總量(USDT) 
            'avg_buy_price': 0.0,   # 平均買價
            'avg_sell_price': 0.0   # 平均賣價
        }
        
        if binance_client:
            try:
                # 獲取今日交易記錄（與XLSX報表使用相同時間範圍）
                today = datetime.now()
                today_start = int(today.replace(hour=0, minute=0, second=0, microsecond=0).timestamp() * 1000)
                today_end = int(today.timestamp() * 1000)
                
                # 獲取所有訂單數據（用於統計委託次數、取消次數、成交次數）
                orders_data = binance_client._make_request('GET', '/fapi/v1/allOrders', {
                    'symbol': 'BTCUSDT',
                    'startTime': today_start,
                    'endTime': today_end
                })
                
                if orders_data:
                    for order in orders_data:
                        status = order.get('status', '')
                        if status in ['NEW', 'PARTIALLY_FILLED', 'FILLED', 'CANCELED', 'REJECTED', 'EXPIRED']:
                            binance_stats['order_count'] += 1  # 所有狀態都算委託次數
                        if status in ['CANCELED', 'EXPIRED']:
                            binance_stats['cancel_count'] += 1  # 取消和過期算取消次數
                        if status in ['FILLED', 'PARTIALLY_FILLED']:
                            binance_stats['fill_count'] += 1  # 成交和部分成交算成交次數
                
                # 獲取今日所有交易記錄（用於統計買賣總量和平均價格）
                trades_data = binance_client._make_request('GET', '/fapi/v1/userTrades', {
                    'symbol': 'BTCUSDT',
                    'startTime': today_start,
                    'endTime': today_end
                })
                
                if trades_data:
                    # 根據交易方向統計（方式一）
                    buy_trades = []   # 所有BUY方向交易
                    sell_trades = []  # 所有SELL方向交易
                    
                    for trade in trades_data:
                        qty = float(trade.get('qty', 0))
                        price = float(trade.get('price', 0))
                        side = trade.get('side', '')
                        trade_value = qty * price
                        
                        if side == 'BUY':
                            # 所有買入交易
                            buy_trades.append({'qty': qty, 'price': price, 'usdt': trade_value})
                            binance_stats['buy_volume'] += trade_value
                        elif side == 'SELL':
                            # 所有賣出交易
                            sell_trades.append({'qty': qty, 'price': price, 'usdt': trade_value})
                            binance_stats['sell_volume'] += trade_value
                    
                    # 計算平均價格
                    if buy_trades:
                        total_buy_qty = sum(t['qty'] for t in buy_trades)
                        total_buy_value = sum(t['usdt'] for t in buy_trades)
                        binance_stats['avg_buy_price'] = total_buy_value / total_buy_qty if total_buy_qty > 0 else 0
                    
                    if sell_trades:
                        total_sell_qty = sum(t['qty'] for t in sell_trades)
                        total_sell_value = sum(t['usdt'] for t in sell_trades)
                        binance_stats['avg_sell_price'] = total_sell_value / total_sell_qty if total_sell_qty > 0 else 0
                        
            except Exception as e:
                logger.error(f"獲取Binance交易數據失敗: {e}")
        
        # 3. 獲取帳戶狀態數據 - 與啟動通知使用相同的數據獲取方式
        account_data = {
            'totalWalletBalance': 0.0,
            'availableBalance': 0.0,
            'totalMarginBalance': 0.0,
            'totalUnrealizedProfit': 0.0,
            'feePaid': 0.0,
            'marginRatio': 0.0,
            'leverageUsage': 0.0,
            'todayPnl': 0.0,
            'week7Pnl': 0.0,
            'month30Pnl': 0.0
        }
        
        # 獲取真實帳戶數據（使用與啟動通知相同的方式）
        if binance_client:
            try:
                # 獲取帳戶基本信息
                account_info = binance_client.get_account_info()
                if account_info:
                    account_data['totalWalletBalance'] = float(account_info.get('totalWalletBalance', 0))
                    account_data['availableBalance'] = float(account_info.get('availableBalance', 0))
                    account_data['totalMarginBalance'] = float(account_info.get('totalMarginBalance', 0))
                    account_data['totalUnrealizedProfit'] = float(account_info.get('totalUnrealizedProfit', 0))
                    
                    # 計算保證金比率和槓桿使用率
                    maintenance_margin = float(account_info.get('totalMaintMargin', 0))
                    initial_margin = float(account_info.get('totalInitialMargin', 0))
                    
                    if maintenance_margin > 0:
                        account_data['marginRatio'] = (account_data['totalMarginBalance'] / maintenance_margin) * 100
                    
                    # 使用與前端一致的槓桿使用率計算邏輯
                    available_balance = float(account_info.get('availableBalance', 0))
                    unrealized_pnl = float(account_info.get('totalUnrealizedProfit', 0))
                    calculated_wallet_balance = available_balance + initial_margin + unrealized_pnl
                    if calculated_wallet_balance > 0:
                        account_data['leverageUsage'] = (initial_margin / calculated_wallet_balance) * 100
                
                # 獲取盈虧數據
                today = datetime.now()
                
                # 今日盈虧
                today_start = int(today.replace(hour=0, minute=0, second=0, microsecond=0).timestamp() * 1000)
                today_end = int(today.timestamp() * 1000)
                
                today_income = binance_client._make_request('GET', '/fapi/v1/income', {
                    'incomeType': 'REALIZED_PNL',
                    'startTime': today_start,
                    'endTime': today_end
                })
                if today_income:
                    account_data['todayPnl'] = sum(float(item.get('income', 0)) for item in today_income)
                
                # 7天盈虧
                week7_start = int((today - timedelta(days=7)).timestamp() * 1000)
                week7_income = binance_client._make_request('GET', '/fapi/v1/income', {
                    'incomeType': 'REALIZED_PNL',
                    'startTime': week7_start,
                    'endTime': today_end
                })
                if week7_income:
                    account_data['week7Pnl'] = sum(float(item.get('income', 0)) for item in week7_income)
                
                # 30天盈虧
                month30_start = int((today - timedelta(days=30)).timestamp() * 1000)
                month30_income = binance_client._make_request('GET', '/fapi/v1/income', {
                    'incomeType': 'REALIZED_PNL',
                    'startTime': month30_start,
                    'endTime': today_end
                })
                if month30_income:
                    account_data['month30Pnl'] = sum(float(item.get('income', 0)) for item in month30_income)
                    
            except Exception as e:
                logger.error(f"獲取BTC帳戶數據失敗: {e}")
        
        # 4. 獲取交易明細（使用JSON配對系統的真實數據）
        trade_details = ""
        try:
            from trade_pairing_BTC import get_btc_cover_trades_for_report
            from trading_config import TradingConfig
            
            # 獲取今日平倉交易明細
            today_date_str = datetime.now().strftime('%Y-%m-%d')
            cover_trades = get_btc_cover_trades_for_report(date_range_days=1)
            
            if cover_trades:
                for trade in cover_trades:
                    # 解析交易時間，確保是今日交易
                    trade_time = datetime.fromisoformat(trade['cover_timestamp'].replace('Z', '+00:00'))
                    if trade_time.date() == datetime.now().date():
                        
                        # 確定多空方向
                        direction = "多單" if trade['cover_action'] == 'SELL' else "空單"
                        
                        # 格式化數量
                        quantity_str = TradingConfig.format_btc_quantity(trade['matched_quantity'])
                        
                        # 計算成交金額
                        open_usdt = trade['open_price'] * trade['matched_quantity']
                        cover_usdt = trade['cover_price'] * trade['matched_quantity']
                        
                        # 構建交易明細字符串
                        trade_details += f"{direction}｜{quantity_str}BTC｜"
                        trade_details += f"{trade['open_price']:,.2f} USDT｜"
                        trade_details += f"{trade['cover_price']:,.2f} USDT\n"
                        trade_details += f"盈虧{trade['pnl']:+.2f} USDT\n"
                        
                        logger.info(f"📊 BTC交易明細: {direction} {quantity_str}BTC 損益${trade['pnl']:+.2f}")
                
                if not trade_details:
                    trade_details = "❌ 無平倉交易"
            else:
                trade_details = "❌ 無平倉交易"
                
        except ImportError:
            logger.warning("BTC配對系統模組未找到，回退到原有邏輯")
            # 回退到原有邏輯（但移除硬編碼）
            if binance_client:
                try:
                    income_data = binance_client._make_request('GET', '/fapi/v1/income', {
                        'incomeType': 'REALIZED_PNL',
                        'startTime': today_start,
                        'endTime': today_end
                    })
                    
                    if income_data and len(income_data) > 0:
                        for income in income_data:
                            pnl = float(income.get('income', 0))
                            if abs(pnl) > 0.01:
                                # 使用API數據但不硬編碼
                                trade_details += f"平倉交易｜API數據｜損益: ${pnl:+.2f} USDT\n"
                                
                    if not trade_details:
                        trade_details = "❌ 無平倉交易"
                        
                except Exception as e:
                    logger.error(f"獲取交易明細失敗: {e}")
                    trade_details = "❌ 無平倉交易"
            else:
                trade_details = "❌ 無平倉交易"
                
        except Exception as e:
            logger.error(f"使用BTC配對系統獲取交易明細失敗: {e}")
            trade_details = "❌ 無平倉交易"
        
        # 5. 獲取持倉狀態
        position_info = ""
        if binance_client:
            try:
                positions = binance_client.get_position_info()
                active_positions = [pos for pos in positions if float(pos.get('positionAmt', 0)) != 0]
                
                if not active_positions:
                    position_info = "❌ 無持倉部位"
                else:
                    for pos in active_positions:
                        position_amt = float(pos.get('positionAmt', 0))
                        entry_price = float(pos.get('entryPrice', 0))
                        unrealized_pnl = float(pos.get('unRealizedProfit', 0))
                        
                        # 判斷持倉方向
                        direction = "多單" if position_amt > 0 else "空單"
                        
                        position_info += f"{direction}｜{abs(position_amt):.8f} BTC｜{entry_price:,.2f} USDT\n未實現盈虧{unrealized_pnl:+.2f} USDT"
                        
            except Exception as e:
                logger.error(f"獲取BTC持倉信息失敗: {e}")
                position_info = "❌ 無持倉部位"
        else:
            position_info = "❌ 無持倉部位"
        
        # 6. 構建統計訊息（與XLSX報表一致）
        message = f"📊 交易統計（{current_date}）\n"
        message += "═════ 交易總覽 ═════\n"
        message += f"委託次數：{binance_stats['order_count']} 筆\n"
        message += f"取消次數：{binance_stats['cancel_count']} 筆\n"
        message += f"成交次數：{binance_stats['fill_count']} 筆\n"
        message += f"買入總量：{binance_stats['buy_volume']:.2f} USDT\n"
        message += f"賣出總量：{binance_stats['sell_volume']:.2f} USDT\n"
        message += f"平均買價：{binance_stats['avg_buy_price']:.2f} USDT\n"
        message += f"平均賣價：{binance_stats['avg_sell_price']:.2f} USDT\n"
        message += "═════ 帳戶狀態 ═════\n"
        message += f"錢包餘額：{account_data['totalWalletBalance']:.8f} USDT\n"
        message += f"可供轉帳：{account_data['availableBalance']:.8f} USDT\n"
        message += f"保證金餘額：{account_data['totalMarginBalance']:.8f} USDT\n"
        message += f"未實現盈虧：{account_data['totalUnrealizedProfit']:+.8f} USDT\n"
        message += f"交易手續費：{account_data['feePaid']:.8f} USDT\n"
        message += f"保證金比率：{account_data['marginRatio']:.2f}%\n"
        message += f"槓桿使用率：{account_data['leverageUsage']:.2f}%\n"
        message += f"本日盈虧：{account_data['todayPnl']:+.8f} USDT\n"
        message += f"7 天盈虧：{account_data['week7Pnl']:+.8f} USDT\n"
        message += f"30天盈虧：{account_data['month30Pnl']:+.8f} USDT\n"
        message += "═════ 交易明細 ═════\n"
        message += trade_details
        message += "═════ 持倉狀態 ═════\n"
        message += position_info
        
        # 7. 發送 Telegram 訊息
        send_btc_telegram_message(message)
        
        # 8. 記錄後端日誌（前端日誌已在send_btc_telegram_message中記錄）
        logger.info(f"BTC交易統計已發送: {current_date}")
        
    except Exception as e:
        logger.error(f"發送BTC交易統計失敗: {e}")
        log_btc_frontend_message(f"Telegram［交易統計］訊息發送失敗: {str(e)}", "error")

def check_btc_daily_trading_statistics():
    """檢查是否需要發送BTC每日交易統計 - BTC 24/7無交易日限制"""
    try:
        logger.info("開始檢查BTC每日交易統計...")
        # BTC 24/7交易，直接發送統計
        send_btc_trading_statistics()
        logger.info("BTC每日交易統計發送完成")
        
        # 延遲生成報表 - 統一時間控制：23:58:30日報，23:59:00月報
        def delayed_generate_btc_reports():
            # 等待30秒後生成日報 (23:58:30)，可中斷睡眠
            for _ in range(30):
                if btc_shutdown_flag.is_set():
                    logger.info("BTC模組關閉中，取消報表生成")
                    return
                time.sleep(1)
                
            logger.info("開始生成BTC交易報表...")
            daily_report_result = generate_btc_daily_report() if not btc_shutdown_flag.is_set() else False
            
            # 如果是月末且日報生成成功，再等待30秒後生成月報 (23:59:00)
            if daily_report_result and is_last_day_of_month() and not btc_shutdown_flag.is_set():
                for _ in range(30):
                    if btc_shutdown_flag.is_set():
                        logger.info("BTC模組關閉中，取消月報生成")
                        return
                    time.sleep(1)
                    
                logger.info("月末檢測，開始生成BTC交易報表...")
                year = datetime.now().year
                month = datetime.now().month
                monthly_result = generate_btc_monthly_report(year, month)
                if monthly_result and monthly_result.get('success'):
                    logger.info(f"📅 BTC月報自動生成並發送成功: {monthly_result.get('filename')}")
                else:
                    logger.error("❌ BTC月報自動生成失敗")
        
        # 在新線程中執行延遲生成報表
        report_thread = threading.Thread(target=delayed_generate_btc_reports, daemon=True, name="BTC報表生成")
        register_btc_thread(report_thread, "BTC報表生成")
        report_thread.start()
        
    except Exception as e:
        logger.error(f"檢查BTC每日交易統計失敗: {e}")


def is_last_day_of_month():
    """檢查是否為月末最後一天"""
    try:
        today = datetime.now()
        tomorrow = today + timedelta(days=1)
        return today.month != tomorrow.month
    except Exception as e:
        logger.error(f"檢查月末日期失敗: {e}")
        return False

def btc_webhook():
    """BTC交易策略接收端點"""
    global btc_processed_signals
    
    try:
        # 優先使用預處理的數據，避免 Content-Type 問題
        from flask import g
        if hasattr(g, 'webhook_data') and g.webhook_data:
            data = g.webhook_data
        else:
            data = request.get_json(force=True)  # 強制解析，忽略 Content-Type
        
        # 驗證請求格式
        if not data:
            return jsonify({'success': False, 'message': '無效的請求格式'})
        
        # 檢查訊號去重
        signal_id = data.get('tradeId') or data.get('id') or f"{data.get('action', '')}-{data.get('symbol', 'BTCUSDT')}-{int(datetime.now().timestamp())}"
        
        # 生成更可靠的訊號ID（使用實際有的字段）
        import hashlib
        signal_content = f"{data.get('action', '')}-{data.get('id', '')}-{data.get('mp', '')}-{data.get('nOrder', '')}-{data.get('price', '')}"
        signal_hash = hashlib.md5(signal_content.encode()).hexdigest()[:8]
        
        # 檢查最近30秒內是否有相同訊號（更嚴格的去重，精確到秒級）
        current_time = datetime.now()
        signal_key = f"{signal_hash}-{current_time.strftime('%Y%m%d%H%M%S')}"  # 精確到秒
        
        with btc_signals_lock:
            if signal_key in btc_processed_signals:
                logger.info(f"BTC重複訊號已跳過: {signal_key}")
                return jsonify({'success': True, 'message': '重複訊號已跳過'})
            
            # 記錄訊號
            btc_processed_signals.add(signal_key)
            
            # 清理超過30秒的舊訊號記錄（秒級精度去重）
            cutoff_time = (current_time - timedelta(seconds=30)).strftime('%Y%m%d%H%M%S')
            btc_processed_signals = {s for s in btc_processed_signals if s.split('-')[-1] >= cutoff_time}
        
        # 記錄接收到的策略信號
        timestamp = datetime.now().isoformat()
        
        # 載入BTC配置
        if not os.path.exists(BTC_ENV_PATH):
            return jsonify({'success': False, 'message': 'BTC配置不存在'})
        
        btc_env = {}
        with open(BTC_ENV_PATH, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    btc_env[key] = value
        
        
        # 處理策略信號並執行交易
        action = data.get('action', '').upper()
        
        # 增強TradingView JSON格式支援
        if not action:
            # 嘗試從direction字段獲取（TradingView常用格式）
            direction = data.get('direction', '')
            if direction:
                action = direction.upper()
        
        # 支援中文訊號解析（TradingView策略常用）
        action_mapping = {
            # 英文訊號
            'LONG': 'LONG', 'SHORT': 'SHORT', 'BUY': 'LONG', 'SELL': 'SHORT',
            'CLOSE': 'CLOSE', 'EXIT': 'CLOSE', 'CLOSE_LONG': 'CLOSE_LONG', 'CLOSE_SHORT': 'CLOSE_SHORT',
            # 數字訊號支援
            'signal: +1': 'LONG', 'signal: -1': 'SHORT', 'signal: 0': 'CLOSE',
            '+1': 'LONG', '-1': 'SHORT', '0': 'CLOSE',
            # 中文訊號支援（包含你的Pine策略格式）
            '開多': 'LONG', '開多單': 'LONG', '多單': 'LONG', 'BTC 開多單': 'LONG',
            '開空': 'SHORT', '開空單': 'SHORT', '空單': 'SHORT', 'BTC 開空單': 'SHORT',
            '平多': 'CLOSE_LONG', '平多單': 'CLOSE_LONG', '多單平倉': 'CLOSE_LONG', '平多倉': 'CLOSE_LONG', 'BTC 平多單': 'CLOSE_LONG',
            '平空': 'CLOSE_SHORT', '平空單': 'CLOSE_SHORT', '空單平倉': 'CLOSE_SHORT', '平空倉': 'CLOSE_SHORT', 'BTC 平空單': 'CLOSE_SHORT',
            '平倉': 'CLOSE', '全平': 'CLOSE'
        }
        
        # 嘗試解析action，支援完整訊息解析和TradingView格式
        original_action = action
        
        # 支援多種訊息來源
        message_sources = [
            data.get('message', ''),
            data.get('action', ''),
            data.get('direction', ''),
            str(data.get('signal', '')),  # 數字訊號支援
            str(data.get('type', ''))     # type字段支援
        ]
        
        raw_message = ' '.join(filter(None, map(str, message_sources)))
        
        
        # 特殊處理signal數字格式（你的Pine策略使用的格式）
        signal_value = data.get('signal', '')
        if signal_value and not action:
            if signal_value in ['+1', '1', 1]:
                action = 'LONG'
            elif signal_value in ['-1', -1]:
                action = 'SHORT' 
            elif signal_value in ['0', 0]:
                action = '平倉'
        
        # 檢查原始訊號
        if action in action_mapping:
            action = action_mapping[action]
        else:
            # 從完整訊息中解析中文訊號
            action_found = False
            for key, value in action_mapping.items():
                if key in raw_message:
                    action = value
                    action_found = True
                    break
            
            if not action_found:
                logger.warning(f"[BTC] 未能解析動作 - 原始action: '{original_action}', 訊息: '{raw_message}'")
                logger.warning(f"[BTC] 完整數據: {data}")
                log_btc_frontend_message(f"BTC動作解析失敗 - action: '{original_action}', message: '{raw_message}'", "error")
                action = 'UNKNOWN'
        
        # 執行交易
        order_result = None
        
        if action in ['LONG', 'SHORT']:
            # 進場信號
            order_result = process_btc_entry_signal(data, action)
        elif action in ['CLOSE', 'CLOSE_LONG', 'CLOSE_SHORT']:
            # 出場信號
            order_result = process_btc_exit_signal(data, action)
        else:
            # 動作類型未知，不執行交易
            logger.warning(f"BTC未知的動作類型，跳過處理: {original_action} (解析後: {action})")
            return None
        
        # 處理交易策略信號記錄
        strategy_data = {
            'timestamp': timestamp,
            'signal': data,
            'action': action,
            'processed': True,
            'order_result': order_result,
            'order_id': order_result.get('orderId') if order_result else None
        }
        
        
        # 發送Telegram通知（如果配置了）
        chat_id = btc_env.get('CHAT_ID_BTC')
        if chat_id:
            try:
                message = f"🔔 BTC交易信號\n"
                message += f"時間: {timestamp}\n"
                message += f"信號: {json.dumps(data, ensure_ascii=False, indent=2)}"
                
                # 這裡可以添加發送Telegram消息的邏輯
                logger.info(f"BTC策略信號: {message}")
                
            except Exception as e:
                logger.error(f"發送Telegram通知失敗: {e}")
        
        return jsonify({
            'success': True, 
            'message': 'BTC策略信號接收成功',
            'timestamp': timestamp
        })
        
    except Exception as e:
        logger.error(f"處理BTC策略信號失敗: {e}")
        return jsonify({'success': False, 'message': f'處理失敗: {str(e)}'})


def get_btc_strategy_status():
    """獲取BTC策略狀態"""
    try:
        if not os.path.exists(BTC_ENV_PATH):
            return jsonify({
                'status': 'inactive',
                'message': 'BTC配置不存在'
            })
        
        # 載入BTC配置
        btc_env = {}
        with open(BTC_ENV_PATH, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    btc_env[key] = value
        
        # 檢查登入狀態
        login_status = btc_env.get('LOGIN_BTC', '0')
        
        if login_status == '1':
            return jsonify({
                'status': 'active',
                'trading_pair': btc_env.get('TRADING_PAIR', ''),
                'leverage': btc_env.get('LEVERAGE', ''),
                'message': 'BTC策略運行中'
            })
        else:
            return jsonify({
                'status': 'inactive',
                'message': 'BTC帳戶未登入'
            })
            
    except Exception as e:
        logger.error(f"獲取BTC策略狀態失敗: {e}")
        return jsonify({
            'status': 'error',
            'message': f'狀態獲取失敗: {str(e)}'
        })

# ========================== 新增的帳戶相關API ==========================

def get_btc_account_balance():
    """獲取BTC帳戶餘額"""
    global binance_client
    
    try:
        if not binance_client:
            return jsonify({'success': False, 'message': '請先登入BTC帳戶'})
        
        # 獲取帳戶信息
        account_info = binance_client.get_account_info()
        if not account_info:
            return jsonify({'success': False, 'message': '無法獲取帳戶信息'})
        
        # 獲取餘額信息
        balance_info = binance_client.get_balance()
        
        # 整理餘額數據
        balances = []
        if balance_info:
            for balance in balance_info:
                if float(balance.get('balance', 0)) > 0:
                    balances.append({
                        'asset': balance.get('asset'),
                        'balance': balance.get('balance'),
                        'available': balance.get('withdrawAvailable', balance.get('balance'))
                    })
        
        return jsonify({
            'success': True,
            'total_wallet_balance': account_info.get('totalWalletBalance', '0'),
            'total_unrealized_pnl': account_info.get('totalUnrealizedProfit', '0'),
            'total_margin_balance': account_info.get('totalMarginBalance', '0'),
            'available_balance': account_info.get('availableBalance', '0'),
            'balances': balances
        })
        
    except Exception as e:
        logger.error(f"獲取BTC帳戶餘額失敗: {e}")
        return jsonify({'success': False, 'message': f'獲取失敗: {str(e)}'})

def get_btc_position():
    """獲取BTC持倉信息"""
    global binance_client
    
    try:
        if not binance_client:
            return jsonify({'success': False, 'message': '請先登入BTC帳戶'})
        
        # 獲取持倉信息
        position_info = binance_client.get_position_info()
        if not position_info:
            return jsonify({'success': False, 'message': '無法獲取持倉信息'})
        
        # 過濾出有持倉的合約
        active_positions = []
        for position in position_info:
            position_amt = float(position.get('positionAmt', 0))
            if position_amt != 0:
                # 調試：打印所有可用的position字段
                
                active_positions.append({
                    'symbol': position.get('symbol'),
                    'positionAmt': position.get('positionAmt'),
                    'entryPrice': position.get('entryPrice'),
                    'markPrice': position.get('markPrice'),
                    'liquidationPrice': position.get('liquidationPrice'),
                    'unRealizedProfit': position.get('unRealizedProfit'),
                    'percentage': position.get('percentage'),
                    'side': 'LONG' if position_amt > 0 else 'SHORT',
                    'leverage': position.get('leverage'),
                    'maintMargin': position.get('maintMargin'),
                    'isolatedMargin': position.get('isolatedMargin'),
                    'isolatedWallet': position.get('isolatedWallet'),
                    'marginRatio': position.get('marginRatio'),
                    'marginType': position.get('marginType')
                })
        
        return jsonify({
            'success': True,
            'positions': active_positions,
            'total_positions': len(active_positions)
        })
        
    except Exception as e:
        logger.error(f"獲取BTC持倉信息失敗: {e}")
        return jsonify({'success': False, 'message': f'獲取失敗: {str(e)}'})

def get_btc_version():
    """獲取幣安版本信息"""
    global binance_client
    
    try:
        if not binance_client:
            return jsonify({'success': False, 'message': '請先登入BTC帳戶'})
        
        # 獲取交易所信息
        exchange_info = binance_client.get_exchange_info()
        if not exchange_info:
            return jsonify({'success': False, 'message': '無法獲取交易所信息'})
        
        # 獲取服務器時間
        server_time = binance_client.get_server_time()
        
        return jsonify({
            'success': True,
            'version': 'Binance Futures API v1',
            'server_time': server_time.get('serverTime') if server_time else None,
            'timezone': exchange_info.get('timezone', 'UTC'),
            'rate_limits': len(exchange_info.get('rateLimits', [])),
            'symbols_count': len(exchange_info.get('symbols', [])),
            'api_status': 'NORMAL'
        })
        
    except Exception as e:
        logger.error(f"獲取幣安版本信息失敗: {e}")
        return jsonify({'success': False, 'message': f'獲取失敗: {str(e)}'})

def get_btc_trading_status():
    """獲取BTC交易狀態"""
    global binance_client, account_info
    
    try:
        if not binance_client:
            return jsonify({
                'success': False,
                'status': 'disconnected',
                'message': '未連接到幣安API'
            })
        
        # 檢查API連接狀態
        connection_test = binance_client.test_connection()
        if not connection_test:
            return jsonify({
                'success': False,
                'status': 'disconnected',
                'message': 'API連接失敗'
            })
        
        # 檢查帳戶狀態
        if account_info:
            can_trade = account_info.get('canTrade', False)
            can_withdraw = account_info.get('canWithdraw', False)
            can_deposit = account_info.get('canDeposit', False)
            
            # 獲取交易平台信息
            exchange_name = "未知"
            try:
                exchange_info = binance_client.get_exchange_info()
                if exchange_info and isinstance(exchange_info, dict):
                    # 檢查回應是否包含正確的交易所信息
                    if 'timezone' in exchange_info and 'serverTime' in exchange_info:
                        exchange_name = "Binance Futures"
                    elif exchange_info.get('symbols'):  # 或者檢查是否有symbols列表
                        exchange_name = "Binance Futures"
                    else:
                        exchange_name = "Binance Futures"  # 只要API有回應就認為連線成功
            except Exception as e:
                logger.error(f"獲取交易所信息失敗: {e}")
                exchange_name = "Binance Futures"  # 既然API連線正常，就顯示Binance Futures
            
            return jsonify({
                'success': True,
                'status': 'connected',
                'exchange_name': exchange_name,
                'can_trade': can_trade,
                'can_withdraw': can_withdraw,
                'can_deposit': can_deposit,
                'account_type': account_info.get('accountType', 'FUTURES'),
                'message': '幣安API連接正常'
            })
        else:
            # 嘗試重新獲取帳戶信息
            try:
                fresh_account_info = binance_client.get_account_info()
                if fresh_account_info:
                    return jsonify({
                        'success': True,
                        'status': 'connected',
                        'can_trade': fresh_account_info.get('canTrade', False),
                        'can_withdraw': fresh_account_info.get('canWithdraw', False),
                        'can_deposit': fresh_account_info.get('canDeposit', False),
                        'account_type': fresh_account_info.get('accountType', 'FUTURES'),
                        'message': '幣安API連接正常'
                    })
                else:
                    return jsonify({
                        'success': False,
                        'status': 'no_account_info',
                        'message': '無法獲取帳戶信息'
                    })
            except:
                return jsonify({
                    'success': False,
                    'status': 'disconnected',
                    'message': '帳戶信息獲取失敗'
                })
        
    except Exception as e:
        logger.error(f"獲取BTC交易狀態失敗: {e}")
        return jsonify({
            'success': False,
            'status': 'error',
            'message': f'狀態檢查失敗: {str(e)}'
        })

def send_btc_daily_startup_notification():
    """發送BTC每日啟動通知 - 00:00發送，使用真實API數據"""
    try:
        # 載入環境配置
        env_data = load_btc_env_data()
        trading_pair = env_data.get('TRADING_PAIR', 'BTCUSDT')
        leverage = env_data.get('LEVERAGE', '10')
        position_size = env_data.get('POSITION_SIZE', '20')
        margin_type = env_data.get('MARGIN_TYPE', 'CROSS')
        contract_type = env_data.get('CONTRACT_TYPE', 'PERPETUAL')
        user_id = env_data.get('BINANCE_USER_ID', '403303357')
        
        # 轉換保證金模式顯示
        margin_mode = "全倉" if margin_type == 'CROSS' else "逐倉"
        
        # 轉換合約類型顯示
        contract_display = "永續"  # PERPETUAL -> 永續
        if contract_type == 'QUARTERLY':
            contract_display = "期貨"
        elif contract_type == 'SPOT':
            contract_display = "現貨"
        
        # 檢查API連線狀態
        api_status = "未連線"
        exchange_name = "Binance Futures"
        
        if binance_client:
            try:
                # 測試API連線
                test_result = binance_client.test_connection()
                api_status = "已連線" if test_result else "未連線"
            except:
                api_status = "未連線"
        
        # 初始化帳戶數據
        account_data = {
            'totalWalletBalance': 0.0,
            'availableBalance': 0.0,
            'totalMarginBalance': 0.0,
            'totalUnrealizedProfit': 0.0,
            'feePaid': 0.0,
            'marginRatio': 0.0,
            'leverageUsage': 0.0,
            'todayPnl': 0.0,
            'todayPnlPercent': 0.0,
            'week7Pnl': 0.0,
            'week7PnlPercent': 0.0,
            'month30Pnl': 0.0,
            'month30PnlPercent': 0.0
        }
        
        # 獲取真實帳戶數據
        if binance_client and api_status == "已連線":
            try:
                # 獲取帳戶基本信息
                account_info = binance_client.get_account_info()
                if account_info:
                    account_data['totalWalletBalance'] = float(account_info.get('totalWalletBalance', 0))
                    account_data['availableBalance'] = float(account_info.get('availableBalance', 0))
                    account_data['totalMarginBalance'] = float(account_info.get('totalMarginBalance', 0))
                    account_data['totalUnrealizedProfit'] = float(account_info.get('totalUnrealizedProfit', 0))
                    
                    # 計算保證金比率和槓桿使用率
                    maintenance_margin = float(account_info.get('totalMaintMargin', 0))
                    initial_margin = float(account_info.get('totalInitialMargin', 0))
                    
                    if maintenance_margin > 0:
                        account_data['marginRatio'] = (account_data['totalMarginBalance'] / maintenance_margin) * 100
                    
                    # 修正：使用與前端一致的槓桿使用率計算邏輯
                    available_balance = float(account_info.get('availableBalance', 0))
                    unrealized_pnl = float(account_info.get('totalUnrealizedProfit', 0))
                    calculated_wallet_balance = available_balance + initial_margin + unrealized_pnl
                    if calculated_wallet_balance > 0:
                        account_data['leverageUsage'] = (initial_margin / calculated_wallet_balance) * 100
                
                # 獲取盈虧數據
                today = datetime.now()
                
                # 今日盈虧
                today_start = int(today.replace(hour=0, minute=0, second=0, microsecond=0).timestamp() * 1000)
                today_end = int(today.timestamp() * 1000)
                
                today_income = binance_client._make_request('GET', '/fapi/v1/income', {
                    'incomeType': 'REALIZED_PNL',
                    'startTime': today_start,
                    'endTime': today_end
                })
                if today_income:
                    account_data['todayPnl'] = sum(float(item.get('income', 0)) for item in today_income)
                    if account_data['totalWalletBalance'] > 0:
                        account_data['todayPnlPercent'] = (account_data['todayPnl'] / account_data['totalWalletBalance']) * 100
                
                # 7天盈虧
                week7_start = int((today - timedelta(days=7)).timestamp() * 1000)
                week7_income = binance_client._make_request('GET', '/fapi/v1/income', {
                    'incomeType': 'REALIZED_PNL',
                    'startTime': week7_start,
                    'endTime': today_end
                })
                if week7_income:
                    account_data['week7Pnl'] = sum(float(item.get('income', 0)) for item in week7_income)
                    if account_data['totalWalletBalance'] > 0:
                        account_data['week7PnlPercent'] = (account_data['week7Pnl'] / account_data['totalWalletBalance']) * 100
                
                # 30天盈虧
                month30_start = int((today - timedelta(days=30)).timestamp() * 1000)
                month30_income = binance_client._make_request('GET', '/fapi/v1/income', {
                    'incomeType': 'REALIZED_PNL',
                    'startTime': month30_start,
                    'endTime': today_end
                })
                if month30_income:
                    account_data['month30Pnl'] = sum(float(item.get('income', 0)) for item in month30_income)
                    if account_data['totalWalletBalance'] > 0:
                        account_data['month30PnlPercent'] = (account_data['month30Pnl'] / account_data['totalWalletBalance']) * 100
                
                # 獲取手續費
                fee_income = binance_client._make_request('GET', '/fapi/v1/income', {
                    'incomeType': 'COMMISSION',
                    'startTime': today_start,
                    'endTime': today_end
                })
                if fee_income:
                    account_data['feePaid'] = sum(abs(float(item.get('income', 0))) for item in fee_income)
                    
            except Exception as e:
                logger.error(f"獲取BTC帳戶信息失敗: {e}")
        
        # 獲取持倉信息
        position_info = ""
        if binance_client and api_status == "已連線":
            try:
                positions = binance_client.get_position_info()
                active_positions = [pos for pos in positions if float(pos.get('positionAmt', 0)) != 0]
                
                if not active_positions:
                    position_info = "❌ 無持倉部位"
                else:
                    for pos in active_positions:
                        position_amt = float(pos.get('positionAmt', 0))
                        entry_price = float(pos.get('entryPrice', 0))
                        unrealized_pnl = float(pos.get('unRealizedProfit', 0))
                        
                        # 判斷持倉方向
                        direction = "多單" if position_amt > 0 else "空單"
                        
                        position_info += f"{direction}｜{abs(position_amt):.8f} BTC｜{entry_price:,.2f} USDT\n未實現盈虧{unrealized_pnl:+.2f} USDT\n"
                        
            except Exception as e:
                logger.error(f"獲取BTC持倉信息失敗: {e}")
                position_info = "❌ 無持倉部位"
        else:
            position_info = "❌ 無持倉部位"
        
        # 構建訊息
        message = "✅ 自動交易比特幣正在啟動中.....\n"
        message += "═════ 系統資訊 ═════\n"
        message += f"交易平台：{exchange_name}\n"
        message += f"綁定帳戶：{user_id}\n"
        message += f"API 狀態：{api_status}\n"
        message += "═════ 選用合約 ═════\n"
        message += f"{trading_pair} {contract_display}（{margin_mode}{leverage}x {position_size}%）\n"
        message += "═════ 帳戶狀態 ═════\n"
        message += f"錢包餘額：{account_data['totalWalletBalance']:.8f} USDT\n"
        message += f"可供轉帳：{account_data['availableBalance']:.8f} USDT\n"
        message += f"保證金餘額：{account_data['totalMarginBalance']:.8f} USDT\n"
        message += f"未實現盈虧：{account_data['totalUnrealizedProfit']:+.8f} USDT\n"
        message += f"交易手續費：{account_data['feePaid']:.8f} USDT\n"
        message += f"保證金比率：{account_data['marginRatio']:.2f}%\n"
        message += f"槓桿使用率：{account_data['leverageUsage']:.2f}%\n"
        message += f"本日盈虧：{account_data['todayPnl']:+.8f} USDT\n"
        message += f"7 天盈虧：{account_data['week7Pnl']:+.8f} USDT\n"
        message += f"30天盈虧：{account_data['month30Pnl']:+.8f} USDT\n"
        message += "═════ 持倉狀態 ═════\n"
        message += position_info
        
        # 發送 Telegram 訊息
        telegram_success = send_btc_telegram_message(message)
        
        if telegram_success:
            # 前端日誌已在send_btc_telegram_message中記錄
            logger.info(f"BTC啟動通知已發送")
            return True
        else:
            # 失敗時需要額外記錄，因為send_btc_telegram_message返回False時不會記錄前端日誌
            log_btc_frontend_message("Telegram［啟動通知］訊息發送失敗", "error")
            return False
        
    except Exception as e:
        logger.error(f"發送BTC啟動通知失敗: {e}")
        # 記錄失敗日誌
        log_btc_frontend_message(f"Telegram［啟動通知］訊息發送失敗: {str(e)}", "error")
        return False

def _format_time_btc(timestamp):
    """格式化BTC時間戳，移除毫秒"""
    if not timestamp:
        return '-'
    try:
        # 移除毫秒部分，只保留到秒
        if '.' in timestamp:
            timestamp = timestamp.split('.')[0]
        # 確保格式為 YYYY-MM-DD HH:MM:SS
        if 'T' in timestamp:
            timestamp = timestamp.replace('T', ' ')
        return timestamp
    except:
        return timestamp



# ========================== API 路由函數 ==========================

def get_monthly_commission():
    """獲取本月手續費"""
    try:
        if not binance_client:
            return 0.0
        
        
        # 獲取本月1號00:00的時間戳
        now = datetime.now(timezone.utc)
        start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        start_time = int(start_of_month.timestamp() * 1000)
        
        # 獲取當前時間戳
        end_time = int(now.timestamp() * 1000)
        
        try:
            # 查詢手續費記錄
            income_data = binance_client._make_request('GET', '/fapi/v1/income', {
                'incomeType': 'COMMISSION',
                'startTime': start_time,
                'endTime': end_time,
                'limit': 1000
            })
            
            if income_data:
                total_commission = sum(abs(float(item.get('income', 0))) for item in income_data)
                return total_commission
        except Exception as e:
            logger.error(f"獲取本月手續費失敗: {e}")
        
        return 0.0
    except Exception as e:
        logger.error(f"計算本月手續費時發生錯誤: {e}")
        return 0.0

def get_monthly_realized_pnl():
    """獲取本月已實現盈虧"""
    try:
        if not binance_client:
            return 0.0
        
        
        # 獲取本月1號00:00的時間戳
        now = datetime.now(timezone.utc)
        start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        start_time = int(start_of_month.timestamp() * 1000)
        
        # 獲取當前時間戳
        end_time = int(now.timestamp() * 1000)
        
        try:
            # 查詢已實現盈虧記錄
            income_data = binance_client._make_request('GET', '/fapi/v1/income', {
                'incomeType': 'REALIZED_PNL',
                'startTime': start_time,
                'endTime': end_time,
                'limit': 1000
            })
            
            if income_data:
                total_realized_pnl = sum(float(item.get('income', 0)) for item in income_data)
                return total_realized_pnl
        except Exception as e:
            logger.error(f"獲取本月已實現盈虧失敗: {e}")
        
        return 0.0
    except Exception as e:
        logger.error(f"計算本月已實現盈虧時發生錯誤: {e}")
        return 0.0

def get_btc_startup_notification_data():
    """獲取BTC啟動通知所需的完整數據"""
    try:
        if not binance_client:
            return {
                'success': False,
                'error': 'BTC客戶端未初始化'
            }
        
        # 1. 獲取環境配置
        env_data = load_btc_env_data()
        trading_pair = env_data.get('TRADING_PAIR', 'BTCUSDT')
        leverage = env_data.get('LEVERAGE', '5')
        margin_type = env_data.get('MARGIN_TYPE', '全倉')
        max_loss_percent = env_data.get('MAX_LOSS_PERCENT', '10')
        
        # 2. 獲取帳戶信息
        account_data = binance_client.get_account_info()
        if not account_data:
            return {'success': False, 'error': '無法獲取帳戶資訊'}
        
        # 提取USDT數據
        usdt_asset = None
        for asset in account_data.get('assets', []):
            if asset.get('asset') == 'USDT':
                usdt_asset = asset
                break
        
        if usdt_asset:
            wallet_balance = float(usdt_asset.get('walletBalance', 0))
            available_balance = float(usdt_asset.get('availableBalance', 0)) 
            margin_balance = float(usdt_asset.get('marginBalance', 0))
            unrealized_pnl = float(usdt_asset.get('unrealizedProfit', 0))
            initial_margin = float(usdt_asset.get('initialMargin', 0))
            maint_margin = float(usdt_asset.get('maintMargin', 0))
        else:
            # 使用頂層數據
            wallet_balance = float(account_data.get('totalWalletBalance', 0))
            available_balance = float(account_data.get('availableBalance', 0))
            margin_balance = float(account_data.get('totalMarginBalance', 0))
            unrealized_pnl = float(account_data.get('totalUnrealizedProfit', 0))
            initial_margin = float(account_data.get('totalInitialMargin', 0))
            maint_margin = float(account_data.get('totalMaintMargin', 0))
        
        # 3. 計算統計數據
        today_commission = get_today_commission()
        margin_ratio = (margin_balance / maint_margin) * 100 if maint_margin > 0 else 0.0
        # 修正：手動計算錢包餘額
        calculated_wallet_balance = available_balance + initial_margin + unrealized_pnl
        leverage_usage = (initial_margin / calculated_wallet_balance) * 100 if calculated_wallet_balance > 0 else 0.0
        
        today_pnl, today_pnl_percent = get_period_total_pnl_binance_formula(1)
        week_pnl, week_pnl_percent = get_period_total_pnl_binance_formula(7)
        month_pnl, month_pnl_percent = get_period_total_pnl_binance_formula(30)
        
        # 4. 獲取持倉信息
        positions = binance_client.get_position_info()
        valid_positions = []
        
        if positions:
            for pos in positions:
                position_amt = float(pos.get('positionAmt', 0))
                if position_amt != 0:
                    side = 'LONG' if position_amt > 0 else 'SHORT'
                    entry_price = float(pos.get('entryPrice', 0))
                    unrealized_pos_pnl = float(pos.get('unRealizedProfit', 0))
                    
                    valid_positions.append({
                        'side': side,
                        'size': abs(position_amt),
                        'entry_price': entry_price,
                        'unrealized_pnl': unrealized_pos_pnl,
                        'symbol': pos.get('symbol', trading_pair)
                    })
        
        # 檢查API連接狀態
        api_connected = binance_client is not None
        if api_connected:
            try:
                # 測試API連接
                test_result = binance_client.get_server_time()
                api_connected = test_result is not None
            except:
                api_connected = False
        
        return {
            'success': True,
            'data': {
                'api_connected': api_connected,
                'trading_pair': trading_pair,
                'leverage': leverage,
                'margin_type': margin_type,
                'max_loss_percent': max_loss_percent,
                'account_id': env_data.get('BINANCE_USER_ID', 'N/A'),
                'wallet_balance': wallet_balance,
                'available_balance': available_balance,
                'margin_balance': margin_balance,
                'unrealized_pnl': unrealized_pnl,
                'today_commission': today_commission,
                'margin_ratio': margin_ratio,
                'leverage_usage': leverage_usage,
                'today_pnl': today_pnl,
                'today_pnl_percent': today_pnl_percent,
                'week_pnl': week_pnl,
                'week_pnl_percent': week_pnl_percent,
                'month_pnl': month_pnl,
                'month_pnl_percent': month_pnl_percent,
                'positions': valid_positions
            }
        }
        
    except Exception as e:
        logger.error(f"獲取BTC啟動通知數據失敗: {e}")
        return {
            'success': False,
            'error': str(e),
            'api_connected': binance_client is not None if 'binance_client' in locals() else False
        }

def get_today_trades():
    """獲取當日交易記錄"""
    try:
        if not binance_client:
            return []
        
        # 計算今天的開始和結束時間戳（毫秒）
        from datetime import datetime, timezone
        today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
        start_time = int(today.timestamp() * 1000)
        end_time = int((today.timestamp() + 86400) * 1000)  # 加一天
        
        # 獲取當日交易記錄
        trades = binance_client.get_account_trades(
            symbol='BTCUSDT',
            startTime=start_time,
            endTime=end_time
        )
        
        return trades if trades else []
        
    except Exception as e:
        logger.error(f"獲取當日交易記錄失敗: {e}")
        return []

def get_btc_message_count_today(message_type):
    """從前端日誌系統統計今天的特定消息數量"""
    try:
        # 動態獲取當前端口
        current_port = 5000
        try:
            import main
            if hasattr(main, 'CURRENT_PORT'):
                current_port = main.CURRENT_PORT
        except:
            pass
        
        # 獲取今天的日期
        today = datetime.now().strftime('%Y-%m-%d')
        
        # 向主系統詢問今天的日誌
        response = requests.get(f'http://127.0.0.1:{current_port}/api/logs', timeout=5)
        if response.status_code != 200:
            logger.warning(f"無法獲取日誌數據: HTTP {response.status_code}")
            return 0
        
        logs_data = response.json()
        logs = logs_data.get('logs', [])
        
        count = 0
        for log in logs:
            # 檢查是否為今天的BTC日誌
            timestamp = log.get('timestamp', '')
            if today in timestamp:
                extra_info = log.get('extra_info', {})
                system = extra_info.get('system', '')
                message = extra_info.get('message', '')
                
                # 確保是BTC系統的消息
                if system == 'BTC' and message_type in message:
                    count += 1
        
        return count
        
    except Exception as e:
        logger.warning(f"統計BTC消息數量失敗 ({message_type}): {e}")
        return 0

def get_btc_trading_statistics_data(date_str=None):
    """獲取BTC交易統計數據"""
    try:
        if not binance_client:
            return {
                'success': False,
                'error': 'BTC客戶端未初始化'
            }
        
        # 如果沒有指定日期，使用今天
        if not date_str:
            from datetime import datetime
            date_str = datetime.now().strftime('%Y-%m-%d')
        
        # 1. 從前端日誌系統統計委託、取消、成交次數
        order_count = get_btc_message_count_today('提交成功')  # 委託次數（提交成功訊息則數）
        cancel_count = get_btc_message_count_today('提交失敗')  # 取消次數（提交失敗訊息則數）
        fill_count = get_btc_message_count_today('成交通知')   # 成交次數（成交通知訊息則數）
        
        # 2. 從JSON配對系統獲取正確的交易統計數據
        try:
            # 使用JSON配對系統獲取當日所有交易記錄
            today_date_str = datetime.now().strftime('%Y-%m-%d')
            
            # 讀取當日交易數據文件 (使用正確的路徑)
            from trading_config import TradingConfig
            import os
            transdata_file = os.path.join(TradingConfig.BTC_DATA_DIR, f"BTCtransdata_{today_date_str.replace('-', '')}.json")
            all_trades = []
            
            try:
                with open(transdata_file, 'r', encoding='utf-8') as f:
                    all_trades = json.load(f)
            except FileNotFoundError:
                logger.warning(f"當日交易記錄文件不存在: {transdata_file}")
            except Exception as e:
                logger.error(f"讀取交易記錄失敗: {e}")
            
            # 初始化統計變數
            buy_volume = 0.0      # 做多BTC數量
            sell_volume = 0.0     # 做空BTC數量  
            buy_total_value = 0.0 # 做多USDT總量
            sell_total_value = 0.0 # 做空USDT總量
            
            # 從JSON交易記錄統計（正確區分開倉/平倉）
            for trade in all_trades:
                try:
                    action = trade.get('action', '').upper()  # BUY/SELL
                    oc_type = trade.get('oc_type', '').upper()  # Open/Cover
                    quantity = float(trade.get('quantity', 0))
                    price = float(trade.get('price', 0))
                    usdt_value = quantity * price
                    
                    # 根據開平倉類型和動作正確分類
                    if oc_type == 'OPEN' or oc_type == 'Open':  # 開倉交易
                        if action == 'BUY':  # 開多倉
                            buy_volume += quantity
                            buy_total_value += usdt_value
                        elif action == 'SELL':  # 開空倉
                            sell_volume += quantity
                            sell_total_value += usdt_value
                    elif oc_type == 'COVER' or oc_type == 'Cover':  # 平倉交易
                        if action == 'SELL':  # 平多倉（賣出平倉）
                            sell_volume += quantity
                            sell_total_value += usdt_value
                        elif action == 'BUY':  # 平空倉（買入平倉）
                            buy_volume += quantity
                            buy_total_value += usdt_value
                            
                except Exception as e:
                    logger.error(f"處理交易記錄失敗: {trade}, 錯誤: {e}")
                    continue
            
            # 計算平均價格
            avg_buy_price = buy_total_value / buy_volume if buy_volume > 0 else 0.0
            avg_sell_price = sell_total_value / sell_volume if sell_volume > 0 else 0.0
            
            logger.info(f"BTC統計數據來源: JSON配對系統")
            logger.info(f"買入總量: {buy_total_value:.2f} USDT, 賣出總量: {sell_total_value:.2f} USDT")
            logger.info(f"平均買價: {avg_buy_price:.2f} USDT, 平均賣價: {avg_sell_price:.2f} USDT")
            
        except Exception as e:
            logger.error(f"使用JSON配對系統統計失敗，回退到Binance API: {e}")
            
            # 回退邏輯：使用Binance API數據（保留原有邏輯作為備份）
            today_trades = get_today_trades()
            
            buy_volume = 0.0
            sell_volume = 0.0  
            buy_total_value = 0.0
            sell_total_value = 0.0
            
            if today_trades:
                for trade in today_trades:
                    qty = float(trade.get('qty', 0))
                    quote_qty = float(trade.get('quoteQty', 0))
                    side = trade.get('side', '')
                    
                    if side == 'BUY':
                        buy_volume += qty
                        buy_total_value += quote_qty
                    elif side == 'SELL':
                        sell_volume += qty
                        sell_total_value += quote_qty
            
            avg_buy_price = buy_total_value / buy_volume if buy_volume > 0 else 0.0
            avg_sell_price = sell_total_value / sell_volume if sell_volume > 0 else 0.0
            
            logger.info(f"BTC統計數據來源: Binance API (回退)")
        
        # 3. 獲取已實現盈虧
        today_realized_pnl = get_today_realized_pnl()
        
        # 4. 獲取當日手續費
        today_commission = get_today_commission()
        
        # 5. 獲取帳戶數據 - 使用內部函數避免Flask上下文問題
        frontend_data = get_btc_account_info_internal()
        if not frontend_data.get('success'):
            return {'success': False, 'error': '無法獲取帳戶數據'}
        
        account_raw = frontend_data['account']
        # 轉換數據格式
        account_data = {
            'totalWalletBalance': float(account_raw.get('walletBalance', '0').replace(',', '')),
            'availableBalance': float(account_raw.get('availableBalance', '0').replace(',', '')),
            'totalMarginBalance': float(account_raw.get('marginBalance', '0').replace(',', '')),
            'totalUnrealizedProfit': float(account_raw.get('unrealizedProfit', '0').replace(',', '')),
            'feePaid': float(account_raw.get('todayCommission', '0').replace(',', '')),
            'marginRatio': float(account_raw.get('marginRatio', '0%').replace('%', '')),
            'leverageUsage': float(account_raw.get('leverageUsage', '0%').replace('%', '')),
            'todayPnl': float(account_raw.get('todayPnl', '0').replace(',', '')),
            'todayPnlPercent': float(account_raw.get('todayPnlPercent', '0%').replace('%', '')),
            'weekPnl': float(account_raw.get('weekPnl', '0').replace(',', '')),
            'weekPnlPercent': float(account_raw.get('weekPnlPercent', '0%').replace('%', '')),
            'monthPnl': float(account_raw.get('monthPnl', '0').replace(',', '')),
            'monthPnlPercent': float(account_raw.get('monthPnlPercent', '0%').replace('%', ''))
        }
        
        # 6. 獲取當日平倉交易記錄
        closed_trades = []
        if today_trades:
            # 這裡需要分析交易記錄來識別平倉交易
            # 簡化版本：假設所有sell交易都是平倉
            for trade in today_trades:
                if not trade.get('isBuyer'):  # 賣出交易
                    side = '多單' if float(trade.get('qty', 0)) > 0 else '空單'
                    size = abs(float(trade.get('qty', 0)))
                    entry_price = 0.0  # 這需要從歷史數據中計算
                    exit_price = float(trade.get('price', 0))
                    pnl = float(trade.get('realizedPnl', 0)) if 'realizedPnl' in trade else 0.0
                    
                    closed_trades.append({
                        'side': side,
                        'size': size,
                        'entry_price': entry_price,
                        'exit_price': exit_price,
                        'pnl': pnl
                    })
        
        return {
            'success': True,
            'data': {
                'date': date_str,
                'order_count': order_count,
                'cancel_count': cancel_count,
                'fill_count': fill_count,
                'buy_volume': buy_total_value,  # 買入總量（USDT）
                'sell_volume': sell_total_value,  # 賣出總量（USDT）
                'avg_buy_price': avg_buy_price,
                'avg_sell_price': avg_sell_price,
                'realized_profit': max(0, today_realized_pnl),  # 只取正數作為獲利
                'realized_pnl': today_realized_pnl,
                'total_realized_pnl': today_realized_pnl,  # 這裡可以累加歷史數據
                'account': account_data,
                'closed_trades': closed_trades,
                'positions': account_data.get('positions', [])  # 添加 positions 欄位
            }
        }
        
    except Exception as e:
        logger.error(f"獲取BTC交易統計數據失敗: {e}")
        return {
            'success': False,
            'error': str(e)
        }

def get_btc_contract_info():
    """獲取真實的BTC合約信息"""
    try:
        if not binance_client:
            return 'BTCUSDT｜永續｜未知'
        
        # 獲取帳戶信息
        account_info = binance_client.get_account_info()
        if not account_info:
            return 'BTCUSDT｜永續｜未知'
        
        # 獲取BTCUSDT持倉信息來確定保證金模式和槓桿
        positions = binance_client.get_position_info()
        btc_position = None
        
        for pos in positions:
            if pos.get('symbol') == 'BTCUSDT':
                btc_position = pos
                break
        
        if btc_position:
            # 獲取保證金模式
            margin_type = btc_position.get('marginType', 'cross')
            margin_mode = '全倉' if margin_type.lower() == 'cross' else '逐倉'
            
            # 獲取槓桿倍數
            leverage = btc_position.get('leverage', '10')
            
            contract_info = f"BTCUSDT｜永續｜{margin_mode}{leverage}X"
        else:
            # 如果沒有持倉，使用帳戶默認設置
            contract_info = 'BTCUSDT｜永續｜全倉20X'
        
        logger.info(f"📊 獲取BTC合約信息: {contract_info}")
        return contract_info
        
    except Exception as e:
        logger.error(f"獲取BTC合約信息失敗: {e}")
        return 'BTCUSDT｜永續｜全倉20X'

def generate_btc_daily_report(date_str=None, custom_filename=None, is_monthly=False, monthly_data=None):
    """生成BTC日報/月報Excel文件 - 四大區塊格式（主要函數）"""
    try:
        if not binance_client:
            return {
                'success': False,
                'error': 'BTC客戶端未初始化'
            }
        
        # 如果沒有指定日期，使用今天
        if not date_str:
            from datetime import datetime
            date_str = datetime.now().strftime('%Y-%m-%d')
        
        # 獲取數據：月報使用傳入的數據，日報查詢API
        if is_monthly and monthly_data:
            data = monthly_data['stats_data']
            account = data['account']
            closed_trades_data = monthly_data['closed_trades_data']
            open_positions_data = monthly_data['open_positions_data']
        else:
            # 獲取統計數據 - 直接使用幣安API與啟動通知/交易統計保持一致
            account = {
                'wallet_balance': 0.0,
                'available_balance': 0.0,
                'margin_balance': 0.0,
                'unrealized_pnl': 0.0,
                'today_commission': 0.0,
                'margin_ratio': 0.0,
                'leverage_usage': 0.0,
                'today_pnl': 0.0,
                'week_pnl': 0.0,
                'month_pnl': 0.0
            }
            
            # 獲取真實帳戶數據（使用與啟動通知相同的方式）
            try:
                # 獲取帳戶基本信息
                account_info = binance_client.get_account_info()
                if account_info:
                    account['wallet_balance'] = float(account_info.get('totalWalletBalance', 0))
                    account['available_balance'] = float(account_info.get('availableBalance', 0))
                    account['margin_balance'] = float(account_info.get('totalMarginBalance', 0))
                    account['unrealized_pnl'] = float(account_info.get('totalUnrealizedProfit', 0))
                    
                    # 計算保證金比率和槓桿使用率
                    maintenance_margin = float(account_info.get('totalMaintMargin', 0))
                    initial_margin = float(account_info.get('totalInitialMargin', 0))
                    
                    if maintenance_margin > 0:
                        account['margin_ratio'] = (account['margin_balance'] / maintenance_margin) * 100
                    
                    # 使用與前端一致的槓桿使用率計算邏輯
                    available_balance = float(account_info.get('availableBalance', 0))
                    unrealized_pnl = float(account_info.get('totalUnrealizedProfit', 0))
                    calculated_wallet_balance = available_balance + initial_margin + unrealized_pnl
                    if calculated_wallet_balance > 0:
                        account['leverage_usage'] = (initial_margin / calculated_wallet_balance) * 100
                
                # 獲取盈虧數據
                today = datetime.now()
                
                # 今日盈虧
                today_start = int(today.replace(hour=0, minute=0, second=0, microsecond=0).timestamp() * 1000)
                today_end = int(today.timestamp() * 1000)
                
                today_income = binance_client._make_request('GET', '/fapi/v1/income', {
                    'incomeType': 'REALIZED_PNL',
                    'startTime': today_start,
                    'endTime': today_end
                })
                if today_income:
                    account['today_pnl'] = sum(float(item.get('income', 0)) for item in today_income)
                
                # 7天盈虧
                week7_start = int((today - timedelta(days=7)).timestamp() * 1000)
                week7_income = binance_client._make_request('GET', '/fapi/v1/income', {
                    'incomeType': 'REALIZED_PNL',
                    'startTime': week7_start,
                    'endTime': today_end
                })
                if week7_income:
                    account['week_pnl'] = sum(float(item.get('income', 0)) for item in week7_income)
                
                # 30天盈虧
                month30_start = int((today - timedelta(days=30)).timestamp() * 1000)
                month30_income = binance_client._make_request('GET', '/fapi/v1/income', {
                    'incomeType': 'REALIZED_PNL',
                    'startTime': month30_start,
                    'endTime': today_end
                })
                if month30_income:
                    account['month_pnl'] = sum(float(item.get('income', 0)) for item in month30_income)
                    
            except Exception as e:
                logger.error(f"獲取BTC帳戶數據失敗: {e}")
                return {'success': False, 'error': f'無法獲取帳戶數據: {str(e)}'}
            
            # 初始化交易總覽數據（與交易統計函數保持一致）
            data = {
                'order_count': 0,
                'cancel_count': 0, 
                'fill_count': 0,
                'buy_volume': 0.0,
                'sell_volume': 0.0,
                'avg_buy_price': 0.0,
                'avg_sell_price': 0.0
            }
            
            # 獲取當日交易統計數據
            try:
                today = datetime.now()
                today_start = int(today.replace(hour=0, minute=0, second=0, microsecond=0).timestamp() * 1000)
                today_end = int(today.timestamp() * 1000)
                
                # 獲取所有訂單數據（用於統計委託次數、取消次數、成交次數）
                orders_data = binance_client._make_request('GET', '/fapi/v1/allOrders', {
                    'symbol': 'BTCUSDT',
                    'startTime': today_start,
                    'endTime': today_end
                })
                
                if orders_data:
                    for order in orders_data:
                        status = order.get('status', '')
                        if status in ['NEW', 'PARTIALLY_FILLED', 'FILLED', 'CANCELED', 'REJECTED', 'EXPIRED']:
                            data['order_count'] += 1  # 所有狀態都算委託次數
                        if status in ['CANCELED', 'EXPIRED']:
                            data['cancel_count'] += 1  # 取消和過期算取消次數
                        if status in ['FILLED', 'PARTIALLY_FILLED']:
                            data['fill_count'] += 1  # 成交和部分成交算成交次數
                
                # 獲取成交記錄數據（用於統計買賣總量和平均價格）
                trades_data = binance_client._make_request('GET', '/fapi/v1/userTrades', {
                    'symbol': 'BTCUSDT',
                    'startTime': today_start,
                    'endTime': today_end
                })
                
                if trades_data:
                    
                    # 根據交易方向統計（方式一）
                    buy_trades = []   # 所有BUY方向交易
                    sell_trades = []  # 所有SELL方向交易
                    
                    for trade in trades_data:
                        qty = float(trade.get('qty', 0))
                        price = float(trade.get('price', 0))
                        side = trade.get('side', '')
                        trade_value = qty * price
                        
                        if side == 'BUY':
                            # 所有買入交易
                            buy_trades.append({'qty': qty, 'price': price, 'usdt': trade_value})
                            data['buy_volume'] += trade_value
                        elif side == 'SELL':
                            # 所有賣出交易
                            sell_trades.append({'qty': qty, 'price': price, 'usdt': trade_value})
                            data['sell_volume'] += trade_value
                    
                    # 計算平均價格
                    if buy_trades:
                        total_buy_qty = sum(t['qty'] for t in buy_trades)
                        total_buy_value = sum(t['usdt'] for t in buy_trades)
                        data['avg_buy_price'] = total_buy_value / total_buy_qty if total_buy_qty > 0 else 0
                    
                    if sell_trades:
                        total_sell_qty = sum(t['qty'] for t in sell_trades)
                        total_sell_value = sum(t['usdt'] for t in sell_trades)
                        data['avg_sell_price'] = total_sell_value / total_sell_qty if total_sell_qty > 0 else 0
                        
            except Exception as e:
                logger.error(f"獲取BTC交易統計失敗: {e}")
            
            # 獲取詳細交易記錄
            closed_trades_data = get_btc_closed_trades_today(date_str)
            open_positions_data = get_btc_open_positions_today()
        
        # 創建Excel文件
        wb = openpyxl.Workbook()
        ws = wb.active
        if custom_filename and '月報' in custom_filename:
            ws.title = f"BTC交易報表_月報"
        else:
            ws.title = f"BTC交易報表_{date_str.replace('-', '')}"
        
        # 設置樣式
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # 背景色
        blue_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # 字體
        white_font = Font(color="FFFFFF", bold=True)
        black_font = Font(color="000000", bold=True)
        
        # 對齊
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # 邊框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        current_row = 1
        
        # ========== 第一區塊：交易總覽 ==========
        # 區塊標題
        ws.merge_cells(f'A{current_row}:K{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "交易總覽"
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = center_alignment
        cell.border = thin_border
        current_row += 1
        
        # 第一標題行
        headers1 = ['委託次數', '取消次數', '成交次數', '買入總量', '賣出總量', '平均買價', '平均賣價']
        for i, header in enumerate(headers1):
            col = get_column_letter(i + 1)
            cell = ws[f'{col}{current_row}']
            cell.value = header
            cell.fill = gray_fill
            cell.font = black_font
            cell.alignment = center_alignment
            cell.border = thin_border
        current_row += 1
        
        # 第一區塊數據
        values1 = [
            f"{data['order_count']} 筆",
            f"{data['cancel_count']} 筆", 
            f"{data['fill_count']} 筆",
            f"{data['buy_volume']:.2f} USDT",
            f"{data['sell_volume']:.2f} USDT",
            f"{data['avg_buy_price']:.2f} USDT",
            f"{data['avg_sell_price']:.2f} USDT"
        ]
        for i, value in enumerate(values1):
            col = get_column_letter(i + 1)
            cell = ws[f'{col}{current_row}']
            cell.value = value
            cell.alignment = center_alignment
            cell.border = thin_border
        current_row += 2  # 空一行
        
        # ========== 第二區塊：帳戶狀態 ==========
        # 區塊標題
        ws.merge_cells(f'A{current_row}:K{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "帳戶狀態"
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = center_alignment
        cell.border = thin_border
        current_row += 1
        
        # 第二標題行
        headers2 = ['錢包餘額', '可供轉帳', '保證金餘額', '未實現盈虧', '交易手續費', '保證金比率', '槓桿使用率', '本日盈虧', '7 天盈虧', '30天盈虧']
        for i, header in enumerate(headers2):
            col = get_column_letter(i + 1)
            cell = ws[f'{col}{current_row}']
            cell.value = header
            cell.fill = gray_fill
            cell.font = black_font
            cell.alignment = center_alignment
            cell.border = thin_border
        current_row += 1
        
        # 第二區塊數據
        values2 = [
            f"{account['wallet_balance']:.8f} USDT",
            f"{account['available_balance']:.8f} USDT",
            f"{account['margin_balance']:.8f} USDT",
            f"{account['unrealized_pnl']:+.8f} USDT",
            f"{account['today_commission']:.8f} USDT",
            f"{account['margin_ratio']:.2f}%",
            f"{account['leverage_usage']:.2f}%",
            f"{account['today_pnl']:+.2f} USDT",
            f"{account['week_pnl']:+.2f} USDT",
            f"{account['month_pnl']:+.2f} USDT"
        ]
        for i, value in enumerate(values2):
            col = get_column_letter(i + 1)
            cell = ws[f'{col}{current_row}']
            cell.value = value
            cell.alignment = center_alignment
            cell.border = thin_border
        current_row += 2  # 空一行
        
        # ========== 第三區塊：交易明細 ==========
        # 區塊標題
        ws.merge_cells(f'A{current_row}:K{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "交易明細"
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = center_alignment
        cell.border = thin_border
        current_row += 1
        
        # 第三標題行
        headers3 = ['平倉時間', '交易單號', '選用合約', '交易動作', '交易類別', '交易方向', '交易數量', '持倉數量', '開倉價格', '平倉價格', '已實現盈虧']
        for i, header in enumerate(headers3):
            col = get_column_letter(i + 1)
            cell = ws[f'{col}{current_row}']
            cell.value = header
            cell.fill = gray_fill
            cell.font = black_font
            cell.alignment = center_alignment
            cell.border = thin_border
        current_row += 1
        
        # 第三區塊數據（平倉交易）
        if closed_trades_data:
            for trade in closed_trades_data:
                # 計算持倉數量(USDT) = BTC數量 * 開倉價格
                btc_quantity = trade.get('quantity', 0)
                entry_price = trade.get('entry_price', 0)
                position_usdt = btc_quantity * entry_price
                
                # 獲取真實合約信息
                margin_info = get_btc_contract_info()  # 獲取真實合約信息
                
                values3 = [
                    trade.get('close_time', 'N/A'),  # 這個已經在上面的close_time處理中設置
                    trade.get('order_id', 'N/A'),   # 這個已經在上面的order_id處理中設置
                    margin_info,
                    "平倉",
                    f"{trade.get('order_type', '市價')}單",
                    '多單' if trade.get('side') == 'LONG' else '空單',
                    f"{btc_quantity:.8f} BTC",
                    f"{position_usdt:.2f} USDT",
                    f"{entry_price:.2f} USDT",
                    f"{trade.get('exit_price', 0):.2f} USDT",
                    f"{trade.get('realized_pnl', 0):+.2f} USDT"
                ]
                for i, value in enumerate(values3):
                    col = get_column_letter(i + 1)
                    cell = ws[f'{col}{current_row}']
                    cell.value = value
                    cell.alignment = center_alignment
                    cell.border = thin_border
                current_row += 1
        else:
            # 無平倉交易時顯示空行
            for i in range(11):
                col = get_column_letter(i + 1)
                cell = ws[f'{col}{current_row}']
                cell.value = ""  # 空白欄位
                cell.alignment = center_alignment
                cell.border = thin_border
            current_row += 1
        
        current_row += 1  # 空一行
        
        # ========== 第四區塊：持倉狀態（動態位置）==========
        # 區塊標題
        ws.merge_cells(f'A{current_row}:K{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "持倉狀態"
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = center_alignment
        cell.border = thin_border
        current_row += 1
        
        # 第四標題行
        headers4 = ['開倉時間', '交易單號', '選用合約', '交易動作', '交易類別', '交易方向', '交易數量', '持倉數量', '開倉價格', '強平價格', '未實現盈虧']
        for i, header in enumerate(headers4):
            col = get_column_letter(i + 1)
            cell = ws[f'{col}{current_row}']
            cell.value = header
            cell.fill = gray_fill
            cell.font = black_font
            cell.alignment = center_alignment
            cell.border = thin_border
        current_row += 1
        
        # 第四區塊數據（持倉狀態）
        if open_positions_data:
            for position in open_positions_data:
                # 獲取原始持倉數量（帶符號）和絕對值
                position_amt_signed = float(position.get('position_amt', 0))
                position_amt = float(position.get('abs_position_amt', abs(position_amt_signed)))
                entry_price = position.get('entry_price', 0)
                position_usdt = position_amt * entry_price
                
                # 獲取真實合約信息
                margin_info = get_btc_contract_info()  # 獲取真實合約信息
                
                values4 = [
                    position.get('open_time', 'N/A'),
                    position.get('order_id', 'N/A'),
                    margin_info,
                    "開倉",
                    f"{position.get('order_type', '市價')}單",
                    '多單' if position_amt_signed > 0 else '空單',  # 使用帶符號的值判斷方向
                    f"{position_amt:.8f} BTC",
                    f"{position_usdt:.2f} USDT",
                    f"{entry_price:.2f} USDT",
                    f"{position.get('liquidation_price', 0):.2f} USDT",
                    f"{position.get('unrealized_pnl', 0):+.2f} USDT"
                ]
                for i, value in enumerate(values4):
                    col = get_column_letter(i + 1)
                    cell = ws[f'{col}{current_row}']
                    cell.value = value
                    cell.alignment = center_alignment
                    cell.border = thin_border
                current_row += 1
        else:
            # 無持倉時顯示空行
            for i in range(11):
                col = get_column_letter(i + 1)
                cell = ws[f'{col}{current_row}']
                cell.value = ""  # 空白欄位
                cell.alignment = center_alignment
                cell.border = thin_border
            current_row += 1
        
        # 設置欄寬
        for col in range(1, 12):  # A-K欄
            ws.column_dimensions[get_column_letter(col)].width = 25
        
        # 保存文件
        report_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'BTC交易報表')
        os.makedirs(report_dir, exist_ok=True)
        if custom_filename:
            filename = custom_filename
        else:
            filename = f"BTC_{date_str}.xlsx"
        file_path = os.path.join(report_dir, filename)
        
        wb.save(file_path)
        
        # 前端日誌記錄
        log_btc_frontend_message(f"{filename} 生成成功！！！", "success")
        
        # 發送 Telegram 通知並附上檔案（優化訊息內容）
        from datetime import datetime
        if custom_filename and '月報' in custom_filename:
            # 月報訊息
            month_str = datetime.now().strftime('%Y年%m月')
            caption = f"{month_str} | 月度交易結算報表已生成！！！"
        else:
            # 日報訊息
            date_str = datetime.now().strftime('%Y年%m月%d日')
            caption = f"{date_str} | 交易結算報表已生成！！！"
        send_btc_telegram_file(file_path, caption)
        
        return file_path
        
    except Exception as e:
        logger.error(f"生成BTC日報失敗: {e}")
        try:
            log_btc_frontend_message(f"交易報表生成失敗：{str(e)[:100]}", "error")
        except:
            pass
        
        return {
            'success': False,
            'error': str(e)
        }

def get_btc_closed_trades_today(date_str):
    """獲取當日平倉交易記錄 - 使用JSON配對數據"""
    try:
        # 🔥 使用BTC JSON配對系統獲取平倉明細
        from trade_pairing_BTC import get_btc_cover_trades_for_report
        
        # 解析日期字串獲取天數差
        target_date = datetime.strptime(date_str, '%Y-%m-%d')
        today = datetime.now()
        days_diff = (today.date() - target_date.date()).days
        
        # 獲取指定日期的平倉交易明細
        cover_trades = get_btc_cover_trades_for_report(date_range_days=max(1, days_diff + 1))
        closed_trades = []
        
        for trade in cover_trades:
            try:
                # 解析時間戳
                cover_timestamp = trade.get('cover_timestamp', '')
                if cover_timestamp:
                    # 處理多種時間戳格式
                    if 'T' in cover_timestamp:
                        cover_time = datetime.fromisoformat(cover_timestamp.replace('Z', '+00:00'))
                    else:
                        cover_time = datetime.fromisoformat(cover_timestamp)
                else:
                    logger.warning(f"BTC平倉記錄缺少時間戳: {trade}")
                    continue
                
                # 判斷是否為指定日期的交易
                if cover_time.date() == target_date.date():
                    closed_trade = {
                        'close_time': cover_time.strftime('%Y/%m/%d %H:%M:%S'),
                        'order_id': trade.get('cover_order_id', 'N/A'),
                        'side': 'SHORT' if trade['cover_action'] == 'BUY' else 'LONG',  # BUY=平空單(顯示空單), SELL=平多單(顯示多單)
                        'quantity': float(trade['matched_quantity']),
                        'position_size': float(trade['matched_quantity']),
                        'entry_price': float(trade['open_price']),  # 從配對數據獲取真實開倉價
                        'exit_price': float(trade['cover_price']),
                        'realized_pnl': float(trade['pnl']),  # 從配對計算獲取真實損益
                        'action': '平倉',
                        'order_type': '市價',
                        'source': trade.get('source', 'manual'),  # 添加交易來源
                        'open_trade_id': trade['open_trade_id'],  # 額外信息
                        'open_timestamp': trade['open_timestamp']
                    }
                    closed_trades.append(closed_trade)
                    
            except Exception as e:
                logger.error(f"處理BTC平倉記錄失敗: {trade}, 錯誤: {e}")
                continue
        
        logger.info(f"✅ 使用JSON配對數據獲取BTC平倉記錄: {len(closed_trades)}筆")
        return closed_trades
        
    except Exception as e:
        logger.error(f"獲取BTC平倉交易記錄失敗: {e}")
        # 如果JSON數據獲取失敗，回退到原有邏輯
        try:
            if not binance_client:
                return []
            
            today_trades = get_today_trades()
            closed_trades = []
            
            for trade in today_trades:
                trade_time = datetime.fromtimestamp(int(trade.get('time', 0)) / 1000)
                
                closed_trade = {
                    'close_time': trade_time.strftime('%Y/%m/%d %H:%M:%S'),
                    'order_id': trade.get('orderId', 'N/A'),
                    'side': 'SHORT' if trade.get('isBuyer') else 'LONG',  # isBuyer=True表示BUY操作=平空單(顯示空單)
                    'quantity': float(trade.get('qty', 0)),
                    'position_size': float(trade.get('qty', 0)),
                    'entry_price': 0.0,
                    'exit_price': float(trade.get('price', 0)),
                    'realized_pnl': float(trade.get('realizedPnl', 0)) if 'realizedPnl' in trade else 0.0,
                    'action': '平倉',
                    'order_type': '市價'
                }
                closed_trades.append(closed_trade)
            
            return closed_trades
        except:
            return []

def get_btc_position_from_json(is_long, position_size, entry_price):
    """從JSON配對系統獲取持倉的開倉信息"""
    try:
        # 載入最近30天的開倉記錄
        from datetime import timedelta
        import os
        import json
        from trading_config import TradingConfig
        
        BTC_RECORDS_DIR = TradingConfig.BTC_RECORDS_DIR
        today = datetime.now()
        
        target_action = 'BUY' if is_long else 'SELL'
        price_tolerance = entry_price * 0.005  # 0.5%的價格容差
        quantity_tolerance = position_size * 0.1  # 10%的數量容差
        
        logger.info(f"🔍 查找BTC持倉開倉記錄: {target_action} {position_size}BTC @ ${entry_price}")
        logger.info(f"🔍 容差範圍: 價格±{price_tolerance:.2f}, 數量±{quantity_tolerance:.8f}")
        
        # 查找最近30天的開倉記錄
        for i in range(30):
            check_date = today - timedelta(days=i)
            date_str = check_date.strftime('%Y%m%d')
            open_file = os.path.join(BTC_RECORDS_DIR, f'btc_open_positions_{date_str}.json')
            
            if os.path.exists(open_file):
                try:
                    with open(open_file, 'r', encoding='utf-8') as f:
                        opens = json.load(f)
                    
                    # 查找匹配的開倉記錄
                    for open_record in opens:
                        record_action = open_record.get('action')
                        record_status = open_record.get('status')
                        record_price = float(open_record.get('price', 0))
                        record_quantity = float(open_record.get('remaining_quantity', 0))
                        
                        
                        action_match = record_action == target_action
                        status_match = record_status == 'open'
                        price_match = abs(record_price - entry_price) <= price_tolerance
                        quantity_match = abs(record_quantity - position_size) <= quantity_tolerance
                        
                        if action_match and status_match and price_match and quantity_match:
                            
                            # 找到匹配的開倉記錄
                            timestamp = open_record.get('timestamp', '')
                            if timestamp:
                                try:
                                    if 'T' in timestamp:
                                        open_time = datetime.fromisoformat(timestamp.replace('Z', '+00:00'))
                                    else:
                                        open_time = datetime.fromisoformat(timestamp)
                                    open_time_str = open_time.strftime('%Y/%m/%d %H:%M:%S')
                                except:
                                    open_time_str = 'N/A'
                            else:
                                open_time_str = 'N/A'
                            
                            order_id = str(open_record.get('order_id', 'N/A'))
                            source = open_record.get('source', 'manual')
                            
                            logger.info(f"✅ 從JSON找到持倉開倉記錄: {open_time_str}, 訂單ID: {order_id}, 來源: {source}")
                            return open_time_str, order_id, source
                            
                except Exception as e:
                    logger.error(f"讀取BTC開倉記錄失敗 {open_file}: {e}")
                    continue
        
        logger.warning(f"未從JSON找到匹配的開倉記錄: {target_action} {position_size}BTC @ ${entry_price}")
        return 'N/A', 'N/A', 'manual'
        
    except Exception as e:
        logger.error(f"從JSON獲取開倉信息失敗: {e}")
        return 'N/A', 'N/A', 'manual'

def get_btc_position_open_info(is_long, position_size, entry_price):
    """從交易歷史獲取持倉的開倉時間和訂單ID"""
    try:
        if not binance_client:
            return 'N/A', 'N/A'
            
        # 獲取最近30天的交易記錄
        end_time = int(datetime.now().timestamp() * 1000)
        start_time = end_time - (30 * 24 * 60 * 60 * 1000)  # 30天前
        
        trades = binance_client._make_request('GET', '/fapi/v1/userTrades', {
            'symbol': 'BTCUSDT',
            'startTime': start_time,
            'endTime': end_time,
            'limit': 1000  # 最多獲取1000筆記錄
        })
        
        if not trades:
            return 'N/A', 'N/A'
        
        # 尋找符合條件的開倉交易
        # 根據方向和價格匹配找到最可能的開倉交易
        target_side = 'BUY' if is_long else 'SELL'
        price_tolerance = entry_price * 0.001  # 0.1%的價格容差
        
        # 從最新的交易開始查找
        for trade in reversed(trades):
            trade_side = trade.get('side', '')
            trade_price = float(trade.get('price', 0))
            trade_qty = float(trade.get('qty', 0))
            
            # 檢查是否是匹配的開倉交易
            if (trade_side == target_side and 
                abs(trade_price - entry_price) <= price_tolerance and
                trade_qty >= position_size * 0.8):  # 數量至少80%匹配
                
                # 轉換時間戳為可讀格式
                trade_time = int(trade.get('time', 0))
                if trade_time:
                    open_time = datetime.fromtimestamp(trade_time / 1000).strftime('%Y/%m/%d %H:%M:%S')
                else:
                    open_time = 'N/A'
                
                order_id = str(trade.get('orderId', 'N/A'))
                
                return open_time, order_id
        
        # 如果沒找到匹配的交易，返回N/A
        return 'N/A', 'N/A'
        
    except Exception as e:
        logger.error(f"獲取BTC持倉開倉信息失敗: {e}")
        return 'N/A', 'N/A'

def get_btc_open_positions_today():
    """獲取當前持倉狀態 - 結合JSON配對數據和真實API數據"""
    try:
        open_positions = []
        
        if binance_client:
            # 獲取真實持倉信息
            positions = binance_client.get_position_info()
            
            for pos in positions:
                if pos.get('symbol') == 'BTCUSDT':
                    position_amt = float(pos.get('positionAmt', 0))
                    entry_price = float(pos.get('entryPrice', 0))
                    
                    if abs(position_amt) > 0.00000001:  # 有持倉（考慮BTC精度）
                        # 獲取更詳細的持倉信息
                        liquidation_price = float(pos.get('liquidationPrice', 0))
                        unrealized_pnl = float(pos.get('unRealizedProfit', 0))
                        margin_type = pos.get('marginType', 'cross')
                        leverage = pos.get('leverage', '10')
                        
                        # 首先嘗試從JSON配對系統獲取開倉信息
                        open_time, order_id, source = get_btc_position_from_json(position_amt > 0, abs(position_amt), entry_price)
                        
                        # 如果JSON系統找不到，回退到API歷史查找
                        if open_time == 'N/A' or order_id == 'N/A':
                            api_open_time, api_order_id = get_btc_position_open_info(position_amt > 0, abs(position_amt), entry_price)
                            if open_time == 'N/A':
                                open_time = api_open_time
                            if order_id == 'N/A':
                                order_id = api_order_id
                        
                        open_position = {
                            'open_time': open_time,  # 從JSON配對系統或交易歷史獲取開倉時間
                            'order_id': order_id,    # 從JSON配對系統或交易歷史獲取開倉訂單ID
                            'position_amt': position_amt,  # 持倉數量（保留符號用於方向判斷）
                            'abs_position_amt': abs(position_amt),  # 持倉數量絕對值（用於顯示）
                            'entry_price': entry_price,         # 開倉價格
                            'liquidation_price': liquidation_price,  # 強平價格
                            'unrealized_pnl': unrealized_pnl,   # 未實現盈虧
                            'action': '開倉',
                            'order_type': '市價',
                            'source': source,  # 保持原始來源標識
                            'symbol': 'BTCUSDT',
                            'margin_type': margin_type,  # 保證金模式
                            'leverage': leverage,        # 槓桿倍數
                            'side': 'LONG' if position_amt > 0 else 'SHORT'  # 多空方向
                        }
                        
                        # 嘗試從JSON配對系統獲取更詳細的開倉信息
                        try:
                            from trade_pairing_BTC import get_btc_trading_statistics
                            stats = get_btc_trading_statistics(date_range_days=30)
                            if stats['total_opens'] > 0:
                                logger.info(f"📊 BTC持倉詳情: {abs(position_amt):.8f}BTC @ ${entry_price:.2f}")
                        except Exception as e:
                            logger.warning(f"獲取JSON配對信息失敗: {e}")
                        
                        open_positions.append(open_position)
                        
                        logger.info(f"✅ 獲取BTC持倉: {abs(position_amt):.8f}BTC @ ${entry_price:.2f} 未實現盈虧:${unrealized_pnl:.2f}")
        
        if not open_positions:
            logger.info("📊 當前無BTC持倉")
            
        return open_positions
        
    except Exception as e:
        logger.error(f"獲取BTC持倉狀態失敗: {e}")
        return []
        
        # 第二區塊：帳戶狀態
        ws.merge_cells(f'A{current_row}:K{current_row}')
        ws[f'A{current_row}'] = "第二區塊 帳戶狀態"
        ws[f'A{current_row}'].fill = blue_fill
        ws[f'A{current_row}'].font = white_font
        ws[f'A{current_row}'].alignment = center_alignment
        current_row += 1
        
        # 第二區塊標題
        headers2 = ["錢包餘額", "可供轉帳", "保證金餘額", "未實現盈虧", "交易手續費", "保證金比率", "槓桿使用率", "本日盈虧", "7天盈虧", "30天盈虧"]
        for i, header in enumerate(headers2, 1):
            cell = ws.cell(row=current_row, column=i, value=header)
            cell.fill = gray_fill
            cell.font = black_font
            cell.alignment = center_alignment
            cell.border = thin_border
        current_row += 1
        
        # 第二區塊數據
        values2 = [
            f"{account['wallet_balance']:.8f}",
            f"{account['available_balance']:.8f}",
            f"{account['margin_balance']:.8f}",
            f"{account['unrealized_pnl']:.8f}",
            f"{account['today_commission']:.8f}",
            f"{account['margin_ratio']:.2f}%",
            f"{account['leverage_usage']:.2f}%",
            f"{account['today_pnl']:.8f}",
            f"{account['week_pnl']:.8f}",
            f"{account['month_pnl']:.8f}"
        ]
        for i, value in enumerate(values2, 1):
            cell = ws.cell(row=current_row, column=i, value=value)
            cell.alignment = center_alignment
            cell.border = thin_border
        current_row += 2
        
        # 第三區塊：交易明細
        ws.merge_cells(f'A{current_row}:K{current_row}')
        ws[f'A{current_row}'] = "第三區塊 交易明細"
        ws[f'A{current_row}'].fill = blue_fill
        ws[f'A{current_row}'].font = white_font
        ws[f'A{current_row}'].alignment = center_alignment
        current_row += 1
        
        # 第三區塊標題
        headers3 = ["平倉時間", "交易單號", "選用合約", "交易動作", "交易類別", "交易方向", "交易數量", "持倉數量", "開倉價格", "平倉價格", "已實現盈虧"]
        for i, header in enumerate(headers3, 1):
            cell = ws.cell(row=current_row, column=i, value=header)
            cell.fill = gray_fill
            cell.font = black_font
            cell.alignment = center_alignment
            cell.border = thin_border
        current_row += 1
        
        # 第三區塊數據
        if closed_trades:
            for trade in closed_trades:
                values3 = [
                    trade.get('close_time', ''),
                    trade.get('trade_id', ''),  # 交易單號
                    trade.get('symbol', 'BTCUSDT'),
                    trade.get('action', '平倉'),  # 交易動作
                    convert_order_type_to_display(trade.get('order_type')),  # 交易類別
                    trade.get('direction', ''),  # 交易方向
                    f"{trade.get('trade_quantity', 0):.8f}",  # 交易數量
                    f"{trade.get('position_size', 0):.8f}",
                    f"{trade.get('entry_price', 0):.2f}",
                    f"{trade.get('exit_price', 0):.2f}",
                    f"{trade.get('realized_pnl', 0):.8f}"
                ]
                for i, value in enumerate(values3, 1):
                    cell = ws.cell(row=current_row, column=i, value=value)
                    cell.alignment = center_alignment
                    cell.border = thin_border
                current_row += 1
        else:
            # 無交易記錄
            ws.merge_cells(f'A{current_row}:K{current_row}')
            ws[f'A{current_row}'] = "❌ 無平倉交易"
            ws[f'A{current_row}'].alignment = center_alignment
            current_row += 1
        
        current_row += 1
        
        # 第四區塊：持倉狀態
        ws.merge_cells(f'A{current_row}:K{current_row}')
        ws[f'A{current_row}'] = "第四區塊 持倉狀態"
        ws[f'A{current_row}'].fill = blue_fill
        ws[f'A{current_row}'].font = white_font
        ws[f'A{current_row}'].alignment = center_alignment
        current_row += 1
        
        # 第四區塊標題
        headers4 = ["開倉時間", "交易單號", "選用合約", "交易動作", "交易類別", "交易方向", "交易數量", "持倉數量", "開倉價格", "平倉價格", "未實現盈虧"]
        for i, header in enumerate(headers4, 1):
            cell = ws.cell(row=current_row, column=i, value=header)
            cell.fill = gray_fill
            cell.font = black_font
            cell.alignment = center_alignment
            cell.border = thin_border
        current_row += 1
        
        # 第四區塊數據
        if open_positions:
            for pos in open_positions:
                values4 = [
                    pos.get('entry_time', ''),  # 開倉時間
                    pos.get('trade_id', ''),    # 交易單號
                    pos.get('symbol', 'BTCUSDT'),
                    pos.get('action', '開倉'),  # 交易動作
                    convert_order_type_to_display(pos.get('order_type')),  # 交易類別
                    pos.get('direction', ''),   # 交易方向
                    f"{pos.get('trade_quantity', 0):.8f}",  # 交易數量
                    f"{pos.get('position_size', 0):.8f}",
                    f"{pos.get('entry_price', 0):.2f}",
                    '',  # 平倉價格空白
                    f"{pos.get('unrealized_pnl', 0):.8f}"
                ]
                for i, value in enumerate(values4, 1):
                    cell = ws.cell(row=current_row, column=i, value=value)
                    cell.alignment = center_alignment
                    cell.border = thin_border
                current_row += 1
        else:
            # 無持倉
            ws.merge_cells(f'A{current_row}:K{current_row}')
            ws[f'A{current_row}'] = "❌ 無持倉部位"
            ws[f'A{current_row}'].alignment = center_alignment
            current_row += 1
        
        # 調整列寬
        for col in range(1, 12):
            ws.column_dimensions[get_column_letter(col)].width = 25
        
        # 保存文件
        report_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'BTC交易報表')
        os.makedirs(report_dir, exist_ok=True)
        filename = f"BTC_{date_str}.xlsx"
        filepath = os.path.join(report_dir, filename)
        
        wb.save(filepath)
        
        return {
            'success': True,
            'filepath': filepath,
            'filename': filename,
            'date': date_str
        }
        
    except Exception as e:
        logger.error(f"生成BTC日報失敗: {e}")
        return {
            'success': False,
            'error': str(e)
        }

def get_detailed_trade_records(date_str):
    """獲取詳細的交易記錄和持倉信息"""
    try:
        # 獲取當日交易記錄
        today_trades = get_today_trades()
        
        # 獲取當前持倉
        positions = binance_client.get_position_info() if binance_client else []
        
        # 分析平倉交易
        closed_trades = []
        if today_trades:
            for trade in today_trades:
                if not trade.get('isBuyer'):  # 賣出交易視為平倉
                    closed_trades.append({
                        'close_time': datetime.fromtimestamp(int(trade.get('time', 0))/1000).strftime('%Y/%m/%d %H:%M:%S'),
                        'trade_id': trade.get('id', ''),  # 改為trade_id
                        'symbol': trade.get('symbol', 'BTCUSDT'),
                        'action': '平倉',  # 交易動作
                        'order_type': '市價',  # 交易類別
                        'direction': '多單' if float(trade.get('qty', 0)) > 0 else '空單',  # 交易方向
                        'trade_quantity': abs(float(trade.get('qty', 0))),  # 交易數量
                        'position_size': abs(float(trade.get('qty', 0))),
                        'entry_price': 0.0,  # 需要從歷史數據計算
                        'exit_price': float(trade.get('price', 0)),
                        'realized_pnl': float(trade.get('realizedPnl', 0)) if 'realizedPnl' in trade else 0.0
                    })
        
        # 當前持倉
        open_positions = []
        if positions:
            for pos in positions:
                position_amt = float(pos.get('positionAmt', 0))
                if position_amt != 0:
                    open_positions.append({
                        'entry_time': '',  # 開倉時間，需要從交易歷史獲取
                        'trade_id': '',    # 交易單號
                        'symbol': pos.get('symbol', 'BTCUSDT'),
                        'action': '開倉',  # 交易動作
                        'order_type': '市價',  # 交易類別
                        'direction': '多單' if position_amt > 0 else '空單',  # 交易方向
                        'trade_quantity': abs(position_amt),  # 交易數量
                        'position_size': abs(position_amt),
                        'entry_price': float(pos.get('entryPrice', 0)),
                        'unrealized_pnl': float(pos.get('unRealizedProfit', 0))
                    })
        
        return {
            'closed_trades': closed_trades,
            'open_positions': open_positions
        }
        
    except Exception as e:
        logger.error(f"獲取詳細交易記錄失敗: {e}")
        return {
            'closed_trades': [],
            'open_positions': []
        }

def generate_btc_monthly_report(year, month):
    """生成BTC月報Excel文件 - 使用與日報相同的格式和標題"""
    try:
        if not binance_client:
            return {
                'success': False,
                'error': 'BTC客戶端未初始化'
            }
        
        # 計算月份的所有日期
        from datetime import datetime, timedelta
        from calendar import monthrange
        
        month_days = monthrange(year, month)[1]
        month_str = f"{year:04d}-{month:02d}"
        
        # 收集每日數據
        daily_data = []
        total_stats = {
            'cancel_count': 0,
            'fill_count': 0, 
            'buy_volume': 0.0,
            'sell_volume': 0.0,
            'realized_profit': 0.0,
            'realized_pnl': 0.0,
            'total_realized_pnl': 0.0,
            'commission': 0.0,
            'all_closed_trades': [],
            'all_open_positions': []
        }
        
        # 收集每一天的數據
        for day in range(1, month_days + 1):
            date_str = f"{year:04d}-{month:02d}-{day:02d}"
            
            # 獲取該日統計數據
            day_stats = get_btc_trading_statistics_data(date_str)
            if day_stats['success']:
                data = day_stats['data']
                
                # 累加統計數據
                total_stats['cancel_count'] += data['cancel_count']
                total_stats['fill_count'] += data['fill_count']
                total_stats['buy_volume'] += data['buy_volume']
                total_stats['sell_volume'] += data['sell_volume']
                total_stats['realized_profit'] += data['realized_profit']
                total_stats['realized_pnl'] += data['realized_pnl']
                total_stats['total_realized_pnl'] += data['total_realized_pnl']
                total_stats['commission'] += data['account']['today_commission']
                
                # 收集整月所有平倉交易
                closed_trades_data = get_btc_closed_trades_today(date_str)
                if closed_trades_data['success']:
                    total_stats['all_closed_trades'].extend(closed_trades_data['data'])
                
        
        # 若無任何交易記錄，不生成月報
        if total_stats['fill_count'] == 0 and len(total_stats['all_closed_trades']) == 0:
            logger.info(f"{month_str} 無交易記錄，不生成月報")
            return {
                'success': False,
                'error': '無交易記錄'
            }
        
        # 獲取當日帳戶狀態（第二區塊使用當日數據）
        try:
            account_info = binance_client.futures_account()
            balance = float(account_info['totalWalletBalance'])
            pnl = float(account_info['totalUnrealizedPnL'])
            margin_ratio = float(account_info['totalMaintMargin']) / balance * 100 if balance > 0 else 0
            
            current_account = {
                'balance': balance,
                'unrealized_pnl': pnl,
                'margin_ratio': margin_ratio,
                'today_commission': 0  # 月報中不顯示當日手續費
            }
        except Exception as e:
            logger.error(f"獲取帳戶狀態失敗: {e}")
            current_account = {
                'balance': 0,
                'unrealized_pnl': 0,
                'margin_ratio': 0,
                'today_commission': 0
            }
        
        # 獲取當日持倉狀態（第四區塊使用當日數據）
        current_positions_data = get_btc_open_positions_today()
        current_positions = current_positions_data.get('data', []) if current_positions_data.get('success') else []
        
        # 檔名
        filename = f'BTC_{month_str}_月報.xlsx'
        
        # 計算平均價格
        avg_buy_price = total_stats['buy_volume'] / total_stats['fill_count'] if total_stats['fill_count'] > 0 else 0.0
        avg_sell_price = total_stats['sell_volume'] / total_stats['fill_count'] if total_stats['fill_count'] > 0 else 0.0
        
        # 組裝月報數據結構，與日報格式完全一致
        monthly_stats_data = {
            'cancel_count': total_stats['cancel_count'],
            'fill_count': total_stats['fill_count'],
            'buy_volume': total_stats['buy_volume'],
            'sell_volume': total_stats['sell_volume'],
            'avg_buy_price': avg_buy_price,
            'avg_sell_price': avg_sell_price,
            'realized_profit': total_stats['realized_profit'],
            'realized_pnl': total_stats['realized_pnl'],
            'total_realized_pnl': total_stats['total_realized_pnl'],
            'account': current_account  # 當日帳戶狀態
        }
        
        # 組裝傳遞給日報函數的數據
        monthly_report_data = {
            'stats_data': monthly_stats_data,
            'closed_trades_data': {'success': True, 'data': total_stats['all_closed_trades']},
            'open_positions_data': {'success': True, 'data': current_positions}
        }
        
        # 使用日報函數生成月報，格式和標題完全一致
        today_str = datetime.now().strftime('%Y-%m-%d')
        report_path = generate_btc_daily_report(
            date_str=today_str,
            custom_filename=filename,
            is_monthly=True,
            monthly_data=monthly_report_data
        )
        
        if report_path:
            logger.info(f"BTC月報生成及發送成功: {report_path}")
            return {
                'success': True,
                'file_path': report_path,
                'filename': filename,
                'year': year,
                'month': month
            }
        else:
            return {
                'success': False,
                'error': 'BTC月報生成失敗'
            }
        
    except Exception as e:
        logger.error(f"生成BTC月報失敗: {e}")
        return {
            'success': False,
            'error': str(e)
        }

def parse_btc_action_and_direction(action, side):
    """解析BTC交易動作和方向
    - new buy: 開倉多單
    - new sell: 開倉空單  
    - cover sell: 平倉多單
    - cover buy: 平倉空單
    """
    # 統一轉換為小寫處理
    action = action.lower() if action else ''
    side = side.lower() if side else ''
    
    # 判斷動作和方向
    if action == 'new' and side == 'buy':
        return '開倉', '多單'
    elif action == 'new' and side == 'sell':
        return '開倉', '空單'
    elif action == 'cover' and side == 'sell':
        return '平倉', '多單'
    elif action == 'cover' and side == 'buy':
        return '平倉', '空單'
    else:
        # 默認情況，嘗試從side判斷
        if side in ['buy', 'long']:
            return '開倉', '多單'
        elif side in ['sell', 'short']:
            return '開倉', '空單'
        else:
            return '交易', '未知'


def store_pending_order(order_id, order_info):
    """存儲待成交訂單"""
    global pending_orders
    pending_orders[str(order_id)] = {
        'order_id': str(order_id),
        'timestamp': datetime.now().isoformat(),
        **order_info
    }
    logger.info(f"存儲待成交訂單: {order_id}")

def remove_pending_order(order_id):
    """移除已處理訂單"""
    global pending_orders
    order_id_str = str(order_id)
    if order_id_str in pending_orders:
        del pending_orders[order_id_str]
        logger.info(f"移除已處理訂單: {order_id}")

def process_order_fill(order_id, fill_price, fill_quantity):
    """處理訂單成交事件"""
    global pending_orders
    order_id_str = str(order_id)
    
    if order_id_str not in pending_orders:
        logger.warning(f"未找到待成交訂單: {order_id}")
        return
    
    order_info = pending_orders[order_id_str]
    action = order_info['action']
    direction = order_info['direction']
    order_type = order_info['order_type']
    source = order_info['source']
    
    try:
        # 前端日誌記錄（成交）- 暫時移除立即輸出，改為延遲輸出
        
        # 🔥 新增：BTC JSON交易記錄系統
        try:
            # 判斷交易類型並記錄到JSON配對系統
            btc_side = direction.upper()  # 多/空 -> BUY/SELL (根據方向)
            if direction == '多':
                trade_action = 'BUY'
            elif direction == '空':
                trade_action = 'SELL'
            else:
                trade_action = 'BUY'  # 預設值
            
            if action in ['開倉', 'open']:
                # 開倉記錄
                trade_id = record_btc_opening_trade(
                    action=trade_action,
                    quantity=float(fill_quantity),
                    price=float(fill_price),
                    order_id=order_id,
                    source=source
                )
                logger.info(f"✅ BTC開倉記錄已保存: {trade_id}")
                
            elif action in ['平倉', 'cover']:
                # 平倉記錄並自動配對
                cover_record = record_btc_covering_trade(
                    action=trade_action,
                    quantity=float(fill_quantity),
                    price=float(fill_price),
                    order_id=order_id,
                    source=source
                )
                if cover_record:
                    logger.info(f"✅ BTC平倉記錄已保存並配對完成: {cover_record['trade_id']}")
                    logger.info(f"   配對{len(cover_record.get('matched_opens', []))}筆開倉，總損益: ${cover_record.get('total_pnl', 0)}")
            
            # 保存到BTCtransdata目錄
            trade_data = {
                'trade_id': f"BTC_{action}_{trade_action}_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
                'timestamp': datetime.now().isoformat(),
                'symbol': 'BTCUSDT',
                'action': trade_action,
                'oc_type': 'Open' if action in ['開倉', 'open'] else 'Cover',
                'quantity': float(fill_quantity),
                'price': float(fill_price),
                'order_id': str(order_id),
                'source': source
            }
            save_btc_transdata(trade_data)
            
        except Exception as e:
            logger.error(f"BTC JSON交易記錄失敗: {e}")
        
        # 🎯 構建成交通知記錄並加入排隊系統 - 確保在提交成功通知(1秒)之後
        fill_trade_record = {
            'symbol': 'BTCUSDT',
            'side': 'BUY' if 'BUY' in action.upper() else 'SELL',
            'quantity': float(fill_quantity),
            'price': float(fill_price),
            'order_id': order_id,
            'order_type': order_type,
            'source': source,
            'action_type': action,
            'reduceOnly': 'cover' in action.lower(),
            'is_manual': source != 'webhook'
        }
        
        logger.info(f"準備將BTC成交通知加入排隊: 訂單{order_id}")
        btc_notification_manager.add_notification('fill', fill_trade_record, True, 5)
        logger.info(f"BTC成交通知已加入排隊系統（延遲5秒）")
        
        # 延遲記錄前端日誌（與成交通知同步）
        def delayed_frontend_log():
            try:
                fill_log = f"{action}成功：{direction}｜{float(fill_quantity):.8f} BTC｜{fill_price:,.2f} USDT｜{order_type}"
                log_btc_frontend_message(fill_log, "success")
                logger.info(f"系統訂單前端日誌已記錄: {order_id}")
            except Exception as e:
                logger.error(f"延遲記錄前端日誌失敗: {e}")
        
        # 使用線程延遲5秒記錄前端日誌
        timer = threading.Timer(5.0, delayed_frontend_log)
        timer.start()
        
        logger.info(f"系統訂單成交處理已安排5秒延遲執行，訂單ID: {order_id}")
        
        # 移除已處理訂單
        remove_pending_order(order_id)
        
    except Exception as e:
        logger.error(f"處理訂單成交通知失敗: {e}")

def handle_btc_order_fill(order_id, action, direction, quantity, order_type, source):
    """處理BTC訂單成交 - 使用WebSocket監控"""
    try:
        # 存儲待成交訂單信息
        order_info = {
            'action': action,
            'direction': direction,
            'quantity': quantity,
            'order_type': order_type,
            'source': source,
            'symbol': 'BTCUSDT'
        }
        store_pending_order(order_id, order_info)
        
        # 如果WebSocket未連接，啟動訂單監控
        if not order_monitor_thread or not order_monitor_thread.is_alive():
            start_btc_order_monitoring()
        
    except Exception as e:
        logger.error(f"處理BTC訂單監控失敗: {e}")

def send_btc_telegram_fill_notification(order_id, action, direction, quantity, price, order_type, source):
    """發送BTC Telegram成交通知 - 完整格式"""
    try:
        # 載入配置
        config = _load_btc_env_config()
        if not config:
            return False
        
        # 重新獲取最新的持倉數據用於成交通知
        position_data = get_btc_position_data_for_notification('BTCUSDT')
        btc_quantity = float(quantity)
        fill_price = float(price)
        
        # 格式化成交通知訊息
        today = datetime.now().strftime('%Y/%m/%d')
        message = f"✅ 成交通知（{today})\n"
        
        # 使用正確的持倉數據
        symbol = position_data.get('symbol', 'BTCUSDT')
        contract_type = position_data.get('contract_type', '永續')
        margin_mode = position_data.get('margin_mode', '全倉')
        leverage = position_data.get('leverage', '10')
        
        message += f"{symbol}｜{contract_type}｜{margin_mode}{leverage}X\n"
        message += f"交易訂單：{order_id}\n"
        message += f"交易類別：{order_type}\n"
        message += f"交易動作：{action}\n"
        message += f"交易方向：{direction}\n"
        # 判斷是否為平倉操作
        is_close_action = action in ['平倉', 'cover']
        
        if is_close_action:
            # 平倉：顯示平倉價格和已平倉狀態
            price_label = "平倉價格(USDT)"
            price_value = f"{fill_price:,.2f}"  # 成交的平倉價格
            liquidation_display = "已平倉"
        else:
            # 開倉：獲取成交後的真實持倉數據
            price_label = "開倉價格(USDT)"
            price_value = f"{fill_price:,.2f}"  # 成交的開倉價格
            
            # 獲取成交後的真實強平價格
            try:
                # 重新獲取最新的持倉信息，確保是成交後的數據
                position_info = binance_client.get_position_risk(symbol='BTCUSDT')
                real_liquidation_price = 0
                
                if position_info:
                    for pos in position_info:
                        if pos['symbol'] == 'BTCUSDT' and float(pos['positionAmt']) != 0:
                            real_liquidation_price = float(pos.get('liquidationPrice', 0))
                            break
                
                if real_liquidation_price > 0:
                    liquidation_display = f"{real_liquidation_price:,.2f}"
                else:
                    # 如果無法獲取真實強平價格，顯示N/A
                    liquidation_display = "N/A"
                    
            except Exception as e:
                logger.error(f"獲取成交後真實強平價格失敗: {e}")
                liquidation_display = "N/A"
        
        message += f"交易數量：{btc_quantity:.8f} BTC\n"
        message += f"持倉數量(USDT)：{btc_quantity * fill_price:,.2f}\n"  # 成交價格 * BTC數量
        message += f"{price_label}：{price_value}\n"
        message += f"強平價格(USDT)：{liquidation_display}"
        
        # 發送Telegram訊息
        return send_btc_telegram_message(message)
        
    except Exception as e:
        logger.error(f"發送BTC Telegram成交通知失敗: {e}")
        return False

def get_btc_position_notification_data(symbol='BTCUSDT'):
    """獲取持倉信息用於Telegram通知 - 從Binance API動態獲取真實數據"""
    try:
        if not binance_client:
            logger.warning("Binance客戶端未初始化，使用預設配置")
            env_data = load_btc_env_data()
            leverage = f"{env_data.get('LEVERAGE', '10')}X"
            margin_mode = '逐倉' if env_data.get('MARGIN_TYPE', 'CROSS') == 'ISOLATED' else '全倉'
            
            return {
                'symbol': symbol,  # 使用傳入的symbol
                'contract_type': '永續',
                'margin_mode': margin_mode,
                'leverage': leverage,
                'entry_price': 0,
                'liquidation_price': 0
            }
        
        # 獲取指定symbol的詳細持倉信息（包含槓桿和保證金模式）
        try:
            position_info = binance_client.get_position_risk(symbol=symbol)
            
            if position_info and len(position_info) > 0:
                # 尋找指定symbol的持倉數據
                pos_data = None
                for pos in position_info:
                    if pos.get('symbol') == symbol:
                        pos_data = pos
                        break
                
                if pos_data:
                    # 從API獲取真實數據
                    actual_symbol = pos_data.get('symbol', symbol)  # 交易對
                    leverage = f"{int(float(pos_data.get('leverage', 10)))}X"  # 槓桿倍數
                    margin_type = '逐倉' if pos_data.get('marginType') == 'isolated' else '全倉'  # 保證金模式
                    entry_price = float(pos_data.get('entryPrice', 0))  # 開倉價格
                    liquidation_price = float(pos_data.get('liquidationPrice', 0))  # 強平價格
                    
                    # 判斷合約類型（Binance期貨通常是永續合約）
                    contract_type = '永續'  # Binance USDT-M 期貨都是永續合約
                    
                    logger.info(f"從Binance API獲取持倉數據: {actual_symbol} | {contract_type} | {margin_type}{leverage}")
                    
                    return {
                        'symbol': actual_symbol,
                        'contract_type': contract_type,
                        'margin_mode': margin_type,
                        'leverage': leverage,
                        'entry_price': entry_price,
                        'liquidation_price': liquidation_price
                    }
            else:
                # 如果無法獲取持倉信息，嘗試從賬戶設置獲取交易配置
                logger.warning(f"無法獲取 {symbol} 的持倉信息，嘗試從賬戶設置獲取")
                
                # 獲取該交易對的當前配置（槓桿和保證金模式）
                try:
                    # 獲取賬戶資訊，嘗試找到該交易對的配置
                    account_info = binance_client.get_account()
                    
                    # 從賬戶信息中尋找該交易對的設置
                    symbol_leverage = 10  # 預設值
                    symbol_margin_type = 'CROSS'  # 預設值
                    
                    # 嘗試獲取交易對的槓桿設置
                    try:
                        leverage_info = binance_client.get_leverage_bracket(symbol=symbol)
                        if leverage_info and len(leverage_info) > 0:
                            # 從槓桿檔位信息中獲取當前設置（通常在第一個bracket中）
                            current_bracket = leverage_info[0].get('brackets', [{}])[0]
                            symbol_leverage = current_bracket.get('initialLeverage', 10)
                    except:
                        logger.warning(f"無法獲取 {symbol} 的槓桿設置，使用預設10倍")
                    
                    # 嘗試獲取保證金模式
                    try:
                        margin_info = binance_client.get_position_mode()
                        # 這裡可以根據需要調整邏輯
                        symbol_margin_type = 'CROSS'  # 多數情況下使用全倉
                    except:
                        logger.warning(f"無法獲取保證金模式，使用預設全倉")
                    
                    leverage = f"{int(symbol_leverage)}X"
                    margin_mode = '逐倉' if symbol_margin_type == 'ISOLATED' else '全倉'
                    
                    # 獲取當前市價
                    current_price = 0
                    try:
                        ticker = binance_client.get_ticker_price(symbol=symbol)
                        current_price = float(ticker.get('price', 0))
                    except:
                        logger.warning(f"無法獲取 {symbol} 的當前價格")
                    
                    logger.info(f"從賬戶設置獲取配置: {symbol} | 永續 | {margin_mode}{leverage}")
                    
                    return {
                        'symbol': symbol,
                        'contract_type': '永續',
                        'margin_mode': margin_mode,
                        'leverage': leverage,
                        'entry_price': current_price,
                        'liquidation_price': 0
                    }
                    
                except Exception as config_error:
                    logger.error(f"獲取賬戶配置失敗: {config_error}")
                    # 最後降級：使用環境配置
                    env_data = load_btc_env_data()
                    leverage = f"{env_data.get('LEVERAGE', '10')}X"
                    margin_mode = '逐倉' if env_data.get('MARGIN_TYPE', 'CROSS') == 'ISOLATED' else '全倉'
                    
                    return {
                        'symbol': symbol,
                        'contract_type': '永續',
                        'margin_mode': margin_mode,
                        'leverage': leverage,
                        'entry_price': 0,
                        'liquidation_price': 0
                    }
                
        except Exception as api_error:
            logger.error(f"調用Binance持倉API失敗: {api_error}")
            
            # API失敗時使用環境配置的預設值
            env_data = load_btc_env_data()
            leverage = f"{env_data.get('LEVERAGE', '10')}X"
            margin_mode = '逐倉' if env_data.get('MARGIN_TYPE', 'CROSS') == 'ISOLATED' else '全倉'
            
            return {
                'symbol': symbol,  # 使用傳入的symbol而非硬編碼
                'contract_type': '永續',
                'margin_mode': margin_mode,
                'leverage': leverage,
                'entry_price': 0,
                'liquidation_price': 0
            }
        
    except Exception as e:
        logger.error(f"獲取BTC持倉通知數據失敗: {e}")
        return {
            'symbol': symbol,  # 使用傳入的symbol而非硬編碼
            'contract_type': '永續', 
            'margin_mode': '全倉',
            'leverage': '10X',
            'entry_price': 0,
            'liquidation_price': 0
        }


def btc_place_order(quantity, action, side, order_type='MARKET', is_auto=False):
    """BTC下單函數 - 支援開倉/平倉判斷"""
    global btc_active_order_requests
    
    logger.info(f"🎯 btc_place_order開始執行: quantity={quantity}, action={action}, side={side}, order_type={order_type}, is_auto={is_auto}")
    
    # 生成下單請求ID用於重複檢查
    order_request_id = f"btc_{action}_{side}_{quantity}_{order_type}_{int(datetime.now().timestamp() * 1000)}"
    
    try:
        if not binance_client:
            logger.error(f"❌ BTC客戶端未初始化")
            return {
                'success': False,
                'error': 'BTC客戶端未初始化'
            }
        
        # 重複下單保護 - 檢查是否有相同的下單請求正在處理
        with btc_order_lock:
            # 檢查是否有相似的下單請求（相同action+side+quantity）
            current_time = int(datetime.now().timestamp() * 1000)
            similar_requests = [req for req in btc_active_order_requests 
                              if req.startswith(f"btc_{action}_{side}_{quantity}_{order_type}_")]
            
            if similar_requests:
                # 檢查是否在3秒內有相似請求（較短時間，因為這是高級API）
                for req in similar_requests:
                    req_time = int(req.split('_')[-1])
                    if current_time - req_time < 3000:  # 3秒內
                        logger.warning(f"3秒內相似BTC下單請求被阻止: {order_request_id}")
                        return {
                            'success': False,
                            'error': '3秒內相似下單請求已被阻止'
                        }
            
            # 添加到活躍請求列表
            btc_active_order_requests.add(order_request_id)
            
            # 清理超過30秒的舊請求記錄
            btc_active_order_requests = {req for req in btc_active_order_requests 
                                       if current_time - int(req.split('_')[-1]) < 30000}
        
        # 解析動作和方向
        logger.info(f"📝 解析動作和方向: action={action}, side={side}")
        parsed_action, direction = parse_btc_action_and_direction(action, side)
        logger.info(f"📝 解析結果: parsed_action={parsed_action}, direction={direction}")
        
        # 判斷訂單來源
        order_source = '自動' if is_auto else '手動'
        
        # 格式化數量
        formatted_quantity = f"{float(quantity):.8f}"
        
        # 記錄委託訂單日誌（前端日誌格式）
        order_type_text = convert_order_type_to_display(order_type)
        commit_log = f"{order_source}{parsed_action}：{direction}｜{formatted_quantity} BTC｜市價｜{order_type_text}"
        
        # 只記錄後端日誌，前端日誌由Websocket成交回調統一處理
        logger.info(f"({order_source}委託) {commit_log}")
        
        # 獲取持倉詳細信息
        position_details = get_btc_position_notification_data('BTCUSDT')
        
        try:
            # 準備下單參數
            order_params = {
                'symbol': 'BTCUSDT',
                'side': 'BUY' if side.lower() == 'buy' else 'SELL',
                'order_type': order_type,
                'quantity': formatted_quantity,
                'reduce_only': (action == 'cover')  # 平倉時設為True
            }
            
            # 如果是限價單，需要價格
            if order_type == 'LIMIT':
                # 獲取當前價格作為限價
                ticker = binance_client.get_ticker_price(symbol='BTCUSDT')
                current_price = float(ticker['price']) if ticker else 0
                # 稍微調整價格以便成交
                limit_price = current_price * 0.999 if side.lower() == 'buy' else current_price * 1.001
                order_params['price'] = limit_price
            
            # 執行下單
            order_result = binance_client.place_order(**order_params)
            
            if order_result:
                order_id = order_result.get('orderId', '')
                client_order_id = order_result.get('clientOrderId', '')
                
                # 立即標記為已處理，避免WebSocket重複處理（必須在獲取價格前）
                with processed_orders_lock:
                    processed_orders.add(f"{order_id}_NEW")
                    if order_type == 'MARKET':
                        processed_orders.add(f"{order_id}_FILLED")
                
                # 立即將訂單加入pending_orders，確保WebSocket能識別為系統訂單
                if str(order_id) not in pending_orders:
                    pending_orders[str(order_id)] = {
                        'order_id': str(order_id),
                        'timestamp': datetime.now().isoformat(),
                        'symbol': 'BTCUSDT',
                        'action': parsed_action,
                        'direction': direction,
                        'quantity': formatted_quantity,
                        'order_type': order_type_text,
                        'source': order_source
                    }
                    logger.info(f"{order_source}下單訂單{order_id}已立即加入pending_orders")
                
                # 獲取提交價格信息
                submitted_price = 0
                if order_type == 'LIMIT':
                    submitted_price = order_params.get('price', 0)
                    logger.info(f"限價單提交價格: {submitted_price}")
                else:
                    # 市價單獲取當前市價作為參考
                    try:
                        ticker = binance_client.get_ticker_price(symbol='BTCUSDT')
                        if ticker and 'price' in ticker:
                            submitted_price = float(ticker['price'])
                            logger.info(f"市價單參考價格獲取成功: {submitted_price}")
                        else:
                            logger.warning(f"獲取市價失敗，ticker回應: {ticker}")
                            submitted_price = 0
                    except Exception as e:
                        logger.error(f"獲取市價異常: {e}")
                        submitted_price = 0
                
                logger.info(f"最終設置的提交價格: {submitted_price}")
                
                # 構建trade_record用於通知
                trade_record = {
                    'symbol': 'BTCUSDT',
                    'side': side,
                    'quantity': formatted_quantity,
                    'price': submitted_price,  # 添加提交價格
                    'order_id': order_id,
                    'order_type': order_type_text,
                    'source': 'manual' if not is_auto else 'webhook',  # 修正source字段用於判斷手動/自動
                    'action_type': parsed_action,
                    'reduceOnly': action == 'cover',
                    'is_manual': not is_auto
                }
                
                
                # 使用排隊系統發送提交成功通知（1秒延遲）
                logger.info(f"準備將BTC提交成功通知加入排隊: {trade_record}")
                btc_notification_manager.add_notification('submit', trade_record, True, 1)
                logger.info(f"BTC提交成功通知已加入排隊系統（延遲1秒）")
                
                # submit_success 已經包含發送結果，通知函數內部已記錄日誌
                
                # 啟動訂單成交監控
                handle_btc_order_fill(order_id, parsed_action, direction, formatted_quantity, order_type_text, order_source)
                
                # 在返回值中需要的變數（這裡需要計算或獲取真實值）
                try:
                    # 獲取當前價格
                    ticker = binance_client.get_ticker_price(symbol='BTCUSDT')
                    current_price = float(ticker['price']) if ticker else 0
                    
                    # 計算持倉價值
                    position_value_usdt = float(formatted_quantity) * current_price
                except:
                    current_price = 0
                    position_value_usdt = 0
                
                # 下單成功，清理請求記錄
                with btc_order_lock:
                    btc_active_order_requests.discard(order_request_id)
                
                return {
                    'success': True,
                    'order_id': order_id,
                    'client_order_id': client_order_id,
                    'action': parsed_action,
                    'direction': direction,
                    'quantity': formatted_quantity,
                    'order_type': order_type_text,
                    'is_auto': is_auto,
                    'position_value_usdt': position_value_usdt,
                    'current_price': current_price,
                    'order_result': order_result,
                    'position_details': position_details
                }
            else:
                # 下單失敗，清理請求記錄
                with btc_order_lock:
                    btc_active_order_requests.discard(order_request_id)
                
                return {
                    'success': False,
                    'error': '下單失敗，未收到訂單結果'
                }
        
        except Exception as order_error:
            # 下單失敗的情況
            error_message = str(order_error)
            
            # 構建失敗的trade_record用於通知
            error_record = {
                'symbol': 'BTCUSDT',
                'side': side,
                'quantity': formatted_quantity,
                'order_id': '--',
                'order_type': order_type_text,
                'source': 'manual' if not is_auto else 'webhook',  # 修正source字段用於判斷手動/自動
                'action_type': parsed_action,
                'reduceOnly': action == 'cover',
                'is_manual': not is_auto,
                'error': error_message
            }
            fail_success = send_btc_order_submit_notification_delayed(error_record, False, 5)
            
            # fail_success 已經包含發送結果，通知函數內部已記錄日誌
            
            # 下單異常，清理請求記錄
            with btc_order_lock:
                btc_active_order_requests.discard(order_request_id)
            
            return {
                'success': False,
                'error': error_message,
                'fail_message_sent': fail_success
            }
            
    except Exception as e:
        logger.error(f"BTC下單失敗: {e}")
        
        # 最終異常，清理請求記錄
        with btc_order_lock:
            btc_active_order_requests.discard(order_request_id)
        
        return {
            'success': False,
            'error': str(e)
        }


def btc_webhook_handler(webhook_data):
    """處理BTC webhook自動交易"""
    try:
        # 解析webhook數據
        action = webhook_data.get('action', '').lower()  # new, cover
        side = webhook_data.get('side', '').lower()      # buy, sell
        quantity = webhook_data.get('quantity', 0)
        
        if not action or not side or not quantity:
            return {
                'success': False,
                'error': '缺少必要的交易參數'
            }
        
        # 執行自動下單
        order_result = btc_place_order(
            quantity=quantity,
            action=action,
            side=side,
            order_type='MARKET',
            is_auto=True
        )
        
        return order_result
        
    except Exception as e:
        logger.error(f"處理BTC webhook失敗: {e}")
        return {
            'success': False,
            'error': str(e)
        }

def get_btc_config():
    """獲取BTC配置"""
    try:
        env_data = load_btc_env_data()
        return jsonify({
            'success': True,
            'config': env_data
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        })

# 已刪除已棄用的get_today_realized_pnl函數

def _load_btc_env_config():
    """通用函數：載入BTC環境配置"""
    btc_env = {}
    if os.path.exists(BTC_ENV_PATH):
        try:
            with open(BTC_ENV_PATH, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith('#') and '=' in line:
                        key, value = line.split('=', 1)
                        btc_env[key] = value
        except Exception as e:
            logger.error(f"載入BTC環境配置失敗: {e}")
    return btc_env

def _send_telegram_api_request(message, files=None):
    """通用函數：發送Telegram API請求"""
    try:
        btc_env = _load_btc_env_config()
        bot_token = btc_env.get('BOT_TOKEN_BTC')
        chat_id = btc_env.get('CHAT_ID_BTC')
        
        if not bot_token or not chat_id:
            return False
        
        import requests
        
        if files:
            # 發送文件
            url = f"https://api.telegram.org/bot{bot_token}/sendDocument"
            with open(files, 'rb') as f:
                files_data = {'document': f}
                data = {'chat_id': chat_id, 'caption': message}
                response = requests.post(url, data=data, files=files_data, timeout=10)
        else:
            # 發送普通消息
            url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
            payload = {
                'chat_id': chat_id,
                'text': message,
                'parse_mode': 'HTML'
            }
            response = requests.post(url, json=payload, timeout=10)
        
        return response.status_code == 200
        
    except Exception as e:
        logger.error(f"發送Telegram API請求失敗: {e}")
        return False

def get_period_realized_pnl(days=1):
    """獲取指定天數的已實現盈虧總計
    Args:
        days (int): 天數，1=本日，7=七天，30=三十天
    Returns:
        tuple: (已實現盈虧, 盈虧百分比)
    """
    try:
        if not binance_client:
            return 0.0, 0.0
        
        # 獲取指定天數的時間範圍（使用UTC+0時間，與幣安統一）
        utc_now = datetime.now(timezone.utc)
        if days == 1:
            # 本日：今天00:00開始到現在
            start_date = utc_now.replace(hour=0, minute=0, second=0, microsecond=0)
        elif days == 7:
            # 7天：7天前00:00開始到現在 (完整7天)
            start_date = (utc_now - timedelta(days=7)).replace(hour=0, minute=0, second=0, microsecond=0)
        elif days == 30:
            # 30天：30天前00:00開始到現在 (完整30天)
            start_date = (utc_now - timedelta(days=30)).replace(hour=0, minute=0, second=0, microsecond=0)
        else:
            # 其他天數：N天前00:00開始到現在
            start_date = (utc_now - timedelta(days=days-1)).replace(hour=0, minute=0, second=0, microsecond=0)
        
        start_time = int(start_date.timestamp() * 1000)
        
        # 方法1: 先嘗試從income API獲取已實現盈虧（更準確）
        try:
            income_data = binance_client._make_request('GET', '/fapi/v1/income', {
                'startTime': start_time,
                'incomeType': 'REALIZED_PNL',
                'limit': 1000
            })
            
            if income_data:
                income_total = 0.0
                for income in income_data:
                    if int(income.get('time', 0)) >= start_time:
                        income_total += float(income.get('income', 0))
                
                # 計算盈虧百分比 - 使用期初餘額作為基準
                percentage = 0.0
                try:
                    # 獲取期初餘額（期間開始時的錢包餘額）
                    period_start_balance = get_period_start_balance(start_time)
                    if period_start_balance > 0:
                        percentage = (income_total / period_start_balance) * 100
                except Exception as calc_e:
                    logger.error(f"百分比計算失敗: {calc_e}")
                    pass
                
                return income_total, percentage
        except Exception as e:
            logger.error(f"收入API失敗，改用交易記錄API: {e}")
        
        # 方法2: 備用方案 - 使用配置的交易對獲取交易記錄
        try:
            with open(BTC_ENV_PATH, 'r', encoding='utf-8') as f:
                content = f.read()
                trading_pair = 'BTCUSDT'  # 默認值
                for line in content.split('\n'):
                    if line.startswith('TRADING_PAIR='):
                        trading_pair = line.split('=')[1].strip()
                        break
        except:
            trading_pair = 'BTCUSDT'  # 備用默認值
        
        # 透過 GET /fapi/v1/userTrades 的 realizedPnl 欄位直接獲取已實現盈虧
        trade_data = binance_client._make_request('GET', '/fapi/v1/userTrades', {
            'symbol': trading_pair,
            'startTime': start_time,
            'limit': 1000
        })
        
        if not trade_data:
            return 0.0, 0.0
        
        # 累計所有交易的已實現盈虧
        total_realized_pnl = 0.0
        trade_count = 0
        
        for trade in trade_data:
            trade_time = int(trade.get('time', 0))
            
            # 確保交易在指定期間範圍內
            if trade_time >= start_time:
                # 直接使用 GET /fapi/v1/userTrades 的 realizedPnl 字段
                # 注意：需要累計所有交易的 realizedPnl，包括0值
                realized_pnl = float(trade.get('realizedPnl', 0))
                total_realized_pnl += realized_pnl
                trade_count += 1
        
        # 計算盈虧百分比（備用方案）- 使用期初餘額作為基準
        percentage = 0.0
        try:
            period_start_balance = get_period_start_balance(start_time)
            if period_start_balance > 0:
                percentage = (total_realized_pnl / period_start_balance) * 100
        except:
            pass
        
        return total_realized_pnl, percentage
        
    except Exception as e:
        logger.error(f"從交易記錄獲取期間已實現盈虧失敗: {e}")
        return 0.0, 0.0

def get_period_start_balance(timestamp):
    """獲取期間開始時的錢包餘額
    Args:
        timestamp (int): 期間開始時間的毫秒時間戳
    Returns:
        float: 期間開始時的錢包餘額
    """
    try:
        if not binance_client:
            return 0.0
        
        # 獲取從期間開始到現在的所有收入記錄
        income_data = binance_client._make_request('GET', '/fapi/v1/income', {
            'startTime': timestamp,
            'limit': 1000
        })
        
        # 獲取當前錢包餘額
        account_info = binance_client.get_account_info()
        if not account_info:
            return 0.0
            
        current_balance = float(account_info.get('totalWalletBalance', 0))
        
        if not income_data:
            return current_balance  # 如果沒有收入記錄，返回當前餘額
        
        # 計算從期間開始到現在的總收入變化
        total_income_change = 0.0
        income_count = 0
        for income in income_data:
            if int(income.get('time', 0)) >= timestamp:
                # 累計所有類型的收入變化（包括已實現盈虧、手續費、資金費率等）
                income_value = float(income.get('income', 0))
                total_income_change += income_value
                income_count += 1
        
        
        # 期初餘額 = 當前餘額 - 期間收入變化
        period_start_balance = current_balance - total_income_change
        
        
        # 確保期初餘額為正數且合理
        if period_start_balance <= 0 or abs(period_start_balance) < 0.01:
            return current_balance
        
        # 如果期初餘額和當前餘額差距過大（超過90%），可能計算有誤
        if abs(period_start_balance - current_balance) / current_balance > 0.9:
            return current_balance
        
        return period_start_balance
        
    except Exception as e:
        logger.error(f"獲取期初餘額失敗: {e}")
        # 如果無法獲取期初餘額，返回當前餘額
        try:
            account_info = binance_client.get_account_info()
            if account_info:
                return float(account_info.get('totalWalletBalance', 0))
        except:
            pass
        return 0.0

def get_period_total_pnl_binance_formula(days=1):
    """使用幣安公式獲取期間總盈虧：已實現盈虧 + 未實現盈虧 - 資金費用
    Args:
        days (int): 天數，1=本日，7=七天，30=三十天
    Returns:
        tuple: (總盈虧, 盈虧百分比)
    """
    try:
        if not binance_client:
            return 0.0, 0.0
        
        # 獲取指定天數的時間範圍（使用UTC+0時間，與幣安統一）
        utc_now = datetime.now(timezone.utc)
        if days == 1:
            # 本日：今天00:00開始到現在
            start_date = utc_now.replace(hour=0, minute=0, second=0, microsecond=0)
        elif days == 7:
            # 7天：7天前00:00開始到現在 (完整7天)
            start_date = (utc_now - timedelta(days=7)).replace(hour=0, minute=0, second=0, microsecond=0)
        elif days == 30:
            # 30天：30天前00:00開始到現在 (完整30天)
            start_date = (utc_now - timedelta(days=30)).replace(hour=0, minute=0, second=0, microsecond=0)
        else:
            start_date = (utc_now - timedelta(days=days-1)).replace(hour=0, minute=0, second=0, microsecond=0)
        
        start_time = int(start_date.timestamp() * 1000)
        
        # 1. 獲取期間已實現盈虧
        realized_pnl = 0.0
        try:
            income_data = binance_client._make_request('GET', '/fapi/v1/income', {
                'startTime': start_time,
                'incomeType': 'REALIZED_PNL',
                'limit': 1000
            })
            
            if income_data:
                for income in income_data:
                    if int(income.get('time', 0)) >= start_time:
                        realized_pnl += float(income.get('income', 0))
        except Exception as e:
            logger.error(f"獲取已實現盈虧失敗: {e}")
        
        # 2. 獲取當前未實現盈虧
        unrealized_pnl = 0.0
        try:
            account_info = binance_client.get_account_info()
            if account_info:
                unrealized_pnl = float(account_info.get('totalUnrealizedProfit', 0))
        except Exception as e:
            logger.error(f"獲取未實現盈虧失敗: {e}")
        
        # 3. 獲取期間資金費用
        funding_fee = 0.0
        try:
            funding_data = binance_client._make_request('GET', '/fapi/v1/income', {
                'startTime': start_time,
                'incomeType': 'FUNDING_FEE',
                'limit': 1000
            })
            
            if funding_data:
                for fee in funding_data:
                    if int(fee.get('time', 0)) >= start_time:
                        funding_fee += float(fee.get('income', 0))
        except Exception as e:
            logger.error(f"獲取資金費用失敗: {e}")
        
        # 計算總盈虧：已實現盈虧 + 未實現盈虧 - 資金費用
        total_pnl = realized_pnl + unrealized_pnl - funding_fee
        
        # 計算盈虧百分比
        percentage = 0.0
        try:
            period_start_balance = get_period_start_balance(start_time)
            
            if period_start_balance > 0:
                percentage = (total_pnl / period_start_balance) * 100
            else:
                # 使用當前餘額作為備用方案
                try:
                    account_info = binance_client.get_account_info()
                    if account_info:
                        current_balance = float(account_info.get('totalWalletBalance', 0))
                        if current_balance > 0:
                            percentage = (total_pnl / current_balance) * 100
                except:
                    pass
        except Exception as e:
            logger.error(f"百分比計算失敗: {e}")
        
        
        return total_pnl, percentage
        
    except Exception as e:
        logger.error(f"獲取{days}天總盈虧失敗: {e}")
        return 0.0, 0.0

def get_today_realized_pnl_from_trades():
    """舊函數包裝器 - 向後兼容"""
    pnl, _ = get_period_realized_pnl(1)
    return pnl

def get_today_commission():
    """獲取本日手續費"""
    try:
        if not binance_client:
            return 0.0
        
        
        # 獲取今日00:00的時間戳
        today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
        start_time = int(today.timestamp() * 1000)
        
        # 獲取收益歷史，只獲取手續費
        income_data = binance_client._make_request('GET', '/fapi/v1/income', {
            'startTime': start_time,
            'incomeType': 'COMMISSION',
            'limit': 100
        })
        
        if not income_data:
            return 0.0
        
        # 計算本日手續費總和（手續費通常是負值）
        today_commission = 0.0
        for income in income_data:
            income_time = int(income.get('time', 0))
            if income_time >= start_time:
                commission_amount = abs(float(income.get('income', 0)))  # 取絕對值
                today_commission += commission_amount
        
        return today_commission
        
    except Exception as e:
        logger.error(f"獲取本日手續費失敗: {e}")
        return 0.0

def get_today_realized_pnl():
    """獲取本日已實現盈虧"""
    try:
        if not binance_client:
            return 0.0
        
        # 獲取今日00:00的時間戳
        today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
        start_time = int(today.timestamp() * 1000)
        
        # 獲取收益歷史，只獲取已實現盈虧
        income_data = binance_client._make_request('GET', '/fapi/v1/income', {
            'startTime': start_time,
            'incomeType': 'REALIZED_PNL',
            'limit': 100
        })
        
        if not income_data:
            return 0.0
        
        # 計算本日已實現盈虧總和
        today_realized_pnl = 0.0
        for income in income_data:
            income_time = int(income.get('time', 0))
            if income_time >= start_time:
                pnl_amount = float(income.get('income', 0))
                today_realized_pnl += pnl_amount
        
        return today_realized_pnl
        
    except Exception as e:
        logger.error(f"獲取本日已實現盈虧失敗: {e}")
        return 0.0

def get_period_pnl(days=1):
    """獲取指定天數的總盈虧 - 使用正確的幣安API"""
    try:
        if not binance_client:
            return 0.0, 0.0  # (總盈虧, 收益率%)
        
        
        # 計算起始時間
        now = datetime.now(timezone.utc)
        start_time = now - timedelta(days=days)
        start_timestamp = int(start_time.timestamp() * 1000)
        end_timestamp = int(now.timestamp() * 1000)
        
        
        # 使用幣安收入歷史API - 只獲取已實現盈虧
        total_pnl = 0.0
        
        # 1. 獲取已實現盈虧 (REALIZED_PNL)
        try:
            realized_pnl_data = binance_client._make_request('GET', '/fapi/v1/income', {
                'incomeType': 'REALIZED_PNL',
                'startTime': start_timestamp,
                'endTime': end_timestamp,
                'limit': 1000
            })
            
            if realized_pnl_data:
                realized_total = sum(float(item.get('income', 0)) for item in realized_pnl_data)
                total_pnl += realized_total
        except Exception as e:
            pass
        
        # 2. 獲取資金費用 (FUNDING_FEE)
        try:
            funding_data = binance_client._make_request('GET', '/fapi/v1/income', {
                'incomeType': 'FUNDING_FEE',
                'startTime': start_timestamp,
                'endTime': end_timestamp,
                'limit': 1000
            })
            
            if funding_data:
                funding_total = sum(float(item.get('income', 0)) for item in funding_data)
                total_pnl += funding_total
        except Exception as e:
            pass
        
        # 3. 對於本日數據，額外添加當前未實現盈虧
        if days == 1:
            try:
                account_info = binance_client.get_account_info()
                if account_info:
                    unrealized_pnl = float(account_info.get('totalUnrealizedProfit', 0))
                    # 注意：不要將未實現盈虧加入總盈虧，因為這是動態變化的
            except Exception as e:
                pass
        
        
        # 計算收益率 - 基於歷史錢包餘額
        percentage = 0.0
        try:
            # 獲取當前錢包餘額
            account_info = binance_client.get_account_info()
            if account_info:
                current_balance = float(account_info.get('totalWalletBalance', 0))
                
                # 計算期初餘額（當前餘額 - 期間盈虧）
                initial_balance = current_balance - total_pnl
                if initial_balance > 0:
                    percentage = (total_pnl / initial_balance) * 100
                else:
                    pass
        except Exception as e:
            pass
        
        return total_pnl, percentage
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return 0.0, 0.0


def get_binance_stats_pnl():
    """從幣安統計接口獲取準確的盈虧數據"""
    try:
        if not binance_client:
            return None
            
        
        # 方法1: 嘗試使用24hr ticker統計API
        try:
            ticker_24hr = binance_client._make_request('GET', '/fapi/v1/ticker/24hr')
        except Exception as e:
            pass
        
        # 方法2: 嘗試獲取帳戶快照
        try:
            # 這個API可能不存在，但值得嘗試
            snapshot = binance_client._make_request('GET', '/fapi/v1/accountSnapshot')
        except Exception as e:
            pass
        
        # 方法3: 嘗試獲取資產流水
        try:
            asset_flow = binance_client._make_request('GET', '/fapi/v1/assetFlow')
        except Exception as e:
            pass
        
        # 方法4: 使用用戶交易記錄計算
        try:
            user_trades = binance_client._make_request('GET', '/fapi/v1/userTrades', {
                'limit': 500
            })
            
            if user_trades:
                # 分析交易記錄，計算盈虧
                return analyze_trades_for_pnl(user_trades)
                
        except Exception as e:
            pass
            
        return None
        
    except Exception as e:
        return None

def analyze_trades_for_pnl(trades):
    """分析交易記錄計算真實盈虧"""
    try:
        
        # 按時間分組計算不同期間的盈虧
        now = datetime.now(timezone.utc)
        periods = {
            'today': now.replace(hour=0, minute=0, second=0, microsecond=0),
            'week': now - timedelta(days=7),
            'month': now - timedelta(days=30),
            'quarter': now - timedelta(days=90)
        }
        
        period_pnl = {}
        
        for period_name, start_time in periods.items():
            start_timestamp = int(start_time.timestamp() * 1000)
            period_total = 0.0
            trade_count = 0
            
            for trade in trades:
                trade_time = int(trade.get('time', 0))
                if trade_time >= start_timestamp:
                    realized_pnl = float(trade.get('realizedPnl', 0))
                    period_total += realized_pnl
                    trade_count += 1
            
            period_pnl[period_name] = {
                'pnl': period_total,
                'trades': trade_count
            }
        
        return period_pnl
        
    except Exception as e:
        return None

def get_today_pnl():
    """獲取本日總盈虧 - 使用本日已實現盈虧"""
    try:
        if not binance_client:
            return 0.0, 0.0
            
        # 本日盈虧就是本日已實現盈虧
        total_pnl = get_today_realized_pnl()
        
        # 獲取錢包餘額計算百分比
        account_info = binance_client.get_account_info()
        percentage = 0.0
        if account_info:
            wallet_balance = float(account_info.get('totalWalletBalance', 0))
            if wallet_balance > 0:
                percentage = (total_pnl / wallet_balance) * 100
        
        return total_pnl, percentage
        
    except Exception as e:
        return 0.0, 0.0


def get_btc_account_info_internal():
    """獲取BTC帳戶資訊 - 內部版本，不使用Flask上下文"""
    try:
        if not binance_client:
            return {
                'success': False,
                'error': 'BTC客戶端未初始化'
            }
        
        # 獲取帳戶資訊
        account_data = binance_client.get_account_info()
        if not account_data:
            return {
                'success': False,
                'error': '無法獲取帳戶資訊'
            }
        
        # 統一使用總餘額數據
        wallet_balance = float(account_data.get('totalWalletBalance', 0))
        available_balance = float(account_data.get('availableBalance', 0))
        margin_balance = float(account_data.get('totalMarginBalance', 0))
        unrealized_pnl = float(account_data.get('totalUnrealizedProfit', 0))
        initial_margin = float(account_data.get('totalInitialMargin', 0))
        maint_margin = float(account_data.get('totalMaintMargin', 0))
        
        # 獲取本日手續費和已實現盈虧
        today_commission = get_today_commission()
        
        # 計算多期間盈虧統計 - 使用幣安正確公式（已實現+未實現-資金費用）
        today_pnl, today_pnl_percent = get_period_total_pnl_binance_formula(1)      # 本日盈虧
        week_pnl, week_pnl_percent = get_period_total_pnl_binance_formula(7)        # 7天盈虧  
        month_pnl, month_pnl_percent = get_period_total_pnl_binance_formula(30)     # 30天盈虧
        
        # 計算保證金比率（預設0.00%）
        margin_ratio = 0.0
        if maint_margin > 0:
            margin_ratio = (margin_balance / maint_margin) * 100
        
        # 計算槓桿使用率（預設0.00%）
        # 修正：手動計算錢包餘額以避免API同步問題
        # 錢包餘額 = 可用餘額 + 已使用保證金 + 未實現盈虧
        calculated_wallet_balance = available_balance + initial_margin + unrealized_pnl
        
        leverage_usage = 0.0
        if calculated_wallet_balance > 0:
            leverage_usage = (initial_margin / calculated_wallet_balance) * 100
            logger.info(f"槓桿使用率計算: {initial_margin:.2f} / {calculated_wallet_balance:.2f} = {leverage_usage:.2f}% (API錢包餘額: {wallet_balance:.2f})")
        
        # 獲取用戶UID（從配置文件獲取，因為Binance API不返回UID）
        try:
            import os
            from dotenv import dotenv_values
            btc_config_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config', 'btc.env')
            if os.path.exists(btc_config_path):
                btc_config = dotenv_values(btc_config_path)
                user_uid = btc_config.get('BINANCE_USER_ID', 'N/A')
                logger.info(f"從配置文件獲取BTC用戶UID: {user_uid}")
            else:
                user_uid = 'N/A'
                logger.warning("BTC配置文件不存在")
        except Exception as e:
            logger.error(f"獲取BTC用戶UID失敗: {e}")
            user_uid = 'N/A'

        from collections import OrderedDict
        organized_account = OrderedDict([
            ('uid', user_uid),                                                       # 幣安用戶UID
            ('walletBalance', f"{wallet_balance:.8f}"),                              # 錢包餘額
            ('availableBalance', f"{available_balance:.8f}"),                        # 可用餘額  
            ('marginBalance', f"{margin_balance:.8f}"),                              # 保證金餘額
            ('unrealizedProfit', f"{unrealized_pnl:.8f}"),                          # 未實現盈虧
            ('initialMargin', f"{initial_margin:.8f}"),                              # 原始保證金（已使用）
            ('maintMargin', f"{maint_margin:.8f}"),                                  # 維持保證金
            ('marginRatio', f"{margin_ratio:.2f}%"),                                 # 保證金率（預設0.00%）
            ('leverageUsage', f"{leverage_usage:.2f}%"),                             # 槓桿使用率（預設0.00%）
            ('todayCommission', f"{today_commission:.8f}"),                          # 手續費
            ('todayPnl', f"{today_pnl:.8f}"),                                        # 本日盈虧
            ('todayPnlPercent', f"{today_pnl_percent:.2f}%"),                       # 本日盈虧百分比
            ('weekPnl', f"{week_pnl:.8f}"),                                          # 7天盈虧
            ('weekPnlPercent', f"{week_pnl_percent:.2f}%"),                         # 7天盈虧百分比
            ('monthPnl', f"{month_pnl:.8f}"),                                        # 30天盈虧
            ('monthPnlPercent', f"{month_pnl_percent:.2f}%"),                       # 30天盈虧百分比
        ])
        
        return {
            'success': True,
            'data': organized_account,  # 前端需要從data字段獲取
            'account': organized_account  # 保持向後兼容
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

def get_btc_account_info():
    """獲取BTC帳戶資訊 - 新版本支援完整字段"""
    try:
        if not binance_client:
            return jsonify({
                'success': False,
                'error': 'BTC客戶端未初始化'
            })
        
        # 獲取帳戶資訊
        account_data = binance_client.get_account_info()
        if not account_data:
            return jsonify({
                'success': False,
                'error': '無法獲取帳戶資訊'
            })
        
        # 嘗試從多個API端點獲取用戶ID
        user_uid = None
        
        # 直接從配置文件獲取用戶ID，因為Binance API無法提供用戶ID
        logger.info("直接從配置文件獲取Binance用戶ID")
        
        # 統一使用總餘額數據
        wallet_balance = float(account_data.get('totalWalletBalance', 0))
        available_balance = float(account_data.get('availableBalance', 0))
        margin_balance = float(account_data.get('totalMarginBalance', 0))
        unrealized_pnl = float(account_data.get('totalUnrealizedProfit', 0))
        initial_margin = float(account_data.get('totalInitialMargin', 0))
        maint_margin = float(account_data.get('totalMaintMargin', 0))
        
        # 獲取本日手續費和已實現盈虧
        today_commission = get_today_commission()
        
        # 計算多期間盈虧統計 - 使用幣安正確公式（已實現+未實現-資金費用）
        today_pnl, today_pnl_percent = get_period_total_pnl_binance_formula(1)      # 本日盈虧
        week_pnl, week_pnl_percent = get_period_total_pnl_binance_formula(7)        # 7天盈虧  
        month_pnl, month_pnl_percent = get_period_total_pnl_binance_formula(30)     # 30天盈虧
        
        
        # 計算保證金比率（預設0.00%）
        margin_ratio = 0.0
        if maint_margin > 0:
            margin_ratio = (margin_balance / maint_margin) * 100
        
        # 計算槓桿使用率（預設0.00%）
        # 修正：手動計算錢包餘額以避免API同步問題
        # 錢包餘額 = 可用餘額 + 已使用保證金 + 未實現盈虧
        calculated_wallet_balance = available_balance + initial_margin + unrealized_pnl
        
        leverage_usage = 0.0
        if calculated_wallet_balance > 0:
            leverage_usage = (initial_margin / calculated_wallet_balance) * 100
            logger.info(f"槓桿使用率計算: {initial_margin:.2f} / {calculated_wallet_balance:.2f} = {leverage_usage:.2f}% (API錢包餘額: {wallet_balance:.2f})")
        
        # totalMaintMargin: 維持保證金
        
        # 獲取用戶UID（從配置文件獲取，因為Binance API不返回UID）
        try:
            import os
            from dotenv import dotenv_values
            btc_config_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config', 'btc.env')
            if os.path.exists(btc_config_path):
                btc_config = dotenv_values(btc_config_path)
                user_uid = btc_config.get('BINANCE_USER_ID', 'N/A')
                logger.info(f"從配置文件獲取BTC用戶UID: {user_uid}")
            else:
                user_uid = 'N/A'
                logger.warning("BTC配置文件不存在")
        except Exception as e:
            logger.error(f"獲取BTC用戶UID失敗: {e}")
            user_uid = 'N/A'

        organized_account = OrderedDict([
            ('uid', user_uid),                                                       # 幣安用戶UID
            ('walletBalance', f"{wallet_balance:.8f}"),                              # 錢包餘額
            ('availableBalance', f"{available_balance:.8f}"),                        # 可用餘額  
            ('marginBalance', f"{margin_balance:.8f}"),                              # 保證金餘額
            ('unrealizedProfit', f"{unrealized_pnl:.8f}"),                          # 未實現盈虧
            ('initialMargin', f"{initial_margin:.8f}"),                              # 原始保證金（已使用）
            ('maintMargin', f"{maint_margin:.8f}"),                                  # 維持保證金
            ('marginRatio', f"{margin_ratio:.2f}%"),                                    # 保證金率（預設0.00%）
            ('leverageUsage', f"{leverage_usage:.2f}%"),                             # 槓桿使用率（預設0.00%）
            ('todayCommission', f"{today_commission:.8f}"),                          # 手續費
            ('todayPnl', f"{today_pnl:.8f}"),                                        # 本日盈虧
            ('todayPnlPercent', f"{today_pnl_percent:.2f}%"),                       # 本日盈虧百分比
            ('weekPnl', f"{week_pnl:.8f}"),                                          # 7天盈虧
            ('weekPnlPercent', f"{week_pnl_percent:.2f}%"),                         # 7天盈虧百分比
            ('monthPnl', f"{month_pnl:.8f}"),                                        # 30天盈虧
            ('monthPnlPercent', f"{month_pnl_percent:.2f}%"),                       # 30天盈虧百分比
        ])
        
        
        return jsonify({
            'success': True,
            'data': organized_account,  # 前端需要從data字段獲取
            'account': organized_account  # 保持向後兼容
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        })

def get_btc_positions():
    """獲取BTC持倉資訊"""
    try:
        if not binance_client:
            return jsonify({
                'success': False,
                'error': 'BTC客戶端未初始化'
            })
        
        # 獲取持倉資訊
        positions = binance_client.get_position_info()
        if positions is None:
            return jsonify({
                'success': False,
                'error': '無法獲取持倉資訊'
            })
        
        
        # 篩選有效持倉（數量不為0的）並計算真實值
        valid_positions = []
        for pos in positions:
            position_amt = float(pos.get('positionAmt', 0))
            if position_amt != 0:
                
                # 計算真實收益率 (unRealizedProfit / isolatedMargin) * 100
                unrealized_pnl = float(pos.get('unRealizedProfit', 0))
                isolated_margin = float(pos.get('isolatedMargin', 0))
                isolated_wallet = float(pos.get('isolatedWallet', 0))
                notional = float(pos.get('notional', 0))
                
                
                # 根據保證金類型選擇保證金值
                if pos.get('marginType') == 'isolated':
                    margin_value = isolated_margin if isolated_margin > 0 else isolated_wallet
                else:
                    margin_value = isolated_wallet
                
                # 計算真實收益率
                if margin_value > 0:
                    percentage = (unrealized_pnl / margin_value) * 100
                else:
                    percentage = 0
                
                # 計算保證金比例
                if margin_value > 0:
                    margin_ratio = (unrealized_pnl / margin_value) * 100
                else:
                    margin_ratio = 0
                
                # 添加計算欄位
                pos['percentage'] = percentage
                pos['marginRatio'] = margin_ratio
                pos['isolatedMargin'] = margin_value
                
                valid_positions.append(pos)
        
        return jsonify({
            'success': True,
            'positions': valid_positions
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        })

# ========================== BTC風險管理功能 ==========================
# 風險警報發送頻率控制
last_risk_alert_time = {}

def calculate_btc_risk_metrics():
    """計算BTC風險指標"""
    global binance_client, account_info
    
    try:
        # 檢查停止標誌
        if btc_shutdown_flag.is_set():
            return None
            
        if not binance_client or not account_info:
            return None
        
        # 獲取最新帳戶和持倉信息
        fresh_account_info = binance_client.get_account_info()
        if btc_shutdown_flag.is_set():  # API調用後再次檢查
            return None
            
        positions = binance_client.get_position_info()
        if btc_shutdown_flag.is_set():  # API調用後再次檢查
            return None
        
        if not fresh_account_info or not positions:
            return None
        
        # 基本帳戶數據
        total_wallet_balance = float(fresh_account_info.get('totalWalletBalance', 0))
        total_margin_balance = float(fresh_account_info.get('totalMarginBalance', 0))
        total_unrealized_pnl = float(fresh_account_info.get('totalUnrealizedProfit', 0))
        total_initial_margin = float(fresh_account_info.get('totalInitialMargin', 0))
        total_maint_margin = float(fresh_account_info.get('totalMaintMargin', 0))
        available_balance = float(fresh_account_info.get('availableBalance', 0))
        
        # 計算風險指標
        risk_metrics = {
            'total_wallet_balance': total_wallet_balance,
            'total_margin_balance': total_margin_balance,
            'available_balance': available_balance,
            'total_unrealized_pnl': total_unrealized_pnl,
            'total_initial_margin': total_initial_margin,
            'total_maint_margin': total_maint_margin,
            'margin_ratio': 0,
            'risk_level': 'SAFE',
            'leverage_usage': 0,
            'position_value': 0,
            'max_drawable': available_balance,
            'positions': []
        }
        
        # 保證金比率計算
        if total_maint_margin > 0:
            risk_metrics['margin_ratio'] = (total_margin_balance / total_maint_margin) * 100
        
        # 槓桿使用率計算 - 修正：手動計算錢包餘額
        calculated_total_wallet_balance = available_balance + total_initial_margin + total_unrealized_pnl
        if calculated_total_wallet_balance > 0:
            risk_metrics['leverage_usage'] = (total_initial_margin / calculated_total_wallet_balance) * 100
            logger.info(f"風險管理槓桿使用率: {total_initial_margin:.2f} / {calculated_total_wallet_balance:.2f} = {risk_metrics['leverage_usage']:.2f}% (API: {total_wallet_balance:.2f})")
        
        # 處理持倉信息
        total_position_value = 0
        active_positions = []
        
        for pos in positions:
            position_amt = float(pos.get('positionAmt', 0))
            if position_amt != 0:
                entry_price = float(pos.get('entryPrice', 0))
                mark_price = float(pos.get('markPrice', 0))
                unrealized_pnl = float(pos.get('unRealizedProfit', 0))
                percentage = float(pos.get('percentage', 0))
                leverage = float(pos.get('leverage', 1))
                
                position_value = abs(position_amt) * mark_price
                total_position_value += position_value
                
                position_info = {
                    'symbol': pos.get('symbol'),
                    'side': 'LONG' if position_amt > 0 else 'SHORT',
                    'size': abs(position_amt),
                    'entry_price': entry_price,
                    'mark_price': mark_price,
                    'unrealized_pnl': unrealized_pnl,
                    'percentage': percentage,
                    'leverage': leverage,
                    'position_value': position_value,
                    'margin_required': position_value / leverage if leverage > 0 else 0
                }
                active_positions.append(position_info)
        
        risk_metrics['position_value'] = total_position_value
        risk_metrics['positions'] = active_positions
        
        # 風險等級評估
        margin_ratio = risk_metrics['margin_ratio']
        leverage_usage = risk_metrics['leverage_usage']
        
        if margin_ratio > 0:
            if margin_ratio < 120:  # 強制平倉風險高
                risk_metrics['risk_level'] = 'HIGH'
            elif margin_ratio < 200:  # 中等風險
                risk_metrics['risk_level'] = 'MEDIUM'
            else:  # 安全
                risk_metrics['risk_level'] = 'SAFE'
        
        # 槓桿使用率風險
        if leverage_usage > 80:
            if risk_metrics['risk_level'] in ['SAFE', 'MEDIUM']:
                risk_metrics['risk_level'] = 'MEDIUM'
        elif leverage_usage > 90:
            risk_metrics['risk_level'] = 'HIGH'
        
        return risk_metrics
        
    except Exception as e:
        logger.error(f"計算BTC風險指標失敗: {e}")
        return None

def check_btc_risk_alerts():
    """檢查BTC風險警報"""
    try:
        risk_metrics = calculate_btc_risk_metrics()
        if not risk_metrics:
            return
        
        risk_level = risk_metrics['risk_level']
        margin_ratio = risk_metrics['margin_ratio']
        leverage_usage = risk_metrics['leverage_usage']
        total_unrealized_pnl = risk_metrics['total_unrealized_pnl']
        
        # 檢查是否需要發送風險警報
        alerts = []
        
        # 強制平倉風險警報
        if margin_ratio > 0 and margin_ratio < 120:
            alerts.append({
                'type': 'LIQUIDATION_RISK',
                'level': 'CRITICAL',
                'message': f'⚠️ 強制平倉風險警報\n保證金比率: {margin_ratio:.1f}%\n請及時補充保證金或減少持倉！'
            })
        elif margin_ratio > 0 and margin_ratio < 200:
            alerts.append({
                'type': 'MARGIN_WARNING',
                'level': 'WARNING',
                'message': f'⚠️ 保證金警告\n保證金比率: {margin_ratio:.1f}%\n建議關注風險控制'
            })
        
        # 槓桿使用率警報
        if leverage_usage > 90:
            alerts.append({
                'type': 'LEVERAGE_RISK',
                'level': 'WARNING',
                'message': f'⚠️ 槓桿使用率過高\n使用率: {leverage_usage:.1f}%\n建議降低槓桿或增加保證金'
            })
        
        # 大額未實現虧損警報
        if total_unrealized_pnl < -1000:  # 虧損超過1000 USDT
            alerts.append({
                'type': 'LARGE_LOSS',
                'level': 'WARNING',
                'message': f'虧損金額: {total_unrealized_pnl:.2f} USDT\n請考慮風險控制措施'
            })
        
        # 發送警報通知
        for alert in alerts:
            send_btc_risk_alert(alert)
            
        # 記錄風險檢查日誌
        
    except Exception as e:
        logger.error(f"BTC風險檢查失敗: {e}")

def send_btc_risk_alert(alert):
    """發送BTC風險警報通知（1小時發送一次）"""
    try:
        global last_risk_alert_time
        
        current_time = datetime.now()
        alert_type = alert['type']
        
        # 檢查是否在1小時內已發送過相同類型的警報
        if alert_type in last_risk_alert_time:
            time_diff = current_time - last_risk_alert_time[alert_type]
            if time_diff.total_seconds() < 3600:  # 3600秒 = 1小時
                minutes_remaining = int((3600 - time_diff.total_seconds()) / 60)
                logger.info(f"BTC風險警報 {alert_type} 跳過發送，還需等待 {minutes_remaining} 分鐘")
                return False
        
        # 更新發送時間
        last_risk_alert_time[alert_type] = current_time
        
        current_time_str = current_time.strftime('%Y/%m/%d')
        
        # 構建警報訊息
        level_emoji = {
            'CRITICAL': '🚨',
            'WARNING': '⚠️',
            'INFO': 'ℹ️'
        }
        
        emoji = level_emoji.get(alert['level'], '⚠️')
        
        if alert['type'] == 'LEVERAGE_RISK':
            leverage_rate = alert['message'].split('使用率: ')[1].split('%')[0]
            message = f"⚠️ BTC風險管理警報 ({current_time_str})\n槓桿使用率過高，目前使用：{leverage_rate}%\n建議降低槓桿或增加保證金\n請及時處理以避免損失！"
        else:
            message = f"{emoji} BTC風險管理警報 ({current_time_str})\n{alert['message']}\n請及時處理以避免損失！"
        
        # 發送Telegram通知
        send_btc_telegram_message(message)
        
        # 記錄到風險日誌
        risk_log = {
            'timestamp': current_time,
            'alert_type': alert['type'],
            'level': alert['level'],
            'message': alert['message']
        }
        
        logger.info(f"BTC風險警報已發送: {alert['type']} - {alert['level']}")
        
    except Exception as e:
        logger.error(f"發送BTC風險警報失敗: {e}")


def get_btc_risk_status():
    """獲取BTC風險狀態API"""
    try:
        risk_metrics = calculate_btc_risk_metrics()
        if not risk_metrics:
            return jsonify({
                'success': False,
                'message': '無法獲取風險數據'
            })
        
        return jsonify({
            'success': True,
            'risk_metrics': risk_metrics
        })
        
    except Exception as e:
        logger.error(f"獲取BTC風險狀態失敗: {e}")
        return jsonify({
            'success': False,
            'message': f'獲取失敗: {str(e)}'
        })

def start_btc_risk_monitoring():
    """啟動BTC風險監控"""
    logger.info("BTC風險監控已啟動")
    
    while not btc_shutdown_flag.is_set():
        try:
            # 檢查停止標誌後再執行風險檢查
            if not btc_shutdown_flag.is_set():
                check_btc_risk_alerts()
            
            # 使用可中斷的睡眠，每5秒檢查一次停止標誌
            for _ in range(12):  # 12 * 5 = 60秒
                if btc_shutdown_flag.is_set():
                    break
                time.sleep(5)
            
        except Exception as e:
            logger.error(f"BTC風險監控異常: {e}")
            # 錯誤時也使用可中斷的睡眠
            for _ in range(12):  # 12 * 5 = 60秒
                if btc_shutdown_flag.is_set():
                    break
                time.sleep(5)
    
    logger.info("BTC風險監控已停止")

# ========================== BTC WebSocket實時數據 ==========================
import websocket
import ssl

# WebSocket全局變量
btc_ws = None
btc_ws_thread = None
btc_real_time_data = {}
btc_ws_connected = False

# Listen Key心跳維持相關變量
btc_listen_key = None
btc_listen_key_timer = None
btc_listen_key_last_keepalive = None

# WebSocket重連控制變量
btc_ws_reconnect_count = 0
btc_ws_last_reconnect_time = None
btc_ws_max_reconnect_attempts = 10
btc_listen_key_reset_count = 0

def btc_listen_key_keepalive():
    """維持BTC Listen Key心跳 - 每30分鐘執行一次"""
    global btc_listen_key, btc_listen_key_timer, btc_listen_key_last_keepalive, binance_client
    
    try:
        # 檢查模組是否關閉
        if btc_shutdown_flag.is_set():
            logger.info("BTC模組關閉中，取消Listen Key心跳")
            return
        
        # 檢查登入狀態
        env_data = load_btc_env_data()
        if env_data.get('LOGIN_BTC', '0') != '1':
            logger.info("BTC未登入，取消Listen Key心跳")
            return
        
        if not binance_client:
            logger.warning("BTC客戶端未初始化，跳過Listen Key心跳")
            return
        
        # 發送心跳請求續期Listen Key
        try:
            if btc_listen_key:
                # 使用PUT方法延長現有Listen Key
                params = {'listenKey': btc_listen_key}
                response = binance_client._make_request('PUT', '/fapi/v1/listenKey', params, signed=False)
                
                # 檢查響應，空對象或無響應都視為成功
                if response is None or response == {} or isinstance(response, dict):
                    btc_listen_key_last_keepalive = datetime.now()
                    logger.info(f"✅ Listen Key心跳成功 - {btc_listen_key[:10]}... 已續期60分鐘")
                else:
                    raise Exception(f"PUT請求返回意外響應: {response}")
                    
            else:
                logger.warning("Listen Key為空，嘗試重新獲取")
                # 如果沒有Listen Key，直接獲取新的
                new_key_response = binance_client._make_request('POST', '/fapi/v1/listenKey')
                if new_key_response and 'listenKey' in new_key_response:
                    btc_listen_key = new_key_response['listenKey']
                    btc_listen_key_last_keepalive = datetime.now()
                    logger.info(f"🔄 Listen Key重新獲取成功 - {btc_listen_key[:10]}...")
                    
                    # 重新啟動訂單監控
                    restart_btc_order_monitoring()
                else:
                    logger.error("獲取新Listen Key失敗")
        
        except Exception as e:
            logger.error(f"Listen Key心跳請求失敗: {e}")
            # 心跳失敗時，嘗試重新獲取新的Listen Key
            try:
                logger.info("心跳失敗，嘗試重新獲取Listen Key...")
                new_key_response = binance_client._make_request('POST', '/fapi/v1/listenKey')
                if new_key_response and 'listenKey' in new_key_response:
                    old_key = btc_listen_key[:10] if btc_listen_key else 'None'
                    btc_listen_key = new_key_response['listenKey']
                    btc_listen_key_last_keepalive = datetime.now()
                    logger.info(f"🔄 Listen Key重新獲取成功 - {old_key}... -> {btc_listen_key[:10]}...")
                    
                    # 重新啟動訂單監控
                    restart_btc_order_monitoring()
                else:
                    logger.error("重新獲取Listen Key失敗")
            except Exception as retry_e:
                logger.error(f"重新獲取Listen Key異常: {retry_e}")
        
        # 安排下一次心跳（30分鐘後）
        if not btc_shutdown_flag.is_set():
            btc_listen_key_timer = threading.Timer(30 * 60, btc_listen_key_keepalive)
            btc_listen_key_timer.daemon = True
            btc_listen_key_timer.start()
            logger.info("⏰ 下次Listen Key心跳已安排在30分鐘後")
        
    except Exception as e:
        logger.error(f"Listen Key心跳維持異常: {e}")

def start_btc_listen_key_heartbeat():
    """啟動Listen Key心跳維持"""
    global btc_listen_key_timer
    
    try:
        # 停止現有計時器
        stop_btc_listen_key_heartbeat()
        
        # 啟動新的心跳計時器（30分鐘後執行第一次）
        btc_listen_key_timer = threading.Timer(30 * 60, btc_listen_key_keepalive)
        btc_listen_key_timer.daemon = True
        btc_listen_key_timer.start()
        
        btc_listen_key_last_keepalive = datetime.now()
        logger.info("💓 Listen Key心跳維持已啟動（每30分鐘續期一次）")
        
    except Exception as e:
        logger.error(f"啟動Listen Key心跳失敗: {e}")

def stop_btc_listen_key_heartbeat():
    """停止Listen Key心跳維持"""
    global btc_listen_key_timer
    
    try:
        if btc_listen_key_timer:
            btc_listen_key_timer.cancel()
            btc_listen_key_timer = None
            logger.info("💔 Listen Key心跳維持已停止")
    except Exception as e:
        logger.error(f"停止Listen Key心跳失敗: {e}")

def restart_btc_order_monitoring():
    """重新啟動BTC訂單監控（當Listen Key更新時）"""
    try:
        logger.info("🔄 重新啟動BTC訂單監控...")
        # 這裡可以添加重新啟動訂單監控的邏輯
        # 由於WebSocket會自動使用新的Listen Key，通常不需要特別處理
        logger.info("✅ BTC訂單監控重啟完成")
    except Exception as e:
        logger.error(f"重啟BTC訂單監控失敗: {e}")

def reset_btc_listen_key():
    """重置Listen Key - 當檢測到可能過期時"""
    global btc_listen_key, btc_listen_key_last_keepalive, binance_client, btc_listen_key_reset_count
    
    try:
        # 防止過於頻繁的重置
        current_time = datetime.now()
        if hasattr(reset_btc_listen_key, 'last_reset_time'):
            time_since_last = (current_time - reset_btc_listen_key.last_reset_time).total_seconds()
            if time_since_last < 30:  # 30秒內不重複重置
                logger.warning(f"Listen Key重置過於頻繁，跳過（{time_since_last:.1f}秒前剛重置）")
                return False
        
        reset_btc_listen_key.last_reset_time = current_time
        btc_listen_key_reset_count += 1
        
        if not binance_client:
            logger.warning("BTC客戶端未初始化，無法重置Listen Key")
            return False
        
        # 停止現有心跳
        stop_btc_listen_key_heartbeat()
        
        # 嘗試刪除舊的Listen Key
        if btc_listen_key:
            try:
                params = {'listenKey': btc_listen_key}
                binance_client._make_request('DELETE', '/fapi/v1/listenKey', params, signed=False)
                logger.info(f"🗑️ 舊Listen Key已刪除: {btc_listen_key[:10]}...")
            except Exception as e:
                logger.warning(f"刪除舊Listen Key失敗: {e}")
        
        # 獲取新的Listen Key
        try:
            listen_key_response = binance_client._make_request('POST', '/fapi/v1/listenKey')
            if listen_key_response and 'listenKey' in listen_key_response:
                old_key = btc_listen_key[:10] if btc_listen_key else 'None'
                btc_listen_key = listen_key_response['listenKey']
                btc_listen_key_last_keepalive = datetime.now()
                logger.info(f"🔄 Listen Key重置成功: {old_key}... -> {btc_listen_key[:10]}...")
                
                # 重新啟動心跳
                start_btc_listen_key_heartbeat()
                
                return True
            else:
                logger.error("重置Listen Key失敗：無效響應")
                return False
                
        except Exception as e:
            logger.error(f"重置Listen Key失敗: {e}")
            return False
            
    except Exception as e:
        logger.error(f"重置Listen Key異常: {e}")
        return False

def check_listen_key_validity():
    """檢查Listen Key有效性"""
    global btc_listen_key_last_keepalive
    
    if not btc_listen_key_last_keepalive:
        return True  # 剛啟動，假設有效
    
    time_since_keepalive = datetime.now() - btc_listen_key_last_keepalive
    minutes_elapsed = time_since_keepalive.total_seconds() / 60
    
    if minutes_elapsed > 55:  # 55分鐘後認為可能過期
        logger.warning(f"⚠️ Listen Key可能已過期（{minutes_elapsed:.1f}分鐘前最後心跳）")
        return False
    
    return True

def start_btc_websocket():
    """啟動BTC WebSocket連接"""
    global btc_ws, btc_ws_thread, btc_ws_connected
    
    try:
        if btc_ws_connected:
            logger.info("BTC WebSocket已連接")
            return True
        
        # 獲取活躍的交易對
        env_data = load_btc_env_data()
        trading_pair = env_data.get('TRADING_PAIR', 'BTCUSDT').lower()
        
        # 構建WebSocket URL - 幣安期貨WebSocket
        ws_url = f"wss://fstream.binance.com/ws/{trading_pair}@ticker"
        
        
        # 創建WebSocket連接
        btc_ws = websocket.WebSocketApp(
            ws_url,
            on_open=on_btc_ws_open,
            on_message=on_btc_ws_message,
            on_error=on_btc_ws_error,
            on_close=on_btc_ws_close
        )
        
        # 在新線程中運行WebSocket
        btc_ws_thread = threading.Thread(
            target=btc_ws.run_forever,
            kwargs={'sslopt': {"cert_reqs": ssl.CERT_NONE}},
            daemon=True,
            name="BTC-WebSocket"
        )
        register_btc_thread(btc_ws_thread, "BTC-WebSocket")
        btc_ws_thread.start()
        
        logger.info("BTC WebSocket線程已啟動")
        return True
        
    except Exception as e:
        logger.error(f"啟動BTC WebSocket失敗: {e}")
        return False

def on_btc_ws_open(ws):
    """WebSocket連接開啟"""
    global btc_ws_connected
    btc_ws_connected = True
    logger.info("BTC WebSocket連接成功")

def on_btc_ws_message(ws, message):
    """處理WebSocket訊息"""
    global btc_real_time_data
    
    try:
        data = json.loads(message)
        
        # 更新實時數據
        btc_real_time_data.update({
            'symbol': data.get('s'),
            'price': float(data.get('c', 0)),  # 最新價格
            'price_change': float(data.get('P', 0)),  # 24小時價格變化百分比
            'price_change_abs': float(data.get('p', 0)),  # 24小時價格變化金額
            'high_price': float(data.get('h', 0)),  # 24小時最高價
            'low_price': float(data.get('l', 0)),  # 24小時最低價
            'volume': float(data.get('v', 0)),  # 24小時交易量
            'quote_volume': float(data.get('q', 0)),  # 24小時交易額
            'timestamp': datetime.now().isoformat()
        })
        
        # 計算更新後的持倉盈虧（如果有持倉）
        update_btc_position_pnl()
        
    except Exception as e:
        logger.error(f"處理BTC WebSocket訊息失敗: {e}")

def on_btc_ws_error(ws, error):
    """WebSocket錯誤"""
    logger.error(f"BTC WebSocket錯誤: {error}")

def on_btc_ws_close(ws, close_status_code, close_msg):
    """WebSocket連接關閉"""
    global btc_ws_connected
    btc_ws_connected = False
    
    logger.warning(f"BTC WebSocket連接關閉 - 狀態碼: {close_status_code}, 訊息: {close_msg}")
    
    # 檢查是否可能是Listen Key過期導致的斷線
    if btc_listen_key_last_keepalive:
        time_since_keepalive = datetime.now() - btc_listen_key_last_keepalive
        if time_since_keepalive.total_seconds() > 3300:  # 55分鐘
            logger.warning("⚠️ Listen Key可能已過期，將重新獲取")
            # 強制重新獲取Listen Key
            reset_btc_listen_key()
    
    # 只有在未設置停止標誌時才嘗試重新連接
    if not btc_shutdown_flag.is_set():
        threading.Timer(5.0, reconnect_btc_websocket).start()

def reconnect_btc_websocket():
    """重新連接WebSocket"""
    # 檢查是否已設置停止標誌
    if btc_shutdown_flag.is_set():
        logger.info("BTC模組正在關閉，取消WebSocket重連")
        return
        
    logger.info("嘗試重新連接BTC WebSocket...")
    start_btc_websocket()

def stop_btc_websocket():
    """停止BTC WebSocket連接"""
    global btc_ws, btc_ws_connected
    
    try:
        if btc_ws:
            btc_ws_connected = False
            btc_ws.close()
            logger.info("BTC WebSocket連接已關閉")
    except Exception as e:
        logger.error(f"關閉BTC WebSocket失敗: {e}")

def update_btc_position_pnl():
    """根據實時價格更新持倉盈虧"""
    global binance_client, btc_real_time_data
    
    try:
        if not binance_client or not btc_real_time_data:
            return
        
        current_price = btc_real_time_data.get('price', 0)
        if current_price <= 0:
            return
        
        # 獲取持倉信息（簡化版，只獲取基本數據）
        positions = binance_client.get_position_info()
        if not positions:
            return
        
        # 更新實時盈虧數據
        for pos in positions:
            position_amt = float(pos.get('positionAmt', 0))
            if position_amt != 0:
                entry_price = float(pos.get('entryPrice', 0))
                if entry_price > 0:
                    # 計算實時盈虧
                    if position_amt > 0:  # 多頭
                        pnl = (current_price - entry_price) * position_amt
                    else:  # 空頭
                        pnl = (entry_price - current_price) * abs(position_amt)
                    
                    # 更新實時數據
                    btc_real_time_data[f'position_pnl_{pos.get("symbol")}'] = pnl
        
    except Exception as e:
        logger.error(f"更新BTC持倉盈虧失敗: {e}")

def check_btc_api_connection():
    """檢查BTC API連接狀態"""
    global binance_client, btc_last_connection_check
    
    try:
        if not binance_client:
            return False
            
        # 嘗試獲取帳戶信息來測試連接
        account_info = binance_client.get_account_info()
        if account_info and 'totalWalletBalance' in account_info:
            btc_last_connection_check = time.time()
            return True
        else:
            return False
            
    except Exception as e:
        logger.error(f"BTC API連接檢查失敗: {e}")
        return False

def btc_reconnect_api():
    """BTC API重連功能"""
    global btc_is_reconnecting, btc_reconnect_attempts, binance_client
    
    if btc_is_reconnecting:
        logger.info("BTC API重連已在進行中，跳過此次重連")
        return False
        
    btc_is_reconnecting = True
    logger.info(f"開始BTC API重連嘗試 (第{btc_reconnect_attempts + 1}次)...")
    
    try:
        # 檢查登入狀態
        env_data = load_btc_env_data()
        if env_data.get('LOGIN_BTC', '0') != '1':
            logger.info("BTC未登入狀態，停止重連嘗試")
            btc_is_reconnecting = False
            return False
        
        # 嘗試重新建立連接
        api_key = env_data.get('BINANCE_API_KEY', '').strip()
        secret_key = env_data.get('BINANCE_SECRET_KEY', '').strip()
        
        if not api_key or not secret_key:
            logger.info("BTC API密鑰不存在，停止重連")
            btc_is_reconnecting = False
            return False
        
        # 重新創建BTC客戶端
        binance_client = BinanceClient(api_key, secret_key)
        
        # 檢查連接
        if check_btc_api_connection():
            logger.info("BTC API重連成功！")
            btc_reconnect_attempts = 0  # 重置重連次數
            btc_is_reconnecting = False
            return True
        else:
            btc_reconnect_attempts += 1
            logger.error(f"BTC API重連失敗，將在30秒後重試 (嘗試次數: {btc_reconnect_attempts})")
            
            # 檢查是否超過最大重連次數（實際上是無限重連）
            if btc_reconnect_attempts < btc_max_reconnect_attempts:
                # 30秒後重試
                threading.Timer(30.0, btc_reconnect_api).start()
            
            btc_is_reconnecting = False
            return False
            
    except Exception as e:
        btc_reconnect_attempts += 1
        logger.error(f"BTC API重連異常: {e}，將在30秒後重試 (嘗試次數: {btc_reconnect_attempts})")
        
        if btc_reconnect_attempts < btc_max_reconnect_attempts:
            threading.Timer(30.0, btc_reconnect_api).start()
        
        btc_is_reconnecting = False
        return False

def start_btc_connection_monitor():
    """啟動BTC連接監控線程"""
    def connection_monitor():
        global btc_last_connection_check
        btc_connection_notified = False  # 追蹤是否已發送斷線通知
        
        while not btc_shutdown_flag.is_set():
            try:
                # 檢查是否登入
                env_data = load_btc_env_data()
                if env_data.get('LOGIN_BTC', '0') != '1':
                    # 可中斷睡眠
                    for _ in range(btc_connection_check_interval):
                        if btc_shutdown_flag.is_set():
                            return
                        time.sleep(5)
                    continue
                
                # 檢查連接狀態
                if not check_btc_api_connection():
                    # 檢測到斷線，發送TG通知（只發送一次）
                    if not btc_connection_notified:
                        logger.info("檢測到BTC API斷線，發送TG通知...")
                        send_btc_telegram_message("⚠️ API連線異常！！！\n正在嘗試重新連線．．．")
                        btc_connection_notified = True
                    
                    logger.info("檢測到BTC API斷線，開始重連...")
                    if btc_reconnect_api():
                        # 重連成功，發送TG通知
                        logger.info("BTC API重連成功，發送TG通知...")
                        send_btc_telegram_message("✅ API連線成功！！！")
                        btc_connection_notified = False  # 重置通知狀態
                else:
                    # 連線正常，重置通知狀態
                    btc_connection_notified = False
                
                # 可中斷的等待下次檢查
                for _ in range(btc_connection_check_interval):
                    if btc_shutdown_flag.is_set():
                        return
                    time.sleep(5)
                
            except Exception as e:
                logger.error(f"BTC連接監控異常: {e}")
                # 可中斷的錯誤恢復睡眠
                for _ in range(btc_connection_check_interval):
                    if btc_shutdown_flag.is_set():
                        return
                    time.sleep(5)
    
    # 創建並啟動監控線程
    monitor_thread = threading.Thread(target=connection_monitor, name="BTC連接監控線程", daemon=True)
    monitor_thread.start()
    btc_active_threads.append(monitor_thread)
    logger.info("BTC連接監控已啟動")

def start_btc_auto_logout_timer():
    """啟動BTC自動登出計時器（每12小時）"""
    global btc_auto_logout_timer
    
    def btc_auto_logout_task():
        global btc_auto_logout_timer
        
        try:
            logger.info("BTC 12小時自動登出重連開始...")
            
            # 檢查是否仍在登入狀態
            env_data = load_btc_env_data()
            if env_data.get('LOGIN_BTC', '0') != '1':
                logger.info("BTC已登出，取消自動重連")
                return
            
            # 嘗試重新連接
            if btc_reconnect_api():
                logger.info("BTC自動重連成功")
            else:
                logger.info("BTC自動重連失敗，將持續重試")
            
            # 設定下次12小時後執行
            btc_auto_logout_timer = threading.Timer(12 * 3600, btc_auto_logout_task)
            btc_auto_logout_timer.daemon = True
            btc_auto_logout_timer.start()
            
        except Exception as e:
            logger.error(f"BTC自動登出重連異常: {e}")
            
            # 即使異常也要設定下次執行
            btc_auto_logout_timer = threading.Timer(12 * 3600, btc_auto_logout_task)
            btc_auto_logout_timer.daemon = True
            btc_auto_logout_timer.start()
    
    # 停止現有計時器
    stop_btc_auto_logout_timer()
    
    # 啟動新的計時器（12小時後執行）
    btc_auto_logout_timer = threading.Timer(12 * 3600, btc_auto_logout_task)
    btc_auto_logout_timer.daemon = True
    btc_auto_logout_timer.start()
    logger.info("BTC自動登出計時器已啟動（12小時後執行）")

def stop_btc_auto_logout_timer():
    """停止BTC自動登出計時器"""
    global btc_auto_logout_timer
    
    try:
        if btc_auto_logout_timer and btc_auto_logout_timer.is_alive():
            btc_auto_logout_timer.cancel()
            logger.info("BTC自動登出計時器已停止")
    except Exception as e:
        logger.error(f"停止BTC自動登出計時器失敗: {e}")
    finally:
        btc_auto_logout_timer = None

def get_btc_realtime_data():
    """獲取BTC實時數據API"""
    global btc_real_time_data, btc_ws_connected
    
    try:
        return jsonify({
            'success': True,
            'connected': btc_ws_connected,
            'data': btc_real_time_data,
            'timestamp': datetime.now().isoformat()
        })
        
    except Exception as e:
        logger.error(f"獲取BTC實時數據失敗: {e}")
        return jsonify({
            'success': False,
            'message': f'獲取失敗: {str(e)}'
        })

def get_btc_listen_key_status():
    """獲取BTC Listen Key狀態監控API"""
    global btc_listen_key, btc_listen_key_last_keepalive, btc_listen_key_reset_count, btc_ws_reconnect_count
    
    try:
        current_time = datetime.now()
        
        # 計算Listen Key狀態
        if btc_listen_key_last_keepalive:
            time_since_keepalive = current_time - btc_listen_key_last_keepalive
            minutes_since_keepalive = time_since_keepalive.total_seconds() / 60
            
            if minutes_since_keepalive > 55:
                key_status = "expired"
                key_health = "critical"
            elif minutes_since_keepalive > 45:
                key_status = "expiring_soon"
                key_health = "warning"
            else:
                key_status = "active"
                key_health = "healthy"
        else:
            key_status = "unknown"
            key_health = "unknown"
            minutes_since_keepalive = None
        
        return jsonify({
            'success': True,
            'listen_key': {
                'exists': btc_listen_key is not None,
                'key_preview': btc_listen_key[:10] + "..." if btc_listen_key else None,
                'status': key_status,
                'health': key_health,
                'last_keepalive': btc_listen_key_last_keepalive.isoformat() if btc_listen_key_last_keepalive else None,
                'minutes_since_keepalive': round(minutes_since_keepalive, 1) if minutes_since_keepalive is not None else None,
                'reset_count': btc_listen_key_reset_count
            },
            'websocket': {
                'reconnect_count': btc_ws_reconnect_count,
                'connected': btc_ws_connected
            },
            'heartbeat': {
                'timer_active': btc_listen_key_timer is not None and btc_listen_key_timer.is_alive() if btc_listen_key_timer else False
            },
            'timestamp': current_time.isoformat()
        })
        
    except Exception as e:
        logger.error(f"獲取Listen Key狀態失敗: {e}")
        return jsonify({
            'success': False,
            'message': f'狀態獲取失敗: {str(e)}'
        })

# ========== BTC訂單監控系統 ==========

def start_btc_order_monitoring():
    """啟動BTC訂單監控WebSocket"""
    global order_monitor_thread, user_data_stream_key
    
    try:
        if not binance_client:
            logger.error("BTC客戶端未初始化，無法啟動訂單監控")
            return False
        
        # 獲取User Data Stream Listen Key
        try:
            listen_key_response = binance_client._make_request('POST', '/fapi/v1/listenKey')
            if not listen_key_response or 'listenKey' not in listen_key_response:
                logger.error("獲取Listen Key失敗")
                return False
            
            user_data_stream_key = listen_key_response['listenKey']
            logger.info(f"BTC Listen Key獲取成功: {user_data_stream_key[:10]}...")
            
            # 保存Listen Key到全局變量
            global btc_listen_key
            btc_listen_key = user_data_stream_key
            
            # 啟動Listen Key心跳維持
            start_btc_listen_key_heartbeat()
            
        except Exception as e:
            logger.error(f"獲取Listen Key失敗: {e}")
            return False
        
        # 啟動WebSocket監控線程
        order_monitor_thread = threading.Thread(
            target=btc_order_websocket_worker,
            name="BTCOrderMonitor",
            daemon=True
        )
        order_monitor_thread.start()
        btc_active_threads.append(order_monitor_thread)
        
        # 啟動備用監控
        start_order_fallback_monitor()
        
        logger.info("BTC訂單監控WebSocket已啟動")
        return True
        
    except Exception as e:
        logger.error(f"啟動BTC訂單監控失敗: {e}")
        return False

def btc_order_websocket_worker():
    """
BTC訂單監控WebSocket工作線程"""
    global order_monitor_ws, user_data_stream_key, btc_listen_key
    
    while not btc_shutdown_flag.is_set():
        try:
            # 檢查Listen Key有效性
            if not btc_listen_key or not check_listen_key_validity():
                logger.warning("Listen Key無效或可能過期，嘗試重置")
                if not reset_btc_listen_key():
                    logger.error("Listen Key重置失敗，等待30秒後重試")
                    # 可中斷的30秒睡眠
                    for _ in range(30):
                        if btc_shutdown_flag.is_set():
                            return
                        time.sleep(5)
                    continue
            
            # 使用全局Listen Key
            current_listen_key = btc_listen_key
            ws_url = f"wss://fstream.binance.com/ws/{current_listen_key}"
            logger.info(f"連接BTC訂單監控WebSocket: {ws_url[:50]}...")
            
            import websocket
            
            def on_message(ws, message):
                try:
                    data = json.loads(message)
                    
                    # 處理訂單更新事件
                    if data.get('e') == 'ORDER_TRADE_UPDATE':
                        order_data = data['o']
                        order_id = str(order_data['i'])
                        status = order_data['X']  # 訂單狀態
                        symbol = order_data['s']  # 交易對
                        side = order_data['S']  # BUY/SELL
                        order_type = order_data['o']  # MARKET/LIMIT
                        quantity = order_data['q']  # 訂單數量
                        price = float(order_data.get('p', 0))  # 訂單價格
                        stop_price = float(order_data.get('sp', 0))  # 止盈止損觸發價格
                        reduce_only = order_data.get('R', False)  # 是否為reduceOnly
                        
                        logger.info(f"收到訂單更新: {order_id}, 狀態: {status}, 交易對: {symbol}, 類型: {order_type}, reduceOnly: {reduce_only}")
                        
                        # 檢查是否已處理過此訂單
                        order_key = f"{order_id}_{status}"
                        with processed_orders_lock:
                            if order_key in processed_orders:
                                logger.info(f"訂單{order_id}狀態{status}已處理過，跳過此訂單")
                                return  # 只跳過這個訂單，不影響後續訊息處理
                        
                        # 標記為已處理
                        with processed_orders_lock:
                            processed_orders.add(order_key)
                        
                        # 檢查是否為系統已知訂單
                        is_system_order = (str(order_id) in pending_orders or 
                                         str(order_id) in btc_active_trades)
                        
                        # 智能判斷：根據訂單特徵推斷是否為WEBHOOK
                        if is_system_order:
                            order_source = '系統'  # 已知系統訂單
                        else:
                            # 如果不在映射中，根據訂單特徵智能判斷
                            # 對於非系統已知訂單，優先判斷為手動（更保守的判斷邏輯）
                            if order_type in ['TAKE_PROFIT_MARKET', 'STOP_MARKET']:
                                # 止盈止損訂單通常是手動設置
                                order_source = '手動'
                                logger.info(f"根據WebSocket訂單特徵推斷為手動交易: {order_type}")
                            else:
                                # 其他所有情況（包括市價單、限價單、平倉單）都優先判斷為手動
                                # 只有快訊系統產生的訂單才應該被識別為WEBHOOK（通過系統訂單列表識別）
                                order_source = '手動'
                                order_desc = convert_order_type_to_display(order_type)
                                action_desc = "平倉" if reduce_only else "開倉"
                                logger.info(f"根據WebSocket訂單特徵推斷為手動交易: {action_desc}{order_desc}")
                        
                        logger.info(f"訂單{order_id}判斷: pending_orders中有={str(order_id) in pending_orders}, active_trades中有={str(order_id) in btc_active_trades}, 判定為={order_source}")
                        logger.info(f"當前pending_orders內容: {list(pending_orders.keys())}")
                        logger.info(f"當前btc_active_trades內容: {list(btc_active_trades.keys())}")
                        
                        # 🔥 新增：檢測訂單修改
                        if order_id in btc_order_history:
                            old_order = btc_order_history[order_id]
                            current_price = float(price)
                            current_quantity = float(quantity)
                            
                            # 檢查是否有價格或數量變更
                            price_changed = abs(old_order.get('price', 0) - current_price) > 0.01
                            quantity_changed = abs(old_order.get('quantity', 0) - current_quantity) > 0.001
                            
                            if (price_changed or quantity_changed) and status not in ['FILLED', 'CANCELED', 'REJECTED', 'EXPIRED']:
                                logger.info(f"🔄 檢測到訂單修改: {order_id}, 價格變更: {price_changed}, 數量變更: {quantity_changed}")
                                
                                # 準備修改通知數據
                                old_data = {
                                    'price': old_order.get('price', 0),
                                    'quantity': old_order.get('quantity', 0)
                                }
                                new_data = {
                                    'price': current_price,
                                    'quantity': current_quantity
                                }
                                
                                # 發送修改通知
                                try:
                                    send_btc_order_modify_notification(order_data, old_data, new_data)
                                    logger.info(f"✅ 訂單修改通知已發送: {order_id}")
                                except Exception as e:
                                    logger.error(f"❌ 發送訂單修改通知失敗: {e}")
                        
                        # 更新訂單歷史記錄
                        btc_order_history[order_id] = {
                            'price': float(price),
                            'quantity': float(quantity),
                            'status': status,
                            'side': side,
                            'order_type': order_type,
                            'symbol': symbol
                        }
                        
                        if status == 'NEW':
                            # 新訂單提交
                            if symbol == 'BTCUSDT':
                                logger.info(f"🔔 WebSocket收到新訂單: {order_id}, 來源判定: {order_source}")
                                handle_manual_binance_order(order_data, True)
                            else:
                                logger.info(f"WebSocket跳過非BTCUSDT訂單: {order_id}, symbol={symbol}")
                                
                        elif status == 'FILLED':
                            # 訂單完全成交
                            fill_price = float(order_data.get('ap', 0))  # 平均成交價
                            fill_quantity = order_data.get('z', '0')  # 成交數量
                            
                            if symbol == 'BTCUSDT':
                                logger.info(f"🔔 WebSocket收到成交通知: {order_id}, 成交價: {fill_price}, 數量: {fill_quantity}")
                                if is_system_order:
                                    # 系統訂單成交 - 使用現有邏輯
                                    process_order_fill(order_id, fill_price, fill_quantity)
                                
                                # 所有BTCUSDT訂單都發送成交通知
                                handle_manual_binance_fill(order_data)
                            
                            # 清理成交訂單的歷史記錄
                            if order_id in btc_order_history:
                                del btc_order_history[order_id]
                                logger.info(f"🧹 已清理成交訂單歷史記錄: {order_id}")
                                
                        elif status == 'REJECTED':
                            # 訂單被拒絕
                            if symbol == 'BTCUSDT':
                                logger.info(f"🔔 WebSocket收到訂單拒絕: {order_id}")
                                # 創建失敗記錄並發送失敗通知
                                reject_reason = order_data.get('r', '未知原因')
                                error_record = {
                                    'symbol': symbol,
                                    'side': side,
                                    'quantity': quantity,
                                    'price': price,
                                    'order_id': order_id,
                                    'order_type': convert_order_type_to_display(order_type),
                                    'source': 'manual',
                                    'action_type': '平倉' if reduce_only else '開倉',
                                    'reduceOnly': reduce_only,
                                    'is_manual': True,
                                    'error': reject_reason
                                }
                                handle_manual_binance_order(error_record, False)
                                    
                        elif status in ['CANCELED', 'REJECTED', 'EXPIRED']:
                            # 🔥 新增：處理訂單失效狀態 - 掛單沒成功，發送提交失敗通知
                            logger.info(f"🔴 WebSocket收到訂單失效: {order_id}, 狀態: {status}")
                            logger.info(f"🔍 WebSocket訂單數據: {order_data}")
                            
                            try:
                                # 構建失敗通知記錄
                                status_mapping = {
                                    'CANCELED': '訂單被取消',
                                    'REJECTED': '訂單被拒絕', 
                                    'EXPIRED': '訂單已過期'
                                }
                                failure_reason = status_mapping.get(status, f'訂單失效({status})')
                                
                                # 解析訂單詳情
                                reduce_only = order_data.get('R', False)
                                action_type = '平倉' if reduce_only else '開倉'
                                
                                # 構建失敗記錄，包含價格和止盈止損信息
                                error_record = {
                                    'symbol': symbol,
                                    'side': side,
                                    'quantity': quantity,
                                    'price': price,  # 添加價格信息
                                    'order_id': order_id,
                                    'order_type': convert_order_type_to_display(order_type),
                                    'source': order_source,
                                    'action_type': action_type,
                                    'reduceOnly': reduce_only,
                                    'is_manual': not is_system_order,
                                    'error': failure_reason
                                }
                                
                                # 🔥 新增：如果有止盈止損價格，加入記錄中
                                if stop_price > 0:
                                    error_record['stopPrice'] = stop_price
                                    logger.info(f"🎯 WebSocket失效訂單包含止盈止損價格: {stop_price}")
                                
                                # 使用排隊系統發送提交失敗通知（5秒延遲）
                                logger.info(f"🚀 WebSocket準備將BTC失效通知加入排隊: {error_record}")
                                btc_notification_manager.add_notification('submit', error_record, False, 5)
                                logger.info(f"📱 WebSocket訂單失效提交失敗通知已加入排隊系統: {order_id} - {failure_reason}")
                                
                                # 如果是系統訂單，從pending_orders中移除
                                if is_system_order and str(order_id) in pending_orders:
                                    remove_pending_order(order_id)
                                
                                # 清理訂單歷史記錄
                                if order_id in btc_order_history:
                                    del btc_order_history[order_id]
                                    logger.info(f"🧹 已清理訂單歷史記錄: {order_id}")
                                
                                # 🔥 新增：清理processed_orders中的相關記錄，允許重新提交相同訂單
                                orders_to_remove = []
                                for processed_key in processed_orders:
                                    if processed_key.startswith(f"{order_id}_"):
                                        orders_to_remove.append(processed_key)
                                
                                for key in orders_to_remove:
                                    processed_orders.discard(key)
                                    logger.info(f"🧹 已清理處理記錄: {key}")
                                
                                logger.info(f"🔄 訂單 {order_id} 已完全清理，允許重新提交")
                                    
                            except Exception as e:
                                logger.error(f"WebSocket處理訂單失效通知失敗: {e}")
                            
                except Exception as e:
                    logger.error(f"處理WebSocket訊息失敗: {e}")
            
            def on_error(ws, error):
                error_str = str(error)
                logger.error(f"BTC訂單監控WebSocket錯誤: {error_str}")
                
                # 檢查是否為Listen Key相關錯誤
                if "1125" in error_str or "listenKey" in error_str.lower() or "does not exist" in error_str.lower():
                    logger.warning("🔑 檢測到Listen Key錯誤，將重新獲取")
                    reset_btc_listen_key()
            
            def on_close(ws, close_status_code, close_msg):
                close_msg_str = str(close_msg) if close_msg else ""
                logger.warning(f"BTC訂單監控WebSocket關閉: {close_status_code}, {close_msg_str}")
                
                # 檢查是否為Listen Key過期導致的關閉
                if close_status_code == 1006 or "listenkey" in close_msg_str.lower():
                    logger.warning("🔑 WebSocket關閉可能與Listen Key有關，將重新獲取")
                    reset_btc_listen_key()
            
            def on_open(ws):
                global btc_ws_reconnect_count
                logger.info("BTC訂單監控WebSocket連接成功")
                btc_ws_reconnect_count = 0  # 重置重連計數
            
            # 創建並啟動WebSocket
            order_monitor_ws = websocket.WebSocketApp(
                ws_url,
                on_message=on_message,
                on_error=on_error,
                on_close=on_close,
                on_open=on_open
            )
            
            order_monitor_ws.run_forever()
            
        except Exception as e:
            logger.error(f"BTC訂單監控WebSocket異常: {e}")
            
        # 指數退避重連機制
        if not btc_shutdown_flag.is_set():
            global btc_ws_reconnect_count
            btc_ws_reconnect_count += 1
            
            # 計算退避時間：5, 10, 20, 40, 60秒（最大）
            backoff_time = min(5 * (2 ** (btc_ws_reconnect_count - 1)), 60)
            logger.info(f"等待{backoff_time}秒後重新連接BTC訂單監控WebSocket（重連次數: {btc_ws_reconnect_count}）...")
            
            # 可中斷的退避睡眠
            for _ in range(int(backoff_time)):
                if btc_shutdown_flag.is_set():
                    return
                time.sleep(5)

def check_pending_orders_fallback():
    """輪詢檢查待成交訂單（備用方案）"""
    global pending_orders
    
    try:
        if not binance_client:
            return
            
        if not pending_orders:
            return
            
        logger.info(f"檢查 {len(pending_orders)} 個待成交訂單...")
        
        orders_to_remove = []
        
        for order_id, order_info in list(pending_orders.items()):
            try:
                logger.info(f"檢查訂單 {order_id} 狀態...")
                
                # 查詢訂單狀態
                order_status = binance_client.get_order_status('BTCUSDT', order_id=order_id)
                
                if order_status:
                    current_status = order_status.get('status')
                    logger.info(f"訂單 {order_id} 狀態: {current_status}")
                    
                    if current_status == 'FILLED':
                        # 訂單成交
                        fill_price = float(order_status.get('avgPrice', 0))
                        fill_quantity = order_status.get('executedQty', order_info['quantity'])
                        
                        logger.info(f"訂單 {order_id} 已成交: 價格={fill_price}, 數量={fill_quantity}")
                        
                        # 處理成交事件
                        process_order_fill(order_id, fill_price, fill_quantity)
                        orders_to_remove.append(order_id)
                        
                    elif current_status in ['CANCELED', 'REJECTED', 'EXPIRED']:
                        # 訂單已取消或失敗
                        logger.info(f"訂單 {order_id} 已取消/失敗: {current_status}")
                        orders_to_remove.append(order_id)
                        
                else:
                    logger.warning(f"無法獲取訂單 {order_id} 狀態")
                    
            except Exception as e:
                logger.error(f"檢查訂單 {order_id} 狀態失敗: {e}")
        
        # 移除已處理的訂單
        for order_id in orders_to_remove:
            remove_pending_order(order_id)
            
    except Exception as e:
        logger.error(f"輪詢檢查待成交訂單失敗: {e}")

def handle_manual_binance_order(order_data, is_success=True):
    """處理手動幣安訂單（提交階段）"""
    try:
        order_id = str(order_data['i'])
        symbol = order_data['s']
        side = order_data['S']  # BUY/SELL
        order_type = order_data['o']  # MARKET/LIMIT
        quantity = order_data['q']
        reduce_only = order_data.get('R', False)  # 是否為reduceOnly
        
        # 解析動作和方向 - 平倉時方向相反
        if reduce_only:
            # 平倉：SELL = 平多單，BUY = 平空單（與get_btc_action_and_direction邏輯一致）
            direction = '多單' if side == 'SELL' else '空單'
            parsed_action = '平倉'
        else:
            # 開倉：BUY = 多單，SELL = 空單
            direction = '多單' if side == 'BUY' else '空單'
            parsed_action = '開倉'
        order_source = '待判斷'  # 初始設定，稍後會根據智能判斷更新
        formatted_quantity = f"{float(quantity):.8f}"
        order_type_text = convert_order_type_to_display(order_type)
        
        # 獲取提交價格
        submitted_price = 0
        if order_type == 'LIMIT':
            submitted_price = float(order_data.get('p', 0))
            logger.info(f"手動限價單提交價格: {submitted_price}")
        else:
            # 市價單獲取當前市價作為參考
            try:
                ticker = binance_client.get_ticker_price(symbol=symbol)
                if ticker and 'price' in ticker:
                    submitted_price = float(ticker['price'])
                    logger.info(f"手動市價單參考價格獲取成功: {submitted_price}")
                else:
                    logger.warning(f"手動獲取市價失敗，ticker回應: {ticker}")
                    submitted_price = 0
            except Exception as e:
                logger.error(f"手動獲取市價異常: {e}")
                submitted_price = 0
        
        logger.info(f"手動訂單最終設置的提交價格: {submitted_price}")
        
        # 智能判斷：根據訂單特徵推斷是否為WEBHOOK（與主要WebSocket邏輯一致）
        is_manual = None  # 初始為None，待智能判斷
        
        # 對於非系統已知訂單，優先判斷為手動（更保守的判斷邏輯）
        # 只有非常明確的WEBHOOK特徵才判定為自動交易
        if order_type in ['TAKE_PROFIT_MARKET', 'STOP_MARKET']:
            # 止盈止損訂單通常是手動設置
            is_manual = True
            order_source = '手動'
            logger.info(f"根據訂單特徵推斷為手動交易: {order_type}")
        else:
            # 其他所有情況（包括市價單、限價單、平倉單）都優先判斷為手動
            # 只有快訊系統產生的訂單才應該被識別為WEBHOOK（通過系統訂單列表識別）
            is_manual = True
            order_source = '手動'
            order_desc = "市價單" if order_type == 'MARKET' else "限價單"
            action_desc = "平倉" if reduce_only else "開倉"
            logger.info(f"根據訂單特徵推斷為手動交易: {action_desc}{order_desc}")
        
        # 記錄後端日誌，包含智能判斷結果
        price_display = "市價" if order_type == 'MARKET' else f"{float(order_data.get('p', 0)):,.2f}"
        commit_log = f"{order_source}{parsed_action}：{direction}｜{formatted_quantity} BTC｜{price_display} USDT｜{order_type_text}"
        logger.info(f"({order_source}委託) {commit_log}")
        
        # 設置顯示文字
        order_source_display = "手動" if is_manual else "WEBHOOK"
        
        # 構建trade_record用於通知
        trade_record = {
            'symbol': symbol,
            'side': side,
            'quantity': formatted_quantity,
            'price': submitted_price,  # 添加提交價格
            'order_id': order_id,
            'order_type': order_type_text,
            'source': 'manual' if is_manual else 'webhook',  # 根據智能判斷設置source
            'action_type': parsed_action,
            'reduceOnly': reduce_only,
            'is_manual': is_manual
        }
        
        # 保存訂單記錄到tradedataBTC（與TX系統一致）
        try:
            save_btc_order_record(order_data, is_manual)
            logger.info(f"✅ {order_source_display}訂單記錄已保存到tradedataBTC")
        except Exception as e:
            logger.error(f"保存{order_source_display}訂單記錄失敗: {e}")
        
        # 使用排隊系統發送提交通知（1秒延遲）
        logger.info(f"準備將{order_source_display}幣安訂單提交通知加入排隊: {trade_record}")
        btc_notification_manager.add_notification('submit', trade_record, is_success, 1)
        logger.info(f"{order_source_display}幣安訂單提交通知已加入排隊系統（延遲1秒）")
        
    except Exception as e:
        logger.error(f"處理手動幣安訂單失敗: {e}")

def handle_manual_binance_fill(order_data):
    """處理手動幣安訂單成交"""
    try:
        order_id = str(order_data['i'])
        symbol = order_data['s']
        side = order_data['S']
        quantity = order_data['z']  # 成交數量
        avg_price = float(order_data.get('ap', 0))  # 平均成交價
        order_type = order_data['o']
        reduce_only = order_data.get('R', False)
        
        # 解析動作和方向 - 平倉時方向相反
        if reduce_only:
            # 平倉：SELL = 平多單，BUY = 平空單（與get_btc_action_and_direction邏輯一致）
            direction = '多單' if side == 'SELL' else '空單'
            action = '平倉'
        else:
            # 開倉：BUY = 多單，SELL = 空單
            direction = '多單' if side == 'BUY' else '空單'
            action = '開倉'
        
        # 智能判斷：根據訂單特徵推斷是否為WEBHOOK（與提交階段邏輯一致）
        # 對於非系統已知訂單，優先判斷為手動（更保守的判斷邏輯）
        if order_type in ['TAKE_PROFIT_MARKET', 'STOP_MARKET']:
            # 止盈止損訂單通常是手動設置
            order_source = '手動'
            logger.info(f"根據成交訂單特徵推斷為手動交易: {order_type}")
        else:
            # 其他所有情況（包括市價單、限價單、平倉單）都優先判斷為手動
            # 只有快訊系統產生的訂單才應該被識別為WEBHOOK（通過系統訂單列表識別）
            order_source = '手動'
            order_desc = "市價單" if order_type == 'MARKET' else "限價單"
            action_desc = "平倉" if reduce_only else "開倉"
            logger.info(f"根據成交訂單特徵推斷為手動交易: {action_desc}{order_desc}")
        
        order_type_text = convert_order_type_to_display(order_type)
        
        # 前端日誌記錄（成交）- 暫時移除立即輸出，改為延遲輸出
        
        # 🔥 新增：BTC JSON交易記錄系統（含智能來源判斷）
        try:
            # 判斷交易類型並記錄到JSON配對系統
            trade_action = side  # BUY/SELL 直接使用
            
            # 轉換來源格式：order_source ("手動"/"WEBHOOK") -> source ("manual"/"webhook")
            source = 'manual' if order_source == '手動' else 'webhook'
            
            if action == '開倉':
                # 開倉記錄
                trade_id = record_btc_opening_trade(
                    action=trade_action,
                    quantity=float(quantity),
                    price=avg_price,
                    order_id=order_id,
                    source=source
                )
                logger.info(f"✅ BTC{order_source}開倉記錄已保存: {trade_id}")
                
            elif action == '平倉':
                # 平倉記錄並自動配對
                cover_record = record_btc_covering_trade(
                    action=trade_action,
                    quantity=float(quantity),
                    price=avg_price,
                    order_id=order_id,
                    source=source
                )
                if cover_record:
                    logger.info(f"✅ BTC{order_source}平倉記錄已保存並配對完成: {cover_record['trade_id']}")
                    logger.info(f"   配對{len(cover_record.get('matched_opens', []))}筆開倉，總損益: ${cover_record.get('total_pnl', 0)}")
            
            # 保存到BTCtransdata目錄
            trade_data = {
                'trade_id': f"BTC_{action}_{trade_action}_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
                'timestamp': datetime.now().isoformat(),
                'symbol': symbol,
                'action': trade_action,
                'oc_type': 'Open' if action == '開倉' else 'Cover',
                'quantity': float(quantity),
                'price': avg_price,
                'order_id': str(order_id),
                'source': order_source
            }
            save_btc_transdata(trade_data)
            
        except Exception as e:
            logger.error(f"BTC手動訂單JSON交易記錄失敗: {e}")
        
        # 🎯 構建成交通知記錄並加入排隊系統 - 確保在提交成功通知(1秒)之後
        fill_trade_record = {
            'symbol': symbol,
            'side': side,
            'quantity': float(quantity),
            'price': float(avg_price),
            'order_id': order_id,
            'order_type': order_type_text,
            'source': 'manual' if order_source == '手動' else 'webhook',
            'action_type': action,
            'reduceOnly': reduce_only,
            'is_manual': order_source == '手動'
        }
        
        logger.info(f"準備將{order_source}BTC成交通知加入排隊: 訂單{order_id}")
        btc_notification_manager.add_notification('fill', fill_trade_record, True, 5)
        logger.info(f"{order_source}BTC成交通知已加入排隊系統（延遲5秒）")
        
        # 延遲記錄前端日誌（與成交通知同步）
        def delayed_frontend_log():
            try:
                fill_log = f"{action}成功：{direction}｜{float(quantity):.8f} BTC｜{avg_price:,.2f} USDT｜{order_type_text}"
                log_btc_frontend_message(fill_log, "success")
                logger.info(f"{order_source}幣安訂單前端日誌已記錄: {order_id}")
            except Exception as e:
                logger.error(f"延遲記錄{order_source}前端日誌失敗: {e}")
        
        # 使用線程延遲5秒記錄前端日誌
        timer = threading.Timer(5.0, delayed_frontend_log)
        timer.start()
        
        logger.info(f"{order_source}幣安訂單成交處理已安排5秒延遲執行，訂單ID: {order_id}")
        
    except Exception as e:
        logger.error(f"處理幣安訂單成交失敗: {e}")

def cleanup_processed_orders():
    """清理過期的已處理訂單記錄和下單請求記錄"""
    global processed_orders, btc_active_order_requests
    
    # 清理過期的訂單記錄
    with processed_orders_lock:
        if len(processed_orders) > 1000:
            processed_list = list(processed_orders)
            # 保留最新的500個
            processed_orders = set(processed_list[-500:])
            logger.info(f"已清理過期訂單記錄，當前記錄數: {len(processed_orders)}")
    
    # 清理過期的下單請求記錄（超過60秒的）
    with btc_order_lock:
        current_time = int(datetime.now().timestamp() * 1000)
        old_request_count = len(btc_active_order_requests)
        btc_active_order_requests = {req for req in btc_active_order_requests 
                                   if current_time - int(req.split('_')[-1]) < 60000}
        new_request_count = len(btc_active_order_requests)
        
        if old_request_count > new_request_count:
            logger.info(f"已清理 {old_request_count - new_request_count} 個過期下單請求記錄，當前記錄數: {new_request_count}")

def start_order_fallback_monitor():
    """啟動訂單備用監控（每10秒檢查一次）"""
    def monitor_worker():
        check_count = 0
        while not btc_shutdown_flag.is_set():
            check_pending_orders_fallback()
            
            # 每60次檢查（10分鐘）清理一次過期記錄
            check_count += 1
            if check_count >= 60:
                cleanup_processed_orders()
                check_count = 0
            
            # 可中斷的10秒睡眠
            for _ in range(10):
                if btc_shutdown_flag.is_set():
                    return
                time.sleep(5)
    
    fallback_thread = threading.Thread(
        target=monitor_worker,
        name="BTCOrderFallbackMonitor",
        daemon=True
    )
    fallback_thread.start()
    btc_active_threads.append(fallback_thread)
    logger.info("BTC訂單備用監控已啟動")

# ========================== BTC定時調度系統 ==========================

def start_btc_daily_scheduler():
    """啟動BTC每日調度器 - 自動生成XLSX報表"""
    global btc_daily_scheduler, btc_scheduler_thread
    
    try:
        # 檢查是否已經啟動
        if btc_scheduler_thread and btc_scheduler_thread.is_alive():
            logger.info("BTC定時調度器已在運行中")
            return
        
        # 檢查登入狀態
        env_data = load_btc_env_data()
        if env_data.get('LOGIN_BTC', '0') != '1':
            logger.info("BTC未登入，跳過啟動定時調度器")
            return
        
        def scheduler_worker():
            """定時調度器工作線程"""
            logger.info("🕐 BTC定時調度器已啟動")
            
            while not btc_shutdown_flag.is_set():
                try:
                    # 檢查登入狀態
                    env_data = load_btc_env_data()
                    if env_data.get('LOGIN_BTC', '0') != '1':
                        logger.info("BTC未登入狀態，暫停調度器")
                        # 等待60秒後再檢查
                        for _ in range(60):
                            if btc_shutdown_flag.is_set():
                                return
                            time.sleep(1)
                        continue
                    
                    # 獲取當前時間
                    now = datetime.now()
                    current_time = now.strftime("%H:%M")
                    
                    # 🚫 移除重複的報表生成觸發，統一由main.py的schedule管理
                    # 原本的 23:58 觸發已移除，避免與main.py的schedule衝突
                    if False:  # current_time == "23:58": - 已停用
                        logger.info("📊 觸發BTC每日報表生成...")
                        try:
                            result = check_btc_daily_trading_statistics()
                            logger.info(f"BTC每日報表生成完成: {result}")
                        except Exception as e:
                            logger.error(f"BTC每日報表生成失敗: {e}")
                        
                        # 等待61秒避免重複觸發
                        for _ in range(61):
                            if btc_shutdown_flag.is_set():
                                return
                            time.sleep(1)
                    else:
                        # 每60秒檢查一次時間
                        for _ in range(60):
                            if btc_shutdown_flag.is_set():
                                return
                            time.sleep(1)
                    
                except Exception as e:
                    logger.error(f"BTC定時調度器異常: {e}")
                    # 發生異常時等待60秒再繼續
                    for _ in range(60):
                        if btc_shutdown_flag.is_set():
                            return
                        time.sleep(1)
            
            logger.info("🔴 BTC定時調度器已停止")
        
        # 創建並啟動調度器線程
        btc_scheduler_thread = threading.Thread(
            target=scheduler_worker,
            name="BTCDailyScheduler",
            daemon=True
        )
        btc_scheduler_thread.start()
        btc_active_threads.append(btc_scheduler_thread)
        
        logger.info("✅ BTC定時調度器已啟動（每日23:58生成XLSX報表）")
        
    except Exception as e:
        logger.error(f"啟動BTC定時調度器失敗: {e}")

def stop_btc_daily_scheduler():
    """停止BTC定時調度器"""
    global btc_daily_scheduler, btc_scheduler_thread
    
    try:
        if btc_scheduler_thread and btc_scheduler_thread.is_alive():
            logger.info("正在停止BTC定時調度器...")
            # 線程會通過 btc_shutdown_flag 自動停止
        
        btc_daily_scheduler = None
        btc_scheduler_thread = None
        logger.info("BTC定時調度器已停止")
        
    except Exception as e:
        logger.error(f"停止BTC定時調度器失敗: {e}")

def get_btc_trading_config():
    """獲取BTC交易配置信息 - 包括實際槓桿設置"""
    global binance_client
    
    try:
        if not binance_client:
            return jsonify({'success': False, 'message': '請先登入BTC帳戶'})
        
        result = {
            'success': True,
            'config_file': {},
            'actual_settings': {},
            'current_position': {}
        }
        
        # 1. 配置文件設置
        env_data = load_btc_env_data()
        result['config_file'] = {
            'leverage': env_data.get('LEVERAGE', '未設定'),
            'position_size_percent': env_data.get('POSITION_SIZE', '未設定'),
            'margin_type': env_data.get('MARGIN_TYPE', '未設定'),
            'trading_pair': env_data.get('TRADING_PAIR', '未設定')
        }
        
        # 2. 實際API設置
        try:
            position_risk = binance_client.get_position_risk(symbol='BTCUSDT')
            if position_risk and len(position_risk) > 0:
                pos_data = position_risk[0]
                result['actual_settings'] = {
                    'leverage': pos_data.get('leverage', '未知'),
                    'margin_type': pos_data.get('marginType', '未知'),
                    'symbol': pos_data.get('symbol', '未知')
                }
            else:
                result['actual_settings'] = {'error': '無法獲取實際設置'}
        except Exception as e:
            result['actual_settings'] = {'error': f'API調用失敗: {str(e)}'}
        
        # 3. 當前持倉
        try:
            positions = binance_client.get_position_info()
            btc_position = None
            for pos in positions:
                if pos.get('symbol') == 'BTCUSDT':
                    position_amt = float(pos.get('positionAmt', 0))
                    if position_amt != 0:
                        btc_position = pos
                        break
            
            if btc_position:
                position_amt = float(btc_position.get('positionAmt', 0))
                entry_price = float(btc_position.get('entryPrice', 0))
                position_value = abs(position_amt) * entry_price
                leverage = float(btc_position.get('leverage', 1))
                margin_used = position_value / leverage if leverage > 0 else 0
                
                result['current_position'] = {
                    'has_position': True,
                    'position_amt': position_amt,
                    'entry_price': entry_price,
                    'position_value_usdt': position_value,
                    'leverage': leverage,
                    'margin_used': margin_used,
                    'unrealized_pnl': float(btc_position.get('unrealizedProfit', 0))
                }
            else:
                result['current_position'] = {'has_position': False}
                
        except Exception as e:
            result['current_position'] = {'error': f'獲取持倉失敗: {str(e)}'}
        
        # 4. 計算配置vs實際的差異
        config_leverage = result['config_file'].get('leverage', '0')
        actual_leverage = result['actual_settings'].get('leverage', '0')
        
        if config_leverage != '未設定' and actual_leverage != '未知':
            result['leverage_mismatch'] = str(config_leverage) != str(actual_leverage)
        else:
            result['leverage_mismatch'] = None
            
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"獲取BTC交易配置失敗: {e}")
        return jsonify({'success': False, 'message': f'獲取配置失敗: {str(e)}'})

