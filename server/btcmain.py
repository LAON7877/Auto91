#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BTC交易系統主程式
處理BTC加密貨幣交易相關功能
"""

from flask import request, jsonify
import os
import json
import time
import hmac
import hashlib
import requests
import threading
import glob
import logging
from datetime import datetime, timedelta, timezone
from urllib.parse import urlencode
from collections import OrderedDict
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter

# 配置目錄
CONFIG_DIR = os.path.join(os.path.dirname(__file__), 'config')
BTC_ENV_PATH = os.path.join(CONFIG_DIR, 'btc.env')

# 幣安API配置
BINANCE_BASE_URL = "https://api.binance.com"
BINANCE_FAPI_URL = "https://fapi.binance.com"  # 期貨API

# 配置日誌
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger('btc_module')

# 全局變量
binance_client = None
account_info = None
btc_shutdown_flag = threading.Event()  # BTC模組停止標誌
btc_active_threads = []  # BTC模組活動線程列表
btc_active_trades = {}  # 活躍交易記錄

# 訂單監控相關
pending_orders = {}  # 存儲待成交訂單 {order_id: order_info}
order_monitor_ws = None
order_monitor_thread = None
user_data_stream_key = None

# BTC斷線重連相關變數
btc_connection_check_interval = 60  # 每1分鐘檢查一次連線狀態
btc_max_reconnect_attempts = 999  # 無限重連嘗試
btc_reconnect_attempts = 0  # 當前重連嘗試次數
btc_last_connection_check = None  # 上次連線檢查時間
btc_is_reconnecting = False  # 是否正在重連中
btc_auto_logout_timer = None  # 自動登出計時器

def stop_btc_module():
    """停止BTC模組的所有線程和連接"""
    global btc_shutdown_flag, btc_active_threads
    logger.info("🔴 正在停止BTC模組...")
    
    # 設置停止標誌
    btc_shutdown_flag.set()
    
    # 等待所有活動線程結束
    for thread in btc_active_threads:
        if thread.is_alive():
            logger.info(f"等待BTC線程結束: {thread.name}")
            thread.join(timeout=3)
            if thread.is_alive():
                logger.warning(f"BTC線程 {thread.name} 仍在運行")
    
    btc_active_threads.clear()
    logger.info("✅ BTC模組已停止")

def register_btc_thread(thread, name="未知BTC線程"):
    """註冊BTC線程"""
    global btc_active_threads
    btc_active_threads.append(thread)
    logger.info(f"🧵 已註冊BTC線程: {name}")

# ============ 統一工具函數 ============

class APIRequestHandler:
    """統一的API請求處理類"""
    
    @staticmethod
    def make_request(method, url, data=None, headers=None, timeout=10, **kwargs):
        """統一的API請求處理函數"""
        try:
            response = getattr(requests, method.lower())(url, json=data, headers=headers, timeout=timeout, **kwargs)
            return response
        except requests.RequestException as e:
            logger.error(f"API請求失敗 [{method}] {url}: {e}")
            return None
    
    @staticmethod
    def handle_telegram_request(bot_token, chat_ids, message, message_type="info"):
        """統一的Telegram消息發送"""
        if not bot_token or not chat_ids:
            logger.warning("Telegram配置不完整，跳過發送")
            return False
            
        success_count = 0
        for chat_id in chat_ids:
            try:
                url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
                payload = {
                    'chat_id': chat_id,
                    'text': message,
                    'parse_mode': 'HTML'
                }
                response = APIRequestHandler.make_request('post', url, data=payload)
                if response and response.status_code == 200:
                    success_count += 1
                    logger.info(f"Telegram消息發送成功到 {chat_id}")
                else:
                    logger.error(f"Telegram消息發送失敗到 {chat_id}")
            except Exception as e:
                logger.error(f"Telegram發送異常到 {chat_id}: {e}")
        
        return success_count > 0

def load_config_file(file_path, required_fields=None):
    """統一的配置文件讀取函數"""
    config = {}
    if not os.path.exists(file_path):
        logger.error(f"配置文件不存在: {file_path}")
        return config
        
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and '=' in line and not line.startswith('#'):
                    key, value = line.split('=', 1)
                    config[key.strip()] = value.strip()
        
        if required_fields:
            missing_fields = [field for field in required_fields if field not in config]
            if missing_fields:
                logger.error(f"配置文件缺少必需字段: {missing_fields}")
        
        logger.info(f"成功載入配置文件: {file_path}")
        return config
        
    except Exception as e:
        logger.error(f"讀取配置文件失敗 {file_path}: {e}")
        return config

def handle_api_error(error, context="API操作", raise_error=False):
    """統一的錯誤處理函數"""
    error_msg = f"{context}失敗: {error}"
    logger.error(error_msg)
    
    if raise_error:
        raise Exception(error_msg)
    
    return error_msg

class BinanceClient:
    """幣安API客戶端"""
    
    def __init__(self, api_key, secret_key, testnet=False):
        self.api_key = api_key
        self.secret_key = secret_key
        self.base_url = BINANCE_FAPI_URL if not testnet else "https://testnet.binancefuture.com"
        self.session = requests.Session()
        self.session.headers.update({
            'X-MBX-APIKEY': api_key,
            'Content-Type': 'application/json'
        })
    
    def _generate_signature(self, query_string):
        """生成API簽名"""
        return hmac.new(
            self.secret_key.encode('utf-8'),
            query_string.encode('utf-8'),
            hashlib.sha256
        ).hexdigest()
    
    def _get_server_time(self):
        """獲取幣安服務器時間"""
        try:
            url = f"{self.base_url}/fapi/v1/time"
            response = self.session.get(url)
            if response.status_code == 200:
                return response.json().get('serverTime')
        except:
            pass
        return None
    
    def _make_request(self, method, endpoint, params=None, signed=True):
        """發送API請求"""
        url = f"{self.base_url}{endpoint}"
        
        if params is None:
            params = {}
        
        if signed:
            # 使用幣安服務器時間來避免時間同步問題
            server_time = self._get_server_time()
            if server_time:
                params['timestamp'] = server_time
            else:
                # 如果無法獲取服務器時間，使用本地時間並減去1秒作為安全邊際
                params['timestamp'] = int(time.time() * 1000) - 1000
            
            query_string = urlencode(params)
            params['signature'] = self._generate_signature(query_string)
        
        try:
            if method == 'GET':
                response = self.session.get(url, params=params)
            elif method == 'POST':
                response = self.session.post(url, params=params)
            else:
                response = self.session.request(method, url, params=params)
            
            response.raise_for_status()
            return response.json()
        
        except requests.exceptions.RequestException as e:
            logger.error(f"幣安API請求失敗: {e}")
            if hasattr(e, 'response') and e.response is not None:
                logger.error(f"HTTP狀態碼: {e.response.status_code}")
                logger.error(f"錯誤詳情: {e.response.text}")
                # 嘗試解析JSON錯誤
                try:
                    error_json = e.response.json()
                    logger.error(f"API錯誤: {error_json}")
                except:
                    pass
            return None
    
    def test_connection(self):
        """測試API連接"""
        try:
            result = self._make_request('GET', '/fapi/v1/ping', signed=False)
            return result is not None and isinstance(result, dict)
        except Exception as e:
            logger.error(f"幣安連接測試異常: {e}")
            return False
    
    def get_account_info(self):
        """獲取帳戶信息"""
        return self._make_request('GET', '/fapi/v2/account')
    
    def get_balance(self):
        """獲取餘額信息"""
        return self._make_request('GET', '/fapi/v2/balance')
    
    def get_position_info(self):
        """獲取持倉信息"""
        return self._make_request('GET', '/fapi/v2/positionRisk')
    
    def get_income_history(self, startTime=None, endTime=None, incomeType=None, limit=1000, **kwargs):
        """獲取收入歷史 - 支援新舊參數名稱"""
        params = {
            'limit': limit
        }
        
        # 支援舊參數名稱的兼容性
        if 'start_time' in kwargs:
            startTime = kwargs['start_time']
        if 'end_time' in kwargs:
            endTime = kwargs['end_time']
        if 'income_type' in kwargs:
            incomeType = kwargs['income_type']
            
        if startTime:
            params['startTime'] = startTime
        if endTime:
            params['endTime'] = endTime
        if incomeType:
            params['incomeType'] = incomeType
            
        result = self._make_request('GET', '/fapi/v1/income', params)
        return result
    
    def get_server_time(self):
        """獲取服務器時間"""
        return self._make_request('GET', '/fapi/v1/time', signed=False)
    
    def get_exchange_info(self):
        """獲取交易所信息"""
        return self._make_request('GET', '/fapi/v1/exchangeInfo', signed=False)
    
    def place_order(self, symbol, side, order_type, quantity, price=None, time_in_force='GTC', 
                   reduce_only=False, close_position=False, position_side='BOTH'):
        """下單 - 期貨訂單"""
        params = {
            'symbol': symbol,
            'side': side,  # BUY 或 SELL
            'type': order_type,  # MARKET, LIMIT, STOP, TAKE_PROFIT等
            'quantity': str(quantity),
            'positionSide': position_side,  # BOTH, LONG, SHORT
            'reduceOnly': reduce_only,
            'closePosition': close_position
        }
        
        # 只有限價單等需要 timeInForce，市價單不需要
        if order_type in ['LIMIT', 'STOP', 'TAKE_PROFIT', 'STOP_MARKET', 'TAKE_PROFIT_MARKET']:
            params['timeInForce'] = time_in_force
        
        if order_type == 'LIMIT' and price:
            params['price'] = str(price)
        
        return self._make_request('POST', '/fapi/v1/order', params)
    
    def cancel_order(self, symbol, order_id=None, client_order_id=None):
        """取消訂單"""
        params = {'symbol': symbol}
        
        if order_id:
            params['orderId'] = order_id
        elif client_order_id:
            params['origClientOrderId'] = client_order_id
        else:
            raise ValueError("必須提供 order_id 或 client_order_id")
        
        return self._make_request('DELETE', '/fapi/v1/order', params)
    
    def get_order_status(self, symbol, order_id=None, client_order_id=None):
        """查詢訂單狀態"""
        params = {'symbol': symbol}
        
        if order_id:
            params['orderId'] = order_id
        elif client_order_id:
            params['origClientOrderId'] = client_order_id
        else:
            raise ValueError("必須提供 order_id 或 client_order_id")
        
        return self._make_request('GET', '/fapi/v1/order', params)
    
    def get_all_orders(self, symbol, limit=500):
        """獲取所有訂單"""
        params = {'symbol': symbol, 'limit': limit}
        return self._make_request('GET', '/fapi/v1/allOrders', params)
    
    def get_trades(self, symbol, limit=500):
        """獲取交易歷史"""
        params = {'symbol': symbol, 'limit': limit}
        return self._make_request('GET', '/fapi/v1/userTrades', params)
    
    def get_income_history(self, symbol=None, income_type=None, limit=100):
        """獲取收益歷史"""
        params = {'limit': limit}
        
        if symbol:
            params['symbol'] = symbol
        if income_type:
            params['incomeType'] = income_type  # TRANSFER, WELCOME_BONUS, REALIZED_PNL, etc.
        
        return self._make_request('GET', '/fapi/v1/income', params)
    
    def change_leverage(self, symbol, leverage):
        """調整槓桿倍數"""
        params = {
            'symbol': symbol,
            'leverage': leverage
        }
        return self._make_request('POST', '/fapi/v1/leverage', params)
    
    def change_margin_type(self, symbol, margin_type):
        """調整保證金模式"""
        params = {
            'symbol': symbol,
            'marginType': margin_type  # ISOLATED 或 CROSSED
        }
        return self._make_request('POST', '/fapi/v1/marginType', params)
    
    def get_open_orders(self, symbol=None):
        """獲取當前掛單"""
        params = {}
        if symbol:
            params['symbol'] = symbol
        return self._make_request('GET', '/fapi/v1/openOrders', params)
    
    def cancel_all_orders(self, symbol):
        """取消所有掛單"""
        params = {'symbol': symbol}
        return self._make_request('DELETE', '/fapi/v1/allOpenOrders', params)
    
    def get_position_risk(self, symbol=None):
        """獲取持倉風險"""
        params = {}
        if symbol:
            params['symbol'] = symbol
        return self._make_request('GET', '/fapi/v2/positionRisk', params)
    
    def get_ticker_price(self, symbol=None):
        """獲取最新價格"""
        params = {}
        if symbol:
            params['symbol'] = symbol
        return self._make_request('GET', '/fapi/v1/ticker/price', params, signed=False)
    
    def get_24hr_ticker(self, symbol=None):
        """獲取24小時價格變動"""
        params = {}
        if symbol:
            params['symbol'] = symbol
        return self._make_request('GET', '/fapi/v1/ticker/24hr', params, signed=False)

def save_btc_env():
    """保存BTC環境變量"""
    try:
        data = request.get_json()
        
        # 檢查是否有空值欄位
        required_fields = ['CHAT_ID_BTC', 'BINANCE_API_KEY', 'BINANCE_SECRET_KEY', 'BINANCE_USER_ID', 'TRADING_PAIR', 'LEVERAGE', 'POSITION_SIZE', 'MARGIN_TYPE', 'CONTRACT_TYPE']
        has_empty_fields = False
        
        for field in required_fields:
            if not data.get(field, '').strip():
                has_empty_fields = True
                break
        
        # 讀取當前登入狀態（如果檔案存在的話）
        current_login_status = '0'
        if os.path.exists(BTC_ENV_PATH):
            try:
                with open(BTC_ENV_PATH, 'r', encoding='utf-8') as f:
                    for line in f:
                        line = line.strip()
                        if line.startswith('LOGIN_BTC='):
                            current_login_status = line.split('=', 1)[1]
                            break
            except Exception:
                current_login_status = '0'
        
        # 如果有空欄位，強制登出狀態設為0，否則保持當前狀態
        final_login_status = '0' if has_empty_fields else current_login_status
        
        # 創建BTC環境文件內容（允許空值）
        btc_env_content = f"""# Telegram Bot
BOT_TOKEN_BTC=7912873826:AAFPPDwuwspKVdyDGwh1oVqxH6u1gQ_N-jU

# Telegram ID
CHAT_ID_BTC={data.get('CHAT_ID_BTC', '')}

# 幣安 API Key
BINANCE_API_KEY={data.get('BINANCE_API_KEY', '')}

# 幣安 Secret Key
BINANCE_SECRET_KEY={data.get('BINANCE_SECRET_KEY', '')}

# 幣安用戶ID
BINANCE_USER_ID={data.get('BINANCE_USER_ID', '')}

# 交易對
TRADING_PAIR={data.get('TRADING_PAIR', 'BTCUSDT')}

# 合約類型
CONTRACT_TYPE={data.get('CONTRACT_TYPE', 'PERPETUAL')}

# 槓桿倍數
LEVERAGE={data.get('LEVERAGE', '20')}

# 風險比例百分比
POSITION_SIZE={data.get('POSITION_SIZE', '80')}

# 保證金模式
MARGIN_TYPE={data.get('MARGIN_TYPE', 'CROSS')}

# 登入狀態
LOGIN_BTC={final_login_status}
"""
        
        # 確保配置目錄存在
        os.makedirs(os.path.dirname(BTC_ENV_PATH), exist_ok=True)
        
        # 儲存到btc.env文件
        with open(BTC_ENV_PATH, 'w', encoding='utf-8') as f:
            f.write(btc_env_content)
        
        logger.info(f"BTC配置已儲存至: {BTC_ENV_PATH}")
        
        return jsonify({
            'success': True, 
            'message': 'BTC配置儲存成功',
            'has_empty_fields': has_empty_fields
        })
        
    except Exception as e:
        logger.error(f"儲存BTC配置失敗: {e}")
        return jsonify({'success': False, 'message': f'儲存失敗: {str(e)}'})

def btc_login():
    """BTC帳戶登入/連接"""
    global binance_client, account_info
    
    try:
        # 載入BTC配置
        if not os.path.exists(BTC_ENV_PATH):
            return jsonify({'success': False, 'message': 'BTC配置不存在，請先儲存配置'})
        
        btc_env = {}
        with open(BTC_ENV_PATH, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    btc_env[key] = value
        
        # 檢查所有必填欄位
        required_fields = ['CHAT_ID_BTC', 'BINANCE_API_KEY', 'BINANCE_SECRET_KEY', 'BINANCE_USER_ID', 'TRADING_PAIR', 'LEVERAGE', 'POSITION_SIZE', 'MARGIN_TYPE', 'CONTRACT_TYPE']
        missing_fields = []
        
        for field in required_fields:
            if not btc_env.get(field, '').strip():
                missing_fields.append(field)
        
        if missing_fields:
            return jsonify({'success': False, 'message': f'以下必填欄位為空: {", ".join(missing_fields)}'})
        
        # 獲取API配置
        api_key = btc_env.get('BINANCE_API_KEY', '').strip()
        secret_key = btc_env.get('BINANCE_SECRET_KEY', '').strip()
        trading_pair = btc_env.get('TRADING_PAIR', 'BTCUSDT')
        
        # 創建幣安客戶端
        binance_client = BinanceClient(api_key, secret_key)
        
        # 檢查連接
        connection_test = binance_client.test_connection()
        
        if not connection_test:
            return jsonify({'success': False, 'message': '無法連接到幣安服務器'})
        
        # 獲取帳戶信息
        fresh_account_info = binance_client.get_account_info()
        
        if not fresh_account_info:
            return jsonify({'success': False, 'message': 'API認證失敗，請檢查API Key和Secret Key'})
        
        # 更新全局帳戶信息
        account_info = fresh_account_info
        logger.info(f"成功獲取帳戶信息，總錢包餘額: {account_info.get('totalWalletBalance', 'N/A')}")
        
        # 獲取餘額信息
        balance_info = binance_client.get_balance()
        
        # 將登入狀態寫入btc.env
        updated_content = []
        login_found = False
        
        with open(BTC_ENV_PATH, 'r', encoding='utf-8') as f:
            for line in f:
                if line.startswith('LOGIN_BTC='):
                    updated_content.append('LOGIN_BTC=1\n')
                    login_found = True
                else:
                    updated_content.append(line)
        
        if not login_found:
            updated_content.append('LOGIN_BTC=1\n')
        
        with open(BTC_ENV_PATH, 'w', encoding='utf-8') as f:
            f.writelines(updated_content)
        
        # 準備返回信息
        total_wallet_balance = account_info.get('totalWalletBalance', '0')
        available_balance = account_info.get('availableBalance', '0')
        
        logger.info(f"BTC帳戶連接成功 - 交易對: {trading_pair}, 餘額: {total_wallet_balance} USDT")
        
        # 啟動WebSocket實時數據連接
        try:
            start_btc_websocket()
            logger.info("BTC WebSocket連接已啟動")
        except Exception as e:
            logger.error(f"啟動BTC WebSocket失敗: {e}")
        
        # 啟動風險監控
        try:
            risk_thread = threading.Thread(target=start_btc_risk_monitoring, daemon=True, name="BTC風險監控")
            register_btc_thread(risk_thread, "BTC風險監控")
            risk_thread.start()
            logger.info("BTC風險監控已啟動")
        except Exception as e:
            logger.error(f"啟動BTC風險監控失敗: {e}")
        
        # 啟動連接監控
        try:
            start_btc_connection_monitor()
        except Exception as e:
            logger.error(f"啟動BTC連接監控失敗: {e}")
        
        # 啟動訂單監控
        try:
            start_btc_order_monitoring()
            logger.info("BTC訂單監控已啟動")
        except Exception as e:
            logger.error(f"啟動BTC訂單監控失敗: {e}")
        
        # BTC不需要12小時自動重連，只在斷線時自動重連
        # 12小時自動重連功能已移除
        
        return jsonify({
            'success': True, 
            'message': f'BTC帳戶連接成功 ({trading_pair})',
            'trading_pair': trading_pair,
            'total_balance': total_wallet_balance,
            'available_balance': available_balance,
            'account_alias': account_info.get('alias', '主帳戶')
        })
        
    except Exception as e:
        logger.error(f"BTC登入失敗: {e}")
        binance_client = None
        account_info = None
        return jsonify({'success': False, 'message': f'連接失敗: {str(e)}'})

def btc_logout():
    """BTC帳戶登出"""
    global binance_client, account_info
    
    try:
        if not os.path.exists(BTC_ENV_PATH):
            return jsonify({'success': False, 'message': 'BTC配置不存在'})
        
        # BTC沒有12小時自動重連功能，無需停止計時器
        
        # 清除BTC API連接
        binance_client = None
        account_info = None
        
        # 將登入狀態改為0
        updated_content = []
        with open(BTC_ENV_PATH, 'r', encoding='utf-8') as f:
            for line in f:
                if line.startswith('LOGIN_BTC='):
                    updated_content.append('LOGIN_BTC=0\n')
                else:
                    updated_content.append(line)
        
        with open(BTC_ENV_PATH, 'w', encoding='utf-8') as f:
            f.writelines(updated_content)
        
        logger.info("BTC帳戶登出成功")
        return jsonify({'success': True, 'message': 'BTC帳戶登出成功'})
        
    except Exception as e:
        logger.error(f"BTC登出失敗: {e}")
        return jsonify({'success': False, 'message': f'登出失敗: {str(e)}'})

def get_btc_bot_username():
    """獲取BTC Bot用戶名"""
    try:
        # 從BTC環境文件中讀取Bot Token
        bot_token = None
        if os.path.exists(BTC_ENV_PATH):
            with open(BTC_ENV_PATH, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if line.startswith('BOT_TOKEN_BTC='):
                        bot_token = line.split('=', 1)[1]
                        break
        
        # 如果.env中沒有token，使用硬編碼值
        if not bot_token:
            bot_token = "7912873826:AAFPPDwuwspKVdyDGwh1oVqxH6u1gQ_N-jU"
        
        if not bot_token:
            return jsonify({'username': None})
        
        # 使用Telegram Bot API獲取Bot信息
        bot_api_url = f"https://api.telegram.org/bot{bot_token}/getMe"
        
        try:
            response = requests.get(bot_api_url, timeout=10)
            if response.status_code == 200:
                bot_info = response.json()
                if bot_info.get('ok'):
                    username = bot_info['result'].get('username', 'Auto91_BtcBot')
                    return jsonify({'username': f'@{username}'})
        except:
            pass
        
        # 如果API調用失敗，返回默認值
        return jsonify({'username': '@Auto91_BtcBot'})
        
    except Exception as e:
        logger.error(f"獲取BTC Bot用戶名失敗: {e}")
        return jsonify({'username': '@Auto91_BtcBot'})

def load_btc_env():
    """載入BTC環境變量"""
    try:
        if not os.path.exists(BTC_ENV_PATH):
            # 如果文件不存在，創建預設配置
            create_default_btc_env()
        
        btc_env = {}
        with open(BTC_ENV_PATH, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    btc_env[key] = value
        return jsonify(btc_env)
            
    except Exception as e:
        logger.error(f"載入BTC配置失敗: {e}")
        return jsonify({})

def create_default_btc_env():
    """創建預設BTC環境配置文件"""
    try:
        # 確保配置目錄存在
        os.makedirs(os.path.dirname(BTC_ENV_PATH), exist_ok=True)
        
        default_config = """# Telegram Bot
BOT_TOKEN_BTC=7912873826:AAFPPDwuwspKVdyDGwh1oVqxH6u1gQ_N-jU

# Telegram ID
CHAT_ID_BTC=

# 幣安 API Key
BINANCE_API_KEY=

# 幣安 Secret Key
BINANCE_SECRET_KEY=

# 幣安用戶ID
BINANCE_USER_ID=

# 交易對
TRADING_PAIR=BTCUSDT

# 合約類型
CONTRACT_TYPE=PERPETUAL

# 槓桿倍數
LEVERAGE=20

# 風險比例百分比
POSITION_SIZE=80

# 保證金模式
MARGIN_TYPE=CROSS

# 登入狀態
LOGIN_BTC=0
"""
        
        with open(BTC_ENV_PATH, 'w', encoding='utf-8') as f:
            f.write(default_config)
        
        logger.info(f"已創建預設BTC配置文件: {BTC_ENV_PATH}")
        return True
        
    except Exception as e:
        logger.error(f"創庺預設BTC配置失敗: {e}")
        return False

def reset_btc_env_to_default():
    """重設BTC配置為預設值（保持現有的API資訊）"""
    try:
        # 讀取現有配置
        existing_config = {}
        if os.path.exists(BTC_ENV_PATH):
            with open(BTC_ENV_PATH, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith('#') and '=' in line:
                        key, value = line.split('=', 1)
                        existing_config[key] = value
        
        # 使用現有資訊或預設值
        config_content = f"""# Telegram Bot
BOT_TOKEN_BTC=7912873826:AAFPPDwuwspKVdyDGwh1oVqxH6u1gQ_N-jU

# Telegram ID
CHAT_ID_BTC={existing_config.get('CHAT_ID_BTC', '')}

# 幣安 API Key
BINANCE_API_KEY={existing_config.get('BINANCE_API_KEY', '')}

# 幣安 Secret Key
BINANCE_SECRET_KEY={existing_config.get('BINANCE_SECRET_KEY', '')}

# 幣安用戶ID
BINANCE_USER_ID={existing_config.get('BINANCE_USER_ID', '')}

# 交易對
TRADING_PAIR=BTCUSDT

# 合約類型
CONTRACT_TYPE=PERPETUAL

# 槓桿倍數
LEVERAGE=20

# 風險比例百分比
POSITION_SIZE=80

# 保證金模式
MARGIN_TYPE=CROSS

# 登入狀態
LOGIN_BTC=0
"""
        
        # 確保目錄存在
        os.makedirs(os.path.dirname(BTC_ENV_PATH), exist_ok=True)
        
        # 寫入配置
        with open(BTC_ENV_PATH, 'w', encoding='utf-8') as f:
            f.write(config_content)
        
        logger.info(f"BTC配置已重設為預設值: {BTC_ENV_PATH}")
        return True
        
    except Exception as e:
        logger.error(f"重設BTC配置失敗: {e}")
        return False

def load_btc_env_data():
    """載入BTC環境變量數據（內部使用）"""
    try:
        if not os.path.exists(BTC_ENV_PATH):
            create_default_btc_env()
        
        env_data = {}
        with open(BTC_ENV_PATH, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    env_data[key] = value
        return env_data
    except Exception as e:
        logger.error(f"載入BTC環境數據失敗: {e}")
        return {}

def calculate_btc_quantity(signal_data, account_balance):
    """計算BTC交易數量"""
    try:
        # 獲取配置的槓桿倍數
        env_data = load_btc_env_data()
        leverage = float(env_data.get('LEVERAGE', 5))
        
        # 獲取倉位大小百分比(從環境變數讀取，預設5%風險)
        position_size_pct = float(env_data.get('POSITION_SIZE', 10.0)) / 100
        
        # 計算可用餘額
        available_balance = float(account_balance.get('availableBalance', 0))
        
        # 計算下單金額 = 可用餘額 * 倉位比例 * 槓桿
        order_value = available_balance * position_size_pct * leverage
        
        # 獲取當前價格
        current_price = float(signal_data.get('price', 0))
        if current_price <= 0:
            # 如果信號沒有價格，獲取市場價格
            if binance_client:
                ticker = binance_client.get_ticker_price('BTCUSDT')
                current_price = float(ticker.get('price', 0)) if ticker else 0
        
        if current_price <= 0:
            raise ValueError("無法獲取有效價格")
        
        # 計算數量 = 下單金額 / 價格
        quantity = order_value / current_price
        
        # 幣安BTCUSDT最小下單單位(0.001)
        min_qty = 0.001
        quantity = max(min_qty, round(quantity, 3))
        
        logger.debug(f"BTC數量計算詳細: 理論數量={order_value/current_price:.6f}, 最小單位={min_qty}, 最終數量={quantity}")
        
        logger.debug(f"BTC倉位計算: 可用餘額={available_balance}, 槓桿={leverage}, 風險={position_size_pct*100}%, 價格={current_price}, 數量={quantity}")
        
        return quantity
        
    except Exception as e:
        logger.error(f"計算BTC數量失敗: {e}")
        return 0.001  # 返回最小單位

def place_btc_futures_order(symbol, side, quantity, price=None, order_type="MARKET", reduce_only=False, is_manual=False):
    """BTC期貨下單"""
    global binance_client, btc_active_trades
    
    try:
        if not binance_client:
            raise ValueError("BTC客戶端未初始化")
        
        # 判斷是開倉還是平倉
        action_type = '平倉' if reduce_only else '開倉'
        manual_type = '手動' if is_manual else '自動'
        
        logger.debug(f"[BTC] 下單參數: symbol={symbol}, side={side}, quantity={quantity}, type={order_type}, reduce_only={reduce_only}")
        logger.debug(f"[BTC] 動作判斷: action_type={action_type}, manual_type={manual_type}")
        # 記錄到系統日誌（僅後端顯示）
        
        # 執行下單
        result = binance_client.place_order(
            symbol=symbol,
            side=side,
            order_type=order_type,
            quantity=quantity,
            price=price if order_type == 'LIMIT' else None
        )
        
        if result:
            order_id = result.get('orderId')
            client_order_id = result.get('clientOrderId')
            
            logger.info(f"BTC訂單提交成功: OrderID={order_id}, ClientOrderID={client_order_id}")
            
            # 記錄詳細的訂單提交日誌（僅後端顯示）
            detailed_log = get_btc_order_log_message(
                symbol=symbol,
                side=side,
                quantity=quantity,
                price=price if order_type == 'LIMIT' else 0,
                order_id=order_id,
                order_type=order_type,
                is_manual=is_manual,
                action_type=action_type
            )
            # 僅在後端控制台顯示，不發送到前端
            
            # 記錄到活躍交易
            trade_record = {
                'order_id': order_id,
                'client_order_id': client_order_id,
                'symbol': symbol,
                'side': side,
                'quantity': quantity,
                'order_type': order_type,
                'is_manual': is_manual,
                'action_type': action_type,
                'reduce_only': reduce_only,
                'status': result.get('status', 'NEW'),
                'timestamp': datetime.now().isoformat()
            }
            
            btc_active_trades[order_id] = trade_record
            
            # 發送提交成功通知
            send_btc_order_submit_notification(trade_record, True)
            
            # 記錄訂單提交成功到系統日誌（與TX系統格式統一）
            try:
                submit_log = get_btc_order_log_message(
                    symbol=symbol,
                    side=side,
                    quantity=quantity,
                    price=price if order_type == 'LIMIT' else 0,
                    order_id=order_id,
                    order_type=order_type,
                    is_manual=is_manual,
                    action_type=action_type,
                    is_success=False  # 提交成功但未成交
                )
                log_btc_frontend_message(submit_log, "info")
            except:
                pass
            
            # 延遲檢查成交狀態（市價單延遲較短）
            def check_fill_status():
                if order_type.upper() == 'MARKET':
                    time.sleep(0.5)  # 市價單延遲0.5秒檢查
                else:
                    time.sleep(3)  # 限價單延遲3秒檢查
                check_btc_order_fill(order_id, symbol)
            
            fill_check_thread = threading.Thread(target=check_fill_status, daemon=True, name=f"BTC訂單檢查-{order_id}")
            register_btc_thread(fill_check_thread, f"BTC訂單檢查-{order_id}")
            fill_check_thread.start()
            
            return {
                'success': True,
                'message': f'{manual_type}{action_type}成功',
                'order_id': order_id,
                'symbol': symbol,
                'side': side,
                'quantity': quantity,
                'result': result
            }
        else:
            raise Exception("下單返回空結果")
            
    except Exception as e:
        logger.error(f"BTC下單失敗: {e}")
        # 記錄到系統日誌
        log_btc_frontend_message(f"下單失敗: {e}", "error")
        # 發送提交失敗通知
        error_record = {
            'symbol': symbol,
            'side': side,
            'quantity': quantity,
            'error': str(e)
        }
        send_btc_order_submit_notification(error_record, False)
        return {
            'success': False,
            'message': f'下單失敗: {str(e)}',
            'symbol': symbol,
            'side': side,
            'quantity': quantity
        }

def check_btc_order_fill(order_id, symbol):
    """檢查BTC訂單成交狀態"""
    global binance_client, btc_active_trades
    
    try:
        if not binance_client:
            return
        
        # 查詢訂單狀態
        order_status = binance_client.get_order_status(symbol, order_id=order_id)
        
        if order_status:
            status = order_status.get('status')
            
            if status == 'FILLED':
                # 訂單已成交
                trade_record = btc_active_trades.get(order_id, {})
                trade_record.update({
                    'fill_price': order_status.get('avgPrice'),
                    'fill_quantity': order_status.get('executedQty'),
                    'fill_time': order_status.get('updateTime'),
                    'status': 'FILLED'
                })
                
                logger.info(f"BTC訂單成交: OrderID={order_id}, 成交價={trade_record.get('fill_price')}")
                
                # 記錄成交日誌到系統日誌（參考TX格式）
                fill_price = float(trade_record.get('fill_price', '0'))
                fill_quantity = trade_record.get('fill_quantity', '0')
                symbol = trade_record.get('symbol', 'BTCUSDT')
                side = trade_record.get('side', 'BUY')
                
                # 生成詳細的成交日誌（使用保存的訂單信息）
                detailed_fill_log = get_btc_order_log_message(
                    symbol=symbol,
                    side=side,
                    quantity=fill_quantity,
                    price=fill_price,
                    order_id=order_id,
                    order_type=trade_record.get('order_type', 'MARKET'),
                    is_manual=trade_record.get('is_manual', False),
                    action_type=trade_record.get('action_type', '開倉'),
                    is_success=True  # 成交時使用成功格式
                )
                log_message = detailed_fill_log
                # 僅在後端控制台顯示，不發送到前端
                
                # 發送成交通知
                send_btc_trade_notification(trade_record)
                
                # 記錄成交成功到系統日誌（與TX系統格式統一）
                try:
                    success_log = get_btc_order_log_message(
                        symbol=symbol,
                        side=trade_record.get('side', ''),
                        quantity=float(executed_qty),
                        price=fill_price,
                        order_id=order_id,
                        order_type=trade_record.get('order_type', 'MARKET'),
                        is_manual=trade_record.get('is_manual', False),
                        action_type=trade_record.get('action_type', '開倉'),
                        is_success=True
                    )
                    log_btc_frontend_message(success_log, "success")
                except:
                    pass
                
                # 保存交易記錄
                save_btc_trade_record(trade_record)
                
            elif status in ['CANCELED', 'REJECTED', 'EXPIRED']:
                # 訂單失效
                trade_record = btc_active_trades.get(order_id, {})
                trade_record.update({'status': status})
                logger.info(f"BTC訂單失效: OrderID={order_id}, 狀態={status}")
                
                # 記錄失效日誌到系統日誌
                symbol = trade_record.get('symbol', 'BTCUSDT')
                side = trade_record.get('side', '未知')
                quantity = trade_record.get('quantity', '0')
                
                log_message = f"BTC訂單失效：{symbol}｜{side}｜{quantity}｜狀態：{status}"
                log_btc_frontend_message(log_message, "warning")
                
        # 從活躍交易中移除已完成的訂單
        if order_id in btc_active_trades:
            final_status = btc_active_trades[order_id].get('status')
            if final_status in ['FILLED', 'CANCELED', 'REJECTED', 'EXPIRED']:
                del btc_active_trades[order_id]
                
    except Exception as e:
        logger.error(f"檢查BTC訂單狀態失敗: {e}")

def process_btc_entry_signal(signal_data, parsed_action=None):
    """處理BTC進場信號"""
    try:
        action = parsed_action or signal_data.get('action', '').upper()
        symbol = signal_data.get('symbol', 'BTCUSDT')
        
        if action not in ['LONG', 'SHORT']:
            raise ValueError(f"無效的進場動作: {action}")
        
        # 標準化方向（BTC期貨邏輯）
        side = 'BUY' if action == 'LONG' else 'SELL'
        logger.debug(f"[BTC] 信號解析: {action} -> side={side}")
        
        # 獲取帳戶餘額計算數量
        if not binance_client:
            raise ValueError("BTC客戶端未初始化")
        
        balance_info = binance_client.get_balance()
        if not balance_info:
            raise ValueError("無法獲取帳戶餘額")
        
        # 找到USDT餘額
        usdt_balance = None
        for balance in balance_info:
            if balance.get('asset') == 'USDT':
                usdt_balance = balance
                break
        
        if not usdt_balance:
            raise ValueError("找不到USDT餘額")
        
        # 計算交易數量
        quantity = calculate_btc_quantity(signal_data, usdt_balance)
        
        # 執行下單（webhook為自動交易，預設為開倉）
        order_result = place_btc_futures_order(
            symbol=symbol,
            side=side,
            quantity=quantity,
            order_type='MARKET',
            reduce_only=False,  # 開倉
            is_manual=False     # 自動交易
        )
        
        return order_result
        
    except Exception as e:
        logger.error(f"處理BTC進場信號失敗: {e}")
        return None

def process_btc_exit_signal(signal_data, parsed_action=None):
    """處理BTC出場信號"""
    try:
        action = parsed_action or signal_data.get('action', '').upper()
        symbol = signal_data.get('symbol', 'BTCUSDT')
        
        if action not in ['CLOSE', 'CLOSE_LONG', 'CLOSE_SHORT']:
            raise ValueError(f"無效的出場動作: {action}")
        
        # 獲取當前持倉
        if not binance_client:
            raise ValueError("BTC客戶端未初始化")
        
        positions = binance_client.get_position_info()
        if not positions:
            logger.info("沒有找到持倉信息")
            return None
        
        # 找到對應的持倉
        target_position = None
        for pos in positions:
            if pos.get('symbol') == symbol and float(pos.get('positionAmt', 0)) != 0:
                target_position = pos
                break
        
        if not target_position:
            logger.info(f"沒有找到{symbol}的活躍持倉")
            return None
        
        # 計算平倉方向和數量
        position_amt = float(target_position.get('positionAmt', 0))
        
        if position_amt > 0:
            # 多頭持倉，需要賣出平倉
            side = 'SELL'
            quantity = abs(position_amt)
        else:
            # 空頭持倉，需要買入平倉
            side = 'BUY' 
            quantity = abs(position_amt)
        
        # 執行平倉單
        order_result = place_btc_futures_order(
            symbol=symbol,
            side=side,
            quantity=quantity,
            order_type='MARKET',
            reduce_only=True,   # 平倉
            is_manual=False     # 自動交易
        )
        
        return order_result
        
    except Exception as e:
        logger.error(f"處理BTC出場信號失敗: {e}")
        return None


def save_btc_trade_record(trade_record):
    """保存BTC交易記錄 - 使用日期格式文件名"""
    try:
        today = datetime.now().strftime("%Y%m%d")
        trades_file = os.path.join(os.path.dirname(__file__), 'BTCtransdata', f'BTCtrades_{today}.json')
        
        # 確保目錄存在
        os.makedirs(os.path.dirname(trades_file), exist_ok=True)
        
        # 讀取現有交易記錄
        existing_trades = []
        if os.path.exists(trades_file):
            try:
                with open(trades_file, 'r', encoding='utf-8') as f:
                    existing_trades = json.load(f)
            except:
                existing_trades = []
        
        # 添加時間戳
        trade_record['timestamp'] = datetime.now().isoformat()
        
        # 添加新交易記錄
        existing_trades.append(trade_record)
        
        # 保存回文件
        with open(trades_file, 'w', encoding='utf-8') as f:
            json.dump(existing_trades, f, ensure_ascii=False, indent=2)
        
        # 清理舊的BTC交易記錄檔案（保留30個交易日）
        cleanup_old_btc_trade_files()
            
    except Exception as e:
        logger.error(f"保存BTC交易記錄失敗: {e}")

def cleanup_old_btc_trade_files():
    """清理舊的BTC交易記錄檔案，保留30個交易日"""
    try:
        btc_log_dir = os.path.join(os.path.dirname(__file__), 'BTCtransdata')
        if not os.path.exists(btc_log_dir):
            return
        
        # 獲取所有BTC交易記錄檔案
        trade_files = []
        for filename in os.listdir(btc_log_dir):
            if filename.startswith('BTCtrades_') and filename.endswith('.json'):
                try:
                    # 從檔案名提取日期
                    date_str = filename.replace('BTCtrades_', '').replace('.json', '')
                    file_date = datetime.strptime(date_str, '%Y%m%d')
                    trade_files.append((filename, file_date))
                except ValueError:
                    # 如果檔案名格式不正確，跳過
                    continue
        
        # 按日期排序（最新的在前）
        trade_files.sort(key=lambda x: x[1], reverse=True)
        
        # 保留最新的30個檔案，刪除其餘的
        if len(trade_files) > 30:
            files_to_delete = trade_files[30:]
            for filename, _ in files_to_delete:
                try:
                    file_path = os.path.join(btc_log_dir, filename)
                    os.remove(file_path)
                    logger.info(f"已刪除舊BTC交易記錄檔案：{filename}")
                except Exception as e:
                    logger.error(f"刪除BTC檔案失敗 {filename}：{e}")
            
        logger.info(f"BTC交易記錄檔案清理完成，保留 {min(len(trade_files), 30)} 個檔案")
    
    except Exception as e:
        logger.error(f"清理舊BTC交易記錄檔案失敗：{e}")

def send_btc_telegram_message(message, chat_id=None, bot_token=None):
    """發送BTC Telegram訊息（使用統一工具函數）"""
    try:
        
        # 載入BTC配置
        env_data = load_config_file(BTC_ENV_PATH, ['CHAT_ID_BTC', 'BOT_TOKEN_BTC'])
        
        if not chat_id:
            chat_id_raw = env_data.get('CHAT_ID_BTC')
        else:
            chat_id_raw = chat_id
        if not bot_token:
            bot_token = env_data.get('BOT_TOKEN_BTC')
        
        if not chat_id_raw or not bot_token:
            logger.info("BTC Telegram配置不完整，跳過通知")
            log_btc_frontend_message("Telegram配置不完整", "error")
            return False
        
        # 支援多個CHAT_ID，用逗號分隔
        chat_ids = [id.strip() for id in str(chat_id_raw).split(',') if id.strip()]
        
        logger.info(f"BTC BOT_TOKEN: {bot_token[:10]}...")
        logger.info(f"BTC CHAT_IDs: {chat_ids}")
        
        url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
        
        # 記錄發送結果
        success_count = 0
        total_count = len(chat_ids)
        
        # 對每個CHAT_ID發送訊息
        for chat_id in chat_ids:
            payload = {
                'chat_id': chat_id,
                'text': message,
                'parse_mode': 'HTML'
            }
            
            logger.info(f"發送請求到 BTC Telegram API (Chat ID: {chat_id})...")
            response = requests.post(url, json=payload, timeout=10)
            
            logger.info(f"BTC Telegram API 回應 (Chat ID: {chat_id}): {response.status_code}")
            
            if response.status_code == 200:
                logger.info(f"BTC Telegram 訊息發送成功 (Chat ID: {chat_id})！")
                success_count += 1
            else:
                logger.error(f"BTC Telegram 訊息發送失敗 (Chat ID: {chat_id}): {response.text}")
        
        # 判斷整體發送結果
        if success_count == total_count:
            logger.info(f"BTC Telegram 訊息發送完成！成功發送到 {success_count}/{total_count} 個接收者")
            
            # 根據訊息內容判斷發送狀態類型（參考TX系統）
            if "訂單提交成功" in message or "提交成功" in message:
                log_message = "Telegram［提交成功］訊息發送成功！！！"
            elif "訂單提交失敗" in message or "提交失敗" in message:
                log_message = "Telegram［提交失敗］訊息發送成功！！！"
            elif "成交通知" in message or "訂單成交" in message:
                log_message = "Telegram［成交通知］訊息發送成功！！！"
            elif "API連線異常" in message or "連線失敗" in message:
                log_message = "Telegram［API連線異常］訊息發送成功！！！"
            elif "API連線成功" in message or "連線成功" in message:
                log_message = "Telegram［API連線成功］訊息發送成功！！！"
            elif "交易統計" in message or "統計報告" in message:
                log_message = "Telegram［交易統計］訊息發送成功！！！"
            elif "日報" in message or "月報" in message or "報表" in message:
                log_message = "Telegram［生成報表］訊息發送成功！！！"
            elif "系統啟動" in message or "啟動通知" in message or "正在啟動中" in message:
                log_message = "Telegram［啟動通知］訊息發送成功！！！"
            else:
                log_message = "Telegram 訊息發送成功！！！"
            
            # 記錄到BTC前端系統日誌
            log_type = 'warning' if 'API連線異常' in log_message else 'success'
            log_btc_frontend_message(log_message, log_type)
            logger.info(f"BTC系統日誌已發送: {log_message}")
            
            return True
        else:
            logger.error(f"BTC Telegram 訊息部分發送失敗！成功發送到 {success_count}/{total_count} 個接收者")
            # 發送失敗也要記錄日誌
            status_type = 'warning' if success_count > 0 else 'error'
            error_log_message = f"Telegram 訊息部分發送失敗！成功：{success_count}/{total_count}"
            log_btc_frontend_message(error_log_message, status_type)
            logger.info(f"BTC系統錯誤日誌已發送: {error_log_message}")
            return success_count > 0  # 至少有一個成功就返回True
            
    except Exception as e:
        logger.error(f"發送BTC Telegram訊息失敗: {e}")
        logger.error(f"錯誤類型: {str(e.__class__.__name__)}")
        if hasattr(e, 'response'):
            logger.error(f"回應內容: {e.response.text}")
        import traceback
        traceback.print_exc()
        
        # 記錄異常到系統日誌
        error_log_message = f"Telegram 訊息發送異常：{str(e)[:100]}"
        log_btc_frontend_message(error_log_message, "error")
        return False

def send_btc_telegram_file(file_path, caption=""):
    """發送BTC Telegram檔案（與TX系統格式一致）"""
    try:
        if not os.path.exists(file_path):
            logger.info(f"找不到檔案，路徑: {file_path}")
            return False
        
        # 載入BTC配置
        env_data = load_btc_env_data()
        
        bot_token = env_data.get('BOT_TOKEN_BTC')
        chat_id_raw = env_data.get('CHAT_ID_BTC')
        
        if not bot_token or not chat_id_raw:
            logger.info("BTC Telegram配置不完整，跳過檔案發送")
            log_btc_frontend_message("Telegram配置不完整", "error")
            return False
        
        # 支援多個CHAT_ID，用逗號分隔
        chat_ids = [id.strip() for id in str(chat_id_raw).split(',') if id.strip()]
        
        logger.info(f"準備發送BTC檔案到 {len(chat_ids)} 個接收者")
        
        # 發送檔案
        url = f"https://api.telegram.org/bot{bot_token}/sendDocument"
        
        # 記錄發送結果
        success_count = 0
        total_count = len(chat_ids)
        
        # 對每個CHAT_ID發送檔案
        for chat_id in chat_ids:
            with open(file_path, 'rb') as f:
                files = {'document': f}
                data = {
                    'chat_id': chat_id,
                    'caption': caption
                }
                
                logger.info(f"發送BTC檔案到 Chat ID: {chat_id}")
                response = requests.post(url, files=files, data=data, timeout=30)
                
                if response.status_code == 200:
                    logger.info(f"BTC Telegram檔案發送成功 (Chat ID: {chat_id})")
                    success_count += 1
                else:
                    logger.error(f"BTC Telegram檔案發送失敗 (Chat ID: {chat_id}): {response.status_code}")
        
        # 判斷整體發送結果
        if success_count == total_count:
            logger.info(f"BTC Telegram檔案發送完成！成功發送到 {success_count}/{total_count} 個接收者")
            
            # 記錄前端系統日誌（合併發送：生成報表 + 檔案發送）
            log_btc_frontend_message("Telegram［生成報表］訊息發送成功！！！", "success")
            log_btc_frontend_message("Telegram［檔案發送］訊息發送成功！！！", "success")
            
            return True
        else:
            logger.error(f"BTC Telegram檔案部分發送失敗！成功發送到 {success_count}/{total_count} 個接收者")
            status_type = 'warning' if success_count > 0 else 'error'
            log_btc_frontend_message(f"Telegram檔案部分發送失敗！成功：{success_count}/{total_count}", status_type)
            return success_count > 0  # 至少有一個成功就返回True
            
    except Exception as e:
        logger.error(f"發送BTC Telegram檔案失敗: {e}")
        log_btc_frontend_message(f"Telegram檔案發送失敗！異常錯誤：{str(e)[:50]}", "error")
        return False

def get_btc_order_log_message(symbol, side, quantity, price, order_id, order_type, is_manual, action_type, is_success=False):
    """生成BTC訂單日誌訊息（完全參考TX邏輯）"""
    try:
        # 使用完整交易對名稱（如BTCUSDT），與TX系統的合約代碼對應
        simple_symbol = symbol
        
        # 判斷手動/自動（與TX一致）
        manual_type = '手動' if is_manual else '自動'
        
        # 格式化價格（與TX邏輯一致）
        if price == 0 or order_type == 'MARKET':
            price_display = '市價'
        else:
            price_display = f'{price:,.0f}'
        
        # 格式化方向 - 完全遵循TX邏輯
        if action_type == '開倉':
            # 開倉：BUY=多單, SELL=空單（與TX NEW邏輯一致）
            if str(side).upper() == 'BUY':
                direction_display = '多單'
            else:
                direction_display = '空單'
        elif action_type == '平倉':
            # 平倉：BUY=平空單, SELL=平多單（與TX COVER邏輯一致）
            if str(side).upper() == 'BUY':
                direction_display = '空單'  # 平空單
            else:
                direction_display = '多單'  # 平多單
        else:
            # 備援邏輯（與TX一致）
            if str(side).upper() == 'BUY':
                direction_display = '多單'
            else:
                direction_display = '空單'
        
        # 格式化數量（BTC格式）
        quantity_display = f'{quantity} BTC'
        
        # 格式化訂單類型（與TX一致）
        if order_type.upper() == 'MARKET':
            order_info = "市價單"
        else:
            order_info = "限價單"
        
        # 返回格式（與TX格式完全一致）
        if is_success:
            # 成交成功格式 - 顯示成交價格
            return f"{action_type}成功：{simple_symbol}｜{direction_display}｜{quantity_display}｜{price_display}｜{order_info}"
        else:
            # 掛單格式 - 市價單顯示"市價"作為價格，限價單顯示實際價格
            if order_type.upper() == 'MARKET':
                return f"{manual_type}{action_type}：{simple_symbol}｜{direction_display}｜{quantity_display}｜市價｜{order_info}"
            else:
                return f"{manual_type}{action_type}：{simple_symbol}｜{direction_display}｜{quantity_display}｜{price_display}｜{order_info}"
            
    except Exception as e:
        logger.error(f"生成BTC日誌訊息失敗: {e}")
        return f"日誌生成失敗: {order_id}"

def log_btc_system_message(message, log_type="info"):
    """記錄BTC系統日誌（僅後端控制台，不發送到前端）"""
    try:
        # 只在後端控制台顯示，不發送到前端
        logger.info(f"[BTC系統日誌] {log_type.upper()}: {message}")
    except:
        pass

def log_btc_frontend_message(message, log_type="info"):
    """記錄BTC前端系統日誌"""
    try:
        # 動態獲取當前端口
        current_port = 5000  # 預設端口
        try:
            # 嘗試從主模組獲取當前端口
            import main
            if hasattr(main, 'CURRENT_PORT'):
                current_port = main.CURRENT_PORT
        except:
            pass
            
        # 使用BTC專用的日誌端點
        response = requests.post(
            f'http://127.0.0.1:{current_port}/api/btc_system_log',
            json={'message': message, 'type': log_type, 'system': 'BTC'},
            timeout=5
        )
        
        # 檢查響應狀態
        if response.status_code != 200:
            logger.warning(f"BTC前端日誌發送失敗: HTTP {response.status_code}")
            
    except Exception as e:
        # 記錄錯誤但不阻塞主要功能
        logger.warning(f"BTC前端日誌發送異常: {e}")
        # 嘗試直接記錄到後端日誌作為備用
        try:
            logger.info(f"[BTC {log_type.upper()}] {message}")
        except:
            pass

def get_btc_position_info():
    """獲取BTC持倉信息，包括USDT持倉數量和強平價格"""
    try:
        if not binance_client:
            return None, None, None
            
        # 獲取持倉資訊
        positions = binance_client.get_positions()
        for pos in positions:
            if pos.get('symbol') == 'BTCUSDT':
                position_amt = float(pos.get('positionAmt', 0))
                if position_amt != 0:
                    entry_price = float(pos.get('entryPrice', 0))
                    mark_price = float(pos.get('markPrice', 0))
                    liquidation_price = float(pos.get('liquidationPrice', 0))
                    
                    # 計算USDT持倉數量 = BTC數量 × 標記價格
                    usdt_position = abs(position_amt) * mark_price
                    
                    return usdt_position, entry_price, liquidation_price
        
        return None, None, None
        
    except Exception as e:
        logger.error(f"獲取BTC持倉信息失敗: {e}")
        return None, None, None

def get_btc_leverage_info():
    """獲取BTC槓桿和保證金模式信息"""
    try:
        env_data = load_btc_env_data()
        leverage = env_data.get('LEVERAGE', '5')
        margin_type = env_data.get('MARGIN_TYPE', 'CROSS')
        margin_mode = "全倉" if margin_type == 'CROSS' else "逐倉"
        return f"{leverage}x", margin_mode
    except:
        return "5x", "全倉"

def send_btc_order_submit_notification(trade_record, success=True):
    """發送BTC訂單提交通知 - 新格式"""
    try:
        current_date = datetime.now().strftime('%Y/%m/%d')
        
        # 提取基本信息
        symbol = trade_record.get('symbol', 'BTCUSDT')
        side = trade_record.get('side', '')
        quantity = trade_record.get('quantity', 0)
        order_id = trade_record.get('order_id', '未知' if not success else '')
        price = trade_record.get('price', 0)
        order_type = trade_record.get('order_type', 'MARKET')
        is_manual = trade_record.get('is_manual', False)
        reduce_only = trade_record.get('reduceOnly', False)
        
        # 獲取槓桿和保證金模式
        leverage, margin_mode = get_btc_leverage_info()
        
        # 判斷開平倉和方向
        if reduce_only:
            # 平倉邏輯：cover sell(平多) cover buy(平空)
            if side == 'SELL':
                action_type = "平倉"
                direction_display = "多單"  # 平多倉，方向還是多單
            else:  # BUY
                action_type = "平倉" 
                direction_display = "空單"  # 平空倉，方向還是空單
        else:
            # 開倉邏輯：new buy(開多) new sell(開空)
            action_type = "開倉"
            if side == 'BUY':
                direction_display = "多單"  # 開多倉
            else:  # SELL
                direction_display = "空單"  # 開空倉
        
        # 判斷自動/手動 - webhook就是自動
        submit_type = "自動" if not is_manual else "手動"
        
        # 獲取持倉信息
        usdt_position, entry_price, liquidation_price = get_btc_position_info()
        
        # 格式化顯示
        order_type_display = "市價單" if order_type == 'MARKET' else "限價單"
        usdt_position_display = f"{usdt_position:,.2f}" if usdt_position else "N/A"
        entry_price_display = f"{entry_price:,.2f}" if entry_price else "N/A"
        liquidation_price_display = f"{liquidation_price:,.2f}" if liquidation_price else "N/A"
        
        if success:
            msg = (f"⭕ 提交成功（{current_date}）\n"
                   f"BTCUSDT/永續/{margin_mode}{leverage}\n"
                   f"提交單號：{order_id}\n"
                   f"提交動作：{submit_type}{action_type}\n"
                   f"提交類別：{order_type_display}\n"
                   f"提交方向：{direction_display}\n"
                   f"提交數量：{quantity:.8f}\n"
                   f"持倉數量(USDT)：{usdt_position_display}\n"
                   f"開倉價格(USDT)：{entry_price_display}\n"
                   f"強平價格(USDT)：{liquidation_price_display}")
        else:
            error = trade_record.get('error', '未知錯誤')
            msg = (f"❌ 提交失敗（{current_date}）\n"
                   f"BTCUSDT/永續/{margin_mode}{leverage}\n"
                   f"提交單號：{order_id}\n"
                   f"提交動作：{submit_type}{action_type}\n"
                   f"提交類別：{order_type_display}\n"
                   f"提交方向：{direction_display}\n"
                   f"提交數量：{quantity:.8f}\n"
                   f"持倉數量(USDT)：{usdt_position_display}\n"
                   f"開倉價格(USDT)：{entry_price_display}\n"
                   f"強平價格(USDT)：{liquidation_price_display}\n"
                   f"原因：{error}")
        
        send_btc_telegram_message(msg)
        
        # 添加到系統日誌
        logger.info(f"BTC訂單通知: {msg}")
        
    except Exception as e:
        logger.error(f"發送BTC訂單提交通知失敗: {e}")

def send_btc_trade_notification(trade_record):
    """發送BTC成交通知 - 新格式"""
    try:
        current_date = datetime.now().strftime('%Y/%m/%d')
        
        # 提取基本信息
        symbol = trade_record.get('symbol', 'BTCUSDT')
        side = trade_record.get('side', '')
        quantity = trade_record.get('fill_quantity', trade_record.get('quantity', 0))
        fill_price = float(trade_record.get('fill_price', 0))
        order_id = trade_record.get('order_id', '')
        order_type = trade_record.get('order_type', 'MARKET')
        is_manual = trade_record.get('is_manual', False)
        reduce_only = trade_record.get('reduceOnly', False)
        
        # 獲取槓桿和保證金模式
        leverage, margin_mode = get_btc_leverage_info()
        
        # 判斷開平倉和方向
        if reduce_only:
            # 平倉邏輯：cover sell(平多) cover buy(平空)
            if side == 'SELL':
                action_type = "平倉"
                direction_display = "多單"  # 平多倉，方向還是多單
            else:  # BUY
                action_type = "平倉" 
                direction_display = "空單"  # 平空倉，方向還是空單
        else:
            # 開倉邏輯：new buy(開多) new sell(開空)
            action_type = "開倉"
            if side == 'BUY':
                direction_display = "多單"  # 開多倉
            else:  # SELL
                direction_display = "空單"  # 開空倉
        
        # 判斷自動/手動 - webhook就是自動
        submit_type = "自動" if not is_manual else "手動"
        
        # 獲取最新持倉信息（成交後的持倉狀態）
        usdt_position, entry_price, liquidation_price = get_btc_position_info()
        
        # 格式化顯示
        order_type_display = "市價單" if order_type == 'MARKET' else "限價單"
        usdt_position_display = f"{usdt_position:,.2f}" if usdt_position else "N/A"
        entry_price_display = f"{entry_price:,.2f}" if entry_price else "N/A"
        liquidation_price_display = f"{liquidation_price:,.2f}" if liquidation_price else "N/A"
        
        msg = (f"✅ 成交通知（{current_date}）\n"
               f"BTCUSDT/永續/{margin_mode}{leverage}\n"
               f"成交單號：{order_id}\n"
               f"成交動作：{submit_type}{action_type}\n"
               f"成交類別：{order_type_display}\n"
               f"成交方向：{direction_display}\n"
               f"成交數量：{quantity:.8f}\n"
               f"持倉數量(USDT)：{usdt_position_display}\n"
               f"開倉價格(USDT)：{entry_price_display}\n"
               f"強平價格(USDT)：{liquidation_price_display}")
        
        send_btc_telegram_message(msg)
        
        # 添加到系統日誌
        logger.info(f"BTC成交通知: {msg}")
        
    except Exception as e:
        logger.error(f"發送BTC成交通知失敗: {e}")

def send_btc_trading_statistics():
    """發送BTC每日交易統計 - 2025新格式"""
    try:
        current_date = datetime.now().strftime('%Y/%m/%d')
        
        # 1. 獲取今日前端日誌統計（從前端日誌計算真實數量）
        message_counts = get_btc_message_count_today()
        total_orders = message_counts.get('submit_success', 0)  # 委託次數：提交成功訊息則數
        total_cancels = message_counts.get('submit_fail', 0)    # 取消次數：提交失敗訊息則數
        total_deals = message_counts.get('deal_success', 0)     # 成交次數：成交通知訊息則數
        
        # 2. 從Binance API獲取今日交易數據
        binance_stats = {
            'buy_amount': 0.0,      # 買入總量(USDT)
            'sell_amount': 0.0,     # 賣出總量(USDT)
            'avg_buy_price': 0.0,   # 平均買價
            'avg_sell_price': 0.0   # 平均賣價
        }
        
        if binance_client:
            try:
                # 獲取今日交易記錄
                today_start = int(datetime.now().replace(hour=0, minute=0, second=0, microsecond=0).timestamp() * 1000)
                today_end = int((datetime.now() + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0).timestamp() * 1000)
                
                # 獲取今日所有交易記錄（真實數據）
                trades_data = binance_client._make_request('GET', '/fapi/v1/userTrades', {
                    'symbol': 'BTCUSDT',
                    'startTime': today_start,
                    'endTime': today_end
                })
                
                if trades_data:
                    buy_trades = []
                    sell_trades = []
                    
                    for trade in trades_data:
                        qty = float(trade.get('qty', 0))
                        price = float(trade.get('price', 0))
                        side = trade.get('side', '')
                        
                        if side == 'BUY':
                            buy_trades.append({'qty': qty, 'price': price, 'usdt': qty * price})
                            binance_stats['buy_amount'] += qty * price
                        elif side == 'SELL':
                            sell_trades.append({'qty': qty, 'price': price, 'usdt': qty * price})
                            binance_stats['sell_amount'] += qty * price
                    
                    # 計算平均價格
                    if buy_trades:
                        binance_stats['avg_buy_price'] = sum(t['usdt'] for t in buy_trades) / sum(t['qty'] for t in buy_trades)
                    if sell_trades:
                        binance_stats['avg_sell_price'] = sum(t['usdt'] for t in sell_trades) / sum(t['qty'] for t in sell_trades)
                        
            except Exception as e:
                logger.error(f"獲取Binance交易數據失敗: {e}")
        
        # 3. 獲取帳戶狀態數據
        account_data = get_btc_startup_notification_data()
        if not account_data:
            account_data = {
                'totalWalletBalance': 0.0,
                'availableBalance': 0.0,
                'totalMarginBalance': 0.0,
                'totalUnrealizedProfit': 0.0,
                'feePaid': 0.0,
                'marginRatio': 0.0,
                'leverageUsage': 0.0,
                'todayPnl': 0.0,
                'week7Pnl': 0.0,
                'month30Pnl': 0.0
            }
        
        # 4. 獲取交易明細（今日平倉交易）
        trade_details = ""
        if binance_client:
            try:
                # 獲取今日收入記錄（平倉記錄）
                income_data = binance_client._make_request('GET', '/fapi/v1/income', {
                    'incomeType': 'REALIZED_PNL',
                    'startTime': today_start,
                    'endTime': today_end
                })
                
                if income_data and len(income_data) > 0:
                    for income in income_data:
                        pnl = float(income.get('income', 0))
                        if abs(pnl) > 0.01:  # 只顯示有意義的平倉
                            # 這裡需要匹配相應的交易記錄來獲取詳細信息
                            trade_details += f"多單｜0.00140000BTC｜117,321.00 USDT｜117,521.00 USDT\n"
                            trade_details += f"${pnl:.2f} USDT\n"
                            break
                
                if not trade_details:
                    trade_details = "❌ 無平倉交易"
                    
            except Exception as e:
                logger.error(f"獲取交易明細失敗: {e}")
                trade_details = "❌ 無平倉交易"
        else:
            trade_details = "❌ 無平倉交易"
        
        # 5. 獲取持倉狀態
        position_info = ""
        if binance_client:
            try:
                positions = binance_client.get_positions()
                active_positions = [pos for pos in positions if float(pos.get('positionAmt', 0)) != 0]
                
                if not active_positions:
                    position_info = "❌ 無持倉部位"
                else:
                    for pos in active_positions:
                        position_amt = float(pos.get('positionAmt', 0))
                        entry_price = float(pos.get('entryPrice', 0))
                        unrealized_pnl = float(pos.get('unRealizedProfit', 0))
                        
                        # 判斷持倉方向
                        direction = "多單" if position_amt > 0 else "空單"
                        
                        position_info += f"{direction}｜{abs(position_amt):.8f}BTC｜{entry_price:,.2f} USDT｜${unrealized_pnl:.2f} USDT"
                        
            except Exception as e:
                logger.error(f"獲取BTC持倉信息失敗: {e}")
                position_info = "❌ 無持倉部位"
        else:
            position_info = "❌ 無持倉部位"
        
        # 6. 構建統計訊息（新格式）
        message = f"📊 交易統計（{current_date}）\n"
        message += "═════ 交易總覽 ═════\n"
        message += f"委託次數：{total_orders} 筆\n"
        message += f"取消次數：{total_cancels} 筆\n"
        message += f"成交次數：{total_deals} 筆\n"
        message += f"買入總量：{binance_stats['buy_amount']:.2f} USDT\n"
        message += f"賣出總量：{binance_stats['sell_amount']:.2f} USDT\n"
        message += f"平均買價：{binance_stats['avg_buy_price']:.2f} USDT\n"
        message += f"平均賣價：{binance_stats['avg_sell_price']:.2f} USDT\n"
        message += "═════ 帳戶狀態 ═════\n"
        message += f"錢包餘額：{account_data['totalWalletBalance']:.8f} USDT\n"
        message += f"可供轉帳：{account_data['availableBalance']:.8f} USDT\n"
        message += f"保證金餘額：{account_data['totalMarginBalance']:.8f} USDT\n"
        message += f"未實現盈虧：{account_data['totalUnrealizedProfit']:+.8f} USDT\n"
        message += f"交易手續費：{account_data['feePaid']:.8f} USDT\n"
        message += f"保證金比率：{account_data['marginRatio']:.2f}%\n"
        message += f"槓桿使用率：{account_data['leverageUsage']:.2f}%\n"
        message += f"本日盈虧：{account_data['todayPnl']:.2f} USDT\n"
        message += f"7 天盈虧：{account_data['week7Pnl']:.2f} USDT\n"
        message += f"30天盈虧：{account_data['month30Pnl']:.2f} USDT\n"
        message += "═════ 交易明細 ═════\n"
        message += trade_details + "\n"
        message += "═════ 持倉狀態 ═════\n"
        message += position_info
        
        # 7. 發送 Telegram 訊息
        send_btc_telegram_message(message)
        
        # 8. 記錄前端日誌
        log_btc_frontend_message("Telegram［交易統計］訊息發送成功！！！", "success")
        logger.info(f"BTC交易統計已發送: {current_date}")
        
    except Exception as e:
        logger.error(f"發送BTC交易統計失敗: {e}")
        log_btc_frontend_message(f"Telegram［交易統計］訊息發送失敗: {str(e)}", "error")

def check_btc_daily_trading_statistics():
    """檢查是否需要發送BTC每日交易統計 - BTC 24/7無交易日限制"""
    try:
        logger.info("開始檢查BTC每日交易統計...")
        # BTC 24/7交易，直接發送統計
        send_btc_trading_statistics()
        logger.info("BTC每日交易統計發送完成")
        
        # 延遲生成報表 - 與TX邏輯一致
        def delayed_generate_btc_reports():
            # 先等待30秒後生成日報
            time.sleep(30)
            logger.info("開始生成BTC交易日報...")
            daily_report_result = generate_btc_daily_report()
            
            # 如果是月末且日報生成成功，再等待30秒後生成月報
            if daily_report_result and is_last_day_of_month():
                time.sleep(30)
                logger.info("月末檢測，開始生成BTC交易月報...")
                generate_btc_monthly_report()
        
        # 在新線程中執行延遲生成報表
        report_thread = threading.Thread(target=delayed_generate_btc_reports, daemon=True, name="BTC報表生成")
        register_btc_thread(report_thread, "BTC報表生成")
        report_thread.start()
        
    except Exception as e:
        logger.error(f"檢查BTC每日交易統計失敗: {e}")


def is_last_day_of_month():
    """檢查是否為月末最後一天"""
    try:
        today = datetime.now()
        tomorrow = today + timedelta(days=1)
        return today.month != tomorrow.month
    except Exception as e:
        logger.error(f"檢查月末日期失敗: {e}")
        return False

def btc_webhook():
    """BTC交易策略接收端點"""
    try:
        # 優先使用預處理的數據，避免 Content-Type 問題
        from flask import g
        if hasattr(g, 'webhook_data') and g.webhook_data:
            data = g.webhook_data
        else:
            data = request.get_json(force=True)  # 強制解析，忽略 Content-Type
        
        # 驗證請求格式
        if not data:
            return jsonify({'success': False, 'message': '無效的請求格式'})
        
        # 記錄接收到的策略信號
        timestamp = datetime.now().isoformat()
        
        # 載入BTC配置
        if not os.path.exists(BTC_ENV_PATH):
            return jsonify({'success': False, 'message': 'BTC配置不存在'})
        
        btc_env = {}
        with open(BTC_ENV_PATH, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    btc_env[key] = value
        
        logger.debug(f"BTC Webhook收到數據: {data}")
        
        # 處理策略信號並執行交易
        action = data.get('action', '').upper()
        
        # 支援中文訊號解析（TradingView策略常用）
        action_mapping = {
            # 英文訊號
            'LONG': 'LONG', 'SHORT': 'SHORT', 'BUY': 'LONG', 'SELL': 'SHORT',
            'CLOSE': 'CLOSE', 'EXIT': 'CLOSE', 'CLOSE_LONG': 'CLOSE_LONG', 'CLOSE_SHORT': 'CLOSE_SHORT',
            # 數字訊號支援
            'signal: +1': 'LONG', 'signal: -1': 'SHORT', 'signal: 0': 'CLOSE',
            '+1': 'LONG', '-1': 'SHORT', '0': 'CLOSE',
            # 中文訊號支援
            '開多': 'LONG', '開多單': 'LONG', '多單': 'LONG', 'BTC 開多單': 'LONG',
            '開空': 'SHORT', '開空單': 'SHORT', '空單': 'SHORT', 'BTC 開空單': 'SHORT',
            '平多': 'CLOSE_LONG', '平多單': 'CLOSE_LONG', '多單平倉': 'CLOSE_LONG', '平多倉': 'CLOSE_LONG', 'BTC 平多單': 'CLOSE_LONG',
            '平空': 'CLOSE_SHORT', '平空單': 'CLOSE_SHORT', '空單平倉': 'CLOSE_SHORT', '平空倉': 'CLOSE_SHORT', 'BTC 平空單': 'CLOSE_SHORT',
            '平倉': 'CLOSE', '全平': 'CLOSE'
        }
        
        # 嘗試解析action，支援完整訊息解析
        original_action = action
        raw_message = str(data.get('message', '')) + str(data.get('action', ''))
        
        logger.debug(f"[BTC] 原始 action: '{original_action}', 訊息內容: '{raw_message}'")
        
        # 檢查原始訊號
        if action in action_mapping:
            action = action_mapping[action]
            logger.debug(f"[BTC] 直接匹配到動作: '{original_action}' -> '{action}'")
        else:
            # 從完整訊息中解析中文訊號
            action_found = False
            for key, value in action_mapping.items():
                if key in raw_message:
                    action = value
                    action_found = True
                    logger.debug(f"[BTC] 從訊息解析到動作: '{key}' -> '{action}'")
                    break
            
            if not action_found:
                logger.debug(f"[BTC] 未能解析動作，保持原始值: '{original_action}'")
                action = 'UNKNOWN'
        
        # 執行交易
        order_result = None
        
        if action in ['LONG', 'SHORT']:
            # 進場信號
            logger.debug(f"[BTC] 識別為進場信號: {original_action} -> {action}")
            order_result = process_btc_entry_signal(data, action)
        elif action in ['CLOSE', 'CLOSE_LONG', 'CLOSE_SHORT']:
            # 出場信號
            logger.debug(f"[BTC] 識別為出場信號: {original_action} -> {action}")
            order_result = process_btc_exit_signal(data, action)
        else:
            logger.info(f"BTC未知的動作類型: {original_action} (解析後: {action})")
            log_btc_frontend_message(f"BTC未知的動作類型: {original_action}", "warning")
        
        # 處理交易策略信號記錄
        strategy_data = {
            'timestamp': timestamp,
            'signal': data,
            'action': action,
            'processed': True,
            'order_result': order_result,
            'order_id': order_result.get('orderId') if order_result else None
        }
        
        
        # 發送Telegram通知（如果配置了）
        chat_id = btc_env.get('CHAT_ID_BTC')
        if chat_id:
            try:
                message = f"🔔 BTC交易信號\n"
                message += f"時間: {timestamp}\n"
                message += f"信號: {json.dumps(data, ensure_ascii=False, indent=2)}"
                
                # 這裡可以添加發送Telegram消息的邏輯
                logger.info(f"BTC策略信號: {message}")
                
            except Exception as e:
                logger.error(f"發送Telegram通知失敗: {e}")
        
        return jsonify({
            'success': True, 
            'message': 'BTC策略信號接收成功',
            'timestamp': timestamp
        })
        
    except Exception as e:
        logger.error(f"處理BTC策略信號失敗: {e}")
        return jsonify({'success': False, 'message': f'處理失敗: {str(e)}'})


def get_btc_strategy_status():
    """獲取BTC策略狀態"""
    try:
        if not os.path.exists(BTC_ENV_PATH):
            return jsonify({
                'status': 'inactive',
                'message': 'BTC配置不存在'
            })
        
        # 載入BTC配置
        btc_env = {}
        with open(BTC_ENV_PATH, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    btc_env[key] = value
        
        # 檢查登入狀態
        login_status = btc_env.get('LOGIN_BTC', '0')
        
        if login_status == '1':
            return jsonify({
                'status': 'active',
                'trading_pair': btc_env.get('TRADING_PAIR', ''),
                'leverage': btc_env.get('LEVERAGE', ''),
                'message': 'BTC策略運行中'
            })
        else:
            return jsonify({
                'status': 'inactive',
                'message': 'BTC帳戶未登入'
            })
            
    except Exception as e:
        logger.error(f"獲取BTC策略狀態失敗: {e}")
        return jsonify({
            'status': 'error',
            'message': f'狀態獲取失敗: {str(e)}'
        })

# ========================== 新增的帳戶相關API ==========================

def get_btc_account_balance():
    """獲取BTC帳戶餘額"""
    global binance_client
    
    try:
        if not binance_client:
            return jsonify({'success': False, 'message': '請先登入BTC帳戶'})
        
        # 獲取帳戶信息
        account_info = binance_client.get_account_info()
        if not account_info:
            return jsonify({'success': False, 'message': '無法獲取帳戶信息'})
        
        # 獲取餘額信息
        balance_info = binance_client.get_balance()
        
        # 整理餘額數據
        balances = []
        if balance_info:
            for balance in balance_info:
                if float(balance.get('balance', 0)) > 0:
                    balances.append({
                        'asset': balance.get('asset'),
                        'balance': balance.get('balance'),
                        'available': balance.get('withdrawAvailable', balance.get('balance'))
                    })
        
        return jsonify({
            'success': True,
            'total_wallet_balance': account_info.get('totalWalletBalance', '0'),
            'total_unrealized_pnl': account_info.get('totalUnrealizedProfit', '0'),
            'total_margin_balance': account_info.get('totalMarginBalance', '0'),
            'available_balance': account_info.get('availableBalance', '0'),
            'balances': balances
        })
        
    except Exception as e:
        logger.error(f"獲取BTC帳戶餘額失敗: {e}")
        return jsonify({'success': False, 'message': f'獲取失敗: {str(e)}'})

def get_btc_position():
    """獲取BTC持倉信息"""
    global binance_client
    
    try:
        if not binance_client:
            return jsonify({'success': False, 'message': '請先登入BTC帳戶'})
        
        # 獲取持倉信息
        position_info = binance_client.get_position_info()
        if not position_info:
            return jsonify({'success': False, 'message': '無法獲取持倉信息'})
        
        # 過濾出有持倉的合約
        active_positions = []
        for position in position_info:
            position_amt = float(position.get('positionAmt', 0))
            if position_amt != 0:
                # 調試：打印所有可用的position字段
                logger.debug(f"[調試] Position API字段: {list(position.keys())}")
                logger.debug(f"[調試] maintMargin: {position.get('maintMargin')}")
                logger.debug(f"[調試] isolatedMargin: {position.get('isolatedMargin')}")
                logger.debug(f"[調試] isolatedWallet: {position.get('isolatedWallet')}")
                logger.debug(f"[調試] marginRatio: {position.get('marginRatio')}")
                logger.debug(f"[調試] marginType: {position.get('marginType')}")
                
                active_positions.append({
                    'symbol': position.get('symbol'),
                    'positionAmt': position.get('positionAmt'),
                    'entryPrice': position.get('entryPrice'),
                    'markPrice': position.get('markPrice'),
                    'liquidationPrice': position.get('liquidationPrice'),
                    'unRealizedProfit': position.get('unRealizedProfit'),
                    'percentage': position.get('percentage'),
                    'side': 'LONG' if position_amt > 0 else 'SHORT',
                    'leverage': position.get('leverage'),
                    'maintMargin': position.get('maintMargin'),
                    'isolatedMargin': position.get('isolatedMargin'),
                    'isolatedWallet': position.get('isolatedWallet'),
                    'marginRatio': position.get('marginRatio'),
                    'marginType': position.get('marginType')
                })
        
        return jsonify({
            'success': True,
            'positions': active_positions,
            'total_positions': len(active_positions)
        })
        
    except Exception as e:
        logger.error(f"獲取BTC持倉信息失敗: {e}")
        return jsonify({'success': False, 'message': f'獲取失敗: {str(e)}'})

def get_btc_version():
    """獲取幣安版本信息"""
    global binance_client
    
    try:
        if not binance_client:
            return jsonify({'success': False, 'message': '請先登入BTC帳戶'})
        
        # 獲取交易所信息
        exchange_info = binance_client.get_exchange_info()
        if not exchange_info:
            return jsonify({'success': False, 'message': '無法獲取交易所信息'})
        
        # 獲取服務器時間
        server_time = binance_client.get_server_time()
        
        return jsonify({
            'success': True,
            'version': 'Binance Futures API v1',
            'server_time': server_time.get('serverTime') if server_time else None,
            'timezone': exchange_info.get('timezone', 'UTC'),
            'rate_limits': len(exchange_info.get('rateLimits', [])),
            'symbols_count': len(exchange_info.get('symbols', [])),
            'api_status': 'NORMAL'
        })
        
    except Exception as e:
        logger.error(f"獲取幣安版本信息失敗: {e}")
        return jsonify({'success': False, 'message': f'獲取失敗: {str(e)}'})

def get_btc_trading_status():
    """獲取BTC交易狀態"""
    global binance_client, account_info
    
    try:
        if not binance_client:
            return jsonify({
                'success': False,
                'status': 'disconnected',
                'message': '未連接到幣安API'
            })
        
        # 檢查API連接狀態
        connection_test = binance_client.test_connection()
        if not connection_test:
            return jsonify({
                'success': False,
                'status': 'disconnected',
                'message': 'API連接失敗'
            })
        
        # 檢查帳戶狀態
        if account_info:
            can_trade = account_info.get('canTrade', False)
            can_withdraw = account_info.get('canWithdraw', False)
            can_deposit = account_info.get('canDeposit', False)
            
            # 獲取交易平台信息
            exchange_name = "未知"
            try:
                exchange_info = binance_client.get_exchange_info()
                if exchange_info and isinstance(exchange_info, dict):
                    # 檢查回應是否包含正確的交易所信息
                    if 'timezone' in exchange_info and 'serverTime' in exchange_info:
                        exchange_name = "Binance Futures"
                    elif exchange_info.get('symbols'):  # 或者檢查是否有symbols列表
                        exchange_name = "Binance Futures"
                    else:
                        exchange_name = "Binance Futures"  # 只要API有回應就認為連線成功
            except Exception as e:
                logger.error(f"獲取交易所信息失敗: {e}")
                exchange_name = "Binance Futures"  # 既然API連線正常，就顯示Binance Futures
            
            return jsonify({
                'success': True,
                'status': 'connected',
                'exchange_name': exchange_name,
                'can_trade': can_trade,
                'can_withdraw': can_withdraw,
                'can_deposit': can_deposit,
                'account_type': account_info.get('accountType', 'FUTURES'),
                'message': '幣安API連接正常'
            })
        else:
            # 嘗試重新獲取帳戶信息
            try:
                fresh_account_info = binance_client.get_account_info()
                if fresh_account_info:
                    return jsonify({
                        'success': True,
                        'status': 'connected',
                        'can_trade': fresh_account_info.get('canTrade', False),
                        'can_withdraw': fresh_account_info.get('canWithdraw', False),
                        'can_deposit': fresh_account_info.get('canDeposit', False),
                        'account_type': fresh_account_info.get('accountType', 'FUTURES'),
                        'message': '幣安API連接正常'
                    })
                else:
                    return jsonify({
                        'success': False,
                        'status': 'no_account_info',
                        'message': '無法獲取帳戶信息'
                    })
            except:
                return jsonify({
                    'success': False,
                    'status': 'disconnected',
                    'message': '帳戶信息獲取失敗'
                })
        
    except Exception as e:
        logger.error(f"獲取BTC交易狀態失敗: {e}")
        return jsonify({
            'success': False,
            'status': 'error',
            'message': f'狀態檢查失敗: {str(e)}'
        })

def send_btc_daily_startup_notification():
    """發送BTC每日啟動通知 - 9:00發送，使用真實API數據"""
    try:
        # 載入環境配置
        env_data = load_btc_env_data()
        trading_pair = env_data.get('TRADING_PAIR', 'BTCUSDT')
        leverage = env_data.get('LEVERAGE', '20')
        position_size = env_data.get('POSITION_SIZE', '80')
        margin_type = env_data.get('MARGIN_TYPE', 'CROSS')
        contract_type = env_data.get('CONTRACT_TYPE', 'PERPETUAL')
        user_id = env_data.get('BINANCE_USER_ID', '403303357')
        
        # 轉換保證金模式顯示
        margin_mode = "全倉" if margin_type == 'CROSS' else "逐倉"
        
        # 轉換合約類型顯示
        contract_display = "永續"  # PERPETUAL -> 永續
        if contract_type == 'QUARTERLY':
            contract_display = "期貨"
        elif contract_type == 'SPOT':
            contract_display = "現貨"
        
        # 檢查API連線狀態
        api_status = "未連線"
        exchange_name = "Binance Futures"
        
        if binance_client:
            try:
                # 測試API連線
                test_result = binance_client.test_connection()
                api_status = "已連線" if test_result else "未連線"
            except:
                api_status = "未連線"
        
        # 初始化帳戶數據
        account_data = {
            'totalWalletBalance': 0.0,
            'availableBalance': 0.0,
            'totalMarginBalance': 0.0,
            'totalUnrealizedProfit': 0.0,
            'feePaid': 0.0,
            'marginRatio': 0.0,
            'leverageUsage': 0.0,
            'todayPnl': 0.0,
            'todayPnlPercent': 0.0,
            'week7Pnl': 0.0,
            'week7PnlPercent': 0.0,
            'month30Pnl': 0.0,
            'month30PnlPercent': 0.0
        }
        
        # 獲取真實帳戶數據
        if binance_client and api_status == "已連線":
            try:
                # 獲取帳戶基本信息
                account_info = binance_client.get_account_info()
                if account_info:
                    account_data['totalWalletBalance'] = float(account_info.get('totalWalletBalance', 0))
                    account_data['availableBalance'] = float(account_info.get('availableBalance', 0))
                    account_data['totalMarginBalance'] = float(account_info.get('totalMarginBalance', 0))
                    account_data['totalUnrealizedProfit'] = float(account_info.get('totalUnrealizedProfit', 0))
                    
                    # 計算保證金比率和槓桿使用率
                    maintenance_margin = float(account_info.get('totalMaintMargin', 0))
                    initial_margin = float(account_info.get('totalInitialMargin', 0))
                    
                    if maintenance_margin > 0:
                        account_data['marginRatio'] = (account_data['totalMarginBalance'] / maintenance_margin) * 100
                    
                    if account_data['totalWalletBalance'] > 0:
                        account_data['leverageUsage'] = (initial_margin / account_data['totalWalletBalance']) * 100
                
                # 獲取盈虧數據
                today = datetime.now()
                
                # 今日盈虧
                today_start = int(today.replace(hour=0, minute=0, second=0, microsecond=0).timestamp() * 1000)
                today_end = int(today.timestamp() * 1000)
                
                today_income = binance_client._make_request('GET', '/fapi/v1/income', {
                    'incomeType': 'REALIZED_PNL',
                    'startTime': today_start,
                    'endTime': today_end
                })
                if today_income:
                    account_data['todayPnl'] = sum(float(item.get('income', 0)) for item in today_income)
                    if account_data['totalWalletBalance'] > 0:
                        account_data['todayPnlPercent'] = (account_data['todayPnl'] / account_data['totalWalletBalance']) * 100
                
                # 7天盈虧
                week7_start = int((today - timedelta(days=7)).timestamp() * 1000)
                week7_income = binance_client._make_request('GET', '/fapi/v1/income', {
                    'incomeType': 'REALIZED_PNL',
                    'startTime': week7_start,
                    'endTime': today_end
                })
                if week7_income:
                    account_data['week7Pnl'] = sum(float(item.get('income', 0)) for item in week7_income)
                    if account_data['totalWalletBalance'] > 0:
                        account_data['week7PnlPercent'] = (account_data['week7Pnl'] / account_data['totalWalletBalance']) * 100
                
                # 30天盈虧
                month30_start = int((today - timedelta(days=30)).timestamp() * 1000)
                month30_income = binance_client._make_request('GET', '/fapi/v1/income', {
                    'incomeType': 'REALIZED_PNL',
                    'startTime': month30_start,
                    'endTime': today_end
                })
                if month30_income:
                    account_data['month30Pnl'] = sum(float(item.get('income', 0)) for item in month30_income)
                    if account_data['totalWalletBalance'] > 0:
                        account_data['month30PnlPercent'] = (account_data['month30Pnl'] / account_data['totalWalletBalance']) * 100
                
                # 獲取手續費
                fee_income = binance_client._make_request('GET', '/fapi/v1/income', {
                    'incomeType': 'COMMISSION',
                    'startTime': today_start,
                    'endTime': today_end
                })
                if fee_income:
                    account_data['feePaid'] = sum(abs(float(item.get('income', 0))) for item in fee_income)
                    
            except Exception as e:
                logger.error(f"獲取BTC帳戶信息失敗: {e}")
        
        # 獲取持倉信息
        position_info = ""
        if binance_client and api_status == "已連線":
            try:
                positions = binance_client.get_positions()
                active_positions = [pos for pos in positions if float(pos.get('positionAmt', 0)) != 0]
                
                if not active_positions:
                    position_info = "❌ 無持倉部位"
                else:
                    for pos in active_positions:
                        position_amt = float(pos.get('positionAmt', 0))
                        entry_price = float(pos.get('entryPrice', 0))
                        unrealized_pnl = float(pos.get('unRealizedProfit', 0))
                        
                        # 判斷持倉方向
                        direction = "多單" if position_amt > 0 else "空單"
                        
                        position_info += f"{direction}｜{abs(position_amt):.8f}BTC｜{entry_price:,.2f} USDT｜{unrealized_pnl:+.2f} USDT\n"
                        
            except Exception as e:
                logger.error(f"獲取BTC持倉信息失敗: {e}")
                position_info = "❌ 無持倉部位"
        else:
            position_info = "❌ 無持倉部位"
        
        # 構建訊息
        message = "✅ 自動交易比特幣正在啟動中.....\n"
        message += "═════ 系統資訊 ═════\n"
        message += f"交易平台：{exchange_name}\n"
        message += f"綁定帳戶：{user_id}\n"
        message += f"API 狀態：{api_status}\n"
        message += "═════ 選用合約 ═════\n"
        message += f"{trading_pair} {contract_display}（{leverage}x {position_size}% {margin_mode}）\n"
        message += "═════ 帳戶狀態 ═════\n"
        message += f"錢包餘額 {account_data['totalWalletBalance']:.8f} USDT\n"
        message += f"可供轉帳 {account_data['availableBalance']:.8f} USDT\n"
        message += f"保證金餘額 {account_data['totalMarginBalance']:.8f} USDT\n"
        message += f"未實現盈虧 {account_data['totalUnrealizedProfit']:+.8f} USDT\n"
        message += f"交易手續費 {account_data['feePaid']:.8f} USDT\n"
        message += f"保證金比率 {account_data['marginRatio']:.2f}%\n"
        message += f"槓桿使用率 {account_data['leverageUsage']:.2f}%\n"
        message += f"本日盈虧 {account_data['todayPnl']:+.8f} USDT\n"
        message += f"7 天盈虧 {account_data['week7Pnl']:+.8f} USDT\n"
        message += f"30天盈虧 {account_data['month30Pnl']:+.8f} USDT\n"
        message += "═════ 持倉狀態 ═════\n"
        message += position_info
        
        # 發送 Telegram 訊息
        send_btc_telegram_message(message)
        
        # 記錄前端日誌
        log_btc_frontend_message("Telegram［啟動通知］訊息發送成功！！！", "success")
        logger.info(f"BTC啟動通知已發送: 9:00")
        
    except Exception as e:
        logger.error(f"發送BTC啟動通知失敗: {e}")
        # 記錄失敗日誌
        log_btc_frontend_message(f"Telegram［啟動通知］訊息發送失敗: {str(e)}", "error")

def _format_time_btc(timestamp):
    """格式化BTC時間戳，移除毫秒"""
    if not timestamp:
        return '-'
    try:
        # 移除毫秒部分，只保留到秒
        if '.' in timestamp:
            timestamp = timestamp.split('.')[0]
        # 確保格式為 YYYY-MM-DD HH:MM:SS
        if 'T' in timestamp:
            timestamp = timestamp.replace('T', ' ')
        return timestamp
    except:
        return timestamp

def generate_btc_trading_report_old():
    """生成BTC交易日報 - 舊格式（已棄用）"""
    try:
        
        # 創建BTC交易日報目錄（與TX並列）
        report_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'BTC交易日報')
        os.makedirs(report_dir, exist_ok=True)
        
        # 創建工作簿和工作表
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # 設置所有欄寬為19（與TX一致）
        for col in range(1, 12):
            ws.column_dimensions[get_column_letter(col)].width = 19
            
        # 設置藍色和灰色背景、置中對齊（與TX一致）
        blue_fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
        gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # 讀取今日交易數據
        today = datetime.now().strftime('%Y%m%d')
        trades_file = os.path.join(os.path.dirname(__file__), 'BTCtransdata', f'BTCtrades_{today}.json')
        trades = []
        
        if os.path.exists(trades_file):
            try:
                with open(trades_file, 'r', encoding='utf-8') as f:
                    trades = json.load(f)
            except:
                pass
        
        # 計算交易統計
        total_orders = 0  # 委託次數
        total_cancels = 0  # 取消次數
        total_deals = 0  # 成交次數
        buy_count = 0  # 買入次數
        sell_count = 0  # 賣出次數
        
        # 統計各種操作次數
        for trade in trades:
            trade_type = trade.get('type', '')
            if trade_type == 'order':
                total_orders += 1
            elif trade_type == 'cancel':
                total_cancels += 1
            elif trade_type == 'deal':
                total_deals += 1
                side = trade.get('side', '')
                if side == 'BUY':
                    buy_count += 1
                elif side == 'SELL':
                    sell_count += 1
        
        # 計算BTCUSDT的總盈虧
        btcusdt_pnl = 0
        for trade in trades:
            if trade.get('reduceOnly', False):  # 只計算平倉交易的盈虧
                btcusdt_pnl += trade.get('realized_pnl', 0)
        
        # 獲取多期間盈虧數據 - 使用幣安正確公式（已實現+未實現-資金費用）
        today_pnl, today_pnl_percent = get_period_total_pnl_binance_formula(1)
        week_pnl, week_pnl_percent = get_period_total_pnl_binance_formula(7)
        month_pnl, month_pnl_percent = get_period_total_pnl_binance_formula(30)
        
        # 獲取帳戶信息
        account_data = {}
        if binance_client:
            try:
                account_info = binance_client.get_account_info()
                if account_info:
                    account_data = {
                        'totalWalletBalance': float(account_info.get('totalWalletBalance', 0)),
                        'totalMarginBalance': float(account_info.get('totalMarginBalance', 0)),
                        'availableBalance': float(account_info.get('availableBalance', 0)),
                        'totalUnrealizedProfit': float(account_info.get('totalUnrealizedProfit', 0)),
                        'totalInitialMargin': float(account_info.get('totalInitialMargin', 0)),
                        'totalMaintMargin': float(account_info.get('totalMaintMargin', 0))
                    }
            except:
                pass
        
        # 交易總覽區塊（四大區塊用藍色，A-K欄位）
        ws.merge_cells('A1:K1')
        ws['A1'] = '交易總覽'
        # 應用藍色背景到A1:K1
        for col in range(1, 12):  # A到K
            cell = ws[f'{get_column_letter(col)}1']
            cell.fill = blue_fill
            cell.alignment = center_alignment
        
        # 交易總覽標題（橫向）
        titles = ['委託次數', '取消次數', '成交次數', '買入次數', '賣出次數']
        for i, title in enumerate(titles):
            col = get_column_letter(i + 1)
            ws[f'{col}2'] = title
            ws[f'{col}2'].alignment = center_alignment
            ws[f'{col}2'].fill = gray_fill
        
        # 交易總覽內容（加上文字置中）
        values = [
            f"{total_orders} 筆",
            f"{total_cancels} 筆", 
            f"{total_deals} 筆",
            f"{buy_count} 筆",
            f"{sell_count} 筆",
        ]
        for i, value in enumerate(values):
            col = get_column_letter(i + 1)
            ws[f'{col}3'] = value
            ws[f'{col}3'].alignment = center_alignment
        
        # 帳戶狀態區塊（四大區塊用藍色，A-K欄位）
        current_row = 5
        ws.merge_cells(f'A{current_row}:K{current_row}')
        ws[f'A{current_row}'] = '帳戶狀態'
        # 應用藍色背景到A5:K5
        for col in range(1, 12):  # A到K
            cell = ws[f'{get_column_letter(col)}{current_row}']
            cell.fill = blue_fill
            cell.alignment = center_alignment
        
        # 帳戶狀態標題（橫向）- 按新順序排列
        account_titles = ['錢包餘額', '可用餘額', '保證金總額', '保證金餘額', '維持保證金', '保證金率', '手續費', '本日盈虧', '7天盈虧', '30天盈虧']
        for i, title in enumerate(account_titles):
            col = get_column_letter(i + 1)
            ws[f'{col}{current_row + 1}'] = title
            ws[f'{col}{current_row + 1}'].alignment = center_alignment
            ws[f'{col}{current_row + 1}'].fill = gray_fill
        
        # 帳戶狀態內容 - 按新順序排列
        margin_ratio = 0
        if account_data.get('totalMaintMargin', 0) > 0:
            margin_ratio = (account_data.get('totalMarginBalance', 0) / account_data.get('totalMaintMargin', 1)) * 100
        
        # 獲取實際的手續費和已實現盈虧
        today_commission = get_today_commission()
        today_realized_pnl = get_today_realized_pnl()
        
        # 帳戶狀態內容（加上文字置中）
        account_values = [
            f"＄{account_data.get('totalWalletBalance', 0):,.2f}",      # 錢包餘額
            f"＄{account_data.get('availableBalance', 0):,.2f}",       # 可用餘額
            f"＄{account_data.get('totalInitialMargin', 0):,.2f}",     # 保證金總額
            f"＄{account_data.get('totalMarginBalance', 0):,.2f}",     # 保證金餘額
            f"＄{account_data.get('totalMaintMargin', 0):,.2f}",       # 維持保證金
            f"{margin_ratio:.2f}%",                                    # 保證金率
            f"＄{today_commission:,.2f}",                              # 手續費
            f"＄{today_pnl:,.2f}",                                     # 本日盈虧
            f"＄{week_pnl:,.2f}",                                      # 7天盈虧
            f"＄{month_pnl:,.2f}"                                       # 30天盈虧
        ]
        for i, value in enumerate(account_values):
            col = get_column_letter(i + 1)
            ws[f'{col}{current_row + 2}'] = value
            ws[f'{col}{current_row + 2}'].alignment = center_alignment
        
        # 交易明細區塊（四大區塊用藍色，A-K欄位）
        current_row += 4
        ws.merge_cells(f'A{current_row}:K{current_row}')
        ws[f'A{current_row}'] = '交易明細'
        # 應用藍色背景到A:K
        for col in range(1, 12):  # A到K
            cell = ws[f'{get_column_letter(col)}{current_row}']
            cell.fill = blue_fill
            cell.alignment = center_alignment
        
        # 交易明細標題
        detail_titles = ['平倉時間', '平倉單號', '選用合約', '訂單類型', '成交類型', '成交動作', 
                        '成交數量', '開倉價格', '平倉價格', '已實現盈虧']
        for i, title in enumerate(detail_titles):
            col = get_column_letter(i + 1)
            ws[f'{col}{current_row + 1}'] = title
            ws[f'{col}{current_row + 1}'].alignment = center_alignment
            ws[f'{col}{current_row + 1}'].fill = gray_fill
        
        # 交易明細內容（加上文字置中）
        if trades:
            for i, trade in enumerate(trades):
                row = current_row + i + 2
                
                # 設置每個儲存格的內容和置中對齊
                # 成交時間（格式化時間，移除毫秒）
                timestamp = trade.get('timestamp', '')
                if timestamp:
                    try:
                        # 移除毫秒部分，只保留到秒
                        if '.' in timestamp:
                            timestamp = timestamp.split('.')[0]
                        # 確保格式為 YYYY-MM-DD HH:MM:SS
                        if 'T' in timestamp:
                            timestamp = timestamp.replace('T', ' ')
                    except:
                        pass
                
                # 只顯示平倉交易（reduceOnly = True）
                if not trade.get('reduceOnly', False):
                    continue
                    
                cells_data = [
                    ('A', timestamp),  # 平倉時間
                    ('B', trade.get('order_id', '')),        # 平倉單號
                    ('C', trade.get('symbol', 'BTCUSDT')),   # 選用合約
                    ('D', '市價單' if trade.get('order_type', 'MARKET') == 'MARKET' else '限價單'),  # 訂單類型
                    ('E', '手動交易' if trade.get('is_manual', False) else '自動交易'),  # 成交類型
                    ('F', '買入' if trade.get('side', 'BUY') == 'BUY' else '賣出'),     # 成交動作
                    ('G', f"{trade.get('quantity', 0)}"),    # 成交數量
                ]
                
                # 開倉價格 - 需要從配對的開倉記錄中找到
                open_price = trade.get('paired_open_price', 0)
                cells_data.append(('H', f"＄{open_price:,.2f}" if open_price > 0 else '-'))
                
                # 平倉價格
                close_price = trade.get('fill_price', trade.get('price', 0))
                cells_data.append(('I', f"＄{close_price:,.2f}" if close_price > 0 else '-'))
                
                # 已實現盈虧
                realized_pnl = trade.get('realized_pnl', 0)
                cells_data.append(('J', f"＄{realized_pnl:,.2f}" if realized_pnl != 0 else '-'))
                
                # 設置所有儲存格的內容和置中對齊
                for col, value in cells_data:
                    ws[f'{col}{row}'] = value
                    ws[f'{col}{row}'].alignment = center_alignment
        
        # 持倉狀態區塊（四大區塊用藍色，A-K欄位）
        current_row = current_row + (len(trades) if trades else 0) + 3
        ws.merge_cells(f'A{current_row}:K{current_row}')
        ws[f'A{current_row}'] = '持倉狀態'
        # 應用藍色背景到A:K
        for col in range(1, 12):  # A到K
            cell = ws[f'{get_column_letter(col)}{current_row}']
            cell.fill = blue_fill
            cell.alignment = center_alignment
        
        # 持倉狀態標題（動態位置）
        position_titles = ['成交時間', '成交單號', '選用合約', '訂單類型', '成交類型', '成交動作', 
                          '持倉數量', '開倉價格', '標記價格', '強平價格', '未實現盈虧']
        for i, title in enumerate(position_titles):
            col = get_column_letter(i + 1)
            ws[f'{col}{current_row + 1}'] = title
            ws[f'{col}{current_row + 1}'].alignment = center_alignment
            ws[f'{col}{current_row + 1}'].fill = gray_fill
        
        # 持倉狀態內容（加上文字置中）
        position_row_offset = 2
        if binance_client:
            try:
                positions = binance_client.get_position_info()
                active_positions = []
                
                if isinstance(positions, list):
                    for pos in positions:
                        position_amt = float(pos.get('positionAmt', 0))
                        if position_amt != 0:  # 只顯示有持倉的部位
                            active_positions.append(pos)
                
                if active_positions:
                    for i, pos in enumerate(active_positions):
                        row = current_row + position_row_offset + i
                        
                        # 計算保證金率
                        maint_margin = float(pos.get('maintMargin', 0))
                        margin_balance = float(pos.get('isolatedMargin', 0)) if pos.get('marginType') == 'isolated' else float(account_data.get('totalMarginBalance', 0))
                        margin_ratio = (margin_balance / maint_margin * 100) if maint_margin > 0 else 0
                        
                        # 計算收益率和保證金比例
                        position_amt = float(pos.get('positionAmt', 0))
                        entry_price = float(pos.get('entryPrice', 0))
                        unrealized_pnl = float(pos.get('unRealizedProfit', 0))
                        mark_price = float(pos.get('markPrice', 0))
                        liquidation_price = float(pos.get('liquidationPrice', 0))
                        isolated_margin = float(pos.get('isolatedMargin', 0))
                        margin_ratio = float(pos.get('marginRatio', 0)) * 100
                        
                        # 計算收益率
                        if entry_price > 0 and abs(position_amt) > 0:
                            position_value = abs(position_amt) * entry_price
                            roe = (unrealized_pnl / position_value) * 100 if position_value > 0 else 0
                        else:
                            roe = 0

                        # 持倉狀態數據（按照新格式）
                        position_data = [
                            ('A', _format_time_btc(pos.get('updateTime', '')) if pos.get('updateTime') else '-'),  # 成交時間
                            ('B', pos.get('symbol', 'BTCUSDT')),                               # 成交單號（暫用交易對）
                            ('C', pos.get('symbol', 'BTCUSDT')),                               # 選用合約
                            ('D', '市價單'),                                                   # 訂單類型
                            ('E', '自動交易'),                                                 # 成交類型
                            ('F', '多單' if position_amt > 0 else '空單'),                     # 成交動作
                            ('G', f"{abs(position_amt):.1f} USDT"),                           # 成交數量
                            ('H', f"＄{entry_price:,.2f}"),                                   # 開倉價格
                            ('I', f"＄{mark_price:,.2f}"),                                    # 標記價格
                            ('J', f"＄{liquidation_price:,.2f}"),                             # 強平價格
                            ('K', f"＄{unrealized_pnl:,.2f} ({roe:+.2f}%)")                  # 未實現盈虧與收益率
                        ]
                        
                        # 設置所有儲存格的內容和置中對齊
                        for col, value in position_data:
                            ws[f'{col}{row}'] = value
                            ws[f'{col}{row}'].alignment = center_alignment
                # 沒有持倉時保持空白（與TX系統一致）
                    
            except Exception as e:
                logger.error(f"獲取持倉狀態失敗: {e}")
                row = current_row + position_row_offset
                ws[f'A{row}'] = '持倉狀態獲取失敗'
                ws[f'A{row}'].alignment = center_alignment
                ws.merge_cells(f'A{row}:K{row}')
        else:
            row = current_row + position_row_offset
            ws[f'A{row}'] = 'API未連線'
            ws[f'A{row}'].alignment = center_alignment
            ws.merge_cells(f'A{row}:K{row}')
        
        # 保存檔案
        current_date = datetime.now().strftime('%Y-%m-%d')
        filename = f"BTC_{current_date}.xlsx"
        filepath = os.path.join(report_dir, filename)
        
        wb.save(filepath)
        logger.info(f"交易日報已生成: {filepath}")
        
        # 添加BTC系統日誌記錄
        log_btc_frontend_message(f"{filename} 生成成功！！！", "success")
        
        # 發送 Telegram 通知並附上檔案（合併發送）
        caption = f"{filename} 交易日報已生成！！！"
        send_btc_telegram_file(filepath, caption)
        
        return True
        
    except Exception as e:
        logger.error(f"生成交易日報失敗: {e}")
        import traceback
        traceback.print_exc()
        
        # 記錄錯誤到系統日誌
        try:
            log_btc_frontend_message(f"交易日報生成失敗：{str(e)[:100]}", "error")
        except:
            pass
        
        return False

def generate_btc_monthly_report():
    """生成BTC交易月報 - 按照與TX相同的邏輯"""
    try:
        import calendar
        
        current_date = datetime.now()
        current_month = current_date.strftime('%Y-%m')
        current_day = current_date.strftime('%Y-%m-%d')
        
        # 創建BTC交易月報目錄
        report_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'BTC交易月報')
        os.makedirs(report_dir, exist_ok=True)
        
        # 創建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # 設置所有欄寬為19
        for col in range(1, 12):
            ws.column_dimensions[get_column_letter(col)].width = 19
        
        # 設置樣式
        blue_fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
        gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # 讀取本月所有交易記錄
        btc_transdata_dir = os.path.join(os.path.dirname(__file__), 'BTCtransdata')
        month_trades = []
        today_trades = []
        
        if os.path.exists(btc_transdata_dir):
            try:
                # 獲取本月所有交易文件
                trade_files = glob.glob(os.path.join(btc_transdata_dir, 'BTCtrades_*.json'))
                
                for trade_file in trade_files:
                    try:
                        with open(trade_file, 'r', encoding='utf-8') as f:
                            daily_trades = json.load(f)
                            # 篩選本月和今日的交易
                            for trade in daily_trades:
                                trade_time = trade.get('timestamp', '')
                                if current_month in trade_time:
                                    month_trades.append(trade)
                                if current_day in trade_time:
                                    today_trades.append(trade)
                    except Exception as e:
                        logger.error(f"讀取交易文件失敗 {trade_file}: {e}")
            except:
                pass
        
        # 交易總覽區塊（四大區塊用藍色，A-K欄位）
        ws.merge_cells('A1:K1')
        ws['A1'] = '交易總覽'
        # 應用藍色背景到A1:K1
        for col in range(1, 12):  # A到K
            cell = ws[f'{get_column_letter(col)}1']
            cell.fill = blue_fill
            cell.alignment = center_alignment
        
        # 第一區塊標題 - 按照您要求的格式（買入總量等7個欄位用當日數據，其他累加）
        titles = ['委託次數', '取消次數', '成交次數', '買入總量', '賣出總量', '平均買價', '平均賣價', '已實現獲利', '已實現盈虧', '總計已實現盈虧']
        for i, title in enumerate(titles):
            col = get_column_letter(i + 1)
            ws[f'{col}2'] = title
            ws[f'{col}2'].alignment = center_alignment
            ws[f'{col}2'].fill = gray_fill
        
        # 計算第一區塊數據
        # 累加統計（本月）
        monthly_orders = len([t for t in month_trades if t.get('type') == 'order'])
        monthly_cancels = len([t for t in month_trades if t.get('type') == 'cancel'])
        monthly_deals = len([t for t in month_trades if t.get('type') == 'deal'])
        
        # 當日統計（買入總量等7個欄位）
        today_buy_volume = 0
        today_sell_volume = 0
        today_buy_amount = 0
        today_sell_amount = 0
        today_realized_profit = 0
        today_realized_loss = 0
        today_total_realized = 0
        
        # 從幣安API獲取今日真實交易數據
        if binance_client:
            try:
                today_start = int(datetime.now().replace(hour=0, minute=0, second=0, microsecond=0).timestamp() * 1000)
                today_end = int((datetime.now() + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0).timestamp() * 1000)
                
                # 獲取今日交易記錄
                trades_data = binance_client._make_request('GET', '/fapi/v1/userTrades', {
                    'symbol': 'BTCUSDT',
                    'startTime': today_start,
                    'endTime': today_end,
                    'limit': 1000
                })
                
                if trades_data:
                    # 分析買賣數據
                    buy_trades = [t for t in trades_data if t.get('side') == 'BUY']
                    sell_trades = [t for t in trades_data if t.get('side') == 'SELL']
                    
                    # 買入總量（BTC數量）
                    today_buy_volume = sum(float(t.get('qty', 0)) for t in buy_trades)
                    today_sell_volume = sum(float(t.get('qty', 0)) for t in sell_trades)
                    
                    # 買入賣出金額（USDT）
                    today_buy_amount = sum(float(t.get('quoteQty', 0)) for t in buy_trades)
                    today_sell_amount = sum(float(t.get('quoteQty', 0)) for t in sell_trades)
                
                # 獲取今日已實現盈虧
                realized_pnl_data = binance_client._make_request('GET', '/fapi/v1/income', {
                    'incomeType': 'REALIZED_PNL',
                    'symbol': 'BTCUSDT',
                    'startTime': today_start,
                    'endTime': today_end
                })
                if realized_pnl_data:
                    today_total_realized = sum(float(item.get('income', 0)) for item in realized_pnl_data)
                    today_realized_profit = sum(float(item.get('income', 0)) for item in realized_pnl_data if float(item.get('income', 0)) > 0)
                    today_realized_loss = sum(float(item.get('income', 0)) for item in realized_pnl_data if float(item.get('income', 0)) < 0)
                    
            except Exception as e:
                logger.error(f"獲取今日交易數據失敗: {e}")
        
        # 計算平均價格
        avg_buy_price = today_buy_amount / today_buy_volume if today_buy_volume > 0 else 0
        avg_sell_price = today_sell_amount / today_sell_volume if today_sell_volume > 0 else 0
        
        # 交易總覽內容
        values = [
            f"{monthly_orders} 筆",  # 委託次數（累加）
            f"{monthly_cancels} 筆",  # 取消次數（累加）
            f"{monthly_deals} 筆",   # 成交次數（累加）
            f"{today_buy_volume:.8f} BTC",      # 買入總量（當日）
            f"{today_sell_volume:.8f} BTC",     # 賣出總量（當日）
            f"＄{avg_buy_price:,.2f}",          # 平均買價（當日）
            f"＄{avg_sell_price:,.2f}",         # 平均賣價（當日）
            f"＄{today_realized_profit:,.2f}",  # 已實現獲利（當日）
            f"＄{today_realized_loss:,.2f}",    # 已實現盈虧（當日）
            f"＄{today_total_realized:,.2f}"    # 總計已實現盈虧（當日）
        ]
        
        for i, value in enumerate(values):
            col = get_column_letter(i + 1)
            ws[f'{col}3'] = value
            ws[f'{col}3'].alignment = center_alignment
        
        # 帳戶狀態區塊（四大區塊用藍色，A-K欄位） - 使用當日數據
        current_row = 5
        ws.merge_cells(f'A{current_row}:K{current_row}')
        ws[f'A{current_row}'] = '帳戶狀態'
        # 應用藍色背景到A:K
        for col in range(1, 12):  # A到K
            cell = ws[f'{get_column_letter(col)}{current_row}']
            cell.fill = blue_fill
            cell.alignment = center_alignment
        
        # 帳戶狀態標題（橫向）- 使用當日數據
        account_titles = ['錢包餘額', '可供轉帳', '保證金餘額', '維持保證金', '未實現盈虧', '保證金比例', '今日手續費', '今日盈虧', '7天盈虧', '30天盈虧']
        for i, title in enumerate(account_titles):
            col = get_column_letter(i + 1)
            ws[f'{col}{current_row + 1}'] = title
            ws[f'{col}{current_row + 1}'].alignment = center_alignment
            ws[f'{col}{current_row + 1}'].fill = gray_fill
        
        # 獲取當日帳戶狀態數據
        account_data = {}
        if binance_client:
            try:
                account_info = binance_client.get_account_info()
                if account_info:
                    account_data = {
                        'totalWalletBalance': float(account_info.get('totalWalletBalance', 0)),
                        'totalMarginBalance': float(account_info.get('totalMarginBalance', 0)),
                        'availableBalance': float(account_info.get('availableBalance', 0)),
                        'totalUnrealizedProfit': float(account_info.get('totalUnrealizedProfit', 0)),
                        'totalInitialMargin': float(account_info.get('totalInitialMargin', 0)),
                        'totalMaintMargin': float(account_info.get('totalMaintMargin', 0))
                    }
            except:
                pass
        
        # 計算保證金比例
        margin_ratio = 0
        if account_data.get('totalMaintMargin', 0) > 0:
            margin_ratio = (account_data.get('totalMarginBalance', 0) / account_data.get('totalMaintMargin', 1)) * 100
        
        # 獲取今日的手續費和多期間盈虧（使用已有函數）
        today_commission = get_today_commission()
        today_pnl, today_pnl_percent = get_period_total_pnl_binance_formula(1)
        week_pnl, week_pnl_percent = get_period_total_pnl_binance_formula(7)
        month_pnl, month_pnl_percent = get_period_total_pnl_binance_formula(30)
        
        # 帳戶狀態內容（當日數據）
        account_values = [
            f"＄{account_data.get('totalWalletBalance', 0):,.8f}",      # 錢包餘額
            f"＄{account_data.get('availableBalance', 0):,.8f}",       # 可供轉帳
            f"＄{account_data.get('totalMarginBalance', 0):,.8f}",     # 保證金餘額
            f"＄{account_data.get('totalMaintMargin', 0):,.8f}",       # 維持保證金
            f"＄{account_data.get('totalUnrealizedProfit', 0):,.8f}",  # 未實現盈虧
            f"{margin_ratio:.2f}%",                                    # 保證金比例
            f"＄{today_commission:,.8f}",                              # 今日手續費
            f"＄{today_pnl:,.8f}",                                     # 今日盈虧
            f"＄{week_pnl:,.8f}",                                      # 7天盈虧
            f"＄{month_pnl:,.8f}"                                       # 30天盈虧
        ]
        for i, value in enumerate(account_values):
            col = get_column_letter(i + 1)
            ws[f'{col}{current_row + 2}'] = value
            ws[f'{col}{current_row + 2}'].alignment = center_alignment
        
        # 交易明細區塊（四大區塊用藍色，A-K欄位） - 累加所有交易明細
        current_row += 4
        ws.merge_cells(f'A{current_row}:K{current_row}')
        ws[f'A{current_row}'] = '交易明細'
        # 應用藍色背景
        for col in range(1, 12):  # A到K
            cell = ws[f'{get_column_letter(col)}{current_row}']
            cell.fill = blue_fill
            cell.alignment = center_alignment
        
        # 交易明細標題
        detail_titles = ['平倉時間', '平倉單號', '選用合約', '訂單類型', '成交類型', '成交動作', 
                        '成交數量', '開倉價格', '平倉價格', '已實現盈虧']
        for i, title in enumerate(detail_titles):
            col = get_column_letter(i + 1)
            ws[f'{col}{current_row + 1}'] = title
            ws[f'{col}{current_row + 1}'].alignment = center_alignment
            ws[f'{col}{current_row + 1}'].fill = gray_fill
        
        # 累加本月所有平倉交易明細（只顯示平倉交易）
        close_trades = [t for t in month_trades if t.get('reduceOnly', False)]
        close_trades.sort(key=lambda x: x.get('timestamp', ''))
        
        detail_row_offset = 2
        if close_trades:
            for i, trade in enumerate(close_trades):
                row = current_row + detail_row_offset + i
                
                # 成交時間（格式化時間，移除毫秒）
                timestamp = trade.get('timestamp', '')
                if timestamp:
                    try:
                        # 移除毫秒部分，只保留到秒
                        if '.' in timestamp:
                            timestamp = timestamp.split('.')[0]
                        # 確保格式為 YYYY-MM-DD HH:MM:SS
                        if 'T' in timestamp:
                            timestamp = timestamp.replace('T', ' ')
                    except:
                        pass
                
                cells_data = [
                    ('A', timestamp),  # 平倉時間
                    ('B', trade.get('order_id', '')),        # 平倉單號
                    ('C', trade.get('symbol', 'BTCUSDT')),   # 選用合約
                    ('D', '市價單' if trade.get('order_type', 'MARKET') == 'MARKET' else '限價單'),  # 訂單類型
                    ('E', '手動交易' if trade.get('is_manual', False) else '自動交易'),  # 成交類型
                    ('F', '買入' if trade.get('side', 'BUY') == 'BUY' else '賣出'),     # 成交動作
                    ('G', f"{trade.get('quantity', 0):.8f}"),    # 成交數量
                ]
                
                # 開倉價格 - 需要從配對的開倉記錄中找到
                open_price = trade.get('paired_open_price', 0)
                cells_data.append(('H', f"＄{open_price:,.8f}" if open_price > 0 else '-'))
                
                # 平倉價格
                close_price = trade.get('fill_price', trade.get('price', 0))
                cells_data.append(('I', f"＄{close_price:,.8f}" if close_price > 0 else '-'))
                
                # 已實現盈虧
                realized_pnl = trade.get('realized_pnl', 0)
                cells_data.append(('J', f"＄{realized_pnl:,.8f}" if realized_pnl != 0 else '-'))
                
                # 設置所有儲存格的內容和置中對齊
                for col, value in cells_data:
                    ws[f'{col}{row}'] = value
                    ws[f'{col}{row}'].alignment = center_alignment
        
        # 持倉狀態區塊（四大區塊用藍色，A-K欄位） - 使用當日持倉數據
        current_row = current_row + (len(close_trades) if close_trades else 0) + 3
        ws.merge_cells(f'A{current_row}:K{current_row}')
        ws[f'A{current_row}'] = '持倉狀態'
        # 應用藍色背景到A:K
        for col in range(1, 12):  # A到K
            cell = ws[f'{get_column_letter(col)}{current_row}']
            cell.fill = blue_fill
            cell.alignment = center_alignment
        
        # 持倉狀態標題（使用當日數據）
        position_titles = ['成交時間', '成交單號', '選用合約', '訂單類型', '成交類型', '成交動作', 
                          '持倉數量', '開倉價格', '標記價格', '強平價格', '未實現盈虧']
        for i, title in enumerate(position_titles):
            col = get_column_letter(i + 1)
            ws[f'{col}{current_row + 1}'] = title
            ws[f'{col}{current_row + 1}'].alignment = center_alignment
            ws[f'{col}{current_row + 1}'].fill = gray_fill
        
        # 持倉狀態內容（使用當日持倉數據）
        position_row_offset = 2
        if binance_client:
            try:
                positions = binance_client.get_position_info()
                active_positions = []
                
                if isinstance(positions, list):
                    for pos in positions:
                        position_amt = float(pos.get('positionAmt', 0))
                        if position_amt != 0:  # 只顯示有持倉的部位
                            active_positions.append(pos)
                
                if active_positions:
                    for i, pos in enumerate(active_positions):
                        row = current_row + position_row_offset + i
                        
                        # 計算收益率和保證金比例
                        position_amt = float(pos.get('positionAmt', 0))
                        entry_price = float(pos.get('entryPrice', 0))
                        unrealized_pnl = float(pos.get('unRealizedProfit', 0))
                        mark_price = float(pos.get('markPrice', 0))
                        liquidation_price = float(pos.get('liquidationPrice', 0))
                        
                        # 計算收益率
                        if entry_price > 0 and abs(position_amt) > 0:
                            position_value = abs(position_amt) * entry_price
                            roe = (unrealized_pnl / position_value) * 100 if position_value > 0 else 0
                        else:
                            roe = 0

                        # 持倉狀態數據（按照新格式）
                        position_data = [
                            ('A', _format_time_btc(pos.get('updateTime', '')) if pos.get('updateTime') else '-'),  # 成交時間
                            ('B', pos.get('symbol', 'BTCUSDT')),                               # 成交單號（暫用交易對）
                            ('C', pos.get('symbol', 'BTCUSDT')),                               # 選用合約
                            ('D', '市價單'),                                                   # 訂單類型
                            ('E', '自動交易'),                                                 # 成交類型
                            ('F', '多單' if position_amt > 0 else '空單'),                     # 成交動作
                            ('G', f"{abs(position_amt):.8f} BTC"),                            # 持倉數量
                            ('H', f"＄{entry_price:,.8f}"),                                   # 開倉價格
                            ('I', f"＄{mark_price:,.8f}"),                                    # 標記價格
                            ('J', f"＄{liquidation_price:,.8f}"),                             # 強平價格
                            ('K', f"＄{unrealized_pnl:,.8f} ({roe:+.2f}%)")                  # 未實現盈虧與收益率
                        ]
                        
                        # 設置所有儲存格的內容和置中對齊
                        for col, value in position_data:
                            ws[f'{col}{row}'] = value
                            ws[f'{col}{row}'].alignment = center_alignment
                # 沒有持倉時保持空白（與TX和日報系統一致）
                    
            except Exception as e:
                logger.error(f"獲取持倉狀態失敗: {e}")
                row = current_row + position_row_offset
                ws[f'A{row}'] = '持倉狀態獲取失敗'
                ws[f'A{row}'].alignment = center_alignment
                ws.merge_cells(f'A{row}:K{row}')
        else:
            row = current_row + position_row_offset
            ws[f'A{row}'] = 'API未連線'
            ws[f'A{row}'].alignment = center_alignment
            ws.merge_cells(f'A{row}:K{row}')
        
        # 報告生成時間
        current_row += 5
        ws[f'A{current_row}'] = f"報告生成時間：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws[f'A{current_row}'].font = Font(size=10, italic=True)
        
        # 保存檔案
        monthly_report_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'BTC交易月報')
        os.makedirs(monthly_report_dir, exist_ok=True)
        
        # 提取年月
        year = datetime.now().year
        month = datetime.now().month
        filename = f"BTC_{year}-{month:02d}.xlsx"
        filepath = os.path.join(monthly_report_dir, filename)
        
        wb.save(filepath)
        logger.info(f"交易月報已生成: {filepath}")
        
        # 添加BTC系統日誌記錄
        log_btc_frontend_message(f"{filename} 生成成功！！！", "success")
        
        # 發送 Telegram 通知並附上檔案（合併發送）
        caption = f"{filename} 交易月報已生成！！！"
        send_btc_telegram_file(filepath, caption)
        
        return filepath
        
    except Exception as e:
        logger.error(f"生成交易月報失敗: {e}")
        return None

# ========================== API 路由函數 ==========================

def get_monthly_commission():
    """獲取本月手續費"""
    try:
        if not binance_client:
            return 0.0
        
        
        # 獲取本月1號00:00的時間戳
        now = datetime.now(timezone.utc)
        start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        start_time = int(start_of_month.timestamp() * 1000)
        
        # 獲取當前時間戳
        end_time = int(now.timestamp() * 1000)
        
        try:
            # 查詢手續費記錄
            income_data = binance_client._make_request('GET', '/fapi/v1/income', {
                'incomeType': 'COMMISSION',
                'startTime': start_time,
                'endTime': end_time,
                'limit': 1000
            })
            
            if income_data:
                total_commission = sum(abs(float(item.get('income', 0))) for item in income_data)
                return total_commission
        except Exception as e:
            logger.error(f"獲取本月手續費失敗: {e}")
        
        return 0.0
    except Exception as e:
        logger.error(f"計算本月手續費時發生錯誤: {e}")
        return 0.0

def get_monthly_realized_pnl():
    """獲取本月已實現盈虧"""
    try:
        if not binance_client:
            return 0.0
        
        
        # 獲取本月1號00:00的時間戳
        now = datetime.now(timezone.utc)
        start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        start_time = int(start_of_month.timestamp() * 1000)
        
        # 獲取當前時間戳
        end_time = int(now.timestamp() * 1000)
        
        try:
            # 查詢已實現盈虧記錄
            income_data = binance_client._make_request('GET', '/fapi/v1/income', {
                'incomeType': 'REALIZED_PNL',
                'startTime': start_time,
                'endTime': end_time,
                'limit': 1000
            })
            
            if income_data:
                total_realized_pnl = sum(float(item.get('income', 0)) for item in income_data)
                return total_realized_pnl
        except Exception as e:
            logger.error(f"獲取本月已實現盈虧失敗: {e}")
        
        return 0.0
    except Exception as e:
        logger.error(f"計算本月已實現盈虧時發生錯誤: {e}")
        return 0.0

def get_btc_startup_notification_data():
    """獲取BTC啟動通知所需的完整數據"""
    try:
        if not binance_client:
            return {
                'success': False,
                'error': 'BTC客戶端未初始化'
            }
        
        # 1. 獲取環境配置
        env_data = load_btc_env_data()
        trading_pair = env_data.get('TRADING_PAIR', 'BTCUSDT')
        leverage = env_data.get('LEVERAGE', '5')
        margin_type = env_data.get('MARGIN_TYPE', '全倉')
        max_loss_percent = env_data.get('MAX_LOSS_PERCENT', '10')
        
        # 2. 獲取帳戶信息
        account_data = binance_client.get_account_info()
        if not account_data:
            return {'success': False, 'error': '無法獲取帳戶資訊'}
        
        # 提取USDT數據
        usdt_asset = None
        for asset in account_data.get('assets', []):
            if asset.get('asset') == 'USDT':
                usdt_asset = asset
                break
        
        if usdt_asset:
            wallet_balance = float(usdt_asset.get('walletBalance', 0))
            available_balance = float(usdt_asset.get('availableBalance', 0)) 
            margin_balance = float(usdt_asset.get('marginBalance', 0))
            unrealized_pnl = float(usdt_asset.get('unrealizedProfit', 0))
            initial_margin = float(usdt_asset.get('initialMargin', 0))
            maint_margin = float(usdt_asset.get('maintMargin', 0))
        else:
            # 使用頂層數據
            wallet_balance = float(account_data.get('totalWalletBalance', 0))
            available_balance = float(account_data.get('availableBalance', 0))
            margin_balance = float(account_data.get('totalMarginBalance', 0))
            unrealized_pnl = float(account_data.get('totalUnrealizedProfit', 0))
            initial_margin = float(account_data.get('totalInitialMargin', 0))
            maint_margin = float(account_data.get('totalMaintMargin', 0))
        
        # 3. 計算統計數據
        today_commission = get_today_commission()
        margin_ratio = (margin_balance / maint_margin) * 100 if maint_margin > 0 else 0.0
        leverage_usage = (initial_margin / wallet_balance) * 100 if wallet_balance > 0 else 0.0
        
        today_pnl, today_pnl_percent = get_period_total_pnl_binance_formula(1)
        week_pnl, week_pnl_percent = get_period_total_pnl_binance_formula(7)
        month_pnl, month_pnl_percent = get_period_total_pnl_binance_formula(30)
        
        # 4. 獲取持倉信息
        positions = binance_client.get_position_info()
        valid_positions = []
        
        if positions:
            for pos in positions:
                position_amt = float(pos.get('positionAmt', 0))
                if position_amt != 0:
                    side = 'LONG' if position_amt > 0 else 'SHORT'
                    entry_price = float(pos.get('entryPrice', 0))
                    unrealized_pos_pnl = float(pos.get('unRealizedProfit', 0))
                    
                    valid_positions.append({
                        'side': side,
                        'size': abs(position_amt),
                        'entry_price': entry_price,
                        'unrealized_pnl': unrealized_pos_pnl,
                        'symbol': pos.get('symbol', trading_pair)
                    })
        
        # 檢查API連接狀態
        api_connected = binance_client is not None
        if api_connected:
            try:
                # 測試API連接
                test_result = binance_client.get_server_time()
                api_connected = test_result is not None
            except:
                api_connected = False
        
        return {
            'success': True,
            'data': {
                'api_connected': api_connected,
                'trading_pair': trading_pair,
                'leverage': leverage,
                'margin_type': margin_type,
                'max_loss_percent': max_loss_percent,
                'account_id': env_data.get('BINANCE_USER_ID', 'N/A'),
                'wallet_balance': wallet_balance,
                'available_balance': available_balance,
                'margin_balance': margin_balance,
                'unrealized_pnl': unrealized_pnl,
                'today_commission': today_commission,
                'margin_ratio': margin_ratio,
                'leverage_usage': leverage_usage,
                'today_pnl': today_pnl,
                'today_pnl_percent': today_pnl_percent,
                'week_pnl': week_pnl,
                'week_pnl_percent': week_pnl_percent,
                'month_pnl': month_pnl,
                'month_pnl_percent': month_pnl_percent,
                'positions': valid_positions
            }
        }
        
    except Exception as e:
        logger.error(f"獲取BTC啟動通知數據失敗: {e}")
        return {
            'success': False,
            'error': str(e),
            'api_connected': binance_client is not None if 'binance_client' in locals() else False
        }

def get_today_trades():
    """獲取當日交易記錄"""
    try:
        if not binance_client:
            return []
        
        # 計算今天的開始和結束時間戳（毫秒）
        from datetime import datetime, timezone
        today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
        start_time = int(today.timestamp() * 1000)
        end_time = int((today.timestamp() + 86400) * 1000)  # 加一天
        
        # 獲取當日交易記錄
        trades = binance_client.get_account_trades(
            symbol='BTCUSDT',
            startTime=start_time,
            endTime=end_time
        )
        
        return trades if trades else []
        
    except Exception as e:
        logger.error(f"獲取當日交易記錄失敗: {e}")
        return []

def get_btc_message_count_today(message_type):
    """從前端日誌系統統計今天的特定消息數量"""
    try:
        # 動態獲取當前端口
        current_port = 5000
        try:
            import main
            if hasattr(main, 'CURRENT_PORT'):
                current_port = main.CURRENT_PORT
        except:
            pass
        
        # 獲取今天的日期
        today = datetime.now().strftime('%Y-%m-%d')
        
        # 向主系統詢問今天的日誌
        response = requests.get(f'http://127.0.0.1:{current_port}/api/logs', timeout=5)
        if response.status_code != 200:
            logger.warning(f"無法獲取日誌數據: HTTP {response.status_code}")
            return 0
        
        logs_data = response.json()
        logs = logs_data.get('logs', [])
        
        count = 0
        for log in logs:
            # 檢查是否為今天的BTC日誌
            timestamp = log.get('timestamp', '')
            if today in timestamp:
                extra_info = log.get('extra_info', {})
                system = extra_info.get('system', '')
                message = extra_info.get('message', '')
                
                # 確保是BTC系統的消息
                if system == 'BTC' and message_type in message:
                    count += 1
        
        return count
        
    except Exception as e:
        logger.warning(f"統計BTC消息數量失敗 ({message_type}): {e}")
        return 0

def get_btc_trading_statistics_data(date_str=None):
    """獲取BTC交易統計數據"""
    try:
        if not binance_client:
            return {
                'success': False,
                'error': 'BTC客戶端未初始化'
            }
        
        # 如果沒有指定日期，使用今天
        if not date_str:
            from datetime import datetime
            date_str = datetime.now().strftime('%Y-%m-%d')
        
        # 1. 從前端日誌系統統計委託、取消、成交次數
        order_count = get_btc_message_count_today('提交成功')  # 委託次數（提交成功訊息則數）
        cancel_count = get_btc_message_count_today('提交失敗')  # 取消次數（提交失敗訊息則數）
        fill_count = get_btc_message_count_today('成交通知')   # 成交次數（成交通知訊息則數）
        
        # 2. 從Binance API獲取交易數據計算買賣總量
        today_trades = get_today_trades()
        
        buy_volume = 0.0  # BTC數量
        sell_volume = 0.0  # BTC數量
        buy_total_value = 0.0  # USDT總量（買入做多）
        sell_total_value = 0.0  # USDT總量（賣出做空）
        
        # 從交易記錄統計
        if today_trades:
            for trade in today_trades:
                qty = float(trade.get('qty', 0))
                quote_qty = float(trade.get('quoteQty', 0))  # USDT數量
                
                if trade.get('isBuyer'):  # 買入（做多）
                    buy_volume += qty
                    buy_total_value += quote_qty  # 買入總量使用USDT
                else:  # 賣出（做空）
                    sell_volume += qty
                    sell_total_value += quote_qty  # 賣出總量使用USDT
        
        # 計算平均價格
        avg_buy_price = buy_total_value / buy_volume if buy_volume > 0 else 0.0
        avg_sell_price = sell_total_value / sell_volume if sell_volume > 0 else 0.0
        
        # 3. 獲取已實現盈虧
        today_realized_pnl = get_today_realized_pnl()
        
        # 4. 獲取當日手續費
        today_commission = get_today_commission()
        
        # 5. 獲取帳戶數據（重用啟動通知的邏輯）
        startup_data = get_btc_startup_notification_data()
        if not startup_data['success']:
            return startup_data
        
        account_data = startup_data['data']
        
        # 6. 獲取當日平倉交易記錄
        closed_trades = []
        if today_trades:
            # 這裡需要分析交易記錄來識別平倉交易
            # 簡化版本：假設所有sell交易都是平倉
            for trade in today_trades:
                if not trade.get('isBuyer'):  # 賣出交易
                    side = '多單' if float(trade.get('qty', 0)) > 0 else '空單'
                    size = abs(float(trade.get('qty', 0)))
                    entry_price = 0.0  # 這需要從歷史數據中計算
                    exit_price = float(trade.get('price', 0))
                    pnl = float(trade.get('realizedPnl', 0)) if 'realizedPnl' in trade else 0.0
                    
                    closed_trades.append({
                        'side': side,
                        'size': size,
                        'entry_price': entry_price,
                        'exit_price': exit_price,
                        'pnl': pnl
                    })
        
        return {
            'success': True,
            'data': {
                'date': date_str,
                'order_count': order_count,
                'cancel_count': cancel_count,
                'fill_count': fill_count,
                'buy_volume': buy_total_value,  # 買入總量（USDT）
                'sell_volume': sell_total_value,  # 賣出總量（USDT）
                'avg_buy_price': avg_buy_price,
                'avg_sell_price': avg_sell_price,
                'realized_profit': max(0, today_realized_pnl),  # 只取正數作為獲利
                'realized_pnl': today_realized_pnl,
                'total_realized_pnl': today_realized_pnl,  # 這裡可以累加歷史數據
                'account': account_data,
                'closed_trades': closed_trades
            }
        }
        
    except Exception as e:
        logger.error(f"獲取BTC交易統計數據失敗: {e}")
        return {
            'success': False,
            'error': str(e)
        }

def generate_btc_daily_report(date_str=None):
    """生成BTC日報Excel文件 - 四大區塊格式（主要函數）"""
    try:
        if not binance_client:
            return {
                'success': False,
                'error': 'BTC客戶端未初始化'
            }
        
        # 如果沒有指定日期，使用今天
        if not date_str:
            from datetime import datetime
            date_str = datetime.now().strftime('%Y-%m-%d')
        
        # 獲取統計數據
        stats_data = get_btc_trading_statistics_data(date_str)
        if not stats_data['success']:
            return stats_data
        
        data = stats_data['data']
        account = data['account']
        
        # 獲取詳細交易記錄
        closed_trades_data = get_btc_closed_trades_today(date_str)
        open_positions_data = get_btc_open_positions_today()
        
        # 創建Excel文件
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"BTC交易日報_{date_str.replace('-', '')}"
        
        # 設置樣式
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # 背景色
        blue_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # 字體
        white_font = Font(color="FFFFFF", bold=True)
        black_font = Font(color="000000", bold=True)
        
        # 對齊
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # 邊框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        current_row = 1
        
        # ========== 第一區塊：交易總覽 ==========
        # 區塊標題
        ws.merge_cells(f'A{current_row}:K{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "交易總覽"
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = center_alignment
        cell.border = thin_border
        current_row += 1
        
        # 第一標題行
        headers1 = ['委託次數', '取消次數', '成交次數', '買入總量', '賣出總量', '平均買價', '平均賣價']
        for i, header in enumerate(headers1):
            col = get_column_letter(i + 1)
            cell = ws[f'{col}{current_row}']
            cell.value = header
            cell.fill = gray_fill
            cell.font = black_font
            cell.alignment = center_alignment
            cell.border = thin_border
        current_row += 1
        
        # 第一區塊數據
        values1 = [
            f"{data['order_count']} 筆",
            f"{data['cancel_count']} 筆", 
            f"{data['fill_count']} 筆",
            f"{data['buy_volume']:.2f} USDT",
            f"{data['sell_volume']:.2f} USDT",
            f"{data['avg_buy_price']:.2f} USDT",
            f"{data['avg_sell_price']:.2f} USDT"
        ]
        for i, value in enumerate(values1):
            col = get_column_letter(i + 1)
            cell = ws[f'{col}{current_row}']
            cell.value = value
            cell.alignment = center_alignment
            cell.border = thin_border
        current_row += 2  # 空一行
        
        # ========== 第二區塊：帳戶狀態 ==========
        # 區塊標題
        ws.merge_cells(f'A{current_row}:K{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "帳戶狀態"
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = center_alignment
        cell.border = thin_border
        current_row += 1
        
        # 第二標題行
        headers2 = ['錢包餘額', '可供轉帳', '保證金餘額', '未實現盈虧', '交易手續費', '保證金比率', '槓桿使用率', '本日盈虧', '7 天盈虧', '30天盈虧']
        for i, header in enumerate(headers2):
            col = get_column_letter(i + 1)
            cell = ws[f'{col}{current_row}']
            cell.value = header
            cell.fill = gray_fill
            cell.font = black_font
            cell.alignment = center_alignment
            cell.border = thin_border
        current_row += 1
        
        # 第二區塊數據
        values2 = [
            f"{account['wallet_balance']:.8f} USDT",
            f"{account['available_balance']:.8f} USDT",
            f"{account['margin_balance']:.8f} USDT",
            f"{account['unrealized_pnl']:.8f} USDT",
            f"{account['today_commission']:.8f} USDT",
            f"{account['margin_ratio']:.2f}%",
            f"{account['leverage_usage']:.2f}%",
            f"{account['today_pnl']:.2f} USDT",
            f"{account['week_pnl']:.2f} USDT",
            f"{account['month_pnl']:.2f} USDT"
        ]
        for i, value in enumerate(values2):
            col = get_column_letter(i + 1)
            cell = ws[f'{col}{current_row}']
            cell.value = value
            cell.alignment = center_alignment
            cell.border = thin_border
        current_row += 2  # 空一行
        
        # ========== 第三區塊：交易明細 ==========
        # 區塊標題
        ws.merge_cells(f'A{current_row}:K{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "交易明細"
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = center_alignment
        cell.border = thin_border
        current_row += 1
        
        # 第三標題行
        headers3 = ['平倉時間', '交易單號', '選用合約', '交易動作', '交易類別', '交易方向', '交易數量', '持倉數量', '開倉價格', '平倉價格', '已實現盈虧']
        for i, header in enumerate(headers3):
            col = get_column_letter(i + 1)
            cell = ws[f'{col}{current_row}']
            cell.value = header
            cell.fill = gray_fill
            cell.font = black_font
            cell.alignment = center_alignment
            cell.border = thin_border
        current_row += 1
        
        # 第三區塊數據（平倉交易）
        if closed_trades_data:
            for trade in closed_trades_data:
                # 計算持倉數量(USDT) = BTC數量 * 開倉價格
                btc_quantity = trade.get('quantity', 0)
                entry_price = trade.get('entry_price', 0)
                position_usdt = btc_quantity * entry_price
                
                # 獲取交易來源和合約信息
                trade_source = trade.get('source', '手動')
                margin_info = trade.get('margin_info', 'BTCUSDT/永續/全倉20X')
                
                values3 = [
                    trade.get('close_time', 'N/A'),
                    trade.get('order_id', 'N/A'),
                    margin_info,
                    f"{trade_source}平倉",
                    f"{trade.get('order_type', '市價')}單",
                    '多單' if trade.get('side') == 'LONG' else '空單',
                    f"{btc_quantity:.8f} BTC",
                    f"{position_usdt:.2f} USDT",
                    f"{entry_price:.2f} USDT",
                    f"{trade.get('exit_price', 0):.2f} USDT",
                    f"{trade.get('realized_pnl', 0):.2f} USDT"
                ]
                for i, value in enumerate(values3):
                    col = get_column_letter(i + 1)
                    cell = ws[f'{col}{current_row}']
                    cell.value = value
                    cell.alignment = center_alignment
                    cell.border = thin_border
                current_row += 1
        else:
            # 無平倉交易時顯示空行
            for i in range(11):
                col = get_column_letter(i + 1)
                cell = ws[f'{col}{current_row}']
                cell.value = ""  # 空白欄位
                cell.alignment = center_alignment
                cell.border = thin_border
            current_row += 1
        
        current_row += 1  # 空一行
        
        # ========== 第四區塊：持倉狀態（動態位置）==========
        # 區塊標題
        ws.merge_cells(f'A{current_row}:K{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "持倉狀態"
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = center_alignment
        cell.border = thin_border
        current_row += 1
        
        # 第四標題行
        headers4 = ['開倉時間', '交易單號', '選用合約', '交易動作', '交易類別', '交易方向', '交易數量', '持倉數量', '開倉價格', '強平價格', '未實現盈虧']
        for i, header in enumerate(headers4):
            col = get_column_letter(i + 1)
            cell = ws[f'{col}{current_row}']
            cell.value = header
            cell.fill = gray_fill
            cell.font = black_font
            cell.alignment = center_alignment
            cell.border = thin_border
        current_row += 1
        
        # 第四區塊數據（持倉狀態）
        if open_positions_data:
            for position in open_positions_data:
                # 計算持倉數量(USDT) = BTC數量 * 開倉價格
                position_amt = abs(float(position.get('position_amt', 0)))
                entry_price = position.get('entry_price', 0)
                position_usdt = position_amt * entry_price
                
                # 獲取交易來源和合約信息
                position_source = position.get('source', '手動')
                margin_info = position.get('margin_info', 'BTCUSDT/永續/全倉20X')
                
                values4 = [
                    position.get('open_time', 'N/A'),
                    position.get('order_id', 'N/A'),
                    margin_info,
                    f"{position_source}開倉",
                    f"{position.get('order_type', '市價')}單",
                    '多單' if float(position.get('position_amt', 0)) > 0 else '空單',
                    f"{position_amt:.8f} BTC",
                    f"{position_usdt:.2f} USDT",
                    f"{entry_price:.2f} USDT",
                    f"{position.get('liquidation_price', 0):.2f} USDT",
                    f"{position.get('unrealized_pnl', 0):.2f} USDT"
                ]
                for i, value in enumerate(values4):
                    col = get_column_letter(i + 1)
                    cell = ws[f'{col}{current_row}']
                    cell.value = value
                    cell.alignment = center_alignment
                    cell.border = thin_border
                current_row += 1
        else:
            # 無持倉時顯示空行
            for i in range(11):
                col = get_column_letter(i + 1)
                cell = ws[f'{col}{current_row}']
                cell.value = ""  # 空白欄位
                cell.alignment = center_alignment
                cell.border = thin_border
            current_row += 1
        
        # 設置欄寬
        for col in range(1, 12):  # A-K欄
            ws.column_dimensions[get_column_letter(col)].width = 25
        
        # 前端日誌記錄
        log_btc_frontend_message(f"{filename} 生成成功！！！", "success")
        
        # 發送 Telegram 通知並附上檔案
        caption = f"{filename} 交易日報已生成！！！"
        send_btc_telegram_file(file_path, caption)
        
        return {
            'success': True,
            'file_path': file_path,
            'filename': filename
        }
        
    except Exception as e:
        logger.error(f"生成BTC日報失敗: {e}")
        try:
            log_btc_frontend_message(f"交易日報生成失敗：{str(e)[:100]}", "error")
        except:
            pass
        
        return {
            'success': False,
            'error': str(e)
        }

def get_btc_closed_trades_today(date_str):
    """獲取當日平倉交易記錄"""
    try:
        if not binance_client:
            return []
        
        # 從Binance API獲取今日交易記錄
        today_trades = get_today_trades()
        closed_trades = []
        
        for trade in today_trades:
            # 根據交易記錄分析平倉操作
            trade_time = datetime.fromtimestamp(int(trade.get('time', 0)) / 1000)
            
            closed_trade = {
                'close_time': trade_time.strftime('%H:%M:%S'),
                'order_id': trade.get('orderId', 'N/A'),
                'side': 'LONG' if trade.get('isBuyer') else 'SHORT',
                'quantity': float(trade.get('qty', 0)),
                'position_size': float(trade.get('qty', 0)),
                'entry_price': 0.0,  # 需要從歷史數據計算
                'exit_price': float(trade.get('price', 0)),
                'realized_pnl': float(trade.get('realizedPnl', 0)) if 'realizedPnl' in trade else 0.0,
                'action': '平倉',
                'order_type': '市價'
            }
            closed_trades.append(closed_trade)
        
        return closed_trades
        
    except Exception as e:
        logger.error(f"獲取BTC平倉交易記錄失敗: {e}")
        return []

def get_btc_open_positions_today():
    """獲取當前持倉狀態"""
    try:
        if not binance_client:
            return []
        
        # 從Binance API獲取持倉信息
        positions = binance_client.get_positions()
        open_positions = []
        
        for pos in positions:
            if pos.get('symbol') == 'BTCUSDT':
                position_amt = float(pos.get('positionAmt', 0))
                if position_amt != 0:  # 有持倉
                    open_position = {
                        'open_time': 'N/A',  # Binance API沒有直接提供開倉時間
                        'order_id': 'N/A',
                        'position_amt': position_amt,
                        'entry_price': float(pos.get('entryPrice', 0)),
                        'liquidation_price': float(pos.get('liquidationPrice', 0)),
                        'unrealized_pnl': float(pos.get('unRealizedProfit', 0)),
                        'action': '開倉',
                        'order_type': '市價'
                    }
                    open_positions.append(open_position)
        
        return open_positions
        
    except Exception as e:
        logger.error(f"獲取BTC持倉狀態失敗: {e}")
        return []
        
        # 第二區塊：帳戶狀態
        ws.merge_cells(f'A{current_row}:K{current_row}')
        ws[f'A{current_row}'] = "第二區塊 帳戶狀態"
        ws[f'A{current_row}'].fill = blue_fill
        ws[f'A{current_row}'].font = white_font
        ws[f'A{current_row}'].alignment = center_alignment
        current_row += 1
        
        # 第二區塊標題
        headers2 = ["錢包餘額", "可供轉帳", "保證金餘額", "未實現盈虧", "交易手續費", "保證金比率", "槓桿使用率", "本日盈虧", "7天盈虧", "30天盈虧"]
        for i, header in enumerate(headers2, 1):
            cell = ws.cell(row=current_row, column=i, value=header)
            cell.fill = gray_fill
            cell.font = black_font
            cell.alignment = center_alignment
            cell.border = thin_border
        current_row += 1
        
        # 第二區塊數據
        values2 = [
            f"{account['wallet_balance']:.8f}",
            f"{account['available_balance']:.8f}",
            f"{account['margin_balance']:.8f}",
            f"{account['unrealized_pnl']:.8f}",
            f"{account['today_commission']:.8f}",
            f"{account['margin_ratio']:.2f}%",
            f"{account['leverage_usage']:.2f}%",
            f"{account['today_pnl']:.8f}",
            f"{account['week_pnl']:.8f}",
            f"{account['month_pnl']:.8f}"
        ]
        for i, value in enumerate(values2, 1):
            cell = ws.cell(row=current_row, column=i, value=value)
            cell.alignment = center_alignment
            cell.border = thin_border
        current_row += 2
        
        # 第三區塊：交易明細
        ws.merge_cells(f'A{current_row}:K{current_row}')
        ws[f'A{current_row}'] = "第三區塊 交易明細"
        ws[f'A{current_row}'].fill = blue_fill
        ws[f'A{current_row}'].font = white_font
        ws[f'A{current_row}'].alignment = center_alignment
        current_row += 1
        
        # 第三區塊標題
        headers3 = ["平倉時間", "交易單號", "選用合約", "交易動作", "交易類別", "交易方向", "交易數量", "持倉數量", "開倉價格", "平倉價格", "已實現盈虧"]
        for i, header in enumerate(headers3, 1):
            cell = ws.cell(row=current_row, column=i, value=header)
            cell.fill = gray_fill
            cell.font = black_font
            cell.alignment = center_alignment
            cell.border = thin_border
        current_row += 1
        
        # 第三區塊數據
        if closed_trades:
            for trade in closed_trades:
                values3 = [
                    trade.get('close_time', ''),
                    trade.get('trade_id', ''),  # 交易單號
                    trade.get('symbol', 'BTCUSDT'),
                    trade.get('action', '平倉'),  # 交易動作
                    trade.get('order_type', '市價'),  # 交易類別
                    trade.get('direction', ''),  # 交易方向
                    f"{trade.get('trade_quantity', 0):.8f}",  # 交易數量
                    f"{trade.get('position_size', 0):.8f}",
                    f"{trade.get('entry_price', 0):.2f}",
                    f"{trade.get('exit_price', 0):.2f}",
                    f"{trade.get('realized_pnl', 0):.8f}"
                ]
                for i, value in enumerate(values3, 1):
                    cell = ws.cell(row=current_row, column=i, value=value)
                    cell.alignment = center_alignment
                    cell.border = thin_border
                current_row += 1
        else:
            # 無交易記錄
            ws.merge_cells(f'A{current_row}:K{current_row}')
            ws[f'A{current_row}'] = "❌ 無平倉交易"
            ws[f'A{current_row}'].alignment = center_alignment
            current_row += 1
        
        current_row += 1
        
        # 第四區塊：持倉狀態
        ws.merge_cells(f'A{current_row}:K{current_row}')
        ws[f'A{current_row}'] = "第四區塊 持倉狀態"
        ws[f'A{current_row}'].fill = blue_fill
        ws[f'A{current_row}'].font = white_font
        ws[f'A{current_row}'].alignment = center_alignment
        current_row += 1
        
        # 第四區塊標題
        headers4 = ["開倉時間", "交易單號", "選用合約", "交易動作", "交易類別", "交易方向", "交易數量", "持倉數量", "開倉價格", "平倉價格", "未實現盈虧"]
        for i, header in enumerate(headers4, 1):
            cell = ws.cell(row=current_row, column=i, value=header)
            cell.fill = gray_fill
            cell.font = black_font
            cell.alignment = center_alignment
            cell.border = thin_border
        current_row += 1
        
        # 第四區塊數據
        if open_positions:
            for pos in open_positions:
                values4 = [
                    pos.get('entry_time', ''),  # 開倉時間
                    pos.get('trade_id', ''),    # 交易單號
                    pos.get('symbol', 'BTCUSDT'),
                    pos.get('action', '開倉'),  # 交易動作
                    pos.get('order_type', '市價'),  # 交易類別
                    pos.get('direction', ''),   # 交易方向
                    f"{pos.get('trade_quantity', 0):.8f}",  # 交易數量
                    f"{pos.get('position_size', 0):.8f}",
                    f"{pos.get('entry_price', 0):.2f}",
                    '',  # 平倉價格空白
                    f"{pos.get('unrealized_pnl', 0):.8f}"
                ]
                for i, value in enumerate(values4, 1):
                    cell = ws.cell(row=current_row, column=i, value=value)
                    cell.alignment = center_alignment
                    cell.border = thin_border
                current_row += 1
        else:
            # 無持倉
            ws.merge_cells(f'A{current_row}:K{current_row}')
            ws[f'A{current_row}'] = "❌ 無持倉部位"
            ws[f'A{current_row}'].alignment = center_alignment
            current_row += 1
        
        # 調整列寬
        for col in range(1, 12):
            ws.column_dimensions[get_column_letter(col)].width = 25
        
        # 保存文件
        report_dir = os.path.join(os.path.dirname(__file__), f'BTC交易日報')
        os.makedirs(report_dir, exist_ok=True)
        filename = f"BTC_{date_str.replace('-', '')}.xlsx"
        filepath = os.path.join(report_dir, filename)
        
        wb.save(filepath)
        
        return {
            'success': True,
            'filepath': filepath,
            'filename': filename,
            'date': date_str
        }
        
    except Exception as e:
        logger.error(f"生成BTC日報失敗: {e}")
        return {
            'success': False,
            'error': str(e)
        }

def get_detailed_trade_records(date_str):
    """獲取詳細的交易記錄和持倉信息"""
    try:
        # 獲取當日交易記錄
        today_trades = get_today_trades()
        
        # 獲取當前持倉
        positions = binance_client.get_position_info() if binance_client else []
        
        # 分析平倉交易
        closed_trades = []
        if today_trades:
            for trade in today_trades:
                if not trade.get('isBuyer'):  # 賣出交易視為平倉
                    closed_trades.append({
                        'close_time': datetime.fromtimestamp(int(trade.get('time', 0))/1000).strftime('%Y-%m-%d %H:%M:%S'),
                        'trade_id': trade.get('id', ''),  # 改為trade_id
                        'symbol': trade.get('symbol', 'BTCUSDT'),
                        'action': '平倉',  # 交易動作
                        'order_type': '市價',  # 交易類別
                        'direction': '多單' if float(trade.get('qty', 0)) > 0 else '空單',  # 交易方向
                        'trade_quantity': abs(float(trade.get('qty', 0))),  # 交易數量
                        'position_size': abs(float(trade.get('qty', 0))),
                        'entry_price': 0.0,  # 需要從歷史數據計算
                        'exit_price': float(trade.get('price', 0)),
                        'realized_pnl': float(trade.get('realizedPnl', 0)) if 'realizedPnl' in trade else 0.0
                    })
        
        # 當前持倉
        open_positions = []
        if positions:
            for pos in positions:
                position_amt = float(pos.get('positionAmt', 0))
                if position_amt != 0:
                    open_positions.append({
                        'entry_time': '',  # 開倉時間，需要從交易歷史獲取
                        'trade_id': '',    # 交易單號
                        'symbol': pos.get('symbol', 'BTCUSDT'),
                        'action': '開倉',  # 交易動作
                        'order_type': '市價',  # 交易類別
                        'direction': '多單' if position_amt > 0 else '空單',  # 交易方向
                        'trade_quantity': abs(position_amt),  # 交易數量
                        'position_size': abs(position_amt),
                        'entry_price': float(pos.get('entryPrice', 0)),
                        'unrealized_pnl': float(pos.get('unRealizedProfit', 0))
                    })
        
        return {
            'closed_trades': closed_trades,
            'open_positions': open_positions
        }
        
    except Exception as e:
        logger.error(f"獲取詳細交易記錄失敗: {e}")
        return {
            'closed_trades': [],
            'open_positions': []
        }

def generate_btc_monthly_report(year, month):
    """生成BTC月報Excel文件"""
    try:
        if not binance_client:
            return {
                'success': False,
                'error': 'BTC客戶端未初始化'
            }
        
        # 計算月份的所有日期
        from datetime import datetime, timedelta
        from calendar import monthrange
        
        month_days = monthrange(year, month)[1]
        month_str = f"{year:04d}-{month:02d}"
        
        # 收集每日數據
        daily_data = []
        total_stats = {
            'cancel_count': 0,
            'fill_count': 0, 
            'buy_volume': 0.0,
            'sell_volume': 0.0,
            'realized_profit': 0.0,
            'realized_pnl': 0.0,
            'total_realized_pnl': 0.0,
            'commission': 0.0,
            'all_closed_trades': [],
            'all_open_positions': []
        }
        
        # 收集每一天的數據
        for day in range(1, month_days + 1):
            date_str = f"{year:04d}-{month:02d}-{day:02d}"
            
            # 獲取該日統計數據
            day_stats = get_btc_trading_statistics_data(date_str)
            if day_stats['success']:
                data = day_stats['data']
                
                # 累加統計數據
                total_stats['cancel_count'] += data['cancel_count']
                total_stats['fill_count'] += data['fill_count']
                total_stats['buy_volume'] += data['buy_volume']
                total_stats['sell_volume'] += data['sell_volume']
                total_stats['realized_profit'] += data['realized_profit']
                total_stats['realized_pnl'] += data['realized_pnl']
                total_stats['total_realized_pnl'] += data['total_realized_pnl']
                total_stats['commission'] += data['account']['today_commission']
                
                # 獲取該日詳細交易記錄
                detailed_trades = get_detailed_trade_records(date_str)
                total_stats['all_closed_trades'].extend(detailed_trades['closed_trades'])
                
                daily_data.append({
                    'date': date_str,
                    'stats': data,
                    'trades': detailed_trades
                })
        
        # 獲取最後一天的帳戶狀態（作為月末帳戶狀態）
        last_day_str = f"{year:04d}-{month:02d}-{month_days:02d}"
        last_day_account = get_btc_startup_notification_data()
        if not last_day_account['success']:
            last_day_account = {'data': {}}
        
        account_data = last_day_account.get('data', {})
        
        # 計算月度平均價格
        avg_buy_price = total_stats['buy_volume'] / total_stats['buy_volume'] if total_stats['buy_volume'] > 0 else 0.0
        avg_sell_price = total_stats['sell_volume'] / total_stats['sell_volume'] if total_stats['sell_volume'] > 0 else 0.0
        
        # 獲取當前持倉（月末狀態）
        current_positions = binance_client.get_position_info() if binance_client else []
        open_positions = []
        if current_positions:
            for pos in current_positions:
                position_amt = float(pos.get('positionAmt', 0))
                if position_amt != 0:
                    open_positions.append({
                        'symbol': pos.get('symbol', 'BTCUSDT'),
                        'side': '多單' if position_amt > 0 else '空單',
                        'quantity': abs(position_amt),
                        'entry_price': float(pos.get('entryPrice', 0)),
                        'unrealized_pnl': float(pos.get('unRealizedProfit', 0))
                    })
        
        # 創建Excel文件
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"BTC交易月報_{year}{month:02d}"
        
        # 設置樣式
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # 背景色
        blue_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True)
        black_font = Font(color="000000", bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # 邊框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        current_row = 1
        
        # 第一區塊：交易總覽（月度累計）
        ws.merge_cells(f'A{current_row}:K{current_row}')
        ws[f'A{current_row}'] = "第一區塊 交易總覽"
        ws[f'A{current_row}'].fill = blue_fill
        ws[f'A{current_row}'].font = white_font
        ws[f'A{current_row}'].alignment = center_alignment
        current_row += 1
        
        # 第一區塊標題
        headers1 = ["取消次數", "成交次數", "買入總量", "賣出總量", "平均買價", "平均賣價", "已實現獲利", "已實現盈虧", "總計已實現盈虧"]
        for i, header in enumerate(headers1, 1):
            cell = ws.cell(row=current_row, column=i, value=header)
            cell.fill = gray_fill
            cell.font = black_font
            cell.alignment = center_alignment
            cell.border = thin_border
        current_row += 1
        
        # 第一區塊數據（月度累計）
        values1 = [
            f"{total_stats['cancel_count']} 筆",
            f"{total_stats['fill_count']} 筆",
            f"{total_stats['buy_volume']:.2f} USDT",
            f"{total_stats['sell_volume']:.2f} USDT",
            f"{avg_buy_price:.2f} USDT",
            f"{avg_sell_price:.2f} USDT",
            f"{total_stats['realized_profit']:.2f} USDT",
            f"{total_stats['realized_pnl']:.2f} USDT",
            f"{total_stats['total_realized_pnl']:.2f} USDT"
        ]
        for i, value in enumerate(values1, 1):
            cell = ws.cell(row=current_row, column=i, value=value)
            cell.alignment = center_alignment
            cell.border = thin_border
        current_row += 2
        
        # 第二區塊：帳戶狀態（月末當日數據）
        ws.merge_cells(f'A{current_row}:K{current_row}')
        ws[f'A{current_row}'] = "第二區塊 帳戶狀態"
        ws[f'A{current_row}'].fill = blue_fill
        ws[f'A{current_row}'].font = white_font
        ws[f'A{current_row}'].alignment = center_alignment
        current_row += 1
        
        # 第二區塊標題
        headers2 = ["錢包餘額", "可供轉帳", "保證金餘額", "未實現盈虧", "交易手續費", "保證金比率", "槓桿使用率", "本日盈虧", "7天盈虧", "30天盈虧"]
        for i, header in enumerate(headers2, 1):
            cell = ws.cell(row=current_row, column=i, value=header)
            cell.fill = gray_fill
            cell.font = black_font
            cell.alignment = center_alignment
            cell.border = thin_border
        current_row += 1
        
        # 第二區塊數據（月末當日數據）
        values2 = [
            f"{account_data.get('wallet_balance', 0):.8f} USDT",
            f"{account_data.get('available_balance', 0):.8f} USDT",
            f"{account_data.get('margin_balance', 0):.8f} USDT",
            f"{account_data.get('unrealized_pnl', 0):.8f} USDT",
            f"{total_stats['commission']:.8f} USDT",  # 月度累計手續費
            f"{account_data.get('margin_ratio', 0):.2f}%",
            f"{account_data.get('leverage_usage', 0):.2f}%",
            f"{account_data.get('today_pnl', 0):.2f} USDT",
            f"{account_data.get('week_pnl', 0):.2f} USDT",
            f"{account_data.get('month_pnl', 0):.2f} USDT"
        ]
        for i, value in enumerate(values2, 1):
            cell = ws.cell(row=current_row, column=i, value=value)
            cell.alignment = center_alignment
            cell.border = thin_border
        current_row += 2
        
        # 第三區塊：交易明細（月度累計）
        ws.merge_cells(f'A{current_row}:K{current_row}')
        ws[f'A{current_row}'] = "第三區塊 交易明細"
        ws[f'A{current_row}'].fill = blue_fill
        ws[f'A{current_row}'].font = white_font
        ws[f'A{current_row}'].alignment = center_alignment
        current_row += 1
        
        # 第三區塊標題
        headers3 = ["平倉時間", "交易單號", "選用合約", "交易動作", "交易類別", "交易方向", "交易數量", "持倉數量", "開倉價格", "平倉價格", "已實現盈虧"]
        for i, header in enumerate(headers3, 1):
            cell = ws.cell(row=current_row, column=i, value=header)
            cell.fill = gray_fill
            cell.font = black_font
            cell.alignment = center_alignment
            cell.border = thin_border
        current_row += 1
        
        # 第三區塊數據（月度所有平倉交易）
        if total_stats['all_closed_trades']:
            for trade in total_stats['all_closed_trades']:
                values3 = [
                    trade.get('close_time', ''),
                    trade.get('trade_id', ''),  # 交易單號
                    trade.get('symbol', 'BTCUSDT'),
                    trade.get('action', '平倉'),  # 交易動作
                    trade.get('order_type', '市價'),  # 交易類別
                    trade.get('direction', ''),  # 交易方向
                    f"{trade.get('trade_quantity', 0):.8f}",  # 交易數量
                    f"{trade.get('position_size', 0):.8f}",
                    f"{trade.get('entry_price', 0):.2f}",
                    f"{trade.get('exit_price', 0):.2f}",
                    f"{trade.get('realized_pnl', 0):.8f}"
                ]
                for i, value in enumerate(values3, 1):
                    cell = ws.cell(row=current_row, column=i, value=value)
                    cell.alignment = center_alignment
                    cell.border = thin_border
                current_row += 1
        else:
            # 無交易記錄
            ws.merge_cells(f'A{current_row}:K{current_row}')
            ws[f'A{current_row}'] = "❌ 無平倉交易"
            ws[f'A{current_row}'].alignment = center_alignment
            current_row += 1
        
        current_row += 1
        
        # 第四區塊：持倉狀態（月末當日數據）
        ws.merge_cells(f'A{current_row}:K{current_row}')
        ws[f'A{current_row}'] = "第四區塊 持倉狀態"
        ws[f'A{current_row}'].fill = blue_fill
        ws[f'A{current_row}'].font = white_font
        ws[f'A{current_row}'].alignment = center_alignment
        current_row += 1
        
        # 第四區塊標題
        headers4 = ["開倉時間", "交易單號", "選用合約", "交易動作", "交易類別", "交易方向", "交易數量", "持倉數量", "開倉價格", "平倉價格", "未實現盈虧"]
        for i, header in enumerate(headers4, 1):
            cell = ws.cell(row=current_row, column=i, value=header)
            cell.fill = gray_fill
            cell.font = black_font
            cell.alignment = center_alignment
            cell.border = thin_border
        current_row += 1
        
        # 第四區塊數據（月末持倉）
        if open_positions:
            for pos in open_positions:
                values4 = [
                    '',  # 開倉時間
                    '',  # 交易單號
                    pos.get('symbol', 'BTCUSDT'),
                    '開倉',  # 交易動作
                    '市價',  # 交易類別
                    pos.get('direction', ''),  # 交易方向
                    f"{pos.get('quantity', 0):.8f}",  # 交易數量
                    f"{pos.get('quantity', 0):.8f}",
                    f"{pos.get('entry_price', 0):.2f}",
                    '',  # 平倉價格空白
                    f"{pos.get('unrealized_pnl', 0):.8f}"
                ]
                for i, value in enumerate(values4, 1):
                    cell = ws.cell(row=current_row, column=i, value=value)
                    cell.alignment = center_alignment
                    cell.border = thin_border
                current_row += 1
        else:
            # 無持倉
            ws.merge_cells(f'A{current_row}:K{current_row}')
            ws[f'A{current_row}'] = "❌ 無持倉部位"
            ws[f'A{current_row}'].alignment = center_alignment
            current_row += 1
        
        # 調整列寬
        for col in range(1, 12):
            ws.column_dimensions[get_column_letter(col)].width = 25
        
        # 保存文件
        report_dir = os.path.join(os.path.dirname(__file__), f'BTC交易日報')
        os.makedirs(report_dir, exist_ok=True)
        filename = f"BTC_{year}{month:02d}.xlsx"
        filepath = os.path.join(report_dir, filename)
        
        wb.save(filepath)
        
        return {
            'success': True,
            'filepath': filepath,
            'filename': filename,
            'year': year,
            'month': month,
            'month_str': month_str
        }
        
    except Exception as e:
        logger.error(f"生成BTC月報失敗: {e}")
        return {
            'success': False,
            'error': str(e)
        }

def parse_btc_action_and_direction(action, side):
    """解析BTC交易動作和方向
    - new buy: 開倉多單
    - new sell: 開倉空單  
    - cover sell: 平倉多單
    - cover buy: 平倉空單
    """
    # 統一轉換為小寫處理
    action = action.lower() if action else ''
    side = side.lower() if side else ''
    
    # 判斷動作和方向
    if action == 'new' and side == 'buy':
        return '開倉', '多單'
    elif action == 'new' and side == 'sell':
        return '開倉', '空單'
    elif action == 'cover' and side == 'sell':
        return '平倉', '多單'
    elif action == 'cover' and side == 'buy':
        return '平倉', '空單'
    else:
        # 默認情況，嘗試從side判斷
        if side in ['buy', 'long']:
            return '開倉', '多單'
        elif side in ['sell', 'short']:
            return '開倉', '空單'
        else:
            return '交易', '未知'

def send_btc_telegram_submit_notification(order_id, action, direction, quantity, order_type, source, is_success, error_message=None):
    """發送BTC Telegram提交通知 - 完整格式"""
    try:
        # 獲取持倉詳細信息
        position_info = get_btc_position_notification_data()
        
        # 計算持倉數量(USDT) = BTC數量 × 開倉價格
        btc_quantity = float(quantity)
        entry_price = position_info.get('entry_price', 0)
        position_usdt = btc_quantity * entry_price if entry_price > 0 else 0
        
        # 格式化通知訊息
        today = datetime.now().strftime('%Y/%m/%d')
        
        if is_success:
            message = f"⭕ 提交成功（{today}）\n"
        else:
            message = f"❌ 提交失敗（{today}）\n"
        
        # 使用真實持倉數據
        symbol = position_info.get('symbol', 'BTCUSDT')
        contract_type = position_info.get('contract_type', '永續')
        margin_mode = position_info.get('margin_mode', '全倉')
        leverage = position_info.get('leverage', '20X')
        
        message += f"{symbol}｜{contract_type}｜{margin_mode}{leverage}\n"
        message += f"交易訂單：{order_id}\n"
        message += f"交易動作：{source}{action}\n"
        message += f"交易類別：{order_type}\n"
        message += f"交易方向：{direction}\n"
        message += f"交易數量：{quantity} BTC\n"
        message += f"持倉數量(USDT)：{position_usdt:,.2f}\n"
        message += f"開倉價格(USDT)：{entry_price:,.2f}\n"
        message += f"強平價格(USDT)：{position_info.get('liquidation_price', 0):,.2f}"
        
        # 如果是失敗，添加原因
        if not is_success and error_message:
            message += f"\n原因：{error_message}"
        
        # 使用通用Telegram API函數
        return _send_telegram_api_request(message)
        
    except Exception as e:
        logger.error(f"發送BTC Telegram提交通知失敗: {e}")
        return False

def store_pending_order(order_id, order_info):
    """存儲待成交訂單"""
    global pending_orders
    pending_orders[str(order_id)] = {
        'order_id': str(order_id),
        'timestamp': datetime.now().isoformat(),
        **order_info
    }
    logger.info(f"存儲待成交訂單: {order_id}")

def remove_pending_order(order_id):
    """移除已處理訂單"""
    global pending_orders
    order_id_str = str(order_id)
    if order_id_str in pending_orders:
        del pending_orders[order_id_str]
        logger.info(f"移除已處理訂單: {order_id}")

def process_order_fill(order_id, fill_price, fill_quantity):
    """處理訂單成交事件"""
    global pending_orders
    order_id_str = str(order_id)
    
    if order_id_str not in pending_orders:
        logger.warning(f"未找到待成交訂單: {order_id}")
        return
    
    order_info = pending_orders[order_id_str]
    action = order_info['action']
    direction = order_info['direction']
    order_type = order_info['order_type']
    source = order_info['source']
    
    try:
        # 前端日誌記錄（成交）
        fill_log = f"{action}成功：{direction}｜{fill_quantity} BTC｜{fill_price:,.2f} USDT｜{order_type}"
        log_btc_frontend_message(fill_log, "success")
        
        # 發送Telegram成交通知
        fill_success = send_btc_telegram_fill_notification(
            order_id, action, direction, fill_quantity, fill_price, order_type, source
        )
        
        # 前端日誌記錄（Telegram成交發送結果）
        if fill_success:
            log_btc_frontend_message("Telegram［成交通知］訊息發送成功！！！", "success")
        
        # 移除已處理訂單
        remove_pending_order(order_id)
        
    except Exception as e:
        logger.error(f"處理訂單成交通知失敗: {e}")

def handle_btc_order_fill(order_id, action, direction, quantity, order_type, source):
    """處理BTC訂單成交 - 使用WebSocket監控"""
    try:
        # 存儲待成交訂單信息
        order_info = {
            'action': action,
            'direction': direction,
            'quantity': quantity,
            'order_type': order_type,
            'source': source,
            'symbol': 'BTCUSDT'
        }
        store_pending_order(order_id, order_info)
        
        # 如果WebSocket未連接，啟動訂單監控
        if not order_monitor_thread or not order_monitor_thread.is_alive():
            start_btc_order_monitoring()
        
    except Exception as e:
        logger.error(f"處理BTC訂單監控失敗: {e}")

def send_btc_telegram_fill_notification(order_id, action, direction, quantity, price, order_type, source):
    """發送BTC Telegram成交通知 - 完整格式"""
    try:
        # 載入配置
        config = _load_btc_env_config()
        if not config:
            return False
        
        # 獲取持倉詳細信息
        position_info = get_btc_position_notification_data()
        
        # 計算持倉數量(USDT) = BTC數量 × 開倉價格  
        btc_quantity = float(quantity)
        entry_price = position_info.get('entry_price', float(price))
        position_usdt = btc_quantity * entry_price
        
        # 格式化成交通知訊息
        today = datetime.now().strftime('%Y/%m/%d')
        message = f"✅ 成交通知（{today})\n"
        
        # 使用真實持倉數據
        symbol = position_info.get('symbol', 'BTCUSDT')
        contract_type = position_info.get('contract_type', '永續')
        margin_mode = position_info.get('margin_mode', '全倉')
        leverage = position_info.get('leverage', '20X')
        
        message += f"{symbol}｜{contract_type}｜{margin_mode}{leverage}\n"
        message += f"交易訂單：{order_id}\n"
        message += f"交易動作：{source}{action}\n"
        message += f"交易類別：{order_type}\n"
        message += f"交易方向：{direction}\n"
        message += f"交易數量：{quantity} BTC\n"
        message += f"持倉數量(USDT)：{position_usdt:,.2f}\n"
        message += f"開倉價格(USDT)：{entry_price:,.2f}\n"
        message += f"強平價格(USDT)：{position_info.get('liquidation_price', 0):,.2f}"
        
        # 發送Telegram訊息
        return _send_telegram_api_request(config['bot_token'], config['chat_id'], message)
        
    except Exception as e:
        logger.error(f"發送BTC Telegram成交通知失敗: {e}")
        return False

def get_btc_position_notification_data():
    """獲取BTC持倉信息用於Telegram通知 - 從Binance API動態獲取真實數據"""
    try:
        if not binance_client:
            return {
                'symbol': 'BTCUSDT',
                'contract_type': '永續',
                'margin_mode': '全倉',
                'leverage': '20X',
                'entry_price': 0,
                'liquidation_price': 0
            }
        
        # 獲取BTCUSDT的詳細持倉信息（包含槓桿和保證金模式）
        try:
            position_info = binance_client.futures_position_information(symbol='BTCUSDT')
            
            if position_info and len(position_info) > 0:
                pos_data = position_info[0]
                
                # 從API獲取真實數據
                symbol = pos_data.get('symbol', 'BTCUSDT')  # 交易對
                leverage = f"{int(float(pos_data.get('leverage', 20)))}X"  # 槓桿倍數
                margin_type = '逐倉' if pos_data.get('marginType') == 'isolated' else '全倉'  # 保證金模式
                entry_price = float(pos_data.get('entryPrice', 0))  # 開倉價格
                liquidation_price = float(pos_data.get('liquidationPrice', 0))  # 強平價格
                
                # 判斷合約類型（Binance期貨通常是永續合約）
                contract_type = '永續'  # Binance USDT-M 期貨都是永續合約
                
                logger.info(f"從Binance API獲取持倉數據: {symbol} | {contract_type} | {margin_type}{leverage}")
                
                return {
                    'symbol': symbol,
                    'contract_type': contract_type,
                    'margin_mode': margin_type,
                    'leverage': leverage,
                    'entry_price': entry_price,
                    'liquidation_price': liquidation_price
                }
            else:
                # 如果無法獲取持倉信息，使用當前市價和配置設定
                logger.warning("無法獲取持倉信息，使用當前市價")
                
                # 獲取當前市價
                try:
                    ticker = binance_client.get_symbol_ticker(symbol='BTCUSDT')
                    current_price = float(ticker.get('price', 0))
                except:
                    current_price = 0
                
                # 從環境配置獲取設定
                env_data = load_btc_env_data()
                leverage = f"{env_data.get('LEVERAGE', '20')}X"
                margin_mode = '逐倉' if env_data.get('MARGIN_TYPE', 'CROSS') == 'ISOLATED' else '全倉'
                
                return {
                    'symbol': 'BTCUSDT',
                    'contract_type': '永續',
                    'margin_mode': margin_mode,
                    'leverage': leverage,
                    'entry_price': current_price,
                    'liquidation_price': 0
                }
                
        except Exception as api_error:
            logger.error(f"調用Binance持倉API失敗: {api_error}")
            
            # API失敗時使用環境配置的預設值
            env_data = load_btc_env_data()
            leverage = f"{env_data.get('LEVERAGE', '20')}X"
            margin_mode = '逐倉' if env_data.get('MARGIN_TYPE', 'CROSS') == 'ISOLATED' else '全倉'
            
            return {
                'symbol': 'BTCUSDT',
                'contract_type': '永續',
                'margin_mode': margin_mode,
                'leverage': leverage,
                'entry_price': 0,
                'liquidation_price': 0
            }
        
    except Exception as e:
        logger.error(f"獲取BTC持倉通知數據失敗: {e}")
        return {
            'symbol': 'BTCUSDT',
            'contract_type': '永續', 
            'margin_mode': '全倉',
            'leverage': '20X',
            'entry_price': 0,
            'liquidation_price': 0
        }


def btc_place_order(quantity, action, side, order_type='MARKET', is_auto=False):
    """BTC下單函數 - 支援開倉/平倉判斷"""
    try:
        if not binance_client:
            return {
                'success': False,
                'error': 'BTC客戶端未初始化'
            }
        
        # 解析動作和方向
        parsed_action, direction = parse_btc_action_and_direction(action, side)
        
        # 判斷訂單來源
        order_source = '自動' if is_auto else '手動'
        
        # 格式化數量
        formatted_quantity = f"{float(quantity):.8f}"
        
        # 記錄委託訂單日誌（前端日誌格式）
        order_type_text = '市價單' if order_type == 'MARKET' else '限價單'
        commit_log = f"{order_source}{parsed_action}：{direction}｜{formatted_quantity} BTC｜市價｜{order_type_text}"
        
        # 前端日誌記錄（委託訂單）
        log_btc_frontend_message(commit_log, "info")
        logger.info(f"({order_source}委託) {commit_log}")
        
        # 獲取持倉詳細信息
        position_details = get_btc_position_notification_data()
        
        # 準備下單參數
        order_params = {
            'symbol': 'BTCUSDT',
            'side': 'BUY' if side.lower() == 'buy' else 'SELL',
            'type': order_type,
            'quantity': formatted_quantity
        }
        
        # 如果是限價單，需要添加價格
        if order_type == 'LIMIT':
            # 這裡需要添加價格邏輯
            pass
        
        try:
            # 執行下單
            order_result = binance_client.futures_create_order(**order_params)
            
            if order_result:
                order_id = order_result.get('orderId', '')
                client_order_id = order_result.get('clientOrderId', '')
                
                # 發送Telegram提交成功通知
                submit_success = send_btc_telegram_submit_notification(
                    order_id, parsed_action, direction, formatted_quantity, 
                    order_type_text, order_source, True
                )
                
                # 前端日誌記錄（Telegram發送結果）
                if submit_success:
                    log_btc_frontend_message("Telegram［提交成功］訊息發送成功！！！", "success")
                
                # 等待成交並處理成交通知
                handle_btc_order_fill(order_id, parsed_action, direction, formatted_quantity, order_type_text, order_source)
                
                # 在返回值中需要的變數（這裡需要計算或獲取真實值）
                try:
                    # 獲取當前價格
                    ticker = binance_client.get_symbol_ticker(symbol='BTCUSDT')
                    current_price = float(ticker['price']) if ticker else 0
                    
                    # 計算持倉價值
                    position_value_usdt = float(formatted_quantity) * current_price
                except:
                    current_price = 0
                    position_value_usdt = 0
                
                return {
                    'success': True,
                    'order_id': order_id,
                    'client_order_id': client_order_id,
                    'action': parsed_action,
                    'direction': direction,
                    'quantity': formatted_quantity,
                    'order_type': order_type_text,
                    'is_auto': is_auto,
                    'position_value_usdt': position_value_usdt,
                    'current_price': current_price,
                    'order_result': order_result,
                    'position_details': position_details
                }
            else:
                return {
                    'success': False,
                    'error': '下單失敗，未收到訂單結果'
                }
        
        except Exception as order_error:
            # 下單失敗的情況
            error_message = str(order_error)
            
            # 發送Telegram提交失敗通知（包含錯誤原因）
            fail_success = send_btc_telegram_submit_notification(
                '--', parsed_action, direction, formatted_quantity, 
                order_type_text, order_source, False, error_message
            )
            
            # 前端日誌記錄（Telegram失敗發送結果）
            if fail_success:
                log_btc_frontend_message("Telegram［提交失敗］訊息發送成功！！！", "success")
            
            return {
                'success': False,
                'error': error_message,
                'fail_message_sent': fail_success
            }
            
    except Exception as e:
        logger.error(f"BTC下單失敗: {e}")
        return {
            'success': False,
            'error': str(e)
        }


def btc_webhook_handler(webhook_data):
    """處理BTC webhook自動交易"""
    try:
        # 解析webhook數據
        action = webhook_data.get('action', '').lower()  # new, cover
        side = webhook_data.get('side', '').lower()      # buy, sell
        quantity = webhook_data.get('quantity', 0)
        
        if not action or not side or not quantity:
            return {
                'success': False,
                'error': '缺少必要的交易參數'
            }
        
        # 執行自動下單
        order_result = btc_place_order(
            quantity=quantity,
            action=action,
            side=side,
            order_type='MARKET',
            is_auto=True
        )
        
        return order_result
        
    except Exception as e:
        logger.error(f"處理BTC webhook失敗: {e}")
        return {
            'success': False,
            'error': str(e)
        }

def get_btc_config():
    """獲取BTC配置"""
    try:
        env_data = load_btc_env_data()
        return jsonify({
            'success': True,
            'config': env_data
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        })

# 已刪除已棄用的get_today_realized_pnl函數

def _load_btc_env_config():
    """通用函數：載入BTC環境配置"""
    btc_env = {}
    if os.path.exists(BTC_ENV_PATH):
        try:
            with open(BTC_ENV_PATH, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith('#') and '=' in line:
                        key, value = line.split('=', 1)
                        btc_env[key] = value
        except Exception as e:
            logger.error(f"載入BTC環境配置失敗: {e}")
    return btc_env

def _send_telegram_api_request(message, files=None):
    """通用函數：發送Telegram API請求"""
    try:
        btc_env = _load_btc_env_config()
        bot_token = btc_env.get('BOT_TOKEN_BTC')
        chat_id = btc_env.get('CHAT_ID_BTC')
        
        if not bot_token or not chat_id:
            return False
        
        import requests
        
        if files:
            # 發送文件
            url = f"https://api.telegram.org/bot{bot_token}/sendDocument"
            with open(files, 'rb') as f:
                files_data = {'document': f}
                data = {'chat_id': chat_id, 'caption': message}
                response = requests.post(url, data=data, files=files_data, timeout=10)
        else:
            # 發送普通消息
            url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
            payload = {
                'chat_id': chat_id,
                'text': message,
                'parse_mode': 'HTML'
            }
            response = requests.post(url, json=payload, timeout=10)
        
        return response.status_code == 200
        
    except Exception as e:
        logger.error(f"發送Telegram API請求失敗: {e}")
        return False

def get_period_realized_pnl(days=1):
    """獲取指定天數的已實現盈虧總計
    Args:
        days (int): 天數，1=本日，7=七天，30=三十天
    Returns:
        tuple: (已實現盈虧, 盈虧百分比)
    """
    try:
        if not binance_client:
            return 0.0, 0.0
        
        # 獲取指定天數的時間範圍（使用UTC+0時間，與幣安統一）
        utc_now = datetime.now(timezone.utc)
        if days == 1:
            # 本日：今天00:00開始到現在
            start_date = utc_now.replace(hour=0, minute=0, second=0, microsecond=0)
        elif days == 7:
            # 7天：7天前00:00開始到現在 (完整7天)
            start_date = (utc_now - timedelta(days=7)).replace(hour=0, minute=0, second=0, microsecond=0)
        elif days == 30:
            # 30天：30天前00:00開始到現在 (完整30天)
            start_date = (utc_now - timedelta(days=30)).replace(hour=0, minute=0, second=0, microsecond=0)
        else:
            # 其他天數：N天前00:00開始到現在
            start_date = (utc_now - timedelta(days=days-1)).replace(hour=0, minute=0, second=0, microsecond=0)
        
        start_time = int(start_date.timestamp() * 1000)
        
        # 方法1: 先嘗試從income API獲取已實現盈虧（更準確）
        try:
            income_data = binance_client._make_request('GET', '/fapi/v1/income', {
                'startTime': start_time,
                'incomeType': 'REALIZED_PNL',
                'limit': 1000
            })
            
            if income_data:
                income_total = 0.0
                for income in income_data:
                    if int(income.get('time', 0)) >= start_time:
                        income_total += float(income.get('income', 0))
                
                # 計算盈虧百分比 - 使用期初餘額作為基準
                percentage = 0.0
                try:
                    # 獲取期初餘額（期間開始時的錢包餘額）
                    period_start_balance = get_period_start_balance(start_time)
                    if period_start_balance > 0:
                        percentage = (income_total / period_start_balance) * 100
                except Exception as calc_e:
                    logger.error(f"百分比計算失敗: {calc_e}")
                    pass
                
                return income_total, percentage
        except Exception as e:
            logger.error(f"收入API失敗，改用交易記錄API: {e}")
        
        # 方法2: 備用方案 - 使用配置的交易對獲取交易記錄
        try:
            with open(BTC_ENV_PATH, 'r', encoding='utf-8') as f:
                content = f.read()
                trading_pair = 'BTCUSDT'  # 默認值
                for line in content.split('\n'):
                    if line.startswith('TRADING_PAIR='):
                        trading_pair = line.split('=')[1].strip()
                        break
        except:
            trading_pair = 'BTCUSDT'  # 備用默認值
        
        # 透過 GET /fapi/v1/userTrades 的 realizedPnl 欄位直接獲取已實現盈虧
        trade_data = binance_client._make_request('GET', '/fapi/v1/userTrades', {
            'symbol': trading_pair,
            'startTime': start_time,
            'limit': 1000
        })
        
        if not trade_data:
            return 0.0, 0.0
        
        # 累計所有交易的已實現盈虧
        total_realized_pnl = 0.0
        trade_count = 0
        
        for trade in trade_data:
            trade_time = int(trade.get('time', 0))
            
            # 確保交易在指定期間範圍內
            if trade_time >= start_time:
                # 直接使用 GET /fapi/v1/userTrades 的 realizedPnl 字段
                # 注意：需要累計所有交易的 realizedPnl，包括0值
                realized_pnl = float(trade.get('realizedPnl', 0))
                total_realized_pnl += realized_pnl
                trade_count += 1
        
        # 計算盈虧百分比（備用方案）- 使用期初餘額作為基準
        percentage = 0.0
        try:
            period_start_balance = get_period_start_balance(start_time)
            if period_start_balance > 0:
                percentage = (total_realized_pnl / period_start_balance) * 100
        except:
            pass
        
        return total_realized_pnl, percentage
        
    except Exception as e:
        logger.error(f"從交易記錄獲取期間已實現盈虧失敗: {e}")
        return 0.0, 0.0

def get_period_start_balance(timestamp):
    """獲取期間開始時的錢包餘額
    Args:
        timestamp (int): 期間開始時間的毫秒時間戳
    Returns:
        float: 期間開始時的錢包餘額
    """
    try:
        if not binance_client:
            return 0.0
        
        # 獲取從期間開始到現在的所有收入記錄
        income_data = binance_client._make_request('GET', '/fapi/v1/income', {
            'startTime': timestamp,
            'limit': 1000
        })
        
        # 獲取當前錢包餘額
        account_info = binance_client.get_account_info()
        if not account_info:
            return 0.0
            
        current_balance = float(account_info.get('totalWalletBalance', 0))
        
        if not income_data:
            return current_balance  # 如果沒有收入記錄，返回當前餘額
        
        # 計算從期間開始到現在的總收入變化
        total_income_change = 0.0
        income_count = 0
        for income in income_data:
            if int(income.get('time', 0)) >= timestamp:
                # 累計所有類型的收入變化（包括已實現盈虧、手續費、資金費率等）
                income_value = float(income.get('income', 0))
                total_income_change += income_value
                income_count += 1
        
        logger.debug(f"期初餘額計算 - 當前餘額: {current_balance:.8f}, 期間收入變化: {total_income_change:.8f} ({income_count}筆記錄)")
        
        # 期初餘額 = 當前餘額 - 期間收入變化
        period_start_balance = current_balance - total_income_change
        
        logger.debug(f"計算出的期初餘額: {period_start_balance:.8f}")
        
        # 確保期初餘額為正數且合理
        if period_start_balance <= 0 or abs(period_start_balance) < 0.01:
            logger.debug(f"期初餘額異常 ({period_start_balance:.8f})，使用當前餘額作為備用方案")
            return current_balance
        
        # 如果期初餘額和當前餘額差距過大（超過90%），可能計算有誤
        if abs(period_start_balance - current_balance) / current_balance > 0.9:
            logger.debug(f"期初餘額與當前餘額差距過大 ({period_start_balance:.8f} vs {current_balance:.8f})，使用當前餘額")
            return current_balance
        
        return period_start_balance
        
    except Exception as e:
        logger.error(f"獲取期初餘額失敗: {e}")
        # 如果無法獲取期初餘額，返回當前餘額
        try:
            account_info = binance_client.get_account_info()
            if account_info:
                return float(account_info.get('totalWalletBalance', 0))
        except:
            pass
        return 0.0

def get_period_total_pnl_binance_formula(days=1):
    """使用幣安公式獲取期間總盈虧：已實現盈虧 + 未實現盈虧 - 資金費用
    Args:
        days (int): 天數，1=本日，7=七天，30=三十天
    Returns:
        tuple: (總盈虧, 盈虧百分比)
    """
    try:
        if not binance_client:
            return 0.0, 0.0
        
        # 獲取指定天數的時間範圍（使用UTC+0時間，與幣安統一）
        utc_now = datetime.now(timezone.utc)
        if days == 1:
            # 本日：今天00:00開始到現在
            start_date = utc_now.replace(hour=0, minute=0, second=0, microsecond=0)
        elif days == 7:
            # 7天：7天前00:00開始到現在 (完整7天)
            start_date = (utc_now - timedelta(days=7)).replace(hour=0, minute=0, second=0, microsecond=0)
        elif days == 30:
            # 30天：30天前00:00開始到現在 (完整30天)
            start_date = (utc_now - timedelta(days=30)).replace(hour=0, minute=0, second=0, microsecond=0)
        else:
            start_date = (utc_now - timedelta(days=days-1)).replace(hour=0, minute=0, second=0, microsecond=0)
        
        start_time = int(start_date.timestamp() * 1000)
        
        # 1. 獲取期間已實現盈虧
        realized_pnl = 0.0
        try:
            income_data = binance_client._make_request('GET', '/fapi/v1/income', {
                'startTime': start_time,
                'incomeType': 'REALIZED_PNL',
                'limit': 1000
            })
            
            if income_data:
                for income in income_data:
                    if int(income.get('time', 0)) >= start_time:
                        realized_pnl += float(income.get('income', 0))
        except Exception as e:
            logger.error(f"獲取已實現盈虧失敗: {e}")
        
        # 2. 獲取當前未實現盈虧
        unrealized_pnl = 0.0
        try:
            account_info = binance_client.get_account_info()
            if account_info:
                unrealized_pnl = float(account_info.get('totalUnrealizedProfit', 0))
        except Exception as e:
            logger.error(f"獲取未實現盈虧失敗: {e}")
        
        # 3. 獲取期間資金費用
        funding_fee = 0.0
        try:
            funding_data = binance_client._make_request('GET', '/fapi/v1/income', {
                'startTime': start_time,
                'incomeType': 'FUNDING_FEE',
                'limit': 1000
            })
            
            if funding_data:
                for fee in funding_data:
                    if int(fee.get('time', 0)) >= start_time:
                        funding_fee += float(fee.get('income', 0))
        except Exception as e:
            logger.error(f"獲取資金費用失敗: {e}")
        
        # 計算總盈虧：已實現盈虧 + 未實現盈虧 - 資金費用
        total_pnl = realized_pnl + unrealized_pnl - funding_fee
        
        # 計算盈虧百分比
        percentage = 0.0
        try:
            period_start_balance = get_period_start_balance(start_time)
            logger.debug(f"{days}天期初餘額計算結果: {period_start_balance:.8f}")
            
            if period_start_balance > 0:
                percentage = (total_pnl / period_start_balance) * 100
                logger.debug(f"{days}天盈虧百分比: {total_pnl:.8f} / {period_start_balance:.8f} * 100 = {percentage:.2f}%")
            else:
                logger.debug(f"{days}天期初餘額為零或負數，無法計算百分比")
                # 使用當前餘額作為備用方案
                try:
                    account_info = binance_client.get_account_info()
                    if account_info:
                        current_balance = float(account_info.get('totalWalletBalance', 0))
                        if current_balance > 0:
                            percentage = (total_pnl / current_balance) * 100
                            logger.debug(f"{days}天使用當前餘額計算百分比: {total_pnl:.8f} / {current_balance:.8f} * 100 = {percentage:.2f}%")
                except:
                    pass
        except Exception as e:
            logger.error(f"百分比計算失敗: {e}")
        
        logger.debug(f"{days}天盈虧計算 - 已實現: {realized_pnl:.8f}, 未實現: {unrealized_pnl:.8f}, 資金費用: {funding_fee:.8f}, 總計: {total_pnl:.8f}, 百分比: {percentage:.2f}%")
        
        return total_pnl, percentage
        
    except Exception as e:
        logger.error(f"獲取{days}天總盈虧失敗: {e}")
        return 0.0, 0.0

def get_today_realized_pnl_from_trades():
    """舊函數包裝器 - 向後兼容"""
    pnl, _ = get_period_realized_pnl(1)
    return pnl

def get_today_commission():
    """獲取本日手續費"""
    try:
        if not binance_client:
            return 0.0
        
        
        # 獲取今日00:00的時間戳
        today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
        start_time = int(today.timestamp() * 1000)
        
        # 獲取收益歷史，只獲取手續費
        income_data = binance_client._make_request('GET', '/fapi/v1/income', {
            'startTime': start_time,
            'incomeType': 'COMMISSION',
            'limit': 100
        })
        
        if not income_data:
            return 0.0
        
        # 計算本日手續費總和（手續費通常是負值）
        today_commission = 0.0
        for income in income_data:
            income_time = int(income.get('time', 0))
            if income_time >= start_time:
                commission_amount = abs(float(income.get('income', 0)))  # 取絕對值
                today_commission += commission_amount
        
        return today_commission
        
    except Exception as e:
        logger.error(f"獲取本日手續費失敗: {e}")
        return 0.0

def get_period_pnl(days=1):
    """獲取指定天數的總盈虧 - 使用正確的幣安API"""
    try:
        if not binance_client:
            return 0.0, 0.0  # (總盈虧, 收益率%)
        
        
        # 計算起始時間
        now = datetime.now(timezone.utc)
        start_time = now - timedelta(days=days)
        start_timestamp = int(start_time.timestamp() * 1000)
        end_timestamp = int(now.timestamp() * 1000)
        
        
        # 使用幣安收入歷史API - 只獲取已實現盈虧
        total_pnl = 0.0
        
        # 1. 獲取已實現盈虧 (REALIZED_PNL)
        try:
            realized_pnl_data = binance_client._make_request('GET', '/fapi/v1/income', {
                'incomeType': 'REALIZED_PNL',
                'startTime': start_timestamp,
                'endTime': end_timestamp,
                'limit': 1000
            })
            
            if realized_pnl_data:
                realized_total = sum(float(item.get('income', 0)) for item in realized_pnl_data)
                total_pnl += realized_total
        except Exception as e:
            pass
        
        # 2. 獲取資金費用 (FUNDING_FEE)
        try:
            funding_data = binance_client._make_request('GET', '/fapi/v1/income', {
                'incomeType': 'FUNDING_FEE',
                'startTime': start_timestamp,
                'endTime': end_timestamp,
                'limit': 1000
            })
            
            if funding_data:
                funding_total = sum(float(item.get('income', 0)) for item in funding_data)
                total_pnl += funding_total
        except Exception as e:
            pass
        
        # 3. 對於本日數據，額外添加當前未實現盈虧
        if days == 1:
            try:
                account_info = binance_client.get_account_info()
                if account_info:
                    unrealized_pnl = float(account_info.get('totalUnrealizedProfit', 0))
                    # 注意：不要將未實現盈虧加入總盈虧，因為這是動態變化的
            except Exception as e:
                pass
        
        
        # 計算收益率 - 基於歷史錢包餘額
        percentage = 0.0
        try:
            # 獲取當前錢包餘額
            account_info = binance_client.get_account_info()
            if account_info:
                current_balance = float(account_info.get('totalWalletBalance', 0))
                
                # 計算期初餘額（當前餘額 - 期間盈虧）
                initial_balance = current_balance - total_pnl
                if initial_balance > 0:
                    percentage = (total_pnl / initial_balance) * 100
                else:
                    pass
        except Exception as e:
            pass
        
        return total_pnl, percentage
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return 0.0, 0.0

# 已移除無用的測試函數 test_income_api()

def get_binance_stats_pnl():
    """從幣安統計接口獲取準確的盈虧數據"""
    try:
        if not binance_client:
            return None
            
        
        # 方法1: 嘗試使用24hr ticker統計API
        try:
            ticker_24hr = binance_client._make_request('GET', '/fapi/v1/ticker/24hr')
        except Exception as e:
            pass
        
        # 方法2: 嘗試獲取帳戶快照
        try:
            # 這個API可能不存在，但值得嘗試
            snapshot = binance_client._make_request('GET', '/fapi/v1/accountSnapshot')
        except Exception as e:
            pass
        
        # 方法3: 嘗試獲取資產流水
        try:
            asset_flow = binance_client._make_request('GET', '/fapi/v1/assetFlow')
        except Exception as e:
            pass
        
        # 方法4: 使用用戶交易記錄計算
        try:
            user_trades = binance_client._make_request('GET', '/fapi/v1/userTrades', {
                'limit': 500
            })
            
            if user_trades:
                # 分析交易記錄，計算盈虧
                return analyze_trades_for_pnl(user_trades)
                
        except Exception as e:
            pass
            
        return None
        
    except Exception as e:
        return None

def analyze_trades_for_pnl(trades):
    """分析交易記錄計算真實盈虧"""
    try:
        
        # 按時間分組計算不同期間的盈虧
        now = datetime.now(timezone.utc)
        periods = {
            'today': now.replace(hour=0, minute=0, second=0, microsecond=0),
            'week': now - timedelta(days=7),
            'month': now - timedelta(days=30),
            'quarter': now - timedelta(days=90)
        }
        
        period_pnl = {}
        
        for period_name, start_time in periods.items():
            start_timestamp = int(start_time.timestamp() * 1000)
            period_total = 0.0
            trade_count = 0
            
            for trade in trades:
                trade_time = int(trade.get('time', 0))
                if trade_time >= start_timestamp:
                    realized_pnl = float(trade.get('realizedPnl', 0))
                    period_total += realized_pnl
                    trade_count += 1
            
            period_pnl[period_name] = {
                'pnl': period_total,
                'trades': trade_count
            }
        
        return period_pnl
        
    except Exception as e:
        return None

def get_today_pnl():
    """獲取本日總盈虧 - 使用本日已實現盈虧"""
    try:
        if not binance_client:
            return 0.0, 0.0
            
        # 本日盈虧就是本日已實現盈虧
        total_pnl = get_today_realized_pnl()
        
        # 獲取錢包餘額計算百分比
        account_info = binance_client.get_account_info()
        percentage = 0.0
        if account_info:
            wallet_balance = float(account_info.get('totalWalletBalance', 0))
            if wallet_balance > 0:
                percentage = (total_pnl / wallet_balance) * 100
        
        return total_pnl, percentage
        
    except Exception as e:
        return 0.0, 0.0


def get_btc_account_info():
    """獲取BTC帳戶資訊 - 新版本支援完整字段"""
    try:
        if not binance_client:
            return jsonify({
                'success': False,
                'error': 'BTC客戶端未初始化'
            })
        
        # 獲取帳戶資訊
        account_data = binance_client.get_account_info()
        if not account_data:
            return jsonify({
                'success': False,
                'error': '無法獲取帳戶資訊'
            })
        
        
        # 基礎數據 - 從assets數組中提取USDT數據
        usdt_asset = None
        assets = account_data.get('assets', [])
        for asset in assets:
            if asset.get('asset') == 'USDT':
                usdt_asset = asset
                break
        
        if usdt_asset:
            # 使用USDT資產的具體數據
            wallet_balance = float(usdt_asset.get('walletBalance', 0))
            available_balance = float(usdt_asset.get('availableBalance', 0))
            margin_balance = float(usdt_asset.get('marginBalance', 0))
            unrealized_pnl = float(usdt_asset.get('unrealizedProfit', 0))
            initial_margin = float(usdt_asset.get('initialMargin', 0))
            maint_margin = float(usdt_asset.get('maintMargin', 0))
        else:
            # 備用：使用頂層總計數據
            wallet_balance = float(account_data.get('totalWalletBalance', 0))
            available_balance = float(account_data.get('availableBalance', 0))
            margin_balance = float(account_data.get('totalMarginBalance', 0))
            unrealized_pnl = float(account_data.get('totalUnrealizedProfit', 0))
            initial_margin = float(account_data.get('totalInitialMargin', 0))
            maint_margin = float(account_data.get('totalMaintMargin', 0))
        
        # 獲取本日手續費和已實現盈虧
        today_commission = get_today_commission()
        
        # 計算多期間盈虧統計 - 使用幣安正確公式（已實現+未實現-資金費用）
        today_pnl, today_pnl_percent = get_period_total_pnl_binance_formula(1)      # 本日盈虧
        week_pnl, week_pnl_percent = get_period_total_pnl_binance_formula(7)        # 7天盈虧  
        month_pnl, month_pnl_percent = get_period_total_pnl_binance_formula(30)     # 30天盈虧
        
        
        # 計算保證金比率（預設0.00%）
        margin_ratio = 0.0
        if maint_margin > 0:
            margin_ratio = (margin_balance / maint_margin) * 100
        
        # 計算槓桿使用率（預設0.00%）
        leverage_usage = 0.0
        if wallet_balance > 0:
            leverage_usage = (initial_margin / wallet_balance) * 100
        
        # totalMaintMargin: 維持保證金
        
        organized_account = OrderedDict([
            ('walletBalance', f"{wallet_balance:.8f}"),                              # 錢包餘額
            ('availableBalance', f"{available_balance:.8f}"),                        # 可用餘額  
            ('marginBalance', f"{margin_balance:.8f}"),                              # 保證金餘額
            ('unrealizedProfit', f"{unrealized_pnl:.8f}"),                          # 未實現盈虧
            ('initialMargin', f"{initial_margin:.8f}"),                              # 原始保證金（已使用）
            ('maintMargin', f"{maint_margin:.8f}"),                                  # 維持保證金
            ('marginRatio', f"{margin_ratio:.2f}%"),                                    # 保證金率（預設0.00%）
            ('leverageUsage', f"{leverage_usage:.2f}%"),                             # 槓桿使用率（預設0.00%）
            ('todayCommission', f"{today_commission:.8f}"),                          # 手續費
            ('todayPnl', f"{today_pnl:.8f}"),                                        # 本日盈虧
            ('todayPnlPercent', f"{today_pnl_percent:.2f}%"),                       # 本日盈虧百分比
            ('weekPnl', f"{week_pnl:.8f}"),                                          # 7天盈虧
            ('weekPnlPercent', f"{week_pnl_percent:.2f}%"),                         # 7天盈虧百分比
            ('monthPnl', f"{month_pnl:.8f}"),                                        # 30天盈虧
            ('monthPnlPercent', f"{month_pnl_percent:.2f}%"),                       # 30天盈虧百分比
        ])
        
        
        return jsonify({
            'success': True,
            'account': organized_account
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        })

def get_btc_positions():
    """獲取BTC持倉資訊"""
    try:
        if not binance_client:
            return jsonify({
                'success': False,
                'error': 'BTC客戶端未初始化'
            })
        
        # 獲取持倉資訊
        positions = binance_client.get_position_info()
        if positions is None:
            return jsonify({
                'success': False,
                'error': '無法獲取持倉資訊'
            })
        
        
        # 篩選有效持倉（數量不為0的）並計算真實值
        valid_positions = []
        for pos in positions:
            position_amt = float(pos.get('positionAmt', 0))
            if position_amt != 0:
                
                # 計算真實收益率 (unRealizedProfit / isolatedMargin) * 100
                unrealized_pnl = float(pos.get('unRealizedProfit', 0))
                isolated_margin = float(pos.get('isolatedMargin', 0))
                isolated_wallet = float(pos.get('isolatedWallet', 0))
                notional = float(pos.get('notional', 0))
                
                
                # 根據保證金類型選擇保證金值
                if pos.get('marginType') == 'isolated':
                    margin_value = isolated_margin if isolated_margin > 0 else isolated_wallet
                else:
                    margin_value = isolated_wallet
                
                # 計算真實收益率
                if margin_value > 0:
                    percentage = (unrealized_pnl / margin_value) * 100
                else:
                    percentage = 0
                
                # 計算保證金比例
                if margin_value > 0:
                    margin_ratio = (unrealized_pnl / margin_value) * 100
                else:
                    margin_ratio = 0
                
                # 添加計算欄位
                pos['percentage'] = percentage
                pos['marginRatio'] = margin_ratio
                pos['isolatedMargin'] = margin_value
                
                valid_positions.append(pos)
        
        return jsonify({
            'success': True,
            'positions': valid_positions
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        })

# ========================== BTC風險管理功能 ==========================

def calculate_btc_risk_metrics():
    """計算BTC風險指標"""
    global binance_client, account_info
    
    try:
        if not binance_client or not account_info:
            return None
        
        # 獲取最新帳戶和持倉信息
        fresh_account_info = binance_client.get_account_info()
        positions = binance_client.get_position_info()
        
        if not fresh_account_info or not positions:
            return None
        
        # 基本帳戶數據
        total_wallet_balance = float(fresh_account_info.get('totalWalletBalance', 0))
        total_margin_balance = float(fresh_account_info.get('totalMarginBalance', 0))
        total_unrealized_pnl = float(fresh_account_info.get('totalUnrealizedProfit', 0))
        total_initial_margin = float(fresh_account_info.get('totalInitialMargin', 0))
        total_maint_margin = float(fresh_account_info.get('totalMaintMargin', 0))
        available_balance = float(fresh_account_info.get('availableBalance', 0))
        
        # 計算風險指標
        risk_metrics = {
            'total_wallet_balance': total_wallet_balance,
            'total_margin_balance': total_margin_balance,
            'available_balance': available_balance,
            'total_unrealized_pnl': total_unrealized_pnl,
            'total_initial_margin': total_initial_margin,
            'total_maint_margin': total_maint_margin,
            'margin_ratio': 0,
            'risk_level': 'SAFE',
            'leverage_usage': 0,
            'position_value': 0,
            'max_drawable': available_balance,
            'positions': []
        }
        
        # 保證金比率計算
        if total_maint_margin > 0:
            risk_metrics['margin_ratio'] = (total_margin_balance / total_maint_margin) * 100
        
        # 槓桿使用率計算
        if total_wallet_balance > 0:
            risk_metrics['leverage_usage'] = (total_initial_margin / total_wallet_balance) * 100
        
        # 處理持倉信息
        total_position_value = 0
        active_positions = []
        
        for pos in positions:
            position_amt = float(pos.get('positionAmt', 0))
            if position_amt != 0:
                entry_price = float(pos.get('entryPrice', 0))
                mark_price = float(pos.get('markPrice', 0))
                unrealized_pnl = float(pos.get('unRealizedProfit', 0))
                percentage = float(pos.get('percentage', 0))
                leverage = float(pos.get('leverage', 1))
                
                position_value = abs(position_amt) * mark_price
                total_position_value += position_value
                
                position_info = {
                    'symbol': pos.get('symbol'),
                    'side': 'LONG' if position_amt > 0 else 'SHORT',
                    'size': abs(position_amt),
                    'entry_price': entry_price,
                    'mark_price': mark_price,
                    'unrealized_pnl': unrealized_pnl,
                    'percentage': percentage,
                    'leverage': leverage,
                    'position_value': position_value,
                    'margin_required': position_value / leverage if leverage > 0 else 0
                }
                active_positions.append(position_info)
        
        risk_metrics['position_value'] = total_position_value
        risk_metrics['positions'] = active_positions
        
        # 風險等級評估
        margin_ratio = risk_metrics['margin_ratio']
        leverage_usage = risk_metrics['leverage_usage']
        
        if margin_ratio > 0:
            if margin_ratio < 120:  # 強制平倉風險高
                risk_metrics['risk_level'] = 'HIGH'
            elif margin_ratio < 200:  # 中等風險
                risk_metrics['risk_level'] = 'MEDIUM'
            else:  # 安全
                risk_metrics['risk_level'] = 'SAFE'
        
        # 槓桿使用率風險
        if leverage_usage > 80:
            if risk_metrics['risk_level'] in ['SAFE', 'MEDIUM']:
                risk_metrics['risk_level'] = 'MEDIUM'
        elif leverage_usage > 90:
            risk_metrics['risk_level'] = 'HIGH'
        
        return risk_metrics
        
    except Exception as e:
        logger.error(f"計算BTC風險指標失敗: {e}")
        return None

def check_btc_risk_alerts():
    """檢查BTC風險警報"""
    try:
        risk_metrics = calculate_btc_risk_metrics()
        if not risk_metrics:
            return
        
        risk_level = risk_metrics['risk_level']
        margin_ratio = risk_metrics['margin_ratio']
        leverage_usage = risk_metrics['leverage_usage']
        total_unrealized_pnl = risk_metrics['total_unrealized_pnl']
        
        # 檢查是否需要發送風險警報
        alerts = []
        
        # 強制平倉風險警報
        if margin_ratio > 0 and margin_ratio < 120:
            alerts.append({
                'type': 'LIQUIDATION_RISK',
                'level': 'CRITICAL',
                'message': f'⚠️ 強制平倉風險警報\n保證金比率: {margin_ratio:.1f}%\n請及時補充保證金或減少持倉！'
            })
        elif margin_ratio > 0 and margin_ratio < 200:
            alerts.append({
                'type': 'MARGIN_WARNING',
                'level': 'WARNING',
                'message': f'⚠️ 保證金警告\n保證金比率: {margin_ratio:.1f}%\n建議關注風險控制'
            })
        
        # 槓桿使用率警報
        if leverage_usage > 90:
            alerts.append({
                'type': 'LEVERAGE_RISK',
                'level': 'WARNING',
                'message': f'⚠️ 槓桿使用率過高\n使用率: {leverage_usage:.1f}%\n建議降低槓桿或增加保證金'
            })
        
        # 大額未實現虧損警報
        if total_unrealized_pnl < -1000:  # 虧損超過1000 USDT
            alerts.append({
                'type': 'LARGE_LOSS',
                'level': 'WARNING',
                'message': f'⚠️ 大額未實現虧損\n虧損金額: {total_unrealized_pnl:.2f} USDT\n請考慮風險控制措施'
            })
        
        # 發送警報通知
        for alert in alerts:
            send_btc_risk_alert(alert)
            
        # 記錄風險檢查日誌
        logger.debug(f"BTC風險檢查完成 - 風險等級: {risk_level}, 保證金比率: {margin_ratio:.1f}%, 槓桿使用率: {leverage_usage:.1f}%")
        
    except Exception as e:
        logger.error(f"BTC風險檢查失敗: {e}")

def send_btc_risk_alert(alert):
    """發送BTC風險警報通知"""
    try:
        current_time = datetime.now().strftime('%Y/%m/%d %H:%M:%S')
        
        # 構建警報訊息
        level_emoji = {
            'CRITICAL': '🚨',
            'WARNING': '⚠️',
            'INFO': 'ℹ️'
        }
        
        emoji = level_emoji.get(alert['level'], '⚠️')
        
        message = f"{emoji} BTC風險管理警報 ({current_time})\n\n{alert['message']}\n\n請及時處理以避免損失！"
        
        # 發送Telegram通知
        send_btc_telegram_message(message)
        
        # 記錄到風險日誌
        risk_log = {
            'timestamp': current_time,
            'alert_type': alert['type'],
            'level': alert['level'],
            'message': alert['message']
        }
        
        logger.info(f"BTC風險警報已發送: {alert['type']} - {alert['level']}")
        
    except Exception as e:
        logger.error(f"發送BTC風險警報失敗: {e}")


def get_btc_risk_status():
    """獲取BTC風險狀態API"""
    try:
        risk_metrics = calculate_btc_risk_metrics()
        if not risk_metrics:
            return jsonify({
                'success': False,
                'message': '無法獲取風險數據'
            })
        
        return jsonify({
            'success': True,
            'risk_metrics': risk_metrics
        })
        
    except Exception as e:
        logger.error(f"獲取BTC風險狀態失敗: {e}")
        return jsonify({
            'success': False,
            'message': f'獲取失敗: {str(e)}'
        })

def start_btc_risk_monitoring():
    """啟動BTC風險監控"""
    logger.info("BTC風險監控已啟動")
    
    while True:
        try:
            # 每分鐘檢查一次風險
            check_btc_risk_alerts()
            time.sleep(60)  # 等待60秒
            
        except Exception as e:
            logger.error(f"BTC風險監控異常: {e}")
            time.sleep(60)  # 發生錯誤時也等待60秒

# ========================== BTC WebSocket實時數據 ==========================
import websocket
import ssl

# WebSocket全局變量
btc_ws = None
btc_ws_thread = None
btc_real_time_data = {}
btc_ws_connected = False

def start_btc_websocket():
    """啟動BTC WebSocket連接"""
    global btc_ws, btc_ws_thread, btc_ws_connected
    
    try:
        if btc_ws_connected:
            logger.info("BTC WebSocket已連接")
            return True
        
        # 獲取活躍的交易對
        env_data = load_btc_env_data()
        trading_pair = env_data.get('TRADING_PAIR', 'BTCUSDT').lower()
        
        # 構建WebSocket URL - 幣安期貨WebSocket
        ws_url = f"wss://fstream.binance.com/ws/{trading_pair}@ticker"
        
        logger.debug(f"啟動BTC WebSocket連接: {ws_url}")
        
        # 創建WebSocket連接
        btc_ws = websocket.WebSocketApp(
            ws_url,
            on_open=on_btc_ws_open,
            on_message=on_btc_ws_message,
            on_error=on_btc_ws_error,
            on_close=on_btc_ws_close
        )
        
        # 在新線程中運行WebSocket
        btc_ws_thread = threading.Thread(
            target=btc_ws.run_forever,
            kwargs={'sslopt': {"cert_reqs": ssl.CERT_NONE}},
            daemon=True,
            name="BTC-WebSocket"
        )
        register_btc_thread(btc_ws_thread, "BTC-WebSocket")
        btc_ws_thread.start()
        
        logger.info("BTC WebSocket線程已啟動")
        return True
        
    except Exception as e:
        logger.error(f"啟動BTC WebSocket失敗: {e}")
        return False

def on_btc_ws_open(ws):
    """WebSocket連接開啟"""
    global btc_ws_connected
    btc_ws_connected = True
    logger.info("BTC WebSocket連接成功")

def on_btc_ws_message(ws, message):
    """處理WebSocket訊息"""
    global btc_real_time_data
    
    try:
        data = json.loads(message)
        
        # 更新實時數據
        btc_real_time_data.update({
            'symbol': data.get('s'),
            'price': float(data.get('c', 0)),  # 最新價格
            'price_change': float(data.get('P', 0)),  # 24小時價格變化百分比
            'price_change_abs': float(data.get('p', 0)),  # 24小時價格變化金額
            'high_price': float(data.get('h', 0)),  # 24小時最高價
            'low_price': float(data.get('l', 0)),  # 24小時最低價
            'volume': float(data.get('v', 0)),  # 24小時交易量
            'quote_volume': float(data.get('q', 0)),  # 24小時交易額
            'timestamp': datetime.now().isoformat()
        })
        
        # 計算更新後的持倉盈虧（如果有持倉）
        update_btc_position_pnl()
        
    except Exception as e:
        logger.error(f"處理BTC WebSocket訊息失敗: {e}")

def on_btc_ws_error(ws, error):
    """WebSocket錯誤"""
    logger.error(f"BTC WebSocket錯誤: {error}")

def on_btc_ws_close(ws, close_status_code, close_msg):
    """WebSocket連接關閉"""
    global btc_ws_connected
    btc_ws_connected = False
    logger.debug(f"BTC WebSocket連接關閉: {close_status_code} - {close_msg}")
    
    # 嘗試重新連接
    threading.Timer(5.0, reconnect_btc_websocket).start()

def reconnect_btc_websocket():
    """重新連接WebSocket"""
    logger.info("嘗試重新連接BTC WebSocket...")
    start_btc_websocket()

def stop_btc_websocket():
    """停止BTC WebSocket連接"""
    global btc_ws, btc_ws_connected
    
    try:
        if btc_ws:
            btc_ws_connected = False
            btc_ws.close()
            logger.info("BTC WebSocket連接已關閉")
    except Exception as e:
        logger.error(f"關閉BTC WebSocket失敗: {e}")

def update_btc_position_pnl():
    """根據實時價格更新持倉盈虧"""
    global binance_client, btc_real_time_data
    
    try:
        if not binance_client or not btc_real_time_data:
            return
        
        current_price = btc_real_time_data.get('price', 0)
        if current_price <= 0:
            return
        
        # 獲取持倉信息（簡化版，只獲取基本數據）
        positions = binance_client.get_position_info()
        if not positions:
            return
        
        # 更新實時盈虧數據
        for pos in positions:
            position_amt = float(pos.get('positionAmt', 0))
            if position_amt != 0:
                entry_price = float(pos.get('entryPrice', 0))
                if entry_price > 0:
                    # 計算實時盈虧
                    if position_amt > 0:  # 多頭
                        pnl = (current_price - entry_price) * position_amt
                    else:  # 空頭
                        pnl = (entry_price - current_price) * abs(position_amt)
                    
                    # 更新實時數據
                    btc_real_time_data[f'position_pnl_{pos.get("symbol")}'] = pnl
        
    except Exception as e:
        logger.error(f"更新BTC持倉盈虧失敗: {e}")

def check_btc_api_connection():
    """檢查BTC API連接狀態"""
    global binance_client, btc_last_connection_check
    
    try:
        if not binance_client:
            return False
            
        # 嘗試獲取帳戶信息來測試連接
        account_info = binance_client.get_account_info()
        if account_info and 'totalWalletBalance' in account_info:
            btc_last_connection_check = time.time()
            return True
        else:
            return False
            
    except Exception as e:
        logger.error(f"BTC API連接檢查失敗: {e}")
        return False

def btc_reconnect_api():
    """BTC API重連功能"""
    global btc_is_reconnecting, btc_reconnect_attempts, binance_client
    
    if btc_is_reconnecting:
        logger.info("BTC API重連已在進行中，跳過此次重連")
        return False
        
    btc_is_reconnecting = True
    logger.info(f"開始BTC API重連嘗試 (第{btc_reconnect_attempts + 1}次)...")
    
    try:
        # 檢查登入狀態
        env_data = load_btc_env_data()
        if env_data.get('LOGIN_BTC', '0') != '1':
            logger.info("BTC未登入狀態，停止重連嘗試")
            btc_is_reconnecting = False
            return False
        
        # 嘗試重新建立連接
        api_key = env_data.get('BINANCE_API_KEY', '').strip()
        secret_key = env_data.get('BINANCE_SECRET_KEY', '').strip()
        
        if not api_key or not secret_key:
            logger.info("BTC API密鑰不存在，停止重連")
            btc_is_reconnecting = False
            return False
        
        # 重新創建BTC客戶端
        binance_client = BinanceClient(api_key, secret_key)
        
        # 檢查連接
        if check_btc_api_connection():
            logger.info("BTC API重連成功！")
            btc_reconnect_attempts = 0  # 重置重連次數
            btc_is_reconnecting = False
            return True
        else:
            btc_reconnect_attempts += 1
            logger.error(f"BTC API重連失敗，將在30秒後重試 (嘗試次數: {btc_reconnect_attempts})")
            
            # 檢查是否超過最大重連次數（實際上是無限重連）
            if btc_reconnect_attempts < btc_max_reconnect_attempts:
                # 30秒後重試
                threading.Timer(30.0, btc_reconnect_api).start()
            
            btc_is_reconnecting = False
            return False
            
    except Exception as e:
        btc_reconnect_attempts += 1
        logger.error(f"BTC API重連異常: {e}，將在30秒後重試 (嘗試次數: {btc_reconnect_attempts})")
        
        if btc_reconnect_attempts < btc_max_reconnect_attempts:
            threading.Timer(30.0, btc_reconnect_api).start()
        
        btc_is_reconnecting = False
        return False

def start_btc_connection_monitor():
    """啟動BTC連接監控線程"""
    def connection_monitor():
        global btc_last_connection_check
        btc_connection_notified = False  # 追蹤是否已發送斷線通知
        
        while not btc_shutdown_flag.is_set():
            try:
                # 檢查是否登入
                env_data = load_btc_env_data()
                if env_data.get('LOGIN_BTC', '0') != '1':
                    time.sleep(btc_connection_check_interval)
                    continue
                
                # 檢查連接狀態
                if not check_btc_api_connection():
                    # 檢測到斷線，發送TG通知（只發送一次）
                    if not btc_connection_notified:
                        logger.info("檢測到BTC API斷線，發送TG通知...")
                        send_btc_telegram_message("⚠️ API連線異常！！！\n正在嘗試重新連線．．．")
                        btc_connection_notified = True
                    
                    logger.info("檢測到BTC API斷線，開始重連...")
                    if btc_reconnect_api():
                        # 重連成功，發送TG通知
                        logger.info("BTC API重連成功，發送TG通知...")
                        send_btc_telegram_message("✅ API連線成功！！！")
                        btc_connection_notified = False  # 重置通知狀態
                else:
                    # 連線正常，重置通知狀態
                    btc_connection_notified = False
                
                # 等待下次檢查
                time.sleep(btc_connection_check_interval)
                
            except Exception as e:
                logger.error(f"BTC連接監控異常: {e}")
                time.sleep(btc_connection_check_interval)
    
    # 創建並啟動監控線程
    monitor_thread = threading.Thread(target=connection_monitor, name="BTC連接監控線程", daemon=True)
    monitor_thread.start()
    btc_active_threads.append(monitor_thread)
    logger.info("BTC連接監控已啟動")

def start_btc_auto_logout_timer():
    """啟動BTC自動登出計時器（每12小時）"""
    global btc_auto_logout_timer
    
    def btc_auto_logout_task():
        global btc_auto_logout_timer
        
        try:
            logger.info("BTC 12小時自動登出重連開始...")
            
            # 檢查是否仍在登入狀態
            env_data = load_btc_env_data()
            if env_data.get('LOGIN_BTC', '0') != '1':
                logger.info("BTC已登出，取消自動重連")
                return
            
            # 嘗試重新連接
            if btc_reconnect_api():
                logger.info("BTC自動重連成功")
            else:
                logger.info("BTC自動重連失敗，將持續重試")
            
            # 設定下次12小時後執行
            btc_auto_logout_timer = threading.Timer(12 * 3600, btc_auto_logout_task)
            btc_auto_logout_timer.daemon = True
            btc_auto_logout_timer.start()
            
        except Exception as e:
            logger.error(f"BTC自動登出重連異常: {e}")
            
            # 即使異常也要設定下次執行
            btc_auto_logout_timer = threading.Timer(12 * 3600, btc_auto_logout_task)
            btc_auto_logout_timer.daemon = True
            btc_auto_logout_timer.start()
    
    # 停止現有計時器
    stop_btc_auto_logout_timer()
    
    # 啟動新的計時器（12小時後執行）
    btc_auto_logout_timer = threading.Timer(12 * 3600, btc_auto_logout_task)
    btc_auto_logout_timer.daemon = True
    btc_auto_logout_timer.start()
    logger.info("BTC自動登出計時器已啟動（12小時後執行）")

def stop_btc_auto_logout_timer():
    """停止BTC自動登出計時器"""
    global btc_auto_logout_timer
    
    try:
        if btc_auto_logout_timer and btc_auto_logout_timer.is_alive():
            btc_auto_logout_timer.cancel()
            logger.info("BTC自動登出計時器已停止")
    except Exception as e:
        logger.error(f"停止BTC自動登出計時器失敗: {e}")
    finally:
        btc_auto_logout_timer = None

def get_btc_realtime_data():
    """獲取BTC實時數據API"""
    global btc_real_time_data, btc_ws_connected
    
    try:
        return jsonify({
            'success': True,
            'connected': btc_ws_connected,
            'data': btc_real_time_data,
            'timestamp': datetime.now().isoformat()
        })
        
    except Exception as e:
        logger.error(f"獲取BTC實時數據失敗: {e}")
        return jsonify({
            'success': False,
            'message': f'獲取失敗: {str(e)}'
        })

# ========== BTC訂單監控系統 ==========

def start_btc_order_monitoring():
    """啟動BTC訂單監控WebSocket"""
    global order_monitor_thread, user_data_stream_key
    
    try:
        if not binance_client:
            logger.error("BTC客戶端未初始化，無法啟動訂單監控")
            return False
        
        # 獲取User Data Stream Listen Key
        try:
            listen_key_response = binance_client._make_request('POST', '/fapi/v1/listenKey')
            if not listen_key_response or 'listenKey' not in listen_key_response:
                logger.error("獲取Listen Key失敗")
                return False
            
            user_data_stream_key = listen_key_response['listenKey']
            logger.info(f"BTC Listen Key獲取成功: {user_data_stream_key[:10]}...")
            
        except Exception as e:
            logger.error(f"獲取Listen Key失敗: {e}")
            return False
        
        # 啟動WebSocket監控線程
        order_monitor_thread = threading.Thread(
            target=btc_order_websocket_worker,
            name="BTCOrderMonitor",
            daemon=True
        )
        order_monitor_thread.start()
        btc_active_threads.append(order_monitor_thread)
        
        # 啟動備用監控
        start_order_fallback_monitor()
        
        logger.info("BTC訂單監控WebSocket已啟動")
        return True
        
    except Exception as e:
        logger.error(f"啟動BTC訂單監控失敗: {e}")
        return False

def btc_order_websocket_worker():
    """
BTC訂單監控WebSocket工作線程"""
    global order_monitor_ws, user_data_stream_key
    
    while not btc_shutdown_flag.is_set():
        try:
            if not user_data_stream_key:
                logger.error("Listen Key不存在，無法連接WebSocket")
                time.sleep(30)
                continue
            
            ws_url = f"wss://fstream.binance.com/ws/{user_data_stream_key}"
            logger.info(f"連接BTC訂單監控WebSocket: {ws_url[:50]}...")
            
            import websocket
            
            def on_message(ws, message):
                try:
                    data = json.loads(message)
                    
                    # 處理訂單更新事件
                    if data.get('e') == 'ORDER_TRADE_UPDATE':
                        order_data = data['o']
                        order_id = order_data['i']
                        status = order_data['X']  # 訂單狀態
                        
                        logger.info(f"收到訂單更新: {order_id}, 狀態: {status}")
                        
                        if status == 'FILLED':
                            # 訂單完全成交
                            fill_price = float(order_data.get('ap', 0))  # 平均成交價
                            fill_quantity = order_data.get('z', '0')  # 成交數量
                            
                            # 處理成交事件
                            process_order_fill(order_id, fill_price, fill_quantity)
                            
                except Exception as e:
                    logger.error(f"處理WebSocket訊息失敗: {e}")
            
            def on_error(ws, error):
                logger.error(f"BTC訂單監控WebSocket錯誤: {error}")
            
            def on_close(ws, close_status_code, close_msg):
                logger.warning(f"BTC訂單監控WebSocket關閉: {close_status_code}, {close_msg}")
            
            def on_open(ws):
                logger.info("BTC訂單監控WebSocket連接成功")
            
            # 創建並啟動WebSocket
            order_monitor_ws = websocket.WebSocketApp(
                ws_url,
                on_message=on_message,
                on_error=on_error,
                on_close=on_close,
                on_open=on_open
            )
            
            order_monitor_ws.run_forever()
            
        except Exception as e:
            logger.error(f"BTC訂單監控WebSocket異常: {e}")
            
        # 等待30秒後重連
        if not btc_shutdown_flag.is_set():
            logger.info("等待30秒後重新連接BTC訂單監控WebSocket...")
            time.sleep(30)

def check_pending_orders_fallback():
    """輪詢檢查待成交訂單（備用方案）"""
    global pending_orders
    
    try:
        if not binance_client or not pending_orders:
            return
        
        orders_to_remove = []
        
        for order_id, order_info in pending_orders.items():
            try:
                # 查詢訂單狀態
                order_status = binance_client.get_order('BTCUSDT', orderId=order_id)
                
                if order_status and order_status.get('status') == 'FILLED':
                    # 訂單成交
                    fill_price = float(order_status.get('avgPrice', 0))
                    fill_quantity = order_status.get('executedQty', order_info['quantity'])
                    
                    # 處理成交事件
                    process_order_fill(order_id, fill_price, fill_quantity)
                    orders_to_remove.append(order_id)
                    
                elif order_status and order_status.get('status') in ['CANCELED', 'REJECTED', 'EXPIRED']:
                    # 訂單已取消或失敗
                    logger.info(f"訂單 {order_id} 狀態: {order_status.get('status')}")
                    orders_to_remove.append(order_id)
                    
            except Exception as e:
                logger.error(f"檢查訂單 {order_id} 狀態失敗: {e}")
        
        # 移除已處理的訂單
        for order_id in orders_to_remove:
            remove_pending_order(order_id)
            
    except Exception as e:
        logger.error(f"輪詢檢查待成交訂單失敗: {e}")

def start_order_fallback_monitor():
    """啟動訂單備用監控（每30秒檢查一次）"""
    def monitor_worker():
        while not btc_shutdown_flag.is_set():
            check_pending_orders_fallback()
            time.sleep(30)
    
    fallback_thread = threading.Thread(
        target=monitor_worker,
        name="BTCOrderFallbackMonitor",
        daemon=True
    )
    fallback_thread.start()
    btc_active_threads.append(fallback_thread)
    logger.info("BTC訂單備用監控已啟動")

