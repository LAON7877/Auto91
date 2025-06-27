# -*- coding: utf-8 -*-
from flask import Flask, request
import requests
import json
import os
from datetime import datetime, timedelta
import schedule
import threading
import time
import pandas as pd
import shioaji as sj
from shioaji.constant import Action, OrderState
import logging
from logging.handlers import TimedRotatingFileHandler
import glob
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from dotenv import load_dotenv
import shutil

# 載入 .env 檔案
dotenv_path = os.path.join(os.path.dirname(__file__), 'admin', '.env')
if not os.path.exists(dotenv_path):
    raise FileNotFoundError(f"找不到 .env 檔案：{dotenv_path}")
load_dotenv(dotenv_path)

# 從 .env 載入 LOG_DIR
LOG_DIR = os.getenv('LOG_DIR')
if not LOG_DIR:
    raise ValueError("環境變數 LOG_DIR 未設置")

# 從 .env 載入 DEBUG_MODE
DEBUG_MODE = os.getenv('DEBUG_MODE', '0') == '1'

# 新增：翻譯永豐API技術訊息的函數
def translate_api_message(message):
    """將永豐API的技術訊息翻譯成中文"""
    translations = {
        # 連線相關
        "Session.*keep-alive.*detected session down": "連線中斷：系統檢測到連線已斷開",
        "Session.*reconnecting": "連線重連中：系統正在嘗試重新連線",
        "Connect attempt.*timed out": "連線超時：嘗試連線到伺服器超時",
        "TCP connection failure.*No route to host": "網路連線失敗：無法連接到伺服器",
        "Protocol or communication error.*are session HOST and PORT correct": "通訊協定錯誤：請檢查連線設定",
        "connected failed, return Not ready": "連線失敗：系統尚未準備就緒",
        "Session connection attempt failed": "連線嘗試失敗：無法建立連線",
        
        # 登入相關
        "error attempting transport connection": "傳輸連線錯誤：嘗試建立傳輸連線時發生錯誤",
        "Protocol or communication error when attempting to login": "登入通訊錯誤：登入時發生通訊協定錯誤",
        "Topic: api/v1/auth/token_login": "登入認證：正在進行API認證登入",
        
        # 一般錯誤
        "Not ready": "系統未就緒：系統尚未準備完成",
        "TimeoutError": "連線超時：請求等待回應超時",
        "No route to host": "無路由到主機：網路路由問題",
        
        # 成功訊息
        "Session.*connected": "連線成功：已成功連接到伺服器",
        "login successful": "登入成功：已成功登入系統",
        
        # 新增：處理更多技術訊息
        "SDK NOTICE.*Session.*error attempting transport connection": "傳輸連線錯誤：API嘗試建立傳輸連線時發生錯誤",
        "SDK NOTICE.*Protocol or communication error": "通訊協定錯誤：登入時發生通訊協定錯誤",
        "SDK NOTICE.*connected failed, return Not ready": "連線失敗：API系統尚未準備就緒",
        "SDK NOTICE.*Session connection attempt failed": "連線嘗試失敗：無法建立API連線",
        "pysolace/core/sol.cpp.*connected failed": "連線失敗：底層連線系統錯誤",
        "pysolace/core/sol.cpp.*Not ready": "系統未就緒：底層系統尚未準備完成",
        "Response Code: 0.*Event Code: 2.*TCP connection failure": "網路連線失敗：TCP連線建立失敗",
        "Response Code: 0.*Event Code: 12.*Session.*reconnecting": "連線重連中：系統正在嘗試重新連線",
        "Event: Session reconnecting": "連線重連中：系統正在嘗試重新連線",
        "Event: Session connection attempt failed": "連線嘗試失敗：無法建立連線",
        
        # 新增：處理 API 超時錯誤
        "TimeoutError: Topic: api/v1/portfolio/margin": "API 超時：獲取保證金資料超時",
        "TimeoutError: Topic: api/v1/portfolio/positions": "API 超時：獲取持倉資料超時",
        "TimeoutError: Topic: api/v1/order/place": "API 超時：下單請求超時",
        "TimeoutError: Topic: api/v1/order/cancel": "API 超時：取消訂單超時",
        "TimeoutError: Topic: api/v1/order/modify": "API 超時：修改訂單超時",
        "TimeoutError: Topic: api/v1/order/query": "API 超時：查詢訂單超時",
        "TimeoutError: Topic: api/v1/order/deal": "API 超時：查詢成交超時",
        "TimeoutError: Topic: api/v1/portfolio/balance": "API 超時：獲取餘額資料超時",
        "TimeoutError: Topic: api/v1/portfolio/account": "API 超時：獲取帳戶資料超時",
        
        # 新增：處理一般 API 超時錯誤
        "TimeoutError: Topic: api/": "API 超時：API 請求等待回應超時",
        
        # 新增：處理餘額更新失敗
        "餘額更新失敗，重新登入無效": "餘額更新失敗：API 連線異常，無法獲取帳戶資料",
        
        # 新增：處理 API 狀態訊息
        "connected=False, futopt_account=": "API 狀態：連線中斷，期貨帳戶未設定",
        "connected=True, futopt_account=": "API 狀態：連線正常，期貨帳戶已設定",
        

    }
    
    for pattern, translation in translations.items():
        if re.search(pattern, message, re.IGNORECASE | re.DOTALL):
            return translation
    
    return message

# 新增：過濾永豐API技術訊息的函數
def should_log_message(message):
    """判斷是否應該記錄此訊息"""
    # 過濾掉一些不必要的技術訊息
    filter_patterns = [
        r'SDK NOTICE.*solClient\.c:\d+',  # SDK內部訊息
        r'pysolace/core/sol\.cpp:\d+',    # pysolace內部訊息
        r'\[thread \d+\]',                # 執行緒訊息
        # 註解：不過濾 Response Code 訊息，讓它們被翻譯顯示
        # r'Response Code: \d+.*Event Code: \d+.*Info:.*Event:',  # 完整的回應碼訊息
        # r'Response Code: \d+.*Event Code: \d+',  # 回應碼訊息
        # r'Event Code: \d+.*Event:',       # 事件碼訊息
        # r'Event: Session up',             # 連線成功事件
        # r'Event: Session.*',              # 所有Session事件
        r'Client: PYAPI/.*',              # 客戶端識別碼
        r'VPN name.*',                    # VPN名稱
        r'peer host.*',                   # 對等主機
        r'local address.*',               # 本地地址
        r'connection.*',                  # 連線類型
        r'fd \d+',                        # 檔案描述符
        r'error = \d+',                   # 錯誤碼
        r'Corr: [a-z0-9]+',              # 相關性ID
        r'payload: \{.*\}',               # 完整的payload內容
        r'token: eyJ[A-Za-z0-9\-_\.]+',   # JWT token
        r'person_id=\'[A-Z0-9]+\'',       # 人員ID
        r'broker_id=\'[A-Z0-9]+\'',       # 券商ID
        r'account_id=\'[0-9]+\'',         # 帳戶ID
        r'signed=[A-Za-z]+',              # 簽署狀態
        r'username=\'[^\']+\'',           # 使用者名稱
        r'account_type=\'[A-Z]\'',        # 帳戶類型
        r'version=\'[0-9\.]+\'',          # 版本號
        r'p2p=\'#[^\']+\'',               # P2P資訊
        r'permissions=\[[^\]]+\]',        # 權限列表
        r'level=\d+',                     # 等級
        r'ca_required=[A-Za-z]+',         # CA要求
        r'Client: PYAPI/[A-Z0-9]+/[0-9]+/[0-9]+/[0-9]+/[0-9\.]+',  # 完整客戶端識別碼
    ]
    
    for pattern in filter_patterns:
        if re.search(pattern, message, re.IGNORECASE):
            return False
    
    return True

# 新增：自定義日誌格式器來翻譯API訊息
class TranslatedFormatter(logging.Formatter):
    def format(self, record):
        # 檢查是否應該記錄此訊息
        if hasattr(record, 'msg') and not should_log_message(str(record.msg)):
            return ""  # 返回空字串，不記錄此訊息
        
        # 翻譯訊息內容
        if hasattr(record, 'msg'):
            record.msg = translate_api_message(str(record.msg))
        if hasattr(record, 'args') and record.args:
            record.args = tuple(translate_api_message(str(arg)) if isinstance(arg, str) else arg for arg in record.args)
        
        result = super().format(record)
        
        # 如果翻譯後是空字串，不記錄
        if not result.strip():
            return ""
        
        return result

# 配置日誌
class CustomTimedRotatingFileHandler(TimedRotatingFileHandler):
    def __init__(self, filename, when='midnight', interval=1, backupCount=30, encoding='utf-8', max_bytes=52428800):
        super().__init__(filename, when=when, interval=interval, backupCount=backupCount, encoding=encoding)
        self.max_bytes = max_bytes  # 50MB = 52428800 bytes
        self.current_size = 0
        self.rollover_reason = ""  # 記錄輪轉原因
        self.daily_sequence = 1  # 當日序號計數器
    
    def shouldRollover(self, record):
        """檢查是否需要輪轉 - 雙重條件檢查"""
        # 1. 檢查時間條件（午夜輪轉）
        t = int(time.time())
        if t >= self.rolloverAt:
            self.rollover_reason = "日期變更"
            return True
        
        # 2. 檢查文件大小條件
        if self.stream and hasattr(self.stream, "tell"):
            if self.stream.tell() + len(record.getMessage()) >= self.max_bytes:
                self.rollover_reason = f"檔案大小超過{self.max_bytes // (1024*1024)}MB"
                return True
        
        # 3. 檢查檔案實際大小（備用檢查）
        if os.path.exists(self.baseFilename):
            file_size = os.path.getsize(self.baseFilename)
            if file_size >= self.max_bytes:
                self.rollover_reason = f"檔案大小超過{self.max_bytes // (1024*1024)}MB"
                return True
        
        return False
    
    def emit(self, record):
        """重寫emit方法，每次寫入前檢查輪轉條件"""
        try:
            if self.shouldRollover(record):
                self.doRollover()
            super().emit(record)
        except Exception:
            self.handleError(record)
    
    def _get_next_sequence_number(self, base_filename):
        """獲取下一個序號"""
        sequence = 1
        while True:
            candidate = base_filename.replace('.log', f'_{sequence}.log')
            if not os.path.exists(candidate):
                return sequence, candidate
            sequence += 1
    
    def doRollover(self):
        """執行輪轉操作"""
        # 先關閉當前檔案流
        if self.stream:
            self.stream.close()
            self.stream = None
        
        if os.path.exists(self.baseFilename):
            try:
                if "日期變更" in self.rollover_reason:
                    # 日期輪轉：更新檔案名到新日期，舊檔案保持原名
                    new_date = datetime.now().strftime('%Y%m%d')
                    new_filename = self.baseFilename.replace(
                        self.baseFilename.split('_')[-1].split('.')[0], 
                        new_date
                    )
                    
                    # 重新設定基礎檔案名為新日期
                    self.baseFilename = new_filename
                    
                    print(f"✅ 日誌檔案已輪替（{self.rollover_reason}）：{os.path.basename(new_filename)}")
                    
                else:
                    # 大小輪轉：使用當日序號
                    self.daily_sequence, backup_filename = self._get_next_sequence_number(self.baseFilename)
                    
                    # 重新命名當前檔案為備份檔案
                    os.rename(self.baseFilename, backup_filename)
                    print(f"✅ 日誌檔案已輪替（{self.rollover_reason}）：{os.path.basename(backup_filename)}")
                
            except (OSError, PermissionError) as e:
                print(f"無法輪替日誌檔案（{self.rollover_reason}）：{e}")
                # 如果重新命名失敗，嘗試清空檔案
                try:
                    with open(self.baseFilename, 'w', encoding='utf-8') as f:
                        f.write(f"# 日誌檔案因無法輪替而清空於 {datetime.now()} (原因：{self.rollover_reason})\n")
                    print(f"日誌檔案已清空：{os.path.basename(self.baseFilename)}")
                except Exception as e2:
                    print(f"無法清空日誌檔案：{e2}")
        
        # 重新開啟檔案流
        self.mode = 'a'
        try:
            self.stream = self._open()
        except Exception as e:
            print(f"無法重新開啟日誌檔案：{e}")
            self.stream = None
        
        # 更新下次輪轉時間（僅針對時間輪轉）
        if "日期變更" in self.rollover_reason:
            self.rolloverAt = self.computeRollover(time.time())
        
        # 重置輪轉原因
        self.rollover_reason = ""

logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG if DEBUG_MODE else logging.INFO)

# 使用自定義格式器
formatter = TranslatedFormatter('%(asctime)s｜%(levelname)s｜%(message)s')

console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
console_handler.setFormatter(formatter)
console_handler.stream.reconfigure(encoding='utf-8', errors='replace')
logger.addHandler(console_handler)

os.makedirs(LOG_DIR, exist_ok=True)
log_date = datetime.now().strftime('%Y%m%d')

# 程式重啟輪轉邏輯
def handle_startup_log_rotation():
    """處理程式重啟時的日誌輪轉"""
    log_filename = os.path.join(LOG_DIR, f"autotx_{log_date}.log")
    
    # 檢查是否存在當日日誌檔案
    if os.path.exists(log_filename):
        file_size = os.path.getsize(log_filename)
        
        # 如果檔案存在且有內容，進行重啟輪轉
        if file_size > 100:  # 100 bytes 以上才輪轉
            # 找到下一個可用的序號
            sequence = 1
            while True:
                backup_filename = log_filename.replace('.log', f'_{sequence}.log')
                if not os.path.exists(backup_filename):
                    break
                sequence += 1
            
            try:
                os.rename(log_filename, backup_filename)
                print(f"日誌檔案已輪替（程式重啟）：{os.path.basename(backup_filename)}")
                return True
            except (OSError, PermissionError) as e:
                print(f"無法輪替重啟日誌檔案：{e}")
                return False
    return False

# 執行重啟輪轉檢查
restart_rotated = handle_startup_log_rotation()

# 單一日誌檔案
log_filename = os.path.join(LOG_DIR, f"autotx_{log_date}.log")
file_handler = CustomTimedRotatingFileHandler(
    filename=log_filename, when='midnight', interval=1, backupCount=30, encoding='utf-8', max_bytes=52428800
)
file_handler.setLevel(logging.DEBUG if DEBUG_MODE else logging.INFO)
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)

# 啟動訊息
if restart_rotated:
    logger.info(f"程式重啟後開始新日誌：autotx_{log_date}.log")
else:
    logger.info(f"開始儲存 日誌更新：autotx_{log_date}.log")

# 設定其他環境變數
BOT_TOKEN = os.getenv('BOT_TOKEN')
CHAT_ID = os.getenv('CHAT_ID')
HOLIDAY_DIR = os.getenv('HOLIDAY_DIR')
API_KEY = os.getenv('API_KEY')
SECRET_KEY = os.getenv('SECRET_KEY')
CA_PATH = os.getenv('CA_PATH')
CA_PASSWD = os.getenv('CA_PASSWD')
PERSON_ID = os.getenv('PERSON_ID')
CERT_START = datetime.strptime(os.getenv('CERT_START'), '%Y-%m-%d %H:%M:%S')
CERT_END = datetime.strptime(os.getenv('CERT_END'), '%Y-%m-%d %H:%M:%S')

app = Flask(__name__)

# ===================== 系統狀態管理類 =====================
class SystemState:
    """統一管理系統的全域狀態"""
    def __init__(self):
        # API 相關
        self.api = sj.Shioaji()
        self.connected = False
        
        # 合約相關
        self.contract_txf = None
        self.contract_mxf = None
        self.contract_tmf = None
        self.delivery_dates = {'txf': None, 'mxf': None, 'tmf': None}
        
        # 交易狀態
        self.has_processed_delivery_exit = False
        self.last_delivery_exit_price = {'txf': None, 'mxf': None, 'tmf': None}
        self.active_trades = {'txf': None, 'mxf': None, 'tmf': None}

        self.order_octype_map = {}
        
        # 餘額與保證金
        self.balance_cache = {'balance': None, 'margin': None, 'last_updated': None}
        self.margin_requirements = {}
        
        # 轉倉相關
        self.use_next_month_contract = False
        self.rollover_reminder_sent = False
        self.rollover_notification_sent = False
        
        # 通知狀態
        self.last_notify_status = "ok"  # "ok"=正常, "error"=API 連線失敗
        self.last_non_trading_log_date = None
        self.last_non_trading_time_log = None
        self.last_notify_time = 0
        self.recent_signals = set()
    
    def reset_daily_flags(self):
        """重置每日標記"""
        self.has_processed_delivery_exit = False
        self.rollover_reminder_sent = False
        self.rollover_notification_sent = False
        self.last_non_trading_log_date = None
        self.last_non_trading_time_log = None
        
    def clear_trade_data(self):
        """清理交易資料"""
        self.active_trades = {'txf': None, 'mxf': None, 'tmf': None}

        self.order_octype_map.clear()
        self.recent_signals.clear()

# 創建全域狀態實例
state = SystemState()

# ===================== 配置管理類 =====================
class AppConfig:
    """統一管理應用程式配置和常數"""
    # 合約映射
    CONTRACT_KEY_MAP = {"大台": "txf", "小台": "mxf", "微台": "tmf"}
    
    # 每點價格（TWD）
    POINT_VALUES = {
        "大台": 200,
        "小台": 50,
        "微台": 10
    }
    
    # 操作訊息翻譯
    OP_MSG_TRANSLATIONS = {
        "Order not found": "訂單未找到",
        "Price not satisfied": "價格未滿足",
        "Insufficient margin": "保證金不足",
        "Invalid quantity": "無效數量",
        "Invalid price": "無效價格",
        "Market closed": "市場已關閉",
        "非該商品可下單時間": "非交易時間",
        "未知錯誤": "未知錯誤"
    }
    
    # API回調狀態翻譯
    CALLBACK_STATUS_TRANSLATIONS = {
        # 訂單狀態翻譯
        "OrderState.Submitted": "已提交",
        "OrderState.FuturesOrder": "期貨訂單", 
        "OrderState.FuturesDeal": "期貨成交",
        "OrderState.Filled": "已成交",
        "OrderState.Failed": "失敗",
        "OrderState.Cancelled": "已取消",
        "OrderState.PartiallyFilled": "部分成交",
        "OrderState.PendingSubmit": "待提交",
        "OrderState.PreSubmitted": "預提交",
        "OrderState.Inactive": "未啟用",
        
        # 操作類型翻譯
        "UpdatePrice": "價格更新",
        "UpdateOrder": "訂單更新", 
        "Cancel": "取消訂單",
        "Deal": "成交",
        "NewOrder": "新訂單",
        "ModifyOrder": "修改訂單",
        
        # 合約類型翻譯
        "FUT": "期貨",
        "OPT": "選擇權",
        "STK": "股票",
        
        # 動作翻譯  
        "Buy": "買進",
        "Sell": "賣出",
        
        # 訂單類型翻譯
        "ROD": "當日有效單",
        "IOC": "立即成交否則取消",
        "FOK": "全部成交否則取消",
        
        # 價格類型翻譯
        "LMT": "限價",
        "MKT": "市價",
        "STP": "停損",
        "UPL": "上限價",
        "DPL": "下限價",
        
        # 開平倉翻譯
        "New": "開倉",
        "Cover": "平倉",
        "Auto": "自動",
        
        # 交易所翻譯
        "TIM": "台指期",
        "TSE": "台股",
        "OTC": "櫃買"
    }
    
    # 系統設定
    NOTIFY_INTERVAL = 1800  # 30分鐘
    ALLOWED_IPS = {"127.0.0.1"}  # 可自行擴充
    
    # 交易時段
    MORNING_START = "08:30"
    MORNING_END = "13:45"
    AFTERNOON_START = "14:50"
    AFTERNOON_END = "05:01"

# 創建配置實例
config = AppConfig()

# 保留向後兼容的變量引用
contract_key_map = config.CONTRACT_KEY_MAP

# 為了向後兼容，保留一些舊的變量引用（後續會逐步替換）
api = state.api
connected = state.connected
contract_txf = state.contract_txf
contract_mxf = state.contract_mxf
contract_tmf = state.contract_tmf
delivery_dates = state.delivery_dates
has_processed_delivery_exit = state.has_processed_delivery_exit
last_delivery_exit_price = state.last_delivery_exit_price
balance_cache = state.balance_cache
margin_requirements = state.margin_requirements
active_trades = state.active_trades
# pending_deals 功能已移除，改為立即發送通知
order_octype_map = state.order_octype_map

# 常數已移至 AppConfig 類中統一管理

# ===================== 系統同步 =====================
global_lock = threading.Lock()

# 通知狀態和系統常數已整合到對應的類中管理

# ===================== 全域狀態監控函數 =====================
def get_system_status():
    """獲取系統狀態總覽（用於調試和監控）"""
    return {
        'api_status': {
            'connected': state.connected,
            'api_ready': hasattr(state.api, 'futopt_account') and state.api.futopt_account is not None
        },
        'contracts': {
            'txf': state.contract_txf.code if state.contract_txf else None,
            'mxf': state.contract_mxf.code if state.contract_mxf else None,
            'tmf': state.contract_tmf.code if state.contract_tmf else None,
            'delivery_dates': state.delivery_dates
        },
        'trading_status': {
            'active_trades': state.active_trades,
            'immediate_notification': '立即通知模式',
            'order_mapping_count': len(state.order_octype_map),
            'recent_signals_count': len(state.recent_signals)
        },
        'balance_info': {
            'balance': state.balance_cache.get('balance'),
            'margin': state.balance_cache.get('margin'),
            'last_updated': state.balance_cache.get('last_updated'),
            'margin_requirements': state.margin_requirements
        },
        'rollover_status': {
            'use_next_month_contract': state.use_next_month_contract,
            'rollover_reminder_sent': state.rollover_reminder_sent,
            'rollover_notification_sent': state.rollover_notification_sent,
            'has_processed_delivery_exit': state.has_processed_delivery_exit
        },
        'notification_status': {
            'last_notify_status': state.last_notify_status,
            'last_notify_time': state.last_notify_time,
            'last_non_trading_log_date': state.last_non_trading_log_date,
            'last_non_trading_time_log': state.last_non_trading_time_log
        }
    }

# 新增輔助函數：推斷合約名稱
def get_contract_name(code):
    if code.startswith('TXF'):
        return "大台"
    elif code.startswith('MXF'):
        return "小台"
    elif code.startswith('TMF'):
        return "微台"
    return "未知"

# 新增輔助函數：生成動作顯示
def get_action_display(octype, direction):
    if octype == 'New':
        return "多單買入" if direction == 'Buy' else "空單買入" if direction == 'Sell' else "未知動作"
    elif octype == 'Cover':
        return "多單賣出" if direction == 'Sell' else "空單賣出" if direction == 'Buy' else "未知動作"
    return "未知動作"

# 新增輔助函數：生成訂單類型顯示
def get_order_type_display(price_type, order_type):
    """獲取訂單類型顯示"""
    price_display = "市價單" if price_type == "MKT" else "限價單"
    return f"{price_display}（{order_type}）"

def notify_and_log(msg, level="info", notify=True, log=True):
    """
    統一處理日誌與 Telegram 通知。
    - msg: 訊息內容（格式自訂，完全保留原本格式）
    - level: "info"、"error"、"warning"、"critical"、"debug"
    - notify: 是否發送 Telegram
    - log: 是否記錄日誌
    """
    if log:
        if level == "info":
            logger.info(msg)
        elif level == "error":
            logger.error(msg)
        elif level == "warning":
            logger.warning(msg)
        elif level == "critical":
            logger.critical(msg)
        elif level == "debug":
            logger.debug(msg)
        else:
            logger.info(msg)
    if notify:
        try:
            send_telegram(msg)
        except Exception as e:
            logger.warning(f"Telegram 發送失敗，已降級為 warning 日誌：{str(e)} | 原訊息：{msg}")

def handle_exception(title, e, notify=True, extra_msg=None):
    """
    統一處理例外狀況的日誌與通知。
    - title: 錯誤標題（如"下單失敗"）
    - e: Exception 物件
    - notify: 是否發送 Telegram
    - extra_msg: 額外補充內容
    """
    msg = f"{title}：{str(e)[:100]}"
    if extra_msg:
        msg += f"\n{extra_msg}"
    notify_and_log(msg, level="error", notify=notify)

def send_telegram(msg, max_retries=3, retry_delay=2, _is_internal=False):
    # logger.info(f"呼叫 send_telegram, msg={msg[:100]}")  # 移除正常流程 log
    if not msg or not msg.strip():
        return False
    if not BOT_TOKEN or not CHAT_ID:
        if not _is_internal:
            logger.error("BOT_TOKEN 或 CHAT_ID 未設置")
        return False
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
    payload = {"chat_id": CHAT_ID, "text": msg}
    last_exception = None
    for attempt in range(max_retries):
        try:
            response = requests.post(url, json=payload, timeout=10)
            if response.status_code == 200:
                logger.info(f"Telegram 訊息發送成功")
                return True
            else:
                error_msg = f"Telegram 發送失敗（狀態：{response.status_code}）"
                if response.text:
                    error_msg += f"｜回應：{response.text[:100]}"
                logger.error(error_msg)
                last_exception = Exception(error_msg)
        except requests.exceptions.Timeout:
            last_exception = Exception("Telegram 發送超時")
            logger.error(f"Telegram 發送超時（嘗試 {attempt + 1}/{max_retries}）")
        except requests.exceptions.ConnectionError:
            last_exception = Exception("Telegram 連線失敗")
            logger.error(f"Telegram 連線失敗（嘗試 {attempt + 1}/{max_retries}）")
        except Exception as e:
            last_exception = e
            logger.error(f"Telegram 發送異常（嘗試 {attempt + 1}/{max_retries}）：{str(e)}")
        if attempt < max_retries - 1:
            time.sleep(retry_delay)
    logger.error(f"Telegram 訊息發送失敗，已重試 {max_retries} 次｜最後錯誤：{last_exception}")
    return False

class HolidayCalendar:
    """處理假期表檔案，判斷交易日狀態."""
    def __init__(self, holiday_dir):
        self.holiday_dir = holiday_dir
        self.holidays = {}  # 儲存日期與交易狀態
        self._load_holidays()

    def _load_holidays(self):
        """從最新假期檔案載入數據."""
        csv_path, minguo_year = self._get_latest_holiday_csv()
        if not csv_path:
            msg = f"未找到假期檔案於 {self.holiday_dir}"
            logger.error(msg)
            send_telegram(msg)
            raise FileNotFoundError("未找到有效假期檔案")
        try:
            ad_year = minguo_year + 1911
            df = pd.read_csv(csv_path, encoding='big5', skiprows=1)
            logger.info(f"成功讀取 假期檔案：{os.path.basename(csv_path)}")
            date_column = '日期'
            remark_column = '備註(* : 市場無交易/僅辦理結算交割作業。o : 交易日。)'
            for _, row in df.iterrows():
                date_val = str(row.get(date_column, '')).strip()
                remark = str(row.get(remark_column, '')).strip()
                try:
                    match = re.match(r'(\d+)月(\d+)', date_val)
                    if match:
                        month, day = match.groups()
                        parsed_date = datetime(ad_year, int(month), int(day)).date()
                        self.holidays[parsed_date] = (remark == 'o')
                except Exception as e:
                    logger.warning(f"無法解析假期日期：{date_val}，錯誤：{str(e)}")
                    continue
        except Exception as e:
            msg = f"讀取假期檔案失敗：{str(e)}"
            logger.error(msg)
            send_telegram(msg)
            raise

    def _get_latest_holiday_csv(self):
        """獲取最新年份的假期檔案."""
        try:
            current_year = datetime.now().year
            csv_files = glob.glob(os.path.join(self.holiday_dir, "holidaySchedule_*.csv"))
            if not csv_files:
                return None, None

            max_year = -1
            latest_csv = None
            for csv_file in csv_files:
                filename = os.path.basename(csv_file)
                try:
                    year_str = filename.split('_')[-1].replace('.csv', '')
                    minguo_year = int(year_str)
                    ad_year = minguo_year + 1911
                    if ad_year <= current_year and minguo_year > max_year:
                        max_year = minguo_year
                        latest_csv = csv_file
                except ValueError:
                    continue

            return latest_csv, max_year if latest_csv else (None, None)
        except Exception as e:
            logger.error(f"掃描假期檔案失敗：{str(e)}")
            return None, None

    def is_trading_day(self, date=None, log_result=False):
        """檢查指定日期是否為交易日."""
        if date is None:
            date = datetime.now().date()
        
        date_str = date.strftime('%Y-%m-%d')
        
        # 週日固定為非交易日（週六有夜盤交易到凌晨05:00，所以週六是交易日）
        if date.weekday() == 6:  # 週日
            if log_result:
                logger.info(f"{date_str} 為週日，非交易日")
            return False
        
        # 檢查假期表
        if date in self.holidays:
            is_trading = self.holidays[date]
            if log_result:
                logger.info(f"{date_str} {'為交易日' if is_trading else '為非交易日'}（備註：{'o' if is_trading else '非交易'}）")
            return is_trading
        
        # 未在假期表中的日期視為交易日
        if log_result:
            logger.info(f"{date_str} 未在假期檔案中，視為交易日")
        return True

def is_delivery_day(date=None):
    """檢查指定日期是否為交割日."""
    if date is None:
        date = datetime.now().date()
    try:
        # 檢查 delivery_dates 是否已初始化
        if not any(delivery_dates.values()):
            logger.warning(f"交割日數據尚未初始化，假設 {date.strftime('%Y-%m-%d')} 非交割日")
            return False
        
        for key, delivery_date_str in delivery_dates.items():
            if delivery_date_str:
                try:
                    delivery_date = datetime.strptime(delivery_date_str, '%Y/%m/%d').date()
                except ValueError:
                    delivery_date = datetime.strptime(delivery_date_str, '%Y%m%d').date()
                if date == delivery_date:
                    logger.info(f"{date.strftime('%Y-%m-%d')} 為交割日（合約型別：{key.upper()}，交割日期：{delivery_date_str})")
                    return True
        
        return False
    except Exception as e:
        logger.error(f"檢查交割日失敗：{str(e)}，交割日狀態：{delivery_dates}")
        send_telegram(f"❌ 無法檢查交割日：{str(e)[:100]}\n交割日狀態：{delivery_dates}")
        return False

def was_yesterday_trading_day():
    yesterday = (datetime.now() - timedelta(days=1)).date()
    return holiday_calendar.is_trading_day(yesterday, log_result=False)

# 初始化 HolidayCalendar
try:
    holiday_calendar = HolidayCalendar(HOLIDAY_DIR)
except Exception as e:
    logger.critical(f"初始化假期日曆失敗：{str(e)}，系統無法啟動")
    send_telegram(f"❌ 初始化假期日曆失敗：{str(e)[:100]}，系統無法啟動")
    os._exit(1)

def login_api(max_retries=5, retry_delay=30, cooldown=300):
    global connected, contract_txf, contract_mxf, contract_tmf, api, error_counter
    error_counter = {}
    for attempt in range(max_retries):
        try:
            current_time = datetime.now()
            if current_time < CERT_START or current_time > CERT_END:
                send_telegram("API 憑證已過期，請盡快更新！")
                os._exit(1)
            if api:
                try:
                    api.logout()
                except Exception as e:
                    logger.warning(f"登出失敗：{str(e)}，將強制重置 API 物件")
                api = None
            logger.info(f"正在登入 永豐 API... (嘗試 {attempt + 1}/{max_retries})")
            api = sj.Shioaji()
            api.login(api_key=API_KEY, secret_key=SECRET_KEY)
            api.activate_ca(ca_path=CA_PATH, ca_passwd=CA_PASSWD, person_id=PERSON_ID)
            time.sleep(2)
            accounts = [acc for acc in api.list_accounts() if acc.account_type == 'F']
            if not accounts:
                logger.error("無期貨帳戶")
                connected = False
                return False
            api.futopt_account = accounts[0]
            current_month = datetime.now().strftime('%Y%m')
            current_date = datetime.now().date()
            for code, var in [("TXF", "contract_txf"), ("MXF", "contract_mxf"), ("TMF", "contract_tmf")]:
                contracts = api.Contracts.Futures.get(code)
                logger.debug(f"[login_api] {code} contracts={contracts}")
                if not contracts:
                    logger.error(f"無法獲取 {code} 合約")
                    connected = False
                    return False
                logger.info(f"{code} 可用合約：")
                for c in contracts:
                    logger.info(f"  - 合約代碼: {c.code}｜交割日期: {c.delivery_date}｜交割月份: {c.delivery_month}｜名稱: {c.name}")
                sorted_contracts = sorted(contracts, key=lambda x: x.delivery_date)
                globals()[var] = sorted_contracts[0]
                delivery_dates[var.lower()] = sorted_contracts[0].delivery_date
            if is_delivery_day():
                # 確保保證金資料已更新
                if not margin_requirements:
                    update_margin_requirements()
                for code, var in [("TXF", "contract_txf"), ("MXF", "contract_mxf"), ("TMF", "contract_tmf")]:
                    contracts = api.Contracts.Futures.get(code)
                    sorted_contracts = sorted(contracts, key=lambda x: x.delivery_date)
                    for c in sorted_contracts:
                        if c.delivery_month > current_month:
                            globals()[var] = c
                            delivery_dates[var.lower()] = c.delivery_date
                            break
                    else:
                        globals()[var] = sorted_contracts[0]
                        delivery_dates[var.lower()] = sorted_contracts[0].delivery_date
                    contract_name = "大台" if code == "TXF" else "小台" if code == "MXF" else "微台"
                    margin = margin_requirements.get(contract_name, 0)
                    logger.info(f"{code} 選用合約：{globals()[var].code}｜交割日：{globals()[var].delivery_date}｜保證金：{margin}")
            else:
                # 確保保證金資料已更新
                if not margin_requirements:
                    update_margin_requirements()
                for code, var in [("TXF", "contract_txf"), ("MXF", "contract_mxf"), ("TMF", "contract_tmf")]:
                    contract_name = "大台" if code == "TXF" else "小台" if code == "MXF" else "微台"
                    margin = margin_requirements.get(contract_name, 0)
                    logger.info(f"{code} 選用合約：{globals()[var].code}｜交割日：{globals()[var].delivery_date}｜保證金：{margin}")
            
            # 所有合約初始化後檢查交割日（系統啟動時顯示一次交易日狀態）
            current_date = datetime.now().date()
            is_trading_today = holiday_calendar.is_trading_day(current_date, log_result=False)
            is_delivery_today = is_delivery_day()
            trading_status = "交易日" if is_trading_today else "非交易日"
            delivery_status = "交割日" if is_delivery_today else "非交割日"
            
            # 週幾的中文對應
            weekday_chinese = {
                'Monday': '週一', 'Tuesday': '週二', 'Wednesday': '週三', 
                'Thursday': '週四', 'Friday': '週五', 'Saturday': '週六', 'Sunday': '週日'
            }
            weekday_eng = current_date.strftime('%A')
            weekday_display = weekday_chinese.get(weekday_eng, weekday_eng)
                
            logger.info(f"今天日期：{current_date.strftime('%Y-%m-%d')} 為{weekday_display}！【{trading_status}】【{delivery_status}】")
            
            connected = True
            logger.info("成功登入 永豐 API，連線建立！")
            
            if state.last_notify_status == "error":
                reconnect_time = datetime.now().strftime('%Y/%m/%d %H:%M')
                send_telegram(f"✅ API 連線成功\n連線時間：{reconnect_time}\n已成功連上")
                state.last_notify_status = "ok"
            
            # 最後更新帳戶餘額
            try:
                update_balance_cache(verbose=False, is_scheduled=False, force_update=True)
                logger.info("成功更新 帳戶餘額，連線有效！")
            except Exception as e:
                translated_error = translate_api_message(str(e))
                logger.error(f"餘額更新失敗，連線可能無效：{translated_error}")
                logger.exception("[login_api] 餘額更新失敗 traceback：")
                raise Exception("餘額更新失敗，重新登入無效")
            
            return True
        except Exception as e:
            err_str = str(e)
            # 錯誤分類和翻譯
            if 'token' in err_str.lower() or 'expired' in err_str.lower() or 'invalid token' in err_str.lower() or '連線失敗' in err_str or 'connection' in err_str.lower() or 'timeout' in err_str.lower() or 'no route to host' in err_str.lower():
                error_type = 'token_or_connection'
                error_desc = "網路連線或認證問題"
            elif 'please check param' in err_str.lower():
                error_type = 'param_error'
                error_desc = "API 參數錯誤"
            else:
                error_type = 'other'
                error_desc = "其他系統錯誤"
            
            error_counter[error_type] = error_counter.get(error_type, 0) + 1
            logger.error(f"登入失敗類型：{error_type}（{error_desc}），次數：{error_counter[error_type]}")
            
            # 翻譯錯誤訊息
            translated_error = translate_api_message(err_str)
            logger.error(f"登入永豐 API 失敗 (嘗試 {attempt + 1}/{max_retries})：{translated_error}")
            
            if DEBUG_MODE:
                logger.exception("[login_api] 登入失敗詳細錯誤：")
            
            if error_type == 'param_error':
                if state.last_notify_status != 'param_error':
                    disconnect_time = datetime.now().strftime('%Y/%m/%d %H:%M')
                    send_telegram(f"❌ API 連線失敗\n斷線時間：{disconnect_time}\n斷線原因：{error_desc}，請檢查帳號權限或API版本\n詳細錯誤：{translated_error[:100]}")
                    state.last_notify_status = 'param_error'
                connected = False
                return False
            elif error_type == 'token_or_connection':
                if attempt < max_retries - 1:
                    logger.info(f"等待 {retry_delay} 秒後重試（{error_desc}）")
                    time.sleep(retry_delay)
                else:
                    disconnect_time = datetime.now().strftime('%Y/%m/%d %H:%M')
                    if state.last_notify_status != 'error':
                        send_telegram(f"❌ API 連線失敗\n斷線時間：{disconnect_time}\n斷線原因：{error_desc}\n詳細錯誤：{translated_error[:100]}\n\n系統將在 {cooldown//60} 分鐘後自動重試")
                        state.last_notify_status = 'error'
                    logger.info(f"達最大重試次數，暫停 {cooldown//60} 分鐘後再嘗試")
                    time.sleep(cooldown)
                    connected = False
                    return False
            else:
                if state.last_notify_status != 'error':
                    disconnect_time = datetime.now().strftime('%Y/%m/%d %H:%M')
                    send_telegram(f"❌ API 連線失敗\n斷線時間：{disconnect_time}\n斷線原因：{error_desc}\n詳細錯誤：{translated_error[:100]}")
                    state.last_notify_status = 'error'
                connected = False
                return False

def update_balance_cache(verbose=False, is_scheduled=False, force_update=False):
    # 使用 state 物件管理狀態，無需 global 聲明
    try:
        # 餘額更新時段控制（只有在排程更新時才檢查）
        if is_scheduled:
            now = datetime.now()
            now_date = now.date()
            now_time = now.time()
            
            # 首先檢查是否為交易日（靜默檢查，避免重複日誌）
            if not holiday_calendar.is_trading_day(now_date, log_result=False):
                # 避免重複日誌：每天只記錄一次非交易日提醒
                if state.last_non_trading_log_date != now_date:
                    logger.info(f"非交易日，暫停即時進度更新。當前日期：{now_date.strftime('%Y-%m-%d')}")
                    state.last_non_trading_log_date = now_date
                return
            
            # 早盤 8:30~13:45，午盤 14:50~05:01
            in_morning = now_time >= datetime.strptime("08:30", "%H:%M").time() and now_time <= datetime.strptime("13:45", "%H:%M").time()
            in_afternoon = now_time >= datetime.strptime("14:50", "%H:%M").time() or now_time <= datetime.strptime("05:01", "%H:%M").time()
            if not (in_morning or in_afternoon):
                # 避免重複日誌：每小時只記錄一次非交易時段提醒
                current_hour = now.strftime('%Y-%m-%d %H')
                if state.last_non_trading_time_log != current_hour:
                    logger.info(f"非交易時段，暫停即時進度更新。當前時間：{now.strftime('%H:%M')}")
                    state.last_non_trading_time_log = current_hour
                return
        margin_data = api.margin()
        balance_cache['balance'] = margin_data.equity_amount
        balance_cache['margin'] = margin_data.available_margin
        balance_cache['last_updated'] = datetime.now()
        total_pnl = 0.0
        positions = api.list_positions(api.futopt_account)
        for pos in positions:
            total_pnl += pos.pnl
        if verbose:
            field_translations = {
                'equity_amount': '權益總額',
                'equity': '權益總值',
                'today_balance': '今日餘額',
                'yesterday_balance': '昨日餘額',
                'available_margin': '可用保證金',
                'initial_margin': '原始保證金',
                'maintenance_margin': '維持保證金',
                'risk_indicator': '風險指標',
                'fee': '手續費',
                'tax': '期交稅',
                'future_settle_profitloss': '本日平倉損益'
            }
            translated_data = {field_translations.get(key, key): value for key, value in margin_data.__dict__.items() if key in field_translations}
            ordered_keys = [
                '權益總值', '權益總額', '今日餘額', '昨日餘額', '可用保證金',
                '原始保證金', '維持保證金', '風險指標', '手續費', '期交稅', '本日平倉損益'
            ]
            logger.info("開始更新帳戶權益：")
            for key in ordered_keys:
                value = translated_data.get(key, 0)
                unit = "%" if key == "風險指標" else ""
                logger.info(f" - {key}: {value}{unit}")
            logger.info(f" - 未實現盈虧: {total_pnl:.0f} TWD")
        elif is_scheduled:
            logger.info(f"[即時進度] 帳戶餘額：{balance_cache['balance']} TWD｜未實現盈虧：{total_pnl:.0f} TWD｜可用保證金：{balance_cache['margin']} TWD")
    except Exception as e:
        translated_error = translate_api_message(str(e))
        logger.error(f"更新帳戶餘額失敗：{translated_error}")
        logger.error(f"詳細錯誤資訊：{repr(e)}")
        logger.error(f"API 狀態：connected={connected}, futopt_account={hasattr(api, 'futopt_account') and api.futopt_account}")
        logger.exception("[update_balance_cache] 更新帳戶餘額失敗 traceback：")
        balance_cache['balance'] = None
        balance_cache['margin'] = None
        raise

def get_next_month_contract(code, current_month):
    contracts = api.Contracts.Futures.get(code)
    if not contracts:
        logger.error(f"無法獲取 {code} 合約")
        return None
    next_contract = None
    for c in sorted(contracts, key=lambda x: x.delivery_date):
        if c.delivery_month > current_month:
            next_contract = c
            break
    if next_contract:
        logger.info(f"選用次月合約：{code} {next_contract.code}，交割日：{next_contract.delivery_date}")
    else:
        logger.error(f"未找到 {code} 的次月合約，無法轉倉")
        send_telegram(f"❌ {code} 無次月合約，轉倉失敗，請檢查合約")
        return None
    return next_contract

def update_margin_requirements(is_scheduled=False):
    global margin_requirements
    try:
        url = "https://openapi.taifex.com.tw/v1/IndexFuturesAndOptionsMargining"
        response = requests.get(url)
        if response.status_code != 200:
            logger.error(f"獲取保證金數據失敗：{response.status_code} - {response.text}")
            send_telegram(f"❌ 獲取保證金數據失敗：HTTP {response.status_code} - {response.text[:100]}")
            return
        data = response.json()
        new_margins = {}
        for item in data:
            contract = item.get('Contract', '')
            margin = int(item.get('InitialMargin', 0))
            if contract == '臺股期貨':
                new_margins['大台'] = margin
            elif contract == '小型臺指':
                new_margins['小台'] = margin
            elif contract == '微型臺指期貨':
                new_margins['微台'] = margin
        if not new_margins:
            logger.warning("未從 API 解析到任何有效保證金數據")
        has_changed = False
        for contract_name in ['大台', '小台', '微台']:
            old_margin = margin_requirements.get(contract_name, 0)
            new_margin = new_margins.get(contract_name, 0)
            if old_margin != new_margin:
                has_changed = True
                break
        margin_requirements.clear()
        margin_requirements.update(new_margins)
        
        # 只在系統啟動或有變動時才記錄日誌
        if not is_scheduled:
            # 系統啟動時靜默更新，不顯示日誌
            pass
        elif is_scheduled and has_changed:
            msg = (
                f"期貨保證金更新\n"
                f"【大台】${margin_requirements.get('大台', 0):,}\n"
                f"【小台】${margin_requirements.get('小台', 0):,}\n"
                f"【微台】${margin_requirements.get('微台', 0):,}"
            )
            send_telegram(msg)
    except Exception as e:
        logger.error(f"獲取保證金數據失敗：{str(e)}")
        send_telegram(f"❌ 獲取保證金數據失敗：{str(e)[:100]}")

def place_order_for_contract(contract, qty, price, action, contract_name, is_entry, is_rollover_entry=False, position=None, order_type="IOC", price_type="LMT", is_manual=False):
    global last_delivery_exit_price, balance_cache, margin_requirements, order_octype_map
    logger.debug(f"[place_order_for_contract] contract={contract}, qty={qty}, price={price}, action={action}, contract_name={contract_name}, is_entry={is_entry}, is_rollover_entry={is_rollover_entry}, position={position}, order_type={order_type}, price_type={price_type}, is_manual={is_manual}")
    if not contract:
        msg = get_formatted_order_message(
            is_success=False,
            order_id="未知",
            contract_name=contract_name,
            qty=qty,
            price=price,
            octype="New" if is_entry else "Cover",
            direction="Buy" if action == Action.Buy else "Sell",
            order_type=order_type,
            price_type=price_type,
            is_manual=is_manual,
            reason="合約未設置",
            contract_code=contract.code if contract else None
        )
        notify_and_log(msg, level="error", notify=True, log=False)
        return f"[{contract_name}]：{qty} 口｜合約未設置\n", False, 0.0, None, None
    contract_key = contract_key_map.get(contract_name, "unknown")
    if contract_key == "unknown":
        msg = get_formatted_order_message(
            is_success=False,
            order_id="未知",
            contract_name=contract_name,
            qty=qty,
            price=price,
            octype="New" if is_entry else "Cover",
            direction="Buy" if action == Action.Buy else "Sell",
            order_type=order_type,
            price_type=price_type,
            is_manual=is_manual,
            reason="無效合約名稱",
            contract_code=contract.code if contract else None
        )
        notify_and_log(msg, level="error", notify=True, log=False)
        return f"[{contract_name}]：{qty} 口｜無效合約名稱\n", False, 0.0, None, None
    if is_entry:
        update_balance_cache(verbose=False, is_scheduled=False, force_update=True)
        required_margin = qty * margin_requirements.get(contract_name, 0)
        required_funds = required_margin + 500
        if balance_cache['balance'] is None or balance_cache['margin'] is None:
            msg = get_formatted_order_message(
                is_success=False,
                order_id="未知",
                contract_name=contract_name,
                qty=qty,
                price=price,
                octype="New" if is_entry else "Cover",
                direction="Buy" if action == Action.Buy else "Sell",
                order_type=order_type,
                price_type=price_type,
                is_manual=is_manual,
                reason="無法獲取餘額資訊",
                contract_code=contract.code if contract else None
            )
            notify_and_log(msg, level="error", notify=True, log=False)
            return f"[{contract_name}]：{qty} 口｜無法獲取餘額資訊\n", False, 0.0, None, None
        if balance_cache['balance'] < required_funds:
            msg = get_formatted_order_message(
                is_success=False,
                order_id="未知",
                contract_name=contract_name,
                qty=qty,
                price=price,
                octype="New" if is_entry else "Cover",
                direction="Buy" if action == Action.Buy else "Sell",
                order_type=order_type,
                price_type=price_type,
                is_manual=is_manual,
                reason=f"所需 {required_funds:.0f} TWD，可用 {balance_cache['balance']:.0f} TWD",
                contract_code=contract.code if contract else None
            )
            notify_and_log(msg, level="error", notify=True, log=False)
            return f"[{contract_name}]：{qty} 口｜餘額不足：所需 {required_funds:.0f} TWD，可用 {balance_cache['balance']:.0f} TWD\n", False, 0.0, None, None
        if balance_cache['margin'] < required_margin:
            msg = get_formatted_order_message(
                is_success=False,
                order_id="未知",
                contract_name=contract_name,
                qty=qty,
                price=price,
                octype="New" if is_entry else "Cover",
                direction="Buy" if action == Action.Buy else "Sell",
                order_type=order_type,
                price_type=price_type,
                is_manual=is_manual,
                reason=f"所需 {required_margin:.0f} TWD，可用 {balance_cache['margin']:.0f} TWD",
                contract_code=contract.code if contract else None
            )
            notify_and_log(msg, level="error", notify=True, log=False)
            return f"[{contract_name}]：{qty} 口｜保證金不足：所需 {required_margin:.0f} TWD，可用 {balance_cache['margin']:.0f} TWD\n", False, 0.0, None, None
    order_action = action if is_entry else (Action.Sell if position and position.direction == Action.Buy else Action.Buy)
    octype = sj.constant.FuturesOCType.New if is_entry else sj.constant.FuturesOCType.Cover
    logger.info(f"設置訂單：contract_name={contract_name}, is_entry={is_entry}, octype={octype}, action={order_action}, is_manual={is_manual}")
    display_action = get_action_display_by_rule(octype, order_action)
    order_type_display = get_order_type_display(price_type, order_type)
    try:
        logger.debug(f"[place_order_for_contract] order_octype_map={order_octype_map}")
        logger.info(f"準備下單：合約={contract_name} | 動作={display_action} | 數量={qty}口 | 類型={order_type_display} | 模式={'手動' if is_manual else '自動'}")
        order = api.Order(
            action=order_action,
            price=0 if price_type == "MKT" else price,
            quantity=qty,
            price_type=price_type,
            order_type=order_type,
            octype=octype,
            account=api.futopt_account
        )
        trade = api.place_order(contract, order)
        logger.debug(f"[place_order_for_contract] trade={trade}")
        if not trade or not hasattr(trade, 'order') or not trade.order.id:
            msg = get_formatted_order_message(
                is_success=False,
                order_id="未知",
                contract_name=contract_name,
                qty=qty,
                price=price,
                octype="New" if octype == sj.constant.FuturesOCType.New else "Cover",
                direction="Buy" if order_action == Action.Buy else "Sell",
                order_type=order_type,
                price_type=price_type,
                is_manual=is_manual,
                reason="訂單提交失敗",
                contract_code=contract.code if contract else None
            )
            notify_and_log(msg, level="error", notify=True, log=False)
            return f"[{contract_name}]：{qty} 口｜訂單無效\n", False, 0.0, None, None
        logger.info(f"訂單提交成功：單號={trade.order.id}")
        
        # 立即建立訂單記錄，避免競爭條件
        order_info = {
            'octype': 'New' if octype == sj.constant.FuturesOCType.New else 'Cover',
            'direction': 'Buy' if order_action == Action.Buy else 'Sell',
            'contract_name': contract_name,
            'order_type': order_type,
            'price_type': price_type,
            'is_manual': is_manual
        }
        with global_lock:
            logger.debug(f"[place_order_for_contract] (with lock) order_octype_map before update: {order_octype_map}")
            order_octype_map[trade.order.id] = order_info
            logger.debug(f"[place_order_for_contract] (with lock) order_octype_map after update: {order_octype_map}")
        logger.debug(f"更新 order_octype_map：order_id={trade.order.id}, octype={order_info['octype']}, is_manual={order_info['is_manual']}, details={order_info}")
        return f"[{contract_name}]：{qty} 口｜訂單已提交\n", True, 0.0, None, trade
    except Exception as e:
        logger.exception("[place_order_for_contract] 下單失敗 traceback：")
        msg = get_formatted_order_message(
            is_success=False,
            order_id="未知",
            contract_name=contract_name,
            qty=qty,
            price=price,
            octype="New" if is_entry else "Cover",
            direction="Buy" if action == Action.Buy else "Sell",
            order_type=order_type,
            price_type=price_type,
            is_manual=is_manual,
            reason=str(e)[:100],
            contract_code=contract.code if contract else None
        )
        notify_and_log(msg, level="error", notify=True, log=False)
        return f"[{contract_name}]：{qty} 口｜下單失敗，原因：{str(e)[:100]}\n", False, 0.0, None, None

def handle_futures_deal(deal):
    # logger.info(f"進入 handle_futures_deal, deal={deal}")  # 移除正常流程 log
    try:
        # logger.info(f"處理成交事件：{deal}")  # 移除正常流程 log
        trade_id = deal.get('trade_id', '')
        if not trade_id or not deal.get('action') or not deal.get('code'):
            logger.error(f"成交數據不完整：{deal}")
            send_telegram(f"❌ 處理成交通知失敗：成交數據不完整，trade_id={trade_id}")
            return
        contract_name = get_contract_name(deal.get('code', ''))
        octype_info = order_octype_map.get(trade_id)
        if octype_info is None:
            # 從交易記錄推斷
            today = datetime.now().strftime("%Y%m%d")
            filename = f"{LOG_DIR}/trades_{today}.json"
            oc_type, direction, order_type, price_type, is_manual = None, None, None, None, True
            if os.path.exists(filename):
                try:
                    with open(filename, 'r') as f:
                        trades = json.load(f)
                    for trade in trades:
                        if trade.get('deal_order_id') == trade_id and trade.get('type') == 'order':
                            raw_order = trade.get('raw_data', {}).get('order', {})
                            oc_type = raw_order.get('oc_type', 'New')
                            direction = raw_order.get('action', 'Sell')
                            order_type = raw_order.get('order_type', 'ROD')
                            price_type = raw_order.get('price_type', 'LMT')
                            is_manual = trade.get('is_manual', True)
                            break
                except Exception as e:
                    handle_exception("讀取交易記錄失敗", e)
            octype_info = {
                'octype': oc_type or 'New',
                'direction': direction or 'Sell',
                'contract_name': contract_name,
                'order_type': order_type or 'ROD',
                'price_type': price_type or 'LMT',
                'is_manual': is_manual
            }
            # 自動推斷訂單資訊（靜默處理，這是正常的備援機制）
        octype = octype_info['octype']
        direction = octype_info['direction']
        order_type = octype_info['order_type']
        price_type = octype_info['price_type']
        is_manual = octype_info.get('is_manual', True)
        octype_display = f"{'手動' if is_manual else '自動'}{'開倉' if octype == 'New' else '平倉'}"
        action_display = get_action_display_by_rule(octype, direction)
        order_type_display = get_order_type_display(price_type, order_type)
        # 獲取合約代碼和交割日期信息
        api_contract_code = deal.get('code', '')  # API返回的合約代碼（可能不完整）
        
        # 使用當前系統選用的完整合約代碼和交割日期
        contract_code = api_contract_code
        delivery_date = deal.get('delivery_date', '')
        
        try:
            # 根據合約名稱匹配當前選用的完整合約
            if contract_name == '大台' and contract_txf:
                contract_code = contract_txf.code
                delivery_date = contract_txf.delivery_date
            elif contract_name == '小台' and contract_mxf:
                contract_code = contract_mxf.code
                delivery_date = contract_mxf.delivery_date
            elif contract_name == '微台' and contract_tmf:
                contract_code = contract_tmf.code
                delivery_date = contract_tmf.delivery_date
            else:
                # 如果無法匹配，嘗試從API合約代碼推斷
                if api_contract_code.startswith('TXF') and contract_txf:
                    contract_code = contract_txf.code
                    delivery_date = contract_txf.delivery_date
                elif api_contract_code.startswith('MXF') and contract_mxf:
                    contract_code = contract_mxf.code
                    delivery_date = contract_mxf.delivery_date
                elif api_contract_code.startswith('TMF') and contract_tmf:
                    contract_code = contract_tmf.code
                    delivery_date = contract_tmf.delivery_date
                    
            logger.debug(f"合約代碼映射：API={api_contract_code} -> 系統={contract_code}, 交割日={delivery_date}")
            
        except Exception as e:
            logger.warning(f"無法獲取合約 {api_contract_code} 的完整資訊：{e}")
            # fallback 使用 API 返回的代碼
            contract_code = api_contract_code
            if not delivery_date:
                delivery_date = '未知'
        
        deal_data = {
            'contract_name': contract_name,
            'trade_id': trade_id,
            'seqno': deal.get('seqno', ''),
            'ordno': deal.get('ordno', ''),
            'exchange_seq': deal.get('exchange_seq', ''),
            'broker_id': deal.get('broker_id', ''),
            'account_id': deal.get('account_id', ''),
            'action': action_display,
            'octype': octype_display,
            'order_type': order_type_display,
            'code': contract_code,
            'price': deal.get('price', 0.0),
            'quantity': deal.get('quantity', 0),
            'subaccount': deal.get('subaccount', ''),
            'security_type': deal.get('security_type', ''),
            'delivery_month': deal.get('delivery_month', '') or get_delivery_month_by_code(contract_code),
            'delivery_date': delivery_date,
            'strike_price': deal.get('strike_price', 0.0),
            'option_right': deal.get('option_right', ''),
            'market_type': deal.get('market_type', ''),
            'combo': deal.get('combo', False),
            'is_manual': is_manual
        }
        deal_data['deal_time'] = datetime.fromtimestamp(deal.get('ts', 0)).strftime('%Y/%m/%d %H:%M')
        
        # 保存交易記錄
        save_trade({
            'type': 'deal',
            'trade_category': 'normal',
            'raw_data': deal,
            'deal_order_id': trade_id,
            f'deal_price_{contract_key_map[contract_name]}': deal_data['price'],
            f'deal_qty_{contract_key_map[contract_name]}': deal_data['quantity'],
            'timestamp': datetime.now().isoformat(),
            'is_manual': is_manual
        })
        
        # 立即發送成交通知
        msg = f"✅ 成交通知（{deal_data['deal_time']}）\n"
        msg += (
            f"選用合約：{format_contract_display(deal_data['code'], deal_data['delivery_date'])}\n"
            f"訂單類型：{deal_data['order_type']}\n"
            f"成交單號：{deal_data['trade_id']}\n"
            f"成交類型：{deal_data['octype']}\n"
            f"成交動作：{deal_data['action']}\n"
            f"成交部位：{deal_data['contract_name']}\n"
            f"成交數量：{deal_data['quantity']} 口\n"
            f"成交價格：{deal_data['price']:.0f}\n"
        )
        
        success = send_telegram(msg)
        if success:
            # 移除詳細的成交日誌訊息，send_telegram函數內已有基本的成功訊息
            pass
        else:
            logger.error(f"Telegram 訊息發送失敗｜成交單號：{trade_id}")
    except Exception as e:
        logger.exception("[handle_futures_deal] 處理成交事件失敗 traceback：")
        send_telegram(f"❌ 處理成交通知失敗：{str(e)[:100]}")

@app.route('/webhook', methods=['POST'])
def webhook():
    global has_processed_delivery_exit, active_trades
    client_ip = request.remote_addr
    if not is_ip_allowed(client_ip):
        logger.warning(f"非法 IP 嘗試存取 webhook：{client_ip}")
        return 'Forbidden', 403
    try:
        raw = request.data.decode('utf-8')
        logger.debug(f"[webhook] raw={raw}")
        # 移除重複日誌，統一在 process_signal() 中記錄
        if '{{strategy.order.alert_message}}' in raw or not raw.strip():
            logger.warning("無效訊號")
            return '無效訊號', 400
        data = json.loads(raw)
        signal_id = data.get('tradeId')
        with global_lock:
            logger.debug(f"[webhook] (with lock) recent_signals={state.recent_signals}")
            if signal_id in state.recent_signals:
                logger.warning(f"重複訊號 {signal_id}，忽略")
                return '重複訊號', 400
            state.recent_signals.add(signal_id)
            threading.Timer(10, lambda: state.recent_signals.discard(signal_id)).start()
        data['receive_time'] = datetime.now()
        process_signal(data)
        return 'OK', 200
    except Exception as e:
        logger.exception("[webhook] 處理錯誤 traceback：")
        logger.error(f"Webhook 處理錯誤：{str(e)}")
        send_telegram(f"❌ Webhook 錯誤：{str(e)[:100]}")
        return f'錯誤：{str(e)}', 500

@app.route('/balance', methods=['GET'])
def get_balance():
    """獲取完整帳戶餘額資訊"""
    client_ip = request.remote_addr
    if not is_ip_allowed(client_ip):
        logger.warning(f"非法 IP 嘗試存取 balance：{client_ip}")
        return 'Forbidden', 403
    
    try:
        # 強制更新餘額，不受交易日和時段限制
        update_balance_cache(verbose=True, is_scheduled=False, force_update=True)
        
        # 獲取持倉資訊
        positions_info = ""
        try:
            positions = api.list_positions(api.futopt_account)
            if positions:
                for pos in positions:
                    contract_type = "大台" if pos.code.startswith("TXF") else "小台" if pos.code.startswith("MXF") else "微台"
                    direction = "多單" if pos.direction == Action.Buy else "空單"
                    positions_info += (
                        f"［{contract_type}］｜"
                        f"動作：{direction}｜"
                        f"數量：{abs(pos.quantity)}口｜"
                        f"均價：{pos.price:.0f}\n"
                        f"未實現盈虧：{pos.pnl:.0f} TWD\n"
                    )
            else:
                positions_info = "無持倉部位\n"
        except Exception as e:
            logger.error(f"獲取持倉資訊失敗：{str(e)}")
            positions_info = "無法獲取持倉資訊\n"
        
        # 發送 Telegram 通知
        current_time = datetime.now().strftime('%Y/%m/%d %H:%M')
        balance_msg = (
            f"帳戶餘額查詢（{current_time}）\n"
            f"═════ 帳戶狀態 ═════\n"
            f"查詢時間：{current_time}\n"
            f"═════ 持倉狀態 ═════\n"
            f"{positions_info}"
        )
        send_telegram(balance_msg)
        
        return {"status": "success", "message": "餘額資訊已發送至 Telegram"}, 200
        
    except Exception as e:
        logger.error(f"獲取帳戶餘額失敗：{str(e)}")
        send_telegram(f"❌ 獲取帳戶餘額失敗：{str(e)[:100]}")
        return {"status": "error", "message": f"獲取餘額失敗：{str(e)}"}, 500

@app.route('/manual_order', methods=['POST'])
def manual_order():
    client_ip = request.remote_addr
    if not is_ip_allowed(client_ip):
        logger.warning(f"非法 IP 嘗試存取 manual_order：{client_ip}")
        return 'Forbidden', 403
    try:
        data = request.get_json()
        logger.debug(f"[manual_order] data={data}")
        contract_name = data.get('contract_name', '').strip()
        qty = int(data.get('quantity', 0))
        action = data.get('action', '').strip().upper()
        is_entry = data.get('is_entry', True)
        order_type = data.get('order_type', '').upper()
        price_type = data.get('price_type', 'LMT').upper()
        price = float(data.get('price', 0))
        if contract_name not in contract_key_map:
            logger.error(format_unified_log(
                log_type="錯誤",
                order_id="未知",
                contract_name=contract_name,
                contract_code="未知",
                qty=qty,
                price=price,
                action=action,
                reason="無效合約名稱",
                auto_fill=False
            ))
            msg = get_formatted_order_message(
                is_success=False,
                order_id="未知",
                contract_name=contract_name,
                qty=qty,
                price=price,
                octype="New",
                direction="Sell",
                order_type=order_type,
                price_type=price_type,
                is_manual=True,
                reason=f"無效合約名稱 {contract_name}",
                contract_code=None
            )
            send_telegram(msg)
            return f"錯誤：無效合約名稱 {contract_name}", 400
        if qty <= 0:
            logger.error(format_unified_log(
                log_type="錯誤",
                order_id="未知",
                contract_name=contract_name,
                contract_code="未知",
                qty=qty,
                price=price,
                action=action,
                reason="無效數量",
                auto_fill=False
            ))
            msg = get_formatted_order_message(
                is_success=False,
                order_id="未知",
                contract_name=contract_name,
                qty=qty,
                price=price,
                octype="New",
                direction="Sell",
                order_type=order_type,
                price_type=price_type,
                is_manual=True,
                reason=f"無效數量 {qty}",
                contract_code=None
            )
            send_telegram(msg)
            return f"錯誤：無效數量 {qty}", 400
        if action not in ['BUY', 'SELL']:
            logger.error(format_unified_log(contract_name=contract_name, qty=qty, price=price, action=action, reason="無效動作"))
            msg = get_formatted_order_message(
                is_success=False,
                order_id="未知",
                contract_name=contract_name,
                qty=qty,
                price=price,
                octype="New",
                direction=action if action in ['Buy', 'Sell'] else "Sell",
                order_type=order_type,
                price_type=price_type,
                is_manual=True,
                reason=f"無效動作 {action}",
                contract_code=None
            )
            send_telegram(msg)
            return f"錯誤：無效動作 {action}", 400
        if order_type not in ['IOC', 'ROD']:
            logger.error(format_unified_log(contract_name=contract_name, qty=qty, price=price, action=action, order_type=order_type, reason="無效訂單類型"))
            msg = get_formatted_order_message(
                is_success=False,
                order_id="未知",
                contract_name=contract_name,
                qty=qty,
                price=price,
                octype="New",
                direction="Sell",
                order_type=order_type if order_type in ['IOC', 'ROD'] else "ROD",
                price_type=price_type,
                is_manual=True,
                reason=f"無效訂單類型 {order_type}",
                contract_code=None
            )
            send_telegram(msg)
            return f"錯誤：無效訂單類型 {order_type}", 400
        if price_type not in ['LMT', 'MKT']:
            logger.error(format_unified_log(contract_name=contract_name, qty=qty, price=price, action=action, price_type=price_type, reason="無效價格類型"))
            msg = get_formatted_order_message(
                is_success=False,
                order_id="未知",
                contract_name=contract_name,
                qty=qty,
                price=price,
                octype="New",
                direction="Sell",
                order_type=order_type,
                price_type=price_type if price_type in ['LMT', 'MKT'] else "LMT",
                is_manual=True,
                reason=f"無效價格類型 {price_type}",
                contract_code=None
            )
            send_telegram(msg)
            return f"錯誤：無效價格類型 {price_type}", 400
        if price <= 0:
            logger.error(format_unified_log(
                log_type="錯誤",
                order_id="未知",
                contract_name=contract_name,
                contract_code="未知",
                qty=qty,
                price=price,
                action=action,
                reason="無效價格",
                auto_fill=False
            ))
            msg = get_formatted_order_message(
                is_success=False,
                order_id="未知",
                contract_name=contract_name,
                qty=qty,
                price=price,
                octype="New",
                direction="Sell",
                order_type=order_type,
                price_type=price_type,
                is_manual=True,
                reason=f"無效價格 {price}",
                contract_code=None
            )
            send_telegram(msg)
            return f"錯誤：無效價格 {price}", 400
        contract = None
        if contract_name == "大台":
            contract = contract_txf
        elif contract_name == "小台":
            contract = contract_mxf
        elif contract_name == "微台":
            contract = contract_tmf
        if not contract:
            logger.error(format_unified_log(contract_name=contract_name, qty=qty, price=price, action=action, reason="合約未設置"))
            msg = get_formatted_order_message(
                is_success=False,
                order_id="未知",
                contract_name=contract_name,
                qty=qty,
                price=price,
                octype="New",
                direction="Sell",
                order_type=order_type,
                price_type=price_type,
                is_manual=True,
                reason=f"合約 {contract_name} 未設置",
                contract_code=None
            )
            send_telegram(msg)
            return f"錯誤：合約 {contract_name} 未設置", 400
        action_map = {'BUY': Action.Buy, 'SELL': Action.Sell}
        result, success, _, _, trade = place_order_for_contract(
            contract=contract,
            qty=qty,
            price=price,
            action=action_map[action],
            contract_name=contract_name,
            is_entry=is_entry,
            order_type=order_type,
            price_type=price_type,
            is_manual=True
        )
        if success:
            logger.info(f"下單提交成功：{result}")
            return "OK", 200
        else:
            logger.error(format_unified_log(contract_name=contract_name, qty=qty, price=price, action=action, reason=str(result)))
            return f"錯誤：{result}", 400
    except Exception as e:
        logger.exception("[manual_order] 下單處理錯誤 traceback：")
        logger.error(f"下單處理錯誤：{str(e)}")
        send_telegram(f"❌ 下單錯誤：{str(e)[:100]}")
        return f"錯誤：{str(e)}", 500

def process_signal(data):
    logger.debug(f"[process_signal] data={data}")
    global has_processed_delivery_exit, active_trades, contract_txf, contract_mxf, contract_tmf, use_next_month_contract, rollover_reminder_sent, rollover_notification_sent
    signal_id = data.get('tradeId')
    msg_type = data.get('type')
    direction = data.get('direction', '未知')
    signal_key = (signal_id, msg_type, direction)
    with global_lock:
        logger.debug(f"[process_signal] (with lock) recent_signals={state.recent_signals}")
        if signal_key in state.recent_signals:
            logger.warning("重複訊號，忽略")
            return
        state.recent_signals.add(signal_key)
        threading.Timer(2, lambda: state.recent_signals.discard(signal_key)).start()
    try:
        logger.info(f"接收到 Webhook 訊號：{data}")
        now = datetime.now()
        today = now.date()
        current_time = now.time()
        is_pre_delivery_trading_day = is_next_trading_day_delivery_day() and holiday_calendar.is_trading_day(today)
        after_15 = current_time >= datetime.strptime("15:00", "%H:%M").time()
        # 若符合條件且尚未切換，則切換合約並發送通知
        if is_pre_delivery_trading_day and after_15 and not use_next_month_contract:
            send_rollover_notification()
        msg_type = data.get('type')
        time_ms = int(data.get('time', 0)) / 1000
        time_str = (datetime.utcfromtimestamp(time_ms) + timedelta(hours=8)).strftime('%Y/%m/%d %H:%M')
        qty_txf = int(float(data.get('txf', 0)))
        qty_mxf = int(float(data.get('mxf', 0)))
        qty_tmf = int(float(data.get('tmf', 0)))
        alert_id = data.get('tradeId', 'unknown')
        price = float(data.get('price', 0))
        order_type = "IOC"  # 強制使用 IOC
        price_type = "MKT"  # 強制使用 市價單

        if price <= 0:
            logger.error(format_unified_log(
                log_type="錯誤",
                order_id="未知",
                contract_name="未知",
                contract_code="未知",
                qty=0,
                price=price,
                action="未知",
                reason="無效價格",
                auto_fill=False
            ))
            msg = get_formatted_order_message(
                is_success=False,
                order_id="未知",
                contract_name="未知",
                qty=0,
                price=price,
                octype="New",
                direction="Sell",
                order_type=order_type,
                price_type=price_type,
                is_manual=False,
                reason=f"價格 {price} 無效",
                contract_code=None
            )
            send_telegram(msg)
            return

        if order_type not in ['IOC', 'ROD']:
            logger.error(format_unified_log(contract_name="未知", qty=0, price=0, action="未知", order_type=order_type, reason="無效訂單類型"))
            msg = get_formatted_order_message(
                is_success=False,
                order_id="未知",
                contract_name="未知",
                qty=0,
                price=price,
                octype="New",
                direction="Sell",
                order_type=order_type,
                price_type=price_type,
                is_manual=False,
                reason=f"無效訂單類型 {order_type}",
                contract_code=None
            )
            send_telegram(msg)
            return
        if price_type not in ['LMT', 'MKT']:
            logger.error(format_unified_log(contract_name="未知", qty=0, price=0, action="未知", price_type=price_type, reason="無效價格類型"))
            msg = get_formatted_order_message(
                is_success=False,
                order_id="未知",
                contract_name="未知",
                qty=0,
                price=price,
                octype="New",
                direction="Sell",
                order_type=order_type,
                price_type=price_type,
                is_manual=False,
                reason=f"無效價格類型 {price_type}",
                contract_code=None
            )
            send_telegram(msg)
            return

        is_delivery = is_delivery_day()
        trade_category = 'normal'
        current_month = datetime.now().strftime('%Y%m')
        positions = api.list_positions(api.futopt_account)
        current_month_positions = [p for p in positions if p.code.endswith(current_month)]
        next_month_positions = [p for p in positions if not p.code.endswith(current_month)]
        should_rollover = False

        if is_delivery:
            trade_category = 'delivery_exit' if msg_type == 'exit' and not has_processed_delivery_exit else 'rollover_entry'
            if msg_type == 'exit' and not has_processed_delivery_exit:
                if current_month_positions:
                    should_rollover = True
                elif next_month_positions:
                    should_rollover = False
                else:
                    should_rollover = False
            elif msg_type == 'entry' and not current_month_positions and not next_month_positions:
                should_rollover = True

        is_delivery_exit = trade_category == 'delivery_exit'

        if msg_type == "entry":
            direction = data.get('direction', '未知')
            if direction not in ["開多", "開空"]:
                logger.error(format_unified_log(contract_name="未知", qty=0, price=price, action=direction, reason="無效進場動作"))
                msg = get_formatted_order_message(
                    is_success=False,
                    order_id="未知",
                    contract_name="未知",
                    qty=0,
                    price=price,
                    octype="New",
                    direction="Sell",
                    order_type=order_type,
                    price_type=price_type,
                    is_manual=False,
                    reason=f"無效進場動作 {direction}",
                    contract_code=None
                )
                send_telegram(msg)
                return
            action = Action.Buy if direction == "開多" else Action.Sell
            expected_action = Action.Buy if direction == "開多" else Action.Sell
            has_opposite = any(p.direction != expected_action and p.quantity != 0 for p in positions)

            if trade_category == 'rollover_entry' or (is_delivery and should_rollover):
                for code, var in [("TXF", "contract_txf"), ("MXF", "contract_mxf"), ("TMF", "contract_tmf")]:
                    next_contract = get_next_month_contract(code, current_month)
                    if next_contract:
                        globals()[var] = next_contract
                        logger.info(f"轉倉：{var} 更新為 {next_contract.code}，交割日：{next_contract.delivery_date}")

            contracts = [(contract_txf, qty_txf, "大台"), (contract_mxf, qty_mxf, "小台"), (contract_tmf, qty_tmf, "微台")]
            for contract, qty, name in contracts:
                if qty > 0:
                    logger.debug(f"[process_signal] 下單前 active_trades={active_trades}")
                    if has_opposite:
                        order_action = "開多" if action == Action.Buy else "開空"
                        contract_code = contract.code if contract else "未知"
                        logger.info(format_unified_log(
                            log_type="取消",
                            order_id="未知",
                            contract_name=name,
                            contract_code=contract_code,
                            qty=qty,
                            price=price,
                            action=order_action,
                            reason="存在相反持倉",
                            auto_fill=False
                        ))
                        msg = get_formatted_order_message(
                            is_success=False,
                            order_id="未知",
                            contract_name=name,
                            qty=qty,
                            price=price,
                            octype="New",
                            direction="Buy" if action == Action.Buy else "Sell",
                            order_type=order_type,
                            price_type=price_type,
                            is_manual=False,
                            reason="存在相反持倉",
                            contract_code=contract.code if contract else None
                        )
                        send_telegram(msg)
                        continue
                    # 使用統一的日誌格式
                    order_action = '開多' if action == Action.Buy else '開空'
                    contract_code = contract.code if contract else "未知"
                    logger.info(format_unified_log(
                        log_type="下單",
                        order_id="待生成",
                        contract_name=name,
                        contract_code=contract_code,
                        qty=qty,
                        price=price,
                        action=order_action,
                        reason=f"類型：{get_order_type_display(price_type, order_type)}",
                        auto_fill=False
                    ))
                    result, success, cost, deal_price, trade = place_order_for_contract(
                        contract=contract,
                        qty=qty,
                        price=price,
                        action=action,
                        contract_name=name,
                        is_entry=True,
                        is_rollover_entry=trade_category == 'rollover_entry',
                        order_type=order_type,
                        price_type=price_type,
                        is_manual=False
                    )
                    logger.info(result)
                    if success:
                        active_trades[contract_key_map[name]] = alert_id
                        logger.debug(f"[process_signal] 下單後 active_trades={active_trades}")

        elif msg_type == "exit":
            has_position = False
            position_txf = next((p for p in positions if p.code.startswith("TXF") and qty_txf > 0), None) if qty_txf > 0 else None
            position_mxf = next((p for p in positions if p.code.startswith("MXF") and qty_mxf > 0), None) if qty_mxf > 0 else None
            position_tmf = next((p for p in positions if p.code.startswith("TMF") and qty_tmf > 0), None) if qty_tmf > 0 else None
            if trade_category == 'delivery_exit':
                has_position = bool(current_month_positions)
            else:
                has_position = bool(position_txf or position_mxf or position_tmf)

            target_positions = current_month_positions if is_delivery_exit else [
                (contract_txf, qty_txf, "大台", position_txf),
                (contract_mxf, qty_mxf, "小台", position_mxf),
                (contract_tmf, qty_tmf, "微台", position_tmf)
            ]
            for item in target_positions:
                if is_delivery_exit:
                    position = item
                    contract = None
                    qty = abs(position.quantity)
                    name = None
                    if position.code.startswith('TXF'):
                        contract = next((c for c in api.Contracts.Futures.get("TXF") if c.code == position.code), None)
                        name = '大台'
                    elif position.code.startswith('MXF'):
                        contract = next((c for c in api.Contracts.Futures.get("MXF") if c.code == position.code), None)
                        name = '小台'
                    elif position.code.startswith('TMF'):
                        contract = next((c for c in api.Contracts.Futures.get("TMF") if c.code == position.code), None)
                        name = '微台'
                else:
                    contract, qty, name, position = item
                if qty > 0:
                    if not contract:
                        logger.info(format_unified_log(
                            log_type="取消",
                            order_id="未知",
                            contract_name=name,
                            contract_code="未設置",
                            qty=qty,
                            price=price,
                            action="平倉",
                            reason="合約未設置",
                            auto_fill=False
                        ))
                        msg = get_formatted_order_message(
                            is_success=False,
                            order_id="未知",
                            contract_name=name,
                            qty=qty,
                            price=price,
                            octype="Cover",
                            direction="Sell",
                            order_type=order_type,
                            price_type=price_type,
                            is_manual=False,
                            reason="合約未設置",
                            contract_code=None
                        )
                        send_telegram(msg)
                        continue
                    action = data.get('action')
                    if action == "反向平倉":
                        action = "平多" if data.get('direction') == "平多" else "平空" if data.get('direction') == "平空" else "未知"
                    action_display = get_action_display_by_rule('Cover', action)
                    if not position:
                        contract_code = contract.code if contract else "未知"
                        logger.info(format_unified_log(
                            log_type="取消",
                            order_id="未知",
                            contract_name=name,
                            contract_code=contract_code,
                            qty=qty,
                            price=price,
                            action="平倉",
                            reason="無對應持倉",
                            auto_fill=False
                        ))
                        msg = get_formatted_order_message(
                            is_success=False,
                            order_id="未知",
                            contract_name=name,
                            qty=qty,
                            price=price,
                            octype="Cover",
                            direction="Sell",
                            order_type=order_type,
                            price_type=price_type,
                            is_manual=False,
                            reason="無對應持倉",
                            contract_code=contract.code if contract else None
                        )
                        send_telegram(msg)
                        continue
                    if action == "未知":
                        contract_code = contract.code if contract else "未知"
                        logger.info(format_unified_log(
                            log_type="取消",
                            order_id="未知",
                            contract_name=name,
                            contract_code=contract_code,
                            qty=qty,
                            price=price,
                            action="平倉",
                            reason="無效平倉動作",
                            auto_fill=False
                        ))
                        msg = get_formatted_order_message(
                            is_success=False,
                            order_id="未知",
                            contract_name=name,
                            qty=qty,
                            price=price,
                            octype="Cover",
                            direction="Sell",
                            order_type=order_type,
                            price_type=price_type,
                            is_manual=False,
                            reason=f"無效平倉動作 {action}",
                            contract_code=contract.code if contract else None
                        )
                        send_telegram(msg)
                        continue
                    if (position.direction == Action.Buy and action == "平多") or \
                       (position.direction == Action.Sell and action == "平空") or \
                       is_delivery_exit:
                        # 使用統一的日誌格式
                        contract_code = contract.code if contract else "未知"
                        logger.info(format_unified_log(
                            log_type="下單",
                            order_id="待生成",
                            contract_name=name,
                            contract_code=contract_code,
                            qty=qty,
                            price=price,
                            action="平倉",
                            reason=f"類型：{get_order_type_display(price_type, order_type)}",
                            auto_fill=False
                        ))
                        result, success, _, deal_price, trade = place_order_for_contract(
                            contract=contract,
                            qty=qty,
                            price=price,
                            action=None,
                            contract_name=name,
                            is_entry=False,
                            position=position,
                            order_type=order_type,
                            price_type=price_type,
                            is_manual=False
                        )
                        logger.info(result)
                        if success:
                            active_trades[contract_key_map[name]] = None
                    else:
                        contract_code = contract.code if contract else "未知"
                        logger.info(format_unified_log(
                            log_type="取消",
                            order_id="未知",
                            contract_name=name,
                            contract_code=contract_code,
                            qty=qty,
                            price=price,
                            action="平倉",
                            reason="存在相反持倉",
                            auto_fill=False
                        ))
                        msg = get_formatted_order_message(
                            is_success=False,
                            order_id="未知",
                            contract_name=name,
                            qty=qty,
                            price=price,
                            octype="Cover",
                            direction="Sell" if action == "平多" else "Buy",
                            order_type=order_type,
                            price_type=price_type,
                            is_manual=False,
                            reason="存在相反持倉",
                            contract_code=contract.code if contract else None
                        )
                        send_telegram(msg)
            if not has_position:
                qty = qty_txf or qty_mxf or qty_tmf
                contract_name = "大台" if qty_txf > 0 else "小台" if qty_mxf > 0 else "微台" if qty_tmf > 0 else "未知"
                logger.info(format_unified_log(
                    log_type="取消",
                    order_id="未知",
                    contract_name=contract_name,
                    contract_code="未知",
                    qty=qty,
                    price=price,
                    action="平倉",
                    reason="無對應持倉",
                    auto_fill=False
                ))
                msg = get_formatted_order_message(
                    is_success=False,
                    order_id="未知",
                    contract_name=contract_name,
                    qty=qty,
                    price=price,
                    octype="Cover",
                    direction="Sell",
                    order_type=order_type,
                    price_type=price_type,
                    is_manual=False,
                    reason="無對應持倉",
                    contract_code=None
                )
                send_telegram(msg)

            if is_delivery_exit and should_rollover:
                has_processed_delivery_exit = True
                for code, var in [("TXF", "contract_txf"), ("MXF", "contract_mxf"), ("TMF", "contract_tmf")]:
                    next_contract = get_next_month_contract(code, current_month)
                    if next_contract:
                        globals()[var] = next_contract
                        logger.info(f"交割日平倉後轉倉：{var} 更新為 {next_contract.code}，交割日：{next_contract.delivery_date}")
            elif is_delivery_exit:
                has_processed_delivery_exit = True
                logger.info("交割日平倉，持次月合約，無需轉倉")

        else:
            logger.error(format_unified_log(contract_name="未知", qty=0, price=0, action=msg_type, reason="無效訊號類型"))
            msg = get_formatted_order_message(
                is_success=False,
                order_id="未知",
                contract_name="未知",
                qty=0,
                price=0,
                octype="New",
                direction="Sell",
                order_type="ROD",
                price_type="LMT",
                is_manual=False,
                reason=f"無效訊號類型 {msg_type}",
                contract_code=None
            )
            send_telegram(msg)
            return

    except Exception as e:
        logger.exception("[process_signal] 處理 Webhook 訊號失敗 traceback：")
        logger.error(f"處理 Webhook 訊號失敗：{str(e)}")
        msg = get_formatted_order_message(
            is_success=False,
            order_id="未知",
            contract_name="未知",
            qty=0,
            price=0,
            octype="New",
            direction="Sell",
            order_type="ROD",
            price_type="LMT",
            is_manual=False,
            reason=str(e)[:100],
            contract_code=None
        )
        send_telegram(msg)
def order_callback(state, deal, order=None):
    try:
        # 使用中文格式化的回調訊息
        formatted_msg = format_callback_message(state, deal, order)
        logger.info(formatted_msg)
        
        # 保留詳細的原始數據日誌（僅在DEBUG模式下顯示）
        if DEBUG_MODE:
            logger.debug(f"原始回調數據：狀態={state}, 成交數據={deal}, 訂單數據={order}")
        
        # 成交回調和訂單回調的數據結構不同，需要分別處理
        if str(state) == 'OrderState.FuturesDeal':
            # 成交回調：使用 deal 的直接欄位
            order_id = deal.get('trade_id', deal.get('order_id', '未知')).strip()
            contract_code = deal.get('code', '')
        else:
            # 訂單回調：使用 order 結構
            order_id = deal.get('order', {}).get('id', '未知').strip()
            contract_code = deal.get('contract', {}).get('code', '')
        
        contract_name = get_contract_name(contract_code)
        
        # 優先使用全域合約物件的完整代碼，而不是回調中的基礎代碼
        if contract_name == "大台" and 'contract_txf' in globals() and contract_txf:
            contract_code = contract_txf.code
        elif contract_name == "小台" and 'contract_mxf' in globals() and contract_mxf:
            contract_code = contract_mxf.code
        elif contract_name == "微台" and 'contract_tmf' in globals() and contract_tmf:
            contract_code = contract_tmf.code
        octype_info = order_octype_map.get(order_id)
        if octype_info is None:
            # 從交易記錄推斷
            today = datetime.now().strftime("%Y%m%d")
            filename = f"{LOG_DIR}/trades_{today}.json"
            oc_type, direction, order_type, price_type, is_manual = None, None, None, None, True
            if os.path.exists(filename):
                try:
                    with open(filename, 'r') as f:
                        trades = json.load(f)
                    for trade in trades:
                        if trade.get('deal_order_id') == order_id and trade.get('type') == 'order':
                            raw_order = trade.get('raw_data', {}).get('order', {})
                            oc_type = raw_order.get('oc_type', 'New')
                            direction = raw_order.get('action', deal.get('order', {}).get('action', 'Sell'))
                            order_type = raw_order.get('order_type', deal.get('order', {}).get('order_type', 'ROD'))
                            price_type = raw_order.get('price_type', 'LMT')
                            is_manual = trade.get('is_manual', True)
                            break
                except Exception as e:
                    handle_exception("讀取交易記錄失敗", e)
            octype_info = {
                'octype': oc_type or ('New' if deal.get('order', {}).get('oc_type', 'New') == 'Auto' else deal.get('order', {}).get('oc_type', 'New')),
                'direction': direction or deal.get('order', {}).get('action', 'Sell'),
                'contract_name': contract_name,
                'order_type': order_type or deal.get('order', {}).get('order_type', 'ROD'),
                'price_type': price_type or deal.get('order', {}).get('price_type', 'LMT'),
                'is_manual': is_manual
            }
            # 自動推斷訂單資訊（靜默處理，這是正常的備援機制）
        
        # 簡化的訂單回調處理日誌（中文顯示）
        if DEBUG_MODE:
            logger.debug(f"訂單回調處理：單號={order_id}, 推斷資訊={octype_info}")
        octype = octype_info['octype']
        direction = octype_info['direction']
        order_type = octype_info['order_type']
        price_type = octype_info['price_type']
        is_manual = octype_info.get('is_manual', False)
        action_display = get_action_display_by_rule(octype, direction)
        order_type_display = get_order_type_display(price_type, order_type)
        qty = deal.get('order', {}).get('quantity', 0)
        op_code = deal.get('operation', {}).get('op_code', '00')
        op_msg = deal.get('operation', {}).get('op_msg', '')
        op_type = deal.get('operation', {}).get('op_type', '')

        if str(state) == 'OrderState.FuturesDeal' and deal.get('code') and deal.get('quantity'):
            handle_futures_deal(deal)
        elif str(state) in ['OrderState.Submitted', 'OrderState.FuturesOrder']:
            if op_type == 'Cancel':
                if order_type == 'IOC' and not op_msg:
                    cancel_reason = "價格未滿足"
                else:
                    cancel_reason = config.OP_MSG_TRANSLATIONS.get(op_msg, op_msg or '未成交或系統取消')
                # 移除舊格式日誌，統一使用 get_formatted_order_message 格式
                save_trade({
                    'type': 'cancel',
                    'trade_category': 'normal',
                    'raw_data': deal,
                    'deal_order_id': order_id,
                    'contract_name': contract_name,
                    'timestamp': datetime.now().isoformat(),
                    'is_manual': is_manual
                })
                price_value = order.get('price', 0.0) if order else deal.get('order', {}).get('price', 0.0)
                msg = get_formatted_order_message(
                    is_success=False,
                    order_id=order_id,
                    contract_name=contract_name,
                    qty=qty,
                    price=price_value,
                    octype=octype,
                    direction=direction,
                    order_type=order_type,
                    price_type=price_type,
                    is_manual=is_manual,
                    reason=f"訂單取消 - {cancel_reason}",
                    contract_code=contract_code
                )
                notify_and_log(msg, level="error", notify=True, log=False)
                with global_lock:
                    order_octype_map.pop(order_id, None)
            elif op_code != '00':
                fail_reason = config.OP_MSG_TRANSLATIONS.get(op_msg, op_msg or '未知錯誤')
                # 移除舊格式日誌，統一使用 get_formatted_order_message 格式
                save_trade({
                    'type': 'order',
                    'trade_category': 'normal',
                    'raw_data': deal,
                    'deal_order_id': order_id,
                    'contract_name': contract_name,
                    'timestamp': datetime.now().isoformat(),
                    'is_manual': is_manual
                })
                price_value = order.get('price', 0.0) if order else deal.get('order', {}).get('price', 0.0)
                msg = get_formatted_order_message(
                    is_success=False,
                    order_id=order_id,
                    contract_name=contract_name,
                    qty=qty,
                    price=price_value,
                    octype=octype,
                    direction=direction,
                    order_type=order_type,
                    price_type=price_type,
                    is_manual=is_manual,
                    reason=fail_reason,
                    contract_code=contract_code
                )
                notify_and_log(msg, level="error", notify=True, log=False)
                with global_lock:
                    order_octype_map.pop(order_id, None)
            else:
                # 移除訂單提交確認的日誌訊息，因為使用者反映不需要
                # state_zh = translate_callback_content(str(state))
                # logger.info(f"訂單提交確認：單號={order_id} | 合約=[{contract_name}] | 狀態={state_zh}")
                save_trade({
                    'type': 'order',
                    'trade_category': 'normal',
                    'raw_data': deal,  # 保存完整的deal對象，包含order信息
                    'deal_order_id': order_id,
                    'contract_name': contract_name,
                    'timestamp': datetime.now().isoformat(),
                    'is_manual': is_manual
                })
                price_value = order.get('price', 0.0) if order else deal.get('order', {}).get('price', 0.0)
                # 嘗試從回調資料中獲取交割月份，轉換為交割日期
                delivery_date = None
                try:
                    delivery_month = None
                    if order and order.get('delivery_month'):
                        delivery_month = order.get('delivery_month')
                    elif deal and deal.get('contract', {}).get('delivery_month'):
                        delivery_month = deal.get('contract', {}).get('delivery_month')
                    
                    if delivery_month and len(delivery_month) == 6:
                        year = int(delivery_month[:4])
                        month = int(delivery_month[4:6])
                        # 簡化為月中作為交割日期
                        delivery_date = f"{year}/{month:02d}/16"
                except:
                    pass
                
                msg = get_formatted_order_message(
                    is_success=True,
                    order_id=order_id,
                    contract_name=contract_name,
                    qty=qty,
                    price=price_value,
                    octype=octype,
                    direction=direction,
                    order_type=order_type,
                    price_type=price_type,
                    is_manual=is_manual,
                    contract_code=contract_code,
                    delivery_date=delivery_date
                )
                notify_and_log(msg, level="info", notify=True, log=False)
        elif str(state) in ['OrderState.Filled', 'OrderState.Failed', 'OrderState.Cancelled']:
            reason = config.OP_MSG_TRANSLATIONS.get(op_msg, op_msg or '未知錯誤')
            state_zh = translate_callback_content(str(state))
            logger.info(f"訂單最終狀態：單號={order_id} | 合約=[{contract_name}] | 狀態={state_zh} | 原因={reason}")
            save_trade({
                'type': 'order',
                'trade_category': 'normal',
                'raw_data': {
                    'order_id': order.get('id', '未知') if order else order_id,
                    'code': order.get('code', '') if order else deal.get('contract', {}).get('code', ''),
                    'action': order.get('action', '') if order else deal.get('order', {}).get('action', ''),
                    'quantity': order.get('quantity', 0) if order else qty,
                    'price': order.get('price', 0.0) if order else deal.get('order', {}).get('price', 0.0),
                    'delivery_month': order.get('delivery_month', '') if order else deal.get('contract', {}).get('delivery_month', ''),
                    'status': str(state),
                    'ts': time.time()
                },
                'deal_order_id': order.get('id', '未知') if order else order_id,
                'contract_name': contract_name,
                'timestamp': datetime.now().isoformat(),
                'is_manual': is_manual
            })
            if str(state) in ['OrderState.Failed', 'OrderState.Cancelled']:
                price_value = order.get('price', 0.0) if order else deal.get('order', {}).get('price', 0.0)
                # 嘗試從回調資料中獲取交割月份，轉換為交割日期
                delivery_date = None
                try:
                    delivery_month = None
                    if order and order.get('delivery_month'):
                        delivery_month = order.get('delivery_month')
                    elif deal and deal.get('contract', {}).get('delivery_month'):
                        delivery_month = deal.get('contract', {}).get('delivery_month')
                    
                    if delivery_month and len(delivery_month) == 6:
                        year = int(delivery_month[:4])
                        month = int(delivery_month[4:6])
                        # 簡化為月中作為交割日期
                        delivery_date = f"{year}/{month:02d}/16"
                except:
                    pass
                
                msg = get_formatted_order_message(
                    is_success=False,
                    order_id=order_id,
                    contract_name=contract_name,
                    qty=qty,
                    price=price_value,
                    octype=octype,
                    direction=direction,
                    order_type=order_type,
                    price_type=price_type,
                    is_manual=is_manual,
                    reason=reason,
                    contract_code=contract_code,
                    delivery_date=delivery_date
                )
                notify_and_log(msg, level="error", notify=True, log=False)
            with global_lock:
                order_octype_map.pop(order_id, None)
    except Exception as e:
        handle_exception("處理訂單回調失敗", e)

def setup_event_handlers():
    api.set_order_callback(order_callback)

def save_trade(data):
    try:
        today = datetime.now().strftime("%Y%m%d")
        filename = f"{LOG_DIR}/trades_{today}.json"
        os.makedirs(LOG_DIR, exist_ok=True)
        try:
            trades = json.load(open(filename, 'r')) if os.path.exists(filename) else []
        except json.JSONDecodeError:
            logger.error(f"交易記錄檔案 {filename} 格式錯誤，重置為空列表")
            send_telegram(f"❌ 交易記錄檔案 {filename} 格式錯誤，已重置")
            trades = []
        data['timestamp'] = datetime.now().isoformat()
        trades.append(data)
        with open(filename, 'w') as f:
            json.dump(trades, f, indent=2)
    except Exception as e:
        logger.error(f"儲存交易記錄失敗：{str(e)}")
        send_telegram(f"❌ 儲存交易記錄失敗：{str(e)[:100]}")

def reset_trade_counter():
    # 使用 state 物件的重置方法
    state.reset_daily_flags()
    state.clear_trade_data()

def clear_old_order_octype():
    # 使用 state 物件管理訂單映射
    state.order_octype_map.clear()

def generate_report(period='', date_str=None, trades=None):
    try:
        if period == 'daily':
            yesterday = (datetime.now() - timedelta(days=1)).date()
            date_str = yesterday.strftime('%Y/%m/%d')
            filename = f"{LOG_DIR}/trades_{yesterday.strftime('%Y%m%d')}.json"
            if not os.path.exists(filename):
                logger.info(f"{date_str} 無交易記錄，跳過日報")
                send_telegram(f"📊 交易統計（{date_str}）\n無交易記錄")
                return
            try:
                with open(filename, 'r') as f:
                    trades = json.load(f)
                # 移除此日誌，因為不需要顯示
                # logger.info(f"交易記錄：{filename}, 總數={len(trades)}, deal類型={len([t for t in trades if t['type'] == 'deal'])}")
            except json.JSONDecodeError:
                logger.error(f"交易記錄檔案 {filename} 格式錯誤，無法生成日報")
                send_telegram(f"❌ 交易記錄檔案 {filename} 格式錯誤，無法生成日報")
                return
            report_file = f"{LOG_DIR}/report_{yesterday.strftime('%Y%m%d')}.xlsx"
            report_title = f"交易報表_{date_str.replace('/', '-')}"
        elif period == 'monthly':
            year = datetime.now().year
            month = datetime.now().month
            date_str = f"{year}/{month:02d}"
            report_file = f"{LOG_DIR}/report_{year}_{month}月份.xlsx"
            report_title = f"交易報表_{year}_{month}月份"
            if not trades:
                logger.error("月報缺少交易數據")
                return

        total = len([t for t in trades if t['type'] == 'deal'])  # 只計成交
        deal_count = total
        order_count = len([t for t in trades if t['type'] == 'order' and t['raw_data']['operation']['op_code'] != '00'])
        cancel_count = len([t for t in trades if t['type'] == 'cancel'])

        paired_trades = []
        open_trades = {}
        
        # 建立訂單ID到訂單信息的映射，用於後續匹配
        order_info_map = {}
        for trade in trades:
            if trade['type'] == 'order':
                order_id = trade.get('deal_order_id', '')
                if order_id:
                    raw_order = trade.get('raw_data', {}).get('order', {})
                    order_info_map[order_id] = {
                        'seqno': raw_order.get('seqno', ''),
                        'ordno': raw_order.get('ordno', ''),
                        'exchange_seq': '',  # 訂單時還沒有exchange_seq
                        'oc_type': raw_order.get('oc_type', 'New'),
                        'action': raw_order.get('action', 'Sell'),
                        'order_type': raw_order.get('order_type', 'ROD'),
                        'price_type': raw_order.get('price_type', 'LMT'),
                        'is_manual': trade.get('is_manual', False)
                    }
        
        for trade in trades:
            if trade['type'] != 'deal':
                continue
            raw_data = trade.get('raw_data', {})
            required_fields = ['trade_id', 'action', 'price', 'quantity', 'ts']
            missing_fields = [f for f in required_fields if f not in raw_data]
            if missing_fields:
                # 移除此日誌，因為不需要顯示
                # logger.warning(f"交易記錄缺少欄位：{missing_fields}, trade_id={raw_data.get('trade_id', '未知')}, 跳過")
                continue
            trade_id = raw_data['trade_id']
            
            # 獲取訂單信息，優先從order_info_map獲取，否則從order_octype_map獲取
            order_info = order_info_map.get(trade_id, {})
            octype_info = order_octype_map.get(trade_id)
            
            if octype_info is None and not order_info:
                # 移除此日誌，因為不需要顯示
                # logger.warning(f"未找到 trade_id {trade_id} 的 order_octype_map 記錄，使用推斷值")
                octype_info = {
                    'octype': raw_data.get('oc_type', 'New'),
                    'direction': raw_data.get('action', 'Sell'),
                    'contract_name': get_contract_name(raw_data.get('code', '')),
                    'order_type': 'ROD',
                    'price_type': 'LMT',
                    'is_manual': trade.get('is_manual', False)
                }
            
            # 合併訂單信息和octype信息
            if order_info:
                octype = order_info.get('oc_type', 'New')
                direction = order_info.get('action', 'Sell')
                order_type = order_info.get('order_type', 'ROD')
                price_type = order_info.get('price_type', 'LMT')
                is_manual = order_info.get('is_manual', False)
            else:
                octype = octype_info['octype']
                direction = octype_info['direction']
                order_type = octype_info['order_type']
                price_type = octype_info['price_type']
                is_manual = octype_info.get('is_manual', trade.get('is_manual', False))
            
            contract_name = get_contract_name(raw_data.get('code', ''))
            octype_display = f"{'手動' if is_manual else '自動'}{'開倉' if octype == 'New' else '平倉'}"
            action_display = get_action_display_by_rule(octype, direction)
            order_type_display = get_order_type_display(price_type, order_type)
            
            # 獲取成交相關的字段
            exchange_seq = raw_data.get('exchange_seq', '') or order_info.get('exchange_seq', '')
            ordno = raw_data.get('ordno', '') or order_info.get('ordno', '')
            
            if octype == 'New':
                open_trades[trade_id] = {
                    'buy_price': raw_data.get('price', 0.0),
                    'quantity': raw_data.get('quantity', 0),
                    'contract_name': contract_name,
                    'code': raw_data.get('code', ''),
                    'delivery_month': raw_data.get('delivery_month', '') or get_delivery_month_by_code(raw_data.get('code', '')),
                    'direction': direction,
                    'deal_time': datetime.fromtimestamp(raw_data.get('ts', 0)).strftime('%Y/%m/%d %H:%M'),
                    'exchange_seq': exchange_seq,
                    'ordno': ordno,
                    'order_type': order_type_display,
                    'is_manual': is_manual
                }
            elif octype == 'Cover':
                for open_trade_id, open_trade in list(open_trades.items()):
                    if (open_trade['contract_name'] == contract_name and
                        open_trade['quantity'] == raw_data.get('quantity', 0) and
                        open_trade['direction'] == direction):
                        buy_price = open_trade['buy_price']
                        sell_price = raw_data.get('price', 0.0)
                        quantity = raw_data.get('quantity', 0)
                        point_value = POINT_VALUES.get(contract_name, 200)
                        profit_loss = (sell_price - buy_price) * quantity * point_value if direction == 'Buy' else (buy_price - sell_price) * quantity * point_value
                        paired_trades.append({
                            'sell_trade_id': trade_id,
                            'action': action_display,
                            'octype': octype_display,
                            'order_type': order_type_display,
                            'buy_price': buy_price,
                            'sell_price': sell_price,
                            'quantity': quantity,
                            'contract_code': open_trade['code'],
                            'delivery_month': open_trade['delivery_month'] or get_delivery_month_by_code(open_trade['code']),
                            'profit_loss': profit_loss,
                            'deal_time': datetime.fromtimestamp(raw_data.get('ts', 0)).strftime('%Y/%m/%d %H:%M'),
                            'time_ms': raw_data.get('ts', 0) if raw_data.get('ts') else 0,
                            'exchange_seq': exchange_seq or open_trade.get('exchange_seq', '') or "未知",
                            'ordno': ordno or open_trade.get('ordno', '') or "未知",
                            'contract_name': contract_name
                        })
                        del open_trades[open_trade_id]
                        break

        # 移除此日誌，因為不需要顯示
        # logger.info(f"配對交易結果：open_trades={len(open_trades)}, paired_trades={len(paired_trades)}")

        if not paired_trades:
            # 移除此日誌，因為不需要顯示
            # logger.info(f"{date_str} 無配對交易，生成報表時跳過交易明細")
            df_trades = pd.DataFrame()
        else:
            paired_trades.sort(key=lambda x: x['time_ms'])
            df_trades = pd.DataFrame(paired_trades)
            if 'time_ms' in df_trades.columns:
                df_trades = df_trades.drop(columns=['time_ms'])
            else:
                # 移除此日誌，因為不需要顯示
                # logger.warning(f"df_trades 中缺少 time_ms 欄位，跳過刪除")
                pass

        if period == 'daily':
            update_balance_cache(force_update=True)
            margin_data = api.margin()
            
            # 簡化為一條日誌記錄
            logger.info(f"已生成每日交易統計（{date_str}）")
            
            # 不再記錄詳細統計到日誌，只發送到Telegram
            balance_msg = (
                f"═════ 帳戶狀態 ═════\n"
                f"權益總值：{margin_data.equity:.0f}\n"
                f"權益總額：{margin_data.equity_amount:.0f}\n"
                f"今日餘額：{margin_data.today_balance:.0f}\n"
                f"昨日餘額：{margin_data.yesterday_balance:.0f}\n"
                f"可用保證金：{margin_data.available_margin:.0f}\n"
                f"原始保證金：{margin_data.initial_margin:.0f}\n"
                f"維持保證金：{margin_data.maintenance_margin:.0f}\n"
                f"風險指標：{margin_data.risk_indicator:.1f}%\n"
                f"手續費：{margin_data.fee:.1f}\n"
                f"期交稅：{margin_data.tax:.1f}\n"
                f"本日平倉損益：{margin_data.future_settle_profitloss:.1f}\n"
            )
            msg = (
                f"📊 交易統計（{date_str}）\n"
                f"═════ 總覽 ═════\n"
                f"成交數量：{deal_count} 筆\n"
                f"訂單提交：{len([t for t in trades if t['type'] == 'order'])} 筆\n"
                f"未成交單：{order_count} 筆\n"
                f"取消單量：{cancel_count} 筆\n"
                f"{balance_msg}"
                f"═════ 交易明細 ═════\n"
            )
            if not paired_trades:
                msg += "❌ 無平倉交易\n"
            else:
                for trade in paired_trades:
                    msg += (
                        f"單號：{trade['sell_trade_id']}｜"
                        f"類型：{trade['octype']}｜"
                        f"動作：{trade['action']}｜"
                        f"部位：{trade['contract_name']}｜"
                        f"數量：{trade['quantity']}口｜"
                        f"均價：{trade['sell_price']:.0f}\n"
                        f"🧧 盈虧：{trade['profit_loss']:.0f} TWD\n\n"
                    )
            msg += "═════ 持倉狀態 ═════\n"
            try:
                positions = api.list_positions(api.futopt_account)
                if positions:
                    for pos in positions:
                        contract_type = "大台" if pos.code.startswith("TXF") else "小台" if pos.code.startswith("MXF") else "微台"
                        direction = "多單" if pos.direction == Action.Buy else "空單"
                        trade_id = next((tid for tid, t in open_trades.items() if t['code'] == pos.code and t['quantity'] == abs(pos.quantity) and t['direction'] == ('Buy' if pos.direction == Action.Buy else 'Sell')), '未知')
                        order_type_display = open_trades.get(trade_id, {}).get('order_type', '限價單（IOC）')
                        is_manual = open_trades.get(trade_id, {}).get('is_manual', False)
                        octype_display = f"{'手動' if is_manual else '自動'}開倉"
                        try:
                            # 先嘗試從 contract 物件獲取
                            delivery_month = getattr(pos.contract, 'delivery_month', None)
                            if not delivery_month:
                                # 如果無法取得，從code解析
                                delivery_month = get_delivery_month_by_code(pos.code)
                        except AttributeError:
                            # 如果都失敗，使用解析方法
                            delivery_month = get_delivery_month_by_code(pos.code)
                            if delivery_month == "未知":
                                logger.warning(f"無法獲取持倉 {pos.code} 的交割月份")
                        msg += (
                            f"［{contract_type}］｜"
                            f"動作：{direction}｜"
                            f"數量：{abs(pos.quantity)}口｜"
                            f"均價：{pos.price:.0f}\n"
                            f"📋 未實現盈虧：{pos.pnl:.0f} TWD\n"
                        )
                else:
                    msg += "無持倉部位\n"
            except Exception as e:
                handle_exception("獲取持倉資訊失敗", e)
                msg += "無法獲取持倉資訊\n"
                        # 只發送到Telegram，不記錄到日誌
            send_telegram(msg)

        update_balance_cache(force_update=True)
        margin_data = api.margin()
        df = pd.DataFrame([{
            "交易日期": date_str,
            "交易總數": total,
            "成交單量": deal_count,
            "未成單量": order_count,
            "取消單量": cancel_count,
            "權益總值": margin_data.equity or 0,
            "權益總額": margin_data.equity_amount or 0,
            "今日餘額": margin_data.today_balance or 0,
            "昨日餘額": margin_data.yesterday_balance or 0,
            "可用保證金": margin_data.available_margin or 0,
            "原始保證金": margin_data.initial_margin or 0,
            "維持保證金": margin_data.maintenance_margin or 0,
            "風險指標": margin_data.risk_indicator or 0,
            "手續費": margin_data.fee or 0,
            "期交稅": margin_data.tax or 0,
            "本日平倉損益": margin_data.future_settle_profitloss or 0
        }])

        wb = Workbook()
        ws_main = wb.active
        ws_main.title = report_title
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        ws_main.append(["總覽區"])
        ws_main.merge_cells('A1:K1')
        ws_main['A1'].fill = header_fill
        ws_main['A1'].font = header_font
        ws_main['A1'].alignment = Alignment(horizontal='center')
        ws_main.append(["交易日期", "交易總數", "成交單量", "未成單量", "取消單量"])
        ws_main.append([df["交易日期"][0], df["交易總數"][0], df["成交單量"][0], df["未成單量"][0], df["取消單量"][0]])
        ws_main.append([])

        ws_main.append(["財務區"])
        ws_main.merge_cells('A5:K5')
        ws_main['A5'].fill = header_fill
        ws_main['A5'].font = header_font
        ws_main['A5'].alignment = Alignment(horizontal='center')
        ws_main.append([
            "權益總值", "權益總額", "今日餘額", "昨日餘額", "可用保證金",
            "原始保證金", "維持保證金", "風險指標", "手續費", "期交稅", "本日平倉損益"
        ])
        ws_main.append([
            df["權益總值"][0], df["權益總額"][0], df["今日餘額"][0], df["昨日餘額"][0], df["可用保證金"][0],
            df["原始保證金"][0], df["維持保證金"][0], f"{df['風險指標'][0]:.1f}%", df["手續費"][0], df["期交稅"][0], df["本日平倉損益"][0]
        ])
        ws_main.append([])

        ws_main.append(["交易明細區"])
        ws_main.merge_cells('A9:K9')
        ws_main['A9'].fill = header_fill
        ws_main['A9'].font = header_font
        ws_main['A9'].alignment = Alignment(horizontal='center')
        ws_main.append([
            "交易所序號", "委託單號", "選用合約", "訂單類型", "成交單號", "成交類型", "成交動作", "成交部位", "成交數量", "成交均價", "盈虧狀態"
        ])
        row_start = 10
        
        if not df_trades.empty:
            for _, row in df_trades.iterrows():
                display_contract_code = row.get("contract_code", "")
                delivery_date = row.get("delivery_month", "")
                
                ws_main.append([
                    row.get("exchange_seq", "") or "未知",
                    row.get("ordno", "") or "未知",
                    format_contract_display(display_contract_code, delivery_date),
                    row.get("order_type", "") or "未知",
                    row.get("sell_trade_id", "") or "未知",
                    row.get("octype", "") or "未知",
                    row.get("action", "") or "未知",
                    row.get("contract_name", "") or "未知",
                    row.get("quantity", 0),
                    row.get("sell_price", 0),
                    f"{row.get('profit_loss', 0):.0f} TWD"
                ])
        row_end = row_start + len(df_trades) - 1 if not df_trades.empty else row_start

        ws_main.append([])
        ws_main.append(["持倉狀態區"])
        ws_main.merge_cells(f'A{row_end + 2}:K{row_end + 2}')
        ws_main[f'A{row_end + 2}'].fill = header_fill
        ws_main[f'A{row_end + 2}'].font = header_font
        ws_main[f'A{row_end + 2}'].alignment = Alignment(horizontal='center')
        ws_main.append([
            "交易所序號", "委託單號", "選用合約", "訂單類型", "成交單號", "成交類型", "成交動作", "成交部位", "成交數量", "成交均價", "盈虧狀態"
        ])
        positions = api.list_positions(api.futopt_account)
        if positions:
            for pos in positions:
                trade_id = next((tid for tid, t in open_trades.items() if t['code'] == pos.code and t['quantity'] == abs(pos.quantity) and t['direction'] == ('Buy' if pos.direction == Action.Buy else 'Sell')), '未知')
                order_type_display = open_trades.get(trade_id, {}).get('order_type', '限價單（IOC）')
                is_manual = open_trades.get(trade_id, {}).get('is_manual', False)
                direction = "多單" if pos.direction == Action.Buy else "空單"
                contract_type = "大台" if pos.code.startswith("TXF") else "小台" if pos.code.startswith("MXF") else "微台"
                octype_display = f"{'手動' if is_manual else '自動'}開倉"
                # 獲取完整合約代碼和交割日期信息
                api_contract_code = pos.code
                delivery_date = ""
                display_contract_code = api_contract_code
                
                try:
                    # 先嘗試從 contract 物件獲取
                    delivery_date = getattr(pos.contract, 'delivery_month', None)
                    if not delivery_date:
                        # 如果無法取得，從code解析
                        delivery_date = get_delivery_month_by_code(pos.code)
                except AttributeError:
                    delivery_date = get_delivery_month_by_code(pos.code)
                    if delivery_date == "未知":
                        logger.warning(f"無法獲取持倉 {pos.code} 的交割日期")
                
                # 從當天交易記錄JSON檔案中查找exchange_seq、ordno、trade_id
                exchange_seq = "未知"
                ordno = "未知"
                actual_trade_id = "未知"
                
                # 優先從open_trades中獲取信息
                if trade_id != '未知' and trade_id in open_trades:
                    open_trade = open_trades[trade_id]
                    exchange_seq = open_trade.get('exchange_seq', '') or "未知"
                    ordno = open_trade.get('ordno', '') or "未知"
                    actual_trade_id = trade_id
                else:
                    # 如果open_trades中沒有，則從當天交易記錄中查找
                    today = datetime.now().strftime("%Y%m%d")
                    filename = f"{LOG_DIR}/trades_{today}.json"
                    if os.path.exists(filename):
                        try:
                            with open(filename, 'r') as f:
                                today_trades = json.load(f)
                            # 尋找匹配的成交記錄
                            for trade_record in today_trades:
                                if (trade_record.get('type') == 'deal' and 
                                    trade_record.get('raw_data', {}).get('code') == pos.code and
                                    trade_record.get('raw_data', {}).get('quantity') == abs(pos.quantity)):
                                    raw_data = trade_record.get('raw_data', {})
                                    exchange_seq = raw_data.get('exchange_seq', '') or "未知"
                                    ordno = raw_data.get('ordno', '') or "未知"
                                    actual_trade_id = raw_data.get('trade_id', '') or "未知"
                                    break
                        except Exception:
                            pass  # 如果讀取失敗，使用預設值"未知"
                
                ws_main.append([
                    exchange_seq or "未知",
                    ordno or "未知", 
                    format_contract_display(display_contract_code, delivery_date),
                    order_type_display,
                    actual_trade_id,
                    octype_display,
                    f"{direction}",
                    contract_type,
                    abs(pos.quantity),
                    pos.price,
                    f"{pos.pnl:.0f} TWD"
                ])

        max_row = ws_main.max_row
        for row in ws_main.iter_rows(min_row=2, max_row=7, min_col=1, max_col=11):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
        for row in ws_main.iter_rows(min_row=10, max_row=max_row, min_col=1, max_col=11):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='center')

        column_widths = {
            'A': 19, 'B': 19, 'C': 19, 'D': 19, 'E': 19, 'F': 19, 'G': 19, 'H': 19, 'I': 19, 'J': 19, 'K': 19
        }
        for col, width in column_widths.items():
            ws_main.column_dimensions[col].width = width

        wb.save(report_file)
        # 修正報表檔案名顯示，只顯示檔案名
        report_filename = os.path.basename(report_file)
        if period == 'daily':
            logger.info(f"✅ 每日報表已生成：{report_filename}")
        elif period == 'monthly':
            logger.info(f"✅ 每月報表已生成：{report_filename}")
            notify_and_log(f"📊 月報生成成功：{report_title}", level="info")
    except Exception as e:
        logger.exception("[generate_report] 生成報表失敗 traceback：")
        send_telegram(f"❌ 生成報表失敗：{str(e)[:100]}")

def auto_daily_report():
    global has_processed_delivery_exit
    has_processed_delivery_exit = False
    # 移除emoji，使用should_log_message過濾
    if should_log_message("已重置交割日旗標：交割平倉"):
        logger.info("已重置交割日旗標：交割平倉")
    generate_report(period='daily')

def auto_monthly_report():
    try:
        current_date = datetime.now()
        year = current_date.year
        month = current_date.month
        all_trades = []
        for day in range(1, 32):
            date = datetime(year, month, 1).replace(day=day)
            if date.month != month or date > current_date:
                break
            filename = f"{LOG_DIR}/trades_{date.strftime('%Y%m%d')}.json"
            if os.path.exists(filename):
                try:
                    with open(filename, 'r') as f:
                        trades = json.load(f)
                    all_trades.extend(trades)
                except json.JSONDecodeError:
                    logger.warning(f"交易記錄檔案 {filename} 格式錯誤，跳過")
        if all_trades:
            generate_report(period='monthly', date_str=f"{year}/{month:02d}", trades=all_trades)
        else:
            logger.info(f"{year}/{month:02d} 無交易數據，跳過月報")
    except Exception as e:
        logger.error(f"生成月報失敗：{str(e)}")
        send_telegram(f"❌ 月報生成失敗：{str(e)[:100]}")

def is_last_trading_day_of_month():
    today = datetime.now().date()
    # 檢查 today 是否為交易日（靜默檢查）
    if not holiday_calendar.is_trading_day(today, log_result=False):
        return False

    # 檢查今天之後是否還有交易日
    for i in range(1, 32): # 最多檢查到月底
        future_date = today + timedelta(days=i)
        if future_date.month != today.month:
            # 已經到下個月了，如果今天沒有找到後續交易日，則今天就是最後一個交易日
            return True
        if holiday_calendar.is_trading_day(future_date, log_result=False):
            return False # 今天之後還有交易日，所以今天不是最後一個交易日
    return True # 如果迴圈結束都沒找到後續交易日，表示今天就是最後一個交易日

def check_and_run_monthly_report():
    if is_last_trading_day_of_month():
        logger.info("今天是本月最後一個交易日，執行月報生成")
        auto_monthly_report()
    else:
        logger.info("今天不是本月最後一個交易日，跳過月報生成")

def send_startup_message(is_reboot=False):
    global balance_cache, margin_requirements
    balance_cache = {'balance': None, 'margin': None, 'last_updated': None}
    update_balance_cache(verbose=True, is_scheduled=False, force_update=True)
    account_info = ""
    try:
        margin_data = api.margin()
        field_translations = {
            'equity': '權益總值',
            'equity_amount': '權益總額',
            'today_balance': '今日餘額',
            'yesterday_balance': '昨日餘額',
            'available_margin': '可用保證金',
            'initial_margin': '原始保證金',
            'maintenance_margin': '維持保證金',
            'risk_indicator': '風險指標',
            'fee': '手續費',
            'tax': '期交稅',
            'future_settle_profitloss': '本日平倉損益'
        }
        translated_data = {field_translations.get(key, key): value for key, value in margin_data.__dict__.items() if key in field_translations}
        ordered_keys = [
            '權益總值', '權益總額', '今日餘額', '昨日餘額', '可用保證金',
            '原始保證金', '維持保證金', '風險指標', '手續費', '期交稅', '本日平倉損益'
        ]
        for key in ordered_keys:
            value = translated_data.get(key, 0)
            unit = "%" if key == "風險指標" else ""
            account_info += f"{key}：{value}{unit}\n"
    except Exception as e:
        logger.error(f"獲取帳戶餘額資訊失敗：{str(e)}")
        send_telegram(f"❌ 無法獲取帳戶餘額資訊：{str(e)[:100]}")
        account_info = "無法獲取帳戶餘額資訊\n"
    positions_info = ""
    try:
        positions = api.list_positions(api.futopt_account)
        if positions:
            for pos in positions:
                contract_type = "大台" if pos.code.startswith("TXF") else "小台" if pos.code.startswith("MXF") else "微台"
                direction = "多單" if pos.direction == Action.Buy else "空單"
                positions_info += (
                    f"［{contract_type}］｜"
                    f"動作：{direction}｜"
                    f"數量：{abs(pos.quantity)}口｜"
                    f"均價：{pos.price:.0f}\n"
                    f"未實現盈虧：{pos.pnl:.0f} TWD\n"
                )
        else:
            positions_info = "無持倉部位\n"
    except Exception as e:
        logger.error(f"獲取持倉資訊失敗：{str(e)}")
        send_telegram(f"❌ 無法獲取持倉資訊：{str(e)[:100]}")
        positions_info = "無法獲取持倉資訊\n"
    notification_type = "✅ 自動交易台指期已重新啟動！！！" if is_reboot else "✅ 自動交易台指期正在啟動中....."
    startup_message = (
        f"{notification_type}\n"
        f"═════ 系統資訊 ═════\n"
        f"啟動時間：{datetime.now().strftime('%Y/%m/%d %H:%M')}\n"
        f"憑證效期：{CERT_START.date()} 至 {CERT_END.date()}\n"
        f"綁定帳戶：{api.futopt_account.account_id if hasattr(api, 'futopt_account') and api.futopt_account else '未設定'}\n"
        f"API 狀態：{'已連線' if connected else '未連線'}\n"
        f"═════ 選用合約 ═════\n"
        f"［大台］{format_contract_display(contract_txf.code, contract_txf.delivery_date) if contract_txf else '未設定'} ${margin_requirements.get('大台', 0):,}\n"
        f"［小台］{format_contract_display(contract_mxf.code, contract_mxf.delivery_date) if contract_mxf else '未設定'} ${margin_requirements.get('小台', 0):,}\n"
        f"［微台］{format_contract_display(contract_tmf.code, contract_tmf.delivery_date) if contract_tmf else '未設定'} ${margin_requirements.get('微台', 0):,}\n"
        f"═════ 帳戶狀態 ═════\n"
        f"{account_info}"
        f"═════ 持倉狀態 ═════\n"
        f"{positions_info}"
    )
    if is_reboot:
        logger.info("自動交易台指期已重新啟動！")
    send_telegram(startup_message)

def monitor_connection():
    global connected, last_error_time, notified
    last_error_time = 0
    notified = False
    time.sleep(60)
    while True:
        try:
            # 只在交易日的指定時段才檢查連線和更新餘額
            now = datetime.now()
            now_date = now.date()
            now_time = now.time()
            
            # 首先檢查是否為交易日（靜默檢查，避免重複日誌）
            if not holiday_calendar.is_trading_day(now_date, log_result=False):
                time.sleep(60)
                continue
            
            # 早盤 8:30~13:45，午盤 14:50~05:01
            in_morning = now_time >= datetime.strptime("08:30", "%H:%M").time() and now_time <= datetime.strptime("13:45", "%H:%M").time()
            in_afternoon = now_time >= datetime.strptime("14:50", "%H:%M").time() or now_time <= datetime.strptime("05:01", "%H:%M").time()
            if not (in_morning or in_afternoon):
                time.sleep(60)
                continue
            
            # 測試連線狀態
            api.list_accounts()
            
            # 如果之前是斷線狀態，現在連線成功
            if not connected and state.last_notify_status == "error":
                try:
                    update_balance_cache(verbose=False, is_scheduled=False, force_update=True)
                    send_telegram(f"✅ API 連線成功")
                    logger.info(f"API 連線已恢復")
                except Exception as e:
                    translated_error = translate_api_message(str(e))
                    send_telegram(f"✅ API 連線成功")
                    logger.warning(f"API 連線恢復但餘額更新失敗：{translated_error}")
                state.last_notify_status = "ok"
            
            with global_lock:
                connected = True
            
            # 定期更新餘額
            update_balance_cache(verbose=False, is_scheduled=True)
            time.sleep(60)
            
        except Exception as e:
            # 連線失敗處理
            if connected and state.last_notify_status != "error":
                translated_error = translate_api_message(str(e))
                # 根據錯誤類型提供不同的說明
                if 'timeout' in str(e).lower():
                    error_desc = "連線超時"
                elif 'no route to host' in str(e).lower():
                    error_desc = "網路路由問題"
                elif 'connection' in str(e).lower():
                    error_desc = "連線中斷"
                else:
                    error_desc = "連線異常"
                send_telegram(f"❌ API 連線中斷\n斷線原因：{error_desc}\n{translated_error[:100]}\n\n系統正在嘗試重新連線．．")
                logger.error(f"API 連線中斷：{error_desc} - {translated_error}")
                state.last_notify_status = "error"
            
            # 嘗試重新登入
            try:
                api.logout()
            except Exception as logout_error:
                logger.warning(f"登出失敗：{translate_api_message(str(logout_error))}，繼續嘗試重新登入")
            
            # 重新登入
            logger.info("開始重新登入 API...")
            login_api(max_retries=5, retry_delay=30)
            
            with global_lock:
                connected = False
            
            time.sleep(60)

# run_scheduler
def run_scheduler():
    # 每日 00:00 重置交易追蹤計數器，清除當日交易記錄
    schedule.every().day.at("00:00").do(reset_trade_counter)
    # pending_deals 功能已移除，改為立即發送通知模式
    # 每日 00:02 清除舊的訂單開平記錄（order_octype_map）- 延後到報表生成後
    schedule.every().day.at("00:02").do(clear_old_order_octype)
   # 每日 14:50 若為交易日，更新保證金需求（margin_requirements），若有變動發送 Telegram 通知
    schedule.every().day.at("14:50").do(lambda: update_margin_requirements(is_scheduled=True) if holiday_calendar.is_trading_day(log_result=False) else None)
    # 每日 00:01 生成前一日交易日報，包含交易統計、帳戶狀態和持倉資訊
    schedule.every().day.at("00:01").do(lambda: auto_daily_report() if was_yesterday_trading_day() else logger.info("📅 昨日非交易日，跳過日報"))
    # 每日 08:30 若為交易日，發送系統啟動訊息（非重啟模式），包含帳戶和持倉狀態
    schedule.every().day.at("08:30").do(lambda: send_startup_message(is_reboot=False) if holiday_calendar.is_trading_day(log_result=False) else None)
    # 每日 08:00 執行 TOKEN 重連（登出後重新登入），更新餘額，記錄日誌，無 Telegram 通知
    schedule.every().day.at("08:00").do(lambda: (logger.info("開始每日 08:00 TOKEN 重連"), login_api() and update_balance_cache(verbose=False, is_scheduled=False, force_update=True) or logger.info("完成每日 08:00 TOKEN 重連"))[1])
    # 每日 20:00 執行 TOKEN 重連（登出後重新登入），更新餘額，記錄日誌，無 Telegram 通知
    schedule.every().day.at("20:00").do(lambda: (logger.info("開始每日 20:00 TOKEN 重連"), login_api() and update_balance_cache(verbose=False, is_scheduled=False, force_update=True) or logger.info("完成每日 20:00 TOKEN 重連"))[1])
    # 每日 08:05 若為交割日，發送 Telegram 通知提醒平倉後轉倉
    # schedule.every().day.at("08:05").do(lambda: notify_and_log("今日為交割日，平倉後將轉倉至次月合約", level="warning") if is_delivery_day() else None)
    # 每日 23:59 若為月底最後一天，生成月報，包含當月交易統計
    schedule.every().day.at("23:59").do(check_and_run_monthly_report)
    schedule.every().day.at("00:05").do(check_and_send_rollover_reminder)
    schedule.every().day.at("13:45").do(force_rollover_if_pre_delivery_trading_day)
    while True:
        schedule.run_pending()
        time.sleep(10)

# 日誌自動清理（保留90天）
def clean_old_logs(log_dir, days=90):
    now = time.time()
    for filename in os.listdir(log_dir):
        file_path = os.path.join(log_dir, filename)
        if os.path.isfile(file_path):
            file_mtime = os.path.getmtime(file_path)
            if now - file_mtime > days * 86400:
                try:
                    os.remove(file_path)
                    logger.info(f"🧹 已自動刪除過舊日誌：{filename}")
                except Exception as e:
                    logger.warning(f"刪除日誌失敗：{filename}，錯誤：{e}")

clean_old_logs(LOG_DIR)

# 憑證到期自動提醒
def check_cert_expiry():
    now = datetime.now()
    days_left = (CERT_END - now).days
    if days_left in [7, 3, 1]:
        send_telegram(f"⚠️ 憑證將於 {CERT_END.strftime('%Y-%m-%d')} 到期，剩餘 {days_left} 天，請盡快更新！")
check_cert_expiry()

# Webhook/手動下單 IP 白名單驗證（使用配置管理）
def is_ip_allowed(ip):
    return ip in config.ALLOWED_IPS

# 新增：判斷下個交易日是否為交割日

def is_next_trading_day_delivery_day():
    today = datetime.now().date()
    # 找到下一個交易日
    for i in range(1, 7):
        next_day = today + timedelta(days=i)
        if holiday_calendar.is_trading_day(next_day, log_result=False):
            return is_delivery_day(next_day)
    return False

# 轉倉相關變數已整合到 SystemState 類中
# use_next_month_contract, rollover_reminder_sent, rollover_notification_sent 等已移至 state 物件

def send_rollover_notification(force=False):
    """發送轉倉通知"""
    global rollover_notification_sent, use_next_month_contract
    if not force and rollover_notification_sent:
        return
    
    try:
        current_month = datetime.now().strftime('%Y%m')
        msg = "轉倉通知\n"
        msg += "今日為交割日前一個交易日，系統將自動切換至次月合約：\n"
        
        for code, var in [("TXF", "contract_txf"), ("MXF", "contract_mxf"), ("TMF", "contract_tmf")]:
            next_contract = get_next_month_contract(code, current_month)
            if next_contract:
                contract_name = "大台" if code == "TXF" else "小台" if code == "MXF" else "微台"
                msg += f"【{contract_name}】：{next_contract.code}（{next_contract.delivery_month}）\n"
        
        msg += "\n下筆交易將使用次月合約進行。"
        send_telegram(msg)
        logger.info("轉倉通知已發送")
        rollover_notification_sent = True
        use_next_month_contract = True
    except Exception as e:
        logger.error(f"發送轉倉通知失敗：{str(e)}")

def force_rollover_if_pre_delivery_trading_day():
    today = datetime.now().date()
    if is_next_trading_day_delivery_day() and holiday_calendar.is_trading_day(today, log_result=False):
        if not rollover_notification_sent:
            send_rollover_notification(force=True)

def check_and_send_rollover_reminder():
    today = datetime.now().date()
    if is_next_trading_day_delivery_day() and holiday_calendar.is_trading_day(today, log_result=False):
        positions = api.list_positions(api.futopt_account)
        has_position = any(abs(pos.quantity) > 0 for pos in positions)
        send_telegram("⚠️ 轉倉提醒\n今日為交割日前一個交易日，下筆交易將進行轉倉。")
        logger.info("轉倉提醒\n今日為交割日前一個交易日，下筆交易將進行轉倉。")
        global rollover_reminder_sent
        rollover_reminder_sent = True
        if not has_position:
            send_rollover_notification()

def translate_callback_content(content):
    """翻譯API回調內容為中文"""
    if not content:
        return content
    
    content_str = str(content)
    translated = content_str
    
    # 翻譯已知的API術語
    for eng_term, chinese_term in config.CALLBACK_STATUS_TRANSLATIONS.items():
        if eng_term in translated:
            translated = translated.replace(eng_term, chinese_term)
    
    return translated

def format_callback_message(state, deal, order=None):
    """格式化API回調訊息為中文顯示"""
    try:
        # 翻譯狀態
        state_str = translate_callback_content(str(state))
        
        # 獲取關鍵信息 - 修復期貨成交時單號顯示"未知"的問題
        if str(state) == 'OrderState.FuturesDeal':
            # 期貨成交：使用 trade_id
            order_id = deal.get('trade_id', deal.get('order_id', '未知'))
        else:
            # 其他訂單狀態：使用 order 結構的 id
            order_id = deal.get('order', {}).get('id', '未知')
        
        contract_code = deal.get('contract', {}).get('code', deal.get('code', ''))
        contract_name = get_contract_name(contract_code)
        
        # 操作信息
        operation = deal.get('operation', {})
        op_type = translate_callback_content(operation.get('op_type', ''))
        op_code = operation.get('op_code', '00')
        op_msg = config.OP_MSG_TRANSLATIONS.get(operation.get('op_msg', ''), operation.get('op_msg', ''))
        
        # 訂單信息 - 期貨成交時優先從deal直接欄位獲取
        if str(state) == 'OrderState.FuturesDeal':
            # 期貨成交：直接從deal獲取數據
            quantity = deal.get('quantity', 0)
            price = deal.get('price', 0.0)
            action = translate_callback_content(deal.get('action', ''))
            # 期貨成交通常沒有訂單類型等信息
            order_type = ''
            price_type = ''
            oc_type = ''
        else:
            # 其他訂單狀態：從order結構獲取
            order_info = deal.get('order', {})
            action = translate_callback_content(order_info.get('action', ''))
            quantity = order_info.get('quantity', 0)
            price = order_info.get('price', 0.0)
            order_type = translate_callback_content(order_info.get('order_type', ''))
            price_type = translate_callback_content(order_info.get('price_type', ''))
            oc_type = translate_callback_content(order_info.get('oc_type', ''))
        
        # 格式化訊息
        msg_parts = [
            f"[{state_str}]",
            f"單號：{order_id}",
            f"合約：[{contract_name}] {contract_code}"
        ]
        
        if quantity > 0:
            msg_parts.append(f"數量：{quantity}口")
        
        if price > 0:
            msg_parts.append(f"價位：{price:.0f}")
        
        if action:
            msg_parts.append(f"動作：{action}")
        
        if oc_type:
            msg_parts.append(f"開平：{oc_type}")
        
        if order_type:
            msg_parts.append(f"類型：{order_type}")
        
        if price_type:
            msg_parts.append(f"價格：{price_type}")
        
        if op_type:
            msg_parts.append(f"操作：{op_type}")
        
        if op_code != '00':
            msg_parts.append(f"狀態碼：{op_code}")
        
        if op_msg:
            msg_parts.append(f"訊息：{op_msg}")
        
        return " | ".join(msg_parts)
        
    except Exception as e:
        logger.warning(f"格式化回調訊息失敗：{e}")
        return f"API回調：{translate_callback_content(str(state))}"

def get_order_context_from_request():
    """從當前請求上下文中獲取訂單信息"""
    try:
        from flask import request, has_request_context
        if has_request_context() and request.is_json:
            data = request.get_json()
            if data:
                return {
                    'contract_name': data.get('contract_name', ''),
                    'qty': data.get('qty', 0),
                    'price': data.get('price', 0),
                    'action': data.get('action', ''),
                    'order_type': data.get('order_type', ''),
                    'price_type': data.get('price_type', '')
                }
    except:
        pass
    return {}

def format_unified_log(log_type="取消", order_id="未知", contract_name="", contract_code="", qty=0, price=0, action="", reason="", octype="", order_type="", price_type="", auto_fill=True):
    """統一格式化日誌訊息 - 動態日誌類型，盡量填充完整訂單信息"""
    try:
        # 如果啟用自動填充且缺少信息，嘗試從請求上下文獲取
        if auto_fill:
            context = get_order_context_from_request()
            if not contract_name and context.get('contract_name'):
                contract_name = context['contract_name']
            if not qty and context.get('qty'):
                qty = context['qty']
            if not price and context.get('price'):
                price = context['price']
            if not action and context.get('action'):
                action = context['action']
        
        # 根據日誌類型決定前綴
        log_prefix_map = {
            "取消": "[期貨取消]",
            "下單": "[期貨下單]", 
            "成交": "[期貨成交]",
            "修改": "[期貨修改]",
            "錯誤": "[期貨錯誤]"
        }
        msg_parts = [log_prefix_map.get(log_type, f"[期貨{log_type}]")]
        
        # 總是顯示單號，即使是未知
        msg_parts.append(f"單號：{order_id}")
        
        # 合約信息 - 嘗試獲取完整的合約代碼
        if contract_name and not contract_code:
            # 如果只有合約名稱，嘗試獲取對應的合約代碼
            try:
                # 從全局變數獲取
                if contract_name == "大台" and 'contract_txf' in globals() and contract_txf:
                    contract_code = contract_txf.code
                elif contract_name == "小台" and 'contract_mxf' in globals() and contract_mxf:
                    contract_code = contract_mxf.code
                elif contract_name == "微台" and 'contract_tmf' in globals() and contract_tmf:
                    contract_code = contract_tmf.code
                # 或者從配置映射嘗試推算（如果上述失敗）
                if not contract_code and contract_name in ["大台", "小台", "微台"]:
                    contract_code = f"推算中..."
            except Exception as e:
                logger.debug(f"獲取合約代碼失敗：{e}")
                pass
        
        if contract_name and contract_code:
            msg_parts.append(f"合約：[{contract_name}] {contract_code}")
        elif contract_name:
            msg_parts.append(f"合約：{contract_name}")
        else:
            msg_parts.append(f"合約：未知")
        
        # 數量 - 安全的類型檢查和轉換
        try:
            qty_num = float(qty) if qty else 0
            if qty_num > 0:
                msg_parts.append(f"數量：{int(qty_num)}口")
            else:
                msg_parts.append(f"數量：未知")
        except (ValueError, TypeError):
            msg_parts.append(f"數量：未知")
        
        # 價位 - 安全的類型檢查和轉換
        try:
            price_num = float(price) if price else 0
            if price_num > 0:
                msg_parts.append(f"價位：{price_num:.0f}")
            else:
                msg_parts.append(f"價位：未知")
        except (ValueError, TypeError):
            msg_parts.append(f"價位：未知")
        
        # 動作 - 如果沒有提供，顯示未知
        if action:
            msg_parts.append(f"動作：{action}")
        else:
            msg_parts.append(f"動作：未知")
        
        # 原因 - 總是顯示
        if reason:
            msg_parts.append(f"原因：{reason}")
        else:
            msg_parts.append(f"原因：未知錯誤")
        
        return " | ".join(msg_parts)
        
    except Exception as e:
        logger.warning(f"格式化統一日誌失敗：{e}")
        log_prefix = f"[期貨{log_type}]" if log_type else "[期貨錯誤]"
        return f"{log_prefix} | 單號：{order_id} | 合約：未知 | 數量：未知 | 價位：未知 | 動作：未知 | 原因：{reason if reason else '格式化失敗'}"

def get_action_display_by_rule(octype, direction):
    if octype == 'New':
        if direction == 'Buy':
            return '多單買入'
        elif direction == 'Sell':
            return '空單買入'
    elif octype == 'Cover':
        if direction == 'Sell':
            return '多單賣出'
        elif direction == 'Buy':
            return '空單賣出'
    return '未知動作'

# 新增輔助函數：根據合約 code 取得交割月份

def get_delivery_month_by_code(code):
    try:
        # 先嘗試從API獲取
        if code.startswith('TXF'):
            contracts = api.Contracts.Futures.get('TXF')
        elif code.startswith('MXF'):
            contracts = api.Contracts.Futures.get('MXF')
        elif code.startswith('TMF'):
            contracts = api.Contracts.Futures.get('TMF')
        else:
            # 如果不是支援的合約類型，直接解析
            match = re.search(r'\d{6}$', code)
            return match.group(0) if match else '未知'
        
        # 從API合約清單中查找
        for c in contracts:
            if c.code == code:
                # 優先取delivery_month，如果沒有則嘗試delivery_date
                if hasattr(c, 'delivery_month') and c.delivery_month:
                    return c.delivery_month
                elif hasattr(c, 'delivery_date') and c.delivery_date:
                    # 從delivery_date轉換為YYYYMM格式
                    try:
                        if isinstance(c.delivery_date, str) and len(c.delivery_date) >= 7:
                            # 假設格式是YYYY/MM/DD或YYYY-MM-DD
                            date_parts = c.delivery_date.replace('-', '/').split('/')
                            if len(date_parts) >= 2:
                                return f"{date_parts[0]}{date_parts[1]:0>2}"
                    except:
                        pass
                break
        
        # fallback 1: 從code後綴解析 (如TXFJ4 -> 202407)
        match = re.search(r'[A-Z](\d)$', code)  # 匹配最後的字母+數字
        if match:
            year_digit = int(match.group(1))
            current_year = datetime.now().year
            
            # 推算完整年份（假設最多跨5年）
            if year_digit <= 5:
                full_year = 2020 + year_digit
            else:
                full_year = 2010 + year_digit
                
            # 從合約代碼中提取月份字母
            month_match = re.search(r'([A-Z])(\d)$', code)
            if month_match:
                month_letter = month_match.group(1)
                # 期貨月份代碼對照（標準國際期貨月份代碼）
                month_map = {
                    'F': '01', 'G': '02', 'H': '03', 'J': '04', 'K': '05', 'M': '06',
                    'N': '07', 'Q': '08', 'U': '09', 'V': '10', 'X': '11', 'Z': '12'
                }
                month = month_map.get(month_letter, '01')
                return f"{full_year}{month}"
        
        # fallback 2: 直接從code尾部提取6位數字
        match = re.search(r'\d{6}$', code)
        if match:
            return match.group(0)
            
        return '未知'
        
    except Exception as e:
        logger.warning(f"get_delivery_month_by_code 失敗：{e}")
        # 最後努力：簡單的字符串解析
        try:
            match = re.search(r'[A-Z](\d)$', code)
            if match:
                year_digit = int(match.group(1))
                return f"202{year_digit}01"  # 簡化為1月
        except:
            pass
        return '未知'

# 新增：格式化合約顯示為 MXFG5 (2025/07/16) 格式
def format_contract_display(code, delivery_date):
    """
    格式化合約顯示
    輸入: code='MXFG5', delivery_date='2025/07/16'
    輸出: 'MXFG5 (2025/07/16)'
    """
    try:
        if not code:
            return '未設定'
        
        if not delivery_date or delivery_date == '未知':
            return f"{code} (日期未知)"
        
        # 確保交割日期格式為 YYYY/MM/DD
        if isinstance(delivery_date, str):
            # 處理不同的日期格式
            if len(delivery_date) == 8 and delivery_date.isdigit():  # YYYYMMDD
                formatted_date = f"{delivery_date[:4]}/{delivery_date[4:6]}/{delivery_date[6:8]}"
            elif len(delivery_date) == 10 and '/' in delivery_date:  # YYYY/MM/DD
                formatted_date = delivery_date
            elif len(delivery_date) == 10 and '-' in delivery_date:  # YYYY-MM-DD
                formatted_date = delivery_date.replace('-', '/')
            else:
                formatted_date = delivery_date
        else:
            formatted_date = str(delivery_date)
            
        # 如果格式化後的日期是空字串，使用日期未知
        if not formatted_date.strip():
            formatted_date = "日期未知"
        
        return f"{code} ({formatted_date})"
    except Exception as e:
        logger.warning(f"format_contract_display 失敗 code={code}, delivery_date={delivery_date}, error={e}")
        return f"{code} (日期錯誤)"

# 新增：根據合約名稱獲取合約代碼的輔助函數
def get_contract_code_by_name(contract_name):
    """根據合約名稱獲取合約代碼"""
    try:
        if contract_name == "大台" and 'contract_txf' in globals() and contract_txf:
            return contract_txf.code
        elif contract_name == "小台" and 'contract_mxf' in globals() and contract_mxf:
            return contract_mxf.code
        elif contract_name == "微台" and 'contract_tmf' in globals() and contract_tmf:
            return contract_tmf.code
        else:
            return None
    except Exception as e:
        logger.warning(f"get_contract_code_by_name 失敗：{e}")
        return None

def get_formatted_order_message(is_success, order_id, contract_name, qty, price, octype, direction, order_type, price_type, is_manual, reason=None, contract_code=None, delivery_date=None):
    """格式化訂單提交訊息"""
    current_time = datetime.now().strftime('%Y/%m/%d %H:%M')
    
    # 獲取完整合約資訊
    try:
        # 如果沒有提供 contract_code，嘗試從合約名稱獲取
        if not contract_code:
            contract_code = get_contract_code_by_name(contract_name)
        
        # 如果有 contract_code，顯示完整格式；否則顯示 "未知"
        if contract_code:
            # 如果有提供 delivery_date，直接使用；否則從 API 獲取
            if delivery_date:
                contract_delivery_date = delivery_date
            else:
                # 從 contract_code 獲取交割日期
                try:
                    if contract_code.startswith('TXF'):
                        contracts = api.Contracts.Futures.get('TXF')
                    elif contract_code.startswith('MXF'):
                        contracts = api.Contracts.Futures.get('MXF')
                    elif contract_code.startswith('TMF'):
                        contracts = api.Contracts.Futures.get('TMF')
                    else:
                        contracts = []
                    
                    contract_delivery_date = "日期未知"
                    for c in contracts:
                        if c.code == contract_code:
                            # 嘗試獲取交割日期，優先使用 delivery_date，然後是 delivery_month
                            if hasattr(c, 'delivery_date') and c.delivery_date:
                                contract_delivery_date = c.delivery_date
                            elif hasattr(c, 'delivery_month') and c.delivery_month:
                                # 從 delivery_month (YYYYMM) 推算交割日期
                                try:
                                    year = int(c.delivery_month[:4])
                                    month = int(c.delivery_month[4:6])
                                    # 台指期交割日通常是第三個週三，這裡用月底簡化
                                    import calendar
                                    last_day = calendar.monthrange(year, month)[1]
                                    contract_delivery_date = f"{year}/{month:02d}/{last_day:02d}"
                                except:
                                    contract_delivery_date = f"{c.delivery_month[:4]}/{c.delivery_month[4:6]}/16"
                            break
                except:
                    # 如果無法連接API，使用預設值
                    contract_delivery_date = "日期未知"
            
            contract_display = format_contract_display(contract_code, contract_delivery_date)
        else:
            # 沒有 contract_code，直接顯示 "未知"
            contract_display = "未知"
    except Exception as e:
        logger.warning(f"獲取合約顯示格式失敗：{e}")
        contract_display = "未知"
    
    # 訂單類型
    try:
        if price_type and price_type in ["MKT", "LMT"]:
            # price_type 正常，使用標準函數
            order_type_display = get_order_type_display(price_type, order_type)
        else:
            # price_type 抓取不到或異常，顯示 "未知 (ROD)" 或 "未知 (IOC)"
            order_type_display = f"未知 ({order_type})" if order_type in ["ROD", "IOC"] else f"未知 (ROD)"
    except Exception as e:
        logger.warning(f"獲取訂單類型顯示失敗：{e}")
        order_type_display = f"未知 ({order_type})" if order_type in ["ROD", "IOC"] else f"未知 (ROD)"
    
    # 提交類型
    manual_type = "手動" if is_manual else "自動"
    octype_display = "開倉" if octype == "New" else "平倉"
    submit_type = f"{manual_type}{octype_display}"
    
    # 提交動作 - 根據開平倉和方向決定，與get_action_display邏輯一致
    if octype == "New":  # 開倉
        if direction == "Buy":
            submit_action = "多單買入"
        else:
            submit_action = "空單買入"
    else:  # 平倉
        if direction == "Buy":
            submit_action = "空單賣出"
        else:
            submit_action = "多單賣出"
    
    # 提交價格
    if price_type == "MKT":
        price_display = "市價"
    else:
        price_display = f"{price:.0f}"
    
    if is_success:
        msg = (f"⭕ 提交成功（{current_time}）\n"
               f"選用合約：{contract_display}\n"
               f"訂單類型：{order_type_display}\n"
               f"提交單號：{order_id}\n"
               f"提交類型：{submit_type}\n"
               f"提交動作：{submit_action}\n"
               f"提交部位：{contract_name}\n"
               f"提交數量：{qty} 口\n"
               f"提交價格：{price_display}")
    else:
        msg = (f"❌ 提交失敗（{current_time}）\n"
               f"選用合約：{contract_display}\n"
               f"訂單類型：{order_type_display}\n"
               f"提交單號：{order_id if order_id != '未知' else '未知'}\n"
               f"提交類型：{submit_type}\n"
               f"提交動作：{submit_action}\n"
               f"提交部位：{contract_name}\n"
               f"提交數量：{qty} 口\n"
               f"提交價格：{price_display}\n"
               f"原因：{reason}")
    
    return msg

# 主程序
if __name__ == '__main__':
    logger.info(f"永豐 API 版本：{sj.__version__}")
    # 先更新保證金資料，再登入API，這樣選用合約時就能顯示正確保證金
    update_margin_requirements()
    login_api()
    setup_event_handlers()
    send_startup_message(is_reboot=True)
    threading.Thread(target=monitor_connection, daemon=True).start()
    threading.Thread(target=run_scheduler, daemon=True).start()
    app.run(host='0.0.0.0', port=5000, debug=False)